# -*- coding: utf-8 -*-

"""
3.11.5                                                        序


https://blog.csdn.net/Hulunbuir/article/details/83715279?spm=1001.2014.3001.5506 csdn上的参照学习文件

 D:/Myprogramsystem/cad/打印服务/printcad_to_pdf.py  二十一世纪不是人工写小说或论文的适宜时代，但的确是适合调用各种手段例如人工智能来实现某

个重大目标任务的时代。程序系统是一个巨大的工程，需要我们综合很多学科技能。其中最重要的事情之一，就是认识理解这个程序系统本身。我建立了自己

的Myprogramsystem 体系，也许在专业人士看来有点可笑。但实际上，正是从这些可能幼稚、错误的做法上，我们可以逐渐正确理解“合理的概念和本质的规

律”。

    《CAD自动化》这个文件，是整个CAD系统的开发文件，是动态的文件。我们会在某个节点复制副本并标注相应的名称保留它们，以便将来的参照使用。就

目前来说，我需要将打印、编目录这些最底层的工作完成，它们最终将封装到《CAD基本操作》这个文件中。

    接下来的任务是方案图的深化、施工图的绘制、工程图纸的识别等研究。在这些研究中，必然要产生、处理天文数字的数据，不可能用word、记事本等

格式保存和引用，这就是我们应该重视数据库操作的原因。我当然知道，过去、现在、将来都会有专业的公司、专业的团队来开发研究这些问题。要知道

这些成果实际上的确是有商业价值的。但是最重要的是这将训练我们的能力。就是扎根在土木工程并充分运用计算机知识的能力。我认为应系统学习

一下计算机课程。实际上，数学仍然是最重要的，我们需要学习大学本科数学教育的课程，具备做研究数学的能力却并不去做数学研究。

    当然，有人会认为这些要求是不是太高了。那我们要人工智能干什么呢。人们不是认为人工智能无所不能吗？我们应该注意到，动用人工智能写一篇百万字

的小说，实际上是对读者的羞辱。当很多公文、工程文件、科学论文的内容是用人工智能完成的时候，只是表明它是一些必不可少、如果需要我们就认真去阅读

并引用的内容，但肯定不是需要读者耗费大量时间精力阅读学习的内容。目前的人工智能只是对天文数字量的数据进行了科学合理处理的结果，它没有意识也没

有灵魂思想，它就是水中、镜中我们人就智力方面而言的影子镜像而已。

    真正有价值的问题不是人工智能有没有意识，会不会统治控制人类。真正有价值的问题是人工智能将迫使我们理解清楚人的本质、社会的本质、政治的本质、

经济的本质。私人拥有过量财富是不是极大的罪恶？拥有天量资本的集团是不是国家民族生死存亡的巨大威胁？社保医保制度是不是恰恰引发一个国家社会人口

日益锐减最后造成整个体系崩溃的根源？如果有人自认为是哲学家或者爱好哲学，他就不得不为所有人类的生存担忧，不得不为所有人活着的尊严、价值、意义

担忧。


完美解决各个脚本公用变量mp,doc,acad等

每个脚本都定义了L1()

在主脚本定义Lx，里面同时运行 L1() cj.L1() sj.L1()所有函数都会激活

eu不需要和脚本绑定

20240820
                                                                                                                                   20250330






"""

















#&&&&%%   CAD基本操作 

#&&&&%%  第一部分  导入、转换、连接等前置程序 
#_____________________________________________________________________________________________________________________________________________


# 导入模块


# —— 脚本最顶部，就写这两行 —— 
import warnings
warnings.filterwarnings(
    "ignore",
    r".*Revert to STA COM threading mode.*",
    category=UserWarning,
    module=r"pywinauto\..*"
)
from pywinauto.application import Application


#A______________________________________________________________

#B______________________________________________________________

import builtins



#C______________________________________________________________

from collections import defaultdict,deque

from concurrent.futures import ThreadPoolExecutor, TimeoutError

import cv2

import ctypes

from contextlib import redirect_stdout, redirect_stderr,suppress
from utils.cad_decorators import cad_operation, wait_cad_idle, wait_document_opened


#D______________________________________________________________


import datetime

#E______________________________________________________________

#F______________________________________________________________
import functools

from fractions import Fraction

from functools import cmp_to_key

from fitz import Rect

import fitz

from functools import wraps

import functools

from functools import wraps

import fire









#G______________________________________________________________

#H______________________________________________________________

#I______________________________________________________________

from itertools import chain

import importlib

import imageio

import inspect

import itertools

import io


#J______________________________________________________________


import  json


#K______________________________________________________________

#L______________________________________________________________

#M______________________________________________________________

import math

import multiprocessing

import mysql.connector

import matplotlib.pyplot as plt


from multiprocessing import Process


#N______________________________________________________________

import networkx as nx

import numpy as np 


#O______________________________________________________________

import os


# ———— 1. 抑制 pygame 欢迎提示 ————
# 必须在任何 import pygame 之前设置
os.environ.setdefault("PYGAME_HIDE_SUPPORT_PROMPT", "1")


from openpyxl import Workbook, load_workbook

from openpyxl.utils import column_index_from_string as col2idx

#P______________________________________________________________

import PyPDF2.errors

import pywintypes

from  pyautocad  import Autocad,APoint,aDouble

import pyautocad.types

import pytesseract

from PIL import ImageGrab

from pythoncom import VT_ARRAY, VT_R8

import pyautogui



from pywinauto import findwindows

import psutil

import pythoncom

import pygame

import pygetwindow as gw

import pyttsx3

from PIL  import Image



import pprint

import pdb

from pathlib import Path

import psutil

from pygetwindow import getWindowsWithTitle

from pypdf import PdfReader, PdfWriter


#Q______________________________________________________________

#R______________________________________________________________

import runpy

import regex as xre

import re

import random

#S______________________________________________________________

from shapely.geometry import Point, Polygon,LineString

import subprocess

import shutil

from sympy import *

import sys

from scipy.spatial import ConvexHull


from subprocess import DETACHED_PROCESS




#T______________________________________________________________

import time

import tempfile

from typing import Tuple, Literal

from typing import Any, Dict, List ,Tuple, Optional

from typing import Sequence, Literal, Optional,Union

from typing import Dict, Set,List

import tkinter as tk

import traceback

import tkinter

import threading


#U______________________________________________________________
import uuid
#V______________________________________________________________

#W______________________________________________________________



# ===== pywin32 / COM 基础 =====
import win32com                      # 确保顶层包存在

import win32com.client

import win32com.client as win32      # 常用入口，统一别名
from win32com.client import CastTo, makepy, constants, VARIANT,Dispatch

# 动态调度（少数场景用）
import win32com.client.dynamic as dyn

# Windows API / GUI / 进程
import win32gui
import win32api
import win32con
import win32process

import pythoncom                     # CoInitialize, COM 错误
import pywintypes                    # com_error


"""
主力用：win32.gencache.EnsureDispatch("AutoCAD.Application")（更稳、配合 MakePy）。

偶尔用：win32.Dispatch(...) 或 dyn.Dispatch(...)（需要强制晚绑定/动态属性时）。

避免把变量命名为 constants，以免遮蔽 win32com.client.constants。

如果你做成脚手架模块（scaffold），可加：

__all__ = [
    "win32com", "win32", "CastTo", "makepy", "constants", "VARIANT",
    "dyn", "win32gui", "win32api", "win32con", "win32process",
    "pythoncom", "pywintypes",
]


"""

#X______________________________________________________________

#Y______________________________________________________________

#Z______________________________________________________________

import psutil
import subprocess
import sys
from subprocess import CREATE_NO_WINDOW, DETACHED_PROCESS


from typing import List, Dict
from pathlib import Path
import datetime
from typing import List, Set, Dict




import datetime
import unicodedata


import psutil
import os
import time
import signal
import subprocess
from subprocess import CREATE_NO_WINDOW, DETACHED_PROCESS


from typing import Tuple
import io
import fitz  # PyMuPDF
from PIL import Image



#自建库


sys.path.append('D:/Myprogramsystem')






#&&% 函数别名
def alias(*names):
    """
    @alias("别名1","别名2",…)
    def foo(...): …
    """
    def decorator(func):
        mod = sys.modules[func.__module__]
        for nm in names:
            setattr(mod, nm, func)
        return func
    return decorator

# —— 保存原始 print ——
_orig_print = builtins.print

# ---------------- 全局劫持 ----------------
_orig_print = builtins.print            # 先保存原 print
def suppress_all_prints():
    builtins.print = lambda *a, **k: None
def restore_all_prints():
    builtins.print = _orig_print

# ---------------- 局部调试控制 ----------------
DEBUG             = False              # 总开关
_DEBUG_CODE_STACK = []                 # ← 新增：调试函数调用栈

def node(msg: str, *args, **kwargs):
    """
    只有 DEBUG = True 且当前帧属于【栈底】函数时才打印。
    """
    if not DEBUG or not _DEBUG_CODE_STACK:
        return
    frame = inspect.currentframe().f_back
    try:
        # 只允许“最外层”调试函数输出
        if frame.f_code is _DEBUG_CODE_STACK[0]:
            _orig_print(msg.format(*args), **kwargs)
    finally:
        del frame

@alias("e")
def enable_debug():
    """开启调试：普通 print → 静默，仅 node() 生效。"""
    global DEBUG
    DEBUG = True
    suppress_all_prints()

@alias("d")
def disable_debug():
    """关闭调试：恢复普通 print，node() 失效。"""
    global DEBUG
    DEBUG = False
    restore_all_prints()

def debuggable(func):
    """
    装饰器：进入时把 func.__code__ 推入栈；离开时弹栈。
    node() 始终只看栈底元素，所以只有第一次进入的函数会真正输出。
    """
    @functools.wraps(func)
    def wrapper(*args, **kw):
        _DEBUG_CODE_STACK.append(func.__code__)     # —— 入栈
        try:
            return func(*args, **kw)
        finally:
            _DEBUG_CODE_STACK.pop()                 # —— 出栈
    return wrapper


#&&% 控制函数运行时间
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --


# 日志配置
LOG_PATH = Path(r"D:/Myprogramsystem/XT/系统运行中断错误记录.xlsx")
LOG_SHEET = "错误记录"

# 初始化日志文件和工作表
if not LOG_PATH.exists():
    wb = Workbook()
    ws = wb.active
    ws.title = LOG_SHEET
    ws.append(["时间", "函数名", "参数repr", "备注"])
    wb.save(LOG_PATH)


def _log(func_name, args, kwargs, note):
    """
    将错误或超时信息写入 Excel 日志。如果工作表不存在则创建。
    """
    from time import strftime
    wb = load_workbook(LOG_PATH)
    # 确保日志表存在
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
    else:
        ws = wb.create_sheet(LOG_SHEET)
        ws.append(["时间", "函数名", "参数repr", "备注"])
    ws.append([strftime("%Y-%m-%d %H:%M:%S"),
               func_name,
               f"args={args!r}, kwargs={kwargs!r}",
               note])
    wb.save(LOG_PATH)


def _kill_acad():
    """
    强制结束所有以 acad 开头的 CAD 进程。
    """
    for proc in psutil.process_iter(("pid", "name")):
        name = proc.info.get("name")
        if name and name.lower().startswith("acad"):
            try:
                proc.kill()
            except Exception:
                pass


def timeout_and_log2(timeout_sec: float):
    """
    装饰器：对关键函数添加双重超时保护。
    - 守护线程监控主函数执行，超过 timeout_sec 秒将
      调用 _kill_acad() 并记录日志。
    - 函数内部可根据 timeout_sec+1 触发二次超时操作。

    用法：
        @timeout_and_log2(20)
        def Fx(..., timeout_sec=20):
            # 可选二次超时监控
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            finished = threading.Event()

            def watchdog():
                # 等待 timeout_sec 秒后，如函数尚未结束，杀CAD并记录
                if not finished.wait(timeout_sec):
                    _kill_acad()
                    _log(func.__name__, args, kwargs,
                         f"超时 {timeout_sec}s，强制终止 CAD")
            threading.Thread(target=watchdog, daemon=True).start()

            try:
                result = func(*args, **kwargs)
            except Exception:
                _log(func.__name__, args, kwargs,
                     f"函数内部异常:\n{traceback.format_exc()}")
                raise
            finally:
                finished.set()

            return result
        return wrapper
    return decorator

# ---------- 示例：双重超时测试函数 ----------
@timeout_and_log2(20)
def test_draw_circle_and_wait(center=(0,0,0), radius=10000, timeout_sec=20):
    """
    测试：在 CAD 中画圆并触发等待命令，
    后台线程在 timeout_sec+1 秒后发送无效命令以触发中断。
    """
    import threading
    from time import sleep, time

    # 获取当前 DWG 文件名
    try:
        current_file = current_dwg_basename()
    except Exception:
        current_file = '<unknown>'

    try:


        # 第1步：画圆
        circ = draw_circle(center, radius)
        if circ is None:
            print("❌ 圆绘制失败，测试中断。")
            return


        # 第2步：触发 CAD 等待
        doc.SendCommand("h\n")

        # 第3步：二次超时线程：timeout_sec+1 秒后发无效命令中断CAD
        def timeout_action():
            start = time()
            while True:
                if time() - start > timeout_sec + 1:
                    try:
                        doc.SendCommand("_.POINT 0,0,0")
                    except Exception as e:
                        print(f"⚠ 第二重超时触发，中断 CAD: {e}")
                    break
                sleep(0.1)

        threading.Thread(target=timeout_action, daemon=True).start()
        return circ

    except Exception as e:
        print(f"❌ test_draw_circle_and_wait 捕获到异常，退出: {e}")
        # 记录日志：函数名、输入参数、当前 DWG、异常信息
        try:
            current_file = getattr(doc, 'Name', '<unknown>')
            _log('test_draw_circle_and_wait', (center, radius), {'timeout_sec': timeout_sec},
                 f"内部异常，文件={current_file}, 异常={e}")
        except Exception:
            pass
        return

#函数计时

def timeit(func):
    """
    运行前后打印耗时，始终使用 _orig_print，
    不会被 enable_debug() 的 print 劫持影响。
    """
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        # 运行前
        _orig_print(f"⏱ 开始 `{func.__name__}` …")
        t0 = time.time()

        result = func(*args, **kwargs)

        elapsed = time.time() - t0
        # 运行后
        _orig_print(f"⏱ 完成 `{func.__name__}`，耗时：{elapsed:.3f} 秒")
        return result

    return wrapper



#&&% 函数打印消息控制
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --



# ---------- 示例：打印消息屏蔽测试函数 ----------

def ceshi_xiaoxi_dayin():

    stc("测试1")
    print("只打印主函数消息")



#&&% JSON数据读取存放设置
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

# ———————— 全局配置 ————————
SAVE_DIR = r"D:/Myprogramsystem/XT/文件字典信息"
os.makedirs(SAVE_DIR, exist_ok=True)

# 全局缓存所有已经加载的打印字典
PRINT_DICTS: dict[str, dict] = {}

# ———————— 加载指定 DWG 的打印字典 ————————
def com_to_handle(obj):
    """
    如果是 COM 对象，则返回它的 Handle（字符串）；否则原样返回 obj。
    """
    try:
        # 大多数 Dispatch 对象都有 .Handle 属性
        if hasattr(obj, "Handle"):
            return obj.Handle
    except pythoncom.com_error:
        pass
    return obj

def serialize(obj):
    """
    递归地将 dict/list/tuple/COMObject 转为只含基础类型(包括 handle 字符串)的结构。
    """
    if isinstance(obj, dict):
        return {k: serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [serialize(v) for v in obj]
    if isinstance(obj, tuple):
        return tuple(serialize(v) for v in obj)
    # 不是容器的，就尝试转换成 handle
    return com_to_handle(obj)

def save_print_dict_generic(dwg_name: str, bind_dict: dict):
    """
    将 bind_dict 序列化并写入 JSON 文件，键为 dwg_name。
    序列化时自动把所有 COM 对象转换为它们的 Handle。
    """
    ser = serialize(bind_dict)
    path = os.path.join(SAVE_DIR, f"{dwg_name}_打印字典.json")
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(ser, f, ensure_ascii=False, indent=2)
    PRINT_DICTS[dwg_name] = ser
    print(f"✅ 已保存打印字典到 {path}")

def load(dwg_name: str) -> dict:
    """
    从 JSON 文件加载之前保存的打印字典（只含基本类型 & handle 字符串）。
    如果内存中已有，则优先返回内存中的。
    """
    if dwg_name in PRINT_DICTS:
        return PRINT_DICTS[dwg_name]

    path = os.path.join(SAVE_DIR, f"{dwg_name}_打印字典.json")
    if os.path.isfile(path):
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        PRINT_DICTS[dwg_name] = data
        print(f"✅ 已加载打印字典自 {path}")
        return data

    print(f"⚠ 未找到 {path}，返回空字典")
    return {}

# ———————— 辅助：获取当前 DWG 的纯文件名 ————————
def current_dwg_basename() -> str:
    """
    acad.ActiveDocument.Name 可能带路径或带 .dwg，
    取不带扩展的文件名。
    """
    name = acad.ActiveDocument.Name
    base = os.path.splitext(os.path.basename(name))[0]
    return base

def current_dwg_folder():
    """
    获取当前打开的 DWG 文件所在的文件夹路径。
    
    """
    try:
        
        full_path = doc.FullName  # e.g. "D:\\Projects\\MyDrawing.dwg"
        folder_path = os.path.dirname(full_path)
        return folder_path
    except Exception as e:
        print(f"❌ 无法获取当前 DWG 文件夹：{e}")
        return None


#&&% 基础函数


def vtpnt(x, y, z):
    """坐标点转化为浮点数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, z))

def vtobj(obj):
    """转化为对象数组"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj)

def vtFloat(lst):
    """列表转化为浮点数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, lst)
    
def vtInt(lst):
    """列表转化为整数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, lst)

def vtVariant(lst):
    """列表转化为变体"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, lst)

def ConvertArrays2Variant(inputdata, vartype):#例如vartype="Variant"
    import pythoncom
    if vartype == "ArrayofObjects":  # 对象数组
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, inputdata)
    if vartype == "Double":  # 双精度
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, inputdata)
    if vartype == "ShortInteger":  # 短整型
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, inputdata)
    if vartype == "LongInteger":  # 长整型
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I4, inputdata)
    if vartype == "Variant":  # 变体
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, inputdata)
    return outputdata

def vtlist(obj_list):#将com对象列表转换为可用win32数据
    """

    将对象列表转为 VARIANT 类型以供 COM 接口使用

    依赖语句 from win32com.client import VARIANT

    """
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj_list)


warnings.filterwarnings("ignore", category=UserWarning)



#&&% 打开天正start_applicationV9

def start_applicationV9(
    PTH: str = r"C:\Tangent\TArchT20V9",
    max_retries: int = 3,
    retry_delay: float = 2.0
) -> subprocess.Popen | None:
    """
    启动天正 TGStart.exe，失败重试。
    返回 Popen 对象（成功）或 None（失败）。
    """
    exe = os.path.join(PTH, "TGStart.exe")
    for attempt in range(1, max_retries + 1):
        try:
            proc = subprocess.Popen(
                [exe],
                creationflags=DETACHED_PROCESS,
                cwd=PTH
            )
            print("🚀 启动天正CAD 成功")
            return proc
        except Exception as e:
            print(f"第 {attempt} 次启动失败: {e}")
            if attempt < max_retries:
                print(f"等待 {retry_delay:.1f} 秒后重试…")
                time.sleep(retry_delay)
    print(f"已达最大重试次数 ({max_retries})，启动失败。")
    return None




def get_acad_process_id(ming):#获取进程的ID
    for process in psutil.process_iter(attrs=['pid', 'name']):
        if str(ming) in process.info['name'].lower():
            return process.info['pid']
    return None
    
def jingchengshu_wenjian():#查看cad进程数

    CAD_PROCESS_NAME = "acad.exe"

    found_process_i = 0
    for process in psutil.process_iter(['pid', 'name']):
        if process.info['name'] == CAD_PROCESS_NAME:
            found_process_i = found_process_i+1            

    return  found_process_i 
    
def close_all_cad_processes():#关闭所有进程
    max_retries = 3
    for attempt in range(max_retries):
        success = True
        for process in psutil.process_iter(['pid', 'name']):
            if process.info['name'] == "acad.exe":
                try:
                    process.kill()
                    time.sleep(2)
                except Exception as e:
                    print(f"尝试关闭 CAD 进程失败: {e}")
                    success = False
                    break
        if success:
            return
        else:
            print(f"关闭 CAD 进程失败，正在重试... 尝试次数: {attempt + 1}")
            time.sleep(2)
    print("多次尝试关闭 CAD 进程失败，请检查系统。")
    
def close_oldest_cad_process(process_name="acad.exe"):#关闭上一个进程
    cad_processes = [p for p in psutil.process_iter(['pid', 'name', 'create_time']) if p.info['name'] == process_name]
        
    # 检查是否有多个CAD进程
    if len(cad_processes) > 1:
        # 按启动时间排序，最早的进程在前
        oldest_process = sorted(cad_processes, key=lambda p: p.info['create_time'])[0]
            
        try:
            # 关闭最早的进程
            psutil.Process(oldest_process.pid).terminate()
            print(f"已关闭最早的CAD进程，PID: {oldest_process.pid}")
        except Exception as e:
            print(f"关闭进程时出错: {e}")
            pass
    else:
        print("没有多个CAD进程运行。")
        pass


#&&%  cad连接、关闭

def li():#这里的li()不需要写成多次尝试
    global acad, doc, mp, sp
    acad, doc = get_acad_doc()
    mp = doc.ModelSpace
    sp = doc.PaperSpace
    print("当前桌面文件：", doc.Name)
    print("win32已经连接正常—CAD基本操作")
    return True





_RPC_BUSY = (-2147417846, -2147418111)  # 应用程序忙/Call rejected
_RPC_DOWN = (-2147023174,)              # RPC 服务器不可用

def com_retry(fn, retries=30, delay=0.05):
    for _ in range(retries):
        try:
            return fn()
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if code in _RPC_BUSY + _RPC_DOWN:
                time.sleep(delay); continue
            raise
    return fn()

def _coinit_once():
    try: pythoncom.CoInitialize()
    except pythoncom.error: pass

def ensure_typelib_from_running():
    """从正在运行的 AutoCAD 直接生成 makepy（不依赖注册表）"""
    _coinit_once()
    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
    tlb, _ = app._oleobj_.GetTypeInfo().GetContainingTypeLib()
    makepy.GenerateFromTypeLibSpec(tlb)

def get_acad_doc(max_wait=8.0):
    _coinit_once()
    t0 = time.time()
    app = None
    while True:
        try:
            app = win32.gencache.EnsureDispatch("AutoCAD.Application")
            doc = app.ActiveDocument
            _ = doc.Name
            return app, doc
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if (code in _RPC_BUSY + _RPC_DOWN) and (time.time()-t0 < max_wait):
                time.sleep(0.05); continue
            # 自愈：若无文档则新建
            try:
                if app is None:
                    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
                doc = app.Documents.Add()
                _ = doc.Name
                return app, doc
            except Exception:
                pass
            raise

class _ComLiveProxy:
    """把 getter 包成“活体”对象：每次访问都拿最新/可用的 COM 实例"""
    def __init__(self, getter): self._getter = getter
    def __getattr__(self, name):
        obj = self._getter()
        return getattr(obj, name)
    # 可选：允许直接作为可迭代或 bool 使用时更自然
    def __dir__(self): return dir(self._getter())

# ===== 你新的 li()（装载代理），兼容旧写法 =====


def li_new():#20250924新版本
    """
    连接/修复 CAD 会话，并把 acad/doc/mp/sp 设置为“活体代理”。
    原有大量函数里继续写 acad./doc./mp./sp. 即可，但它们会自动取到当前有效对象。
    """
    global acad, doc, mp, sp
    # 先确保 makepy 存在（CastTo/强类型更稳）
    try:
        ensure_typelib_from_running()
    except Exception:
        pass

    # 安装代理（getter 内部包含重试与自愈）
    acad = _ComLiveProxy(lambda: get_acad_doc()[0])
    doc  = _ComLiveProxy(lambda: get_acad_doc()[1])
    mp   = _ComLiveProxy(lambda: get_acad_doc()[1].ModelSpace)
    sp   = _ComLiveProxy(lambda: get_acad_doc()[1].PaperSpace)

    # 轻量校验 & 提示
    name = com_retry(lambda: doc.Name)
    print("当前桌面文件：", name)
    print("win32已经连接正常—CAD基本操作")
    return True
    



"""
@require_doc，它最强大的地方在于自动跟随cad文件的切换，跟踪激活文件？更快的_.doc=get_cad_doc()模式则是确定没有文件切换时更快




"""

#装饰器动态模式

def require_doc(fn):
    def wrapper(*args, doc=None, **kw):
        if doc is None:
            _, doc = get_acad_doc()
        return fn(*args, doc=doc, **kw)
    return wrapper

def rpc_safe(fn):
    def wrapper(*a, **k):
        return com_retry(lambda: fn(*a, **k))
    return wrapper

#只读类函数模板

@require_doc
@rpc_safe
def read_active_doc_info(doc=None):
    """示例：读取当前图名、空间实体数量"""
    name = com_retry(lambda: doc.Name)
    msp  = doc.ModelSpace
    count = com_retry(lambda: msp.Count)
    return {"name": name, "model_count": count}

#修改类函数模板（带撤销、失败也能收尾
@require_doc
def move_selected_by(dx, dy, dz=0.0, doc=None):
    """示例：交互拾取若干对象并平移"""
    ents = pmxz(prompt="\n选择要平移的对象，回车结束：")
    if not ents:
        print("未选择对象"); return 0

    com_retry(lambda: doc.StartUndoMark())
    moved = 0
    try:
        for e in ents:
            com_retry(lambda e=e: e.Move((0,0,0), (dx,dy,dz)))
            moved += 1
    finally:
        com_retry(lambda: doc.EndUndoMark())
    print(f"✅ 已平移 {moved} 个对象")
    return moved

#命令行交互模板（SendCommand + 等待）
"""
主要是为了稳定性和可控性，区别在于：

直接用 doc.SendCommand()

最快，直接把字符串送进 CAD 命令行。

但要自己加 \n (chr(13))，否则命令不会生效。

没有等待/确认逻辑，CAD 还没执行完，你的 Python 可能就继续往下跑了。

封装过的模板（SendCommand + 等待）

会在 SendCommand 后自动 time.sleep(x) 或循环检测 CAD 是否空闲（如 While acad.GetAcadState().IsQuiescent == False）。

可以加日志：执行了什么命令、耗时多久。

遇到 CAD 忙碌 / 拒绝调用时，可以自动重试。

一般还会做“换行清理”，避免上一条命令残留影响下一条。

返回结构：尽量返回明确的字典/元组/数量，便于上层组合使用。

异常：普通 COM 波动交给 com_retry；真正的逻辑错误（类型不符、数据不全）直接抛 RuntimeError 或返回 None 并打印友好信息。





"""

def send_cmd(cmd: str, wait_s: float = 0.3):
    """安全发送命令并小等待，避免命令堆积"""
    _, doc = get_acad_doc()
    doc.SendCommand(cmd if cmd.endswith("\n") else (cmd + "\n"))
    time.sleep(wait_s)

@require_doc
def zoom_window(x1, y1, x2, y2, pad_ratio=0.1, doc=None):
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    pad = pad_ratio * ((x_hi-x_lo + y_hi-y_lo)/2.0)
    send_cmd(f"_.ZOOM\n_W\n{x_lo-pad},{y_lo-pad}\n{x_hi+pad},{y_hi+pad}\n", wait_s=0.6)



#选择 + Cast + 读取属性模板

@require_doc
def pick_line_and_report(doc=None):
    """交互选直线并输出起终点、长度"""
    ents = ss_select("onscreen", filter_types=[0], filter_data=["LINE"], autocast=True,
                     prompt="\n请选择一根直线：")
    if not ents:
        print("未选到直线"); return None
    ln = ents[0]  # IAcadLine（autocast 已转换）
    sp = com_retry(lambda: ln.StartPoint)
    ep = com_retry(lambda: ln.EndPoint)
    L  = com_retry(lambda: ln.Length)
    info = {"start": tuple(sp), "end": tuple(ep), "length": L}
    print(info)
    return info


#批处理模板（不交互、按过滤批量处理）

"""
🔹 为什么我单独提出“批处理模板”？

因为光有 for 循环 还不够，实际在 CAD 中会遇到一些坑：

撤销保护
如果你循环 1000 个对象，用户后悔时，Ctrl+Z 可能要点 1000 次。
→ 用 doc.StartUndoMark() / doc.EndUndoMark() 包起来，只需撤销一次。

RPC 错误 / CAD 忙
在循环里有时会报 RPC 服务器不可用，所以每个对象的属性访问最好包一层 com_retry。

性能
如果每次循环都重新 get_acad_doc()，会很慢。
→ 批处理模板里只获取一次 doc，然后在循环中直接用。

安全清理
SelectionSet 要记得 Delete()，否则下一次可能报错。

🔹 一个最简单的批处理例子
@require_doc
def batch_move_all_lines_to_layer0(doc=None):
    
    # 1) 选择
    ents = ss_select("all", filter_types=[0], filter_data=["LINE"], autocast=True)
    if not ents:
        print("❌ 没有找到直线"); return 0

    # 2) 撤销块
    com_retry(lambda: doc.StartUndoMark())
    moved = 0
    try:
        # 3) 循环处理
        for e in ents:
            com_retry(lambda: setattr(e, "Layer", "0"))
            moved += 1
    finally:
        # 4) 结束撤销块
        com_retry(lambda: doc.EndUndoMark())

    print(f"✅ 批处理完成，共移动 {moved} 条直线到图层 0")
    return moved


"""



@require_doc
def flatten_all_circles_to_z0(doc=None):
    """把所有圆的 Z 置 0（示例）"""
    circles = ss_select("all", filter_types=[0], filter_data=["CIRCLE"], autocast=True)
    if not circles:
        print("未找到圆"); return 0
    com_retry(lambda: doc.StartUndoMark())
    n = 0
    try:
        for c in circles:
            # 移到 Z=0（Circle 没有 Elevation，就 Transform）
            sp = com_retry(lambda: c.Center)
            com_retry(lambda: c.Move(sp, (sp[0], sp[1], 0.0)))
            n += 1
    finally:
        com_retry(lambda: doc.EndUndoMark())
    print(f"✅ 处理圆 {n} 个")
    return n


#&&&% RGB色彩

def aci_to_rgb(ci: int):
    # 只给出常用 ACI 的近似色；其余返回 None
    table = {
        1:(255,0,0), 2:(255,255,0), 3:(0,255,0), 4:(0,255,255),
        5:(0,0,255), 6:(255,0,255), 7:(255,255,255)
    }
    return table.get(int(ci)) if ci is not None else None
def get_entity_rgb(ent):
    """返回 (rgb_tuple_or_None, source_str)"""
    # 1) 先看 TrueColor 的方式
    tc = ent.TrueColor                 # IAcadAcCmColor
    method = getattr(tc, "ColorMethod", None)
    # acColorMethodByRGB 大多为 3（不同版本枚举值可能不同，但逻辑相同）
    if method == 3 or (hasattr(tc, "Red") and (tc.Red or tc.Green or tc.Blue)):
        return (tc.Red, tc.Green, tc.Blue), "ByRGB(TrueColor)"

    # 2) 看 ColorIndex / Color（ACI / ByLayer / ByBlock）
    ci = getattr(ent, "ColorIndex", None)
    if ci is None:
        ci = getattr(ent, "Color", None)

    # ByLayer
    if ci == 256:
        layer_name = ent.Layer
        layer = ent.Document.Layers.Item(layer_name)
        ltc = layer.TrueColor
        return (ltc.Red, ltc.Green, ltc.Blue), f"ByLayer({layer_name})"

    # ByBlock
    if ci == 0:
        # 真正的显示颜色取决于引用它的块参照/上级容器
        return None, "ByBlock(需根据上级块参照解析)"

    # 3) 纯 ACI（1..255）
    if isinstance(ci, int) and 1 <= ci <= 255:
        rgb = aci_to_rgb(ci)
        return (rgb if rgb else None), f"ACI({ci})"

    # 兜底
    return None, "Unknown"


"""
lj=pmxz()
pl=lj[0]
rgb, src = get_entity_rgb(pl)
print("RGB:", rgb, "| 来源:", src)
RGB: (255, 0, 0) | 来源: ByRGB(TrueColor)
"""

    
def  guanbi_cad_doc():#关闭所有cad文件

        
    shuliang_cad = acad.Documents.Count

    print("当前桌面打开的cad文件数量:",shuliang_cad)

    for i in range(0,shuliang_cad):

        CourentDoc = acad.ActiveDocument

        CourentDoc.Close()

        li()
    
def guanbi_cad_one():#关闭所有cad文件但保留一个并连接
    shuliang_cad = acad.Documents.Count
    print("当前桌面打开的cad文件数量:", shuliang_cad)

    if shuliang_cad > 1:
        for i in range(shuliang_cad - 1):  # 保留最后一个文件
            courent_doc = acad.ActiveDocument
            courent_doc.Close()

        # 重新连接到最后一个文件
        if not li():
            print("重新连接到最后一个文件失败，请检查系统。")
    else:
        print("只有一个CAD文件打开，无需关闭其他文件。")

def st():#启动或连上CAD开始工作
    """
    归零重新归于标准态的处理

    """
    cad_process_count = jingchengshu_wenjian()

    if cad_process_count > 1:
        close_all_cad_processes()
        start_applicationV9()
        li()#生成全局变量acad,mp,doc
    elif cad_process_count == 0:
        start_applicationV9()
        li()#生成全局变量acad,mp,doc
    elif cad_process_count == 1:
        if not li():
            close_all_cad_processes()
            start_applicationV9()



def huifu_xitong():#启动或连上CAD开始工作
    """
    归零重新归于标准态的处理

    """
    close_all_cad_processes()

    start_applicationV9()




#&&&%  * 重复多次调用函数

def chongfu_caozuo(
    Fx,
    dwg_instance=None,
    args: tuple=(),
    kwargs: dict=None,
    max_retries: int=3,
    failure_value=None
):
    """
    对指定函数/方法进行重复调用，直到成功或耗尽重试次数。

    :param Fx:             待调用的函数对象（或方法），不要在这里写 Fx()
    :param dwg_instance:   如果要调用类的方法，就把实例传进来；否则为 None
    :param args:           位置参数，按顺序传给 Fx
    :param kwargs:         关键字参数 dict，传给 Fx
    :param max_retries:    最多尝试次数（>=1）
    :param failure_value:  全部尝试失败时的返回值

    :return: (result, attempts, error)
        - result:   Fx 的返回值，或 failure_value
        - attempts: 成功时为那次尝试编号（1-based），失败时等于 max_retries
        - error:    最后一次捕获到的 Exception 实例；成功时为 None
    """
    if kwargs is None:
        kwargs = {}

    last_exception = None

    for attempt in range(1, max_retries + 1):
        try:
            # 决定如何调用 Fx
            if dwg_instance is not None and not getattr(Fx, "__self__", None):
                # Fx 很可能是类中定义的函数，但还未绑定到实例
                result = Fx(dwg_instance, *args, **kwargs)
            else:
                # Fx 已经是普通函数或已绑定的方法
                result = Fx(*args, **kwargs)

            # 成功：返回 (结果, 尝试次数, None)
            return result, attempt, None

        except Exception as e:
            last_exception = e
            if attempt < max_retries:
                print(f"⚠ 操作出错，正在第 {attempt} 次重试…")
                # （可选）在重试前做些环境刷新、移动窗口等
                try:    li()  # 示例：重新排列 CAD/IDLE 窗口
                except: pass
                time.sleep(2)
            else:
                print("❌ 多次尝试均失败，请检查环境或参数是否正确。")
                # 返回 (failure_value, 用尽的尝试次数, 最后一个异常)
                return failure_value, attempt, last_exception


#  自动计时装饰器

"""
用法
@simple_timer
def heavy_func():
    ...

heavy_func()

"""
def simple_timer(func):
    def wrapper(*args, **kwargs):
        t0 = time.time()
        result = func(*args, **kwargs)
        t1 = time.time()
        print(f"⏱️ {func.__name__} 耗时：{t1 - t0:.4f} 秒")
        return result
    return wrapper            









#&&&&%% 第二部分 列表处理


"""
该模块研究列表的基本问题，特别是com对象列表 


有一个针对全部程序系统的基本文件处理文件Basic_oper.py，例如字符串，列表，集合的操作等等，便于我们展开别的专项工作时引用而不必重复生成。

在每个专项工作文件中，仍然有一套自己的基本处理文件，放在脚本的最前面，这两者并不重复，而且很合理。我们将在专项工作中集中这些基本处理，

将来再移植到Basic_oper.py。虽然函数重复出现在不同的脚本，但并不会混乱，也不会降低效率。这反映了内在的本质规律：Basic_oper.py

不是专门开发研究的成果，而是来自各个专项工作的总结。


"""


# 按com实体对象中提取的坐标排序


def sort_tuples(lst,cha_Y =2000):#对列表按插入点xy坐标排序
    
    """
    这是很有用的一个双值排序函数，对于COM对象，可以先将其转换为元组，即可使用这个函数

    它的价值在于，很容易拓展到n值排序
    """
    
    # 先按照m[1]降序排序
    lst.sort(key=lambda x: -x[2][1])

    i = 0
    while i < len(lst) - 1:
        j = i + 1
        # 查找所有m[1]差距在chaY以内的元素
        while j < len(lst) and abs(lst[i][2][1] - lst[j][2][1]) < cha_Y:
            j += 1
        
        # 如果找到了m[1]值相近的元素，根据m[0]值进行排序
        if j - i > 1:
            lst[i:j] = sorted(lst[i:j], key=lambda x: x[2][0])
        
        i = j

    return lst

def multi_dim_tolerance_sort(lst, key_index=2, tolerances=[10000, 1000, 0]):#高维排序
    """
    对 lst 列表中的元组按多维坐标字段排序，考虑每个维度的容差进行逐层排序。

    参数：
        lst: [(id, name, (x, y, z)), ...]
        key_index: 坐标在元组中的索引（默认是第3项，即元组[2]）
        tolerances: 每个维度允许的容差，例如 [Z差, Y差, X差]

    返回：
        排好序的新列表
    """
    # 最外层排序：按最高维降序排（Z从上往下）
    dim = len(tolerances)
    lst.sort(key=lambda x: -x[key_index][0])  # Z 倒序

    def recursive_sort(sublist, level):
        if level >= dim - 1:
            # 最后一级直接按该维升序
            return sorted(sublist, key=lambda x: x[key_index][level + 1])
        
        i = 0
        while i < len(sublist) - 1:
            j = i + 1
            while j < len(sublist) and abs(sublist[i][key_index][level] - sublist[j][key_index][level]) < tolerances[level]:
                j += 1
            if j - i > 1:
                sublist[i:j] = recursive_sort(sublist[i:j], level + 1)
            i = j
        return sublist

    return recursive_sort(lst, 0)


def get_ll_pt(ent):#提取函数，对象左下角点
    minpt, _ = ent.GetBoundingBox()
    return minpt[0], minpt[1],0

def get_center(ent):#提取函数，中心点
    minpt, maxpt = ent.GetBoundingBox()
    return ((minpt[0]+maxpt[0])/2, (minpt[1]+maxpt[1])/2)

def sort_entities_by_position( entity_list, extract_func, cha_Y=2000):#对com对象按提取坐标分别沿y,x方向排序
        """
        对实体列表根据其坐标（通过 extract_func 获取）进行排序：
        - 先按 Y 值降序（从上到下）
        - Y 值接近（差值 < cha_Y）者再按 X 值升序（从左到右）

        参数：
            entity_list: COM 实体对象列表
            extract_func: 提取坐标函数，返回 (x, y) 元组的函数即可
            cha_Y: 同一行判定的 Y 方向容差

        返回：按坐标顺序排列的新实体对象列表

        调用示例

        sorted_objs = sort_entities_by_position(LB, extract_func=get_ll_pt)
        
        """
        triples = [(ent, *extract_func(ent)) for ent in entity_list]

        # 按 Y 值降序排列
        triples.sort(key=lambda t: -t[2])

        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])  # 按 X 升序
            i = j

        return [t[0] for t in triples]

def get_line_start(ent):
    """
    提取一条直线的起点 (x, y)
    ent: AcDbLine 或类似对象，具有 .StartPoint 属性
    """
    pt = ent.StartPoint   # 假设是一个 (x, y, z) 或 [x, y, z]
    return pt[0], pt[1]

#&&% * 对列表实体进行从上到下、从左到右的排序

def sort_coms_by_llcorner(com_list, cha_Y=2000):
    """
    按 BoundingBox 左下角坐标排序：
      · 先按 y 降序（越大越靠前，即自上而下）
      · 同一行(Δy<cha_Y)内按 x 升序（自左向右）
    """
    wrapped = []
    for ent in com_list:
        try:
            p1, _ = ent.GetBoundingBox()      # p1 已是左下
            x_ll, y_ll = p1[0], p1[1]
        except Exception:
            x_ll = y_ll = float('-inf')       # 取不到的一律放最后
        wrapped.append((ent, x_ll, y_ll))     # (实体, x, y)

    # 先按 y 降序
    wrapped.sort(key=lambda t: -t[2])

    i = 0
    while i < len(wrapped) - 1:
        j = i + 1
        while j < len(wrapped) and abs(wrapped[i][2] - wrapped[j][2]) < cha_Y:
            j += 1
        # 行内再按 x 升序
        if j - i > 1:
            wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: t[1])
        i = j

    return [ent for ent, _, _ in wrapped]




def sort_coms_by_rbcorner(com_list, *, cha_X=100):
    """
    竖向图框（或已整体旋转 -90° 的图纸）使用 ——  
    · 先按列：x₁ 降序（右 → 左）  
    · 列内：  y₂ 降序（上 → 下）

    参数
    ----
    com_list : list[COM object]
        文字实体列表
    cha_X    : float
        判定“同一列”的容差，单位与图纸坐标一致
    """
    wrapped = []
    for ent in com_list:
        try:
            (x1, _, _), (_, y2, _) = ent.GetBoundingBox()  # x1 = 左, y2 = 上
        except Exception:
            x1 = y2 = float("-inf")                        # 失败的排最后
        wrapped.append((ent, x1, y2))

    # — ① 按 x₁ 降序：最右列在前 —
    wrapped.sort(key=lambda t: -t[1])

    # — ② 同列内再按 y₂ 降序：上 → 下 —
    i = 0
    while i < len(wrapped) - 1:
        j = i + 1
        while j < len(wrapped) and abs(wrapped[i][1] - wrapped[j][1]) < cha_X:
            j += 1
        wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: -t[2])
        i = j

    return [t[0] for t in wrapped]


def sort_coms_by_llcorner_custom(objs, tol_x=100):
    """
    按左下角 x, y 坐标对 COM 对象列表 objs 排序：
      1. 先按 x 升序分组：相邻 x 差 ≤ tol_x 视为同一组
      2. 每组内按 y 降序排列（y 大的排前）
      3. 最终把各组按它们的 x 升序依次拼接

    参数
    ----
    objs : List[COMObject]
        要排序的 COM 对象列表，每个对象必须支持 GetBoundingBox()
    tol_x : float
        x 方向上的容差，小于等于此值的 x 差视为同组

    返回
    ----
    List[COMObject]
        排序后的对象列表
    """
    # 提取 (obj, llx, lly)
    items = []
    for e in objs:
        ll, ur = e.GetBoundingBox()
        llx, lly, _ = ll
        items.append((e, llx, lly))

    # 按 x 升序初排
    items.sort(key=lambda t: t[1])

    # 分组：相邻 x 差 ≤ tol_x 的归一组
    clusters = []
    rep_x, current = items[0][1], [items[0]]
    for item in items[1:]:
        _, x, _ = item
        if abs(x - rep_x) <= tol_x:
            current.append(item)
        else:
            clusters.append((rep_x, current))
            rep_x, current = x, [item]
    clusters.append((rep_x, current))

    # 按组的 x 升序排好后，组内按 y 降序，再平铺
    result = []
    for _, group in sorted(clusters, key=lambda c: c[0]):
        group_sorted = sorted(group, key=lambda t: t[2], reverse=True)
        result.extend([t[0] for t in group_sorted])

    return result


def sort_coms_by_center(objs, tol_x=100):
    """
    按外包盒中心坐标对 COM 对象列表 objs 排序：
      1. 先计算每个对象的 bbox 中心 (cx, cy)
      2. 按 cx 升序初排序
      3. 将相邻 cx 差 ≤ tol_x 的对象归同一组
      4. 各组按 cy 降序排列（cy 大的排前）
      5. 最后按组的 cx 升序依次拼接各组内对象

    参数
    ----
    objs : List[COMObject]
        要排序的 COM 对象列表，每个对象必须支持 GetBoundingBox()
    tol_x : float
        x 方向上的容差，小于等于此值的 cx 差视为同组

    返回
    ----
    List[COMObject]
        排序后的对象列表
    """
    # 1. 提取 (obj, cx, cy)
    items = []
    for e in objs:
        ll, ur = e.GetBoundingBox()
        cx = (ll[0] + ur[0]) / 2.0
        cy = (ll[1] + ur[1]) / 2.0
        items.append((e, cx, cy))

    if not items:
        return []

    # 2. 按 cx 升序初排
    items.sort(key=lambda t: t[1])

    # 3. 分组：相邻 cx 差 ≤ tol_x 的归一组
    clusters = []
    rep_x, current = items[0][1], [items[0]]
    for item in items[1:]:
        _, x, _ = item
        if abs(x - rep_x) <= tol_x:
            current.append(item)
        else:
            clusters.append((rep_x, current))
            rep_x, current = x, [item]
    clusters.append((rep_x, current))

    # 4. 各组内部按 cy 降序排列，再平铺
    result = []
    for rep_x, group in sorted(clusters, key=lambda c: c[0]):
        group_sorted = sorted(group, key=lambda t: t[2], reverse=True)
        result.extend([t[0] for t in group_sorted])

    return result


##对列表实体进行正序或逆序编号

def number_entities_by_order(entity_list, prefix="", start=1, k=0):
    """
    对排序好的 COM 实体对象列表进行编号。

    参数：
        entity_list: 实体对象列表
        prefix: 编号前缀，默认为空字符串
        start: 编号起始数值，默认为 1
        k: 排序方向控制变量：
            - k = 0 正序编号
            - k = 1 逆序编号

    返回：
        编号字符串列表（如 ["1", "2", "3"] 或 ["图1", "图2", "图3"]）
    """
    n = len(entity_list)
    index_list = range(n) if k == 0 else reversed(range(n))
    
    result = []
    for i, idx in enumerate(index_list):
        label = f"{prefix}{start + i}"
        result.append(label)
    
    return result


# 重复操作列表对象

def apply_to_each1(obj_list,  action_func):#重复操作列表对象
    """
    对 obj_list 中的每个对象，
    传入 action_func 中处理。

    参数：
        obj_list: 对象列表
        
        action_func: 用于处理提取结果的函数（如 srhd）
    """
    for obj in obj_list:
        
        action_func(obj)
        
def apply_to_each2(obj_list, extract_func, action_func):#双层嵌套重复操作列表对象
    """
    对 obj_list 中的每个对象，先通过 extract_func 提取值，
    再将该值传入 action_func 中处理。

    参数：
        obj_list: 对象列表
        extract_func: 用于提取 (x, y) 或其他值的函数
        action_func: 用于处理提取结果的函数（如 srhd）
    """
    for obj in obj_list:
        value = extract_func(obj)
        action_func(value)

#&&&&%% 第三部分 文件夹文件处理


"""
该模块是有关一般性的文件夹、文件上的操作 

"""
#&&% 确保文件被删除
def ensure_file_absent(save_path: str, ty: float = 1.0) -> None:
    """
    确保指定路径的文件不存在。如果存在，则删除并等待 ty 秒。

    :param save_path: 带路径的文件名，例如 "D:/output/result.txt"
    :param ty: 删除后等待的秒数，默认为 1 秒
    """
    try:
        if os.path.isfile(save_path):
            os.remove(save_path)
            # 等待 ty 秒，以确保文件系统完成删除操作
            time.sleep(ty)
            print(f"✅ 已删除文件：{save_path}，并等待了 {ty} 秒")
        else:
            print(f"ℹ 文件不存在，无需删除：{save_path}")
    except Exception as e:
        print(f"❌ 删除文件时出错：{e}")


def traverse_with_os_walk(root_dir: str):
    """
    遍历 root_dir 及其所有子目录，打印每个目录和文件的完整路径。
    """
    for dirpath, dirnames, filenames in os.walk(root_dir):
        # dirpath: 当前遍历到的目录路径
        print(f"Directory: {dirpath}")
        for dirname in dirnames:
            print(f"  Sub-dir: {os.path.join(dirpath, dirname)}")
        for filename in filenames:
            print(f"  File   : {os.path.join(dirpath, filename)}")



def find_files_with_extensions(directory, extensions):
    #找到以[".dwg"]结尾的文件directory为文件夹及其路径，extensions为后缀或中间位置字符列表
    matching_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if any(file.lower().endswith(ext) for ext in extensions):
                matching_files.append(os.path.join(root, file))
    return matching_files

def get_filename_without_extension(FileandPath):
    # 从完整路径中获取文件名（包含扩展名）
    filename_with_extension = os.path.basename(FileandPath)
    
    # 分离文件名和扩展名
    filename_without_extension, _ = os.path.splitext(filename_with_extension)

    return filename_without_extension


def delete_files_with_patterns(folder_path, patterns):
    """
    删除文件夹中符合指定模式的文件。

    Args:
        folder_path (str): 文件夹的路径。
        patterns (list): 包含要匹配的模式的列表，例如["_t7", "_t3", "_t8"]。

    Returns:
        None
    """
    # 遍历文件夹中的所有文件
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        # 检查文件名是否包含任何一个指定的模式
        if any(pattern in filename for pattern in patterns):
            # 如果符合条件，就删除该文件
            os.remove(file_path)
            print(f"Deleted file: {filename}")
            pass

    print("File deletion completed.")
    pass


#确保文件夹中名字含有特殊字符的文件被清空

def clear_files_with_prefix(folder: str, filename_prefix: str = "区域导出", delay: float = 0.5):
    """
    清除指定文件夹中所有文件名包含给定前缀的文件。若文件正在被占用或删除失败，会尝试重试一次。

    :param folder:           要清理的文件夹路径
    :param filename_prefix:  需要匹配的文件名前缀（只要文件名中包含该字符串，就会被删除）
    :param delay:            删除失败时的等待时间（秒），默认 0.5 秒
    """
    if not os.path.isdir(folder):
        print(f"❌ 目标路径不是有效文件夹：{folder}")
        return

    # 列出文件夹中所有条目
    entries = os.listdir(folder)
    # 过滤出文件名中包含指定前缀的文件
    to_delete = [fname for fname in entries if filename_prefix in fname and os.path.isfile(os.path.join(folder, fname))]

    if not to_delete:
        print(f"ℹ️ 文件夹中未发现文件名包含 “{filename_prefix}” 的文件。")
        return

    for fname in to_delete:
        full_path = os.path.join(folder, fname)
        try:
            os.remove(full_path)
            print(f"✅ 已删除：{fname}")
        except Exception as e:
            print(f"⚠ 删除失败（第一次尝试）：{fname}，错误：{e}，稍后重试……")
            time.sleep(delay)
            # 再试一次
            try:
                os.remove(full_path)
                print(f"✅ 重试成功删除：{fname}")
            except Exception as e2:
                print(f"❌ 再次删除仍失败：{fname}，错误：{e2}")



def find_files_with_string(directory, search_string):
    #找到文件夹中含有指定字符串的文件
    matched_files = []
    for file in os.listdir(directory):
        if search_string in file:
            matched_files.append(file)
    return matched_files

#使路径名和文件名合并后合乎预期
def join_paths(p1, p2):
    # 使用 os.path.join 合并路径
    result = os.path.join(p1, p2)
    # 替换反斜杠为正斜杠
    return result.replace("\\", "/")



#&&&&%% 第四部分 选择方法 


# ===================== 前置工具（统一处理） =====================

# —— 通用重试：吞“忙/拒绝/RPC down”
_RPC_BUSY = (-2147417846, -2147418111)
_RPC_DOWN = (-2147023174,)

def com_retry(fn, retries=30, delay=0.05):
    for _ in range(retries):
        try:
            return fn()
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if code in _RPC_BUSY + _RPC_DOWN:
                time.sleep(delay); continue
            raise
    return fn()

# —— 现取现用：获取当前激活文档（无文档则新建）
def get_acad_doc(max_wait=8.0):
    try: pythoncom.CoInitialize()
    except pythoncom.error: pass
    t0 = time.time()
    app = None
    while True:
        try:
            app = win32.gencache.EnsureDispatch("AutoCAD.Application")
            doc = app.ActiveDocument
            _ = doc.Name
            return app, doc
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if (code in _RPC_BUSY + _RPC_DOWN) and (time.time()-t0 < max_wait):
                time.sleep(0.05); continue
            # 无打开图 → 新建
            try:
                if app is None:
                    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
                doc = app.Documents.Add(); _ = doc.Name
                return app, doc
            except Exception:
                pass
            raise

# —— 过滤数组（避免与你库已有 vtInt/vtVariant 冲突）
def to_vt_int(seq):
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, list(seq))

def to_vt_variant(seq):
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, list(seq))

# —— 常见对象自动 Cast：拿到专接口（StartPoint/Coordinates/Contents…）
_CAST_MAP = {
    # 基础几何
    "AcDbLine":"IAcadLine", "AcDbCircle":"IAcadCircle", "AcDbArc":"IAcadArc","AcDbPoint":"IAcadPoint",
    "AcDbEllipse":"IAcadEllipse", "AcDbSpline":"IAcadSpline",
    # 多段线
    "AcDbPolyline":"IAcadLWPolyline", "AcDb2dPolyline":"IAcadPolyline", "AcDb3dPolyline":"IAcad3DPolyline",
    # 文字
    "AcDbText":"IAcadText", "AcDbMText":"IAcadMText",
    # 块/属性
    "AcDbBlockReference":"IAcadBlockReference",
    "AcDbAttribute":"IAcadAttributeReference", "AcDbAttributeDefinition":"IAcadAttribute",
    # 引线/标注（常用）
    "AcDbLeader":"IAcadLeader", "AcDbMLeader":"IAcadMLeader",
    "AcDbDimension":"IAcadDimension", "AcDbAlignedDimension":"IAcadDimAligned",
    "AcDbRotatedDimension":"IAcadDimRotated", "AcDbRadialDimension":"IAcadDimRadial",
    "AcDbDiametricDimension":"IAcadDimDiametric", "AcDbArcDimension":"IAcadDimArc",
    "AcDb3PointAngularDimension":"IAcadDim3PointAngular", "AcDb2LineAngularDimension":"IAcadDim2LineAngular",
    "AcDbOrdinateDimension":"IAcadDimOrdinate",
    # 其它
    "AcDbHatch":"IAcadHatch", "AcDbTable":"IAcadTable",
}

def _maybe_cast(ent):
    try:
        name = com_retry(lambda: ent.ObjectName)
        iface = _CAST_MAP.get(name)
        if iface:
            try: return CastTo(ent, iface)
            except Exception: return ent
        return ent
    except Exception:
        return ent

# —— 统一选择器：all/window/crossing/onscreen
# mode: "all" | "window" | "crossing" | "onscreen"
def ss_select(mode="all", p1=None, p2=None, filter_types=None, filter_data=None, autocast=True, prompt=None):
    _, doc = get_acad_doc()
    ss_name = f"SS_{uuid.uuid4().hex[:8]}"
    # 清理旧名 + 新建
    try: doc.SelectionSets.Item(ss_name).Delete()
    except Exception: pass
    ss = doc.SelectionSets.Add(ss_name)

    try:
        if mode == "onscreen":
            if prompt:
                try: doc.Utility.Prompt(prompt)
                except Exception: pass
            ss.SelectOnScreen()
        elif mode in ("all","window","crossing"):
            selmode = 5 if mode=="all" else (3 if mode=="window" else 4)
            ss.Select(
                selmode,
                p1 if mode!="all" else None,
                p2 if mode!="all" else None,
                to_vt_int(filter_types) if filter_types else None,
                to_vt_variant(filter_data) if filter_data else None
            )
        else:
            raise ValueError("未知选择模式")

        n = com_retry(lambda: ss.Count)
        items = [com_retry(lambda i=i: ss.Item(i)) for i in range(n)]
        if autocast:
            items = [_maybe_cast(e) for e in items]
        return items
    finally:
        try: ss.Delete()
        except Exception: pass


# ===================== 选择函数（重写版） =====================

# 1) 选择“包围框覆盖指定点”的对象
def select_entities_through_point(p, tol=0.1):
    """
    查找所有其 BoundingBox 覆盖点 p 的图元。
    不依赖全局 mp；自动重试；失败的对象跳过。
    """
    _, doc = get_acad_doc()
    msp = doc.ModelSpace
    px, py, _ = p
    selected = []
    # 遍历模型空间；大图慎用（必要时可改为小窗口交叉选择）
    for ent in msp:
        try:
            p1, p2 = com_retry(lambda: ent.GetBoundingBox())
            x_min, x_max = sorted([p1[0], p2[0]])
            y_min, y_max = sorted([p1[1], p2[1]])
            x_min -= tol; x_max += tol; y_min -= tol; y_max += tol
            if x_min <= px <= x_max and y_min <= py <= y_max:
                selected.append(_maybe_cast(ent))
        except Exception:
            continue
    print(f"✅ 共找到 {len(selected)} 个对象经过点 {p}")
    return selected


# 2) 按图层选择（支持单个或列表）
def select_tuceng(layer_names, max_retries=5, delay=0.5, autocast=True):
    if isinstance(layer_names, str):
        layers = [layer_names]
    else:
        layers = list(layer_names)
    last = None
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(
                mode="all",
                filter_types=[8],                # 8 = Layer
                filter_data=[layers if len(layers)>1 else layers[0]],
                autocast=autocast
            )
            print(f"✅ 第 {k} 次尝试：选到图层 {layers} 上 {len(ents)} 个对象")
            return ents
        except Exception as e:
            last = e
            print(f"⚠ 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc()
                doc.SendCommand("RE\nZ\nE\n")
            except Exception:
                pass
            time.sleep(delay)
    print(f"❌ 重试 {max_retries} 次后仍失败：{last!r}")
    return []

def stc(layer_names, **kwargs):
    return select_tuceng(layer_names, **kwargs)


# 3) 选择所有块（INSERT）
def select_kuai(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(
                mode="all",
                filter_types=[0],               # 0 = 实体类型
                filter_data=["INSERT"],
                autocast=autocast
            )
            print(f"✅ select_kuai 成功（第 {k} 次），耗时 {time.time()-t0:.3f}s，共 {len(ents)} 个块")
            return ents
        except Exception as e:
            last = e
            print(f"⚠ select_kuai 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"❌ select_kuai 在 {max_retries} 次尝试后仍失败：{last!r}")
    return []


# 4) 选择所有 TEXT
def select_text(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["TEXT"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

# 5) 选择所有 MTEXT
def select_mtext(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["MTEXT"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

# 6) PUB_TEXT 图层上的“天正文字”分类（按 ObjectName）
def select_pub_text_entities():
    LAYER_NAME = "PUB_TEXT"
    ents = select_tuceng(LAYER_NAME, autocast=False)  # 天正对象通常是代理/专有类，先别 Cast
    tdb_texts, tdb_mtexts = [], []
    for ent in ents:
        name = getattr(ent, "ObjectName", None) or getattr(ent, "EntityName", "")
        if name == "TDbText":
            tdb_texts.append(ent)
        elif name == "TDbMText":
            tdb_mtexts.append(ent)
    return tdb_texts, tdb_mtexts

def collect_all_texts():
    """
    同时收集天正与原生 CAD 文本，并统一图层为 'PUB_TEXT'。
    返回：tianzheng_texts / tianzheng_mtexts / cad_texts / cad_mtexts
    """
    LAYER_NAME = "PUB_TEXT"
    tz_texts, tz_mtexts = select_pub_text_entities()
    cad_texts  = select_text(autocast=True)
    cad_mtexts = select_mtext(autocast=True)

    for ent in (tz_texts + tz_mtexts + cad_texts + cad_mtexts):
        try:
            ent.Layer = LAYER_NAME
        except Exception:
            pass
    return tz_texts, tz_mtexts, cad_texts, cad_mtexts


# 7) 选择 LINE / CIRCLE / ELLIPSE / SPLINE
def select_line(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["LINE"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

def select_circle(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["CIRCLE"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

def select_ellipse(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["ELLIPSE"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

def select_spline(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["SPLINE"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents


# 8) 传统多段线（POLYLINE）与轻量多段线（LWPOLYLINE）
def select_polyline_chuantong(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(mode="all", filter_types=[0], filter_data=["POLYLINE"], autocast=autocast)
            print(f"✅ select_polyline_chuantong 成功（第 {k} 次），耗时 {time.time()-t0:.3f}s，共 {len(ents)} 条")
            return ents
        except Exception as e:
            last = e
            print(f"⚠ select_polyline_chuantong 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"❌ select_polyline_chuantong 在 {max_retries} 次后仍失败：{last!r}")
    return []

def select_polyline(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(mode="all", filter_types=[0], filter_data=["LWPOLYLINE"], autocast=autocast)
            print(f"✅ select_polyline 成功（第 {k} 次），耗时 {time.time()-t0:.3f}s，共 {len(ents)} 条")
            return ents
        except Exception as e:
            last = e
            print(f"⚠ select_polyline 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"❌ select_polyline 在 {max_retries} 次后仍失败：{last!r}")
    return []


#&&% ##  6  屏幕选择





# --- 新版选择：不依赖全局 doc；不 for-in；自动清理；可选自动 CastTo ---
def pmxz(prompt="\n请在屏幕拾取图元，以Enter键结束：", autocast=True):
    """
    交互拾取对象（显式 on-screen 选择），返回实体列表。
    内部直接复用统一内核 ss_select，保持与其它选择函数一致的行为。
    """
    return ss_select(
        mode="onscreen",
        p1=None, p2=None,
        filter_types=None,     # 如需限制类型，可传 [0] 并配合 filter_data=["LINE"] 等
        filter_data=None,
        autocast=autocast,
        prompt=prompt
    )

"""
如果你想“只允许选某类对象”（比如直线），就传过滤参数即可：

def pmxz_line(prompt="选一根直线："):
    return ss_select("onscreen", filter_types=[0], filter_data=["LINE"], autocast=True, prompt=prompt)


"""


# 8) 隐性 → 显性选择（安全版）
def yin_to_xian_xuanze(LB, wait_s=0.6):
    """
    将 COM 选中的对象列表（LB）转换为命令窗口中的蓝色“高亮选中”状态。
    实现：StartUndoMark → 批量 Delete → U 撤销 → SELECT P。
    """
    _, doc = get_acad_doc()
    deleted = 0
    com_retry(lambda: doc.StartUndoMark())
    try:
        for x in LB:
            try:
                com_retry(lambda x=x: x.Delete())
                deleted += 1
            except Exception:
                pass
        print(f"🧹 尝试删除 {len(LB)} 个对象，成功删除 {deleted} 个。")

        # 撤销删除（恢复并建立 _P）
        doc.SendCommand("_U\n\n")
        time.sleep(wait_s)

        # 使用 P（previous selection）建立显性选择
        doc.SendCommand("_.SELECT\nP\n\n")
        time.sleep(0.5)
        print("✅ 已将隐性选择转换为显性（蓝色高亮）")
    finally:
        try: com_retry(lambda: doc.EndUndoMark())
        except Exception: pass


# 9) 显性 → 隐性选择（保留你原法，但收敛成函数）
def xian_to_yin_xuanze():
    _, doc = get_acad_doc()
    # 基于你的流程：CopyBase → 全选 → Erase → PasteClip → 取 SelectionSetAll
    doc.SendCommand("_COPYBASE\n0,0,0\n\n"); time.sleep(3)
    doc.SendCommand("_AI_SELALL\n\n");       time.sleep(1.5)
    doc.SendCommand("_ERASE\n\n");           time.sleep(0.8)
    doc.SendCommand("_PASTECLIP\n0,0,0\n\n");time.sleep(2.0)

    # 取隐性选择（全部）
    try:
        doc.SelectionSets.Item("MySelectionSet").Delete()
    except Exception:
        pass
    ss = doc.SelectionSets.Add("MySelectionSet")
    ss.Select(5)  # 全选
    lb = [ss.Item(i) for i in range(ss.Count)]
    try: ss.Delete()
    except Exception: pass
    print(f"✅ 共获取对象 {len(lb)} 个，转换为可操作选择列表")
    return lb


# 10) 隐性窗口选择（优先 SelectionSet 窗口；失败兜底遍历包围盒）
def select_objects_in_window_area(x1, y1, x2, y2, max_retry=5):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    doc.SetVariable("TILEMODE", 1)

    # 临时选择集
    try: doc.SelectionSets.Item("MyWindowSelection").Delete()
    except Exception: pass
    ss = doc.SelectionSets.Add("MyWindowSelection")

    entities = []
    # ① 重试窗口选择
    p1, p2 = pt3(x_lo,y_lo), pt3(x_hi,y_hi)
    for _ in range(max_retry):
        try:
            ss.Clear()
            ss.Select(1, p1, p2)  # 1 = Window
            if ss.Count > 0: break
        except Exception:
            pass
        time.sleep(0.25)

    if ss.Count > 0:
        for i in range(ss.Count):
            try: entities.append(_maybe_cast(ss.Item(i)))
            except Exception: continue
        try: ss.Delete()
        except Exception: pass
        print(f"✅ 窗口选择成功，共 {len(entities)} 个对象。")
        return entities

    # ② 兜底：遍历 ModelSpace 的包围盒角点
    try:
        msp = doc.ModelSpace
        def in_win(pt): return x_lo <= pt[0] <= x_hi and y_lo <= pt[1] <= y_hi
        for ent in msp:
            try:
                a,b = com_retry(lambda: ent.GetBoundingBox())
                if in_win(a) or in_win(b):
                    entities.append(_maybe_cast(ent))
            except Exception:
                continue
        print(f"[FALLBACK] 遍历模型空间得到 {len(entities)} 个对象。")
    except Exception as e:
        print(f"❌ 遍历模型空间失败: {e}")

    try: ss.Delete()
    except Exception: pass
    return entities


# 11) 隐显结合的区域选择（高亮选择并返回 PickfirstSelectionSet）
def select_entities_in_window(x1, y1, x2, y2, ty: float = 1.0, select_mode: str = "_W"):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)

    # 清空 Pickfirst
    try: doc.Pickenabled = False
    except Exception: pass
    try: doc.PickfirstSelectionSet.Clear()
    except Exception: pass

    # Zoom（加 20% 缓冲）
    buf = 0.20 * ((x_hi-x_lo) + (y_hi-y_lo)) / 2.0
    doc.SendCommand(f"_.ZOOM\n_W\n{x_lo-buf},{y_lo-buf}\n{x_hi+buf},{y_hi+buf}\n")
    time.sleep(ty)

    # Select（显性，蓝色高亮）
    doc.SendCommand(f"_.SELECT\n{select_mode}\n{x_lo},{y_lo}\n{x_hi},{y_hi}\n\n")
    time.sleep(ty/2)

    selset = doc.PickfirstSelectionSet
    com_list = [ent for ent in selset]
    try: selset.Clear()
    except Exception: pass
    return com_list


# 12) 显性区域选择（蓝色高亮）
def highlight_entities_in_window(x1, y1, x2, y2):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    h = 0.1 * ((abs(x_hi-x_lo)+abs(y_hi-y_lo))/2.0)

    # ZOOM
    doc.SendCommand(f"_.ZOOM\n_W\n{x_lo-h},{y_lo-h}\n{x_hi+h},{y_hi+h}\n"); time.sleep(1)

    # 显性 SELECT
    doc.SendCommand(f"_.SELECT\n_W\n{x_lo},{y_lo}\n{x_hi},{y_hi}\n\n"); time.sleep(0.5)
    print(f"✅ 已高亮选择区域 ({x_lo},{y_lo}) ~ ({x_hi},{y_hi}) 的对象")


# ====== 辅助：按偏移量扩张包围框（p1,p2 都是 (x,y,?) 序列）======
def expand_rectangle(p1, p2, offset):
    x1, y1 = float(p1[0]), float(p1[1])
    x2, y2 = float(p2[0]), float(p2[1])
    x_lo, x_hi = sorted((x1, x2))
    y_lo, y_hi = sorted((y1, y2))
    return (x_lo - offset, y_lo - offset), (x_hi + offset, y_hi + offset)
# 1) 隔离对象（模型空间区域）
def isolate_modelspace_area(x1, y1, x2, y2):
    """
    在模型空间，将窗口区域内对象隔离显示：
      - 隐性选区（COM SelectionSet/包围盒兜底）→
      - 转显性（蓝色高亮，供 SendCommand 识别）→
      - 发送 _IsolateObjects
    """
    _, doc = get_acad_doc()
    # 确保模型空间
    try: doc.SetVariable("TILEMODE", 1)
    except Exception: pass

    LB = select_objects_in_window_area(x1, y1, x2, y2)
    if not LB:
        print("❌ 没有选择到对象，终止操作")
        return

    print(f"✅ 选中对象 {len(LB)} 个，准备隔离")
    yin_to_xian_xuanze(LB, wait_s=0.6)  # 已实现：隐性→显性
    time.sleep(0.4)
    doc.SendCommand("_.IsolateObjects\n")
    print("🎯 已发送隔离命令")
# 2) 图纸空间的区域选择（隐性列表返回）
def select_paperspace_objects_in_window(x1, y1, x2, y2):
    """
    在图纸空间内，选择窗口 (x1,y1)-(x2,y2) 覆盖的对象；优先用 SelectionSet 窗口，
    若为空则遍历 PaperSpace 的包围盒兜底。返回 COM 对象列表（已自动 Cast）。
    """
    _, doc = get_acad_doc()
    try: doc.SetVariable("TILEMODE", 0)  # 切到图纸空间
    except Exception: pass

    (x_lo, y_lo), (x_hi, y_hi) = normalize_rect(x1, y1, x2, y2)
    # 优先：窗口选择
    try:
        items = ss_select(
            mode="window",
            p1=pt3(x_lo, y_lo),
            p2=pt3(x_hi, y_hi),
            filter_types=None,
            filter_data=None,
            autocast=True
        )
        if items:
            print(f"✅ 图纸空间窗口选择 {len(items)} 个对象。")
            return items
    except Exception:
        pass

    # 兜底：遍历 PaperSpace 包围盒
    selected = []
    ps = doc.PaperSpace
    def intersects(a1, a2, b1, b2):
        return (a1 <= b2) and (a2 >= b1)
    for i in range(ps.Count):
        try:
            ent = com_retry(lambda i=i: ps.Item(i))
            p1, p2 = com_retry(lambda: ent.GetBoundingBox())
            ex1, ey1 = p1[0], p1[1]
            ex2, ey2 = p2[0], p2[1]
            if intersects(ex1, ex2, x_lo, x_hi) and intersects(ey1, ey2, y_lo, y_hi):
                selected.append(_maybe_cast(ent))
        except Exception:
            continue
    print(f"[FALLBACK] 图纸空间包围盒遍历得到 {len(selected)} 个对象。")
    return selected
# 3) 将“单个对象”显性高亮（蓝色夹点）
def highlight_entity_by_bbox(entity):
    """
    对指定 COM 对象 entity，通过扩大其 bounding box 来进行显性高亮选中。
    会自动 Zoom 到对象附近，然后用窗口选择把它高亮出来。
    """
    try:
        # 用对象自身的 App/Doc，避免多文档切换干扰
        doc = com_retry(lambda: entity.Application.ActiveDocument)
        p1, p2 = com_retry(lambda: entity.GetBoundingBox())

        x1, y1 = float(p1[0]), float(p1[1])
        x2, y2 = float(p2[0]), float(p2[1])

        # Zoom 到对象 + 10% 缓冲
        h = 0.1 * ((abs(x1 - x2) + abs(y1 - y2)) / 2.0)
        doc.SendCommand(f"_.ZOOM\n_W\n{x1-h},{y1-h}\n{x2+h},{y2+h}\n")
        time.sleep(1.0)

        # 计算略扩张的选择窗口（墙体特殊偏移）
        x_len = abs(x2 - x1); y_len = abs(y2 - y1)
        max_len = max(x_len, y_len)
        offset = 130 if (getattr(entity, "ObjectName", "") == "TDbWall") else (max_len * 0.1)
        (X1,Y1),(X2,Y2) = expand_rectangle(p1, p2, offset)

        # 显性窗口选择
        highlight_entities_in_window(X1, Y1, X2, Y2)
        print("✅ 已高亮目标对象")
    except Exception as e:
        print("❌ 无法高亮该对象:", e)
# 4) 利用 Visible 进行“显选→隐选”转换的便捷封装（保持你的流程）
def select_visible(x1, y1, x2, y2):
    """
    先显性高亮窗口对象，再调用 xian_to_yin_xuanze() 转换为可操作的隐性列表。
    返回隐性列表（COM 对象集合）。
    """
    highlight_entities_in_window(x1, y1, x2, y2)
    time.sleep(1.0)
    return xian_to_yin_xuanze()
# 5) 显示指定空间中所有隐藏对象（可过滤类型，可选择高亮）
def unhide_all(space=None, filter_names=None, highlight=False):
    """
    显示 space 中所有 Visible=False 的对象。
    参数：
      - space: 可传 doc.ModelSpace / doc.PaperSpace / None / "model" / "paper"
      - filter_names: 仅恢复这些 ObjectName 的对象（例如 ["AcDbPolyline", "AcDbBlockReference"]）
      - highlight: True 则对恢复的对象调用 .Highlight(True)
    返回：revealed 列表（已自动 Cast）
    """
    _, doc = get_acad_doc()

    # 解析空间参数
    if space is None:
        target = doc.ModelSpace
    elif isinstance(space, str):
        s = space.strip().lower()
        if s in ("model", "m", "modelspace"):
            target = doc.ModelSpace
        elif s in ("paper", "p", "paperspace", "layout"):
            target = doc.PaperSpace
        else:
            target = doc.ModelSpace
    else:
        # 传入的是 COM 集合（有 Count/Item）
        target = space

    revealed = []
    cnt = com_retry(lambda: target.Count)
    for i in range(cnt):
        try:
            obj = com_retry(lambda i=i: target.Item(i))
            # 有些对象没有 Visible，跳过
            if not hasattr(obj, "Visible"):
                continue
            if obj.Visible:
                continue
            name = getattr(obj, "ObjectName", "")
            if (filter_names is None) or (name in filter_names):
                obj.Visible = True  # 设为可见
                if highlight and hasattr(obj, "Highlight"):
                    try: obj.Highlight(True)
                    except Exception: pass
                revealed.append(_maybe_cast(obj))
                print(f"✅ 显示对象：{name} | Handle={getattr(obj, 'Handle', '?')}")
        except Exception as e:
            print(f"⚠️ 跳过索引 {i}：{e}")

    print(f"\n📊 共显示 {len(revealed)} 个隐藏对象。")
    return revealed


#&&% #  组
"""

创建组
group = doc.Groups.Add("mygroup")
LB=pmxz()
请在屏幕拾取图元，以Enter键结束

group.AppendItems(vtobj([LB[0], LB[1], LB[2]]))


从组名获取组合组中对象

group = doc.Groups.Item("G001")

entities = [group.Item(i) for i in range(group.Count)]  # 遍历组内对象

entities[0].Handle
'2C3'

entities[0].Move(vtpnt(0,0,0),vtpnt(0,10000,0))


解除组
group.Delete()


group1 = doc.Groups.Add("mygroup")
group1.AppendItems(vtobj([LB[0], LB[1], LB[2]]))
group2 = doc.Groups.Add("mygroupA")
group2.AppendItems(vtobj([LB[3], LB[4], LB[5],LB[6]]))
group3 = doc.Groups.Add("mygroupB")
group3.AppendItems(vtobj([group1,group2]))非法操作
group3.AppendItems(vtobj([LB[0],LB[1],LB[2],LB[3], LB[4], LB[5],LB[6]]))


# 提取 group1 和 group2 的所有成员
group1_entities = [group1.Item(i) for i in range(group1.Count)]
group2_entities = [group2.Item(i) for i in range(group2.Count)]

# 合并成新的列表
all_entities = group1_entities + group2_entities

# 创建 group3
group3 = doc.Groups.Add("mygroupB")
group3.AppendItems(vtobj(all_entities))

 get_boundingbox_from_objects(objs)
"""

# 建立全部列表com对象的最小边界框

def get_boundingbox_from_objects(objs):#从列表com对象建立最小边界框
    """
    从一组图形对象（如 LB）中获取整体包围盒
    返回值：(min_x, min_y, min_z), (max_x, max_y, max_z)
    """
    min_point, max_point = None, None

    for obj in objs:
        try:
            min_pt, max_pt = obj.GetBoundingBox()
            if min_point is None:
                min_point, max_point = list(min_pt), list(max_pt)
            else:
                min_point = [min(min_point[i], min_pt[i]) for i in range(3)]
                max_point = [max(max_point[i], max_pt[i]) for i in range(3)]
        except Exception as e:
            print(f"跳过无法获取边界的对象: {obj.ObjectName}")
            continue

    return tuple(min_point), tuple(max_point)
        

# 建立组的最小边界框

def chuangjian_zu(group_name):

    group = doc.Groups.Add(group_name)

    return group

def nametogroup(group_name):#从组名获取实体com组对象
    group_obj = doc.Groups.Item(group_name)

    return group_obj

##获取所有组

def get_all_group_names():
    """
    获取当前 DWG 文档中所有组的名称列表。
    
    返回:
      List[str] — 包含所有组名称的列表
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    groups = doc.Groups
    return [groups.Item(i).Name for i in range(groups.Count)]

def get_all_groups():
    """
    获取当前 DWG 文档中所有组的 COM 对象列表及其名称。
    
    返回:
      List[Tuple[str, COMObject]] — 每项为 (组名称, 组对象)
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    groups = doc.Groups
    result = []
    for i in range(groups.Count):
        grp = groups.Item(i)
        result.append((grp.Name, grp))
    return result



#将多个com对象对象加入名为group_name的组
def add_objects_to_group(group_name, obj_list):

    """
    将 obj_list 中的所有图形对象加入名为 group_name 的组中
    如果组已存在，使用原组；否则新建
    返回：Group 对象
    """
    groups = doc.Groups
    try:
        group = groups.Item(group_name)
    except:
        group = groups.Add(group_name)

    group.AppendItems(vtlist(obj_list))
    return group


#将单独com对象对象加入名为group_name的组中

def add_object_to_group(group_name, obj):
    """
    将单个图形对象 obj 加入名为 group_name 的组中。
    如果组已存在，则使用该组；否则新建一个新组。
    
    参数：
      group_name (str)：组名称
      obj：要加入组的 COM 对象（如多段线、线段、块参照等）
    
    返回：
      COM Group 对象
    """
    groups = doc.Groups
    try:
        # 尝试获取已存在的组
        group = groups.Item(group_name)
    except Exception:
        # 不存在则新建
        group = groups.Add(group_name)
    
    # vtlist 工具将 Python list 转成 VBA 可接受的 SAFEARRAY
    group.AppendItems(vtlist([obj]))
    return group

#将单独com对象对象移出名为group_name的组
def remove_object_from_group(group_name, obj):
    """
    将单个 COM 对象 obj 从名为 group_name 的组中移出。
    如果组不存在或对象不在组中，则会打印错误信息但不抛异常。
    
    参数:
      group_name: 组名（字符串）
      obj:         要移出的 COM 对象
    
    返回:
      如果组存在，返回该 Group 对象；否则返回 None。
    """
    try:
        group = doc.Groups.Item(group_name)
    except Exception:
        print(f"❌ 组 '{group_name}' 不存在")
        return None

    # 把单个对象包装成长度为1的 COM SafeArray
    variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, [obj])
    try:
        group.RemoveItems(variant)
        print(f"✅ 已从组 '{group_name}' 中移除对象 {obj.Handle}")
    except Exception as e:
        print(f"❌ 从组 '{group_name}' 移除对象失败：{e}")

    return group

#将多个com对象对象移出名为group_name的组
def remove_objects_from_group(group_name, obj_list):
    """
    将 obj_list 中的所有图形对象从名为 group_name 的组中移出。
    如果组不存在，会打印提示并返回 None；否则返回该组对象。
    
    :param group_name: 组名称
    :param obj_list: 要移除的 COM 对象列表
    :return: Group 对象 或 None
    """
    groups = doc.Groups
    try:
        group = groups.Item(group_name)
    except Exception:
        print(f"组 '{group_name}' 不存在，无法移除对象。")
        return None

    # 把 Python 列表包装成 VARIANT SafeArray，VT_DISPATCH 表示对象类型
    arr = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj_list)
    try:
        group.RemoveItems(arr)
        print(f"已从组 '{group_name}' 中移除 {len(obj_list)} 个对象。")
    except Exception as e:
        print(f"移除对象时发生错误：{e}")
    return group



#&&% 从名为group_name的组获取内部包含的实体对象


def get_com_from_groupname(group_name):
    """
    根据组名获取对应实体列表。
    - 若组不存在、或组中无实体，均返回空列表，不抛异常。
    """

    try:
        group = nametogroup(group_name)
    except Exception:
        # nametogroup 本身失败（组不存在等）
        return []

    if not group:
        # group 为 None 或空，也直接返回空列表
        return []

    entities = [group.Item(i) for i in range(group.Count)] #从com组中获取全部对象
    
    return entities

#从名为group_name的组返回按类型分类的字典
def get_com_from_groupname_by_type(group_name):
    """
    根据组名获取对应实体，并按类型名分类返回。

    :param group_name: 组名称
    :return: dict，键为实体类型名（ObjectName），值为该类型的实体列表
    """
    # nametogroup 是你已有的“组名→Group 对象”函数
    group = nametogroup(group_name)
    if group is None:
        print(f"组 '{group_name}' 不存在")
        return {}

    by_type = {}
    # Group.Count 是实体数量，Item(i) 取出第 i 个实体
    for i in range(group.Count):
        ent = group.Item(i)
        # AutoCAD COM 对象一般有 ObjectName 属性
        typ = getattr(ent, "ObjectName", None) or getattr(ent, "EntityName", "Unknown")
        by_type.setdefault(typ, []).append(ent)

    # 打印一下各类型数量，方便调试
    for typ, lst in by_type.items():
        print(f"  类型 {typ} ：{len(lst)} 个实体")

    return by_type

#从名为group_name的组返回按类型分类的字典，且类型按各自位置提取函数排好序
def get_group_entities_sorted(group_name, type_extractors, cha_Y=0.5):
    """
    从组中按类型获取实体，并对指定类型按坐标排序。

    :param group_name: str，组名
    :param type_extractors: dict，{ type_name: extract_func }，
           extract_func(ent) 返回 (x, y) 坐标，用于排序。
    :param cha_Y: float，同一行 Y 方向的容差
    :return: dict，{ type_name: [ent, ...] }，已排序或原序

    对天正单行多行文字可以使用GetBoundingBox获取的左下角点
    
    """
    def get_cadtext_pos(ent):
        # CAD单行、多行文字：Position 属性返回 (x, y, z)
        return float(ent.InsertionPoint[0]), float(ent.InsertionPoint[1])


    # 先按类型获取全部实体
    by_type = get_com_from_groupname_by_type(group_name)
    sorted_by_type = {}

    for typ, ents in by_type.items():
        if typ in type_extractors:
            extract_func = type_extractors[typ]
            # 排序
            sorted_list = sort_entities_by_position(ents, extract_func, cha_Y=cha_Y)
            sorted_by_type[typ] = sorted_list
            print(f"Type '{typ}' sorted with {len(sorted_list)} entities")
        else:
            # 保持原序
            sorted_by_type[typ] = list(ents)
            print(f"Type '{typ}' left unsorted ({len(ents)} entities)")

    return sorted_by_type


#从名为group_name的组返回按类型分类的字典，各类型统一按boundingbox中心排好序


def get_group_entities_sorted_by_type_and_bbox(group_name, cha_Y=0.5):
    """
    将组 group_name 中的实体按类型分类，并对每种类型内部按包围盒中心排序：
      1) 先按 center_y 降序（从上到下）
      2) 同一“行”内（|ΔY|<cha_Y）再按 center_x 升序（从左到右）

    参数：
      group_name: 要操作的组名
      cha_Y: 同一“行”Y 方向容差

    返回：
      一个 dict，key=类型名(ObjectName)，value=排序后的实体列表
    """
    # 1. 取组
    group = nametogroup(group_name)
    ents = [group.Item(i) for i in range(group.Count)]

    # 2. 按类型分组
    by_type = {}
    for ent in ents:
        typ = getattr(ent, "ObjectName", None) or ent.EntityName
        by_type.setdefault(typ, []).append(ent)

    # 3. 辅助：计算包围盒中心点

    # 4. 对每个类型内部排序
    for typ, lst in by_type.items():
        triples = [(e, *bbox_center_2(e)) for e in lst]
        # Y 降序
        triples.sort(key=lambda t: -t[2])
        # 同行内按 X 升序
        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])
            i = j
        # 覆盖原列表
        by_type[typ] = [t[0] for t in triples]

    return by_type


# 获取两个组中共有的实体，按类型分类并按包围盒中心排序
def common_group_entities_sorted(group_name1, group_name2, cha_Y=0.5):
    """
    获取两个组中共有的实体，按类型分类并按包围盒中心排序。

    参数：
      group_name1, group_name2: 要比较的两个组名
      cha_Y: 同一行判定的 Y 方向容差

    返回：
      dict => key: ObjectName 类型名, value: 排序后的实体列表
    """
    # 1. 取组
    g1 = nametogroup(group_name1)
    g2 = nametogroup(group_name2)

    ents1 = [g1.Item(i) for i in range(g1.Count)]
    ents2 = [g2.Item(i) for i in range(g2.Count)]

    # 2. 建立 handle->entity 映射
    map1 = {e.Handle: e for e in ents1}
    map2 = {e.Handle: e for e in ents2}

    # 3. 找共有的 handles
    common_handles = set(map1.keys()) & set(map2.keys())

    # 4. 收集共有实体（这里取自 map1）
    common_ents = [map1[h] for h in common_handles]

    # 5. 按类型分组
    by_type = {}
    for ent in common_ents:
        typ = getattr(ent, "ObjectName", None) or ent.EntityName
        by_type.setdefault(typ, []).append(ent)

    # 6. 包围盒中心点计算
    def bbox_center(e):
        min_pt, max_pt = e.GetBoundingBox()
        x1, y1, _ = tuple(min_pt)
        x2, y2, _ = tuple(max_pt)
        return ((x1 + x2) / 2, (y1 + y2) / 2)

    # 7. 排序：Y 降序，同一行内(|ΔY|<cha_Y)按 X 升序
    for typ, lst in by_type.items():
        triples = [(e, *bbox_center(e)) for e in lst]
        triples.sort(key=lambda t: -t[2])
        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])
            i = j
        by_type[typ] = [t[0] for t in triples]

    return by_type

def get_boundingbox_from_group(group):#从com对象group建立最小边界框

    """
    并非组的实际BoundingBOX数据

    """

    entities = [group.Item(i) for i in range(group.Count)] #从组中获取全部对象

    p1,p2 = get_boundingbox_from_objects(entities)

    return p1,p2


def select_group_entities(group_obj): #选择组中对象，转换为高亮显性选择
    """
    让组中的所有对象进入蓝色高亮状态（通过反选法）
    """
    try:
        handles = [entity.Handle for entity in group_obj]
        objs = []

        for obj in mp:
            if obj.Handle in handles:
                objs.append(obj)

        yin_to_xian_xuanze(objs)  # 你已有的高亮函数
        print("✅ 已显性选中组内图元")
        return True
    except Exception as e:
        print("❌ 显性选择失败:", e)
        return False

#&&% #  Handle和Label


"""
Handle身份信息和天正标签Label

与列表和组不同，身份信息的标识使我们可以从字典存储的名称信息精准控制每个对象，而不是总依赖对全部对象的遍历


可以在对象创建时就打上标签，使对象归为某个类

可以在一批对象完成后调整标签

可以临时标记一组标签

可以通过Label给天正对象加标签
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)#自动获取图纸空间之前最后一个绘制的对象


从Handle回溯com对象

doc.HandleToObject('2BD')
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
q1.Label="A1"
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
q1.Label="B1"
LB=pmxz()
请在屏幕拾取图元，以Enter键结束
LB[0].Label
'B1'
LB[1].Label
'A1'


"""

# 将列表对象按分类将其handle身份标识存入字典

@alias("h")
def HandleToObject(ZF):#从Handle身份信息值回溯com对象

    """
    对连接在墙上的门窗测试无效

    """

    obj = doc.HandleToObject(ZF)

    return obj

def print_coms_handle(LB):

    LC=[]

    for x in LB:

        LC.append(x.Handle)

    print(f"com对象列表对应的Handle句柄列表：{LC} ")




@alias("H")
def handles_to_coms(LB_handles):

    """
    对连接在墙上的门窗测试无效

    """
    LC=[]

    for xx in LB_handles:

        obj = doc.HandleToObject(xx)
        LC.append(obj)

    return LC


def get_all_handles():#获取所有Handle
    """
    获取当前图纸中所有对象（通常在 ModelSpace）的 Handle 值列表。

    返回：
        handle_list: 所有图元的 Handle 字符串列表
    """
    handle_list = []

    for obj in mp:  # 使用你已经定义好的全局 ModelSpace mp
        try:
            handle_list.append(obj.Handle)
        except:
            continue  # 跳过无 Handle 或异常对象

    print(f"✅ 已获取 {len(handle_list)} 个对象的 Handle")
    return handle_list

def find_entity_by_handle(handle_str):#从Handle获取实体（适合包括天正的文件）
    """
    遍历当前图纸所有对象，手动比对 Handle 值，找到指定的实体对象。

    参数：
        handle_str: 目标 Handle（字符串）

    返回：
        对象（若找到），否则 None
    """
    for obj in mp:  # 可扩展：遍历 sp 也行
        try:
            if obj.Handle == handle_str:
                return obj
        except:
            continue

    return None



def group_objects_by_type_and_handle(LB):#将列表对象的Handle身份信息分类存入字典返回
    """
    将com对象列表 LB 中的对象按 ObjectName 分类，并存储其 Handle。
    每类按 LB 中出现顺序编号。

    参数：
        LB - AutoCAD 实体对象列表（如 select_objects_in_window_area() 返回）

    返回：
        ZD - dict 格式 {ObjectName: [Handle1, Handle2, ...]}
    """
    ZD = {}  # 初始化字典

    for obj in LB:
        try:
            obj_type = obj.ObjectName
            handle = obj.Handle

            if obj_type not in ZD:
                ZD[obj_type] = []

            ZD[obj_type].append(handle)

        except Exception as e:
            print(f"⚠️ 跳过对象，原因: {e}")
            continue

    # 输出提示信息
    for obj_type, handles in ZD.items():
        print(f"✅ {obj_type}: 共 {len(handles)} 个对象")

    return ZD

# 通过名称存储对象信息反回溯对象

def record_handle_with_type(LB, typename, prefix="OBJ"):#将一批对象的 Handle 存储到结构化的字典中，并记录类型名和编号
    """
    替代 XData 方法：记录对象 Handle、类型名、编号，返回结构化字典。
    """
    ZD = {typename: {}}
    for i, obj in enumerate(LB, start=1):
        try:
            h = obj.Handle
            tag = f"{prefix}_{i:03d}"
            ZD[typename][h] = tag
        except:
            continue
    print(f"✅ 已记录 {len(ZD[typename])} 个“{typename}”对象（Handle+编号）")
    return ZD

def convert_named_dict(ZD, typename):# 构建编号 → COM 对象 的映射字典
    """
    将 ZD["门"] 的结构由 Handle: 编号 转换为 编号: COM对象
    返回：新的字典 {编号: COM实体}
    """
    doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument
    named_dict = {}

    handle_map = ZD.get(typename, {})
    for handle, name in handle_map.items():
        try:
            obj = doc.HandleToObject(handle)
            named_dict[name] = obj
        except:
            print(f"⚠️ 无法找到对象（handle={handle}）")
            continue

    return named_dict

def get_named_object(tag, ZD, typename="门"):#从标签获取对象
    named = convert_named_dict(ZD, typename)
    return named.get(tag)



def draw_tags_on_objects_fixed(named_dict, height=250, offset=(1000, 1000, 0)):#直接将编号文字写在每个图元上，通常居中或偏移一点点
    """
    在每个对象的中心点附近绘制标注文字。
    
    参数:
        named_dict - 如 {"Men_001": <COMObject>, ...}
        height     - 文字高度
        offset     - 偏移量（用于防止文字盖住对象）
    """
    import win32com.client
    import pythoncom

    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    ms = doc.ModelSpace

    for name, obj in named_dict.items():
        try:
            # 使用 GetBoundingBox 获取中心点
            min_pt, max_pt = obj.GetBoundingBox()
            center_pt = (
                (min_pt[0] + max_pt[0]) / 2,
                (min_pt[1] + max_pt[1]) / 2,
                (min_pt[2] + max_pt[2]) / 2
            )

            # 加上偏移量
            label_pt = (
                center_pt[0] + offset[0],
                center_pt[1] + offset[1],
                center_pt[2] + offset[2]
            )

            # 确保插入点是三维点
            pt = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, label_pt)
            
            # 添加文字
            ms.AddText(name, pt, height)
            print(f"✅ 已标注对象: {name}")

        except Exception as e:
            print(f"⚠️ 标注失败: {name}, 错误: {e}")

# 给天正对象打上标签存入字典，用于以名称反向回溯操作

def label_tarch_doors(LB1, typename="门", prefix="men"):#给选中的天正门打上标签并存入字典返回，非天正图元没有Label属性
    """
    从对象列表 LB1 中筛选出天正门 (ObjectName == 'TDbOpening')，
    并为其按顺序打上 .Label 标签（如 men001, men002 ...）。

    返回：
        ZD = {typename: {编号: 对象}}
    """
    ZD = {typename: {}}
    LB2 = []

    for obj in LB1:
        try:
            if hasattr(obj, "ObjectName") and obj.ObjectName == "TDbOpening":
                LB2.append(obj)
        except:
            continue

    for i, obj in enumerate(LB2, start=1):
        try:
            tag = f"{prefix}{i:03d}"
            obj.Label = tag
            ZD[typename][tag] = obj
            print(f"✅ 已标记: {tag}")
        except Exception as e:
            print(f"⚠️ 设置标签失败：{e}")

    print(f"\n📦 共找到并标注 {len(LB2)} 个天正门")
    return ZD


# 获取模型空间上绘制的最后一个图元

def last_obj():

    obj = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)

    return obj

# 获取模型空间上的Handle


"""
target_handles = ['5F', '60', '61']
map = get_handle_object_map(doc.ModelSpace)
objs = [map[h] for h in target_handles if h in map]
这比每次都遍历 ModelSpace 快得多，尤其是大图纸中上千个图元时。

"""

def get_handle_object_map(ms):
    """返回 {handle: object} 映射"""
    return {ent.Handle: ent for ent in ms}



#&&% XData

"""

在 RegAppTable 中注册,在第一次给图元附加 XData 时，AutoCAD 内部会检查 RegAppTable（注册应用程序表）中是否已经存在 “TestApp” 这个名称。

如果不存在，AutoCAD 会自动往 RegAppTable 里插入一条记录，把 “TestApp” 注册进去。

如果你希望手动控制，也可以先调用 doc.Application.RegistryModes.Add("TestApp")（或使用 AutoLISP：(regapp "TestApp")）

app_name    = "TestApp"            # 自定义的应用程序名
data_types  = [1000, 1040, 1070]
data_values = ["示例文字", 3.14159, 12345]
set_xdata(lineObj, app_name, data_types, data_values)
types_out, data_out = get_xdata(lineObj, app_name)
types_out
[1001, 1000, 1040, 1070]
data_out
['TestApp', '示例文字', 3.14159, 12345]


"""
def set_xdata(
    com_obj,
    app_name: str,
    data_types: list[int],
    data_values: list,
):
    """
    向任意 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）附加 XData。

    参数：
      com_obj        -- 任意支持 SetXData() 方法的 COM 对象
      app_name       -- 注册过的应用程序名（字符串）；第一个 DataType 一定是 1001，对应的第一个 Data 存放此 app_name
      data_types     -- 后续的 DataType 列表（不含第一项 1001）；例如 [1000, 1040, 1070] 等
      data_values    -- 与 data_types 对应的数据值列表；长度与 data_types 一一对应。例如 ["文字串", 3.14, 42]

    说明：
      AutoCAD 规定 XData 的第一对元素必须是 (1001, 应用程序名)。后面才是按顺序出现的其他 (DataType, Data)。
      因此实际发送给 SetXData 的 DataType 数组第一个元素要放 1001，Data 数组第一个元素要放 app_name。
    """
    def vtint(val):
        """
        将 Python 列表转换为 VARIANT 类型的整数数组，
        以便传给 COM 对象作为 XData 的 DataType。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, val)

    def vtvariant(var):
        """
        将 Python 列表转换为 VARIANT 类型的 VARIANT 数组，
        以便传给 COM 对象作为 XData 的 Data。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, var)


    # 1. 拼接完整的 DataType 列表：第一个为 1001
    full_types = [1001] + data_types

    # 2. 拼接完整的 Data 列表：第一个为 app_name
    full_data = [app_name] + data_values

    # 3. 将 Python 列表转换为 VARIANT 数组
    vt_types = vtint(full_types)
    vt_data  = vtvariant(full_data)

    # 4. 调用 COM 的 SetXData 方法
    com_obj.SetXData(vt_types, vt_data)

def get_xdata(
    com_obj,
    app_name: str,
):
    """
    从任意 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）读取 XData。

    参数：
      com_obj   -- 任意支持 GetXData() 方法的 COM 对象
      app_name  -- 申请读取的应用程序名（必须与 set_xdata 时使用的相同）

    返回：
      (type_codes, data_values) 二元组，其中
        type_codes: Python 列表，对应每个 XData 项目的 DataType（包括第一项 1001）
        data_values: Python 列表，对应每个 XData 项目的 Data（包括第一项 app_name）
    
    如果该对象没有附加此 app_name 下的 XData，则 GetXData 会抛出错误；建议调用前先用 Error Handling 包裹或
    通过 com_obj.GetXData(app_name) 进行捕获并返回 None。
    """
    def vtint(val):
        """
        将 Python 列表转换为 VARIANT 类型的整数数组，
        以便传给 COM 对象作为 XData 的 DataType。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, val)

    def vtvariant(var):
        """
        将 Python 列表转换为 VARIANT 类型的 VARIANT 数组，
        以便传给 COM 对象作为 XData 的 Data。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, var)


    try:
        type_codes, data_values = com_obj.GetXData(app_name)
        # 注意：返回的 type_codes 和 data_values 都是 tuple，转换为 list 更易处理
        return list(type_codes), list(data_values)
    except pythoncom.com_error:
        # 对象上不存在该 app_name 的 XData，或读取失败
        return None, None


#&&% Xdata标记

def set_xdata_tab(entitycom):

    app_name    = "PrintApp"
    data_types  = [1000]
    data_values = ["增补目录模板"]
    set_xdata(entitycom, app_name, data_types, data_values)

    return

def is_printApp_xdata_com(entitycom):

    try:

        get_xdata( entitycom, "PrintApp")

        return True

    except:

        return  False


#&&&&%%  第五部分  线面分析 


#_____________________________________________________________________________________________________________________________________________

#  模块使用说明

"""
该模块研究dwg图纸中的线段、圆曲线、平面等基本几何问题 

"""
# 线段分析

def compute_line_angle(line):#按绘制顺序度量线段角度
    """
    计算直线的方向角（单位：度），基于 StartPoint / EndPoint。
    非直线对象将抛出异常。
    0-360，从起点处画横线，旋转到终点的角度
    """

    try:
        x1, y1, _ = line.StartPoint
        x2, y2, _ = line.EndPoint
        dx = x2 - x1
        dy = y2 - y1
        angle_rad = math.atan2(dy, dx)
        angle_deg = math.degrees(angle_rad)
        if angle_deg < 0:
            angle_deg += 360
        return angle_deg
    except AttributeError:
        print("❌ 该对象不具备 StartPoint / EndPoint")
        return None


def draw_point(pt):
    """
    在模型空间绘制一个 AutoCAD 点实体。

    参数：
        pt: (x, y, z) 三维坐标元组

    返回：
        新创建的 Point 对象；失败时返回 None
    """
    try:
        # AutoCAD 的“点”由 AddPoint 创建，需传 VARIANT
        obj = mp.AddPoint(vtpnt(*pt))
        return obj
    except Exception as e:
        print(f"❌ 无法绘制点: {e}")
        return None

def draw_line(p1, p2):#从两点坐标返回直线段
    """
    在模型空间中绘制从 p1 到 p2 的直线段。

    参数：
        p1, p2: 三维坐标元组 (x, y, z)

    返回：
        新创建的直线对象（COM 对象）
    """
    try:
        line_obj = mp.AddLine(vtpnt(*p1), vtpnt(*p2))
        return line_obj
    except Exception as e:
        print(f"❌ 无法绘制直线: {e}")
        return None


def draw_circle(center, radius):
    """
    以 center 为圆心、radius 为半径绘制圆。

    参数：
        center: (x, y, z)
        radius: 浮点半径

    返回：
        新创建的 Circle 对象；失败时返回 None
    """
    try:
        obj = mp.AddCircle(vtpnt(*center), radius)
        return obj
    except Exception as e:
        print(f"❌ 无法绘制圆: {e}")
        return None


def draw_regular_polygon(center, radius, sides):
    """
    绘制正多边形（LWPolyline，已闭合）
    :param center: 圆心 (x,y,z)
    :param radius: 外接圆半径
    :param sides : 边数 ≥3
    """
    if sides < 3:
        print("❌ 边数必须 ≥ 3"); return None
    cx, cy, cz = (center + (0.0,))[:3]

    pts_flat = []
    for k in range(sides):
        ang = 2 * math.pi * k / sides
        pts_flat.extend([cx + radius*math.cos(ang),
                         cy + radius*math.sin(ang)])

    try:
        v_pts = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            pts_flat
        )
        poly = mp.AddLightWeightPolyline(v_pts)
        poly.Closed = True
        return poly
    except Exception as e:
        print(f"❌ 无法绘制正多边形: {e}")
        return None



def prioritize_horizontal(lines, tol=0.5):
    """
    将列表中所有“水平”直线段（起点和终点的 y 差小于 tol）放在最前面，
    其它直线保留原有相对顺序。

    :param lines: 直线对象列表，每个对象具有 .StartPoint 和 .EndPoint 属性，
                  这两个属性应返回 (x, y, z) 或类似可下标的三元组。
    :param tol:   判定为水平的 y 方向容差（默认 0.5）
    :return:      新列表，水平直线段在前，非水平在后
    """
    horizontals = []
    non_horizontals = []
    for ln in lines:
        y1 = ln.StartPoint[1]
        y2 = ln.EndPoint[1]
        if abs(y1 - y2) < tol:
            horizontals.append(ln)
        else:
            non_horizontals.append(ln)
    return horizontals,non_horizontals

    
def get_spline_length_by_conversion(spline_entity):#返回样条曲线的长度（按默认10分断拟合）
    
    """
    将样条曲线对象复制、高亮并通过 _SPLINEDIT 转为多段线，
    然后读取长度，并删除该多段线。

    返回：
        样条曲线转换后的长度值（float）
    """
    try:
        # Step 1：复制 spline 对象
        new_spline = spline_entity.Copy()

        # Step 2：显性高亮
        highlight_entity_by_bbox(new_spline)

        # Step 3：模拟命令 SPLINEDIT → P → Enter → Enter
        doc.SendCommand("_SPLINEDIT\nP\n\n\n")
        time.sleep(1.2)  # 等待 CAD 完成处理（可根据机器速度调整）

        # Step 4：获取新生成的对象（最后一个）
        last_index = doc.ModelSpace.Count - 1
        poly = doc.ModelSpace.Item(last_index)

        # Step 5：检查 Length 属性
        if hasattr(poly, "Length"):
            length = poly.Length
            poly.Delete()  # 删除临时 polyline
            return length
        else:
            print("❌ 转换后对象没有 Length 属性")
            return None

    except Exception as e:
        print(f"❌ 获取样条曲线长度失败：{e}")
        return None

def estimate_ellipse_length(ellipse):#返回椭圆长度
    """
    估算椭圆对象的长度（周长），使用 Ramanujan 公式。
    """
    try:
        a = ellipse.MajorRadius
        b = ellipse.MinorRadius

        pi = math.pi
        h = 3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b))
        length = pi * h
        return length
    except Exception as e:
        print(f"❌ 无法估算椭圆长度: {e}")
        return None



def get_entity_geometry_info(obj):#返回图形关键几何信息
    """
    根据图元类型返回其关键几何信息：
    - 点：坐标
    - 直线：起点、终点、长度
    - 圆：圆心、半径、长度、面积
    - 椭圆：中心、主轴、次轴、长度、面积
    - 多段线：起点、终点、长度、面积（若闭合）
    - 样条曲线：起点、终点、长度（需转换），面积（若闭合）
    """
    try:
        name = obj.ObjectName.lower()

        # 点
        if "point" in name:
            return {"type": "Point", "position": obj.Coordinates}

        # 直线
        elif "line" in name and "xline" not in name:
            p1 = obj.StartPoint
            p2 = obj.EndPoint
            length = math.dist(p1, p2)
            return {"type": "Line", "start": p1, "end": p2, "length": length}

        # 圆
        elif "circle" in name:
            center = obj.Center
            radius = obj.Radius
            length = 2 * math.pi * radius
            area = math.pi * radius ** 2
            return {"type": "Circle", "center": center, "radius": radius, "length": length, "area": area}

        # 椭圆
        elif "ellipse" in name:
            center = obj.Center
            a = obj.MajorRadius
            b = obj.MinorRadius
            area = math.pi * a * b
            h = 3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b))
            length = math.pi * h  # Ramanujan 公式
            return {
                "type": "Ellipse",
                "center": center,
                "major_radius": a,
                "minor_radius": b,
                "length": length,
                "area": area
            }

        # 多段线
        elif "polyline" in name:
            coords = obj.Coordinates
            start = (coords[0], coords[1], 0)
            end = (coords[-2], coords[-1], 0)
            length = getattr(obj, "Length", 0)
            area = obj.Area if obj.Closed else 0
            return {
                "type": "Polyline",
                "start": start,
                "end": end,
                "length": length,
                "area": area
            }

        # 样条曲线（需转换测量）
        elif "spline" in name:
            p1 = obj.GetFitPoint(0)
            p2 = obj.GetFitPoint(obj.NumberOfFitPoints - 1)
            length = get_spline_length_by_conversion(obj)
            area = obj.Area if obj.Closed else 0
            return {
                "type": "Spline",
                "start": p1,
                "end": p2,
                "length": length,
                "area": area
            }

        else:
            return {"type": "Unknown", "ObjectName": obj.ObjectName}

    except Exception as e:
        return {"type": "Error", "message": str(e)}



##在两点确定的方向上，返回与对象点指定距离的点

def points_on_line_at_distance_3d(
    p1: Tuple[float, float, float],
    p2: Tuple[float, float, float],
    px: Tuple[float, float, float],
    distance: float
) -> List[Tuple[float, float, float]]:
    """
    已知 px 在由 p1->p2 确定的直线上，返回在该直线上与 px 距离为 distance 的两个点。

    :param p1: 起点 (x1, y1, z1)
    :param p2: 终点 (x2, y2, z2)
    :param px: 参考点 (x, y, z)，位于直线上
    :param distance: 与 px 的目标距离
    :return: [(ax, ay, az), (bx, by, bz)]，分别为正向和反向移动 distance 后的点
    """
    x1, y1, z1 = p1
    x2, y2, z2 = p2
    xx, yy, zz = px

    # 1) 计算方向向量并归一化
    dx, dy, dz = x2 - x1, y2 - y1, z2 - z1
    length = math.sqrt(dx*dx + dy*dy + dz*dz)
    if length == 0:
        raise ValueError("p1 和 p2 重合，无法确定方向向量")
    ux, uy, uz = dx / length, dy / length, dz / length

    # 2) 沿正向和反向各移动 distance
    ax = xx + ux * distance
    ay = yy + uy * distance
    az = zz + uz * distance

    bx = xx - ux * distance
    by = yy - uy * distance
    bz = zz - uz * distance

    return [(ax, ay, az), (bx, by, bz)]


# 找出一组直线段内的伪相交区域

def find_fake_intersection_regions(lines, tol=10, real_tol=0.01):
    """
    查找伪相交区域：对于任意线段 A 的端点 P，若：
    - 存在其他线段 B 满足 P 到 B 距离 < tol，且
    - 对所有 B，P 到 B 的距离 >= real_tol
    则判定为伪相交点。
    在模型空间中绘制圆（半径 1000）表示这些点。
    """
    ensure_layer("测试辅助")
    ms = doc.ModelSpace
    added = []

    def point_to_line_distance(p, a1, a2):
        x0, y0 = p[:2]
        x1, y1 = a1[:2]
        x2, y2 = a2[:2]
        dx, dy = x2 - x1, y2 - y1
        if dx == dy == 0:
            return math.hypot(x0 - x1, y0 - y1)
        t = max(0, min(1, ((x0 - x1) * dx + (y0 - y1) * dy) / (dx*dx + dy*dy)))
        proj_x = x1 + t * dx
        proj_y = y1 + t * dy
        return math.hypot(x0 - proj_x, y0 - proj_y)

    for A in lines:
        try:
            p1 = A.StartPoint
            p2 = A.EndPoint
        except Exception:
            continue

        for pt in [p1, p2]:
            pt_key = tuple(round(c, 3) for c in pt)
            if pt_key in added:
                continue

            min_dist = 1e10
            has_near = False
            has_real_near = False

            for B in lines:
                if B == A:
                    continue
                try:
                    b1, b2 = B.StartPoint, B.EndPoint
                    dist = point_to_line_distance(pt, b1, b2)
                    if dist < tol:
                        has_near = True
                    if dist < real_tol:
                        has_real_near = True
                        break
                except:
                    continue

            if has_near and not has_real_near:
                ms.AddCircle(vtpnt(*pt), 1000).Layer = "测试辅助"
                added.append(pt_key)
                print(f"✅ 伪相交区域点: {pt}")

    print("✅ 伪相交区域绘制完成")



# 把区域内的直线段交点打断

def lines_daduan(start_point,end_point):#全部脚本统一采用三维坐标点模式

    """
    这个命令对于避免天正墙体没有出现不相交的覆盖是非常重要的，直接应用天正的tlinebk

    还要先处理假相交点区域待优化20250409

    """

    # 使用 f-string 语法将三维坐标变量插入命令字符串中
    start_point_str = f"{start_point[0]},{start_point[1]},{start_point[2]}"

    end_point_str = f"{end_point[0]},{end_point[1]},{end_point[2]}"

    command = f"tlinebk{chr(13)}{start_point_str}{chr(13)}{end_point_str}{chr(13)}{chr(13)}{chr(13)}"

    acad.ActiveDocument.SendCommand(command)






#找出一组直线段中的所有直线段中所有重复的线段并删除

def delete_duplicate_lines(lines, tol=0.01):
    """
    删除重复的直线段，仅保留每组中一条。

    参数：
        lines: 模型空间中所有线段对象列表（ObjectName 为 'AcDbLine'）
        tol: 距离容差，小于此值认为两点重合
    """
    def is_same_point(p1, p2):
        return all(abs(a - b) < tol for a, b in zip(p1, p2))

    def is_duplicate(line1, line2):
        try:
            a1, a2 = line1.StartPoint, line1.EndPoint
            b1, b2 = line2.StartPoint, line2.EndPoint
            return (
                (is_same_point(a1, b1) and is_same_point(a2, b2)) or
                (is_same_point(a1, b2) and is_same_point(a2, b1))
            )
        except:
            return False

    keep = []
    to_delete = []

    for i, line in enumerate(lines):
        is_dup = False
        for j in range(i):
            if is_duplicate(line, lines[j]):
                is_dup = True
                break
        if is_dup:
            to_delete.append(line)
        else:
            keep.append(line)

    count = 0
    for dup in to_delete:
        try:
            dup.Delete()
            count += 1
        except:
            continue

    print(f"✅ 删除了 {count} 条重复直线段，保留 {len(keep)} 条。")
    return keep



#删除完全或局部重复线段

def delete_redundant_lines(lines, tol=0.01):
    """
    删除重复线段和局部重复线段，只保留每组中的一条。
    """
    def is_same_point(p1, p2):
        return abs(p1[0] - p2[0]) < tol and abs(p1[1] - p2[1]) < tol

    def point_on_segment(p, a, b):
        ax, ay = a[:2]
        bx, by = b[:2]
        px, py = p[:2]
        cross = abs((bx - ax) * (py - ay) - (by - ay) * (px - ax))
        if cross > tol:
            return False
        dot = (px - ax) * (bx - ax) + (py - ay) * (by - ay)
        if dot < 0:
            return False
        sq_len = (bx - ax)**2 + (by - ay)**2
        if dot > sq_len:
            return False
        return True

    def is_completely_duplicate(l1, l2):
        try:
            p1, p2 = l1.StartPoint, l1.EndPoint
            q1, q2 = l2.StartPoint, l2.EndPoint
            return (
                (is_same_point(p1, q1) and is_same_point(p2, q2)) or
                (is_same_point(p1, q2) and is_same_point(p2, q1))
            )
        except:
            return False

    def is_locally_duplicate(short_line, long_line):
        try:
            p1, p2 = short_line.StartPoint, short_line.EndPoint
            q1, q2 = long_line.StartPoint, long_line.EndPoint
            return point_on_segment(p1, q1, q2) and point_on_segment(p2, q1, q2)
        except:
            return False

    to_delete_handles = set()
    total = len(lines)

    for i in range(total):
        l1 = lines[i]
        h1 = l1.Handle
        if h1 in to_delete_handles:
            continue
        for j in range(i + 1, total):
            l2 = lines[j]
            h2 = l2.Handle
            if h2 in to_delete_handles:
                continue

            if is_completely_duplicate(l1, l2):
                to_delete_handles.add(h2)
            elif is_locally_duplicate(l2, l1):
                to_delete_handles.add(h2)
            elif is_locally_duplicate(l1, l2):
                to_delete_handles.add(h1)
                break

    deleted = 0
    for ent in lines:
        if ent.Handle in to_delete_handles:
            try:
                ent.Delete()
                deleted += 1
            except:
                continue

    print(f"✅ 删除重复/局部重复线段 {deleted} 条，保留 {total - deleted} 条。")

#找出一组直线段中的孤立线段产生的交点


def find_isolated_intersections(LB, tol=0.5):
    """
    找出线段列表 LB 中的孤立线段，并计算它们与其它线段的所有交点。

    用于人工标记的门窗位置

    参数：
      LB:   线段列表，每个元素是 [(x1,y1,z1), (x2,y2,z2)]
      tol:  端点重合判断容差

    返回：
      intersections: 交点列表，每个元素是 (x, y, z)
    """
    def segment_intersection(seg1, seg2, tol):
        """
        计算线段 seg1=(A,B) 与 seg2=(C,D) 的交点（二维），
        若相交且唯一，返回 (x, y, z)，否则返回 None。
        z 取 seg1 第一端点的 z。
        """
        (x1, y1, z1), (x2, y2, _) = seg1
        (x3, y3, _),  (x4, y4, _) = seg2

        # 方向向量
        r = (x2-x1, y2-y1)
        s = (x4-x3, y4-y3)
        # 叉积 r × s
        rxs = r[0]*s[1] - r[1]*s[0]
        if abs(rxs) < tol:
            return None  # 平行或共线，不处理
        # 解 t, u
        qp = (x3-x1, y3-y1)
        t = (qp[0]*s[1] - qp[1]*s[0]) / rxs
        u = (qp[0]*r[1] - qp[1]*r[0]) / rxs
        # 只考虑严格交于段内
        if -tol <= t <= 1+tol and -tol <= u <= 1+tol:
            xi = x1 + t*r[0]
            yi = y1 + t*r[1]
            zi = z1
            return (xi, yi, zi)
        return None

    # 1. 找出孤立线段
    isolated = []
    for i, seg in enumerate(LB):
        p1, p2 = seg
        shared = False
        for j, other in enumerate(LB):
            if i == j:
                continue
            q1, q2 = other
            if same_point(p1, q1, tol) or same_point(p1, q2, tol) \
            or same_point(p2, q1, tol) or same_point(p2, q2, tol):
                shared = True
                break
        if not shared:
            isolated.append(seg)

    # 2. 对每根孤立线段，与其余线段求交点
    intersections = []
    for seg in isolated:
        for other in LB:
            if other is seg:
                continue
            ip = segment_intersection(seg, other, tol)
            if ip is not None:
                intersections.append(ip)

    #删除孤立线段20250420
    for seg in isolated:

        seg.Delete()

    return intersections    


##doc.sendcommand("TSpOutline"+chr(13)+"41849.69465957, 12250.50102376, 0"+chr(13)+chr(13)+chr(13))

##doc.sendcommand("TRoflna"+chr(13)+"0"+chr(13))


def get_inner_point_of_polygon(polygon: Polygon):
    """
    获取给定 polygon 的一个保证在其内部的点。

    参数：
        polygon (shapely.geometry.Polygon): 目标多边形

    返回：
        (x, y): 内部点坐标元组
    """
    if not isinstance(polygon, Polygon):
        raise ValueError("❌ 输入必须是 shapely.geometry.Polygon")

    inner_point = polygon.representative_point()
    return inner_point.x, inner_point.y


#&&%____________________________________________________       获取一组直线段所有的封闭多边形和外轮廓线       ________________
#…………………………………………………………………………………………………………………………………………………………………
# 线面分析 - 获取一组直线段所有的封闭多边形和外轮廓线

#__________________________________________________________________________________________________________________________
#…………………………………………………………………………………………………………………………………………………………………
#  使用说明

"""
获取一组直线段所有的封闭多边形和外轮廓线

主要函数有：
(1)寻找直线段组中最右下角封闭多边性
(2)从com边或顶点坐标列表用PL复线绘制多边形
(3)获取一组直线段的外轮廓线
(4)获取最右下角的封闭多边形不影响其他封闭多边形的连续边的顶点列表
(5)获取全部封闭多边形，但不完全
(6)获取全部封闭多边形

"""
def get_room_outline_from_point(x, y, z=0):## 获取输入点所在房间的轮廓
    """
    自动发送 TSpOutline 命令，从指定点获取房间轮廓。

    参数:
        x, y, z: 点的坐标，z 默认为 0。

    也许会有用处，能从多边形内点迅速得到多边形
        
    """
    try:
        point_str = f"{x},{y},{z}"
        cmd = (
            "TSpOutline" + chr(13) +  # 启动命令
            point_str + chr(13) +     # 输入点
            chr(13) +                 # 确认默认设置
            chr(13)                   # 确认生成
        )
        doc.SendCommand(cmd)
        print(f"✅ 已请求获取点 ({x},{y},{z}) 的房间轮廓。")

    except Exception as e:
        print(f"❌ 获取房间轮廓失败：{e}")


def connect_lines_to_polyline_if_closed(lines, tol=0.5):##判断一组闭合直线段是否构成封闭多段线，是就返回PL复线
    """
    判断线段是否首尾连接成闭合多边形，如果是，则绘制PL多段线。
    
    参数:
        lines: AutoCAD中选中的 AcDbLine 对象列表。
        tol: 容许的端点闭合误差。
    
    返回:
        多段线对象，或 None。
    """

    try:
        # 提取二维端点集合（忽略z）
        segments = []
        for ln in lines:
            try:
                p1 = (ln.StartPoint[0], ln.StartPoint[1])
                p2 = (ln.EndPoint[0], ln.EndPoint[1])
                segments.append((p1, p2))
            except Exception:
                continue

        if not segments:
            print("⚠️ 无有效线段")
            return None

        # 构建连接链条
        used = set()
        sequence = [segments[0][0]]
        current = segments[0][1]
        used.add(0)

        while True:
            found = False
            for idx, (a, b) in enumerate(segments):
                if idx in used:
                    continue
                if Point(current).distance(Point(a)) < tol:
                    sequence.append(current)
                    current = b
                    used.add(idx)
                    found = True
                    break
                elif Point(current).distance(Point(b)) < tol:
                    sequence.append(current)
                    current = a
                    used.add(idx)
                    found = True
                    break
            if not found:
                break

        # 检查是否闭合
        if Point(current).distance(Point(sequence[0])) > tol:
            print("❌ 线段未构成闭合区域")
            return None
        sequence.append(sequence[0])  # 闭合环

        # 构造二维点数组
        pts = []
        for pt in sequence:
            pts.extend([pt[0], pt[1]])

        # 绘制PL
        poly = doc.ModelSpace.AddLightWeightPolyline(vtFloat(pts))
        poly.Closed = True
        print("✅ 成功绘制封闭PL线")
        return poly

    except Exception as e:
        print(f"❌ Polyline 创建失败: {e}")
        return None


def is_closed_polygon_from_lines(lines, tol=0.5):##判断一组闭合直线段是否构成封闭多段线，不返回PL复线
    """
    判断一组 AutoCAD 直线段是否首尾连接形成闭合多边形。
    
    参数:
        lines: AcDbLine 类型的 COM 对象列表
        tol: 闭合判断容差，单位与CAD一致（如mm）

    返回:
        True 表示首尾闭合形成多边形，False 否则
    """
    try:
        # 提取二维端点 (x, y)
        segments = []
        for ln in lines:
            try:
                p1 = (ln.StartPoint[0], ln.StartPoint[1])
                p2 = (ln.EndPoint[0], ln.EndPoint[1])
                segments.append((p1, p2))
            except Exception:
                continue

        if not segments:
            return False

        # 构造首尾连接链
        used = set()
        sequence = [segments[0][0]]
        current = segments[0][1]
        used.add(0)

        while True:
            found = False
            for idx, (a, b) in enumerate(segments):
                if idx in used:
                    continue
                if Point(current).distance(Point(a)) < tol:
                    sequence.append(current)
                    current = b
                    used.add(idx)
                    found = True
                    break
                elif Point(current).distance(Point(b)) < tol:
                    sequence.append(current)
                    current = a
                    used.add(idx)
                    found = True
                    break
            if not found:
                break

        # 判断是否回到起点（闭合）
        if Point(current).distance(Point(sequence[0])) < tol and len(used) == len(segments):
            return True
        else:
            return False

    except Exception as e:
        print(f"❌ 判断失败: {e}")
        return False

def same_point(p1, p2, tol=0.5):
    """判断两个点是否在容差范围内相同（只比较 X、Y 坐标）"""
    return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol


def same_line(ln1, ln2, tol=0.5):
    """
    判断两条线段 ln1 和 ln2 是否“相同”
    线段被认为相同的条件是：
      ln1 的 StartPoint 与 ln2 的 StartPoint 近似相同且 ln1 的 EndPoint 与 ln2 的 EndPoint 近似相同，
      或者 ln1 的 StartPoint 与 ln2 的 EndPoint 近似相同且 ln1 的 EndPoint 与 ln2 的 StartPoint 近似相同。
    """
    p1 = tuple(ln1.StartPoint)
    p2 = tuple(ln1.EndPoint)
    q1 = tuple(ln2.StartPoint)
    q2 = tuple(ln2.EndPoint)

    return (same_point(p1, q1, tol) and same_point(p2, q2, tol)) or \
           (same_point(p1, q2, tol) and same_point(p2, q1, tol))


def calculate_absolute_angle(line, P, tol=0.5):
    """
    计算线段（line）从点 P 出发的绝对角度（0-360°）
    如果 line 的起点等于 P，则返回从 P 到终点的角度，否则返回从 P 到起点的角度。
    """
    sp = tuple(line.StartPoint)
    ep = tuple(line.EndPoint)
    if same_point(sp, P, tol):
        dx = ep[0] - P[0]
        dy = ep[1] - P[1]
    else:
        dx = sp[0] - P[0]
        dy = sp[1] - P[1]
    return math.degrees(math.atan2(dy, dx)) % 360

def calculate_relative_angle(line, P, current_line, tol=0.5):
    """
    计算当前参考线（current_line）与候选线段（line）之间的相对角度，
    角度是以共点 P 为中心，当前线从 P 到其非 P 的端点的绝对角度为基准，
    计算 candidate line 与该基准角度之间顺时针或逆时针的角度差（逆时针方向）。
    结果为 0 到 360 之间的数值。
    """
    sp_current = tuple(current_line.StartPoint)
    ep_current = tuple(current_line.EndPoint)
    sp = tuple(line.StartPoint)
    ep = tuple(line.EndPoint)

    # 选择当前线段中不等于 P 的端点作为参考
    if same_point(sp_current, P, tol):
        current_point = ep_current
    else:
        current_point = sp_current

    def angle(p1, p2):
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        return math.degrees(math.atan2(dy, dx)) % 360

    angle_current = angle(P, current_point)
    # 对候选线段，选择不等于 P 的端点
    if same_point(sp, P, tol):
        target_point = ep
    else:
        target_point = sp
    angle_target = angle(P, target_point)
    angle_diff = (angle_target - angle_current) % 360
    return angle_diff

#####################
# 函数：查找给定点P经过的线段，按照绝对角度排序

def find_lines_angle(lines, P, tol=0.5):
    """
    查找与指定点 P 共端点的所有线段，并按从 P 出发离开的绝对几何角度排序。

    参数:
        lines: 直线段对象列表，每个对象要求具备 StartPoint, EndPoint, Handle 属性。
        P: 三元组 (x, y, z)，目标共点。
        tol: 判断共点的容差（仅比较 x 和 y 坐标）。

    返回:
        按绝对角度从小到大排序的共点线段列表。
    """
    shared_lines = []
    P = tuple(P)
    print(f"调试：目标共点 P = {P}")
    
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            dx_sp = abs(sp[0] - P[0])
            dy_sp = abs(sp[1] - P[1])
            dx_ep = abs(ep[0] - P[0])
            dy_ep = abs(ep[1] - P[1])
            print(f"线段 {ln.Handle}: sp={sp} 差值=({dx_sp:.4f},{dy_sp:.4f}), ep={ep} 差值=({dx_ep:.4f},{dy_ep:.4f})")
            if (dx_sp <= tol and dy_sp <= tol) or (dx_ep <= tol and dy_ep <= tol):
                shared_lines.append(ln)
        except Exception as e:
            print(f"⚠️ 跳过无效线段 {getattr(ln, 'Handle', '未知')} : {e}")
            continue

    shared_lines.sort(key=lambda ln: calculate_absolute_angle(ln, P, tol))
    print("调试：按绝对角度排序后的共点线段：")
    for ln in shared_lines:
        print(f"  线段 {ln.Handle}：角度 = {calculate_absolute_angle(ln, P, tol):.2f}°")
    return shared_lines

#####################
# 函数：查找与P共点的线段，按照与当前线段的相对角度排序

def find_lines_sharing_point(lines, P, current_line, tol=0.5):
    """
    查找与指定点 P 共端点的所有线段，并按从 current_line 逆时针旋转到其他线段的相对角度排序。
    其中 current_line 的角度定义为 0°。
    
    参数:
        lines: 直线段对象列表，每个对象要求具备 StartPoint, EndPoint, Handle 属性。
        P: 共点，三元组 (x, y, z)。
        current_line: 当前参考线段，对应角度定义为 0°。
        tol: 判断共点的容差（只比较 x 和 y 坐标）。
        
    返回:
        按相对角度从小到大排序的经过 P 的线段列表（current_line 亦包括其中）。
    """
    shared_lines = []
    P = tuple(P)
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            if (abs(sp[0]-P[0])<=tol and abs(sp[1]-P[1])<=tol) or (abs(ep[0]-P[0])<=tol and abs(ep[1]-P[1])<=tol):
                shared_lines.append(ln)
        except Exception as e:
            print(f"⚠️ 跳过无效线段 {getattr(ln, 'Handle', '未知')} : {e}")
            continue
    # 根据从当前线段旋转（逆时针）的相对角度排序
    shared_lines.sort(key=lambda ln: calculate_relative_angle(ln, P, current_line, tol))
    return shared_lines

#####################
# 函数：根据当前线段和共点 P，选择下一条后继线段（选择相对角度最大的那条），返回 (后继线段, 新共点)

def find_successor_line_max(current_line, lines, P, tol=0.5):
    """
    在给定共点 P 处，排除当前线段（current_line）后，
    选择在该点处与当前线段相对旋转角度最大的那条线段作为后继线段，
    返回 (后继线段, 新共点)。

    参数：
      current_line: 当前线段对象。
      lines: 所有直线段对象列表，每个对象必须具备 StartPoint, EndPoint, Handle 属性。
      P: 共点（三元组）。
      tol: 共点判断的容差（默认 0.5）。

    返回：
      (后继线段, 新共点)，若找不到合适的后继线段，则返回 (None, P)。
    """
    # 调用 find_lines_sharing_point 获取所有经过 P 的线段，并使用 current_line 的相对角度排序
    candidates = find_lines_sharing_point(lines, P, current_line, tol)
    # 排除当前线段
    candidates = [ln for ln in candidates if ln.Handle != current_line.Handle]

    if not candidates:
        print(f"❌ 在点 {P} 处找不到除当前线外的候选后继线段")
        return None, P

    best_line = None
    max_angle = -1
    new_point = P
    for ln in candidates:
        relative_angle = calculate_relative_angle(ln, P, current_line, tol)
        if relative_angle > max_angle:
            max_angle = relative_angle
            if same_point(tuple(ln.StartPoint), P, tol):
                new_point = tuple(ln.EndPoint)
            else:
                new_point = tuple(ln.StartPoint)
            best_line = ln

    if best_line is None:
        print(f"❌ 没有找到有效的后继线段")
        return None, P

    print(f"选中后继线段 {best_line.Handle}，新共点 {new_point}")
    return best_line, new_point

#&&%#####################
# 辅助函数：从所有线段中找出最右下角的点
def find_rightbottom_point(lines, tol=0.5):
    """
    从所有线段端点中，找出 y 值最小的点；若有多个，则选 x 最大的点作为最右下角点。
    """
    all_points = []
    for line in lines:
        if hasattr(line, "StartPoint") and hasattr(line, "EndPoint"):
            all_points.append(tuple(line.StartPoint))
            all_points.append(tuple(line.EndPoint))
    if not all_points:
        return None
    min_y = min(p[1] for p in all_points)
    candidates = [p for p in all_points if abs(p[1]-min_y) <= tol]
    rb = max(candidates, key=lambda p: p[0])
    print(f"✅ 最右下角点为：{rb}")
    return rb



#  主函数
#  (1)
# 寻找直线段组中最右下角封闭多边性

#  该函数系列包括如下一些函数


"""
辅助函数

same_point(p1, p2, tol) 判断两点是否在容差范围内相同

calculate_absolute_angle(line, P, tol) 计算从点 P 出发到某条线段（取与 P 不相同的端点）的绝对角度

calculate_relative_angle(line, P, current_line, tol) 计算当前线（以 P 为起点，选取与 P 不相同的端点）

和候选线（以 P 为起点）的角度差（逆时针方向，结果在 0～360 之间）

函数 find_lines_angle
用于初始阶段：给定一个共点 P，找到所有经过该点的线段，并按它们的绝对角度排序（从小到大）。

函数 find_lines_sharing_point
给定共点 P 与一个“当前线段”（作为参考），返回所有经过 P 的线段，并按照从当前线段出发旋转（逆时针）的相对角度排序。

函数 find_successor_line_max
根据当前线段和共点 P，从通过 P 的候选线中选取“转角”最大的作为后继线段，并返回后继线段及该线段另一端的新共点。

函数 find_rightbottom_closed_polygon
利用上述函数构造封闭多边形。初始时先根据所有线段端点确定“最右下角”点（函数 find_rightbottom_point），然后在该点处调用

find_lines_angle 得到所有经过该点的线段，选取绝对角度最小的那一根作为第一条边（初始 current_line），再依次调用

find_successor_line_max 推进多边形直到闭合或达到最大步数。


"""

def find_rightbottom_closed_polygon(lines, tol=0.5, max_steps=50):
    """
    利用所有线段构造封闭多边形：
      1. 先查找所有线段端点中的最右下角点（RB）。
      2. 在 RB 处获取所有经过该点的线段（按照绝对角度排序），选择绝对角度最小的那根作为初始边。
      3. 以该初始边为第一条边，之后依次利用 find_successor_line_max 选择后继边，
         直到新共点回到初始点（闭合）或超过最大步数。
         
    返回：
      构成封闭多边形的点列表（依次为每条边的终点），若无法构成则返回 None。
    """
    # 定位最右下角点 RB
    rb = find_rightbottom_point(lines, tol)
    if rb is None:
        print("❌ 无右下角点")
        return None

    # 初始共点
    current_point = rb
    # 从 RB 处按绝对角度排序取候选线段
    candidates = find_lines_angle(lines, rb, tol)
    if not candidates:
        print(f"❌ 在右下角点 {rb} 处没有找到经过的线段")
        return None
    # 选择绝对角度最小的线段作为初始边
    current_line = candidates[0]
    print(f"调试：选中初始线段 {current_line.Handle}，绝对角度 = {calculate_absolute_angle(current_line, rb, tol):.2f}°")

    # 得到初始边的另一端
    sp = tuple(current_line.StartPoint)
    ep = tuple(current_line.EndPoint)
    if same_point(sp, rb, tol):
        next_point = ep
    else:
        next_point = sp

    polygon_points = [rb, next_point]
    visited_handles = {current_line.Handle}
    current_point = next_point
    steps = 1

    while steps < max_steps:
        # 调用 find_successor_line_max 得到下一条线段及其另一端的点
        successor, new_point = find_successor_line_max(current_line, lines, current_point, tol)
        if successor is None:
            print("❌ 无后继线段，构造失败")
            return None
        # 检查是否闭合（新共点与起始点接近）
        if same_point(new_point, rb, tol):
            polygon_points.append(rb)
            print(f"✅ 成功构建封闭多边形，步数 = {steps}")
            return polygon_points

        if successor.Handle in visited_handles:
            print("🔁 检测到重复线段，构造失败")
            return None

        polygon_points.append(new_point)
        visited_handles.add(successor.Handle)
        current_line = successor
        current_point = new_point
        steps += 1

    print("⚠️ 达到最大步数，未能构造出闭合多边形")
    return None


# 从com边或顶点坐标列表用PL复线绘制多边形


def draw_polygon_as_polyline(polygon, layer_name="测试辅助", tol=0.5):
    """
    将构造的多边形（polygon）转换为顶点序列，并在当前 AutoCAD 文档 doc 的 ModelSpace 中添加
    一个多段线（PLINE）。如果顶点序列的首尾两点重合，则绘制闭合多段线，否则绘制开放多段线。
    
    参数:
      polygon:
        1. 如果是由线段组成的列表，每个元素要求具备 StartPoint 和 EndPoint 属性，
           则按照这些线段的端点顺序构造顶点序列。
        2. 如果是顶点列表，例如 [(x, y, z), (x, y, z), ...]，则直接使用。
      layer_name: 绘制多段线所在的图层名称（默认为 "测试辅助"）。
      tol: 判断点是否重合的容差值（仅比较 x 和 y 坐标）。
      
    返回:
      成功时返回创建的多段线对象（PLINE），否则返回 None。
    """

    # 内部函数：判断两个点是否近似相等（仅比较 x 和 y 坐标）
##    def same_point(p1, p2, tol=tol):
##        return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol

    if not polygon:
        print("❌ 未提供有效的 polygon 数据")
        return None

    vertices = []
    is_closed = False  # 是否绘制闭合多段线

    # 判断 polygon 是顶点列表还是线段列表
    if isinstance(polygon[0], (tuple, list)):
        # 判断传入是否为顶点列表：检查第一个元素为 tuple/list 且长度>=3
        if len(polygon[0]) < 3:
            print("❌ 顶点数据格式错误")
            return None
        # 直接使用顶点列表（转换为 tuple 形式）
        vertices = [tuple(pt) for pt in polygon]
        # 如果首尾已经重合，则视为闭合；否则保持开放（不自动补充首点）
        if same_point(vertices[0], vertices[-1], tol):
            is_closed = True
        else:
            is_closed = False
    else:
        # 假定 polygon 是线段列表，每个元素具有 StartPoint 和 EndPoint 属性
        first_line = polygon[0]
        start_pt = tuple(first_line.StartPoint)
        last_line = polygon[-1]
        # 判断哪一端与起始点相连，作为初始点
        if same_point(start_pt, tuple(last_line.StartPoint), tol) or same_point(start_pt, tuple(last_line.EndPoint), tol):
            # start_pt 可以作为起点
            pass
        else:
            # 否则选用第一条线段的 EndPoint作为起始点
            start_pt = tuple(first_line.EndPoint)
        vertices.append(start_pt)
        current_pt = start_pt
        # 遍历每条线段构造顶点序列
        for line in polygon:
            sp = tuple(line.StartPoint)
            ep = tuple(line.EndPoint)
            if same_point(current_pt, sp, tol):
                next_pt = ep
            elif same_point(current_pt, ep, tol):
                next_pt = sp
            else:
                print(f"❌ 线段 {line.Handle} 与当前点 {current_pt} 未连接，构造多边形失败")
                return None
            vertices.append(next_pt)
            current_pt = next_pt
        # 检查是否闭合（首尾重合）
        if same_point(vertices[0], vertices[-1], tol):
            is_closed = True
        else:
            is_closed = False

    # 输出调试信息：打印顶点序列
    print("调试：多段线顶点序列：")
    for i, pt in enumerate(vertices):
        print(f"  顶点 {i}: {pt}")

    # 将顶点序列转换为一维坐标数组：[x1, y1, z1, x2, y2, z2, …]
    coords = []
    for pt in vertices:
        coords.extend([pt[0], pt[1], pt[2]])
    coords_tuple = tuple(coords)
    coords_variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, coords_tuple)

    # 确保图层存在，否则创建新图层
    try:
        _ = doc.Layers.Item(layer_name)
    except Exception:
        doc.Layers.Add(layer_name)

    # 在 ModelSpace 中添加多段线并设置相应属性
    try:
        ms = doc.ModelSpace
        polyline = ms.AddPolyline(coords_variant)
        polyline.Closed = is_closed
        polyline.Layer = layer_name
        # 可选设置颜色和宽度，按需要调整
        polyline.Color = 1
        polyline.ConstantWidth = 20
        polyline.Update()
        doc.Regen(0)
        print(f"✅ 成功在图层 '{layer_name}' 绘制多段线, Closed={is_closed}")
        return polyline
    except Exception as e:
        print("❌ 绘制多段线失败：", e)
        return None



#&&%###外轮廓线

# -----------------------------------------------------
# 辅助函数：判断两个点是否近似相等（仅比较 x 和 y 坐标）
def is_nearly_equal(p1, p2, tol):
    return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol



# -----------------------------------------------------
# 寻找后继线段：在共点 P 处，从当前边之外的候选边中选择相对角度最小的边
def find_successor_line_min(current_line, lines, P, tol=0.5):
    # 先获取共点 P 的所有线段，使用 find_lines_sharing_point 逻辑，但这里直接采用 find_lines_angle 排序后筛选
    candidates = []
    P = tuple(P)
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            if (abs(sp[0]-P[0]) <= tol and abs(sp[1]-P[1]) <= tol) or (abs(ep[0]-P[0]) <= tol and abs(ep[1]-P[1]) <= tol):
                # 排除当前线段自身
                if ln.Handle == current_line.Handle:
                    continue
                candidates.append(ln)
        except Exception as e:
            continue

    if not candidates:
        print(f"❌ 在点 {P} 处找不到候选后继线段")
        return None, P

    best_line = None
    min_angle = 360
    new_point = P
    for candidate in candidates:
        angle_diff = calculate_relative_angle(candidate, P, current_line, tol)
        # 选择相对角度最小的候选
        if angle_diff < min_angle:
            min_angle = angle_diff
            best_line = candidate
            # 更新新共点：选择 candidate 中不等于 P 的端点
            sp_cand = tuple(candidate.StartPoint)
            ep_cand = tuple(candidate.EndPoint)
            if is_nearly_equal(sp_cand, P, tol):
                new_point = ep_cand
            else:
                new_point = sp_cand

    if best_line is None:
        print(f"❌ 没有找到合适的后继线段")
        return None, P

    print(f"选中后继线段 {best_line.Handle}，新共点 {new_point}，相对转角 = {min_angle:.2f}°")
    return best_line, new_point



#  主函数
#  (3)
# 获取一组直线段的外轮廓线

def get_outer_contour(lines, tol=0.5, max_steps=50):
    """
    获取一组直线段的外轮廓线
    规则：
      1. 计算所有线段端点的最右下角点 P（绝对方法）。
      2. 在 P 处，按绝对角度排序，选择绝对角度最小的边作为第一条边；
      3. 以当前边的另一端点作为新的共点，从当前边出发，选择相对于当前边逆时针转角最小的候选边，
         直到新共点回到初始 P 点（闭合）或达到最大步数为止。
    返回：
      外轮廓线构成的线段列表，如果无法构成封闭轮廓则返回空列表。
    """
    # 内部函数：获取所有线段的最右下角点
    rb = find_rightbottom_point(lines, tol)
    if rb is None:
        print("❌ 无最右下角点")
        return []

    # 设置初始共点为 rb
    P = rb
    print(f"调试：最右下角点为 {P}")

    # 在 P 处，按照绝对角度排序获取所有共点的线段
    candidate_lines = find_lines_angle(lines, P, tol)
    if not candidate_lines:
        print(f"❌ 在点 {P} 处未找到共点线段")
        return []

    # 根据题目要求，第一条边选择绝对角度最小的线段（即排序后第 1 根线段）
    initial_line = candidate_lines[0]
    # 确定初始线段的另一端点
    sp_init = tuple(initial_line.StartPoint)
    ep_init = tuple(initial_line.EndPoint)
    if is_nearly_equal(sp_init, P, tol):
        next_point = ep_init
    else:
        next_point = sp_init
    print(f"调试：选中初始线段 {initial_line.Handle}，起点 {P} -> 终点 {next_point}")
    
    contour_lines = [initial_line]
    visited_handles = {initial_line.Handle}
    current_line = initial_line
    current_point = next_point
    steps = 0

    # 迭代构造外轮廓
    while steps < max_steps:
        print(f"调试：目标共点 P = {current_point}")
        # 在当前共点处查找候选的后继线段，使用最小相对角度策略
        successor, new_point = find_successor_line_min(current_line, lines, current_point, tol)
        if successor is None:
            print("❌ 无后继线段，构造失败")
            return []
        if successor.Handle in visited_handles:
            print(f"🔁 检测到重复线段 {successor.Handle}，构造失败")
            return []
        contour_lines.append(successor)
        visited_handles.add(successor.Handle)
        print(f"步 {steps+1}: 选中线段 {successor.Handle}，新共点 {new_point}")
        # 检查是否闭合：如果新共点与最右下角点近似相等，则认为闭合
        if is_nearly_equal(new_point, rb, tol):
            print("✅ 成功构造封闭轮廓线")
            return contour_lines
        # 更新当前线段及共点
        current_line = successor
        current_point = new_point
        steps += 1

    print("⚠️ 达到最大步数，未能构造封闭轮廓线")
    return []


#&&%##获取所有封闭多边形

#删除多边形的一些边

"""
有lines中多于两条直线段经过，就称该点为多分枝点

编制函数，PL是lines中的封闭多边形，用顶点列表表示，p1是lines的最右下角点，

且p1是PL的一个顶点，用坐标点表示。将PL的顶点排序，从p1按逆时针方向推进，

将遇到的非多分枝顶点放入列表LN，直到遇到一个多分枝点D1结束；再从p1按顺时

针方向推进，将遇到的顶点放入列表LN，直到遇到一个多分枝点D2结束。顺时针方

向推进时，把p1考虑在内，即p1如果为多分枝点则立即结束，结束时遇到的多分枝

点D2就是p1。逆时针方向推进时，不把p1考虑在内，即p1为多分枝点仍然推进，遇

到一个多分枝点D1结束。函数返回多分枝点D1，D2，以及它们之间的非多分枝顶点

列表LN。


"""
def deduplicate_vertices(vertices, tol=0.5):
    """
    去掉顶点列表中相邻重复的顶点：
    如果两个相邻顶点（顺序出现）之间的二维距离小于 tol，则认为它们重复，只保留前一个，
    但如果这个重复点是列表中的最后一个且与第一个顶点相同（表示闭合多边形），则保留此重复点。
    
    参数:
      vertices: 顶点列表，每个顶点格式为 (x, y, z) 的元组。
      tol: 距离阈值。
      
    返回:
      处理后的顶点列表，只有中间连续重复的点被去除，保留闭合多边形的首尾重复。
    """
    if not vertices:
        return []
    
    deduped = [vertices[0]]
    n = len(vertices)
    
    for i in range(1, n):
        pt = vertices[i]
        prev = deduped[-1]
        dx = pt[0] - prev[0]
        dy = pt[1] - prev[1]
        dist = math.sqrt(dx*dx + dy*dy)
        
        # 如果距离大于等于 tol，则认为是不同的顶点，保留它
        if dist >= tol:
            deduped.append(pt)
        else:
            # 如果当前顶点是最后一个，并且与第一个顶点相同，则保留它（表示闭合）
            if i == n - 1 and same_point(pt, vertices[0], tol):
                deduped.append(pt)
            # 否则跳过该点
    return deduped


#  主函数
#  (4)
# 获取最右下角的封闭多边形不影响其他封闭多边形的连续边的顶点列表 

def analyze_polygon_branches(PL, lines, p1, tol=0.5):
    """
    分析封闭多边形 PL 的分枝情况。PL 为封闭多边形的顶点列表（按逆时针顺序排列），
    p1 为最右下角点，必在 PL 中。
    
    规则：
      1. 从 p1 出发：
         - 沿逆时针方向（不将 p1 视为候选）推进，累计遇到的非多分枝顶点，直至遇到第一个多分枝点 D1；
         - 沿顺时针方向（将 p1 也考虑在内）推进，累计遇到的非多分枝顶点，直至遇到第一个多分枝点 D2。
      2. 将这两个方向累计得到的非多分枝顶点 LN（其中顺时针方向得到的顶点需要反转后与逆时针方向的顶点相接）打印出来。
      3. 由于 PL 是闭合多边形，从 D1 到 D2有两条连续顶点序列，从中选择包含 LN 的那条作为最终返回结果。

    返回：
      返回从 D1 到 D2（包含 D1 和 D2）的连续顶点序列，此序列包含了 LN 的顶点。
      如果任一方向未能找到多分枝点，则打印提示，并返回 None。
    """


    # 内部辅助：判断两个点是否近似相等（仅比较 x,y 坐标）
##    def same_point(a, b, tol=tol):
##        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol
##


    # 内部辅助：判断某顶点是否为多分枝点
    def is_multi_branch(vertex):
        cnt = 0
        for ln in lines:
            try:
                sp = tuple(ln.StartPoint)
                ep = tuple(ln.EndPoint)
            except Exception:
                continue
            if same_point(vertex, sp, tol) or same_point(vertex, ep, tol):
                cnt += 1
        return cnt > 2

    # 找出 p1 在 PL 中的索引
    try:
        idx = PL.index(p1)
    except Exception as e:
        print("错误：p1 不在 PL 中")
        return None

    n = len(PL)

    # 沿逆时针方向推进（不包括 p1），累计非多分枝顶点 LN_ccw，直到遇到第一个多分枝点 D1
    LN_ccw = []
    D1 = None
    i = (idx + 1) % n
    while i != idx:
        v = PL[i]
        if is_multi_branch(v):
            D1 = v
            break
        else:
            LN_ccw.append(v)
        i = (i + 1) % n
    if D1 is None:
        print("调试：逆时针方向未遇到多分枝点")
        return None

    # 沿顺时针方向推进（包含 p1），累计非多分枝顶点 LN_cw，直到遇到第一个多分枝点 D2
    LN_cw = []
    if is_multi_branch(PL[idx]):
        D2 = PL[idx]
    else:
        D2 = None
        i = (idx - 1) % n
        while i != idx:
            v = PL[i]
            if is_multi_branch(v):
                D2 = v
                break
            else:
                LN_cw.append(v)
            i = (i - 1) % n
        if D2 is None:
            print("调试：顺时针方向未遇到多分枝点")
            return None

    # 合并顺时针方向的顶点（需要反转）和逆时针方向的顶点
    LN = list(reversed(LN_cw)) + LN_ccw

    # 调试打印
    print("调试信息：")
    print("目标共点 p1 =", p1)
    print("逆时针方向收集的非多分枝顶点 (LN_ccw):")
    for pt in LN_ccw:
        print("  ", pt)
    print("顺时针方向收集的非多分枝顶点 (LN_cw, 原顺序):")
    for pt in LN_cw:
        print("  ", pt)
    print("合并后的非多分枝顶点 LN (从 D2 到 D1):")
    for pt in LN:
        print("  ", pt)
    print("逆时针方向遇到的多分枝点 D1:", D1)
    print("顺时针方向遇到的多分枝点 D2:", D2)

    # 计算在 PL 中 D1 和 D2 的索引
    try:
        idx_D1 = PL.index(D1)
        idx_D2 = PL.index(D2)
    except Exception as e:
        print("错误：无法在 PL 中查找 D1 或 D2:", e)
        return None

    # 由于 PL 是闭合的，我们有两条连接 D1 和 D2：
    if idx_D1 <= idx_D2:
        branch_A = PL[idx_D1: idx_D2 + 1]
        branch_B = PL[idx_D2:] + PL[:idx_D1 + 1]
    else:
        branch_A = PL[idx_D1:] + PL[:idx_D2 + 1]
        branch_B = PL[idx_D2: idx_D1 + 1]

    # 选择包含 LN 中顶点的那个分枝
    selected_branch = None
    if LN:
        found_in_A = any(same_point(v, LN[0], tol) for v in branch_A)
        found_in_B = any(same_point(v, LN[0], tol) for v in branch_B)
        if found_in_A and not found_in_B:
            selected_branch = branch_A
        elif found_in_B and not found_in_A:
            selected_branch = branch_B
        elif found_in_A and found_in_B:
            # 如果两条分枝都包含，则选择较短的那条，通常这才是“局部”的分枝
            selected_branch = branch_A if len(branch_A) <= len(branch_B) else branch_B
        else:
            print("调试：LN 中的顶点不在任一分枝上，默认选择 branch_A")
            selected_branch = branch_A
    else:
        print("调试：LN 为空，无法确定包含 LN 的分枝，默认选择 branch_A")
        selected_branch = branch_A

    print("最终返回的从 D1 到 D2（包含 LN 的分枝）顶点序列：")
    for pt in selected_branch:
        print("  ", pt)


    #去重
    selected_branch=deduplicate_vertices(selected_branch, tol=tol)

    return selected_branch

##根据输入的顶点列表判断，将lines的其顶点在该顶点列表的线段移出列表


def remove_lines_in_LBv(lines, LB_v, tol=0.1):
    """
    从 COM 线段列表 lines 中移除那些其两个顶点都在 LB_v 中的线段。
    
    参数：
      lines: COM 线段对象列表，每个对象要求具有 StartPoint 与 EndPoint 属性，
             其值为形如 (x, y, z) 的可迭代对象。
      LB_v: 顶点列表，每个顶点为 (x, y, z) 的元组。
      tol: 判断两顶点近似相等的容差值（默认 0.1）。
    
    返回：
      返回移除满足条件（即两端点都在 LB_v 中）的线段后剩余的线段列表。
    """

    def same_point(pt1, pt2, tol=tol):
        # 只比较 x, y, z 坐标的差值，判断两点是否近似相等
        return abs(pt1[0] - pt2[0]) <= tol and abs(pt1[1] - pt2[1]) <= tol and abs(pt1[2] - pt2[2]) <= tol

    remaining_lines = []
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
        except Exception as e:
            print(f"⚠️ 跳过无效线段，原因: {e}")
            continue

        # 判断该线段的起点和终点是否均存在于 LB_v 中
        sp_in = any(same_point(sp, lb, tol) for lb in LB_v)
        ep_in = any(same_point(ep, lb, tol) for lb in LB_v)
        if not (sp_in and ep_in):
            remaining_lines.append(ln)
        else:
            print(f"删除线段 {ln.Handle}，其两个端点均在 LB_v 中：sp={sp}, ep={ep}")
    return remaining_lines

#  主函数
#  (5)
#&&% 获取全部封闭多边形，但不完全

def process_polygons(lines, tol=0.5, max_steps=50, layer_name="测试辅助"):
    """
    递归提取并绘制直线段集 lines 中所有封闭多边形。
    
    流程：
      1. 从 lines 中提取最右下角封闭多边形（调用 find_rightbottom_closed_polygon）。
      2. 对该多边形：
            - 获取最右下角点 p1（调用 find_rightbottom_point）；
            - 将多边形（polygon）加入列表 LB；
            - 调用 analyze_polygon_branches 分析多边形分枝，得到用来判断移除线段的顶点列表 Lv；
            - 调用 remove_lines_in_LBv，将所有两端点均在 Lv 中的直线段从 lines 中移除；
            - 调用 draw_polygon_as_polyline 绘制该多边形，将生成的多段线 COM 对象加入 LBcom；
            - 然后调用 draw_polygon_as_polyline 绘制顶点列表 Lv（蓝色绘制），将生成的多段线 COM 对象加入 LB_yc；
      3. 重复上述过程直至无法提取出封闭多边形；
    
    返回：
      (LB, LBcom, LB_yc)，其中：
         LB: 封闭多边形顶点列表集合（每个为顶点序列）。
         LBcom: 绘制出的多段线 COM 对象列表。
         LB_yc: 绘制蓝色辅助多段线（用于检查移除的线段顶点）的 COM 对象列表。
    """
    LB = []
    LBcom = []
    LB_yc = []  # 存储蓝色辅助多段线 COM 对象，用于检查移除的顶点
    Ly = []  # 用来记录每次移除的直线段
    iteration = 0

    while True:
        iteration += 1
        print(f"\n【迭代 {iteration}】剩余直线段数量 = {len(lines)}")
        p1 = find_rightbottom_point(lines, tol)
        if p1 is None:
            print("未找到最右下角点，结束迭代")
            break
        print(f"当前最右下角点 p1 = {p1}")
        
        polygon = find_rightbottom_closed_polygon(lines, tol=tol, max_steps=max_steps)
        if polygon is None:
            print("无法提取封闭多边形，结束迭代")
            break
        LB.append(polygon)
        print("提取的封闭多边形顶点：")
        for pt in polygon:
            print("  ", pt)
        
        Lv = analyze_polygon_branches(polygon, lines, p1, tol=tol)
        if Lv is None:
            print("分析多分枝失败，结束本次迭代")
        else:
            print("用于移除线段的顶点列表 Lv：")
            for pt in Lv:
                print("  ", pt)
            # 移除 lines 中两端点均在 Lv 中的直线段（调用 remove_lines_in_LBv）
            new_lines = remove_lines_in_LBv(lines, Lv, tol=0.1)
            removed_count = len(lines) - len(new_lines)
            print(f"本次移除直线段数: {removed_count}")
            # 记录被移除的直线段
            for ln in lines:
                if ln not in new_lines:
                    Ly.append(ln)
            lines = new_lines
        
        # 绘制提取出的封闭多边形
        polyline = draw_polygon_as_polyline(polygon, layer_name=layer_name, tol=tol)
        if polyline:
            LBcom.append(polyline)
        
        # 调用 draw_polygon_as_polyline 绘制 Lv（辅助多段线，注意Lv是顶点列表）
        if Lv is not None and len(Lv) > 0:
            poly_blue = draw_polygon_as_polyline(Lv, layer_name=layer_name, tol=tol)
            if poly_blue:
                LB_yc.append(poly_blue)
        
        if len(lines) < 3:
            print("剩余直线段不足以构成封闭多边形，退出")
            break

    return LB, LBcom, LB_yc



def extract_polygon_from_lines(lines, tol=0.5):
    """
    将表示封闭多边形边缘的线段（COM对象列表）转换为顶点列表（按顺序排列），
    消除中间相邻的重复顶点。若能成功构成闭合多边形，则返回顶点列表（闭环不重复），否则返回 None。
    
    参数：
      lines: 直线段对象列表，每个对象要求具有 StartPoint 和 EndPoint 属性
      tol: 判断顶点相等的容差（仅比较 x 和 y 坐标）。
    
    返回：
      顶点列表，如 [p1, p2, ..., pn]，其中 p1 表示多边形的起点（不重复列出闭合顶点）。
    """
    if not lines:
        return None
    
    # 定义同点判断函数
##    def same_point(a, b, tol=tol):
##        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol
    
    # 拷贝线段列表，方便修改（注意：此处不会复制COM对象，仅复制引用）
    remaining_lines = list(lines)
    
    try:
        start_pt = tuple(remaining_lines[0].StartPoint)
    except Exception as e:
        print("提取第一条线段起点失败：", e)
        return None
    vertices = [start_pt]
    current_pt = start_pt

    # 由于多边形是闭合的，最多尝试 2*len(remaining_lines) 次
    for _ in range(len(remaining_lines) * 2):
        found_edge = False
        for ln in remaining_lines:
            try:
                sp = tuple(ln.StartPoint)
                ep = tuple(ln.EndPoint)
            except Exception:
                continue

            nxt = None
            if same_point(current_pt, sp) and (not same_point(current_pt, ep)):
                nxt = ep
            elif same_point(current_pt, ep) and (not same_point(current_pt, sp)):
                nxt = sp

            if nxt is None:
                continue

            # 一旦找到与当前点相连的边，移除此边
            remaining_lines.remove(ln)
            # 如果得到的 nxt 与起点相同，则说明闭合结束
            if same_point(nxt, start_pt):
                vertices.append(start_pt)
                return deduplicate_vertices(vertices, tol)
            else:
                vertices.append(nxt)
                current_pt = nxt
                found_edge = True
                break
        if not found_edge:
            # 未能找到与当前点相连的边，构成闭合失败
            break
    # 如果循环结束后闭合未完成，则返回 None
    return None



#将多段线列表炸开为线段，返回线段列表



def explode_polylines(LB):
    """
    对多段线列表 LB 中的每一个多段线，调用 .Explode() 方法，
    并将炸开得到的所有线段合并为一个新的线段列表返回。

    参数:
      LB: 多段线 COM 对象列表，每个对象应支持 .Explode() 方法，
          如 pl = polyline_explode_object.Explode() 返回该多段线炸开后的线段集合。
          
    返回:
      一个包含所有炸开后直线段 COM 对象的列表。
    """
    exploded_lines = []
    for pl in LB:
        try:
            # 调用 Explode() 方法，返回一个集合（例如 COM Collection）
            exploded = pl.Explode()
            # 遍历返回的集合，将每根线段添加到列表中
            # 注意：遍历 COM Collection 的方法可能因环境不同而不同，此处假设可以直接迭代
            for ln in exploded:
                exploded_lines.append(ln)
        except Exception as e:
            handle = getattr(pl, 'Handle', '未知')
            print(f"⚠️ 处理多段线 {handle} 时出错: {e}")
    return exploded_lines


#lines1 中那些不在 lines2 中的线段

def subtract_line_sets(lines1, lines2, tol=0.5):
    """
    比较两个线段集合 lines1 和 lines2，返回 lines1 中那些不在 lines2 中的线段。

    参数：
      lines1: 第一组线段（候选集合），每个对象须具有 StartPoint 和 EndPoint 属性。
      lines2: 第二组线段，参照集合。
      tol: 判断点是否相同的容差，默认值为 0.5。
      
    返回：
      lines1 中所有不与 lines2 中任意线段相同的线段构成的列表。
    """
    result = []
    for ln in lines1:
        found = False
        for ln2 in lines2:
            if same_line(ln, ln2, tol):
                found = True
                break
        if not found:
            result.append(ln)
    return result

#  主函数
#  (6)
#&&% 获取全部封闭多边形


def process_final(lines, tol=0.5, max_steps=50, layer_name="测试辅助"):

    print("len(lines):",len(lines))

    L1, L2, L3 = process_polygons(lines, tol=tol, max_steps=max_steps, layer_name=layer_name)

    print(f"process_polygons 完成：\n  L1 数量 = {len(L1)}\n  L2 数量 = {len(L2)}\n  L3 数量 = {len(L3)}")
    
    print(">>> 对 L3 中的多段线执行 Explode() 操作...")

    exploded_lines=explode_polylines(L3)



    shengyu = subtract_line_sets(lines, exploded_lines, tol=tol)

    for x in shengyu:

        print(x.StartPoint,x.EndPoint)

    L_shengyu=extract_polygon_from_lines(shengyu, tol=tol)
   

    ld = draw_polygon_as_polyline(L_shengyu, layer_name=layer_name, tol=tol)

    L2.append(ld)

    L1.append(L_shengyu)
    for x in exploded_lines:
        x.Delete()

    return L1, L2, L3



#&&&% PL复线处理


"""
研究PL复线处理的问题

(1)多段线的基本操作 get_unique_vertices_from_pl_com

(2)PL打印线 generate_name_and_ratio_from_polyline(polyline,A3dy=0)

(3)将正交六边形多段线分成两个矩形区域 split_hexagon_combined(polygon, tol=0.1, simplify_tol=0.5)

(4)获取多段线的上下左右边界的直线段，返回线段端点列表 get_bbox_edge_segments(pl, tol=0.5)

(5)获取多段线的内部的文字 get_texts_in_polyline(com_pl, tol=0.5)

(6)多段线上的均分插入 distribute_points_on_entity(entity, n, block, scale_factor, ys)

(7)返回 pl1 中与 pl2 “共线且有重叠”的区段列表 common_segments_between_polylines(pl1, pl2, tol=0.5)

(8)找到全部“两根多段线耦合成一个矩形”的多段线 two_plines_making_rectangle(pl1, pl2, tol=0.5)

"""
#  主函数
#  (1)
# 多段线的基本操作

#  该函数系列包括如下一些函数
"""

标准顶点坐标列表,是像[(0, 0, 0), (0, 100, 0), (0, 900, 0)]这样的列表

它表示3个连续顶点2根连续线段，公共点坐标不重复

"""
## 0 轻量多段线和一般传统多段线各有用处，后者才能在三维中使用，更广泛




@alias("画轻量多段线")
def draw_lwpolyline(
    coords3d: list[tuple[float, float, float]],
    layer_name: str = "0",
    width: float = 0.0,
    color: int = 256,
    closed: bool = False
):
    """
    根据一组 (x, y, z) 坐标点绘制轻量级多段线（LWPOLYLINE）。

    :param coords3d: 形如 [(x1, y1, z1), (x2, y2, z2), …] 的点列表，
                     仅使用 x,y 坐标，忽略 z。
    :param layer_name: 目标图层名称，不存在则自动创建。
    :param width:      多段线恒宽 (ConstantWidth)。
    :param color:      颜色索引 (AutoCAD Color Index)，256=BYLAYER。
    :param closed:     是否闭合多段线（首尾相连）。

    :return:           新建的轻量级多段线对象 (COM AddLightWeightPolyline)。

    pts = [
        (0.0, 0.0, 0.0),
        (100.0, 0.0, 0.0),
        (100.0, 50.0, 0.0),
        (0.0, 50.0, 0.0),
    ]
    poly = draw_lwpolyline(
        coords3d=pts,
        layer_name="dy_quyu",
        width=0.0,
        color=1,      # 红色
        closed=True
    )
    poly.Coordinates
    (0.0, 0.0, 100.0, 0.0, 100.0, 50.0, 0.0, 50.0)

    len((0.0, 0.0, 100.0, 0.0, 100.0, 50.0, 0.0, 50.0))
    8

    """
    # 1️⃣ 连接 AutoCAD

    # 2️⃣ 确保图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    # Optional: 开启图层
    lyr.LayerOn = True

    # 3️⃣ 准备坐标数组：扁平化 x,y
    raw = []
    for x, y, _ in coords3d:
        raw.extend((x, y))
    # 转 COM VARIANT 数组
    arr = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        raw
    )

    # 4️⃣ 绘制轻量级多段线

    try:

        pline = mp.AddLightWeightPolyline(arr)
        pline.Layer         = layer_name
        pline.ConstantWidth = width
        pline.color         = color
        pline.Closed        = bool(closed)

        print(f"✅ 已在图层『{layer_name}』绘制多段线，Closed = {closed}")
        return pline
    except Exception as e:
        print("❌ 绘制多段线失败:", e) 

   # 5️⃣ 返回新对象

    return pline



# 1 从com复线获取标准顶点坐标列表

def get_unique_vertices_from_pl_com(pl_com):
    """
    提取多段线的顶点列表，不重复连续线段的公共顶点，返回顶点列表。
    
    参数:
        pl_com: AutoCAD 中的 Polyline COM 对象（AcDbPolyline）
        
    返回:
        顶点列表，每两个点构成一条线段，格式：[ (x1, y1, z1), (x2, y2, z2), ... ]
        
    """
    # 获取二维坐标数据
    coords = pl_com.Coordinates
    vertices = []

    # 以每两个坐标为一组，构成线段
    for i in range(0, len(coords) - 1, 2):
        x1, y1 = coords[i], coords[i + 1]
        z1 = 0  # 假设z坐标为0，如果需要，可以通过某种方式获取真实z坐标
        if not vertices:
            vertices.append((x1, y1, z1))
        else:
            # 如果当前点与上一个点不重复，添加到顶点列表
            if (x1, y1, z1) != vertices[-1]:
                vertices.append((x1, y1, z1))
    
    # 添加最后一个点，避免遗漏
    last_x, last_y = coords[-2], coords[-1]
    z_last = 0  # 同样，假设最后的z坐标为0
    if (last_x, last_y, z_last) != vertices[-1]:
        vertices.append((last_x, last_y, z_last))
    
    return vertices

# 2 将com线段转成顶点坐标列表，一根线段一个列表

def convert_lines_to_points(segments):
    """
    将com线段列表转换为顶点列表，每条线段的两个端点作为一个独立的列表。

    参数:
      segments: 线段对象列表，每个元素具有 StartPoint 和 EndPoint 属性。

    返回:
      包含多个线段顶点的列表，每个线段是一个包含两个端点坐标的列表。
    """
    points_list = []

    for segment in segments:
        # 提取线段的两个端点
        start_point = tuple(segment.StartPoint)
        end_point = tuple(segment.EndPoint)

        # 将线段的两个端点存入一个列表，添加到结果列表
        points_list.append([start_point, end_point])

    return points_list

# 3 合并顶点列表表示的连续线段，允许多根断开的连续线段

def merge_segments_new(LB, tol=0.5):
    """
    使用convert_lines_to_points 将线段实体转成顶点列表表达式后就可以使用此命令

    不断合并连接，能连接的都会连接   

    """
    def same(p, q):
        return abs(p[0]-q[0]) <= tol and abs(p[1]-q[1]) <= tol

    # 1) 为端点做哈希 — 用 round() 把 tol 纳入 key，避免浮点字典键难比较
    def key(pt):
        return (round(pt[0]/tol), round(pt[1]/tol))   # 只 hash XY

    buckets = defaultdict(list)   # key(pt)  ->  [(seg_index, dir), ...]
    for idx, seg in enumerate(LB):
        a, b = seg[0], seg[-1]
        buckets[key(a)].append((idx, +1))   #  +1 表示 seg[0] 方向
        buckets[key(b)].append((idx, -1))   #  -1 表示 seg[-1] 方向

    used = [False]*len(LB)
    sequences = []

    while True:
        # 找到尚未使用的第一条线段
        try:
            seed_idx = next(i for i,u in enumerate(used) if not u)
        except StopIteration:
            break                           # 全部用完
        used[seed_idx] = True
        seq = deque(LB[seed_idx])           # 双端队列便于首尾增长

        # 函数: 把可接的线段拼到 deque 的一头
        def grow(at_tail: bool):
            while True:
                end_pt = seq[-1] if at_tail else seq[0]
                bucket = buckets[key(end_pt)]
                # 移除已用完的骨牌
                bucket[:] = [pair for pair in bucket if not used[pair[0]]]
                if not bucket:              # 再也接不上
                    break
                idx, direction = bucket.pop()
                used[idx] = True
                seg = LB[idx]
                # 根据 direction 决定正向还是反向加入
                if direction == +1:         # bucket 点是 seg[0]
                    add = seg[1:]           # 去掉公共点再拼
                else:                       # bucket 点是 seg[-1]
                    add = seg[-2::-1]       # 反向、去掉公共点
                if at_tail:
                    seq.extend(add)
                else:
                    seq.extendleft(add[::-1])   # extendleft 要反转

        # 先往尾巴拼，再往头拼（顺序无所谓，都会做到极限）
        grow(True)
        grow(False)
        sequences.append(list(seq))

    return sequences

# 4 绘制连续PL多段线，断开的PL多段线要分开绘制

def draw_polyline(vertices,
                  layer_name="测试辅助",
                  tol=0.5,
                  width=20,
                  color=1):
    """
    复线和多边形都应该按标准顶点坐标列表表达
    根据顶点序列 vertices 在当前 AutoCAD 文档 (全局变量 doc) 的 ModelSpace
    绘制多段线 (PLine)。

    参数
    ----
    vertices : list[tuple]
        仅支持顶点列表形式：[(x, y, z), (x, y, z), ...]。
    layer_name : str
        目标图层名；不存在则自动创建。
    tol : float
        判断首尾是否闭合的容差(只比较 X / Y)。
    width : float
        ConstantWidth 设置。
    color : int
        AutoCAD 颜色号。

    返回
    ----
    acad_polyline : PLine COM 对象 | None
    """

    # ----------------- 内部工具 -----------------
##    def same_point(p1, p2, _tol=tol):
##        """只比较 x、y 坐标的近似相等"""
##        return abs(p1[0] - p2[0]) <= _tol and abs(p1[1] - p2[1]) <= _tol
    # --------------------------------------------

    # ---------- 输入校验 ----------
    if not vertices or not isinstance(vertices, (list, tuple)):
        print("❌ 请输入有效的顶点列表")
        return None
    if not isinstance(vertices[0], (list, tuple)) or len(vertices[0]) < 3:
        print("❌ 顶点格式错误，应为 (x, y, z)")
        return None

    # --------- 处理闭合性 ---------
    is_closed = same_point(vertices[0], vertices[-1])

    # --------- 打印调试信息 --------
    print("调试：绘制多段线的顶点序列")
    for idx, pt in enumerate(vertices):
        print(f"  {idx}: {pt}")

    # --------- 生成 SAFEARRAY ------
    flat = []
    for x, y, z in vertices:
        flat.extend([x, y, z])
    coords_var = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, tuple(flat))

    # ---------- 确保图层存在 -------
    try:
        _ = doc.Layers.Item(layer_name)
    except Exception:
        doc.Layers.Add(layer_name)

    # ---------- 绘制多段线 ----------
    try:
        pl = doc.ModelSpace.AddPolyline(coords_var)
        pl.Closed = is_closed
        pl.Layer = layer_name
        pl.Color = color
        pl.ConstantWidth = width
        pl.Update()
        doc.Regen(0)
        print(f"✅ 已在图层『{layer_name}』绘制多段线，Closed = {is_closed}")
        return pl
    except Exception as e:
        print("❌ 绘制多段线失败:", e)
        return None



# 5 获取多段线后的线段列表，原来的多段线仍然存在


"""
pl=  Pl_obj.Explode()

"""



# 6 将多条直线段（允许不连续）连接成PL复线

def lines_to_polylines(Lc,
                       tol=0.5,
                       layer_name="某某图层",
                       width=20,
                       color=1):
    """
    将若干直线段 (COM Line 对象) 连续合并成多段线：
      1. 直线段 → 顶点对列表  (convert_lines_to_points)
      2. 连续段合并         (merge_segments_new)
      3. 生成多段线         (draw_polyline)
      4. 删除原直线段

    参数
    ----
    Lc : list[Line COM]
        要合并的 AcadLine 对象列表。
    tol : float
        顶点判同容差。
    layer_name / width / color
        传递给 draw_polyline 的控制参数。

    返回
    ----
    PLs : list[PLine COM]
        新生成的多段线对象列表。
    """

    # ---------- 0. 边界检查 ----------
    if not Lc:
        print("❌ 输入线段列表为空")
        return []

    # ---------- 1. 线段 → 顶点对 --------
    # convert_lines_to_points 应返回形如 [[p1,p2],[p3,p4] ...]
    LB = convert_lines_to_points(Lc)

    # ---------- 2. 合并连续顶点 --------
    # merge_segments_new 会把 LB 合并成若干连续顶点序列
    LK = merge_segments_new(LB, tol=tol)

    # ---------- 3. 生成多段线 ----------
    PLs = []
    for verts in LK:
        pl = draw_polyline(verts,
                           layer_name=layer_name,
                           tol=tol,
                           width=width,
                           color=color)
        if pl:
            PLs.append(pl)

    # ---------- 4. 删除原直线 ----------
    for ln in Lc:
        try:
            ln.Delete()
        except Exception:
            pass          # 若已被删除或无效则忽略

    print(f"✅ 已生成 {len(PLs)} 条多段线，并删除 {len(Lc)} 条原直线")
    return PLs





# 7 找到多段线的最左下角的点

def find_min_point(obj):
    """
    获取任意对象的左下角坐标（通过其外包盒）。

    :param obj: 支持 GetBoundingBox() 方法的 COM 对象（如多段线、块参照等）。
    :return:    (min_x, min_y) 元组，表示对象外包盒的左下角坐标
    """
    try:
        ll_point, _ = obj.GetBoundingBox()
        min_x, min_y, _ = ll_point
        return min_x, min_y
    except Exception as e:
        print(f"❌ 获取外包盒失败: {e}")
        return None, None

# 8 找到多段线的最右上角的点

def find_max_point(obj):
    """
    获取任意对象的右上角坐标（通过其外包盒）。

    :param obj: 支持 GetBoundingBox() 方法的 COM 对象（如多段线、块参照等）。
    :return:    (max_x, max_y) 元组，表示对象外包盒的右上角坐标
    """
    try:
        _, ur_point = obj.GetBoundingBox()
        max_x, max_y, _ = ur_point
        return max_x, max_y
    except Exception as e:
        print(f"❌ 获取外包盒失败: {e}")
        return None, None


def distance(point1, point2):
    """计算两点之间的距离"""
    return ((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)**0.5

# 9 删除多段线列表中重复的多段线

def remove_duplicate_polylines(LB_1):
    """处理多段线列表，删除重复的多段线，并将处理后的多段线添加到新列表中。"""
    LB_2 = []

    for XX in LB_1:

        biaoji=0
        try:
            XX.color = 6  # 尝试设置颜色，如果出错则跳过
        except :
            biaoji=1

        if biaoji == 0:
           
            # 检查是否有重复的多段线
            for YY in LB_1:

                try:
                    
                    if XX != YY:
                        
                        min_distance = distance(find_min_point(XX), find_min_point(YY))
                        
                        max_distance = distance(find_max_point(XX), find_max_point(YY))

                        if min_distance < 10 and max_distance < 10:
                            
                            YY.Delete()  # 尝试删除重复的多段线
                           
                except:

                    pass
           
            LB_2.append(XX)

    return LB_2

# 10 定义矩形

def define_rectangle_by_diagonal(p1, p2):
    """
    使用两个对角顶点定义矩形，矩形的边与坐标轴平行或垂直。
    p1, p2: 分别为矩形的一对对角顶点，格式为 (x, y)。
    返回矩形的四个顶点、长和宽。
    """
    x1, y1 = p1
    x2, y2 = p2

    # 计算两个边长
    length_x = abs(x2 - x1)
    length_y = abs(y2 - y1)

    # 确定长和宽
    length = max(length_x, length_y)
    width = min(length_x, length_y)

    # 确定矩形的四个顶点
    rectangle_points = [(x1, y1), (x1, y2), (x2, y2), (x2, y1)]

    return rectangle_points, length, width

def define_rectangle_by_diagonal_x(p1, p2):
    """
    使用两个对角顶点定义矩形，矩形的边与坐标轴平行或垂直。
    p1, p2: 分别为矩形的一对对角顶点，格式为 (x, y)。
    返回矩形的四个顶点、长和宽。
    """
    x1, y1 = p1
    x2, y2 = p2

    # 计算两个边长
    length_x = abs(x2 - x1)
    length_y = abs(y2 - y1)

    # 确定长和宽
    length = max(length_x, length_y)
    width = min(length_x, length_y)

    # 确定矩形的四个顶点
    rectangle_points = [x1, y1, x2, y1, x2, y2, x1, y2]

    return rectangle_points




# 11 矩形框的扩张

def expand_rectangle(p1, p2, offset=130):
    """
    给定矩形框的两个对角点（p1 和 p2），
    返回在四个方向扩展 offset 后的新对角点 P1 和 P2。
    """

    x1, y1, z1 = p1
    x2, y2, z2 = p2

    # 确保 p1 是左下角，p2 是右上角（即使输入反了）
    x_min, x_max = sorted([x1, x2])
    y_min, y_max = sorted([y1, y2])
    z = z1  # z 坐标保持一致

    # 四向扩展 offset，构造新的矩形框
    P1 = (x_min - offset, y_min - offset, z)
    P2 = (x_max + offset, y_max + offset, z)

    return P1, P2


# 12 矩形标准化

def parse_rectangle_points(*args):
    """
    接收多种坐标格式输入，统一解析为矩形四角点：
    返回：
        (左下, 右上, 左上, 右下)，每个为三元组 (x, y, z)
    
    合法输入形式：
        - (x1, y1, z1), (x2, y2, z2)
        - [(x1, y1, z1), (x2, y2, z2)]
        - (x1, y1, x2, y2)
        - (x1, y1, 0, x2, y2, 0)
    """
    try:
        # 解包列表输入
        if len(args) == 1 and isinstance(args[0], list):
            args = args[0]

        # 标准化为两个三维点
        if len(args) == 2 and all(isinstance(pt, (tuple, list)) and len(pt) == 3 for pt in args):
            p1, p2 = args
        elif len(args) == 4:
            x1, y1, x2, y2 = args
            p1, p2 = (x1, y1, 0), (x2, y2, 0)
        elif len(args) == 6:
            x1, y1, z1, x2, y2, z2 = args
            p1, p2 = (x1, y1, z1), (x2, y2, z2)
        else:
            raise ValueError("输入格式不合法")

        # 解析 min/max 点坐标
        x_min = min(p1[0], p2[0])
        x_max = max(p1[0], p2[0])
        y_min = min(p1[1], p2[1])
        y_max = max(p1[1], p2[1])
        z = p1[2] if len(p1) > 2 else 0

        # 四角坐标
        left_bottom = (x_min, y_min, z)
        right_top = (x_max, y_max, z)
        left_top = (x_min, y_max, z)
        right_bottom = (x_max, y_min, z)

        return left_bottom, right_top, left_top, right_bottom

    except Exception as e:
        print(f"❌ 解析矩形点失败: {e}")
        return None





#  主函数
#  (2)
#  PL打印线

#  该函数系列包括如下一些函数


#&&%  1 判断多段线是否合乎打印要求的多段线返回其图号和比例或0

def generate_name_and_ratio_from_polyline(comobj,A3dy=0,Fandy=("ISO_A3_(420.00_x_297.00_MM)","0:0","A3"),tol=10):



    """
    函数已经修改为针对任意对象，按其外包盒来分析

    当一般选择为空时自动作比例选择。不再重新编制函数和修改该函数，直接取值A3dy=1即可


    1)在使用该函数之前，应使用多段线快速选择并逐步去掉多余的PL线

    2)这一段表达这样的思想：对于那些不是很接近但误差也不是很大的情况，我们会使用已有图框中最接近的一个去作为结论，但这个可能也是错误的
    elif diff1 < 200 and diff2 < 200:
                    total_diff = diff1 + diff2  # 计算总差值
                    if total_diff < closest_diff:
                        closest_diff = total_diff
                        index_pl = i


    index_pl和closest_diff都在不断地更新直到得到最接近的.当48基元有靠近对象时，会记录下靠近痕迹，但不会影响取值。仅当其未匹配偏差到更准确的基元时菜有必要打印这个参考消息。

    另外，包含着5点合法打印线的可能。因此逻辑仍待优化，但从根本上没问题了。20250507

    3)define_rectangle_by_diagonal所定义的length就是长边，不是x方向的距离

    4)Fandy=("ISO_A3_(420.00_x_297.00_MM)","0:0","A3"),("ISO_A2_(594.00_x_420.00_MM)","0:0","A2"),("ISO_A1_(841.00_x_594.00_MM)","0:0","A1"),("ISO_A0_(1189.00_x_841.00_MM)","0:0","A0")
        
    """
    #函数返回打印多段线的图幅，比例，图号信息。对于不合法的PL线返回0值。

    # 1） 定义基本列表
    
    LB_dayingkuang = [
            (118900,84100,100),(178350,126150,150),(59450,42050,50),(29725,21025,25),
            (133763,84100,100),(200644.5,126150,150),(66881.5,42050,50),(33440.75,21025,25),
            (148625,84100,100),(222937.5,126150,150),(74312.5,42050,50),(37156.25,21025,25),
            (84100,59400,100),(126150,89100,150),(42050,29700,50),(21025,14850,25),
            (105125,59400,100),(157687.5,89100,150),(52562.5,29700,50),(26281.25,14850,25),            
            (126150,59400,100),(189225,89100,150),(63075,29700,50),(31537.5,14850,25),            
            (147175,59400,100),(220762.5,89100,150),(73587.5,29700,50),(36793.75,14850,25),
            (59400,42000,100),(89100,63000,150),(29700,21000,50),(14850,10500,25),            
            (74250,42000,100),(111375,63000,150),(37125,21000,50),(18562.5,10500,25),            
            (89100,42000,100),(133650,63000,150),(44550,21000,50),(22275,10500,25),            
            (103950,42000,100),(155925,63000,150),(51975,21000,50),(25987.5,10500,25),            
            (42000,29700,100),(63000,44550,150),(21000,14850,50),(10500,7425,25)            
        ]
    
    drawing_map_ml = [("A0","1:100"),("A0","1:150"),("A0","1:50"),("A0","1:25"),
                      ("A0+1/8","1:100"),("A0+1/8","1:150"),("A0+1/8","1:50"),("A0+1/8","1:25"),
                      ("A0+1/4","1:100"),("A0+1/4","1:150"),("A0+1/4","1:50"),("A0+1/4","1:25"),
                      ("A1","1:100"),("A1","1:150"),("A1","1:50"),("A1","1:25"),
                      ("A1+1/4","1:100"),("A1+1/4","1:150"),("A1+1/4","1:50"),("A1+1/4","1:25"),
                      ("A1+1/2","1:100"),("A1+1/4","1:150"),("A1+1/2","1:50"),("A1+1/2","1:25"),
                      ("A1+3/4","1:100"),("A1+3/4","1:150"),("A1+3/4","1:50"),("A1+3/4","1:25"),
                      ("A2","1:100"),("A2","1:150"),("A2","1:50"),("A2","1:25"),
                      ("A2+1/4","1:100"),("A2+1/4","1:150"),("A2+1/4","1:50"),("A2+1/4","1:25"),
                      ("A2+1/2","1:100"),("A2+1/2","1:150"),("A2+1/2","1:50"),("A2+1/2","1:25"),
                      ("A2+3/4","1:100"),("A2+3/4","1:150"),("A2+3/4","1:50"),("A2+3/4","1:25"),
                      ("A3","1:100"),("A3","1:150"),("A3","1:50"),("A3","1:25")
        ]                      
                      
    drawing_map = ["ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)",
                   "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)","UserDefinedMetric (1337.63 x 841.00毫米)","UserDefinedMetric (1337.63 x 841.00毫米)",
                   "UserDefinedMetric (1486.25 x 841.00毫米)","UserDefinedMetric (1486.25 x 841.00毫米)","UserDefinedMetric (1486.25 x 841.00毫米)","UserDefinedMetric (1486.25 x 841.00毫米)",
                   "ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)",
                   "UserDefinedMetric (1051.25 x 594.00毫米)","UserDefinedMetric (1051.25 x 594.00毫米)","UserDefinedMetric (1051.25 x 594.00毫米)","UserDefinedMetric (1051.25 x 594.00毫米)",
                   "UserDefinedMetric (1261.50 x 594.00毫米)","UserDefinedMetric (1261.50 x 594.00毫米)","UserDefinedMetric (1261.50 x 594.00毫米)","UserDefinedMetric (1261.50 x 594.00毫米)",
                   "UserDefinedMetric (1471.75 x 594.00毫米)","UserDefinedMetric (1471.75 x 594.00毫米)","UserDefinedMetric (1471.75 x 594.00毫米)","UserDefinedMetric (1471.75 x 594.00毫米)",
                   "ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)",
                   "UserDefinedMetric (742.50 x 420.00毫米)","UserDefinedMetric (742.50 x 420.00毫米)","UserDefinedMetric (742.50 x 420.00毫米)","UserDefinedMetric (742.50 x 420.00毫米)",                   
                   "UserDefinedMetric (891.00 x 420.00毫米)","UserDefinedMetric (891.00 x 420.00毫米)","UserDefinedMetric (891.00 x 420.00毫米)","UserDefinedMetric (891.00 x 420.00毫米)",
                   "UserDefinedMetric (1039.50 x 420.00毫米)","UserDefinedMetric (1039.50 x 420.00毫米)","UserDefinedMetric (1039.50 x 420.00毫米)","UserDefinedMetric (1039.50 x 420.00毫米)",
                   "ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)"]
                      
    # 2） 确定多段线的长与宽（长度值大的为length）
    PL_min = find_min_point(comobj)
    
    PL_max = find_max_point(comobj)
    
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)


    # 3） 确定序号

    index_pl = ""
    closest_diff = float('inf')  # 初始化最接近的差值为无穷大
    if A3dy == 0:

        print("length,width",length,width)
        for i in range(len(LB_dayingkuang)):
            obj = LB_dayingkuang[i]
            obj_length, obj_width, _ = obj
            diff1 = abs(length - obj_length)
            diff2 = abs(width - obj_width)

            if diff1 < tol and diff2 < tol:
                index_pl = i
                break  # 如果找到了非常接近的，就不再继续查找


            

            elif diff1 < 200 and diff2 < 200:

                print("偏差值在10--200之间,请注意该尺寸")
                total_diff = diff1 + diff2  # 计算总差值
                if total_diff < closest_diff:
                    closest_diff = total_diff
                    index_pl = i

        # 退出循环后统一判断
        if index_pl != "":

            print("位置：",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            print("数据：",drawing_map[index_pl], drawing_map_ml[index_pl][1], drawing_map_ml[index_pl][0])
            
            return drawing_map[index_pl], drawing_map_ml[index_pl][1], drawing_map_ml[index_pl][0]
        else:


            print("位置：",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            #对原0值对象加广义判断

            fanhui = get_print_template_info(comobj, tol=tol)

            return fanhui 

    # 4） 返回值

    elif A3dy == 1:

        print("length,width",length,width)
        
        if abs(width/length - 0.707) <= 0.01:

            print("位置：",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            return Fandy[0],Fandy[1],Fandy[2]

        else:            

            print("位置：",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            return 0

    else:

        print("参数输入错误")

        return 0
##新的四元信息分析检测函数

def generate_name_and_ratio_from_com(
    comobj,
    A3dy=0,
    Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3",0),
    tol=10
):
    """
    函数已经修改为针对任意对象，按其外包盒来分析。
    返回值由原来的三元组 (图幅, 比例, 图号) 扩展为四元组 (图幅, 比例, 图号, 竖向标志)：
      - 竖向标志 = 1 表示外包盒竖向 (height > width)
      - 否则返回 0（横向或正方形）

    参数：
      comobj -- 任意支持 GetBoundingBox() 的多段线 COM 对象
      A3dy   -- 当 A3dy=1 时，直接按照 Fandy 返回。否则使用 LB_dayingkuang 进行匹配。
      Fandy  -- 默认 ("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3",0)，规模 A3 情况下返回的信息
      tol    -- 与标准尺寸比较的容差，默认为 10

    返回：
      (图幅, 比例, 图号, 竖向标志) 或者 0
    """
    # 1）定义基本尺寸库
    LB_dayingkuang = [
        (118900, 84100, 100),   (178350, 126150, 150),  (59450, 42050, 50),   (29725, 21025, 25),
        (133763, 84100, 100),   (200644.5, 126150, 150), (66881.5, 42050, 50), (33440.75, 21025, 25),
        (148625, 84100, 100),   (222937.5, 126150, 150),(74312.5, 42050, 50), (37156.25, 21025, 25),
        (84100, 59400, 100),    (126150, 89100, 150),   (42050, 29700, 50),   (21025, 14850, 25),
        (105125, 59400, 100),   (157687.5, 89100, 150), (52562.5, 29700, 50), (26281.25, 14850, 25),
        (126150, 59400, 100),   (189225, 89100, 150),   (63075, 29700, 50),   (31537.5, 14850, 25),
        (147175, 59400, 100),   (220762.5, 89100, 150), (73587.5, 29700, 50), (36793.75, 14850, 25),
        (59400, 42000, 100),    (89100, 63000, 150),    (29700, 21000, 50),   (14850, 10500, 25),
        (74250, 42000, 100),    (111375, 63000, 150),   (37125, 21000, 50),   (18562.5, 10500, 25),
        (89100, 42000, 100),    (133650, 63000, 150),   (44550, 21000, 50),   (22275, 10500, 25),
        (103950, 42000, 100),   (155925, 63000, 150),   (51975, 21000, 50),   (25987.5, 10500, 25),
        (42000, 29700, 100),    (63000, 44550, 150),    (21000, 14850, 50),   (10500, 7425, 25)
    ]

    drawing_map_ml = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"),  ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"),  ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"),  ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"),  ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"),  ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"),  ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"),  ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"),  ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"),  ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"),  ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"),  ("A3", "1:25")
    ]

    drawing_map = [
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)",
        "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)",
        "UserDefinedMetric (1486.25 x 841.00毫米)", "UserDefinedMetric (1486.25 x 841.00毫米)",
        "UserDefinedMetric (1486.25 x 841.00毫米)", "UserDefinedMetric (1486.25 x 841.00毫米)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "UserDefinedMetric (1051.25 x 594.00毫米)", "UserDefinedMetric (1051.25 x 594.00毫米)",
        "UserDefinedMetric (1051.25 x 594.00毫米)", "UserDefinedMetric (1051.25 x 594.00毫米)",
        "UserDefinedMetric (1261.50 x 594.00毫米)", "UserDefinedMetric (1261.50 x 594.00毫米)",
        "UserDefinedMetric (1261.50 x 594.00毫米)", "UserDefinedMetric (1261.50 x 594.00毫米)",
        "UserDefinedMetric (1471.75 x 594.00毫米)", "UserDefinedMetric (1471.75 x 594.00毫米)",
        "UserDefinedMetric (1471.75 x 594.00毫米)", "UserDefinedMetric (1471.75 x 594.00毫米)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "UserDefinedMetric (742.50 x 420.00毫米)", "UserDefinedMetric (742.50 x 420.00毫米)",
        "UserDefinedMetric (742.50 x 420.00毫米)", "UserDefinedMetric (742.50 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (1039.50 x 420.00毫米)", "UserDefinedMetric (1039.50 x 420.00毫米)",
        "UserDefinedMetric (1039.50 x 420.00毫米)", "UserDefinedMetric (1039.50 x 420.00毫米)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)"
    ]

    # 2）计算外包盒的长、宽，以及横竖信息
    PL_min = find_min_point(comobj)
    PL_max = find_max_point(comobj)
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)

    # 计算 dx, dy 以判断竖向或横向
    dx = abs(PL_max[0] - PL_min[0])
    dy = abs(PL_max[1] - PL_min[1])
    orientation_flag = 1 if dy > dx else 0

    # 3）确定匹配序号 index_pl
    index_pl = ""
    closest_diff = float('inf')

    if A3dy == 0:
        print("length, width =", length, width)
        for i, (obj_length, obj_width, _) in enumerate(LB_dayingkuang):
            diff1 = abs(length - obj_length)
            diff2 = abs(width - obj_width)

            # 完全匹配
            if diff1 < tol and diff2 < tol:
                index_pl = i
                break

            # 近似匹配的候选
            elif diff1 < 200 and diff2 < 200:
                total_diff = diff1 + diff2
                if total_diff < closest_diff:
                    closest_diff = total_diff
                    index_pl = i

        # 判断并返回
        if index_pl != "":
            print("位置：", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return (
                drawing_map[index_pl],
                drawing_map_ml[index_pl][1],
                drawing_map_ml[index_pl][0],
                orientation_flag
            )
        else:
            print("位置：", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            fanhui = get_print_template_info(comobj, tol=tol)
            # 如果 fanhui 是三元组，则拼接 orientation_flag；否则返回 0
            if isinstance(fanhui, tuple) and len(fanhui) == 3:
                return (*fanhui, orientation_flag)
            return 0

    # 4）当 A3dy == 1 时，直接根据长宽比返回 Fandy 并附加 orientation_flag
    elif A3dy == 1:
        print("length, width =", length, width)
        if abs(width / length - 0.707) <= 0.01:
            print("位置：", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return (*Fandy, orientation_flag)
        else:
            print("位置：", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return 0

    else:
        print("参数输入错误")
        return 0


##20250711修改
def generate_name_and_ratio_from_com(
    comobj,
    A3dy=0,
    Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0),
    tol=10
):
    """
    返回 (图幅, 比例, 图号, 竖向标志)。未命中时返回 0。

    ① 精确命中 → 原样返回  
    ② 命中 “×1.2”  → 返回原标准数据，并把对象改成蓝色  
    ③ 近似命中     → 返回原标准数据，并把对象改成红色 + 宽度 (200 / 2)

    对拟合标准框线，加了红色警告和加粗提示，对放大1.2倍的打印框线加了蓝色提示

    """

    # ————————————— 1. 基础数据区 —————————————
    LB_dayingkuang = [
        (118900, 84100, 100),   (178350, 126150, 150),  (59450, 42050, 50),   (29725, 21025, 25),
        (133763, 84100, 100),   (200644.5, 126150, 150), (66881.5, 42050, 50), (33440.75, 21025, 25),
        (148625, 84100, 100),   (222937.5, 126150, 150),(74312.5, 42050, 50), (37156.25, 21025, 25),
        (84100, 59400, 100),    (126150, 89100, 150),   (42050, 29700, 50),   (21025, 14850, 25),
        (105125, 59400, 100),   (157687.5, 89100, 150), (52562.5, 29700, 50), (26281.25, 14850, 25),
        (126150, 59400, 100),   (189225, 89100, 150),   (63075, 29700, 50),   (31537.5, 14850, 25),
        (147175, 59400, 100),   (220762.5, 89100, 150), (73587.5, 29700, 50), (36793.75, 14850, 25),
        (59400, 42000, 100),    (89100, 63000, 150),    (29700, 21000, 50),   (14850, 10500, 25),
        (74250, 42000, 100),    (111375, 63000, 150),   (37125, 21000, 50),   (18562.5, 10500, 25),
        (89100, 42000, 100),    (133650, 63000, 150),   (44550, 21000, 50),   (22275, 10500, 25),
        (103950, 42000, 100),   (155925, 63000, 150),   (51975, 21000, 50),   (25987.5, 10500, 25),
        (42000, 29700, 100),    (63000, 44550, 150),    (21000, 14850, 50),   (10500, 7425, 25)
    ]

    drawing_map_ml = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"),  ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"),  ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"),  ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"),  ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"),  ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"),  ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"),  ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"),  ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"),  ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"),  ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"),  ("A3", "1:25")
    ]

    drawing_map = [
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)",
        "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)",
        "UserDefinedMetric (1486.25 x 841.00毫米)", "UserDefinedMetric (1486.25 x 841.00毫米)",
        "UserDefinedMetric (1486.25 x 841.00毫米)", "UserDefinedMetric (1486.25 x 841.00毫米)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "UserDefinedMetric (1051.25 x 594.00毫米)", "UserDefinedMetric (1051.25 x 594.00毫米)",
        "UserDefinedMetric (1051.25 x 594.00毫米)", "UserDefinedMetric (1051.25 x 594.00毫米)",
        "UserDefinedMetric (1261.50 x 594.00毫米)", "UserDefinedMetric (1261.50 x 594.00毫米)",
        "UserDefinedMetric (1261.50 x 594.00毫米)", "UserDefinedMetric (1261.50 x 594.00毫米)",
        "UserDefinedMetric (1471.75 x 594.00毫米)", "UserDefinedMetric (1471.75 x 594.00毫米)",
        "UserDefinedMetric (1471.75 x 594.00毫米)", "UserDefinedMetric (1471.75 x 594.00毫米)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "UserDefinedMetric (742.50 x 420.00毫米)", "UserDefinedMetric (742.50 x 420.00毫米)",
        "UserDefinedMetric (742.50 x 420.00毫米)", "UserDefinedMetric (742.50 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (1039.50 x 420.00毫米)", "UserDefinedMetric (1039.50 x 420.00毫米)",
        "UserDefinedMetric (1039.50 x 420.00毫米)", "UserDefinedMetric (1039.50 x 420.00毫米)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)"
    ]

    # —————————— 2. 外包盒与朝向 ——————————
    PL_min = find_min_point(comobj)
    PL_max = find_max_point(comobj)
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)
    dx, dy = abs(PL_max[0] - PL_min[0]), abs(PL_max[1] - PL_min[1])
    orientation_flag = 1 if dy > dx else 0

    # —————————— 3. 匹配类型判定 ——————————
    best_i, best_type = None, None      # 'exact' | 'scale12' | 'approx'
    best_score = float('inf')           # 仅用于 approx

    if A3dy == 0:
        for i, (std_len, std_wid, _) in enumerate(LB_dayingkuang):
            # 精确
            if abs(length - std_len) < tol and abs(width - std_wid) < tol:
                best_i, best_type = i, 'exact'
                break

            # ×1.2
            if abs(length - 1.2*std_len) < tol and abs(width - 1.2*std_wid) < tol:
                best_i, best_type = i, 'scale12'
                break

            # 近似
            diff = max(abs(length - std_len), abs(width - std_wid))
            if diff < 200 and diff < best_score:
                best_i, best_type = i, 'approx'
                best_score = diff

        # —————————— 4. 命中后的处理 ——————————
        if best_i is not None:
            zhi = (
                drawing_map[best_i],
                drawing_map_ml[best_i][1],
                drawing_map_ml[best_i][0],
                orientation_flag
            )

            if best_type == 'scale12':          # 蓝
                try:
                    comobj.Color = 5
                except Exception:
                    pass

            elif best_type == 'approx':         # 红 + 加粗
                try:
                    comobj.Color = 1
                except Exception:
                    pass
                try:
                    total_len = float(getattr(comobj, "Length", 0))
                except Exception:
                    total_len = 0
                new_wid = 2 if total_len and total_len < 10000 else 200
                try:
                    comobj.ConstantWidth = new_wid
                except Exception:
                    try:
                        comobj.SetWidth(new_wid, new_wid)
                    except Exception:
                        pass

            return zhi

        # —— 未命中：调用外部模板判定 —— 
        fanhui = get_print_template_info(comobj, tol=tol)
        return (*fanhui, orientation_flag) if isinstance(fanhui, tuple) else 0

    # —————————— 5. A3dy == 1 直接用长宽比判定 ——————————
    elif A3dy == 1:
        if abs(width / length - 0.707) <= 0.01:
            return (*Fandy, orientation_flag)
        return 0

    # —————————— 6. 参数错误 ——————————
    else:
        print("参数输入错误")
        return 0





#  2 com复线从上到下从左到右排序


def polyline_sort(polyline_list):
    """对com多段线按照特定规则进行排序"""

    # 存储多段线及其最左下角点
    polylines_with_min_points = [(pl, find_min_point(pl)) for pl in polyline_list]

    # 先按照y值降序排序
    polylines_with_min_points.sort(key=lambda item: -item[1][1])

    # 再对y值差距在1000以内的多段线按照x值升序排序
    i = 0
    while i < len(polylines_with_min_points) - 1:
        j = i + 1
        # 查找所有y值差距在1000以内的多段线
        while j < len(polylines_with_min_points) and abs(polylines_with_min_points[i][1][1] - polylines_with_min_points[j][1][1]) < 1000:
            j += 1
        
        # 如果找到了y值相近的多段线，根据x值进行排序
        if j - i > 1:
            polylines_with_min_points[i:j] = sorted(polylines_with_min_points[i:j], key=lambda item: item[1][0])
        
        i = j

    # 只返回多段线对象
    return [item[0] for item in polylines_with_min_points]




#&&&%  *** 3 将PLcom线列表的坐标信息存储


def plcom_to_coor(plines):
    """
    接受多根轻量级多段线或常规多段线的 COM 对象列表，返回它们的坐标列表及闭合状态。

    :param plines: 可迭代的一组 LWPOLYLINE 或 POLYLINE COM 对象
    :return: 列表，每个元素为 (pts, closed_flag)：
             - pts: 顶点列表 [(x0, y0), (x1, y1), …]
             - closed_flag: 1 表示闭合，0 表示未闭合

    兼容两种情况：
      1. LWPOLYLINE.Coordinates → 偶数长度，例如 [x0,y0, x1,y1, x2,y2, …]
      2. POLYLINE.Coordinates  → 3 的倍数长度，例如 [x0,y0,z0, x1,y1,z1, …]

    如果既不是偶数也不是 3 的倍数，将跳过该条多段线并打印 WARN 提示。
    """
    plines = ensure_list(plines)
    all_info = []

    for pl in plines:
        raw = list(pl.Coordinates)  # 可能是偶数长度 (LWPOLYLINE) 或 3 的倍数长度 (POLYLINE)

        pts = []
        # —— 情况 A：如果长度能被 3 整除，认为是常规 POLYLINE → 每 3 个数一组 (x,y,z)
        if len(raw) % 3 == 0 and len(raw) > 0:
            for i in range(0, len(raw), 3):
                x = raw[i]
                y = raw[i + 1]
                pts.append((x, y))

        # —— 情况 B：否则如果长度能被 2 整除，认为是轻量级 LWPOLYLINE → 每 2 个数一组 (x,y)
        elif len(raw) % 2 == 0 and len(raw) > 0:
            for i in range(0, len(raw), 2):
                x = raw[i]
                y = raw[i + 1]
                pts.append((x, y))

        else:
            # 既不是 2 的倍数也不是 3 的倍数：坐标数据异常，跳过这一条，打印 WARN
            handle = getattr(pl, "Handle", "<unknown>")
            print(f"[WARN] plcom_to_coor：跳过 Handle={handle} 的多段线，"
                  f"Coordinates 长度={len(raw)} 既非 2 的倍数也非 3 的倍数。")
            continue

        # 读取 Closed 属性，True 表示闭合
        closed_flag = 1 if getattr(pl, "Closed", False) else 0

        all_info.append((pts, closed_flag))

    return all_info


# 4 从坐标信息列表返回PLcom线列表

    
def plcoor_to_com(coord_info, layer_name="测试辅助", width=0, color=256):
    """
    在当前 DWG 中根据坐标和封闭标志绘制多条轻量级多段线。

    :param coord_info: 列表，每个元素为 (pts, closed_flag)，
                       pts 为 [(x0,y0),…] 顶点列表，
                       closed_flag 为 1（闭合）或 0（不闭合）。
    :param layer_name: 目标图层名称（不存在则创建），默认 "测试辅助"
    :param width:      多段线宽度，默认 0
    :param color:      颜色索引，默认 256（BYLAYER）
    :return:           绘制的多段线对象列表
    """
    # 1) 连接 AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    ms   = doc.ModelSpace

    # 2) 确保图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    lyr.LayerOn = True

    created = []
    for pts, closed_flag in coord_info:
        # 将 pts 展平为 [x0,y0,x1,y1,…]
        raw = []
        for x, y in pts:
            raw.extend((x, y))
        # 转为 COM 数组
        arr = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            raw
        )
        # 添加轻量级多段线
        lw = ms.AddLightWeightPolyline(arr)
        lw.Layer         = layer_name
        lw.ConstantWidth = width
        lw.Color         = color
        lw.Closed        = bool(closed_flag)
        created.append(lw)

    # 可选：缩放到可见范围
    acad.ZoomExtents()
    print(f"✅ 已绘制 {len(created)} 条轻量级多段线到图层 “{layer_name}”")
    return created






# 5 确定多段线打印框是否竖向

def panduan_shuxiangkuang(polyline):

    PL_min = find_min_point(polyline)
    
    PL_max = find_max_point(polyline)

    cha_x  = abs(PL_max[0] - PL_min[0])   

    cha_y  = abs(PL_max[1] - PL_min[1])

    if  cha_y > cha_x:

        return True

    else :

        return False


# 6 统一为A2的图幅打印

def tongyi_tufu(LB,TFname):


    """
    将打印线列表的每根线对应的图纸尺寸统一为一个TFname
        
    """

    LB_xin = []

    for ob in LB:

        LB_xin.append(TFname)#"ISO_A2_(594.00_x_420.00_MM)","ISO_A3_(420.00_x_297.00_MM)"

    return LB_xin



#  主函数
#  (3)
#  将正交六边形多段线分成两个矩形区域

#  该函数系列包括如下一些函数

"""
该函数将正交六边形多段线分成两个矩形区域，只针对标准六点六边的PL多边形
在处理六点六边的正交PL形多边形之前，使用simplify_polygon(polygon, tol=0.5)将伪边点清除
            
"""

# 消除多边形的伪边点

def simplify_polygon(poly, tol=0.5):
    """
    简化多边形顶点列表：如果某顶点 P 与其前后两点共线（在容差 tol 范围内），则将其移除。
    首尾相连处理：第一个点的“前点”是最后一个点，最后一个点的“次点”是第一个点。
    
    参数:
        poly: [(x, y, z), …]  原始顶点列表（可能首尾重复或有多余顶点）
        tol:  共线判断的容差（对应叉积绝对值）
    
    返回:
        简化后的顶点列表（同样保留首尾是否闭合的形式，不会自动去重首尾）
    """
    # 先做一次“去掉首尾重复”以免无限循环
    if len(poly) > 1 and poly[0] == poly[-1]:
        poly = poly[:-1]

    def is_colinear(p_prev, p, p_next):
        # 只比较 x,y，计算 (p - p_prev) × (p_next - p) 的“叉积”
        x1, y1 = p[0] - p_prev[0], p[1] - p_prev[1]
        x2, y2 = p_next[0] - p[0],   p_next[1] - p[1]
        cross = x1 * y2 - y1 * x2
        return abs(cross) <= tol

    # 重复扫一遍能删就删，直到一轮下来没有删除
    changed = True
    while changed and len(poly) >= 3:
        changed = False
        n = len(poly)
        for i in range(n):
            prev_idx = (i - 1) % n
            next_idx = (i + 1) % n
            if is_colinear(poly[prev_idx], poly[i], poly[next_idx]):
                # 删除第 i 个点，重置一轮扫描
                del poly[i]
                changed = True
                break
    return poly



# 1. 标准化多边形顶点列表，去掉相邻和首尾重复点
def normalize_polygon(polygon):
    """
    标准化多边形顶点列表：  
      - 去掉任意相邻重复的点  
      - 如果首尾相同，则去掉末尾那个  
      
    参数:
        polygon: 原始顶点列表，每个点为 (x, y, z)
    返回:
        去重后的顶点列表
    """
    if not polygon:
        return []
    normalized = [polygon[0]]
    for pt in polygon[1:]:
        if pt != normalized[-1]:
            normalized.append(pt)
    # 首尾相同则删尾
    if len(normalized) > 1 and normalized[0] == normalized[-1]:
        normalized.pop()
    return normalized


# 2. 找到某顶点的前驱/后继（按循环多边形）
def get_adjacent_points(polygon, p):
    """
    在多边形 polygon 中返回顶点 p 的前后相邻点（支持循环）。
    会先调用 normalize_polygon 清理重复点。
    """
    poly = normalize_polygon(polygon)
    if not poly:
        raise ValueError("多边形为空")
    try:
        idx = poly.index(p)
    except ValueError:
        raise ValueError(f"点 {p} 不在多边形顶点列表中")
    prev_pt = poly[idx - 1] if idx > 0 else poly[-1]
    next_pt = poly[idx + 1] if idx < len(poly) - 1 else poly[0]
    return prev_pt, next_pt


# 3. 点是否在多边形内部（射线法，仅在 XY 平面判断）
def point_in_polygon(pt, polygon):
    """
    判断三维点 pt=(x,y,z) 在多边形 polygon 的 XY 投影内否。
    polygon 中点格式为 (x,y,z)，首尾可重复或不重复均可。
    """
    x, y, _ = pt
    poly2d = [(q[0], q[1]) for q in normalize_polygon(polygon)]
    inside = False
    n = len(poly2d)
    for i in range(n):
        x1, y1 = poly2d[i]
        x2, y2 = poly2d[(i + 1) % n]
        if (y1 > y) != (y2 > y):
            xint = x1 + (y - y1) * (x2 - x1) / (y2 - y1)
            if xint > x:
                inside = not inside
    return inside


# 4. 无穷直线 vs 线段在 XY 平面相交
def line_segment_intersection_2d(p, d, a, b, tol=1e-8):
    """
    计算射线 L(t)=p + t·d 与线段 AB 在 XY 平面上的交点。
    p,d,a,b 皆为 (x,y,z)，但只取 x,y 分量参与计算。
    返回 (xi, yi, t) 或 None。
    """
    px, py, _ = p
    dx, dy, _ = d
    ax, ay, _ = a
    bx, by, _ = b
    ux, uy = bx - ax, by - ay

    det = dx * (-uy) - dy * (-ux)
    if abs(det) < tol:
        return None

    rhsx, rhsy = ax - px, ay - py
    t_param = (rhsx * (-uy) - rhsy * (-ux)) / det
    u_param = (dx * rhsy - dy * rhsx) / det

    if -tol <= u_param <= 1 + tol:
        xi = px + t_param * dx
        yi = py + t_param * dy
        return xi, yi, t_param
    return None


# 5. 计算 p 和其相邻点中点 c，如果 c 内部则返回 c，否则沿 p->c 的射线找到第一个交点
def get_auxiliary_point(p, p_prev, p_next, polygon, tol=1e-8):
    """
    对于多边形顶点 p 及其前后相邻点 p_prev, p_next，
    返回一个位于多边形内部的辅助点 q：
      1. 先取 c = (p_prev + p_next)/2；若 c 在内部，则返回 c
      2. 否则沿射线 p->c 与多边形其它边求最靠近 p 的交点
    返回点格式为 (x,y,z)。
    """
    # 1) 中点 c
    cx = (p_prev[0] + p_next[0]) / 2
    cy = (p_prev[1] + p_next[1]) / 2
    cz = (p_prev[2] + p_next[2]) / 2
    c = (cx, cy, cz)

    if point_in_polygon(c, polygon):
        return c

    # 2) 构造方向 d = c - p 并归一化
    dx, dy = cx - p[0], cy - p[1]
    mag = math.hypot(dx, dy)
    if mag < tol:
        raise RuntimeError("辅助点方向向量过小")
    d = (dx / mag, dy / mag, 0.0)

    # 在每条不含 p 的边上求交
    poly = normalize_polygon(polygon)
    intersects = []
    for i in range(len(poly)):
        a = poly[i]
        b = poly[(i + 1) % len(poly)]
        if a == p or b == p:
            continue
        res = line_segment_intersection_2d(p, d, a, b, tol)
        if res:
            xi, yi, t_param = res
            if abs(t_param) > tol:
                # 插值出对应的 z
                zi = p[2] + t_param * d[2]
                intersects.append((xi, yi, zi, t_param))

    if not intersects:
        raise RuntimeError("未找到有效交点")
    # 取最小 |t| 的那一个
    intersects.sort(key=lambda it: abs(it[3]))
    xi, yi, zi, _ = intersects[0]
    return (xi, yi, zi)


# 6. 计算 p 点的“凹凸度量角”
def concavity_measure(p, p_prev, p_next, q):
    """
    给定 p, p_prev, p_next, q（均为 (x,y,z)），
    计算度量角：
      angle = 360 - larger_angle + smaller_angle  
    其中 smaller/larger 是 pq→p_prev 与 pq→p_next 的逆时针夹角。
    凸点约 ~90°，凹点约 ~270°。
    """
    def angle_of(vx, vy):
        a = math.degrees(math.atan2(vy, vx))
        return a if a >= 0 else a + 360

    # 构造 2D 向量
    vq = (q[0] - p[0], q[1] - p[1])
    v1 = (p_prev[0] - p[0], p_prev[1] - p[1])
    v2 = (p_next[0] - p[0], p_next[1] - p[1])

    a_q = angle_of(*vq)
    a1  = angle_of(*v1)
    a2  = angle_of(*v2)

    d1 = (a1 - a_q) % 360
    d2 = (a2 - a_q) % 360

    small, large = (d1, d2) if d1 < d2 else (d2, d1)
    return 360 - large + small


# 7. 直接给出 p 在多边形上的度量角
def concavity_angle(p, polygon):
    """
    直接计算多边形 polygon 上顶点 p 的凹凸度量角。
    """
    p_prev, p_next = get_adjacent_points(polygon, p)
    q = get_auxiliary_point(p, p_prev, p_next, polygon)
    return concavity_measure(p, p_prev, p_next, q)





# 8.合理分割PL正交六边形

def split_orthogonal_hexagon(polygon, tol=0.1):#水平分割
    """
    将正交六边形 polygon 按凹顶点所在水平线切成两个矩形。
    polygon: 6 个 (x,y,z) 顶点的列表，允许首尾重合或相邻重复，会自动规范化。
    """
    # 1. 规范化，去掉相邻重复和首尾同点
    poly = normalize_polygon(polygon)
    if len(poly) != 6:
        raise ValueError("必须传入6点正交多边形")
    # 2. 找出唯一的凹点 p
    concaves = [pt for pt in poly
                if abs(concavity_angle(pt, poly) - 270) < tol]
    if len(concaves) != 1:
        raise RuntimeError(f"没能唯一定位凹点，找到 {len(concaves)} 个")
    p = concaves[0]
    y0 = p[1]

    # 3. 只对真正“跨越” y=y0 的边求交
    intersections = []
    n = len(poly)
    for i in range(n):
        a = poly[i]
        b = poly[(i+1) % n]
        y1, y2 = a[1], b[1]
        # 仅当严格跨越才算：一端在上 (y>y0)，一端在下 (y<y0)
        if (y1 - y0) * (y2 - y0) < -tol**2:
            t = (y0 - y1) / (y2 - y1)
            xi = a[0] + t * (b[0] - a[0])
            zi = a[2] + t * (b[2] - a[2])
            intersections.append((xi, y0, zi, i))

    # 4. 应该只剩一个真的 crossing 交点
    if len(intersections) != 1:
        raise RuntimeError(f"没能唯一定位 q，找到 {len(intersections)} 个候选点")
    xi, yi, zi, edge_idx = intersections[0]
    q = (xi, yi, zi)

    # 5. 把 q 插回那条边之后
    newpoly = []
    for i in range(n):
        newpoly.append(poly[i])
        if i == edge_idx:
            newpoly.append(q)
    # now len(newpoly)==7

    # 6. 找 p, q 的索引
    i_p = newpoly.index(p)
    i_q = newpoly.index(q)

    # 7. 分割成两段多边形
    if i_q < i_p:
        rect1 = newpoly[i_q:i_p+1]
        rect2 = newpoly[i_p:] + newpoly[:i_q+1]
    else:
        rect1 = newpoly[i_p:i_q+1]
        rect2 = newpoly[i_q:] + newpoly[:i_p+1]

    # 8. 计算2D面积
    def area2d(pts):
        s = 0
        m = len(pts)
        for j in range(m):
            x1,y1,_ = pts[j]
            x2,y2,_ = pts[(j+1)%m]
            s += x1*y2 - x2*y1
        return abs(s)/2

    A1, A2 = area2d(rect1), area2d(rect2)
    # 面积小的放前面
    return (rect1, rect2) if A1 <= A2 else (rect2, rect1)


def split_orthogonal_hexagon_vertical(polygon, tol=0.1):#竖向分割
    """
    将正交六边形 polygon 按凹顶点所在竖线切成两个矩形。
    polygon: 6 个 (x,y,z) 顶点的列表，允许首尾重合或相邻重复，会自动规范化。
    tol: 用于识别凹点和跨越判断的容差。
    返回: (rect1, rect2)，面积小的放前面。
    """
    # 规范化：去掉相邻重复和首尾同点
    poly = normalize_polygon(polygon)
    if len(poly) != 6:
        raise ValueError("必须传入6点正交多边形")

    # 找唯一凹点
    concaves = [pt for pt in poly
                if abs(concavity_angle(pt, poly) - 270) < tol]
    if len(concaves) != 1:
        raise RuntimeError(f"没能唯一定位凹点，找到 {len(concaves)} 个")
    p = concaves[0]
    x0 = p[0]

    # 求竖线 x=x0 与真正跨越边的交点
    intersections = []
    n = len(poly)
    for i in range(n):
        a = poly[i]
        b = poly[(i+1)%n]
        x1, x2 = a[0], b[0]
        # 仅对严格跨越的边求交
        if (x1 - x0)*(x2 - x0) < -tol**2:
            # 插值比例
            t = (x0 - x1)/(x2 - x1)
            yi = a[1] + t*(b[1] - a[1])
            zi = a[2] + t*(b[2] - a[2])
            intersections.append((x0, yi, zi, i))

    if len(intersections) != 1:
        raise RuntimeError(f"没能唯一定位 q，找到 {len(intersections)} 个候选点")
    xi, yi, zi, edge_idx = intersections[0]
    q = (xi, yi, zi)

    # 把 q 插回那条边
    newpoly = []
    for i in range(n):
        newpoly.append(poly[i])
        if i == edge_idx:
            newpoly.append(q)
    # newpoly 长度应为7

    # 定位 p、q 索引
    i_p = newpoly.index(p)
    i_q = newpoly.index(q)

    # 分割两段
    if i_q < i_p:
        rect1 = newpoly[i_q:i_p+1]
        rect2 = newpoly[i_p:] + newpoly[:i_q+1]
    else:
        rect1 = newpoly[i_p:i_q+1]
        rect2 = newpoly[i_q:] + newpoly[:i_p+1]

    # 面积计算（2D）
    def area2d(pts):
        s = 0
        m = len(pts)
        for j in range(m):
            x1,y1,_ = pts[j]
            x2,y2,_ = pts[(j+1)%m]
            s += x1*y2 - x2*y1
        return abs(s)*0.5

    A1, A2 = area2d(rect1), area2d(rect2)
    return (rect1, rect2) if A1 <= A2 else (rect2, rect1)


# 合理分割PL正交六边形

def area_of(verts):
    """多边形面积计算（顶点首尾闭合或不闭合均可）"""
    s = 0
    n = len(verts)
    for i in range(n):
        x1, y1, *_ = verts[i]
        x2, y2, *_ = verts[(i+1) % n]
        s += x1 * y2 - x2 * y1
    return abs(s) * 0.5



def split_hexagon_combined(polygon, tol=0.1, simplify_tol=0.5):# 合理分割PL正交六边形
    """
    合理分割一个正交（近似）六边形 PL：
      1) 如果传入的是 COM PL 对象，则先提取唯一顶点列表；
      2) 对顶点列表做简化（去除伪顶点/伪边）；
      3) 先做横向分割，再做竖向分割，比较两种分割最小矩形面积，
         将最小的那对矩形排在前面，返回四个矩形顶点列表：
         [min1, partner1, min2, partner2]
    参数:
      polygon: 要分割的多段线，既可以是 [(x,y,z),...] 顶点列表，也可以是 COM PL 对象
      tol: 分割时判断凹角的容差
      simplify_tol: 简化多边形时去除“伪顶点/伪边”的容差
    返回:
      四个矩形的顶点列表：最小矩形、其同组矩形、次小矩形、其同组矩形
    """
    # —— 1. 如果是 COM PL 对象，先提取顶点列表 —— 
    # （假定已有 get_unique_vertices_from_pl_com(com_pl) -> [(x,y,z),...]）
    if not isinstance(polygon, list):
        polygon = get_unique_vertices_from_pl_com(polygon)

    # —— 2. 简化顶点列表 —— 
    # （假定已有 simplify_polygon(verts, tol) -> 清理后的顶点列表）
    polygon = simplify_polygon(polygon, simplify_tol)

    # —— 3. 横向与竖向分割 —— 
    rects_h = split_orthogonal_hexagon(polygon, tol)
    rects_v = split_orthogonal_hexagon_vertical(polygon, tol)

    A_h1, A_h2 = rects_h
    A_v1, A_v2 = rects_v

    # 计算每对的“更小”矩形和“配对”矩形
    if area_of(A_h1) <= area_of(A_h2):
        min_h, partner_h = A_h1, A_h2
    else:
        min_h, partner_h = A_h2, A_h1

    if area_of(A_v1) <= area_of(A_v2):
        min_v, partner_v = A_v1, A_v2
    else:
        min_v, partner_v = A_v2, A_v1

    # —— 4. 根据最小面积值决定输出顺序 —— 
    if area_of(min_h) <= area_of(min_v):
        return [min_h, partner_h, min_v, partner_v]
    else:
        return [min_v, partner_v, min_h, partner_h]



#  主函数
#  (4)
#  获取多段线的上下左右边界的直线段，返回线段端点列表

#  该函数系列包括如下一些函数

def get_bbox_edge_segments(pl, tol=0.5):
    """
    获取对象 pl 的包围盒四条边，分别作为独立的列表返回：
      top    = [(xmin, ymax, z), (xmax, ymax, z)]
      bottom = [(xmin, ymin, z), (xmax, ymin, z)]
      left   = [(xmin, ymin, z), (xmin, ymax, z)]
      right  = [(xmax, ymin, z), (xmax, ymax, z)]
    并打印调试信息。
    """
    # ----- 1. 调用 GetBoundingBox -----
    try:
        min_pt, max_pt = pl.GetBoundingBox()
    except Exception:
        mins = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))
        maxs = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))
        pl.GetBoundingBox(mins, maxs)
        min_pt = tuple(mins.value)
        max_pt = tuple(maxs.value)

    xmin, ymin, zmin = min_pt
    xmax, ymax, _    = max_pt

    print(f"▶ BoundingBox Min 点: {min_pt}")
    print(f"▶ BoundingBox Max 点: {max_pt}")

    # ----- 2. 构造四个顶点（顺时针） -----
    p1 = (xmin, ymin, zmin)
    p2 = (xmax, ymin, zmin)
    p3 = (xmax, ymax, zmin)
    p4 = (xmin, ymax, zmin)

    print("▶ 矩形四个顶点 (顺时针):")
    for i, pt in enumerate((p1, p2, p3, p4), 1):
        print(f"   {i}: {pt}")

    # ----- 3. 四条边，各自用列表表达 -----
    top    = [p4, p3]    # y = ymax
    bottom = [p1, p2]    # y = ymin
    left   = [p1, p4]    # x = xmin
    right  = [p2, p3]    # x = xmax

    return top, bottom, left, right



#  主函数
#  (5)
#&&%  获取多段线的内部的文字

#  该函数系列包括如下一些函数



def get_texts_in_polyline(com_pl, tol=0.5):
    """
    在多段线 com_pl 内部筛选文字，并返回文字对象列表和对应的文字内容列表。

    参数:
      com_pl:  COM 多段线对象
      tol:     点-in-多边形时的容差（目前未用到，可留作将来扩展）

    返回:
      inside:   落在 com_pl 内部的文字 COM 对象列表
      contents: 对应 inside 中每个对象的文字内容列表
    """
    # 1) 多段线转顶点列表（标准化后的三维点）
    poly = get_unique_vertices_from_pl_com(com_pl)

    # 2) 收集所有文字实体（天正＋原生 CAD）
    tzh_text, tzh_mtext, cad_text, cad_mtext = collect_all_texts()
    all_texts = tzh_text + tzh_mtext + cad_text + cad_mtext

    inside = []
    contents = []

    for txt in all_texts:
        # 用左下角点判断是否在多段线内
        min_pt, _ = txt.GetBoundingBox()
        if point_in_polygon(min_pt, poly):
            inside.append(txt)

            # 根据对象类型取出内容
            name = getattr(txt, "ObjectName", "") or getattr(txt, "EntityName", "")
            if name in ("AcDbText", "AcDbMText"):
                # AutoCAD 原生单/多行
                contents.append(txt.TextString)
            elif name == "TDbText":
                # 天正单行
                contents.append(txt.Text)
            elif name == "TDbMText":
                # 天正多行，需要炸开取内容
                contents.append( TDbMText_content(txt))
            else:
                # 其它（万一有），留空或用 repr
                contents.append("")

    print(f"总共找到 {len(inside)} 条落在多段线内部的文字。")
    return inside, contents




# 获取单独一行的天正多行文字内容

def TDbMText_content(comobj):

    """
    对复杂多行文本内容，需要专门分析

    本函数仅针对单个文字，获取其文字内容

    不改变其本身属性和其它对象的关系

    """

    highlight_entity_by_bbox( comobj)
    cmd = "x\n"
    doc.SendCommand(cmd)
    try:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        wait_cad_idle(acad)
    except Exception:
        pass

    ob=doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
    neirong=ob.TextString

    cmd = "u\n"

    doc.SendCommand(cmd)

    return neirong







#  主函数
#  (6)
#  多段线上的均分插入


"""
该函数用于在dwg文件沿着PL线快速均衡放置树木等图块

            
"""
def distribute_points_on_entity(entity, n, block, scale_factor, ys):

    def distance(p1, p2):
        return math.sqrt((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2)

    model_space = doc.ModelSpace
    block_name = block.Name  # 获取块的名称
    
    # 如果实体是直线
    if entity.ObjectName == "AcDbLine":
        start_point = entity.StartPoint
        end_point = entity.EndPoint
        for i in range(n):
            x = start_point[0] + i * (end_point[0] - start_point[0]) / (n - 1)
            y = start_point[1] + i * (end_point[1] - start_point[1]) / (n - 1)
            inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
            inserted_block.Color = ys  # 设置颜色为红色

    # 如果实体是圆弧
    elif entity.ObjectName == "AcDbArc":
        start_angle = entity.StartAngle
        end_angle = entity.EndAngle
        center = entity.Center
        radius = entity.Radius
        for i in range(n):
            angle = start_angle + i * (end_angle - start_angle) / (n - 1)
            x = center[0] + radius * math.cos(angle)
            y = center[1] + radius * math.sin(angle)
            inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
            inserted_block.Color = ys  # 设置颜色为红色

    # 如果实体是多段线
    elif entity.ObjectName == "AcDbPolyline":
        coords = entity.Coordinates
        points = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        total_length = sum(distance(points[i], points[i+1]) for i in range(len(points)-1))
        segment_length = total_length / n
        current_length = 0

        for i in range(n):
            accumulated_length = 0
            for j in range(len(points) - 1):
                segment = distance(points[j], points[j+1])
                if accumulated_length + segment > current_length:
                    ratio = (current_length - accumulated_length) / segment
                    x = points[j][0] + ratio * (points[j+1][0] - points[j][0])
                    y = points[j][1] + ratio * (points[j+1][1] - points[j][1])
                    inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
                    inserted_block.Color = ys  # 设置颜色为红色
                    break
                accumulated_length += segment
            current_length += segment_length
            




#  主函数
#  (7)
# 返回 pl1 中与 pl2 “共线且有重叠”的区段列表

#  该函数系列包括如下一些函数

# 1 判断一条直线是否完全在另一条直线上

def is_segment_contained(seg_a, seg_b, tol=1e-4):
    """
    判断 seg_a 是否完全位于 seg_b 上（包含端点）。
    
    参数
    ----
    seg_a, seg_b :  ( (x1,y1,z1), (x2,y2,z2) )   或   AcDbLine
        两条待比较的线段。先判断 seg_a 是否被 seg_b 覆盖；
        若想双向判断，调用两次并对调顺序即可。
    tol : float
        距离与投影误差容差，默认 1e‑4 (CAD 单位)。

    返回
    ----
    bool
        True  —— seg_a 整段落在 seg_b 上  
        False —— 否则
    """
    # -------- 把输入统一转成端点元组 ----------
    def get_endpoints(entity):
        if hasattr(entity, "StartPoint"):            # COM 线段
            return (tuple(entity.StartPoint), tuple(entity.EndPoint))
        else:                                        # 纯坐标二元组
            return (tuple(entity[0]), tuple(entity[1]))

    a1, a2 = get_endpoints(seg_a)
    b1, b2 = get_endpoints(seg_b)

    # -------- 基本几何工具 ----------
    def dist(p, q):
        return math.hypot(p[0]-q[0], p[1]-q[1])

    def dot(u, v):
        return u[0]*v[0] + u[1]*v[1]

    def colinear(p, q, r, tol):
        """三点是否近似共线（面积≈0）"""
        return abs( (q[0]-p[0])*(r[1]-p[1]) - (q[1]-p[1])*(r[0]-p[0]) ) <= tol

    # -------- 先检测共线性 ----------
    if not (colinear(b1, b2, a1, tol) and colinear(b1, b2, a2, tol)):
        return False          # seg_a 端点不与 seg_b 共线 → 不可能包含

    # -------- 再检测投影是否在 b1‑b2 区间内 ----------
    # 令 b1 为原点，b_dir 为方向向量
    b_dir = (b2[0]-b1[0], b2[1]-b1[1])
    b_len2 = dot(b_dir, b_dir)
    if b_len2 == 0:           # seg_b 长度为 0 → 无法包含他段
        return False

    def proj_param(p):
        # 标准化投影参数 t，若 0<=t<=1 则投影落在 seg_b 上
        return dot( (p[0]-b1[0], p[1]-b1[1]), b_dir ) / b_len2

    t_a1 = proj_param(a1)
    t_a2 = proj_param(a2)

    # 允许介于 0±tol 到 1±tol 之间
    inside_1 = -tol <= t_a1 <= 1+tol
    inside_2 = -tol <= t_a2 <= 1+tol

    return inside_1 and inside_2

#&&% 2 返回 PL线pl1 中与 pl2 “共线且有重叠”的区段列表

def common_segments_between_polylines(pl1, pl2, tol=0.5):
    """
    返回 pl1 中与 pl2 “共线且有重叠”的区段列表，每个区段用
      [(x1,y1,0.0),(x2,y2,0.0)] 表示。

    参数
    ----
    pl1, pl2 : AutoCAD AcDbPolyline COM 对象 (或伪造对象，只需有 .Coordinates / .Closed)
    tol      : 共线 & 距离容差 (同 CAD 单位)

    返回
    ----
    overlaps : list[ list[(x,y,z),(x,y,z)] ]
    """

    # ────────── 内部小工具 ──────────
    def coords_to_xy_pairs(coords):
        """把 (x1,y1,x2,y2,…) 转成 [(x,y),…]；自动丢弃尾部残值"""
        pairs = []
        for i in range(0, len(coords) - 1, 2):
            pairs.append((coords[i], coords[i + 1]))
        return pairs

    def build_segments(verts, closed=False):
        """由顶点顺序生成线段 [P,Q] 列表"""
        segs = []
        for i in range(len(verts) - 1):
            segs.append([verts[i], verts[i + 1]])
        if closed and len(verts) > 2:
            segs.append([verts[-1], verts[0]])
        return segs

    def dist(p, q):
        return math.hypot(p[0] - q[0], p[1] - q[1])

    def colinear(p, q, r):
        """三点共线判定 |叉积| < tol · max(边长)"""
        return abs((q[0] - p[0]) * (r[1] - p[1]) -
                   (q[1] - p[1]) * (r[0] - p[0])) <= tol * max(dist(p, q), dist(p, r), dist(q, r), 1)

    def project(p, axis):
        """根据 axis==0 用 x，否则 y"""
        return p[axis]

    def segment_overlap(seg_a, seg_b):
        """
        若共线且区间有重叠，返回实际重叠区间端点 (pa, pb)（二维点）。
        否则返回 None
        """
        (p1, p2), (q1, q2) = seg_a, seg_b
        # 共线检测
        if not (colinear(p1, p2, q1) and colinear(p1, p2, q2)):
            return None

        # 选投影轴
        axis = 0 if abs(p2[0] - p1[0]) >= abs(p2[1] - p1[1]) else 1
        a1, a2 = project(p1, axis), project(p2, axis)
        b1, b2 = project(q1, axis), project(q2, axis)
        # 使 a1 <= a2, b1 <= b2
        if a1 > a2:
            p1, p2 = p2, p1
            a1, a2 = a2, a1
        if b1 > b2:
            q1, q2 = q2, q1
            b1, b2 = b2, b1

        # 计算 1‑D 重叠区间
        left, right = max(a1, b1), min(a2, b2)
        if right - left <= tol:          # “长度”≈0 视为无重叠
            return None

        # 把投影点还原到 2D 端点 —— 在线段 p1‑p2 上按比例取值
        def interp(t):
            """t 为 0~1"""
            return (p1[0] + t * (p2[0] - p1[0]),
                    p1[1] + t * (p2[1] - p1[1]))

        len_p = a2 - a1 if a2 != a1 else 1e-9
        pa = interp((left - a1) / len_p)
        pb = interp((right - a1) / len_p)
        return pa, pb

    # ────────── 主流程 ──────────
    v1 = coords_to_xy_pairs(pl1.Coordinates)
    v2 = coords_to_xy_pairs(pl2.Coordinates)

    # 若为闭合 polyline，补一个首尾顶点
    if getattr(pl1, "Closed", False) and v1 and v1[0] != v1[-1]:
        v1.append(v1[0])
    if getattr(pl2, "Closed", False) and v2 and v2[0] != v2[-1]:
        v2.append(v2[0])

    segs1 = build_segments(v1, closed=False)
    segs2 = build_segments(v2, closed=False)

    overlaps = []

    for s1 in segs1:
        for s2 in segs2:
            ov = segment_overlap(s1, s2)
            if ov:
                pa, pb = ov
                overlaps.append([(pa[0], pa[1], 0.0), (pb[0], pb[1], 0.0)])
                break        # 一条 s1 找到重叠就够了，可跳出

    # --- 打印摘要 ---
    print("★ 与 pl2 重叠 (或被包含) 的 pl1 线段数：", len(overlaps))
    for idx, seg in enumerate(overlaps, 1):
        print(f"  {idx}. {seg[0]}  →  {seg[1]}")

    return overlaps




#  主函数
#  (8)
# 找到全部“两根多段线耦合成一个矩形”的多段线

#  该函数系列包括如下一些函数

"""
函数是用来分析主房间带卫生间这种情况的，因此对输入变量是有较严格假定的，并非针对任意情况

"""
# 1 判断矩形是否包含另一个矩形

def is_rect_inside_rect(rect_outer, rect_inner, tol=1e-6):
    """
    判定 axis‑aligned 的矩形 rect_inner 是否被（含边界）完全包在 rect_outer 内。

    参数
    ----
    rect_outer : ((xmin, ymin), (xmax, ymax))
    rect_inner : ((xmin, ymin), (xmax, ymax))
        两个元组分别给出矩形左下、右上坐标（假定 Z 全 0）。
    tol        : float
        容差（允许轻微数值误差；AutoCAD double 转 python float 时推荐 1e‑6～1e‑4）。

    返回
    ----
    bool   ——  rect_inner ⊆ rect_outer ？
    """
    (ox0, oy0), (ox1, oy1) = rect_outer
    (ix0, iy0), (ix1, iy1) = rect_inner

    return (
        ix0 >= ox0 - tol and
        iy0 >= oy0 - tol and
        ix1 <= ox1 + tol and
        iy1 <= oy1 + tol
    )




# 2 判断两条正交多段线拼在一起后是否正好是一个矩形


def two_plines_making_rectangle(pl1, pl2, tol=0.5):#
    """
    判断两条正交多段线拼在一起后是否正好是一个矩形。
    假设：两 PLine 没有面积重叠，只可能共用完整边或边的一部分。
    """

    import math

    def same_point(a, b, tol):
        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol

    def pline_vertices(pl):
        # 从 pl.Coordinates 提取 (x,y) 顶点列表，自动闭合
        c = pl.Coordinates
        verts = []
        for i in range(0, len(c), 2):
            verts.append((c[i], c[i+1]))
        if not same_point(verts[0], verts[-1], tol):
            verts.append(verts[0])
        return verts

    def poly_area(verts):
        # 计算首尾闭合的多边形面积
        s = 0
        for i in range(len(verts)-1):
            x1,y1 = verts[i]
            x2,y2 = verts[i+1]
            s += x1*y2 - x2*y1
        return abs(s)*0.5

    def collect_segments(verts):
        # 从顶点列表生成线段列表 [((x1,y1),(x2,y2)), ...]
        segs = []
        for i in range(len(verts)-1):
            segs.append((verts[i], verts[i+1]))
        return segs

    def covers_edge(edge, segs):
        """
        判断给定的 bbox 边 edge=((x0,y0),(x1,y1)) 能否被 segs 中若干共线线段
        连续覆盖（无缝隙）。
        """
        (x0,y0),(x1,y1) = edge
        # 算出方向向量和长度
        dx, dy = x1-x0, y1-y0
        L = math.hypot(dx, dy)
        if L < tol:
            return False
        ux, uy = dx/L, dy/L  # 单位方向向量

        # 投影每条 seg 到 [0,L] 参数区间
        intervals = []
        for (ax,ay),(bx,by) in segs:
            # 判断端点是否在同一直线上
            # cross == 0
            cross = (ax-x0)*dy - (ay-y0)*dx
            if abs(cross) > tol*L:  
                continue
            # 计算两端投影参数
            t1 = (  (ax-x0)*ux + (ay-y0)*uy )
            t2 = (  (bx-x0)*ux + (by-y0)*uy )
            a, b = min(t1,t2), max(t1,t2)
            # 只保留与 [0,L] 有交集的部分
            if b < -tol or a > L+tol:
                continue
            intervals.append((max(0.0, a), min(L, b)))

        if not intervals:
            return False

        # 合并所有区间
        intervals.sort(key=lambda iv: iv[0])
        cur_start, cur_end = intervals[0]
        for a,b in intervals[1:]:
            if a > cur_end + tol:
                # 出现间隙
                return False
            cur_end = max(cur_end, b)
        # 最后检测是否覆盖了整个 [0, L]
        return (cur_start <= tol) and (cur_end >= L - tol)

    # 1) 提取顶点及面积
    v1 = pline_vertices(pl1)
    v2 = pline_vertices(pl2)
    A1 = poly_area(v1)
    A2 = poly_area(v2)

    # 2) 计算公共外包矩形
    xs = [p[0] for p in v1[:-1]] + [p[0] for p in v2[:-1]]
    ys = [p[1] for p in v1[:-1]] + [p[1] for p in v2[:-1]]
    xmin, xmax = min(xs), max(xs)
    ymin, ymax = min(ys), max(ys)
    A_bbox = (xmax-xmin)*(ymax-ymin)

    # 3) 面积校验
    if abs((A1 + A2) - A_bbox) > tol:
        return False

    # 4) 检查四条边被覆盖
    bbox_edges = [
        ((xmin,ymin),(xmax,ymin)),
        ((xmax,ymin),(xmax,ymax)),
        ((xmax,ymax),(xmin,ymax)),
        ((xmin,ymax),(xmin,ymin))
    ]
    segs = collect_segments(v1) + collect_segments(v2)

    hits = 0
    for edge in bbox_edges:
        if covers_edge(edge, segs):
            hits += 1
    return hits == 4



#  主函数
#  (9)
#  判断多段线PL2是否在多段线PL1多边形中

#  该函数系列包括如下一些函数

def are_all_vertices_inside(pl1, pl2):
    """
    判断多段线 pl2 的所有顶点是否都在多段线 pl1 构成的多边形内部。

    参数：
      pl1, pl2: COM 多段线对象（LWPOLYLINE 或 POLYLINE），视为封闭正多边形。
    返回：
      (all_inside, outside_pts)
      - all_inside: 如果 pl2 的每个顶点都在 pl1 内部，返回 True，否则 False。
      - outside_pts: 列表，包含所有落在 pl1 外部的 pl2 顶点 (x, y, z)。
    """
    # 先把 COM 多段线转成顶点列表 [(x,y,z), ...]
    verts1 = get_unique_vertices_from_pl_com(pl1)
    verts2 = get_unique_vertices_from_pl_com(pl2)

    outside_pts = []
    for pt in verts2:
        if not point_in_polygon(pt, verts1):
            outside_pts.append(pt)

    all_inside = len(outside_pts) == 0
    if all_inside:
        print("✅ pl2 的所有顶点都在 pl1 的内部。")
    else:
        print(f"❌ pl2 有 {len(outside_pts)} 个顶点不在 pl1 内部：")
        for p in outside_pts:
            print("   ", p)

    return all_inside, outside_pts




#&&&&%% 第六部分 综合控制管理

#  模块使用说明

"""



估计马工那样的文件打开时会遇到大量的窗口，需要按照插入图签函数那样采用双线程进行7次取消处理 20250720





没有那么复杂，根本上讲，使用代码将一个作业过程完成，再将其流程标准化梳理，这个过程的自动化就完成了。不需要太多耗费心神研究这个过程本身。





通过输入输出字典，不仅收集了信息，统一了信息处理平台，将来在构建大函数节点时，也可以通过查看字典的信息输出，判断每一个环节的重要参数

控制系统运转

抓住主要核心部分，忽略次要的要素，免得陷入无意义的劳作

先把整体大框架骨架建立好

关键概念，思想，节点，消息要打印出来分析

加强数据结构和算法的训练，这是要点

我们不仅仅是解决了某个问题，更是研究清楚了其包含的内在规律。这些骨架，结构，相关的思想方法概念就是内在规律的表达。

要深刻确定每个功能函数的设计思想和骨架结构，不能含糊依赖人工智能，否则就会飘忽不确定，没力度。有了本质的明确的认知，需要的时候，无惧重来


系统尽可能完全自动化   尽可能少干预系统自动化 

关于整个系统处理的优化分析

文件处理之前，应该对所有对象选择后分类放入各自的预设图层，这样将来可以快速找到它们
同样文件操作中，将相应对象放入它们的图层，也便于找到它们
图层的选择很快，也可以使用Handle标签等
如果对象有很强的逻辑关联，可以建立组用来关联它们

20250423

对自动化操作有了新的提升。我们有pywin32的API控制，还有SendComman发送窗口命令控制，还有pyautogui等拟人操作控制，
还有autolisp控制。同时还要注意整个电脑系统 的控制设置。它们是综合解决问题的。

把CAD窗口设置为1/4，把IDLE设置到右侧，这是使用拟人操作的基础。

select_objects_in_window_area 实际选择不可靠且慢，可行方法是使用高亮显示，再通过pick方法转隐性选择 pick方法转隐性选择
也会改变文件之间的复制粘贴

基于在一次操作运算中获得的信息应该进行充分的整合分析，减少重复选择计算和在CAD复杂环境操作的风险


应该总结经验教训，针对一个任务的深入分析，分解成函数，数据结构，各个结构的咬合，精准到位。比如编目录，它的要点，每个任务的实现，组合实现。转场，参照起点，雷边，删除，选择

插图纸目录模板时跳出shx字体对话框 

activate_window_by_title("缺少 SHX 文件")
✅ 已激活窗口：“缺少 SHX 文件” 位置(1060,510) 大小424×318
(1060, 510, 424, 318)

已获取坐标：(1232, 733)  点击忽略
(1232, 733)




activate_window_by_title("AutoCAD 错误中断", click_titlebar= True)
✅ 已激活窗口：“AutoCAD 错误中断” 位置(485,227) 大小364×211
(485, 227, 364, 211)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置… 确定
已获取坐标：(775, 404)
(775, 404)


activate_window_by_title("AutoCAD 警告", click_titlebar= True)
✅ 已激活窗口：“AutoCAD 警告” 位置(392,247) 大小497×227
(392, 247, 497, 227)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…  否
已获取坐标：(811, 434)
(811, 434)


✅ 已激活窗口：“ AutoCAD 错误报告” 位置(1055,425) 大小450×549
(1055, 425, 450, 549)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…  x（关闭）
已获取坐标：(1483, 440)
(1483, 440)

activate_window_by_title("错误报告 - 已取消", click_titlebar= True)
✅ 已激活窗口：“错误报告 - 已取消” 位置(1127,610) 大小307×179
(1127, 610, 307, 179)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置… 确定
已获取坐标：(1364, 753)
(1364, 753)


出问题以后的CAD重新打开

activate_window_by_title("图形修复", click_titlebar = True)
✅ 已激活窗口：“图形修复” 位置(1089,562) 大小366×213
(1089, 562, 366, 213)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…关闭(汉字按钮)
已获取坐标：(1397, 724)
(1397, 724)



"""
#&&&% 如果获取不到cad文字的处理办法


from win32com.client import CastTo

# —— 可选：简单清洗 MText 控制码（保留数字用）
#_MTEXT_CTRL_RE = re.compile(r"(\\[A-Za-z])|({|})|(;)|(\\\S[^;]*;)")
def _clean_mtext(s: str) -> str:
    # 把 \P 换成空格，去掉常见控制码；如需保留换行，可改成 '\n'
    s = str(s).replace("\\P", " ")
    s = _MTEXT_CTRL_RE.sub("", s)
    return s.strip()

def get_text_and_ip(ent):
    """
    通吃 AcDbText / AcDbMText / Attribute(Reference)：
    返回 (text, (x,y,z)) 或 (None, None)
    """
    try:
        name = ent.ObjectName
    except Exception:
        name = ""

    # 1) 先尝试“就地访问”，兼容以前版本（有些环境直接能用）
    for attr in ("TextString", "Contents"):  # TEXT / MTEXT
        try:
            txt = getattr(ent, attr)
            ip  = ent.InsertionPoint  # 这一步有时也需要 CastTo，下面会兜底
            if txt:
                return str(txt), (float(ip[0]), float(ip[1]), float(ip[2]) if len(ip) > 2 else 0.0)
        except Exception:
            pass

    # 2) 按类型 CastTo 到具体接口再取
    try:
        if name == "AcDbText":
            obj = CastTo(ent, "IAcadText")
            return str(obj.TextString), tuple(map(float, obj.InsertionPoint))
        elif name == "AcDbMText":
            obj = CastTo(ent, "IAcadMText")
            return _clean_mtext(obj.Contents), tuple(map(float, obj.InsertionPoint))
        elif name in ("AcDbAttribute", "AcDbAttributeReference"):
            # 属性文字（块里的标签）
            # 两种接口都试一下
            try:
                obj = CastTo(ent, "IAcadAttributeReference")
            except Exception:
                obj = CastTo(ent, "IAcadAttribute")
            return str(obj.TextString), tuple(map(float, obj.InsertionPoint))
    except Exception:
        pass

    return None, None


# —— 你原来“按插入点 y 从上到下配对 J1..Jn”的封装（适配 get_text_and_ip）
#_num_re = re.compile(r"-?\d+(?:\.\d+)?")

def _first_number(s: str) -> float:
    m = _num_re.search(str(s))
    if not m:
        raise ValueError(f"无法解析数字: {s!r}")
    return float(m.group(0))

def build_J_points_from_selected_texts(LB, n_points=61, prefix_x="30", prefix_y="37"):
    """
    输入：LB=select_text() 返回的选择集（含 TEXT/MTEXT/属性文字等）
    规则：
      - 文本以 '30' 开头的是 X；以 '37' 开头的是 Y
      - 按“插入点 y”从大到小（即从上到下）排序；同一行按 x 升序稳定
      - 依序配给 J1..Jn
    返回：
      pts_dict: {'J1': (X, Y), ...}
      pts_list: [('J1', X, Y), ...]  # 已按编号排序
    """
    items = []
    for ent in LB:
        txt, ip = get_text_and_ip(ent)
        if not txt or not ip:
            continue
        x0, y0 = float(ip[0]), float(ip[1])
        t = str(txt).strip().replace(" ", "")

        if t.startswith(prefix_x):
            items.append(("x", _first_number(t), y0, x0))
        elif t.startswith(prefix_y):
            items.append(("y", _first_number(t), y0, x0))

    # 上->下：y 降序；同行 x 升序
    items.sort(key=lambda it: (-it[2], it[3]))

    xs = [v for typ, v, _, _ in items if typ == "x"]
    ys = [v for typ, v, _, _ in items if typ == "y"]

    if len(xs) != n_points or len(ys) != n_points:
        raise RuntimeError(f"X={len(xs)} / Y={len(ys)} 与期望 {n_points} 不符，请检查选择集或前缀。")

    pts_dict = {f"J{i+1}": (xs[i], ys[i]) for i in range(n_points)}
    pts_list = [(f"J{i+1}", xs[i], ys[i]) for i in range(n_points)]
    return pts_dict, pts_list

"""

把你项目里所有直接用：

obj.TextString

obj.Contents

obj.InsertionPoint

的地方，统一替换为：

txt, ip = get_text_and_ip(obj)
# txt 为字符串（已做了基本清洗），ip 为 (x, y, z)


"""
import math
import traceback
from win32com.client import CastTo, gencache, Dispatch

def _flatten_coords(obj):
    """将各种返回格式（tuple/list of tuples OR flat list）统一为平面坐标 [x,y,x,y,...]"""
    if obj is None:
        return None
    # 如果是扁平数值列表
    try:
        if isinstance(obj, (list, tuple)) and obj and all(isinstance(x, (int,float)) for x in obj):
            return list(obj)
    except Exception:
        pass
    # 如果是 list/tuple of points [(x,y,z),...]
    try:
        coords = []
        for p in obj:
            if p is None:
                continue
            # p 可以是 (x,y,z)
            coords.extend([float(p[0]), float(p[1])])
        if coords:
            return coords
    except Exception:
        pass
    return None

def poly_length_from_coords_flat(coords):
    if not coords or len(coords) < 4:
        return 0.0
    total = 0.0
    for i in range(0, len(coords)-2, 2):
        dx = coords[i+2] - coords[i]
        dy = coords[i+3] - coords[i+1]
        total += math.hypot(dx, dy)
    return total

def try_read_coords_by_props(ent):
    """尝试各种常见属性/方法读取多段线顶点或坐标列表"""
    cand_props = ("Coordinates","coordinates","GetCoordinates","Points","PointsXY","Vertexes","Vertices")
    for p in cand_props:
        try:
            val = getattr(ent, p)
            coords = val() if callable(val) else val
            flat = _flatten_coords(coords)
            if flat:
                return flat
        except Exception:
            continue
    return None

def try_cast_and_read(ent):
    """尝试 CastTo 到常见 Polyline 接口并读取 .Coordinates"""
    try:
        # IAca dLWPolyline 或 IAcadPolyline（在 makepy 中常见）
        for iface in ("IAcadLWPolyline","IAcadPolyline"):
            try:
                pev = CastTo(ent, iface)
                # 常见属性
                for p in ("Coordinates","coordinates","GetCoordinates"):
                    try:
                        val = getattr(pev, p)
                        coords = val() if callable(val) else val
                        flat = _flatten_coords(coords)
                        if flat:
                            return flat
                    except Exception:
                        pass
            except Exception:
                continue
    except Exception:
        pass
    return None

def get_length_via_explode(ent, doc):
    """通过复制+Explode的方式收集爆炸后分段的长度（对线段/弧段处理），并清理临时对象
       返回 (length, had_temp_entities_flag)
    """
    created = []
    total = 0.0
    try:
        # 1) 复制实体（Copy 会返回新的 COM 对象）
        try:
            ent_copy = ent.Copy()
            created.append(ent_copy)
        except Exception:
            # 有些 wrapper Copy 方法需要文档操作，尝试 alternative：Add copy via clone? 若失败直接 explode 原对象（危险）
            ent_copy = ent  # 退而求其次（但我们不想直接 explode 原件）
        # 2) Explode（可能返回一个集合或产生新的实体在 ModelSpace）
        try:
            res = ent_copy.Explode()
        except Exception as e:
            # 如果 Explode 失败，尝试直接读取 Length 属性（最后的兜底）
            try:
                return float(ent.Length), False
            except Exception:
                raise

        # res 可能是一个 sequence（生成的实体集合）或 None（有些接口直接在 ModelSpace 插入）
        segs = []
        try:
            # 如果 res 本身是一个 Python iterable
            for item in (res or []):
                segs.append(item)
        except Exception:
            # 当 res 为 None 时，尝试扫描最近创建的实体（不能可靠）
            segs = []

        # 如果没有通过返回的 res 获取到分段，尝试在 ModelSpace 中查找由 Copy 产生的后续对象
        # 为简单稳妥，我们也把 ent_copy 加入 created（若 ent_copy != ent）
        for s in segs:
            # 对每个分段，优先读取 Length 属性（圆弧也可能有 Length）
            try:
                L = getattr(s, "Length")
                if L is not None:
                    total += float(L)
                    created.append(s)
                    continue
            except Exception:
                pass
            # 对直线段，使用 StartPoint/EndPoint
            try:
                sp = getattr(s, "StartPoint", None)
                ep = getattr(s, "EndPoint", None)
                if sp and ep:
                    dx = float(ep[0]) - float(sp[0])
                    dy = float(ep[1]) - float(sp[1])
                    total += math.hypot(dx, dy)
                    created.append(s)
                    continue
            except Exception:
                pass
            # 对弧段，尝试用 Radius 和 StartAngle/EndAngle 计算
            try:
                r = getattr(s, "Radius", None)
                sa = getattr(s, "StartAngle", None)
                ea = getattr(s, "EndAngle", None)
                if r is not None and sa is not None and ea is not None:
                    # 角度可能以弧度给出
                    ang = abs(float(ea) - float(sa))
                    total += abs(float(r)) * ang
                    created.append(s)
                    continue
            except Exception:
                pass
            # 若无法识别，忽略该段
        # 3) 清理：擦除临时创建的实体（注意：如果 ent_copy is original ent, 不要擦除原实体）
        for obj in created:
            try:
                # 不擦除原对象
                if obj is ent:
                    continue
                obj.Erase()
            except Exception:
                pass
        # 如果 ent_copy was the original ent (we didn't want that), we didn't erase it above
        return total, (len(created) > 0)
    except Exception as e:
        # 尝试兜底读 Length
        try:
            return float(ent.Length), False
        except Exception:
            # 最终失败
            # 打印 traceback 以便调试
            traceback.print_exc()
            return None, False

def get_polyline_coords_and_length(ent, doc=None):
    """
    主要对外函数：先尝试直接取顶点坐标（返回 flat coords list），并计算长度；
    如果无法得到顶点，则用 explode 回退计算长度。
    返回 (coords_flat_or_None, length_or_None, used_explode_flag)
    """
    # 1) 直接按常见属性名尝试
    coords = try_read_coords_by_props(ent)
    if coords:
        length = poly_length_from_coords_flat(coords)
        return coords, length, False

    # 2) 尝试 CastTo 常见接口
    coords = try_cast_and_read(ent)
    if coords:
        length = poly_length_from_coords_flat(coords)
        return coords, length, False

    # 3) Explode 回退（需要提供 doc 用于清理）
    if doc is None:
        # 尝试获取 doc（实体有 Document 属性）
        try:
            doc = ent.Document
        except Exception:
            doc = None
    length, used_temp = get_length_via_explode(ent, doc)
    if length is not None:
        return None, length, used_temp

    # 4) 最后兜底：尝试读取 Length 属性直接返回
    try:
        L = getattr(ent, "Length")
        return None, float(L), False
    except Exception:
        return None, None, False



import win32com.client as win32

def ensure_acad():
    # 早绑定：生成并加载类型库（比 Dispatch 稳定，属性/方法更齐全）
    acad = win32.gencache.EnsureDispatch("AutoCAD.Application")
    return acad

def cast_polyline(ent):
    """把选到的多段线实体转成更具体的接口，方便取专有属性。"""
    name = ent.ObjectName  # 如 'AcDbPolyline' / 'AcDb2dPolyline' / 'AcDb3dPolyline'
    if name == "AcDbPolyline":     # 轻量 2D 多段线
        try:
            return win32.CastTo(ent, "IAcadLWPolyline")
        except Exception:
            return ent  # 退化为动态调度
    elif name == "AcDb2dPolyline":
        try:
            return win32.CastTo(ent, "IAcadPolyline")  # 旧 2D 多段线
        except Exception:
            return ent
    elif name == "AcDb3dPolyline":
        try:
            return win32.CastTo(ent, "IAcad3DPolyline")
        except Exception:
            return ent
    else:
        return ent

def get_ent_truecolor_rgb(ent):
    """
    返回实体实际显示颜色(尽力)：(r,g,b)
    如果是 ByLayer / ByBlock，会回退去查图层或给出 None。
    """
    try:
        tc = ent.TrueColor  # AcadAcCmColor
        # ColorMethod：0=ByBlock,1=ByLayer,2=ByColor,3=ByACI,4=ByRGB,5=ByPen
        cm = getattr(tc, "ColorMethod", None)
        if cm in (4,):  # ByRGB
            return (tc.Red, tc.Green, tc.Blue)
        if cm in (2, 3):  # 直接颜色/ACI
            # 直接给出 ACI 对应的 RGB 可能需要映射；简单返回 None
            return None
        if cm == 1:  # ByLayer
            try:
                layer = ent.Document.Layers.Item(ent.Layer)
                ltc = layer.TrueColor
                if getattr(ltc, "ColorMethod", None) == 4:
                    return (ltc.Red, ltc.Green, ltc.Blue)
            except Exception:
                pass
        # 其它情况回退
        return None
    except Exception:
        return None

def polyline_vertices(pl):
    """
    返回多段线顶点列表：
    - LWPolyline: 用 Coordinates（扁平数组 x,y[,x,y...]）
    - Polyline/3DPolyline: 用 Coordinate(i)
    """
    name = pl.ObjectName
    pts = []
    try:
        if name == "AcDbPolyline":  # LW
            coords = list(pl.Coordinates)  # [x1,y1, x2,y2, ...]
            # LWPolyline 固定在平面，Z 一般等于 Elevation
            z = getattr(pl, "Elevation", 0.0)
            for i in range(0, len(coords), 2):
                pts.append((coords[i], coords[i+1], z))
        elif name in ("AcDb2dPolyline", "AcDb3dPolyline"):
            n = pl.NumberOfVertices
            for i in range(n):
                p = pl.Coordinate(i)  # (x, y, z)
                pts.append((p[0], p[1], p[2]))
        else:
            # 兜底尝试 Coordinates
            coords = list(pl.Coordinates)
            z = getattr(pl, "Elevation", 0.0)
            for i in range(0, len(coords), 2):
                pts.append((coords[i], coords[i+1], z))
    except Exception:
        pass
    return pts

def inspect_polyline(ent):
    """打印多段线的关键信息：类型、图层、颜色、是否闭合、长度、起点、顶点列表等。"""
    pl = cast_polyline(ent)
    info = {}
    info["ObjectName"] = ent.ObjectName
    info["Layer"] = ent.Layer
    # Color: 256 表示 ByLayer；更可靠的是 TrueColor
    try:
        info["ColorIndex_or_Int"] = ent.Color
    except Exception:
        info["ColorIndex_or_Int"] = None
    info["TrueColorRGB"] = get_ent_truecolor_rgb(ent)

    # 是否闭合
    closed = None
    for key in ("Closed", "Closed2d", "Closed3d"):
        if hasattr(pl, key):
            try:
                closed = bool(getattr(pl, key))
                break
            except Exception:
                pass
    info["Closed"] = closed

    # 长度/面积
    for key in ("Length", "Area"):
        try:
            info[key] = float(getattr(pl, key))
        except Exception:
            info[key] = None

    # 顶点与起点
    verts = polyline_vertices(pl)
    info["VertexCount"] = len(verts)
    info["StartPoint"] = verts[0] if verts else None
    info["Vertices"] = verts

    return info
















#&&&% 重要基础知识（python、算法、数据结构）

#&&% ***🧠  抛出异常的处理逻辑

"""
1 try+except+finaly

成功后执行finaly后语句，失败后也会执行finaly后语句，必定预期整个代码段要做的就是finaly提供的功能


2 try+except+else

成功后执行else后语句，失败后不会会执行else后语句，必定预期成功情况下 要做的就是else提供的功能

3
try:
    risky_operation()
except SomeError as e:
    handle_error(e)
    # 这段只在 try 失败时运行
    try:
        do_special_on_failure()
    except Exception as e2:
        log("failure-only block failed:", e2)
else:
    # 成功时走这里
    do_if_success()
finally:
    cleanup()

do_special_on_failure() 就是专门为代码段失败时预期必定要执行的语句。它不同于直接放在except中，因为except可能的失败操作不能保证它被预期必然执行。

我们需要在某个操作失败时，必须执行某个操作，就使用这个逻辑



"""




#&&&% 脚本注释标准模块
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

#📌 节点 1：形成正式“打印框线”并重排图形
#─────────────────────────────────────

"""🔍 输入:"""

"""🔧 关键函数："""
"""   ⮞ F1: `find_standard_printframes()` – 返回标准图框列表"""

"""🧠 处理逻辑:"""

"""📤 输出:"""


 

"""
🄐  🄑  🄒  🄓  🄔  🄕  🄖  🄗  🄘  🄙  🄚  🄛  🄜  🄝  🄞  🄟  🄠  🄡  🄢  🄣  🄤  🅅  🅆  🅇   🅈   🅉







⓿ ① ② ③ ④ ⑤ ⑥ ⑦ ⑧ ⑨ ⑩ ⑪ ⑫ ️⃣  

# ❖❖❖—— 重要函数：foo() ——❖❖❖

# ✺✺✺—— 关键逻辑：process_data()——✺✺✺

# ⚡⚡⚡—— 高优先：init_config() ——⚡⚡⚡

# ★★★—— 核心入口：main() ——★★★
常用符号推荐：

❖ （四角菱形）

✺ （放射状花瓣）

⚡ ★ （闪电）

★ （实心星）

✦ （中实菱形）

❗ （粗感叹号）

 
+--------------------------+
| 方法1：从打印图签块或打印线获取DX |
+--------------------------+
      ↓ 成功          ↓ 失败
+------------------+   +--------------------------+
| 从图签块重建框线 |   | 方法2：从"公司图签"图块提取DX |
+------------------+   +--------------------------+
           ↓                     ↓ 成功    ↓ 失败


+--------------------------------+
| 初始状态：未开始打印           |
+--------------------------------+
               ↓


        +---------+
        | F3函数  | ← 识别拟合框线对应图形
        +---------+
             ↓
 
📌 节点 1：确定“准打印框线”
─────────────────────────────────────
🔁 初始化阶段：
    ⮞ 清空 "准打印框线" 图层
    ⮞ 若已存在旧对象，则全部删除

🔍 提取打印框候选集合 DX 的三种方法：

① 方法一（推荐方式）
    ⮞ 查找 “图签块” + “打印框线” 块
    ⮞ 若存在图签块：
        ⮞ 提取包围盒 → 重绘为矩形 → 存入“准打印框线”
        ⮞ 清除原“打印框线”对象
        ⮞ ✅ 完成本节点目标


📦 输出：DX 对象集合 + “准打印框线”图层绘制

------------------------------------------------------

📌 节点 2：形成正式“打印框线”并重排图形
─────────────────────────────────────
🔍 输入：DX 集合（来自上节点）

🧠 图形分析与处理：
    ⮞ 识别标准打印框（F1）
    ⮞ 识别拟合打印框（F2）

🧱 图框替换与排版：
    ⮞ 拟合框 → 替换为最近 A1 标准比例图框
    ⮞ 重新排列打印图框，避免重叠

📦 输出：
    ⮞ 正式“打印框线”写入目标图层
    ⮞ 准打印框图层清空
    ⮞ Handle → 信息存储用于后续匹配

🔧 函数：
    ⮞ F1: 标准框识别
    ⮞ F2: 拟合框识别
    ⮞ F3: 图形识别归属（局部）
    ⮞ F4: 图形重排整体逻辑

┌────────────┐⮞ F1: 标准框识别 
│ 节点 1：   │
│ 准打印框线 │
└────┬───────┘
     │ 清空图层
     ▼
┌────────────┐
│ 方法一：图签块或框线 │◄─────┐
└────┬──────────────┘      │
     │ 找不到             │
     ▼                    │

┌────────────┐
│ 排序 + 移动图形 F3/F4 │
└────┬────────┘
     ▼
┌────────────────────┐
│ 写入“打印框线”图层 │
└────────────────────┘

📌 节点 X：<简要标题>
─────────────────────────────────────
🔍 输入：
    ⮞ 输入数据或图元说明（如 DX 集合）
    ⮞ 来源说明（如来自上游节点）
    
🧠 处理逻辑：
    ⮞ 步骤说明 1（如：识别标准打印框）
    ⮞ 步骤说明 2（如：判断拟合框）

🔧 关键函数：
    ⮞ F1: 函数名 – 功能说明
    ⮞ F2: 函数名 – 功能说明

📤 输出：
    ⮞ 输出数据目标（如写入正式图层）
    ⮞ 副作用（如清除中间层、更新 handle 记录）

⚠️ 异常处理：
    ⮞ 未匹配图框 → 输出警告
    ⮞ 多图重叠 → 排序并修正重排

🗂 示例调用：
    process_printframes(DX)


📌	节点起始	📌 节点 2：形成正式“打印框线”并重排图形
────────────	分割线	──────── 节点分隔 ────────
🧭	逻辑导向	🧭 下一步识别拟合图框
🔀	路径分支	🔀 分流处理：标准框 vs 拟合框
⤵ / ⤴	子流程跳转	⤵ 进入子流程 F3
🔂	循环处理	🔂 针对所有 F2 框执行替换逻辑

🔧 功能或模块说明
符号	用途	示例
🔧	函数定义	🔧 F1: 标准框识别函数
🧠	处理逻辑	🧠 图形分析与处理
🧪	测试/调试块	🧪 临时打印 bbox 信息
📦	输出结果	📦 输出图层对象
📥 / 📤	输入/输出	📤 输出写入 LAYER_PRINT 正式图框
🧱	数据结构/实体对象	🧱 实体对象替换操作

🔍 分析与识别相关
符号	用途	示例
🔍	搜索/识别	🔍 查找符合 A1 比例的图框
🧬	特征提取	🧬 提取图形外包盒坐标与宽高比
🔎	精细识别	🔎 判断是否为完整封闭矩形

✅ 状态反馈
符号	用途	示例
✅	成功	✅ 图框替换完成
⚠️	警告	⚠️ 找不到合适图框匹配项
❌	错误	❌ 无法识别多段线边界
⏳	等待	⏳ 等待 CAD 对象响应

🔁 控制结构（可嵌入注释中）
符号	用途	示例
⬅️ ➡️ ⬆️ ⬇️	方位、移动方向	⬆️ 上移图框 1000 单位避免重叠
🔄	重复执行	🔄 对所有图框执行一次旋转检测
↪️	返回	↪️ 返回原图层处理状态

📊 数据操作与计算
符号	用途	示例
📐	几何计算	📐 计算图框宽高比
📏	测量	📏 测量距离与偏移
🧮	数值处理	🧮 总计图框数量 = len(F1) + len(F2)



"""


#&&%  *** 常用功能代码块

## 1 如果某个操作流程出错，最多间隔1秒尝试3次即无视之，进入下一流程


##try:
##    feng_lines = stc("bianmulu_lp")
##    print(f"🗑 将 {len(split_pairs)} 条分裂线标记为红色 (尝试 {attempt})")
##    for A, B in split_pairs:
##        _, yA, _ = bbox_center_3(A)
##        _, yB, _ = bbox_center_3(B)
##        y_mid = (yA + yB) / 2.0
##
##        for ent in feng_lines:
##            try:
##                obj = dyn.Dispatch(ent._oleobj_)
##            except:
##                continue
##            if obj.ObjectName not in ("AcDbLine", "AcDbPolyline", "AcDb2dPolyline"):
##                continue
##            _, cy, _ = bbox_center_3(obj)
##            if abs(cy - y_mid) <= y_tol:
##                print(f"  🔴 标记 Handle={obj.Handle} centerY={cy:.1f}")
##                try:
##                    obj.Color = 1
##                    obj.Layer = "测试辅助"
##                except:
##                    try:
##                        obj.TrueColor = 0xFF0000
##                    except:
##                        print(f"    ⚠️ 上色失败 Handle={obj.Handle}")
##    # 如果执行到这里不出异常，就退出重试循环
##    break
##except Exception as e:
##    print(f"⚠ 分裂线标记第 {attempt} 次失败: {e}")
##    if attempt < 3:
##        time.sleep(1)
##    else:
##        print("ℹ 分裂线标记跳过，进入下一流程")


def kill_wps(verbose: bool = False):
    """
    结束所有 WPS/金山办公相关进程，特别是 wpspdf.exe。
    verbose=True 时打印被终止的映像名。
    """
    # 需要关掉的映像清单（全部小写）
    targets = {
        "wps.exe",          # WPS 主进程
        "wpspdf.exe",       # ★ PDF 预览器
        "wpp.exe",          # 演示
        "et.exe",           # 表格
        "ksolaunch.exe",    # 启动器
        "wpscloudsvr.exe",  # 云同步
        "wpsupdate.exe",    # 更新器
        "wpscenter.exe",    # 消息中心
    }

    killed = set()
    for name in targets:
        # /T 连子进程一起；/F 强制；/IM 按映像名
        rc = subprocess.call(
            f'taskkill /T /F /IM {name}',
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        if rc == 0:      # 返回 0 表示确有此进程并已结束
            killed.add(name)

    if verbose:
        if killed:
            print("✓ 已结束进程:", ", ".join(sorted(killed)))
        else:
            print("ℹ️ 未检测到 WPS 相关进程")


#&&% 人机结合的函数


"""

通过input输入等待，在复杂环节引入人工操作，而其余部分仍交给计算机自动完成
因为最终的目的是节省时间
从普通图签块制作标准属性块就是一个典范
"""

#&&% 函数运行节点控制


"""
通过打印对象的Handle我们能很方便检测函数运行过程中的结果，而不需要从函数返回值或麻烦的调试中检查错误
print(f"从块{blk.Name}中对象按比例和面积特征过滤得到的com对象Handle句柄")

print_coms_handle(filtered)

"""


#&&% 函数状态控制

# 函数运行状态的标准化重构(重装软件使重启状态标准化)

"""
如果重启天正CAD系统5次都未能运行，则系统将自己重新安装天正和CAD

天正画图错误被强制中断CAD进程消息窗口
activate_window_by_title("图形修复", click_titlebar= True)
✅ 已激活窗口：“图形修复” 位置(1089,562) 大小366×213
(1089, 562, 366, 213)
同一测试的其窗口“关闭”按钮的点击位置 (1394, 724)，点击后即正常运行



"""

# 函数时间控制

"""
如果一个系统函数运转受阻，将会被时间监控中断进程，天正CAD系统会被中断重启


"""
def cad_zhongduan_ceshi():

    i=0

    while True:

        LP=[(0, i, 0.0), (90509.17232155753, 38080.72757883748, 0.0), (72659.40193916814, 48550.64173869454, 0.0), (61087.391734572855, 35617.55453017057, 0.0), (62182.914098626905, 23340.76471079199, 0.0), (43443.713317261485, 21669.276832727846, 0.0), (32018.9079127094, 29207.203355399004, 0.0), (22430.211826160492, 19081.580148860856, 0.0)]

        draw_lwpolyline(
    LP,
    layer_name= "0",
    width = 0.0,
    color= 256,
    closed = False
)

    i=i+10000




# 函数运行消息打印和结构标准控制

"""
在系统的更高级层面，函数不需要显示底层函数的运转信息，从提速的角度讲应该去掉print语句，但这个可以在稳定运转一段时间之后处理

"""


# 函数逻辑结构

"""
应该设置函数的明确逻辑构成，在各个环节打印消息用于分析运转是否错误。通过定义内部函数或小函数，建立正确的概念和模块，使得人工智能系统更容易被驾驭

"""


# 函数案例生成

"""
生成足够多的测试文件，节省程序测试时间，并通过完整案例使函数逻辑更合理

"""

# 打印案例生成器

"""
generate_test_cases(
    r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/自动化(动态)/目录图签/生成测试文件/目录图签.dwg",
    num = 3,
    jianju_x = 10000,
    juli_tukuang = 6000,
    juli_y = 400000,
)

"""

# 文字候选
TEXT_OPTIONS = [
    "一层平面图", "二层平面图", "1-18轴立面图", "2-2剖面图",
    "1-1剖面图", "楼梯大样图", "门窗详图图"
]

# 辅助：创建文字实体
def create_text(txt: str, position: Tuple[float, float, float], height: float, layer: str):
    vpt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, list(position))
    ent = mp.AddText(txt, vpt, height)
    try:
        ent.Layer = layer
    except Exception:
        pass
    return ent

# 随机绘制基本图形，置于layer_tu
def draw_random_shape(xmin: float, xmax: float, ymin: float, ymax: float, layer_tu: str):
    """
    在给定范围内随机绘制一点、线段、圆或正多边形，
    并放置到指定图层。
    """
    typ = random.choice(["point", "line", "circle", "polygon"])
    shape = None
    if typ == "point":
        pt = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        shape = draw_point(pt)
    elif typ == "line":
        p1 = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        p2 = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        shape = draw_line(p1, p2)
    elif typ == "circle":
        center = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        r = random.uniform(1000, min(xmax - xmin, ymax - ymin) / 4)
        shape = draw_circle(center, r)
    else:
        center = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        r = random.uniform(1000, min(xmax - xmin, ymax - ymin) / 4)
        sides = random.randint(3, 8)
        shape = draw_regular_polygon(center, r, sides)
    # 设置图层
    try:
        if shape is not None:
            shape.Layer = layer_tu
    except Exception:
        pass
def generate_test_cases(
    output_path: str,
    num: int = 3,
    jianju_x: float = 10000,
    juli_tukuang: float = 6000,
    juli_y: float = 400000,
):
    """
    生成测试 DWG：多行图框+图形+边缘文字，文字区域由 get_frame_edge 计算。
    """
    sizes = {"A3": (42000, 29700), "A2": (59400, 42000), "A1": (84000, 59400)}
    seq_fixed = [
        ("A3", False), ("A2", False), ("A1", False),
        ("A3", True),  ("A2", False), ("A1", True),
        ("A2", False), ("A2", False),
    ]

    layer_kuang = "测试绘制_图框"
    layer_tu    = "测试绘制_图形"
    layer_wz    = "测试绘制_文字"
    #不覆盖从当前时间给带路径文件名命名
    output_path = rename_time(output_path)
    
    create_new_dwg_file_no(output_path)
    time.sleep(1)

    base_x = random.uniform(0, 1000000)
    base_y = random.uniform(0, 1000000)

    for row in range(num):
        y0 = base_y + row * juli_y
        x = base_x
        seq = seq_fixed if row == 0 else [
            (random.choice(list(sizes.keys())), random.choice([False, True]))
            for _ in seq_fixed
        ]

        for name, vertical in seq:
            w, h = sizes[name]
            if vertical:
                pts = [(x, y0, 0), (x, y0 + w, 0), (x + h, y0 + w, 0),
                       (x + h, y0, 0), (x, y0, 0)]
            else:
                pts = [(x, y0, 0), (x + w, y0, 0), (x + w, y0 + h, 0),
                       (x, y0 + h, 0), (x, y0, 0)]
            # 绘制并获取多段线实体
            poly_ent = draw_polyline(pts, layer_name=layer_kuang, tol=0.5, width=20, color=1)
            # 计算边域四角点
            edge_pts = get_frame_edge(poly_ent, juli_tukuang)
            if not edge_pts:
                continue
            ex0, ey0, _ = edge_pts[0]
            _, ey1, _ = edge_pts[1]
            ex1, _, _ = edge_pts[2]

            # 图框内部随机图形，保留内边缘
            # 计算内部区域
            if vertical:
                inner_xmin, inner_xmax = x, x + h
                inner_ymin, inner_ymax = y0 + juli_tukuang, y0 + w
                x_shift = h
            else:
                inner_xmin, inner_xmax = x, x + w
                inner_ymin, inner_ymax = y0, y0 + h - juli_tukuang
                x_shift = w
            # 随机图形
            for _ in range(random.randint(5, 10)):
                ensure_layer_current(layer_name="测试绘制_图形")                
                draw_random_shape(
                    inner_xmin, inner_xmax,
                    inner_ymin, inner_ymax,
                    layer_tu
                )

            # 边域文字：1-6 文本，分 1-3 排，限定在 edge 区域
            n_texts = random.randint(1, 6)
            rows_t = random.randint(1, min(3, n_texts))
            counts = [n_texts // rows_t] * rows_t
            for i in range(n_texts % rows_t):
                counts[i] += 1
            text_height = 300
            line_spacing = text_height + 100

            for r, cnt in enumerate(counts):
                choices = random.sample(TEXT_OPTIONS, cnt)
                line_str = '、'.join(choices)
                if not vertical:
                    tx = random.uniform(ex0, ex1 - text_height)
                    ty = ey1 - (r + 1) * line_spacing
                    ty = min(max(ty, ey0), ey1 - text_height)
                else:
                    ty = random.uniform(ey0 + text_height, ey1)
                    tx = ex0 + r * line_spacing
                    tx = min(max(tx, ex0), ex1 - text_height)
                print(f"[DEBUG] 写入文字 '{line_str}' 在 ({tx:.2f},{ty:.2f}) 图层 {layer_wz}")
                ensure_layer_current("测试绘制_文字")                
                create_text(line_str, (tx, ty, 0), text_height, layer_wz)

            x += x_shift + jianju_x

    time.sleep(1)
    doc.SendCommand("Z\nE\n")
    savefile()
    guanbifile()

def get_frame_edge(poly_ent, juli_tukuang: float = 6000.0):
    """
    给定一个闭合多段线实体，判断其横向/竖向，并返回边域四角坐标。

    :param poly_ent:    闭合的 polyline 实体（AcDbPolyline）
    :param juli_tukuang:边域宽度
    :return: [(x1,y1,0), (x1,y2,0), (x2,y2,0), (x2,y1,0)]
    """
    try:
        (pmin, pmax) = poly_ent.GetBoundingBox()
    except Exception as e:
        print(f"[ERROR] 获取包围盒失败: {e}")
        return None
    xmin, ymin, _ = pmin
    xmax, ymax, _ = pmax
    w = xmax - xmin
    h = ymax - ymin
    if w >= h:
        orientation = 'horizontal'
        # 横向：边域在右侧
        x1 = xmax - juli_tukuang
        x2 = xmax
        y1 = ymin
        y2 = ymax
    else:
        orientation = 'vertical'
        # 竖向：边域在底部
        x1 = xmin
        x2 = xmax
        y1 = ymin
        y2 = ymin + juli_tukuang
    print(f"[DEBUG] 框为 {orientation}，边域坐标: ({x1:.2f},{y1:.2f}), ({x1:.2f},{y2:.2f}), ({x2:.2f},{y2:.2f}), ({x2:.2f},{y1:.2f})")
    return [(x1, y1, 0), (x1, y2, 0), (x2, y2, 0), (x2, y1, 0)]

#&&% 打印设置和pyautogui基本操作

"""
run_py("测鼠标位置.py")#测位置

pyautogui.moveTo(1415, 796)#定位

pyautogui.write("EXPLODE", interval=0.1)#写文字

pyautogui.press("enter")#回车

pyautogui.doubleClick()#双击

pyautogui.hotkey('ctrl', 'a')#全选

pyautogui.press('delete')#删除

# 然后从该位置向上滚动 200
pyautogui.scroll(200)

# 假设已经 moveTo(x,y)
pyautogui.mouseDown()
pyautogui.dragRel(200, 100, duration=0.3)  # 向右200，下100的框选
pyautogui.mouseUp()

如果你的“预先有的数据”不是全在一个文本框里，而是一个列表/表格中的多行，你可以
在单击一次之后，用 pyautogui.dragTo() 或 pyautogui.dragRel() 做框选
然后按 delete

# 方法一：直接右键
pyautogui.click(button='right')
# 或者使用快捷函数
# pyautogui.rightClick()

# 如果需要在弹出的上下文菜单里选择某项，比如向下移动两次再回车：
pyautogui.press('down', presses=2, interval=0.1)
pyautogui.press('enter')


# 发送 Ctrl+1
pyautogui.hotkey('ctrl', '1')
# 发送 Ctrl+9
pyautogui.hotkey('ctrl', '9')
# 发送 Shift+PrintScreen
# 注意 PyAutoGUI 里 PrintScreen 键通常叫 'printscreen' 或 'prtsc'
pyautogui.hotkey('shift', 'printscreen')

# Ctrl+1 手动版
pyautogui.keyDown('ctrl')
pyautogui.press('1')
pyautogui.keyUp('ctrl')
pyautogui.press("esc")

currentLayout = acad.ActiveDocument.ActiveLayout
currentLayout.CanonicalMediaName = "ISO_A1_(841.00_x_594.00_MM)"
currentLayout.CanonicalMediaName#查看当前图纸设置


标准A0,A1,A2,A3的可打印区域是可单独修改的，其0-0-0-0效果已经打印出来 


"""

dy_yonghu=[

                                           #A0 1189.00_x_841.00

    ("1337.63","841.00","17","17","5","5"),#A0+1/8


    ("1486.25","841.00","17","17","5","5"),#A0+1/4

                                           #A1 841.00_x_594.00

    ("1051.25","594.00","17","17","5","5"),#A1+1/4


    ("1261.50","594.00","17","17","5","5"),#A1+1/2


    ("1471.75","594.00","17","17","5","5"),#A1+3/4

                                           #A2 594.00_x_420.00


    ("742.50","420.00","17","17","5","5"), #A2+1/4

                                         

    ("891.00","420.00","17","17","5","5"), #A2+1/2


    ("1039.50","420.00","17","17","5","5") #A2+3/4


                                           #A3 420.00_x_297.00
]



def 批量设置用户打印尺寸(dy_yonghu):
    """
    先连接，再人工打开ctr+P选择DWG TO PDF打印后再运行此函数

    """

    #按用户给定数据批量生成用户打印尺寸
    for i in range(0,len(dy_yonghu)):
        dyshu = dy_yonghu[i]
 
        pyautogui.moveTo(1331, 569)#点击打印窗口的 特性 按钮
        pyautogui.click(1331, 569)        
        time.sleep(2)


        pyautogui.moveTo(1189, 627)#点击打印窗口的 自定义图纸尺寸 按钮
        pyautogui.click(1189, 627)        
        time.sleep(2)

        pyautogui.moveTo(1415, 796)#点击打印窗口的 编辑 按钮
        pyautogui.click(1415, 796)        
        time.sleep(2)

        pyautogui.moveTo(1161, 690)#点击宽度窗口位置
        pyautogui.click(1161, 690)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[0], interval=0.1)#写宽度
        time.sleep(2)

        pyautogui.moveTo(1171, 745)#点击高度窗口位置
        pyautogui.click(1171, 745)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[1], interval=0.1)#写高度
        time.sleep(2)
        
        pyautogui.moveTo(1452, 915)#点击下一页
        pyautogui.click(1452, 915)        
        time.sleep(2)

        pyautogui.moveTo(1144, 684)#点击上边距窗口位置
        pyautogui.click(1144, 684)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[2], interval=0.1)#写上边距
        time.sleep(2)

        pyautogui.moveTo(1144, 729)#点击下边距窗口位置
        pyautogui.click(1144, 729)        
        time.sleep(2)


        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[3], interval=0.1)#写下边距
        time.sleep(2)

        pyautogui.moveTo(1143, 770)#点击左边距窗口位置
        pyautogui.click(1143, 770)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[4], interval=0.1)#写左边距
        time.sleep(2)

        pyautogui.moveTo(1147, 809)#点击右边距窗口位置
        pyautogui.click(1147, 809)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[5], interval=0.1)#写右边距
        time.sleep(2)

        pyautogui.moveTo(1443, 919)#点击下一页
        pyautogui.click(1443, 919)        
        time.sleep(2)

        pyautogui.moveTo(1443, 922)#点击下一页
        pyautogui.click(1443, 922)        
        time.sleep(2)

        pyautogui.moveTo(1460, 921)#点击完成
        pyautogui.click(1460, 921)        
        time.sleep(2)


        pyautogui.moveTo(1239, 920)#点击确定
        pyautogui.click(1239, 920)        
        time.sleep(2)

##确保删除对象

def safe_delete(ob, retries: int = 5, delay: float = 1.0) -> bool:
    """
    尝试删除 CAD 对象 ob，最多重试 retries 次，每次间隔 delay 秒。
    只捕获 COM 错误，成功返回 True，否则返回 False。
    """


    for attempt in range(1, retries + 1):
        try:
            ob.Delete()
            return True
        except pywintypes.com_error:
            time.sleep(delay)
    return False





#&&&%  * 高亮选择转隐性移动区域内全部对象


"""
高亮选择窗口操作更可靠从而更快，但要考虑窗口不能挡

"""

def move_entities_in_region(coms, target=(0,0,0), ty=1, max_iter=3):
    """
    将 `coms` 对象的包围盒内所有实体，沿向量 (target - 左下角) 移动，
    每轮等待 `ty` 秒，最多循环 `max_iter` 次。

    参数：
      coms      -- 支持 GetBoundingBox() 的 COM 对象（如多段线、块参照等）
      target    -- 目标基点坐标，默认为 (0,0,0)
      ty        -- 每轮移动后等待秒数
      max_iter  -- 最多尝试的轮数
    """
    # 1. 读取包围盒，计算左下角 (x1,y1) 和右上角 (x2,y2)
    p1, p2 = coms.GetBoundingBox()
    x1, y1 = min(p1[0], p2[0]), min(p1[1], p2[1])
    x2, y2 = max(p1[0], p2[0]), max(p1[1], p2[1])

    # 2. 计算位移向量 Δ = target - (x1,y1)
    dx = target[0] - x1
    dy = target[1] - y1
    dz = target[2] - 0.0

    for i in range(1, max_iter + 1):
        # 清空显性选择集
        try:
            doc.PickfirstSelectionSet.Clear()
        except Exception:
            pass

        # 3. 在 CAD 里用窗口选择高亮区域内对象
        highlight_entities_in_window(x1, y1, x2, y2)

        # 4. 拿到显性选择集里的实体
        pick = doc.PickfirstSelectionSet
        count = pick.Count if hasattr(pick, 'Count') else len(list(pick))
        if count == 0:
            print(f"✅ 第 {i} 轮：区域已清空，停止。")
            break

        print(f"♻️ 第 {i} 轮：检测到 {count} 个对象，正在移动…")
        # 5. 对每个实体按计算好的向量执行 Move
        for ent in pick:
            try:
                # 从 (x1,y1,0) 移动到 (x1+dx, y1+dy, dz)
                ent.Move(vtpnt(x1, y1, 0.0),
                         vtpnt(x1 + dx, y1 + dy, dz))
            except Exception as e:
                print(f"  ⚠ 对象 {ent.Handle} 移动失败：{e}")

        # 6. 等待 CAD 完成命令再进入下一轮
        time.sleep(ty)
    else:
        print("⚠ 达到最大迭代次数，可能仍有残留对象。")





###矩形标准化及整体优化
"""
使用parse_rectangle_points函数接受用户对矩形对角点两种方式的自由理解,使用它

改造已有函数，从而修复这种因为对角点不同理解的混乱

jd()应该在连接时就设置，控制绘图精度

有些函数处理还可以优化，显性选择不要在窗口上遮挡，不要动窗口

稳定性和强壮性随使用反馈变化

复制一个参照版本便于查阅函数

以数据库为基础不断扩展数据，扩张整个绘图系统的想法是对的


应该有一个文件初始化函数，将处理文件转入标准设置状态，从精度，字体，墙线显示加粗，打印空间安全，打印配置，视图辅助环境等等。


允许输入一条多段线或多根多段线，改善函数接口，这种错误是合乎人类思维的


在两个字典之间定义函数，通过字典查找确定信息








"""


#&&% 字典的生成、使用、扩展

"""
##{
##  # 顶层键：每个图签框的 BlockReference.Handle（字符串）
##  "2BC3": {
##      # ① 这条记录对应的打印框信息，统一放到一个子字典里
##      "frame_info": {
##          # 4 个角点（3D）    
##          "corners": [
##              (x1, y1, z1),
##              (x1, y2, z1),
##              (x2, y2, z1),
##              (x2, y1, z1),
##          ],
##          # ISO 编号或名称
##          "format": "ISO_A2_(594.00_x_420.00_MM)",
##          # 表示比例
##          "scale":  "1:100",
##          # 纸张规格（A2/A3/A1…）
##          "paper_size": "A2",
##      },
##      # ② 插入成功后的标题块引用  
##      "title_block_handle": "2C10",         # Handle 字符串
##      "title_block_entity": <COMObject ...> # （可选，方便立刻操作）
##  },
##
##  "2BC4": {
##    # ……
##  }
##}
##
##在上面示例里：
##
##顶层 key 用 Handle（字符串形式）来索引每一条记录；
##
##frame_info：一个“打印框”相关的子字典，里面再细分
##
##corners：四个 3D 坐标点
##
##format：图幅名称（ISO_…）
##
##scale：打印比例（比如 "1:100"）
##
##paper_size：规格（比如 "A2"、"A3"）
##
##title_block_handle：新插入的图签块引用的 Handle，用于快速查回或二次处理；
##
##title_block_entity（可选）：保留一份 COM 对象，以免你后面还想马上调用它的属性／方法。
##
##为什么这样命名更易维护？
##语义明确
##
##format、scale、paper_size 一看就知道分别代表图幅、比例、规格。
##
##层次分明
##
##所有跟“打印框”相关的内容都放 frame_info，以后要扩展（比如加 “偏移量 offset”）也可以放进去。
##
##与之平行的是标题块信息 (title_block_…)。
##
##Handle 做主键
##
##外层直接用 BlockReference.Handle 字符串，既不易冲突，也方便序列化、存盘。
##
##保留实体或句柄
##
##存下 title_block_entity 可直接调用；存下 title_block_handle 可在重载文档后重新查找实例。
##
##你可以据此把已有的 zd_tuqian 做一次重命名／重构，例如：
##
##python
##复制
##编辑
##def normalize_tuqian_dict(raw: dict) -> dict:
##    
##    将原始的 zd_tuqian 转换为更清晰的结构：
##    {
##      handle: {
##        "frame_info": {
##          "corners":    …,
##          "format":     …,
##          "scale":      …,
##          "paper_size": …,
##        },
##        "title_block_handle": …,
##        "title_block_entity": …
##      }, …
##    }
##    
##    new = {}
##    for handle, info in raw.items():
##        fi = info["frame_info"]
##        new[handle] = {
##            "frame_info": {
##                "corners":    fi["corners"],
##                "format":     fi["drawing_frame"],
##                "scale":      fi["ratio"],
##                "paper_size": fi["spec"],
##            },
##            "title_block_handle": getattr(info["title_block"], "Handle", None),
##            "title_block_entity": info["title_block"],
##        }
##    return new
##这样，后续所有基于 zd_tuqian 读取比例、规格等，都能写成：
##
##entry = zd_tuqian[handle]
##scale     = entry["frame_info"]["scale"]
##paper_sz  = entry["frame_info"]["paper_size"]
##blk_handle = entry["title_block_handle"]
##希望这个命名规范能帮助你把字典结构理得更清晰，也更容易维护和扩展。
##
##
##entities = [info["entity"] for info in res.values()]
##specs     = [info["spec"]   for info in res.values()]
##ratios    = [info["ratio"]  for info in res.values()]
##
##info =bindy.get("2D07")
##info['title_block']
##字典的产生，保存，调用非常有价值，它将为我们存取大量信息，也改变了整个程序结构
##
### 拿到某个 title_block_handle 的 info
##info = bind_dict[handle]
##
### 移动图签块
##tb = info['title_block']
##tb.Move(vtpnt(0,0,0), vtpnt(0,10000,0))
##
### 移动对应的打印框 polyline
##pl = info['frame_info']['entity']
##pl.Move(vtpnt(0,0,0), vtpnt(0,10000,0))


"""

def 圆点(tz=1):


    """
    控制点的显示

    """

    zhi=0

    if tz ==1:

        zhi=35

    acad.ActiveDocument.SetVariable("PDMODE", zhi)#就是点最好的圆加十字形显示



def 图纸背景(zhi = 16777215 ):

    acad.ActiveDocument.Application.preferences.Display.GraphicsWinModelBackgrndColor = zhi#0即变成黑色





#&&&%  *** 按区域调整视图

def  shitu_region(x1,y1,x2,y2):

    """
    按按对象外包盒调整视图
    使用时可shitu_region(*p),p=(x1,y1,x2,y2)

    """

    h=0.3*(abs(x1-x2)+abs(y1-y2))/2

    # 1️⃣ 缩放视图到合适窗口
    zoom_cmd = (
        f"_.ZOOM\n"      # 调用 Zoom
        f"_W\n"          # 窗口选项
        f"{x1-h},{y1-h}\n"   # 第一个角点
        f"{x2+h},{y2+h}\n"   # 第二个角点
    )
    doc.SendCommand(zoom_cmd)

#按对象外包盒调整视图

def  shitu_entity(obj):

    """
    按按对象外包盒调整视图

    """
    p1,p2=obj.GetBoundingBox()

    x1,y1=p1[0],p1[1]

    x2,y2=p2[0],p2[1]

    h=0.3*(abs(x1-x2)+abs(y1-y2))/2

    # 1️⃣ 缩放视图到合适窗口
    zoom_cmd = (
        f"_.ZOOM\n"      # 调用 Zoom
        f"_W\n"          # 窗口选项
        f"{x1-h},{y1-h}\n"   # 第一个角点
        f"{x2+h},{y2+h}\n"   # 第二个角点
    )
    doc.SendCommand(zoom_cmd)


def ensure_list(obj, element_type=None):
    """
    如果 obj 已经是列表，则原样返回；
    否则，如果没有指定 element_type，或 obj 是 element_type 的实例，
      则将其包装成单元素列表返回；
    否则抛出 TypeError。

    :param obj: 单个元素或元素列表
    :param element_type: 列表元素应有的类型（可选）
    :return: 元素列表
    """
    # 1) 如果已经是列表，直接返回
    if isinstance(obj, list):
        return obj

    # 2) 要么不限制类型，要么 obj 是指定类型
    if element_type is None or isinstance(obj, element_type):
        return [obj]

    # 3) 其它情况报错
    raise TypeError(
        f"期望类型 {element_type.__name__} 或 List[{element_type.__name__}]，"
        f"但收到 {type(obj).__name__}"
    )


#  主函数
#  (1)
# 某某功能作用的函数

#  该函数系列包括如下一些函数

"""
文件处理之前执行
collect_all_texts()
将所有文字放入天正"PUB_TEXT"图层，后续无需再去重复

执行大文件变文件夹+每张图一个文件，减少不必要的选择运算


"""


#&&% 窗口键盘控制
"""
# 1. 截取全屏，返回一个 PIL.Image 对象
img = pyautogui.screenshot()

# 2. 如果想直接保存到文件：
pyautogui.screenshot('full_screen.png')

# 3. 只截取屏幕的一部分（x, y, width, height）
#    例如：从左上角 (100,100) 开始，截取 300×200 的区域
region_img = pyautogui.screenshot(region=(100, 100, 300, 200))
region_img.save('partial.png')


"""

#GIF录屏

def record_screen_gif(
    output_path: str,
    duration: float = 5.0,
    fps: int = 10,
    region: tuple[int,int,int,int] | None = None
):
    """
    录制屏幕并保存为 GIF。

    :param output_path: 输出 GIF 文件路径，比如 "demo.gif"
    :param duration:    录制时长（秒）
    :param fps:         帧率（每秒截多少帧）
    :param region:      可选，(left, top, width, height) 只录制这一区域
    """
    import imageio 

    frames = []
    interval = 1.0 / fps
    end_time = time.time() + duration

    print(f"▶ 开始录制：时长={duration}s，帧率={fps}fps，区域={region or '全屏'}")
    while time.time() < end_time:
        img = pyautogui.screenshot(region=region)  # PIL.Image 对象
        frames.append(img)
        time.sleep(interval)

    print(f"🛑 录制结束，共捕获 {len(frames)} 帧，正在合成 GIF……")
    # imageio 能直接接受 PIL.Image
    imageio.mimsave(output_path, frames, fps=fps)
    print(f"✅ 已保存为 {output_path}")


#动画录屏










def minimize_all_windows():
    """
    模拟按下 Win+M，将所有窗口最小化。

    """

    import ctypes

    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B   # 左 Win 键
    VK_M    = 0x4D   # M 键
    KEYEVENTF_KEYUP = 0x2

    # 按下 Win
    user32.keybd_event(VK_LWIN, 0, 0, 0)
    # 按下 M
    user32.keybd_event(VK_M,    0, 0, 0)
    # 松开 M
    user32.keybd_event(VK_M,    0, KEYEVENTF_KEYUP, 0)
    # 松开 Win
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)

    # 给系统一点反应时间
    time.sleep(0.1)



#&&%#控制CAD屏幕窗口在左上角


def set_autocad_window_to_top_left(resize_half: bool = True):
    """
    将 AutoCAD 窗口还原并移动到屏幕左上角，可选将其调整为半屏大小。
    """
    # 1️⃣ 找到可见或最小化的 AutoCAD 窗口
    windows = [w for w in gw.getWindowsWithTitle('AutoCAD')]
    if not windows:
        print("❌ 未找到 AutoCAD 窗口！")
        return
    win = windows[0]

    # 2️⃣ 如果窗口最小化，先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.3)

    # 3️⃣ 激活窗口，确保我们操作的是真正的前台窗口
    try:
        win.activate()
    except Exception:
        # 有时 activate 也会失败，短暂等待再试
        time.sleep(0.2)
        win.activate()
    time.sleep(0.2)

    # 4️⃣ 移动到左上角 (0,0)
    win.moveTo(0, 0)
    time.sleep(0.2)

    # 5️⃣ 可选：将窗口调整为屏幕一半大小
    if resize_half:
        screen_w, screen_h = pyautogui.size()
        win.resizeTo(screen_w // 2, screen_h // 2)
        time.sleep(0.2)
        print(f"✅ AutoCAD 窗口已恢复并移动到左上角，尺寸设为 {screen_w//2} x {screen_h//2}")
    else:
        print("✅ AutoCAD 窗口已恢复并移动到左上角")



def l():

    set_autocad_window_to_top_left()



#&&% 更合理控制窗口函数
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

def minimize_all_windows_d():
    """
    模拟 Win + D，将所有窗口最小化（切换）。
    """
    # VK_LWIN = 0x5B, VK_D = 0x44, KEYEVENTF_KEYUP = 0x2
    ctypes.windll.user32.keybd_event(0x5B, 0, 0, 0)       # 按下 Win
    ctypes.windll.user32.keybd_event(0x44, 0, 0, 0)       # 按下 D
    ctypes.windll.user32.keybd_event(0x44, 0, 0x2, 0)     # 松开 D
    ctypes.windll.user32.keybd_event(0x5B, 0, 0x2, 0)     # 松开 Win
    time.sleep(0.3)


def minimize_all_windows_m():
    """
    模拟按下 Win+M，将所有窗口最小化。
    """
    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B   # 左 Win 键
    VK_M    = 0x4D   # M 键
    KEYEVENTF_KEYUP = 0x2

    # 按下 Win
    user32.keybd_event(VK_LWIN, 0, 0, 0)
    # 按下 M
    user32.keybd_event(VK_M,    0, 0, 0)
    # 松开 M
    user32.keybd_event(VK_M,    0, KEYEVENTF_KEYUP, 0)
    # 松开 Win
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)

    # 给系统一点反应时间
    time.sleep(0.1)


def restore_and_position(
    name: str = "AutoCAD",
    width_ratio: float = 0.5,
    height_ratio: float = 0.5,
    x: int = 0,
    y: int = 0
) -> bool:
    """
    将第一个标题中包含 name 的窗口恢复、激活，并调整到指定位置和大小。

    :param name:          窗口标题关键字，默认 "AutoCAD"
    :param width_ratio:   窗口宽度占屏幕宽度的比例 (0 < ratio ≤ 1)
    :param height_ratio:  窗口高度占屏幕高度的比例 (0 < ratio ≤ 1)
    :param x:             窗口左上角 X 坐标，默认 0
    :param y:             窗口左上角 Y 坐标，默认 0
    :return:              找到并操作成功返回 True，否则 False
    """
    # ① 查找窗口
    candidates = [w for w in gw.getWindowsWithTitle(name) if w.title]
    if not candidates:
        print(f"❌ 未找到标题包含 “{name}” 的窗口。")
        return False
    win = candidates[0]

    # ② 如果最小化，则还原
    if win.isMinimized:
        try:
            win.restore()
            time.sleep(0.2)
        except Exception as e:
            print(f"⚠ 无法还原窗口：{e}")

    # ③ 激活窗口（置于最前）
    try:
        win.activate()
    except Exception:
        time.sleep(0.1)
        try:
            win.activate()
        except Exception as e:
            print(f"⚠ 激活窗口失败：{e}")

    time.sleep(0.2)

    # ④ 移动到指定位置
    try:
        win.moveTo(x, y)
    except Exception as e:
        print(f"⚠ 移动窗口失败：{e}")

    time.sleep(0.2)

    # ⑤ 调整窗口大小
    sw, sh = pyautogui.size()
    # 限制比例范围
    wr = max(0.01, min(1.0, width_ratio))
    hr = max(0.01, min(1.0, height_ratio))
    new_w = int(sw * wr)
    new_h = int(sh * hr)
    try:
        win.resizeTo(new_w, new_h)
        time.sleep(0.2)
    except Exception as e:
        print(f"⚠ 调整窗口大小失败：{e}")

    print(f"✅ 已将窗口“{win.title}”移动到 ({x},{y})，并调整为 {new_w}×{new_h}（占屏幕 {wr*100:.0f}% × {hr*100:.0f}%）")

    return True


"""
restore_and_position_cad(
    "微信",
    width_ratio = 0.5,
    height_ratio = 0.5,
    x= 0,
    y= 0
)

"""

def list_open_window_titles() -> list[str]:
    """
    获取当前所有可见窗口的标题列表包括子窗口。

    :return: 一个字符串列表，每一项都是一个非空窗口标题。
    """
    titles = []
    for w in gw.getAllWindows():
        title = w.title.strip()
        if title:
            titles.append(title)
    return titles
#&&% * 测鼠标位置
def ceshubiao_weizhi():
    """
    提示用户 5 秒内将鼠标移动到 AutoCAD 命令栏输入位置，
    然后采集当前鼠标坐标并返回 (x, y)。
    """
    print("请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…")
    time.sleep(5)
    x, y = pyautogui.position()
    print(f"已获取坐标：({x}, {y})")
    return x, y

def run_idle_background(script_path: str):
    """
    用后台模式启动 IDLE 去运行某个脚本，返回 Popen 实例。
    """
    # Windows 上隐藏窗口的 flag
    CREATE_NO_WINDOW = 0x08000000
    proc = subprocess.Popen([
            sys.executable, "-m", "idlelib", "-r", script_path
        ],
        creationflags=CREATE_NO_WINDOW,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
    )
    return proc



"""
idle_proc = run_idle_background(r"D:/Myprogramsystem/cad/CAD基本操作.py")
# 结束刚才启动的 IDLE
idle_proc.terminate()
idle_proc.wait(timeout=5)
"""

#按下鼠标左键再拖动

def click_and_drag(x: int, y: int, juli: int):
    """
    在屏幕坐标 (x, y) 按下左键，然后向纵向拖动距离 juli。
    若 juli 为正值，则向上拖动；若为负值，则向下拖动。
    """
    # 1. 移动到 (x, y)
    pyautogui.moveTo(x, y)

    time.sleep(2) 

    # 2. 按住左键
    pyautogui.mouseDown(button='left')

    # 3. 拖动：拖动到 (x, y + juli)
    #    如果仅需相对拖动，也可用 dragRel(0, juli)。
    dest_x = x
    dest_y = y - juli  # 注意：屏幕坐标里，向上是 y 减小，向下是 y 增大
    pyautogui.moveTo(dest_x, dest_y, duration=0.2)

    # 4. 松开左键
    pyautogui.mouseUp(button='left')

#寻找指定图片形状
def click_and_find_image_shape(x: int, y: int, tupian_path: str, timeout: float = 10.0):
    """
    在 (x, y) 单击一次，然后不断在整个屏幕上查找与 tupian_path 对应的图片形状，
    一旦找到，就把鼠标移到它的中心并返回中心坐标 (Python int)；超时仍未找到则返回 None。

    :param x, y:        单击的初始坐标（可触发界面更新）
    :param tupian_path: 要识别的目标图片路径（如微信笑脸）
    :param timeout:     最长等待时间（秒）
    :return:            找到时返回 (x, y)；否则返回 None
    """
    if not os.path.isfile(tupian_path):
        raise FileNotFoundError(f"找不到图片文件：{tupian_path}")

    # 1. 右键点击 (x, y)
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.3)  # 等待界面刷新

    start = time.time()
    while True:
        # 2. 搜索图片并获取中心坐标
        loc = pyautogui.locateCenterOnScreen(tupian_path, confidence=0.9)
        if loc:
            cx, cy = loc
            # 转为 Python int 再移动
            cx, cy = int(cx), int(cy)
            pyautogui.moveTo(cx, cy)
            return cx, cy

        # 3. 检查超时
        if time.time() - start > timeout:
            print("❌ 超时，未能识别到指定图片形状。")
            return None

        time.sleep(0.2)

#右键菜单
def right_click_and_move(x: int, y: int, x_xiangdui: int, y_xiangdui: int):
    """
    在屏幕坐标 (x, y) 处执行右键点击，然后将鼠标相对于 (x, y) 
    移动到 (x + x_xiangdui, y + y_xiangdui) 并停留。

    :param x:            初始水平坐标
    :param y:            初始垂直坐标
    :param x_xiangdui:   相对水平偏移（正值向右，负值向左）
    :param y_xiangdui:   相对垂直偏移（正值向下，负值向上）
    """
    # 移动到目标位置并右键点击
    pyautogui.moveTo(x, y)
    pyautogui.click(button='right')
    time.sleep(2)
    # 计算相对目标位置
    dest_x = x + x_xiangdui
    dest_y = y + y_xiangdui
    # 平滑移动到新位置
    pyautogui.moveTo(dest_x, dest_y, duration=0.2)

def kill_all_idle():
    """
    终止所有名为 'idle' 或 'idle.exe' 的进程（不再需要任务管理器）。
    """
    for p in psutil.process_iter(("pid", "name")):
        name = (p.info["name"] or "").lower()
        if name.startswith("idle"):
            try:
                p.terminate()
            except Exception:
                pass






##控制IDLE屏幕窗口在右上角
def set_idle_window_to_top_right():

    # 获取IDLE窗口句柄
    windows = [w for w in gw.getWindowsWithTitle('IDLE') if w.visible]
    
    if not windows:
        print("❌ 未找到IDLE窗口！")
        return
    
    # 获取IDLE主窗口
    win = windows[0]
    
    # 获取屏幕分辨率
    screen_width, screen_height = pyautogui.size()

    # 计算窗口大小：横竖各占1/2
    half_width = screen_width // 2
    half_height = screen_height // 2

    # 将窗口移动到屏幕的右上角 (屏幕宽度的一半，顶部为0)
    win.moveTo(half_width, 0)

    # 调整窗口大小，设置为屏幕的1/2宽度和1/2高度
    win.resizeTo(half_width, half_height)
    
    print(f"✅ IDLE窗口已调整到屏幕右上角，尺寸为 {half_width} x {half_height}")

def r():

    set_idle_window_to_top_right()

##控制OBS窗口在右下角
def place_obs_bottom_right():
    """
    将 OBS Studio 主窗口移动到屏幕右下角，并缩放为屏幕宽高的一半。
    - 若找不到 OBS 窗口，会打印错误信息。
    - 若有多个 OBS 窗口，仅操作第一个可见窗口。
    """
    # 1️⃣ 获取 OBS 主窗口
    obs_windows = [w for w in gw.getWindowsWithTitle('OBS') if w.visible]
    if not obs_windows:
        print('❌ 未找到可见的 OBS Studio 窗口')
        return

    obs = obs_windows[0]            # 取第一个
    print(f'🔍 找到窗口: {obs.title}')

    # 2️⃣ 计算目标尺寸与位置 —— 右下角 1/4 区域
    screen_w, screen_h = pyautogui.size()
    half_w, half_h = screen_w // 2, screen_h // 2
    target_left = screen_w - half_w      # 右半屏起点
    target_top  = screen_h - half_h      # 下半屏起点

    # 3️⃣ 移动并缩放
    obs.moveTo(target_left, target_top)
    time.sleep(0.1)                      # 给系统一点缓冲
    obs.resizeTo(half_w, half_h)

    print(f'✅ 已将 OBS 窗口调整到右下角 {half_w}×{half_h}')

def r2():

    place_obs_bottom_right()




##最小、最大化窗口
def minimize_window(window_keyword: str = 'OBS') -> bool:
    """
    通用：最小化第一个标题包含 window_keyword 的可见窗口。

    :param window_keyword: 要匹配的窗口标题关键字（子串匹配），默认 'OBS'
    :return: 如果成功最小化返回 True，否则返回 False
    """
    # 1) 找到所有匹配的可见窗口
    windows = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not windows:
        print(f'❌ 未找到标题包含 "{window_keyword}" 的可见窗口')
        return False

    # 2) 取第一个并最小化
    win = windows[0]
    print(f'🔍 找到窗口: "{win.title}"，执行最小化')
    win.minimize()
    return True

def maximize_autocad_window(window_keyword: str = 'AutoCAD') -> bool:
    """
    强制最大化第一个标题包含 window_keyword 的可见窗口。
    优先尝试使用 win32gui，如不可用则退回 ctypes 调用 user32。
    """
    # 1) 找到目标窗口
    wins = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not wins:
        print(f"❌ 未找到标题包含 “{window_keyword}” 的可见窗口")
        return False

    win = wins[0]
    hWnd = win._hWnd

    # 2) 先恢复（避免最小化状态），再最大化
    #    尝试使用 win32gui
    try:
        import win32gui, win32con
        win32gui.ShowWindow(hWnd, win32con.SW_RESTORE)
        time.sleep(0.1)
        win32gui.ShowWindow(hWnd, win32con.SW_MAXIMIZE)
    except ImportError:
        # 如果没有 pywin32，就退回 ctypes
        SW_RESTORE  = 9
        SW_MAXIMIZE = 3
        ctypes.windll.user32.ShowWindow(hWnd, SW_RESTORE)
        time.sleep(0.1)
        ctypes.windll.user32.ShowWindow(hWnd, SW_MAXIMIZE)

    time.sleep(0.2)  # 确保窗口完成最大化
    print(f"✅ 已将窗口 “{win.title}” 最大化")
    return True

##控制屏幕位置开始录制或关闭OBS
def start_obs_recording_by_click(x: int = 1768, y: int = 872,
                                 button: str = 'left',
                                 clicks: int = 1,
                                 move_duration: float = 0.2):
    """
    通过鼠标点击屏幕上 (x,y) 坐标来控制 OBS 开始/停止录制。
    
    :param x: 目标点击点 X 坐标
    :param y: 目标点击点 Y 坐标
    :param button: 鼠标按钮，'left'、'right' 或 'middle'
    :param clicks: 点击次数，默认为单击
    :param move_duration: 从当前位置移动到目标的耗时（秒）
    """
    # 安全设置（可选）
    pyautogui.FAILSAFE = True
    pyautogui.PAUSE    = 0.1

    # 1) 平滑移动到目标
    pyautogui.moveTo(x, y, duration=move_duration)
    # 2) 点击
    pyautogui.click(x=x, y=y, clicks=clicks, button=button)
    print(f"✅ 已点击 ({x}, {y})，请检查 OBS 是否已开始/停止录制。")


#&&&% *** 录屏
def fs(x1,y1):
    """
    微信调到0.5窗口

    """
    
    pyautogui.moveTo(x1+595,y1+190)
    pyautogui.click(x1+595,y1+190)
    pyautogui.press("enter")

def xuanqun(x1,y1,neirong):
    
    copy_to_clipboard(neirong)

    pyautogui.moveTo(x1+128,y1+33) 
    pyautogui.click(x1+128,y1+33)   
    time.sleep(2)
    activate_window_by_title("微信")

    time.sleep(2)

    pyautogui.moveTo(x1+128,y1+33)

    activate_window_by_title("微信")

    time.sleep(2)

    pyautogui.click(x1+128,y1+33)
  

    pyautogui.hotkey('ctrl', 'v') 


    pyautogui.moveTo(x1+158,y1+49) 
    pyautogui.click(x1+158,y1+49)   
 

    pyautogui.press("enter")


def copy_to_clipboard(text: str):
    """
    将传入的 text 文本写入系统剪贴板，供后续右键→粘贴使用。
    
    """
    
    # 创建一个隐藏的 tk 根窗口
    r = tkinter.Tk()
    r.withdraw()  # 隐藏主窗口

    # 清空剪贴板并写入新的文本
    r.clipboard_clear()
    r.clipboard_append(text)
    # 必须 update() 一下，确保数据真正存到剪贴板
    r.update()

    # 销毁隐藏窗口
    r.destroy()

def xieweixin(x1,y1,neirong):
        
        copy_to_clipboard(neirong)

        time.sleep(2)
   
        pyautogui.moveTo(x1+532,y1+151)

        time.sleep(2)
        activate_window_by_title("微信")

        
        pyautogui.moveTo(x1+532,y1+151)

        time.sleep(2)

        pyautogui.click(x1+532,y1+151)
      

        pyautogui.hotkey('ctrl', 'v')        

        pyautogui.click(x1+532,y1+151)
    
def 主操作函数():
    restore_and_position(
        name = "微信",
        width_ratio = 0.5,
        height_ratio = 0.5,
        x = 0,
        y = 0
    )  
 
    time.sleep(1)


    x1,y1,_,_ = activate_window_by_title("微信")

    xuanqun(x1,y1,"华新工作群")

    time.sleep(3)

    x2,y2=click_and_find_image_shape(358, 594, r"D:/Myprogramsystem/XT/weixin_xiaolian.png", timeout = 10.0)

    time.sleep(5)

    pyautogui.click(x2,y2) 

    time.sleep(5)

    x3,y3=click_and_find_image_shape(358, 594, r"D:/Myprogramsystem/XT/weixin_daixao_biaoqingbao.png", timeout = 10.0)

    time.sleep(2)
    pyautogui.click(x3,y3) 
    time.sleep(2)


 
    xieweixin(x1,y1,"Hello!我是公司管理员小化身，从今天起将参与公司管理，与各位一起奋进！")

    fs(x1,y1)

    
    time.sleep(1)


def  main_func(folder_path=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/测试"):

    打印输出PDF() 




def luping_1(main_func):
    """
    1) 开始 OBS 录制
    2) 最小化所有窗口；恢复并定位 OBS
    3) 调用传入的 main_func() 执行主操作
    4) 恢复 OBS，停止录制

    main_func: 零参数函数，负责执行插入或其他主操作
    """
    # 最小化所有窗口
    minimize_all_windows_d()

    # 恢复并定位 OBS 窗口
    restore_and_position(
        name="OBS",
        width_ratio=0.5,
        height_ratio=0.5,
        x=0,
        y=0
    )
    time.sleep(1)

    # 点击 OBS “开始录制”按钮（相对坐标）
    x0, y0, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x0 + 824, y0 + 328)
    time.sleep(1)
    pyautogui.click(x0 + 824, y0 + 328)

    # 最小化所有窗口，切换到 CAD
    minimize_all_windows_d()

    # 3️⃣ 执行主操作

    time.sleep(1)
    restore_and_position(
        name="AutoCAD",
        width_ratio=1,
        height_ratio=1,
        x=0,
        y=0
    )
 
    
    main_func()

    print("主操作函数执行完毕")

    # 4️⃣ 恢复并激活 OBS，停止录制
    time.sleep(0.5)
    minimize_all_windows_d()
    restore_and_position(
        name="OBS",
        width_ratio=0.5,
        height_ratio=0.5,
        x=0,
        y=0
    )
    time.sleep(0.2)

    x4, y4, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x4 + 824, y4 + 328)
    time.sleep(1)
    pyautogui.click(x4 + 824, y4 + 328)

    print("录制已停止")


def luping(main_func, *args, **kwargs):
    """
    1) 开始 OBS 录制  
    2) 最小化所有窗口；恢复并定位 OBS  
    3) 切换到 CAD 窗口，调用 main_func(*args, **kwargs）  
    4) 切换回 OBS，停止录制  

    main_func : 任意可调用对象  
    *args, **kwargs : 会原样传给 main_func  
    """
    # —— 最小化所有窗口 ——  
    minimize_all_windows_d()

    # —— 定位并激活 OBS ——  
    restore_and_position(name="OBS", width_ratio=0.5, height_ratio=0.5, x=0, y=0)
    time.sleep(1)

    # 点击 OBS “开始录制”  
    x0, y0, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x0 + 824, y0 + 328)
    time.sleep(0.2)
    pyautogui.click(x0 + 824, y0 + 328)

    # —— 切换到 CAD ——  
    minimize_all_windows_d()
    time.sleep(0.5)
    restore_and_position(name="AutoCAD", width_ratio=1.0, height_ratio=1.0, x=0, y=0)
    time.sleep(1)

    # —— 3️⃣ 执行主操作 ——  
    main_func(*args, **kwargs)
    print("主操作函数执行完毕")

    # —— 停止录制 ——  
    time.sleep(0.5)
    minimize_all_windows_d()
    restore_and_position(name="OBS", width_ratio=0.5, height_ratio=0.5, x=0, y=0)
    time.sleep(0.2)

    x4, y4, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x4 + 824, y4 + 328)
    time.sleep(0.2)
    pyautogui.click(x4 + 824, y4 + 328)

    print("录制已停止")













def mccs():
    
    dw=DWG()

    MC_com =[]

    shitu_entity(kuang)

    for x in LD:

        Ent=dw.insert_door(x,width=600,btn_x=351, btn_y=129, cmd_x=471, cmd_y=455)    

        MC_com.append(Ent)
        
    time.sleep(10)

    for yuan in LByuan:
        
        for x in MC_com:

            transfer_props_by_matchprop(yuan, x, delay=0.5)

            shitu_entity(x)

            time.sleep(2)

        name = (yuan.Layer)[3:]

        print("************************************************************")

        print(f"**************       替换门成{name} 构件       *************")

        print("************************************************************")


        shitu_entity(kuang)        

        time.sleep(10)


def 魔方():

    import 魔方分析

    魔方分析.魔方控制台(r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/自动化(动态)/魔方/分析")



#&&%#运行指定程序名的程序

def run_py(pyname):
    try:
        # 运行指定的 Python 程序，隐藏命令行窗口
        result = subprocess.run(
            [sys.executable, pyname],
            check=True,
            text=True,
            capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        print(f"✅ 程序 {pyname} 执行成功。输出:\n{result.stdout}")
    except subprocess.CalledProcessError as e:
        print(f"❌ 运行 {pyname} 时发生错误: {e}")
        print(f"错误信息: {e.stderr}")
    except FileNotFoundError:
        print(f"❌ 未找到程序 {pyname}。请检查文件名和路径。")




##把鼠标移动到命令行窗口
def focus_cmdline(cmd_x, cmd_y, delay=0.2):
    """
    把鼠标移到命令行并单击，确保焦点回到 AutoCAD 命令栏。

    
    """
    pyautogui.moveTo(cmd_x, cmd_y, duration=delay)
    pyautogui.click()



#&&% 激活窗口和子窗口

def activate_window_by_title(title: str, click_titlebar: bool = True) -> bool:
    """
    激活一个指定标题的窗口。
    
    1. 用 pygetwindow 找到窗口对象；
    2. 如果窗口被最小化，先还原；
    3. 将窗口置为前台（优先使用 win.activate()，失败时用 ctypes 强制）；
    4. （可选）在标题栏中点击一次，确保焦点。
    
    :param title: 要激活的窗口标题（部分匹配）。
    :param click_titlebar: 是否模拟一次点击标题栏，以确保窗口获得焦点。
    :return: True=激活成功，False=未找到窗口。
    """
    USER32 = ctypes.windll.user32
    SW_RESTORE = 9


    # ① 查找标题包含关键字的窗口
    wins = [w for w in gw.getWindowsWithTitle(title) if title in w.title]
    if not wins:
        print(f"❌ 未找到标题包含 “{title}” 的窗口")
        return False
    win = wins[0]

    # ② 如果最小化，先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.2)

    
    # ③ 置为前台：先尝试 pygetwindow.activate()
    try:
        win.activate()
        time.sleep(0.2)
    except Exception:
        # ctypes 强制还原并置前
        hwnd = win._hWnd
        USER32.ShowWindow(hwnd, SW_RESTORE)          # 还原
        USER32.SetForegroundWindow(hwnd)             # 置前
        time.sleep(0.2)

    # ④ 可选：在标题栏中点一下
    if click_titlebar:
        x = win.left + win.width // 2
        y = win.top + max(1, min(30, win.height // 10))
        pyautogui.click(x, y)
        time.sleep(0.1)

    print(f"✅ 已激活窗口：“{win.title}” 位置({win.left},{win.top}) 大小{win.width}×{win.height}")

    return  win.left,win.top,win.width,win.height


def click_in_window(title: str, offset_x: float, offset_y: float, click_titlebar: bool = False) -> bool:
    """
    在指定窗口的某个相对像素位置点击一次（相对于窗口左上角的偏移量）。

    :param title:         窗口标题关键字（部分匹配）
    :param offset_x:      从窗口左上角算起的水平像素偏移量（0 表示左边缘，width 表示右边缘）
    :param offset_y:      从窗口左上角算起的垂直像素偏移量（0 表示顶边缘，height 表示底边缘）
    :param click_titlebar: 是否先在标题栏点击一次以确保窗口获取焦点
    :return:              True=点击成功，False=未找到窗口
    """
    # 1) 查找窗口
    wins = [w for w in gw.getWindowsWithTitle(title) if title in w.title]
    if not wins:
        print(f"❌ 未找到标题包含 “{title}” 的窗口")
        return False
    win = wins[0]

    # 2) 如果最小化，就先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.2)

    # 3) 尝试激活
    try:
        win.activate()
        time.sleep(0.2)
    except Exception:
        # ctypes 强制还原并置前
        hwnd = win._hWnd
        SW_RESTORE = 9
        ctypes.windll.user32.ShowWindow(hwnd, SW_RESTORE)
        ctypes.windll.user32.SetForegroundWindow(hwnd)
        time.sleep(0.2)

    # 4) 可选：点击标题栏让焦点真正到窗口
    if click_titlebar:
        tx = win.left + win.width // 2
        ty = win.top + 10  # 标题栏一般在窗口顶部 10px 左右
        pyautogui.click(tx, ty)
        time.sleep(0.1)

    # 5) 计算目标点的绝对屏幕坐标
    abs_x = int(win.left + offset_x)
    abs_y = int(win.top  + offset_y)

    # 6) 点击
    pyautogui.moveTo(abs_x, abs_y, duration=0.2)
    pyautogui.click(abs_x, abs_y)
    time.sleep(0.1)
    print(f"🔘 已在窗口 “{win.title}” 内部点击 ({offset_x}, {offset_y}) → 绝对 ({abs_x}, {abs_y})")
    return True

"""
click_in_window("图形导出", offset_x=600-10, offset_y=600-10, click_titlebar=True)

"""

#&&% 链接和关闭艾可云

def activate_and_click_aikeyun():


    try:
    
        left, top, width, height = activate_window_by_title("艾可云", click_titlebar=True)
        
        # Calculate relative offsets from provided data
        rel_x1 = 887 - 682  # 205
        rel_y1 = 542 - 93   # 449
        
        # Move to first click position and left click
        pyautogui.moveTo(left + rel_x1, top + rel_y1)
        pyautogui.click(button='left')
        
        # Wait for 2 seconds
        time.sleep(2)
        
        # Calculate close position: using width - 22 for x, as it aligns with top-right close button
        rel_x2 = width - 22  # Equivalent to 400 - 22 = 378 in data
        rel_y2 = 18          # 111 - 93
        
        # Move to close position and left click
        pyautogui.moveTo(left + rel_x2, top + rel_y2)
        pyautogui.click(button='left')
    
    
    
    except:
    
        pass
    
    









def drag_in_window_simple(
    title: str,
    start: tuple[float,float],
    offset: tuple[float,float],
    duration: float = 0.3,
    button: str = 'left',
    absolute_start: bool = False
):
    """
    拖拽函数，支持相对或绝对起点：

    :param title: 窗口标题关键字
    :param start: 起点坐标，(x, y)；
                  如果 absolute_start==False，则当成“窗口内部”坐标
                  如果 absolute_start==True，则当成“屏幕”坐标
    :param offset: 拖拽向量 (dx, dy)
    :param duration: 拖动时长
    :param button: 'left' 或 'right'
    :param absolute_start: True 则 start 视为屏幕绝对坐标
    """
    left, top, w, h = activate_window_by_title(title)

    if absolute_start:
        x1, y1 = start
    else:
        x1 = left + start[0]
        y1 = top  + start[1]

    x2 = x1 + offset[0]
    y2 = y1 + offset[1]

    pyautogui.moveTo(x1, y1)
    pyautogui.mouseDown(button=button)
    pyautogui.moveTo(x2, y2, duration=duration)
    pyautogui.mouseUp(button=button)
    time.sleep(0.1)
    print(f"已在“{title}”窗口拖拽：起点{(x1,y1)} → 终点{(x2,y2)}")


"""
drag_in_window_simple(
    "图形导出",
    start=(10,10),
    offset=(100,50),
    absolute_start=False
)

"""

#&&% 屏幕截图快照

"""
img = pyautogui.screenshot()

img.save('D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/CAD打印/01建筑/screenshot.png')

局部

# region=(left, top, width, height)
img = pyautogui.screenshot(region=(0, 0, 800, 600))

left=0 表示从屏幕最左边开始
top=0 表示从屏幕最上边开始
截取从左上角往右 800 像素、往下 600 像素范围内的区域。

应该是对角点 测算多次得CAD屏幕
img = pyautogui.screenshot(region=(175,101,2350,1250))
img.save('D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/CAD打印/01建筑/screenshot.png')


"""


##自动炸开区域内对象
    
def run_auto_explode_area(x1, y1, x2, y2, cmd_x, cmd_y, delay=2):

    """
    这是一个未使用pywin32API控制天正CAD的典型函数
    它炸开区域x1, y1, x2, y2的所有对象，不适合反复操作

    合理的处理方式还是配合视窗调整（发送z\na或e\n）先slect显性选择对象，再发送命令

    """
    
    script = Path(__file__).with_name('auto_EXPLODE.py')
    cmd = [
        sys.executable, str(script),
        f'--x1={x1}', f'--y1={y1}',
        f'--x2={x2}', f'--y2={y2}',
        f'--cmd_x={cmd_x}', f'--cmd_y={cmd_y}',
        f'--delay={delay}'
    ]
    # 隐藏子进程窗口
    flags = subprocess.CREATE_NO_WINDOW
    try:
        subprocess.run(cmd, check=True, creationflags=flags)
        print('▶ 子程序完成')
    except subprocess.CalledProcessError as e:
        print(f'❌ 子程序退出码 {e.returncode}')





##确保CAD窗口输入法为英文
"""
本函数研究了ctr+空格的转换，如需要可进一步分析ctr+shift转换的问题        

它真正的意义是提供了ocr截图分析方法，这让自动化处理窗口问题更为强大

"""
def ensure_english_input_method(cmd_x: int, cmd_y: int, delay: float = 0.2):
    """
    聚焦 AutoCAD 命令行 -> 输入 a -> OCR 判断是否出现 A(ARC)/AA(AREA)。
    若未出现则执行 Win+Space 强制切换输入法一次。

    参数
    ----
    cmd_x, cmd_y : 命令行输入框的屏幕坐标
    delay        : 鼠标移动耗时（秒）
    """

    # ---------- 辅助 ---------- #
    def click_cmdline():
        pyautogui.moveTo(cmd_x, cmd_y, duration=delay)
        pyautogui.click()

    def set_to_english():
        pyautogui.hotkey("win", "space")       # 系统输入法切换
        time.sleep(2.0)                        # 给系统充足时间

    def is_english() -> bool:
        """输入 a 截图 → OCR → 宽松匹配"""
        pyautogui.write("a")                   # 测试字符
        time.sleep(1.2)                        # 等列表弹出
        img = ImageGrab.grab(bbox=(310, 381, 510, 581))   # 按需调整
        raw = pytesseract.image_to_string(img)
        print(f"OCR 捕获: {raw!r}")

        # 去空白→大写，统一匹配
        cleaned = re.sub(r"\s+", "", raw).upper()
        return ("A(ARC)" in cleaned) or ("AA(AREA)" in cleaned)

    # ---------- 主流程 ---------- #
    click_cmdline()                # 聚焦命令栏

    if is_english():
        print("✅ 已是英文输入法，直接结束。")
        return

    print("❌ 检测到非英文，执行切换 …")
    set_to_english()
    click_cmdline()                # 切换后重新聚焦

    # 切换一次后不再循环；如仍失败可手动检查
    if is_english():
        print("✅ 切换成功。")
    else:
        print("⚠️ 切换后仍未检测到英文，可能 OCR 失效或坐标需调整。")



##列出所有当前窗口

def list_all_windows():
    # 获取所有可见的窗口
    windows = gw.getWindowsWithTitle('')
    
    if not windows:
        print("❌ 没有找到任何可见窗口。")
    else:
        print("当前桌面上的所有窗口：")
        for win in windows:
            print(f"窗口标题: {win.title}, 窗口大小: {win.width}x{win.height}, 位置: ({win.left}, {win.top})")


#关闭垃圾干扰窗口
def close_360_popup_window():
    # 获取所有窗口标题
    windows = gw.getWindowsWithTitle('360')  # '360' 为弹出窗口的部分标题，具体根据实际情况调整

    # 遍历所有窗口，找到匹配的窗口
    for window in windows:
        print(f"找到窗口: {window.title}")

        if '360' in window.title:  # 确认窗口属于360软件（你可以根据实际窗口标题调整）
            # 激活窗口
            window.activate()
            time.sleep(1)

            # 通过模拟按键关闭窗口（如果是弹出的提示框，常见的关闭按钮是 'ESC' 或 'Alt+F4'）
            pyautogui.hotkey('alt', 'f4')  # 模拟 Alt + F4 来关闭窗口
            print(f"✅ 关闭了窗口: {window.title}")
            return

    print("❌ 没有找到匹配的360窗口")







#  主函数
#  (1)
# 天正墙中轴线显示与隐藏

#  该函数系列包括如下一些函数

"""
墙是否加粗边线，墙中线显示，墙中线隐藏是文件属性，因此我们用D:/Myprogramsystem/cad/xitongjicuwenjian/墙基线打开.dwg，

D:/Myprogramsystem/cad/xitongjicuwenjian/墙基线关闭.dwg，D:/Myprogramsystem/cad/xitongjicuwenjian/墙线加粗.dwg三个空文件来

控制文件和程序的运行。因为获取天正墙信息的函数依赖墙中线的显示，所以对不熟悉的文件进行处理时，就可以通过这几个基本文件实现确定

的控制，即我们预期要文件的墙中线显示出来，或者墙边线要加粗，不加粗，依赖实际的需要。




"""
## 测试示例

##__________







def toggle_wall_centerline(conf=0.8, wait=2):
    """
    若 CAD 底栏中 “墙中线显示” 图标为灰色则点击启用，
    否则直接报告已开启。
    """
    import pyautogui as pg

    pg.FAILSAFE = True
    pg.PAUSE    = 0.1           # 全局轻微停顿

    # 模板文件
    BASE = Path(__file__).parent
    BAR_IMG  = BASE / 'bar_block.png'     # 底栏块
    OFF_IMG  = BASE / 'icon_off.png'      # 灰  ≡
    ON_IMG   = BASE / 'icon_on.png'       # 蓝  ≡


    # 1️⃣ 找到底栏大块坐标
    region = pg.locateOnScreen(str(BAR_IMG), confidence=conf)
    if not region:
        print('❌ 未找到底部状态栏块，请确认模板 bar_block.png')
        return
    bar_region = (region.left, region.top, region.width, region.height)

    # 2️⃣ 先查“蓝色已开”图标
    if pg.locateOnScreen(str(ON_IMG), region=bar_region, confidence=conf):
        print('✅ 天正墙中线已经显示（蓝色）。')
        return

    # 3️⃣ 查灰色图标
    off = pg.locateOnScreen(str(OFF_IMG), region=bar_region, confidence=conf)
    if not off:
        print('❌ 未找到灰色图标，可能界面皮肤不同或模板需重截。')
        return

    # 4️⃣ 点击灰色图标
    pg.click(pg.center(off))
    time.sleep(wait)

    # 5️⃣ 再次验证
    if pg.locateOnScreen(str(ON_IMG), region=bar_region, confidence=conf):
        print('✅ 已点击，天正墙中线现在已显示（蓝色）。')
    else:
        print('⚠️ 点击后仍未检测到蓝色图标，请手动检查。')





##使用辅助大矩形控制视图范围

def draw_shitu_rectangle_lw(length=500000.0, width=500000.0, center=(0.0, 0.0), layer_name="shitu"):
    """
    在 AutoCAD 中绘制一个以 center 为中心、长 length、宽 width 的闭合轻量级多段线矩形，
    并放到 layer_name 图层（不存在则新建）。

    参数
    ----
    length, width : float
        矩形的长和宽
    center : tuple(float, float)
        矩形中心点 (x, y)
    layer_name : str
        指定图层名称
    """

    # 2️⃣ 确保目标图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
        lyr.Color = 3  # 设置图层颜色（可选）

    # 3️⃣ 计算矩形顶点（轻量级多段线只要二维坐标）
    cx, cy = center
    hl = length / 2.0
    hw = width  / 2.0

    pts = [
        cx - hl, cy - hw,
        cx + hl, cy - hw,
        cx + hl, cy + hw,
        cx - hl, cy + hw,
        cx - hl, cy - hw,  # 回到起点闭合
    ]

    # 4️⃣ 把 Python 列表转换为 COM-safe 的双精度数组
    variant_pts = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        pts
    )

    # 5️⃣ 添加轻量级多段线并设置属性
    lwpoly = mp.AddLightWeightPolyline(variant_pts)
    lwpoly.Closed = True
    lwpoly.Layer  = layer_name

    # 6️⃣ 可选：缩放到图形 extents
    acad.ZoomExtents()

    print(f"✅ 已在图层 “{layer_name}” 上绘制 {length}×{width} 矩形（轻量级多段线）")














#&&%设置长度角度单位精度

def set_dwg_units_precision():
    """
    设置当前 DWG 文件的单位及精度：
    - 长度单位：单位类型不变，仅设置精度为 0.00000000
    - 角度单位：单位类型不变，精度为 0.00000000
    """

    try:
        vars = doc.GetVariable

        # 设置长度精度（LUPREC = 8 表示 8 位小数）
        doc.SetVariable("LUPREC", 8)

        # 设置角度精度（AUPREC = 8 表示 8 位小数）
        doc.SetVariable("AUPREC", 8)

        print("✅ 已将长度和角度单位精度设置为 8 位小数 (0.00000000)")
    except Exception as e:
        print(f"❌ 设置失败: {e}")

def jd():
    set_dwg_units_precision()


##标注样式

def list_dim_styles():
    """
    列出当前 DWG 文件中所有标注样式名称。
    """
    try:
        styles = doc.DimStyles
        names = [styles.Item(i).Name for i in range(styles.Count)]
        print("📐 当前标注样式列表：")
        for name in names:
            print(" -", name)
        return names
    except Exception as e:
        print(f"❌ 获取标注样式失败：{e}")
        return []



def set_current_dimstyle_via_command(style_name="_TCH_ARCH"):
    """
    使用命令行方式设置当前标注样式，兼容天正。

    参数:
        style_name (str): 要设为当前标注样式的名称（如 "_TCH_ARCH"）
    """
    try:
        doc.SendCommand(f"-DIMSTYLE\nR\n{style_name}\n")
        print(f"✅ 已尝试通过命令行设置标注样式为：{style_name}")
    except Exception as e:
        print(f"❌ 命令行设置标注样式失败：{e}")



#&&% 文字左对齐到垂直线

def align_texts_to_x_by_llcorner(texts, x_target):
    """
    将 texts 列表中每个文字对象的外包盒左下角 X 对齐到 x_target，Y、Z 不变。

    texts: List[COMObject]    单行或多行文字对象
    x_target: float           目标 X 坐标
    """
    base_pt = vtpnt(0, 0, 0)
    for txt in texts:
        try:
            # 取左下角点
            ll_pt, _ = txt.GetBoundingBox()
            x0, y0, z0 = ll_pt
            # 计算平移向量
            dx = x_target - x0
            move_vec = vtpnt(dx, 0, 0)
            # 平移
            txt.Move(base_pt, move_vec)
        except Exception as e:
            h = getattr(txt, "Handle", "?")
            print(f"⚠ 对象 {h} 对齐失败：{e}")




#&&% 获取文字内容

def extract_text_content(ent):
    """
    根据实体类型获取文本内容，并仅移除文字首尾的空格：
      - AcDbText 或 AcDbMText: ent.TextString
      - TDbText: ent.Text
      - TDbMText: TDbMText_content(ent)
    只调用 strip() 删除首尾空白，不影响中间空格。
    """
    obj = ent.ObjectName
    if obj in ("AcDbText", "AcDbMText"):
        raw = ent.TextString
    elif obj == "TDbText":
        raw = ent.Text
    elif obj == "TDbMText":
        raw = TDbMText_content(ent)
    else:
        return ""
    # 仅删除首尾空白
    return raw.strip()





#&&% 设置当前使用的字体样式

        
def set_current_text_style(style_name="Standard"):
    """
    设置当前文字样式（通过 COM 接口方式）。

    参数:
        style_name (str): 要设置为当前的文字样式名称
    """
    try:
        text_styles = doc.TextStyles
        style = text_styles.Item(style_name)  # 获取指定样式
        doc.ActiveTextStyle = style           # 设置为当前样式
        print(f"✅ 当前文字样式已设置为：{style_name}")
    except Exception as e:
        print(f"❌ 设置文字样式失败：{e}")

#&&% 获取所有当前字体样式

def huoqu_ziti_style():

    styles = {ts.Name for ts in doc.TextStyles}
    
    return styles




#&&% 设置字体样式

"""
1，可多次对某种样式设置再更新

2，对汉字字体设置，不存在要设大字体。可用字体在C:/Windows/Fonts查找，直接用汉字名，使用create_text_style命令

3， 可以单独设置shx字体，使用set_text_style_onlyshx(style_name="style01", font_file="gbenor.shx", big_font_file=None)

所造字体缺失大字体，对汉字会产生问号

4，如不是单独设置汉字字体，完整的字体设置应包括英文部分的设置和大字体设置，用于中文显示而且是shx字体

 

"""
def create_text_style(sty_name="style01", ziti="宋体"):
    """
    在当前 DWG 中创建（或更新）一个中文文字样式。
    

    :param sty_name: 样式名称，默认 "style01"
    :param ziti:     字体名称，AutoCAD 会在系统字体中查找，例如 "宋体"
    # acad.ActiveDocument.ActiveTextStyle.SetFont(Typeface, Bold, Italic, charSet, PitchandFamily)
    # Typeface 字体名称；
    # Bold 加粗，布尔值，False为不加粗字体；
    # Italic 倾斜，布尔值，False为倾斜字体；
    # CharSet 字体字符集，1为默认字符集；
    # PitchAndFamily 字节及笔画形式。

    ts = doc.TextStyles.Item("HIT_TxtStyle")对shx

    ts.fontFile = "bigfont.shx"



    """

    # 1️⃣ 确保样式存在
    styles = doc.TextStyles
    try:
        ts = styles.Item(sty_name)
        print(f"⚠ 样式 '{sty_name}' 已存在，正在更新其属性。")
    except Exception:
        ts = styles.Add(sty_name)
        print(f"✅ 已创建文字样式 '{sty_name}'。")

    # 2️⃣ 设置字体
    try:
        acad.ActiveDocument.TextStyles.Item(sty_name).SetFont(ziti, False, False, 1, 0 or 0)
    except Exception:
        try:
            acad.ActiveDocument.TextStyles.Item(sty_name).SetFont(ziti, False, False, 1, 0 or 0)
        except Exception as e:
            print(f"⚠ 无法设置字体为 '{ziti}'：{e}")

    # 3️⃣ 置为当前

    acad.ActiveDocument.ActiveTextStyle = acad.ActiveDocument.TextStyles.Item(sty_name)
    
    # 5️⃣ 通知用户
    print(f"✅ 样式 '{sty_name}' 属性已更新")




def set_text_style_onlyshx(style_name="style01", font_file="gbenor.shx", big_font_file=None):
    """
    C:/Program Files/Autodesk/AutoCAD 2021/Fonts查找可用shx字体    
    设置 AutoCAD 的文字样式：
    - font_file：英文字体（.shx 或 .ttf）
    - big_font_file：可选的大字体（如 gbcbig.txt），默认为 None 表示不设置，仅设置英文字体
    set_text_style_onlyshx(style_name="TEST_STYLE", font_file="gbxxx.shx", big_font_file=None)
    不成功消息可用于判定shx字体在不在左侧英文字体


    """
    try:
        import win32com.client

        acad = win32com.client.Dispatch("AutoCAD.Application")
        doc = acad.ActiveDocument
        styles = doc.TextStyles

        # 查找或创建字体样式
        if style_name in [s.Name for s in styles]:
            text_style = styles.Item(style_name)
        else:
            text_style = styles.Add(style_name)

        text_style.FontFile = font_file

        # 只有传入合法的大字体路径才设置 BigFontFile，否则跳过设置
        if big_font_file and isinstance(big_font_file, str):
            text_style.BigFontFile = big_font_file

        print(f"字体样式 '{style_name}' 设置成功：英文字体 = {font_file}，大字体 = {big_font_file or '未设置'}")
        return True

    except Exception as e:
        print(f"设置字体样式失败：{e}")
        return False



def set_text_style(style_name="style01", font_file="gbenor.shx", big_font_file="gbcbig.shx"):
    """
    设置CAD中文字样式：英文shx文件 + 中文大字体文件
    """
    try:

        styles = doc.TextStyles

        # 判断是否已有该样式
        if style_name in [s.Name for s in styles]:
            text_style = styles.Item(style_name)
        else:
            text_style = styles.Add(style_name)

        # 设置字体和大字体
        text_style.FontFile = font_file
        text_style.BigFontFile = big_font_file

        print(f"字体样式 '{style_name}' 设置成功，英文字体: {font_file}, 中文大字体: {big_font_file}")
        return True

    except Exception as e:
        print(f"设置字体样式失败：{e}")
        return False

#&&% 列出可用shx非大字表 shx大字表 

"""

到CAD字体设置下拉菜单找

"""

##&&% 问号字体替换
"""
网友的ftst方案已经彻底解决此问题，我们需要的是备份好文件和进一步编制点击窗口的函数
我建议使用Standard样式的gbenor.shx和大字gbcbig.shx替换更安全
"""



#&&% ## 处理字体样式同名问题

def rename_conflicting_text_styles(file1_path: str,
                                   file2_path: str,
                                   suffix: str = "_1",
                                   retry_delay: float = 0.2,
                                   max_retries: int = 10):
    """
    在两个 DWG 中找出同名（用户）文字样式，
    并在第一个文件中将它们重命名（原名 + suffix）：
      1) 通过 -RENAME 命令重命名样式
      2) 确认样式表里旧名已消失、新名已出现
      3) REGEN 强制刷新
      4) 遍历 ModelSpace，将引用旧样式的实体指向新样式
      5) 保存并关闭
    系统默认样式会被自动跳过。
    """
    SYSTEM_STYLES = {"Standard", "ASHADE", "Annotative", "BigFont"}

    acad = win32com.client.Dispatch("AutoCAD.Application")
    acad.Visible = True

    doc1 = acad.Documents.Open(os.path.abspath(file1_path))\n    wait_document_opened(acad, os.path.abspath(file1_path))\n    wait_cad_idle(acad)
    doc2 = acad.Documents.Open(os.path.abspath(file2_path))\n    wait_document_opened(acad, os.path.abspath(file2_path))\n    wait_cad_idle(acad)\n
    try:
        # 1) 收集两文件的样式
        styles1 = {ts.Name for ts in doc1.TextStyles}
        styles2 = {ts.Name for ts in doc2.TextStyles}
        conflicts = (styles1 & styles2) - SYSTEM_STYLES
        if not conflicts:
            print("✅ 未发现需要重命名的用户样式。")
            return

        print(f"⚠ 发现同名用户样式：{conflicts}，将在 “{os.path.basename(file1_path)}” 中重命名：")
        ms = doc1.ModelSpace

        for old_name in conflicts:
            new_name = old_name + suffix
            # 如果新名也冲突，就多加后缀，直到独一
            while new_name in styles1:
                new_name += suffix

            # 2) 发送 RENAME 命令
            cmd = f"-RENAME\nStyle\n{old_name}\n{new_name}\n\n"
            doc1.SendCommand(cmd)
            # 等待命令被处理、样式表更新
            for attempt in range(max_retries):
                time.sleep(retry_delay)
                # 强制刷新图形状态
                doc1.SendCommand("REGEN\n")
                try:
                    # 如果旧名已不存在并且新名存在，就跳出重试
                    doc1.TextStyles.Item(new_name)
                    try:
                        doc1.TextStyles.Item(old_name)
                        # 旧名仍然存在，继续等
                        continue
                    except Exception:
                        # 旧名被正确移除
                        break
                except Exception:
                    # 新名还没出现，继续等
                    continue
            else:
                print(f"  ❗ 重命名 “{old_name}” → “{new_name}” 可能未生效（超时）。")

            # 3) 再把所有实体中引用旧样式的改成新样式
            for ent in ms:
                try:
                    ename = getattr(ent, "EntityName", "").upper()
                    oname = getattr(ent, "ObjectName", "")
                    # 原生 TEXT/MTEXT
                    if ename in ("TEXT", "MTEXT"):
                        if ent.TextStyle == old_name:
                            ent.TextStyle = new_name
                    # 天正文字
                    elif oname in ("TDbText", "TDbMText"):
                        if ent.TextStyle == old_name:
                            ent.TextStyle = new_name
                except Exception:
                    # 有些实体可能不允许改样式，忽略它们
                    pass

            # 更新本地样式集合
            styles1.discard(old_name)
            styles1.add(new_name)
            print(f"  · 样式 “{old_name}” → “{new_name}”")

        # 4) 最后保存并反馈
        doc1.Save()
        print(f"✅ 已保存改动到 “{os.path.basename(file1_path)}”。")

    finally:
        # 关闭文档，不保存对第二个文件的任何改动
        doc1.Close(False)
        doc2.Close(False)




















##将一个对象属性传给多个对象

def transfer_props_by_matchprop(entity, Ob, max_try=3, delay=0.4):

    CR = chr(13)

    def wait_idle(acad, dt=0.2, n=50):
        """轮询 IsQuiescent，最多 n 次，每次 dt 秒"""
        for _ in range(n):
            if acad.GetAcadState().IsQuiescent:
                return
            time.sleep(dt)

    def expand_rectangle(p1, p2, offset):
        return (p1[0]-offset, p1[1]-offset), (p2[0]+offset, p2[1]+offset)


    """
    把 entity 的属性批量复制到 Ob。若 Layer 未变化则重试，最多 3 次。
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    src_layer   = entity.Layer
    orig_layer  = Ob.Layer        # 复制前目标的图层

    # 目标包围盒窗口
    p1, p2 = Ob.GetBoundingBox()
    x1, y1, x2, y2 = p1[0], p1[1], p2[0], p2[1]
    h = 0.1 * (abs(x1 - x2) + abs(y1 - y2)) / 2
    (wx1, wy1), (wx2, wy2) = expand_rectangle(p1, p2, h)

    match_cmd = (
        "_MATCHPROP" + CR +
        "P"          + CR +            # Previous 作为源
        "_W"         + CR +
        f"{wx1},{wy1}" + CR +
        f"{wx2},{wy2}" + CR + CR
    )

    for attempt in range(1, max_try + 1):
        try:
            # ——— 1. 设定源对象为 Previous ———
            try:
                highlight_entity_by_bbox(entity)
            except Exception:
                entity.Select(True)        # 退而求其次
            time.sleep(delay)

            # ——— 2. 发送 MATCHPROP ———
            doc.SendCommand(match_cmd)
            wait_idle(acad)

            # ——— 3. 判断是否成功 ———
            if Ob.Layer == src_layer:
                print(f"[OK] 第 {attempt} 次匹配成功，Layer 改为 {src_layer}")
                return True

            print(f"[WARN] 第 {attempt} 次后 Layer 未变，重试…")
            time.sleep(delay)

        except Exception as e:
            print(f"[ERR] 第 {attempt} 次匹配异常：{e}")

    print(f"[FAIL] 连续 {max_try} 次仍未把属性复制给目标")
    return False


#视图合理化控制
"""
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_swiso"+chr(13))#西南轴测
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_seiso"+chr(13))#东南轴测
acad.ActiveDocument.SendCommand(
    "_-view" + chr(13) +
    "_nwiso" + chr(13)
)#西北轴测
acad.ActiveDocument.SendCommand(
    "_-view" + chr(13) +
    "_neiso" + chr(13)
)#东北轴测
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_top"+chr(13))#俯视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_bottom"+chr(13))#仰视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_front"+chr(13))#前视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Right"+chr(13))#右视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Back"+chr(13))#后视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Left"+chr(13))#左视图
acad.ActiveDocument.SendCommand("_vscurrent"+chr(13)+"_R"+chr(13))#真实视觉样式

##合理显示对象

acad.ActiveDocument.SendCommand(
    "_z" + chr(13) +
    "_e" + chr(13)
)

acad.ActiveDocument.SendCommand(
    "_zoom" + chr(13) +
    "s"     + chr(13) +
    "0.8xp" + chr(13)
)




"""
#&&% 双线程生成器

def run_dual_threads_1(f1,                     # 线程1 函数

                     f2,                     # 线程2 函数

                     f1_args=(), f1_kwargs=None,

                     f2_args=(), f2_kwargs=None,

                     timeout_sec=180):



    """
    通用“双线程-GUI”调度器

    - f1 必须负责触发一个阻塞性窗口（如 SendCommand、ShowModalDlg 等）。
    - f2 负责侦测并自动化处理该窗口，当处理完毕后通知 f1 继续或退出。
    - f1、f2 的函数签名都应以 (timeout_event, done_event, …) 开头：
    
      def f1(timeout_event, done_event, …):
          pythoncom.CoInitialize()
          try:
              # ……主体代码：例如触发打印对话框
              timeout_event.wait()    # 等待线程2通知或超时
          except Exception:
              # 打印日志或其他处理
          finally:
              pythoncom.CoUninitialize()
              done_event.set()        # 通知“双线程”总控：我结束了

      def f2(timeout_event, done_event, …):
          pythoncom.CoInitialize()
          try:
              # ……主体代码：例如等待“创建打印文件”对话框出现并点击“回车”键
          except Exception:
              # 打印日志或其他处理
          finally:
              pythoncom.CoUninitialize()
              timeout_event.set()    # 通知 f1 线程：窗口已处理完，可以结束等待
              done_event.set()       # 通知“双线程”总控：我结束了

    - 参数说明：
        f1, f2：传入上面那种签名的函数
        f1_args、f1_kwargs：给 f1 传递的位置参数和关键字参数
        f2_args、f2_kwargs：给 f2 传递的位置参数和关键字参数
        timeout_sec：最多等待多少秒，如果超过则报告超时并返回 False

    - 返回值：
        True  ：两个子线程在 time_limit 内都完成了
        False ：超时，至少一个线程没及时调用 done_event.set()
    """
    if f1_kwargs is None:
        f1_kwargs = {}
    if f2_kwargs is None:
        f2_kwargs = {}

    # ① 为线程 1、2 准备同一个事件对
    timeout_event = threading.Event()
    done_event    = threading.Event()

    # ② 启动“线程1”：负责弹出、阻塞窗口
    t1 = threading.Thread(
        target=f1,
        args=(timeout_event, done_event, *f1_args),
        kwargs=f1_kwargs,
        daemon=True
    )
    # ③ 启动“线程2”：负责侦测窗口并点击、关闭
    t2 = threading.Thread(
        target=f2,
        args=(timeout_event, done_event, *f2_args),
        kwargs=f2_kwargs,
        daemon=True
    )

    start_time = time.time()
    t1.start()
    t2.start()

    # ④ 等待线程1 在 timeout_sec 秒内结束
    t1.join(timeout=timeout_sec)
    
    t2.join(timeout=timeout_sec)
    
        
        

    # ⑥ 检查 done_event 是否已被 set()
    if not done_event.is_set():
        # 超时：由调度器负责给线程1 发送退出信号
        timeout_event.set()
        print(f"⚠ 双线程执行超过 {timeout_sec}s —— 已触发 timeout_event")
        return False
    else:
        print("✅ 双线程任务在时限内完成")
        return True



#&&% CAD取消选择操作



def cancel_cad_selection(attempts: int = 3, delay: float = 0.5) -> bool:

    for i in range(1, attempts + 1):
        try:
            highlight_entities_in_window(0, 0, 0, 0)
            print(f"✅ 第{i}次尝试：cancel_cad_selection 成功")
            return True
        except Exception as e:
            print(f"⚠ 第{i}次尝试失败：{e}")
            if i < attempts:
                time.sleep(delay)
    print("❌ 已重试多次，仍未能执行 cancel_cad_selection")
    return False






#&&&&%% 第七部分  测试辅助




#   CAD基本操作-测试辅助
#####&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&




#最小化窗口
def min_w():
    import ctypes
    
    VK_MENU = 0x12  # Alt键
    VK_TAB = 0x09   # Tab键
    VK_LWIN = 0x5B  # 左Win键
    VK_M = 0x4D     # M键
    KEYEVENTF_KEYUP = 0x2

    # 模拟Alt + Tab
    ctypes.windll.user32.keybd_event(VK_MENU, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_TAB, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_TAB, 0, KEYEVENTF_KEYUP, 0)
    ctypes.windll.user32.keybd_event(VK_MENU, 0, KEYEVENTF_KEYUP, 0)

    # 模拟Win + M
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_M, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_M, 0, KEYEVENTF_KEYUP, 0)
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)




def ql():#清除测试辅助图层上的对象

    ensure_layer("测试辅助")

    




    
def srhd(*args):#根据输入坐标在模型空间画点
    """
    在模型空间绘制点并标注序号，支持以下调用形式：
    - srhd((x1,y1,z1), (x2,y2,z2))   # 多个点元组
    - srhd([(x1,y1,z1), (x2,y2,z2)]) # 列表形式
    - srhd((x1,y1,z1))               # 单点
    """
    doc = acad.ActiveDocument
    ms = doc.ModelSpace
    layer_name = "测试辅助"

    # 创建图层（如果不存在）
    try:
        doc.Layers.Item(layer_name)
    except:
        doc.Layers.Add(layer_name)
        print(f"✅ 已创建图层：{layer_name}")

    # 统一格式处理：允许单点、多个点、列表传入
    if len(args) == 1:
        if isinstance(args[0], (list, tuple)):
            if len(args[0]) == 3 and all(isinstance(i, (int, float)) for i in args[0]):
                points = [args[0]]
            else:
                points = args[0]
        else:
            print("❌ 输入格式不正确")
            return
    else:
        points = args

    # 绘制点与编号
    for idx, P in enumerate(points, 1):
        try:
            pt = vtpnt(*P)
            point = ms.AddPoint(pt)
            point.Layer = layer_name

            text = ms.AddText(str(idx), pt, 10)
            text.Layer = layer_name
        except Exception as e:
            print(f"❌ 添加点失败: {e}")

    return "✅ 点与编号已绘制"



def srhd_p(*args):#根据输入坐标在图纸空间画点
    """
    在图纸空间绘制点和编号，支持：
    - 多个坐标元组：srhd_p((x1,y1,z1), (x2,y2,z2))
    - 单个列表：srhd_p([(x1,y1,z1), (x2,y2,z2)])
    - 单个点：srhd_p((x1,y1,z1))
    """
    ps = doc.PaperSpace
    layer_name = "测试辅助"

    # 创建图层（如果不存在）
    try:
        doc.Layers.Item(layer_name)
    except:
        doc.Layers.Add(layer_name)
        print(f"✅ 已创建图层：{layer_name}")

    # 判断参数结构
    if len(args) == 1:
        if isinstance(args[0], (list, tuple)):
            if len(args[0]) == 3 and all(isinstance(i, (int, float)) for i in args[0]):
                # 单个点元组
                points = [args[0]]
            else:
                # 列表形式
                points = args[0]
        else:
            print("❌ 输入格式不正确")
            return
    else:
        # 多个点元组
        points = args

    for idx, P in enumerate(points, 1):
        try:
            pt = vtpnt(*P)
            point = ps.AddPoint(pt)
            point.Layer = layer_name

            text = ps.AddText(str(idx), pt, 10)
            text.Layer = layer_name
        except Exception as e:
            print(f"❌ 添加点失败: {e}")

    return "✅ 图纸空间中的点与编号已绘制"


def comtomath(LBcom):#将com点列表转为数学点列表

    LB_point=[]

    for i in range(0,len(LBcom)):

        point = LBcom[i].Coordinates

        LB_point.append(point)

    return LB_point        

        

#&&% 隔远查看

def fuzhi_chakan(LBcom,K=1):#K为放大倍数

    LK=[]

    for xx in LBcom:

        try:

            copy_obj = xx.Copy()
    
            point1 = vtpnt(0,0,0)
    
            point2 = vtpnt(0,K*1000000,0)
    
            copy_obj.Move(point1,point2)


            LK.append(copy_obj)

        except Exception as e:

            print(f"❌ 移动对象{xx.Handle}失败: {e}")               

    return LK



#测量已有文字长度

def celiang_wenzichangdu(TEXTCOM):

    text_copy = TEXTCOM.Copy()

    text_copy.Alignment = 2

    text_copy.TextAlignmentPoint =vtpnt(0,0,0)

    chang = abs(text_copy.InsertionPoint[0])

    text_copy.Delete()

    return chang

#测量新写文字长度

def celiang_wenzichangdu_write(ZF,style="图签",height=270,scalefactor=0.8):

    #根据字符串按样式字高宽度因子写入cad后的测量长度

    text_obj = acad.ActiveDocument.ModelSpace.AddText(ZF, vtpnt(0,0,0), height)

    text_obj.StyleName = style

    text_obj.ScaleFactor =scalefactor #宽度因子

    chang = celiang_wenzichangdu(text_obj)

    text_obj.Delete()

    return chang




##清空文件夹
def qingkong_wenjianjia(FolderPath):

     #清空文件夹B
    folder_path_1 = FolderPath 

    # 遍历文件夹中的所有文件
    for filename in os.listdir(folder_path_1):
        
        file_pathx = os.path.join(folder_path_1, filename)
        
        # 确保它是一个文件而不是文件夹
        if os.path.isfile(file_pathx):
            
            os.remove(file_pathx)  # 删除文件

        print(f"{FolderPath}文件夹已清空")


#&&% 返回对象外包盒的长，宽，横竖向，角点信息

def get_bbox_info(com_obj):
    """
    获取传入 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）的外包盒信息，
    并计算其长（length）、宽（width）以及横向或竖向（orientation）状态。

    参数：
      com_obj -- 任意支持 GetBoundingBox() 方法的 AutoCAD COM 对象

    返回：
      (length, width, orientation)
        length      -- 外包盒较长一边的长度（在 X/Y 平面上）
        width       -- 外包盒较短一边的长度（在 X/Y 平面上）
        orientation -- 字符串：
                         "horizontal" 表示 X 方向跨度 ≥ Y 方向跨度，
                         "vertical"   表示 Y 方向跨度 >  X 方向跨度
    如果对象不支持 GetBoundingBox，会抛出异常；也可根据需要自行捕获处理。
    """
 


   # 调用 GetBoundingBox 方法，返回两个点：minPt、maxPt
    # minPt、maxPt 都是 3 元素的 tuple 或 list，形如 (x, y, z)
    




    try:

        minPt, maxPt = com_obj.GetBoundingBox()

    except Exception as e:

        print(f"获取外包盒失败: {e}")               

        return None


    # 计算 X、Y 方向跨度
    dx = maxPt[0] - minPt[0]
    dy = maxPt[1] - minPt[1]

    # 将较大值定义为 length，较小值定义为 width
    length = max(dx, dy)
    width  = min(dx, dy)

    # 判断横向（X 跨度 ≥ Y 跨度）还是竖向（Y 跨度 > X 跨度）
    if dx >= dy:
        orientation = "horizontal"
    else:
        orientation = "vertical"

    return minPt, maxPt,length, width, orientation

#&&% 判断对象外包盒的横竖

def bbox_orientation_flag(com_obj):
    """
    判断任意 COM 对象的外包盒是竖向、横向还是正方形：
      - 如果 Y 方向跨度 > X 方向跨度，返回 1（竖向）
      - 否则（包括 X 方向跨度 >= Y 方向跨度），返回 0
        —— 即当外包盒为正方形（X 跨度 == Y 跨度）时，也返回 0

    参数：
      com_obj -- 支持 GetBoundingBox() 的 AutoCAD COM 对象

    返回：
      int -- 竖向返回 1，横向或正方形返回 0
    """
    # 获取外包盒的最小点和最大点
    min_pt, max_pt = com_obj.GetBoundingBox()
    # 计算 X、Y 方向跨度
    dx = abs(max_pt[0] - min_pt[0])
    dy = abs(max_pt[1] - min_pt[1])
    # 如果是竖向（dy > dx），返回 1；否则（横向或正方形）返回 0
    return 1 if dy > dx else 0

#&&% 获取多个对象的外包盒数据
def group_bbox_corners(com_objs):
    """
    计算一组 COM 对象的整体外包盒，并按顺序返回四个角点坐标：
      1. 左下角 (minX, minY)
      2. 右上角 (maxX, maxY)
      3. 左上角 (minX, maxY)
      4. 右下角 (maxX, minY)

    参数：
      com_objs -- 可迭代的一组支持 GetBoundingBox() 方法的 COM 对象列表

    返回：
      四元组：(
        (minX, minY, z),
        (maxX, maxY, z),
        (minX, maxY, z),
        (maxX, minY, z)
      )
      其中 z 取自各对象外包盒的 z 值范围，统一使用最小 z（如需不同，可按需调整）
    """
    # 初始化为极端值
    global_min_x = float('inf')
    global_min_y = float('inf')
    global_max_x = float('-inf')
    global_max_y = float('-inf')
    global_min_z = float('inf')  # 如果需要统一 z 值，可使用最小 z
    # 如果不关心 z，只返回 0 即可。这里以最小 z 作为统一 z
    for obj in com_objs:
        try:
            min_pt, max_pt = obj.GetBoundingBox()
        except Exception:
            # 如果对象不支持 GetBoundingBox，跳过
            continue

        x1, y1, z1 = min_pt
        x2, y2, z2 = max_pt

        # 更新 X/Y extremes
        if x1 < global_min_x:
            global_min_x = x1
        if y1 < global_min_y:
            global_min_y = y1
        if x2 > global_max_x:
            global_max_x = x2
        if y2 > global_max_y:
            global_max_y = y2

        # 更新 Z extremes（如果需要统一使用最小 z）
        if z1 < global_min_z:
            global_min_z = z1

    # 如果所有对象都被跳过（列表为空或都不支持 GetBoundingBox），直接返回 None
    if global_min_x == float('inf'):
        return None

    # 采用 global_min_z 作为所有角点的 z 分量
    z = global_min_z

    # 左下、右上、左上、右下
    bottom_left  = (global_min_x, global_min_y, z)
    top_right    = (global_max_x, global_max_y, z)
    top_left     = (global_min_x, global_max_y, z)
    bottom_right = (global_max_x, global_min_y, z)

    return bottom_left, top_right, top_left, bottom_right

#&&% zip的用法
"""
seq_x0  = (x0, y0, z0)
P_start = (dx, dy, dz)
要得到新坐标 (x0+dx, y0+dy, z0+dz)，就可以写：
seq_x0 = tuple(a + b for a, b in zip(seq_x0, P_start))
zip(seq_x0, P_start) 会产出 (x0, dx), (y0, dy), (z0, dz)
a + b for a, b in ... 就对每一对做相加
最后用 tuple(...) 把结果收成三元组
对两个列表求和
xs = [10, 20, 30]
ys = [1, 2, 3]
sums = [x + y for x, y in zip(xs, ys)]
# sums == [11, 22, 33]
并行遍历三组数据
names = ["A", "B", "C"]
ages  = [30, 25, 40]
scores= [85, 92, 78]
for n, a, s in zip(names, ages, scores):
    print(f"{n} 年龄{a} 分数{s}")
解压
如果有一个列表 pairs = [(1,4),(2,5),(3,6)]，要拆成两个列表：
a, b = zip(*pairs)
# a == (1,2,3), b == (4,5,6)

"""

#&&% 从两点绘制矩形

def draw_rectangle_by_corners(p1: tuple[float, float, float],
                              p2: tuple[float, float, float],
                              layer_name: str = "测试辅助",
                              width: float = 0.0,
                              color: int = 256) -> object:
    """
    基于两点绘制一个闭合矩形：
      - p1: (x_min, y_min, z)
      - p2: (x_max, y_max, z)
    调用 draw_lwpolyline 绘制四边闭合多段线，返回新创建的多段线对象。
    """
    # 左下、右下、右上、左上
    x1, y1, z1 = p1
    x2, y2, z2 = p2
    coords = [
        (x1, y1, z1),
        (x2, y1, z1),
        (x2, y2, z1),
        (x1, y2, z1)
    ]
    # 绘制闭合矩形
    rect = draw_lwpolyline(
        coords3d=coords,
        layer_name=layer_name,
        width=width,
        color=color,
        closed=True
    )
    return rect

#&&% 获取外包盒中心
def bbox_center_2(e):
    # GetBoundingBox 返回两个 Point (minPt, maxPt)
    min_pt, max_pt = e.GetBoundingBox()
    x1, y1, _ = tuple(min_pt)
    x2, y2, _ = tuple(max_pt)
    return ((x1 + x2) / 2, (y1 + y2) / 2)

def bbox_center_3(ent):
    """返回实体外包盒中心 (cx, cy, cz)"""
    mn, mx = ent.GetBoundingBox()
    return ((mn[0] + mx[0]) / 2.0,
            (mn[1] + mx[1]) / 2.0,
            (mn[2] + mx[2]) / 2.0)


#&&&&%% 第八部分   CAD文件操作


#_________________________________________________________________________________________________________________________

#  模块使用说明

"""
该模块解决CAD文件的转换、打开、关闭、文件之间的复制、分解组合等问题 

"""

#  主函数
#  (1)
# 文件转成t7、t3格式

#  该函数系列包括如下一些函数
"""
zhuancheng_t7()
zhuancheng_t3()
在li()连接激活文件后，直接执行该命令即可转换
一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
"""    
#&&% 对获得的两行大地坐标数据建立字典

"""
大量数据的整理读取，AI不能绝对保证正确

"""

def build_J_points_from_selected_texts(LB, n_points=61, prefix_x="30", prefix_y="37"):
    num_re = re.compile(r"-?\d+(?:\.\d+)?")

    def first_number(s):
        m = num_re.search(str(s))
        if not m:
            raise ValueError(f"无法解析数字: {s!r}")
        return float(m.group(0))

    items = []
    for ent in LB:
        txt, ip = get_text_and_ip(ent)
        if not txt or not ip:
            continue
        txt = str(txt).strip().replace(" ", "")
        x0, y0, *_ = ip
        if txt.startswith(prefix_x):
            items.append(("x", first_number(txt), float(y0), float(x0)))
        elif txt.startswith(prefix_y):
            items.append(("y", first_number(txt), float(y0), float(x0)))

    # 从上到下（y降序），同行按x升序稳定
    items.sort(key=lambda t: (-t[2], t[3]))
    xs = [v for typ, v, _, _ in items if typ == "x"]
    ys = [v for typ, v, _, _ in items if typ == "y"]

    if len(xs) != n_points or len(ys) != n_points:
        raise RuntimeError(f"X={len(xs)} / Y={len(ys)} 与期望 {n_points} 不符。")

    pts_dict = {f"J{i+1}": (xs[i], ys[i]) for i in range(n_points)}
    pts_list = [(f"J{i+1}", xs[i], ys[i]) for i in range(n_points)]
    return pts_dict, pts_list


#从界点大地坐标计算经纬度
def convert_pts_dict_to_latlon(pts_dict, central_lon=111):
    """
    输入: pts_dict = {'J1': (N, E), 'J2': (N, E), ...}
          N=北坐标, E=东坐标 (CGCS2000 高斯-克吕格投影)
    输出: geo_dict = {'J1': (lon, lat), 'J2': (lon, lat), ...}  (经度, 纬度, 单位:度)
    """
    a = 6378137.0
    f = 1 / 298.257222101
    FE = 500000
    FN = 0
    e2 = 2*f - f*f
    e_prime_2 = e2 / (1.0 - e2)

    def one_point(N, E):
        m = N - FN
        mu = m / (a * (1 - e2/4 - 3*e2**2/64 - 5*e2**3/256))
        e1 = (1 - math.sqrt(1 - e2)) / (1 + math.sqrt(1 - e2))
        Bf = (mu
              + (3*e1/2 - 27*e1**3/32) * math.sin(2*mu)
              + (21*e1**2/16 - 55*e1**4/32) * math.sin(4*mu)
              + (151*e1**3/96) * math.sin(6*mu)
              + (1097*e1**4/512) * math.sin(8*mu))
        x_prime = E - FE
        cosBf = math.cos(Bf)
        tanBf = math.tan(Bf)
        eta2 = e_prime_2 * (cosBf**2)
        Nf = a / math.sqrt(1 - e2 * (math.sin(Bf)**2))
        Mf = a * (1 - e2) / (1 - e2 * (math.sin(Bf)**2))**1.5
        lat = (Bf
               - tanBf/(2* Mf* Nf) * (x_prime**2)
               + tanBf/(24* Mf* Nf**3) * (5 + 3*tanBf**2 + eta2 - 9*tanBf**2 * eta2) * (x_prime**4)
               - tanBf/(720* Mf* Nf**5) * (61 + 90*tanBf**2 + 45*tanBf**4) * (x_prime**6))
        lon = (math.radians(central_lon)
               + x_prime/(Nf* cosBf)
               - (1 + 2*tanBf**2 + eta2) * (x_prime**3)/(6* Nf**3* cosBf)
               + (5 + 28*tanBf**2 + 24*tanBf**4 + 6* eta2 + 8*tanBf**2 * eta2) * (x_prime**5)/(120* Nf**5* cosBf))
        return math.degrees(lon), math.degrees(lat)

    geo_dict = {}
    for k, (N, E) in pts_dict.items():
        geo_dict[k] = one_point(N, E)

    return geo_dict

# 文件基本操作



"""
使用共同的文件全局变量acad,mp,doc,sp是我们编制脚本控制不同文件的基础            


"""


    

def rename_time(output_path):#不覆盖从当前时间给带路径文件名命名

    from datetime import datetime
    now = datetime.now()
    time_str = now.strftime("%m-%d-%H-%M")

    dir_name, file_name = os.path.split(output_path)
    name, ext = os.path.splitext(file_name)
    new_file_name = f"{name}_{time_str}{ext}"
    new_path = os.path.join(dir_name, new_file_name)
    return new_path

#供参考优化
def open_dwg(path: str, visible: bool = True):
    # 初始化 COM（尤其在多线程或非脚本交互环境下推荐调用）
    pythoncom.CoInitialize()

    # 启动并连接到 AutoCAD 应用
    # Dispatch 会在已有实例上复用，DispatchEx 会强制新开一个实例
    acad = Dispatch("AutoCAD.Application")  # 或者 DispatchEx("AutoCAD.Application")

    # 可见性（True 时会弹出界面）
    acad.Visible = visible

    # 打开 DWG 文档
    doc = acad.Documents.Open(path)

    print(f"已打开：{doc.Name}")

    return acad, doc









##打开文件
def Open_By_Omission_wenjian(file_path):

    """
    不能指望通过这函数控制所有打开文件时遇到的窗口跳出，报错等，它们应该在别的地方解决，例如字体问题，报错窗口，代理错误等等

    """
    
    t1 = time.time()
    
    max_retries = 3

    for attempt in range(max_retries):
        
        try:
            
            # 尝试打开文件
            new_doc = acad.Documents.Open(file_path)
            
            print(f"Opened file: {file_path}")
            
            t2 = time.time()
            
            print("文件打开耗时:", t2 - t1, "秒")

            li() #所有的函数都使用同样的全局变量acad,mp,doc等，这样在不同的文件上函数仍然通用
            
            return new_doc
        
        except Exception as e:
            
            print(f"Attempt {attempt + 1} failed: {e}")
            pass
            
            time.sleep(2)  # 等待2秒后再次尝试

    # 如果所有尝试都失败
    t2 = time.time()
    
    print("文件打开耗时:", t2 - t1, "秒")
    
    return  None



##打开文件时跳过缺字体窗口


def f1_openfile_getwindaow(timeout_event, done_event,
                     file_path):
    pythoncom.CoInitialize()
    try:

        

        print("线程1启动")

        li()

        time.sleep(2)

        Open_By_Omission_wenjian(file_path)

        time.sleep(10)

        timeout_event.wait()          # 等待线程2完工 / 或调度器超时

    except Exception as e:
        print("f1_openfile_getwindaow:", e, traceback.format_exc())
    finally:
        pythoncom.CoUninitialize()
        done_event.set()
        


def f2_delwindaow(timeout_event, done_event):
    pythoncom.CoInitialize()
    try:
        print("线程2启动")
        # 闪动窗口（示例：刷新一下窗口列表）
       
        time.sleep(1)

        BT = list_open_window_titles()
        print("当前窗口标题：", BT)

        if '缺少 SHX 文件' in BT:
            # 点击对话框中的“忽略”按钮，假设它在对话框中的大致坐标 (148, 220)
            click_in_window("缺少 SHX 文件", 148, 220, click_titlebar=True)
            print("🖱 已点击“忽略”按钮")
        else:
            print("ℹ️ 未检测到“缺少 SHX 文件”对话框")

    except Exception as e:
        print("f2_delwindaow 异常:", e, traceback.format_exc())
    finally:
        # 先发信号给线程1 退出阻塞，再通知总控已经完成
        timeout_event.set()
        pythoncom.CoUninitialize()
        done_event.set()




def Open_file_nic(file_path ):

    """
    实测似乎人工打开文件时会跳出SHX字体缺失窗口，而Open_By_Omission_wenjian命令不会，也许这个命令是多余的，但仍然保留这个机制，因为我们不能断定
    
    """

    ok = run_dual_threads(

        f1=f1_openfile_getwindaow,

        f2=f2_delwindaow,

        f1_args=(file_path,),

        f2_args=(),

        timeout_sec=600
    )
    if ok:
        print(f"🎉 成功打开文件 → {file_path}")
    else:
        print("🚨 打开文件超时 / 失败")





##保存文件#另存就是doc.SaveAs()

def savefile():

    doc.Save()


##关闭文件(别乱删)

def guanbifile():

    doc.Close()



##确保关闭当前文件


def close_current_drawing_safely():
    """
    安全关闭当前 DWG 文件，确保确实关闭并重新连接。
    最多尝试 3 次。
    """
    
    try:
        Name1 = doc.Name
    except:
        print("⚠️ 当前 doc 无法获取名称，可能未连接。")

        li()

        Name1 = doc.Name
    
    for attempt in range(1, 4):  # 最多尝试3次
        print(f"🔄 第 {attempt} 次尝试关闭 '{Name1}'")
        close_dwg_by_name(Name1)

        li()  # 重新连接 acad、doc、mp、sp 等
        try:
            Name2 = doc.Name
        except:
            Name2 = None

        if Name2 != Name1:
            print(f"🟢 已确认文件 '{Name1}' 关闭，当前打开文件为 '{Name2}'")
            li()  # 再执行一次，确保变量正确
            return
        else:
            print("⚠️ 文件仍未关闭，继续尝试...")

    print(f"❌ 多次尝试仍未成功关闭 '{Name1}'，请手动检查。")



#&&% 2 两个文件A,B的操作
"""
文件作为块插入不稳定，甚至在交互模式可以同样的代码转为函数模式却不行，反倒是使用全选复制粘贴比较靠谱

除了使用这个命令合并文件，还有转换为纯数据命令

        
    """

#新建一个空白文件(不打开)
def create_new_dwg_file(name_with_path):
    """
    创建一个新的 DWG 文件，并保存为指定完整路径 name_with_path（应以 .dwg 结尾）。

    参数：
        name_with_path - 完整的文件保存路径（例如 D:\CADXT\Export\新图01.dwg）
    """
   
    # 检查并创建目录（如果不存在）
    folder = os.path.dirname(name_with_path)
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"📁 已创建文件夹：{folder}")

    # 新建 DWG 文件
    acad.Documents.Add()
    time.sleep(1)

    # 连接当前文档
    new_doc = acad.ActiveDocument

    # 保存到指定路径
    new_doc.SaveAs(name_with_path)

    new_doc.Close()
    li()
    print(f"✅ 新建并保存 DWG 文件：{name_with_path}")

#新建一个空白文件
def create_new_dwg_file_no(name_with_path):
    """
    创建一个新的 DWG 文件，并保存为指定完整路径 name_with_path（应以 .dwg 结尾）。

    参数：
        name_with_path - 完整的文件保存路径（例如 D:\CADXT\Export\新图01.dwg）
    """
   
    # 检查并创建目录（如果不存在）
    folder = os.path.dirname(name_with_path)
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"📁 已创建文件夹：{folder}")

    # 新建 DWG 文件
    acad.Documents.Add()
    time.sleep(1)

    # 连接当前文档
    new_doc = acad.ActiveDocument

    # 保存到指定路径
    new_doc.SaveAs(name_with_path)
  
    li()
    print(f"✅ 新建并保存 DWG 文件：{name_with_path}")



#  主函数
#  (1)
# 关闭文件

#  该函数系列包括如下一些函数

def close_all_except_active_safe():#关闭除当前激活文档外的所有 DWG 文件
    """
    更稳定地关闭除当前激活文档外的所有 DWG 文件，避免 COM 对象断链。
    """
    try:
        active_name = acad.ActiveDocument.Name
        all_names = [acad.Documents.Item(i).Name for i in range(acad.Documents.Count)]
        closed = 0

        for name in all_names:
            if name != active_name:
                try:
                    doc = acad.Documents.Item(name)
                    doc.Close(False)  # 不保存直接关闭
                    print(f"🗂️ 已关闭：{name}")
                    closed += 1
                except Exception as e:
                    print(f"⚠️ 无法关闭 {name}：{e}")

        print(f"✅ 成功关闭 {closed} 个文档，仅保留 {active_name}")

    except Exception as e:
        print(f"❌ 安全关闭失败：{e}")


## 测试示例
##close_all_except_active_safe()
##🗂️ 已关闭：Drawing3.dwg
##🗂️ 已关闭：Drawing4.dwg
##🗂️ 已关闭：空白.dwg
##🗂️ 已关闭：测试1.dwg
##✅ 成功关闭 4 个文档，仅保留 cs.dwg
##__________
    

# 双文件操作
"""
刚开始启动天正，默认打开了一个drawing1的文件，使用create_new_dwg_file新建一个文件后，新建的文件会关闭，仍然是drawing1在，仍需li()连接一下

此时使用Open_By_Omission_wenjian打开新文件，就会同时自动关闭drawing1。同样需要li()连接一下
因此我们面临两种可能，一个是天正的drawing1作为基本连接状态，此时文件数为1，但要注意create_new_dwg_file新建 不改变这种状态，而 Open_By_Omission_wenjian打开桌面就会变成只有1个文件

另一种情况时我们用自己系统的“空白.dwg”作为当前桌面的唯一基础文件。
事实上，在这个桌面上只有一个文件的情况下，我们需要有两个文件的状态。继续打开一个新文件即可。但此时，就需要在打开新文件之前，运行doc1=acad.ActiveDocument，用来标记当前文件。再打开新文件
li()新文件，运行doc2=acad.ActiveDocument，记录当前文件。此时，连接的是doc2,但我们可以执行doc1.Close()关闭前面的文件，就恢复到了单一文件状态。
当然，通过dir(doc1)可以查阅文件的属性和方法，不仅可以查看文件的名字，还可以获取这个文件的其它属性。例如doc1.Groups，PlotConfigurations,我们是否可以在两个文件都在的情况下拷贝一个到另一个来。


#当前激活桌面文件的激活布局
acad.ActiveDocument.SetVariable("TILEMODE",0)

layouts = doc.Layouts

existing_names = [layout.Name for layout in layouts]

doc.ActiveLayout = layouts.Item("布局1")
    
layout = source_doc.ActiveLayout

#获取当前激活文件的打印配置

zg_pdf_config = source_doc.PlotConfigurations.Item("ZG_PDF")

# 使用 CopyFrom 方法将 "ZG_PDF" 的设置复制到指定布局

layout.CopyFrom(zg_pdf_config)

也就是说，两个文件同时打开的操作是必不可少的。虽然具体运行时我们尽量保持一个文件。但两个文件需要交换信息。

"""

def copy_doc_to_current(file_path):#将目标文件原位复制粘贴到当前文件

    """
    若操作不当，竟至不能复制粘贴
    将目标文件粘贴进来，当然是最自然的需求
    不要用“空白.dwg”作测试，也不要用天正开启时默认生成的文件，因为它们都会破坏整个流程
    很遗憾，块插入命令不能生效

    """   

    li()
    doc.Save()
    file_path1=file_path
    file_path2=doc.FullName#获取当前文件带路径的文件名
    
    
    name2=doc.Name
    doc2=get_doc_by_name(name2)   
    
    #打开辅助文件
    
    Open_By_Omission_wenjian("D:/Myprogramsystem/cad/xitongjicuwenjian/空白.dwg")

    li()

    name3=doc.Name

    doc3=get_doc_by_name(name3)     

    doc2.Close()

    insert_dwg_2(file_path2,file_path1, insert_point=(0, 0, 0))

    time.sleep(2)

    Open_By_Omission_wenjian(file_path2)

    li()

    doc.SendCommand("z\ne\n")

    doc3.Close()

def insert_dwg_2(new_file_path_A,new_file_path_B, insert_point=(0, 0, 0)):#文件B合成到文件A中

    """
    如果刚打开天正 ，会默认产生一个dwg1文件，也会激活，用来进行后续操作。可当我们操作到后面关闭新建的文件A时，注意新打开A新建A时这个dwg1也关闭了。?

    """

    

    li()

    Open_By_Omission_wenjian(new_file_path_B) #打开文件B       

    li()

    acad.ActiveDocument.SendCommand('_ai_selall'+chr(13))

    time.sleep(4)

    acad.ActiveDocument.SendCommand('_copybase'+chr(13)+'0,0,0'+chr(13)+chr(13))#全选复制

    time.sleep(4)

    guanbifile()
    
    li()

    Open_By_Omission_wenjian(new_file_path_A)#打开文件A

    li()        
    
    acad.ActiveDocument.SendCommand('TPasteClip'+chr(13)+'0,0,0'+chr(13)+chr(13))#粘贴

    time.sleep(4)

    doc.save()

    time.sleep(4)
   
    doc.close()
    
    li()



def copy_group_S1_from_doc1_to_doc2(doc1, doc2, group_name="S1"):#将名为“S1”的组复制到当前桌面上另一个文档
    """
    将 doc1 中名为 group_name 的组复制到 doc2 中，并重新组装组。
    粘贴点为 0,0,0。粘贴后通过 handle 差集识别新对象。

    mp要不断更新

    def refresh_modelspace(doc):
        return doc.ModelSpace
    逻辑更清晰

    直接发送命令复制粘贴或许更好
    """
    try:
        # 1. 激活源文档
        set_active_doc(doc1)
        li()

        group = doc.Groups.Item(group_name)
        handles = [ent.Handle for ent in group]
        objs = [ent for ent in mp if ent.Handle in handles]

        yin_to_xian_xuanze(objs)

        doc.SendCommand("_copybase\n0,0,0\n")
        time.sleep(0.5)
        doc.SendCommand("_copyclip\n\n")
        time.sleep(1)

        # 2. 激活目标文档
        set_active_doc(doc2)
        li()
        ms2 = doc.ModelSpace

        # 3. 粘贴前记录已有对象
        pre_map = get_handle_object_map(ms2)

        # 4. 粘贴
        doc.SendCommand("_pasteclip"+chr(13)+"0,0,0"+chr(13)+chr(13))#
        time.sleep(1.5)

        # 5. 粘贴后重新记录对象
        ms2 = doc.ModelSpace
        
        post_map = get_handle_object_map(ms2)

        # 6. 取出新对象（通过 handle 差集）
        new_handles = set(post_map) - set(pre_map)
        new_objs = [post_map[h] for h in new_handles]

        print(f"✅ 粘贴完成，识别出 {len(new_objs)} 个新图元")

        # 7. 添加这些对象到组中（使用你提供的方法）
        add_objects_to_group(group_name, new_objs)

        print(f"✅ 成功将粘贴对象加入组 '{group_name}'")

    except Exception as e:
        print(f"❌ 复制组失败: {e}")

        



#&&% 合并两个不重叠的文件    def insert_dwg_pyautocad(new_file_path_A,new_file_path_B, insert_point=(0, 0, 0)):#文件B合成到新建文件A中

    """
    如果刚打开天正 ，会默认产生一个dwg1文件，也会激活，用来进行后续操作。可当我们操作到后面关闭新建的文件A时，注意新打开A新建A时这个dwg1也关闭了。?

    """

    create_new_dwg_file(new_file_path_A)#新建文件命令已经默认了关闭新建的这个文件的操作

    li()

    Open_By_Omission_wenjian(new_file_path_B) #打开文件B       

    li()

    acad.ActiveDocument.SendCommand('_ai_selall'+chr(13))

    time.sleep(4)

    acad.ActiveDocument.SendCommand('_copybase'+chr(13)+'0,0,0'+chr(13)+chr(13))#全选复制

    time.sleep(4)

    guanbifile()
    
    li()

    Open_By_Omission_wenjian(new_file_path_A)#打开文件A

    li()        
    
    acad.ActiveDocument.SendCommand('TPasteClip'+chr(13)+'0,0,0'+chr(13)+chr(13))#粘贴

    time.sleep(4)

    doc.save()

    time.sleep(4)
   
    doc.close()
    
    li()

#批量合成文件def insert_multiple_dwgs_to_new_file(new_file_path_A, source_files_list):
    """
    创建一个新 DWG 文件 A，并将多个已有 DWG 文件（B1, B2, ...）的内容
    原位复制粘贴到 A 中，最终保存并关闭 A，返回初始 dwg 状态。

    参数：
        new_file_path_A      - 要创建的新 DWG 文件路径（完整路径）
        source_files_list    - 要复制进来的多个 DWG 文件路径列表
    """
    # Step 1: 创建新文件 A（新建后自动关闭）
    create_new_dwg_file(new_file_path_A)
    time.sleep(1)

    # Step 2: 逐个处理每个源文件 B
    for idx, fileB in enumerate(source_files_list):
        print(f"📥 处理第 {idx+1} 个源文件：{fileB}")
        Open_By_Omission_wenjian(fileB)
        li()
        time.sleep(1)

        acad.ActiveDocument.SendCommand('_ai_selall' + chr(13))
        time.sleep(2)

        acad.ActiveDocument.SendCommand('_copybase' + chr(13) + '0,0,0' + chr(13) + chr(13))
        time.sleep(3)

        guanbifile()
        time.sleep(1)

        Open_By_Omission_wenjian(new_file_path_A)
        li()
        time.sleep(1)

        acad.ActiveDocument.SendCommand('TPasteClip' + chr(13) + '0,0,0' + chr(13) + chr(13))
        time.sleep(3)

        acad.ActiveDocument.Save()
        time.sleep(1)

        acad.ActiveDocument.Close(False)
        time.sleep(1)

    # Step 3: 操作结束后回到 dwg1 初始状态
    li()
    print(f"✅ 已成功将 {len(source_files_list)} 个 DWG 合并至文件：{new_file_path_A}")





    

def xianshi_yincangtuxing():#  显示文件中可能隐藏的对象
        
    acad.ActiveDocument.SendCommand("HFKJ"+chr(13)+chr(13))#在V4状态下可能要改成"HFKJ"+chr(13)+"Y"+chr(13)

    doc.Save()



# 创建一个事件对象

timeout_event = threading.Event()

event = threading.Event()

def run_cad_program(timeout_event, event):
    pythoncom.CoInitialize()
    
    try:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        print(acad.ActiveDocument.Name)
        
        acad.ActiveDocument.SendCommand("TSaveAs"+chr(13))
        
        print("CAD命令已发送，等待窗口操作完成...")
        
        # 等待timeout_event信号，如果收到信号，则退出
        timeout_event.wait()

    except Exception as e:
        print(f"run_cad_program 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        event.set()  # 通知主线程完成

        
def automate_window_with_pywinauto_t7(timeout_event, event):
    pythoncom.CoInitialize()

    try:
        jincheng_acad = get_acad_process_id('acad.exe')
        handle = findwindows.find_windows(process=jincheng_acad)[0]

        app = Application().connect(handle=handle)
        time.sleep(2)

        try:
            window = app.window(title="图形导出", class_name="#32770")
        except Exception as e:
            print("尝试重新获取窗口:", e)
            time.sleep(2)
            window = app.window(title="图形导出", class_name="#32770")

        if window.exists():
            print("窗口存在")
            child_windows = window.children()

            for child_window in child_windows:
                try:
                    child_window.select("天正7文件 (*.dwg) ")
                    print("正在导出t7")
                    pass
                except Exception as e:
                    pass
        else:
            print("窗口不存在")
            pass

        save_button = window.child_window(title="保存(&S)", class_name="Button")

        try:
            save_button.set_focus()
            
            save_button.click()


        except Exception as e:
            print("pywinauto处理窗口出现问题:", e)
            time.sleep(2)
            save_button.set_focus()
            save_button.click()

        print("已使用pywinauto自动化窗口")
        pass

    except Exception as e:
        print(f"automate_window_with_pywinauto_t7 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # 通知主线程完成

#&&% # 天正转t7    
def zhuancheng_t7():
    
    """
    在li()连接激活文件后，直接执行该命令即可转换
    一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
    C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
    """    
    t1 = time.time()

    chaoshibiaoji = 1

    # 创建一个Event，用于通知子线程终止
    timeout_event = threading.Event()
    
    # 创建一个Event，用于通知主线程完成
    event = threading.Event()

    # 创建两个线程，分别执行run_cad_program和automate_window_with_pywinauto_t7
    thread1 = threading.Thread(target=run_cad_program, args=(timeout_event, event))
    thread2 = threading.Thread(target=automate_window_with_pywinauto_t7, args=(timeout_event, event))

    thread1.start()
    thread2.start()

    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("操作超时，正在中断...")

        pass

        
        # 设置timeout_event，通知run_cad_program()终止
        timeout_event.set()

    print("文件转T7格式操作结束")

    t2 = time.time()
    print("文件转T7格式操作总共用时", t2 - t1, "秒")


# 创建一个事件对象

timeout_event = threading.Event()

event = threading.Event()


def automate_window_with_pywinauto_t3(timeout_event, event):
    pythoncom.CoInitialize()

    try:
        jincheng_acad = get_acad_process_id('acad.exe')
        handle = findwindows.find_windows(process=jincheng_acad)[0]

        app = Application().connect(handle=handle)
        time.sleep(2)

        try:
            window = app.window(title="图形导出", class_name="#32770")
        except Exception as e:
            print("尝试重新获取窗口:", e)
            time.sleep(2)
            window = app.window(title="图形导出", class_name="#32770")

        if window.exists():
            print("窗口存在")
            pass
            child_windows = window.children()

            for child_window in child_windows:
                try:
                    child_window.select("天正3文件 (*.dwg) ")
                    print("正在导出t3")
                    pass
                except Exception as e:
                    pass
        else:
            print("窗口不存在")
            pass

        save_button = window.child_window(title="保存(&S)", class_name="Button")

        try:
            save_button.set_focus()
            save_button.click()
        except Exception as e:
            print("pywinauto处理窗口出现问题:", e)
            pass
            time.sleep(2)
            save_button.set_focus()
            save_button.click()

        print("已使用pywinauto自动化窗口")
        pass

    except Exception as e:
        print(f"automate_window_with_pywinauto_t3 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # 通知主线程完成
#&&% # 天正转t3
        
def zhuancheng_t3():

    """
    在li()连接激活文件后，直接执行该命令即可转换
    一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
    C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
    """    

    t1 = time.time()

    chaoshibiaoji = 1

    # 创建一个Event，用于通知子线程终止
    timeout_event = threading.Event()
    
    # 创建一个Event，用于通知主线程完成
    event = threading.Event()

    # 创建两个线程，分别执行run_cad_program和automate_window_with_pywinauto_t3
    thread1 = threading.Thread(target=run_cad_program, args=(timeout_event, event))
    thread2 = threading.Thread(target=automate_window_with_pywinauto_t3, args=(timeout_event, event))

    thread1.start()
    thread2.start()

    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("操作超时，正在中断...")

        pass

        
        # 设置timeout_event，通知run_cad_program()终止
        timeout_event.set()

    print("文件转T3格式操作结束")

    t2 = time.time()
    print("文件转T3格式操作总共用时", t2 - t1, "秒")


#  主函数
#  (2)
# 确定一个文件中所有的对象类型

#  该函数系列包括如下一些函数


def get_all_object_types():
    """
    获取当前文件中模型空间中所有图形对象的类型名称（如 AcDbText, AcDbLine 等）

    返回：
        类型名列表（去重，已排序）
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    modelspace = doc.ModelSpace

    types_set = set()

    for obj in modelspace:
        try:
            types_set.add(obj.ObjectName)
        except:
            continue

    types_list = sorted(types_set)
    print(f"✅ 共发现 {len(types_list)} 种对象类型：")
    for t in types_list:
        print(f" - {t}")

    return types_list



#  主函数
#  (3)
# 关闭文件名为Name的文件

#  该函数系列包括如下一些函数

def close_dwg_by_name(Name):
    """
    关闭当前桌面中名为 Name 的 DWG 文件。
    如果文件已打开，则关闭该文件。
    
    参数：
        Name: 要关闭的 DWG 文件的名称（如 "example.dwg"）不含路径的名
    """
    try:
        acad = win32com.client.Dispatch("AutoCAD.Application")  # 获取 AutoCAD 应用
        doc = acad.Documents.Item(Name)  # 获取指定名称的文档
        
        if doc:
            doc.Close(False)  # 关闭文件，不提示保存
            print(f"✅ 文件 '{Name}' 已关闭")
        else:
            print(f"❌ 未找到名为 '{Name}' 的文件")
    except Exception as e:
        print(f"❌ 关闭文件 '{Name}' 失败: {e}")




def set_active_doc(doc):# 设置文件doc为激活对象

    """
    doc为com对象
    将指定文档 doc 设置为当前激活窗口。
    将其中一个打开时，要使用li()，然后执行doc1=acad.ActiveDocument
    再打开另一个，使用li()连接，执行doc2=acad.ActiveDocument就实际上获取了两个文件com实体对象
    然后就可以使用这个函数随时激活其中一个文件，激活后仍然要使用li()连接才能正确运行
    """
    try:
        doc.Activate()
        time.sleep(0.3)  # 稍作延时，确保激活生效
        print("✅ 当前激活文档：", doc.Name)
        return True
    except Exception as e:
        print("❌ 激活文档失败：", e)
        return False


def get_doc_by_name(name): #从文件名获取文件对象
    """
    通过文件名获取 AutoCAD 文档对象，例如 '空白.dwg'
    如果未找到，返回 None
    """
    for doc in acad.Documents:
        if doc.Name.lower() == name.lower():
            return doc
    return None

def get_open_document_names():#返回所有打开的文件名
    return [doc.Name for doc in acad.Documents]



#&&% 当前CAD文件数

def dwgs_count():#当前桌面的dwg文件数量

    shu = acad.Documents.Count

    return shu



#  主函数
#  (3)
# 跨文件复制粘贴

#&&%  文件按打印区域分解成多个文件


def export_region_to_new_file(x1, y1, x2, y2, filename,ty=1):##将区域(x1,y1,x2,y2)内的对象原位剪切粘贴到一个新文件中。
    """
    将当前文件中指定窗口区域内的图形对象，剪切粘贴到一个新 DWG 文件中，并保存为 filename.dwg。
    最后关闭新建文件、重新连接当前文件。
    """


    # 聚焦区域
    shitu_region(x1,y1,x2,y2)


    # 步骤 1：选择区域对象并转为蓝色夹点选中


    
    li()
    LB = select_objects_in_window_area(x1, y1, x2, y2)
    if not LB:
        print("❌ 区域内未找到任何对象，操作中止")
        return

    yin_to_xian_xuanze(LB,ty=ty)  # 关键：转换为命令行状态选中

    # 步骤 2：复制选中对象
    
    
    original_doc = doc  # 记录原始文档句柄

    doc.SendCommand("_copybase" + chr(13) + "0,0,0" + chr(13) + chr(13))
    time.sleep(3)

    # 步骤 3：新建 DWG 文件并连接
    acad.Documents.Add()
    time.sleep(1)
    li()
    new_doc = acad.ActiveDocument

    # 步骤 4：粘贴对象
    new_doc.SendCommand("TPasteClip" + chr(13) + "0,0,0" + chr(13) + chr(13))
    time.sleep(3)

    #视点聚焦
    shitu_region(x1,y1,x2,y2)

    # 步骤 5：保存新文件
    original_path = original_doc.FullName
    folder = os.path.dirname(original_path)
    save_path = os.path.join(folder, filename + ".dwg")


    new_doc.SaveAs(save_path)
    print(f"✅ 对象已导出到：{save_path}")

    # ✅ 新增部分 ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓

    # 步骤 6：关闭新建文件
    new_doc.Close(False)  # 不提示保存

    # 步骤 7：重新连接原始文件
    li()
    time.sleep(1)
    active_doc = acad.ActiveDocument

    



def batch_export_regions_to_files(region_list, filename_prefix="区域导出",ty=1):##文件分解成多个小文件存储
    """
    批量导出多个区域为 DWG 文件。

    参数：
        region_list: 列表，每项为一个 (x1, y1, x2, y2) 元组
        filename_prefix: 保存文件名前缀
        ty---通过删除再恢复对象以将隐性选择转变为显性选择，留一个延迟时间控制参数
    """
    #确保文件夹中名含"区域导出"的文件被删除

    folder_1 = current_dwg_folder()
    clear_files_with_prefix(folder_1, filename_prefix="区域导出", delay = ty)
    time.sleep(ty)



    name1 = current_dwg_basename()

    for idx, region in enumerate(region_list):
        if len(region) != 4:
            print(f"❌ 区域第 {idx+1} 个格式错误，跳过。")
            continue

        x1, y1, x2, y2 = region
        filename = f"{filename_prefix}_{idx+1:02d}"
        filename = name1 + filename

        chongfu_caozuo(
            export_region_to_new_file,
            dwg_instance=None,
            args=(x1, y1, x2, y2, filename),
            kwargs={'ty':ty},
            max_retries=3,
            failure_value=None
        )




#&&% #双文件操作

def copy_region_from_doc1_to_doc2_absolute(doc1, doc2, x1, y1, x2, y2):#从doc1的一个区域原位粘贴到doc2
    """
    将 doc1 中指定区域 (x1, y1)-(x2, y2) 内的对象复制，
    粘贴到 doc2 的相同位置 (0,0,0)。（原位复制）

    当未能完全控制激活文件时，这些函数运行并不稳定
    
    要求：
    - li() 可更新当前全局连接
    - select_objects_in_window_area() 返回 COM 对象列表
    - yin_to_xian_xuanze() 将对象转为命令行选中状态
    """
    try:
        # Step 1: 激活并连接源文档
        set_active_doc(doc1)
        li()

        # Step 2: 选择区域内对象
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("❌ 区域内未找到对象，复制中止")
            return

        # Step 3: 将对象蓝色选中状态
        yin_to_xian_xuanze(LB)

        # Step 4: 执行 _copybase 和 _copyclip
        doc1.SendCommand("_copybase" + chr(13) + "0,0,0" + chr(13) + chr(13))
        time.sleep(2)
        
        """        
        如果出现粘贴回源文档的混乱，可以考虑将源文件在已经执行了set_active_doc(doc2)后关闭doc1

        """        

        # Step 5: 激活并连接目标文档
        set_active_doc(doc2)
        li()

        # Step 6: 粘贴到相同位置
        doc2.SendCommand("TPasteClip" + chr(13) + "0,0,0" + chr(13) + chr(13))#_pastclip效果一样
        time.sleep(1.5)

        print(f"✅ 区域对象已从 {doc1.Name} 粘贴到 {doc2.Name}")

    except Exception as e:
        print(f"❌ 复制区域对象失败: {e}")

def copy_region_from_doc1_to_doc2_relative(doc1, doc2, x1, y1, x2, y2):#从doc1区域零点粘贴到doc2整个空间零点
    """
    将 doc1 中指定区域 (x1, y1)-(x2, y2) 内的对象复制，
    粘贴到 doc2 的相同位置 (0,0,0)。（原位复制）
    
    要求：
    - li() 可更新当前全局连接
    - select_objects_in_window_area() 返回 COM 对象列表
    - yin_to_xian_xuanze() 将对象转为命令行选中状态
    """
    try:
        # Step 1: 激活并连接源文档
        set_active_doc(doc1)
        li()

        # Step 2: 选择区域内对象
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("❌ 区域内未找到对象，复制中止")
            return

        # Step 3: 将对象蓝色选中状态
        yin_to_xian_xuanze(LB)

        # Step 4: 执行 _copybase 和 _copyclip
        doc1.SendCommand("_copybase\n0,0,0\n\n")
        time.sleep(0.5)
        doc1.SendCommand("_copyclip\n\n")
        time.sleep(2)

        
        """        
        如果出现粘贴回源文档的混乱，可以考虑将源文件在已经执行了set_active_doc(doc2)后关闭doc1

        """        


        # Step 5: 激活并连接目标文档
        set_active_doc(doc2)
        li()

        # Step 6: 粘贴到相同位置
        doc2.SendCommand("_pasteclip\n0,0,0\n\n")
        time.sleep(1.5)

        print(f"✅ 区域对象已从 {doc1.Name} 粘贴到 {doc2.Name}")

    except Exception as e:
        print(f"❌ 复制区域对象失败: {e}")


def copy_region_from_doc1_to_doc2_at_point(doc1, doc2, x1, y1, x2, y2, x0, y0, z0=0):#将doc1指定区域的最低点，粘贴到doc2的目标点
    """
    将 doc1 中指定区域 (x1, y1)-(x2, y2) 内的对象复制，
    粘贴到 doc2 的指定位置 (x0, y0, z0)。
    """
    try:
        # Step 1: 激活并连接源文档
        set_active_doc(doc1)
        li()

        # Step 2: 选择区域内对象
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("❌ 区域内未找到对象，复制中止")
            return

        # Step 3: 蓝色高亮
        yin_to_xian_xuanze(LB)

        # Step 4: 拷贝
        doc1.SendCommand("_copybase\n0,0,0\n\n")
        time.sleep(0.5)
        doc1.SendCommand("_copyclip\n\n")
        time.sleep(2)

        # Step 5: 激活目标文档
        set_active_doc(doc2)

        # 尝试多次连接目标文档
        for _ in range(5):
            try:
                li()
                break
            except:
                time.sleep(0.5)
        else:
            print("❌ 激活目标文档失败，无法连接")
            return

        # Step 6: 粘贴到指定点
        paste_cmd = f"_pasteclip\n{x0},{y0},{z0}\n\n"
        doc2.SendCommand(paste_cmd)
        time.sleep(1.5)

        print(f"✅ 区域对象已从 {doc1.Name} 粘贴到 {doc2.Name} 的指定点 ({x0}, {y0}, {z0})")

    except Exception as e:
        print(f"❌ 复制区域对象失败：{e}")





def insert_as_block(p,block_path = r"D:/Myprogramsystem/XT/MC_yuan.dwg"):#以块的插入将MC_yuan.dwg插入当前激活文件
    """
    将 MC_yuan.dwg 文件以块形式插入当前文档，并立即分解。

    MC_yuan.dwg 文件分别将9种门窗基元放入各自的图层

    jz-danmen jz-shuangmen jz-tuilamen jz-juanlianmen jz-zimumen jz-pingchuang jz-tuchuang jz-baiyechuang jz-gaochuang

    墙都在jizhunwall
    p   (x,y,z)三维坐标
    
    """
    
    if not os.path.exists(block_path):
        print("❌ 文件不存在：", block_path)
        return

    ensure_layer(layer_name="jizhunwall")

    x=p[0]

    y=p[1]

    z=p[2]

    # 插入块（插入点0,0,0，比例1）
    cmd = f"-insert\n{block_path}\n{x},{y},{z}\n1\n1\n0\n"
    doc.SendCommand(cmd)

    # 稳定等待图层对象出现（最多等 3 秒）
    LB = []
    for _ in range(30):
        time.sleep(0.1)
        LB = select_tuceng("jizhunwall")
        if LB:
            break
    else:
        print("❌ 图层中未能及时检测到对象")
        return

    print("选到的jizhutuceng对象数量", len(LB))

    block = LB[0]

    for attempt in range(3):
        try:
            block.Explode()
            print(f"✅ 第 {attempt+1} 次炸块成功")
            break
        except Exception as e:
            print(f"⚠️ 第 {attempt + 1} 次炸块失败：{e}")
            time.sleep(0.2)
    else:
        print("❌ 多次尝试炸块失败")
        return

    try:
        block.Delete()
        print("🗑️ 原块对象已删除")
    except:
        print("⚠️ 删除原块失败")



#&&&&%%  第九部分 CAD图块 

##给属性块赋予新值，可以局部赋值


def huoqukuai_shuxing_zhi(XX):#XX为属性块实体


    attributes = XX.GetAttributes()

    tags=[]
    values=[]

    for attr in attributes:
               
        tag = attr.TagString

            
        value = attr.TextString

        tags.append(tag)

        values.append(value)

    return tags,values


def set_attributes_values(block, tags_order, new_values):
    """
    为属性块的标签设置新的值。

    参数:
    - block: 要修改的属性块（COM BlockReference 对象）。
    - tags_order: 一个列表，包含你希望按照哪种顺序为属性块的标签设置新的值。
    - new_values: 一个列表，包含按标签顺序排列的新值。

    返回:
    None
    """
    # 先尝试获取属性列表
    try:
        attributes = block.GetAttributes()
    except Exception as e:
        print(f"⚠ 实体 {block.ObjectName}({getattr(block, 'Handle', '?')}) 无法获取属性，跳过: {e}")
        return

    index = 0
    for tag in tags_order:
        # 找到对应标签的属性
        found = False
        for attr in attributes:
            if attr.TagString == tag:
                found = True
                try:
                    attr.TextString = new_values[index]
                    print(f"标签: {tag}  新值: {new_values[index]}")
                except Exception as e:
                    print(f"⚠ 设置标签 '{tag}' 时出错: {e}")
                index += 1
                break
        if not found:
            print(f"⚠ 未找到属性标签 '{tag}'，已跳过")

    # 更新块
    try:
        block.Update()
    except Exception as e:
        print(f"⚠ 更新块时出错: {e}")              

##>>> tags_order=["1.0","施工图","2023.10","1:100","1.0","专业名称"]
##>>> new_values=["1.2版","初步设计","2021.07","1:25","JS-09","建施"]



def resize_block_attribute(block_ref, tag: str, *, height: float = 200.0, width: float = 4500.0):
    """
    将块参照 block_ref 中指定 Tag 的属性文字改成给定字高并设置边界宽度。

    适用对象
    --------
    block_ref : AcadBlockReference
        必须是包含属性 (HasAttributes=True) 的块参照
    tag       : str
        目标属性 TagString（不区分大小写）
    height    : float
        目标字高（Drawing Units）
    width     : float
        多行属性 (MText attribute) 的边界宽度；若属性不是多行，
        尝试设置 WidthFactor 以近似效果。

    返回
    ----
    bool
        True  : 至少找到并修改了一个属性
        False : 没有找到指定 tag 或调整失败
    """
    if (getattr(block_ref, "ObjectName", "") != "AcDbBlockReference"
            or not getattr(block_ref, "HasAttributes", False)):
        print("⚠ 传入对象不是带属性的块参照")
        return False

    modified = False
    target_tag = tag.strip().upper()

    try:
        for attr in block_ref.GetAttributes():
            if attr.TagString.strip().upper() != target_tag:
                continue

            # ——— 字高 ———
            try:
                attr.Height = height
            except Exception as e:
                print(f"⚠ 设置 Height 失败: {e}")

            # ——— 边界宽度 / 宽度因子 ———
            # 多行属性是 AcDbAttributeReference，但内核中仍带 MText，
            # COM 暴露 'Width'；若没有就退而求其次改 WidthFactor
            if hasattr(attr, "Width"):
                try:
                    attr.Width = width
                except Exception as e:
                    print(f"⚠ 设置 Width 失败: {e}")
            else:
                try:
                    # 估算一个宽度因子使单行文本占据近似宽度
                    # 【经验】WidthFactor * 字符数 * 字高 ≈ 宽度
                    char_count = max(len(attr.TextString.replace("\\P", "")), 1)
                    wf = width / (char_count * height)
                    attr.WidthFactor = wf
                except Exception as e:
                    print(f"⚠ 设置 WidthFactor 失败: {e}")

            modified = True
    except Exception as e:
        print(f"⚠ GetAttributes() 失败: {e}")

    return modified










## 3 获取属性块里多段线矩形的坐标值

def huoqu_kuai_pl(blocka):#输入实体块，得到实体块中多段线矩形的坐标，其坐标以插入点的定义点为原点
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    kuaiming=blocka.Name

    # 获取块定义
    block_def = doc.Blocks.Item(str(kuaiming))

    # 获取块定义中的所有对象
    block_objects = list(block_def)

    # 查找三角形并删除
    for obj in block_objects:

##        print(obj.ObjectName)
        
        if obj.ObjectName == "AcDbPolyline":

            print(obj.Coordinates)

##>>> huoqu_kuai_pl(kuai[0])
##AcDbPolyline
##(-6000.0, 0.0, 0.0, 0.0, 0.0, 57400.0, -6000.0, 57400.0)
## 



#定义基点的块

def create_block_with_basepoint():
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 定义块的基点位置
    base_point = vtpnt(10, 10, 0)

    # 创建一个新的块
    block = doc.Blocks.Add(base_point, "MyBlock")

    # 在块中添加一个圆形实体
    block.AddCircle(base_point, 5)

#块的添加


def create_block_with_triangle_and_text():
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 创建新块
    grip = vtpnt(0, 0)
    blockObj = doc.Blocks.Add(grip, "MyBlock")

    # 在块中添加三角形
    pt1 = vtpnt(0, 0, 0)
    pt2 = vtpnt(10, 0, 0)
    pt3 = vtpnt(5, 10, 0)
    blockObj.AddLine(pt1, pt2)
    blockObj.AddLine(pt2, pt3)
    blockObj.AddLine(pt3, pt1)

    # 在块中添加文字对象
    text_point = vtpnt(2, 2, 0)
    blockObj.AddText("太美了", text_point, 2)

    print("块 'MyBlock' 创建成功")



def huoqu_kuai_pl(blocka):
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    kuaiming=blocka.Name

    # 获取块定义
    block_def = doc.Blocks.Item(str(kuaiming))

    # 获取块定义中的所有对象
    block_objects = list(block_def)

    # 查找三角形并删除
    for obj in block_objects:

        print(obj.ObjectName)
        if obj.ObjectName == "AcDbPolyline":

            print(obj.Coordinates)

    

# 块的边界

def get_bounding_box_of_block(block_name):
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 获取块定义
    block_def = doc.Blocks.Item(block_name)

    min_x, min_y, min_z = float('inf'), float('inf'), float('inf')
    max_x, max_y, max_z = float('-inf'), float('-inf'), float('-inf')

    # 遍历块定义中的所有对象
    for obj in block_def:
        try:
            # 尝试获取对象的边界框
            lower_left, upper_right = obj.GetBoundingBox()
            
            min_x = min(min_x, lower_left[0])
            min_y = min(min_y, lower_left[1])
            min_z = min(min_z, lower_left[2])
            
            max_x = max(max_x, upper_right[0])
            max_y = max(max_y, upper_right[1])
            max_z = max(max_z, upper_right[2])
        except:
            pass

    return ((min_x, min_y, min_z), (max_x, max_y, max_z))


def create_new_block_with_insert_and_line():
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 检查块名称"块1"是否已经存在
    if "块3" in [blk.Name for blk in doc.Blocks]:
        print("块名称'块3'已经存在。请选择一个新的名称或删除现有的块。")
        return

    # 创建新块的插入点
    grip = vtpnt(0, 0, 0)
    blockObj1 = doc.Blocks.Add(grip, "块3")

    # 在块1中插入MyBlock块
    insertion_point_for_myblock = vtpnt(10, 10, 0)
    blockObj1.InsertBlock(insertion_point_for_myblock, "MyBlock", 1, 1, 1, 0)

    # 在块1中添加一根直线段
    start_point = vtpnt(0, 0, 0)
    end_point = vtpnt(50, 50, 0)
    blockObj1.AddLine(start_point, end_point)

    print("块1已创建并添加了MyBlock和直线段")

##块的搜索
    
def copy_and_move_blocks_from_layer(layer_name, block_prefix):
    
        
    # 使用select_tuceng函数选择指定图层上的所有对象
    all_objects = select_tuceng(layer_name)
    
    # 过滤出块对象，且块名的前两个字母与指定的前缀匹配
    blocks = [obj for obj in all_objects if obj.ObjectName == "AcDbBlockReference" and obj.Name[:2] == block_prefix]
    
    # 定义移动的起始点和结束点
    vtpnt_from = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 0, 0])
    vtpnt_to = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 2000000, 0])
    
    # 对每个块进行复制和移动操作
    for block in blocks:
        # 复制块
        copied_block = block.Copy()
        
        # 移动复制的块
        copied_block.Move(vtpnt_from, vtpnt_to)

    print(f"Copied and moved {len(blocks)} blocks from layer {layer_name} with prefix {block_prefix}.")

#块名的清除

def delete_block_name(block_name):#先删除块名为block_name的实体再执行该命令删除块名，免得将来发生替换警告

    t1=time.time()

    # 获取块定义的集合
    blocks = acad.ActiveDocument.Database.Blocks

    print("len(blocks)",len(blocks))

    # 遍历所有块定义
    for block in blocks:
        # 检查块名称是否匹配
        if block.Name == block_name:
            try:
                # 尝试删除块
                block.Delete()
                print(f"块 '{block_name}' 已被删除。")
            except Exception as e:
                print(f"删除块 '{block_name}' 时发生错误: {str(e)}")
            break
    else:
        # 如果未找到块
        print(f"未找到名为 '{block_name}' 的块。")

    t2 = time.time()

    print("删除块名耗时：",t2-t1)

#更改实体块名

def rename_block_entity(ent, new_name):
    """
    将给定块参照实体 ent 的块名改为 new_name。
    如果 new_name 在 Block 表中已存在，则该实体将指向已有定义；
    否则将重命名它当前所引用的块定义。
    
    参数：
      ent       -- 一个 COM 块参照对象（如 BlockReference）
      new_name  -- 目标块名（字符串）
    """
    # 获取当前文档和块表

    blocks = doc.Blocks
    old_name = ent.Name

    try:
        # 尝试查找 new_name 是否已存在
        blocks.Item(new_name)
        # 存在：直接让该实体引用此定义
        ent.Name = new_name
    except Exception:
        # 不存在：重命名它当前所引用的定义
        blk_def = blocks.Item(old_name)
        blk_def.Name = new_name

#&&% 由块名选择实例

def get_block_instances(block_name: str, max_retries: int = 5):
    """
    根据给定的块定义名，检索当前图形中所有对应的块参照实例（BlockReference），
    返回这些 COM 对象的列表。

    参数：
      block_name     – 块定义的名称（字符串），如 "MyBlock"
      max_retries    – 调用 select_kuai 时的最大重试次数

    返回：
      instances     – 包含所有匹配块参照的列表（如果未找到或者出错，返回空列表）
    """
    # ① 先调用 select_kuai() 拿到所有块实例 COM 对象（扁平列表）
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        print(f"❌ 调用 select_kuai 失败：{e}")
        return []

    instances = []
    # ② select_kuai 已经是所有块实例的列表，直接遍历
    for ent in all_blocks:
        try:
            # 对于块参照，EntityName 通常是 "AcDbBlockReference"
            # 且我们要筛选 Name 恰好等于 block_name 的那些
            if getattr(ent, "EntityName", "") == "AcDbBlockReference" and getattr(ent, "Name", "") == block_name:
                instances.append(ent)
        except Exception:
            # 如果某个 COM 对象没有 EntityName/Name 属性，就跳过
            continue

    print(f"选择到名为{block_name}的实例块{len(instances)}个")

    return instances



#&&% 从块实体对象获取其内部com对象

def get_entities_from_block_reference(block_ref):
    """
    获取块引用对象中的所有子实体（COM对象形式）。

    参数：
        block_ref: 块引用对象（AcDbBlockReference）
        doc: 当前 AutoCAD 文档对象（Document）

    返回：
        entities: 子实体列表
    """
    try:
        block_name = block_ref.EffectiveName
        block_def = doc.Blocks.Item(block_name)
        entities = [ent for ent in block_def]
        print(f"✅ 获取到 {len(entities)} 个子对象")
        return entities
    except Exception as e:
        print(f"❌ 获取失败：{e}")
        return []

    




#以块的方式插入文件
def insert_block_into_autocad(block_file_path, insertion_point=(0, 0, 0), scale=(1, 1, 1), rotation=0):
    """
    以块的方式插入 DWG 文件到 AutoCAD，并等待命令完成。
    """
    try:
        insertion_point = (insertion_point[0], insertion_point[1], insertion_point[2])
        # 插入块（COM）
        block = ms.InsertBlock(insertion_point, block_file_path, scale[0], scale[1], scale[2], rotation)
        # 等待 CAD 空闲
        try:
            acad = win32com.client.Dispatch("AutoCAD.Application")
            wait_cad_idle(acad)
        except Exception:
            pass
        print(f"✔ 块已插入，文件：{block_file_path}，插入点：{insertion_point}，缩放：{scale}，旋转角度：{rotation}")
    except Exception as e:
        print(f"⚠ 插入块时出错：{e}")
    def _area(b):
        x1, y1, x2, y2 = b
        return max(0.0, (x2 - x1) * (y2 - y1))

    def _contains(b_big, b_small, tol=1e-6):
        x1, y1, x2, y2 = b_big
        a1, b1, a2, b2 = b_small
        return (a1 >= x1 - tol and b1 >= y1 - tol and a2 <= x2 + tol and b2 <= y2 + tol)

    def _almost_same(b1, b2, tol=1e-6):
        return all(abs(x - y) <= tol for x, y in zip(b1, b2))

    last = None
    t0 = time.time()
    for k in range(1, max_retries + 1):
        try:
            # 取两类多段线
            lw = ss_select(mode="all", filter_types=[0], filter_data=["LWPOLYLINE"], autocast=autocast)
            old = ss_select(mode="all", filter_types=[0], filter_data=["POLYLINE"], autocast=autocast)
            ents = list(lw) + list(old)

            # 1) 根据顶点数与是否为轴对齐矩形筛选
            cands = []  # (bbox, area, ent)
            for ent in ents:
                try:
                    pl = cast_polyline(ent)
                    verts = polyline_vertices(pl)
                    if not verts:
                        continue
                    vnum = len(verts)
                    if vnum not in (4, 5):
                        continue
                    if not _closed_flag(pl):
                        continue
                    verts = _unique_vertices(verts)
                    if len(verts) != 4:
                        continue
                    if not _is_axis_aligned_rect(verts):
                        continue
                    bbox = _bbox_from_pts(verts)
                    cands.append((bbox, _area(bbox), ent))
                except Exception:
                    continue

            if not cands:
                print("⚠ 未找到矩形多段线候选。")
                return []

            # 2) 面积降序，去包含与去重：仅保留“极大矩形”
            cands.sort(key=lambda x: x[1], reverse=True)
            selected = []  # (bbox, area, ent)
            for bbox, ar, ent in cands:
                if any(_almost_same(bbox, b) for b, _, _ in selected):
                    continue
                if any(_contains(b, bbox) for b, _, _ in selected):
                    continue
                selected.append((bbox, ar, ent))

            result = [ent for _, _, ent in selected]

            # 3) 统一设置为红色，便于检查
            for ent in result:
                try:
                    ent.Color = 1  # ACI=1 红色
                    ent.Update()
                except Exception:
                    pass

            print(f"✅ select_polyline_chuantong 成功（第 {k} 次），耗时 {time.time() - t0:.3f}s，共 {len(result)} 条")
            return result
        except Exception as e:
            last = e
            print(f"⚠ select_polyline_chuantong 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception:
                pass
            time.sleep(0.5)
    print(f"❌ select_polyline_chuantong 在 {max_retries} 次后仍失败：{last!r}")
    return []

# === Cad wait helpers (auto-wait until CAD completes operations) ===

def _normpath(p: str) -> str:
    try:
        return os.path.normcase(os.path.abspath(p))
    except Exception:
        return p


def _doc_exists(acad, path: str) -> bool:
    want = _normpath(path)
    try:
        for d in acad.Documents:
            try:
                full = getattr(d, 'FullName', None) or getattr(d, 'Name', None)
                if full and _normpath(full) == want:
                    return True
            except Exception:
                continue
    except Exception:
        pass
    return False


def wait_cad_idle(acad, *, min_quiet: float = 0.4, poll: float = 0.2, timeout: float | None = None) -> bool:
    """轮询 AutoCAD 空闲直到连续 min_quiet 秒处于空闲。"""
    t0 = time.time()
    quiet_start = None
    while True:
        try:
            state = acad.GetAcadState()
            if bool(getattr(state, 'IsQuiescent', True)):
                if quiet_start is None:
                    quiet_start = time.time()
                if time.time() - quiet_start >= min_quiet:
                    return True
            else:
                quiet_start = None
        except Exception:
            pass
        if timeout is not None and (time.time() - t0) > timeout:
            return False
        time.sleep(poll)


def wait_document_opened(acad, path: str, *, poll: float = 0.25, timeout: float | None = None) -> bool:
    """等待指定 DWG 加入 acad.Documents 集合。"""
    t0 = time.time()
    while True:
        if _doc_exists(acad, path):
            return True
        if timeout is not None and (time.time() - t0) > timeout:
            return False
        time.sleep(poll)








