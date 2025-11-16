# -*- coding: utf-8 -*-

#&&&&%%   CADåŸºæœ¬æ“ä½œ 

#&&&&%%  ç¬¬ä¸€éƒ¨åˆ†  å¯¼å…¥ã€è½¬æ¢ã€è¿æ¥ç­‰å‰ç½®ç¨‹åº 
#_____________________________________________________________________________________________________________________________________________


# å¯¼å…¥æ¨¡å—


# â€”â€” è„šæœ¬æœ€é¡¶éƒ¨ï¼Œå°±å†™è¿™ä¸¤è¡Œ â€”â€” 
import warnings
warnings.filterwarnings(
    "ignore",
    r".*Revert to STA COM threading mode.*",
    category=UserWarning,
    module=r"pywinauto\..*"
)
from pywinauto.application import Application


#A______________________________________________________________

#B______________________________________________________________

import builtins



#C______________________________________________________________

from collections import defaultdict,deque

from concurrent.futures import ThreadPoolExecutor, TimeoutError

import cv2

import ctypes

from contextlib import redirect_stdout, redirect_stderr,suppress


#D______________________________________________________________


import datetime

#E______________________________________________________________

#F______________________________________________________________
import functools

from fractions import Fraction

from functools import cmp_to_key

from fitz import Rect

import fitz

from functools import wraps

import functools

from functools import wraps

import fire









#G______________________________________________________________

#H______________________________________________________________

#I______________________________________________________________

from itertools import chain

import importlib

import imageio

import inspect

import itertools

import io


#J______________________________________________________________


import  json


#K______________________________________________________________

#L______________________________________________________________

#M______________________________________________________________

import math

import multiprocessing

import mysql.connector

import matplotlib.pyplot as plt


from multiprocessing import Process


#N______________________________________________________________

import networkx as nx

import numpy as np 


#O______________________________________________________________

import os


# â€”â€”â€”â€” 1. æŠ‘åˆ¶ pygame æ¬¢è¿æç¤º â€”â€”â€”â€”
# å¿…é¡»åœ¨ä»»ä½• import pygame ä¹‹å‰è®¾ç½®
os.environ.setdefault("PYGAME_HIDE_SUPPORT_PROMPT", "1")


from openpyxl import Workbook, load_workbook

from openpyxl.utils import column_index_from_string as col2idx

#P______________________________________________________________

import PyPDF2.errors

import pywintypes

from  pyautocad  import Autocad,APoint,aDouble

import pyautocad.types

import pytesseract

from PIL import ImageGrab

from pythoncom import VT_ARRAY, VT_R8

import pyautogui



from pywinauto import findwindows

import psutil

import pythoncom

import pygame

import pygetwindow as gw

import pyttsx3

from PIL  import Image



import pprint

import pdb

from pathlib import Path

import psutil

from pygetwindow import getWindowsWithTitle

from pypdf import PdfReader, PdfWriter


#Q______________________________________________________________

#R______________________________________________________________

import runpy

import regex as xre

import re

import random

#S______________________________________________________________

from shapely.geometry import Point, Polygon,LineString

import subprocess

import shutil

from sympy import *

import sys

from scipy.spatial import ConvexHull


from subprocess import DETACHED_PROCESS




#T______________________________________________________________

import time

import tempfile

from typing import Tuple, Literal

from typing import Any, Dict, List ,Tuple, Optional

from typing import Sequence, Literal, Optional,Union

from typing import Dict, Set,List

import tkinter as tk

import traceback

import tkinter

import threading


#U______________________________________________________________
import uuid
#V______________________________________________________________

#W______________________________________________________________



# ===== pywin32 / COM åŸºç¡€ =====
import win32com                      # ç¡®ä¿é¡¶å±‚åŒ…å­˜åœ¨

import win32com.client

import win32com.client as win32      # å¸¸ç”¨å…¥å£ï¼Œç»Ÿä¸€åˆ«å
from win32com.client import CastTo, makepy, constants, VARIANT,Dispatch

# åŠ¨æ€è°ƒåº¦ï¼ˆå°‘æ•°åœºæ™¯ç”¨ï¼‰
import win32com.client.dynamic as dyn

# Windows API / GUI / è¿›ç¨‹
import win32gui
import win32api
import win32con
import win32process

import pythoncom                     # CoInitialize, COM é”™è¯¯
import pywintypes                    # com_error


"""
ä¸»åŠ›ç”¨ï¼šwin32.gencache.EnsureDispatch("AutoCAD.Application")ï¼ˆæ›´ç¨³ã€é…åˆ MakePyï¼‰ã€‚

å¶å°”ç”¨ï¼šwin32.Dispatch(...) æˆ– dyn.Dispatch(...)ï¼ˆéœ€è¦å¼ºåˆ¶æ™šç»‘å®š/åŠ¨æ€å±æ€§æ—¶ï¼‰ã€‚

é¿å…æŠŠå˜é‡å‘½åä¸º constantsï¼Œä»¥å…é®è”½ win32com.client.constantsã€‚

å¦‚æœä½ åšæˆè„šæ‰‹æ¶æ¨¡å—ï¼ˆscaffoldï¼‰ï¼Œå¯åŠ ï¼š

__all__ = [
    "win32com", "win32", "CastTo", "makepy", "constants", "VARIANT",
    "dyn", "win32gui", "win32api", "win32con", "win32process",
    "pythoncom", "pywintypes",
]


"""

#X______________________________________________________________

#Y______________________________________________________________

#Z______________________________________________________________

import psutil
import subprocess
import sys
from subprocess import CREATE_NO_WINDOW, DETACHED_PROCESS


from typing import List, Dict
from pathlib import Path
import datetime
from typing import List, Set, Dict




import datetime
import unicodedata


import psutil
import os
import time
import signal
import subprocess
from subprocess import CREATE_NO_WINDOW, DETACHED_PROCESS


from typing import Tuple
import io
import fitz  # PyMuPDF
from PIL import Image



#è‡ªå»ºåº“


sys.path.append('D:/Myprogramsystem')






#&&% å‡½æ•°åˆ«å
def alias(*names):
    """
    @alias("åˆ«å1","åˆ«å2",â€¦)
    def foo(...): â€¦
    """
    def decorator(func):
        mod = sys.modules[func.__module__]
        for nm in names:
            setattr(mod, nm, func)
        return func
    return decorator

# â€”â€” ä¿å­˜åŸå§‹ print â€”â€”
_orig_print = builtins.print

# ---------------- å…¨å±€åŠ«æŒ ----------------
_orig_print = builtins.print            # å…ˆä¿å­˜åŸ print
def suppress_all_prints():
    builtins.print = lambda *a, **k: None
def restore_all_prints():
    builtins.print = _orig_print

# ---------------- å±€éƒ¨è°ƒè¯•æ§åˆ¶ ----------------
DEBUG             = False              # æ€»å¼€å…³
_DEBUG_CODE_STACK = []                 # â† æ–°å¢ï¼šè°ƒè¯•å‡½æ•°è°ƒç”¨æ ˆ

def node(msg: str, *args, **kwargs):
    """
    åªæœ‰ DEBUG = True ä¸”å½“å‰å¸§å±äºã€æ ˆåº•ã€‘å‡½æ•°æ—¶æ‰æ‰“å°ã€‚
    """
    if not DEBUG or not _DEBUG_CODE_STACK:
        return
    frame = inspect.currentframe().f_back
    try:
        # åªå…è®¸â€œæœ€å¤–å±‚â€è°ƒè¯•å‡½æ•°è¾“å‡º
        if frame.f_code is _DEBUG_CODE_STACK[0]:
            _orig_print(msg.format(*args), **kwargs)
    finally:
        del frame

@alias("e")
def enable_debug():
    """å¼€å¯è°ƒè¯•ï¼šæ™®é€š print â†’ é™é»˜ï¼Œä»… node() ç”Ÿæ•ˆã€‚"""
    global DEBUG
    DEBUG = True
    suppress_all_prints()

@alias("d")
def disable_debug():
    """å…³é—­è°ƒè¯•ï¼šæ¢å¤æ™®é€š printï¼Œnode() å¤±æ•ˆã€‚"""
    global DEBUG
    DEBUG = False
    restore_all_prints()

def debuggable(func):
    """
    è£…é¥°å™¨ï¼šè¿›å…¥æ—¶æŠŠ func.__code__ æ¨å…¥æ ˆï¼›ç¦»å¼€æ—¶å¼¹æ ˆã€‚
    node() å§‹ç»ˆåªçœ‹æ ˆåº•å…ƒç´ ï¼Œæ‰€ä»¥åªæœ‰ç¬¬ä¸€æ¬¡è¿›å…¥çš„å‡½æ•°ä¼šçœŸæ­£è¾“å‡ºã€‚
    """
    @functools.wraps(func)
    def wrapper(*args, **kw):
        _DEBUG_CODE_STACK.append(func.__code__)     # â€”â€” å…¥æ ˆ
        try:
            return func(*args, **kw)
        finally:
            _DEBUG_CODE_STACK.pop()                 # â€”â€” å‡ºæ ˆ
    return wrapper


#&&% æ§åˆ¶å‡½æ•°è¿è¡Œæ—¶é—´
# â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- -- -- -- -- â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- --


# æ—¥å¿—é…ç½®
LOG_PATH = Path(r"D:/Myprogramsystem/XT/ç³»ç»Ÿè¿è¡Œä¸­æ–­é”™è¯¯è®°å½•.xlsx")
LOG_SHEET = "é”™è¯¯è®°å½•"

# åˆå§‹åŒ–æ—¥å¿—æ–‡ä»¶å’Œå·¥ä½œè¡¨
if not LOG_PATH.exists():
    wb = Workbook()
    ws = wb.active
    ws.title = LOG_SHEET
    ws.append(["æ—¶é—´", "å‡½æ•°å", "å‚æ•°repr", "å¤‡æ³¨"])
    wb.save(LOG_PATH)


def _log(func_name, args, kwargs, note):
    """
    å°†é”™è¯¯æˆ–è¶…æ—¶ä¿¡æ¯å†™å…¥ Excel æ—¥å¿—ã€‚å¦‚æœå·¥ä½œè¡¨ä¸å­˜åœ¨åˆ™åˆ›å»ºã€‚
    """
    from time import strftime
    wb = load_workbook(LOG_PATH)
    # ç¡®ä¿æ—¥å¿—è¡¨å­˜åœ¨
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
    else:
        ws = wb.create_sheet(LOG_SHEET)
        ws.append(["æ—¶é—´", "å‡½æ•°å", "å‚æ•°repr", "å¤‡æ³¨"])
    ws.append([strftime("%Y-%m-%d %H:%M:%S"),
               func_name,
               f"args={args!r}, kwargs={kwargs!r}",
               note])
    wb.save(LOG_PATH)


def _kill_acad():
    """
    å¼ºåˆ¶ç»“æŸæ‰€æœ‰ä»¥ acad å¼€å¤´çš„ CAD è¿›ç¨‹ã€‚
    """
    for proc in psutil.process_iter(("pid", "name")):
        name = proc.info.get("name")
        if name and name.lower().startswith("acad"):
            try:
                proc.kill()
            except Exception:
                pass


def timeout_and_log2(timeout_sec: float):
    """
    è£…é¥°å™¨ï¼šå¯¹å…³é”®å‡½æ•°æ·»åŠ åŒé‡è¶…æ—¶ä¿æŠ¤ã€‚
    - å®ˆæŠ¤çº¿ç¨‹ç›‘æ§ä¸»å‡½æ•°æ‰§è¡Œï¼Œè¶…è¿‡ timeout_sec ç§’å°†
      è°ƒç”¨ _kill_acad() å¹¶è®°å½•æ—¥å¿—ã€‚
    - å‡½æ•°å†…éƒ¨å¯æ ¹æ® timeout_sec+1 è§¦å‘äºŒæ¬¡è¶…æ—¶æ“ä½œã€‚

    ç”¨æ³•ï¼š
        @timeout_and_log2(20)
        def Fx(..., timeout_sec=20):
            # å¯é€‰äºŒæ¬¡è¶…æ—¶ç›‘æ§
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            finished = threading.Event()

            def watchdog():
                # ç­‰å¾… timeout_sec ç§’åï¼Œå¦‚å‡½æ•°å°šæœªç»“æŸï¼Œæ€CADå¹¶è®°å½•
                if not finished.wait(timeout_sec):
                    _kill_acad()
                    _log(func.__name__, args, kwargs,
                         f"è¶…æ—¶ {timeout_sec}sï¼Œå¼ºåˆ¶ç»ˆæ­¢ CAD")
            threading.Thread(target=watchdog, daemon=True).start()

            try:
                result = func(*args, **kwargs)
            except Exception:
                _log(func.__name__, args, kwargs,
                     f"å‡½æ•°å†…éƒ¨å¼‚å¸¸:\n{traceback.format_exc()}")
                raise
            finally:
                finished.set()

            return result
        return wrapper
    return decorator

# ---------- ç¤ºä¾‹ï¼šåŒé‡è¶…æ—¶æµ‹è¯•å‡½æ•° ----------
@timeout_and_log2(20)
def test_draw_circle_and_wait(center=(0,0,0), radius=10000, timeout_sec=20):
    """
    æµ‹è¯•ï¼šåœ¨ CAD ä¸­ç”»åœ†å¹¶è§¦å‘ç­‰å¾…å‘½ä»¤ï¼Œ
    åå°çº¿ç¨‹åœ¨ timeout_sec+1 ç§’åå‘é€æ— æ•ˆå‘½ä»¤ä»¥è§¦å‘ä¸­æ–­ã€‚
    """
    import threading
    from time import sleep, time

    # è·å–å½“å‰ DWG æ–‡ä»¶å
    try:
        current_file = current_dwg_basename()
    except Exception:
        current_file = '<unknown>'

    try:


        # ç¬¬1æ­¥ï¼šç”»åœ†
        circ = draw_circle(center, radius)
        if circ is None:
            print("[é”™è¯¯] åœ†ç»˜åˆ¶å¤±è´¥ï¼Œæµ‹è¯•ä¸­æ–­ã€‚")
            return


        # ç¬¬2æ­¥ï¼šè§¦å‘ CAD ç­‰å¾…
        doc.SendCommand("h\n")

        # ç¬¬3æ­¥ï¼šäºŒæ¬¡è¶…æ—¶çº¿ç¨‹ï¼štimeout_sec+1 ç§’åå‘æ— æ•ˆå‘½ä»¤ä¸­æ–­CAD
        def timeout_action():
            start = time()
            while True:
                if time() - start > timeout_sec + 1:
                    try:
                        doc.SendCommand("_.POINT 0,0,0")
                    except Exception as e:
                        print(f"[è­¦å‘Š] ç¬¬äºŒé‡è¶…æ—¶è§¦å‘ï¼Œä¸­æ–­ CAD: {e}")
                    break
                sleep(0.1)

        threading.Thread(target=timeout_action, daemon=True).start()
        return circ

    except Exception as e:
        print(f"[é”™è¯¯] test_draw_circle_and_wait æ•è·åˆ°å¼‚å¸¸ï¼Œé€€å‡º: {e}")
        # è®°å½•æ—¥å¿—ï¼šå‡½æ•°åã€è¾“å…¥å‚æ•°ã€å½“å‰ DWGã€å¼‚å¸¸ä¿¡æ¯
        try:
            current_file = getattr(doc, 'Name', '<unknown>')
            _log('test_draw_circle_and_wait', (center, radius), {'timeout_sec': timeout_sec},
                 f"å†…éƒ¨å¼‚å¸¸ï¼Œæ–‡ä»¶={current_file}, å¼‚å¸¸={e}")
        except Exception:
            pass
        return

#å‡½æ•°è®¡æ—¶

def timeit(func):
    """
    è¿è¡Œå‰åæ‰“å°è€—æ—¶ï¼Œå§‹ç»ˆä½¿ç”¨ _orig_printï¼Œ
    ä¸ä¼šè¢« enable_debug() çš„ print åŠ«æŒå½±å“ã€‚
    """
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        # è¿è¡Œå‰
        _orig_print(f"â± å¼€å§‹ `{func.__name__}` â€¦")
        t0 = time.time()

        result = func(*args, **kwargs)

        elapsed = time.time() - t0
        # è¿è¡Œå
        _orig_print(f"â± å®Œæˆ `{func.__name__}`ï¼Œè€—æ—¶ï¼š{elapsed:.3f} ç§’")
        return result

    return wrapper



#&&% å‡½æ•°æ‰“å°æ¶ˆæ¯æ§åˆ¶
# â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- -- -- -- -- â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- --



# ---------- ç¤ºä¾‹ï¼šæ‰“å°æ¶ˆæ¯å±è”½æµ‹è¯•å‡½æ•° ----------

def ceshi_xiaoxi_dayin():

    stc("æµ‹è¯•1")
    print("åªæ‰“å°ä¸»å‡½æ•°æ¶ˆæ¯")



#&&% JSONæ•°æ®è¯»å–å­˜æ”¾è®¾ç½®
# â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- -- -- -- -- â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- --

# â€”â€”â€”â€”â€”â€”â€”â€” å…¨å±€é…ç½® â€”â€”â€”â€”â€”â€”â€”â€”
SAVE_DIR = r"D:/Myprogramsystem/XT/æ–‡ä»¶å­—å…¸ä¿¡æ¯"
os.makedirs(SAVE_DIR, exist_ok=True)

# å…¨å±€ç¼“å­˜æ‰€æœ‰å·²ç»åŠ è½½çš„æ‰“å°å­—å…¸
PRINT_DICTS: dict[str, dict] = {}

# â€”â€”â€”â€”â€”â€”â€”â€” åŠ è½½æŒ‡å®š DWG çš„æ‰“å°å­—å…¸ â€”â€”â€”â€”â€”â€”â€”â€”
def com_to_handle(obj):
    """
    å¦‚æœæ˜¯ COM å¯¹è±¡ï¼Œåˆ™è¿”å›å®ƒçš„ Handleï¼ˆå­—ç¬¦ä¸²ï¼‰ï¼›å¦åˆ™åŸæ ·è¿”å› objã€‚
    """
    try:
        # å¤§å¤šæ•° Dispatch å¯¹è±¡éƒ½æœ‰ .Handle å±æ€§
        if hasattr(obj, "Handle"):
            return obj.Handle
    except pythoncom.com_error:
        pass
    return obj

def serialize(obj):
    """
    é€’å½’åœ°å°† dict/list/tuple/COMObject è½¬ä¸ºåªå«åŸºç¡€ç±»å‹(åŒ…æ‹¬ handle å­—ç¬¦ä¸²)çš„ç»“æ„ã€‚
    """
    if isinstance(obj, dict):
        return {k: serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [serialize(v) for v in obj]
    if isinstance(obj, tuple):
        return tuple(serialize(v) for v in obj)
    # ä¸æ˜¯å®¹å™¨çš„ï¼Œå°±å°è¯•è½¬æ¢æˆ handle
    return com_to_handle(obj)

def save_print_dict_generic(dwg_name: str, bind_dict: dict):
    """
    å°† bind_dict åºåˆ—åŒ–å¹¶å†™å…¥ JSON æ–‡ä»¶ï¼Œé”®ä¸º dwg_nameã€‚
    åºåˆ—åŒ–æ—¶è‡ªåŠ¨æŠŠæ‰€æœ‰ COM å¯¹è±¡è½¬æ¢ä¸ºå®ƒä»¬çš„ Handleã€‚
    """
    ser = serialize(bind_dict)
    path = os.path.join(SAVE_DIR, f"{dwg_name}_æ‰“å°å­—å…¸.json")
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(ser, f, ensure_ascii=False, indent=2)
    PRINT_DICTS[dwg_name] = ser
    print(f"[OK] å·²ä¿å­˜æ‰“å°å­—å…¸åˆ° {path}")

def load(dwg_name: str) -> dict:
    """
    ä» JSON æ–‡ä»¶åŠ è½½ä¹‹å‰ä¿å­˜çš„æ‰“å°å­—å…¸ï¼ˆåªå«åŸºæœ¬ç±»å‹ & handle å­—ç¬¦ä¸²ï¼‰ã€‚
    å¦‚æœå†…å­˜ä¸­å·²æœ‰ï¼Œåˆ™ä¼˜å…ˆè¿”å›å†…å­˜ä¸­çš„ã€‚
    """
    if dwg_name in PRINT_DICTS:
        return PRINT_DICTS[dwg_name]

    path = os.path.join(SAVE_DIR, f"{dwg_name}_æ‰“å°å­—å…¸.json")
    if os.path.isfile(path):
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        PRINT_DICTS[dwg_name] = data
        print(f"[OK] å·²åŠ è½½æ‰“å°å­—å…¸è‡ª {path}")
        return data

    print(f"[è­¦å‘Š] æœªæ‰¾åˆ° {path}ï¼Œè¿”å›ç©ºå­—å…¸")
    return {}

# â€”â€”â€”â€”â€”â€”â€”â€” è¾…åŠ©ï¼šè·å–å½“å‰ DWG çš„çº¯æ–‡ä»¶å â€”â€”â€”â€”â€”â€”â€”â€”
def current_dwg_basename() -> str:
    """
    acad.ActiveDocument.Name å¯èƒ½å¸¦è·¯å¾„æˆ–å¸¦ .dwgï¼Œ
    å–ä¸å¸¦æ‰©å±•çš„æ–‡ä»¶åã€‚
    """
    name = acad.ActiveDocument.Name
    base = os.path.splitext(os.path.basename(name))[0]
    return base

def current_dwg_folder():
    """
    è·å–å½“å‰æ‰“å¼€çš„ DWG æ–‡ä»¶æ‰€åœ¨çš„æ–‡ä»¶å¤¹è·¯å¾„ã€‚
    
    """
    try:
        
        full_path = doc.FullName  # e.g. "D:\\Projects\\MyDrawing.dwg"
        folder_path = os.path.dirname(full_path)
        return folder_path
    except Exception as e:
        print(f"[é”™è¯¯] æ— æ³•è·å–å½“å‰ DWG æ–‡ä»¶å¤¹ï¼š{e}")
        return None


#&&% æ•°æ®ç±»å‹è½¬æ¢å‡½æ•°


def vtpnt(x, y, z):
    """åæ ‡ç‚¹è½¬åŒ–ä¸ºæµ®ç‚¹æ•°"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, z))

def vtobj(obj):
    """è½¬åŒ–ä¸ºå¯¹è±¡æ•°ç»„"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj)

def vtFloat(lst):
    """åˆ—è¡¨è½¬åŒ–ä¸ºæµ®ç‚¹æ•°"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, lst)
    
def vtInt(lst):
    """åˆ—è¡¨è½¬åŒ–ä¸ºæ•´æ•°"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, lst)

def vtVariant(lst):
    """åˆ—è¡¨è½¬åŒ–ä¸ºå˜ä½“"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, lst)

def ConvertArrays2Variant(inputdata, vartype):#ä¾‹å¦‚vartype="Variant"
    import pythoncom
    if vartype == "ArrayofObjects":  # å¯¹è±¡æ•°ç»„
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, inputdata)
    if vartype == "Double":  # åŒç²¾åº¦
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, inputdata)
    if vartype == "ShortInteger":  # çŸ­æ•´å‹
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, inputdata)
    if vartype == "LongInteger":  # é•¿æ•´å‹
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I4, inputdata)
    if vartype == "Variant":  # å˜ä½“
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, inputdata)
    return outputdata

def vtlist(obj_list):#å°†comå¯¹è±¡åˆ—è¡¨è½¬æ¢ä¸ºå¯ç”¨win32æ•°æ®
    """

    å°†å¯¹è±¡åˆ—è¡¨è½¬ä¸º VARIANT ç±»å‹ä»¥ä¾› COM æ¥å£ä½¿ç”¨

    ä¾èµ–è¯­å¥ from win32com.client import VARIANT

    """
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj_list)


warnings.filterwarnings("ignore", category=UserWarning)



#&&% æ‰“å¼€å¤©æ­£start_applicationV9

def start_applicationV9(
    PTH: str = r"C:\Tangent\TArchT20V9",
    max_retries: int = 3,
    retry_delay: float = 2.0
) -> subprocess.Popen | None:
    """
    å¯åŠ¨å¤©æ­£ TGStart.exeï¼Œå¤±è´¥é‡è¯•ã€‚
    åŒæ—¶å¯åŠ¨ cad_dialog_killer.pyã€‚
    CADç•Œé¢å¿…é¡»å¯è§ï¼ˆä¸åœ¨åå°è¿è¡Œï¼‰ã€‚
    è¿”å› Popen å¯¹è±¡ï¼ˆæˆåŠŸï¼‰æˆ– Noneï¼ˆå¤±è´¥ï¼‰ã€‚
    """
    exe = os.path.join(PTH, "TGStart.exe")
    for attempt in range(1, max_retries + 1):
        try:
            # å¯åŠ¨CADï¼ˆç•Œé¢å¯è§ï¼Œä¸ä½¿ç”¨DETACHED_PROCESSï¼‰
            proc = subprocess.Popen([exe], cwd=PTH)
            print("[å¯åŠ¨] å¯åŠ¨å¤©æ­£CAD æˆåŠŸ")

            # å¯åŠ¨ cad_dialog_killer.py
            script_dir = os.path.dirname(os.path.dirname(__file__))  # è¿”å›åˆ°cadç›®å½•
            killer_script = os.path.join(script_dir, "system", "cad_dialog_killer.py")

            if not os.path.exists(killer_script):
                print(f"[é”™è¯¯] å¼¹çª—æ²»ç†è„šæœ¬ä¸å­˜åœ¨: {killer_script}")
            else:
                try:
                    killer_proc = subprocess.Popen(
                        [sys.executable, killer_script],
                        creationflags=subprocess.CREATE_NO_WINDOW
                    )
                    print(f"[å¯åŠ¨] cad_dialog_killer å·²å¯åŠ¨ (PID: {killer_proc.pid})")

                    # éªŒè¯è¿›ç¨‹æ˜¯å¦çœŸæ­£è¿è¡Œ
                    time.sleep(0.5)
                    if killer_proc.poll() is None:
                        print("[éªŒè¯] cad_dialog_killer æ­£åœ¨è¿è¡Œ")
                    else:
                        print("[è­¦å‘Š] cad_dialog_killer æ„å¤–é€€å‡º")

                except Exception as e:
                    print(f"[è­¦å‘Š] cad_dialog_killer å¯åŠ¨å¤±è´¥: {e}")

            return proc
        except Exception as e:
            print(f"ç¬¬ {attempt} æ¬¡å¯åŠ¨å¤±è´¥: {e}")
            if attempt < max_retries:
                print(f"ç­‰å¾… {retry_delay:.1f} ç§’åé‡è¯•â€¦")
                time.sleep(retry_delay)
    print(f"å·²è¾¾æœ€å¤§é‡è¯•æ¬¡æ•° ({max_retries})ï¼Œå¯åŠ¨å¤±è´¥ã€‚")
    return None




def get_acad_process_id(ming):#è·å–è¿›ç¨‹çš„ID
    for process in psutil.process_iter(attrs=['pid', 'name']):
        if str(ming) in process.info['name'].lower():
            return process.info['pid']
    return None
    
def jingchengshu_wenjian():#æŸ¥çœ‹cadè¿›ç¨‹æ•°

    CAD_PROCESS_NAME = "acad.exe"

    found_process_i = 0
    for process in psutil.process_iter(['pid', 'name']):
        if process.info['name'] == CAD_PROCESS_NAME:
            found_process_i = found_process_i+1            

    return  found_process_i 
    
def close_all_cad_processes():#å…³é—­æ‰€æœ‰è¿›ç¨‹ï¼ˆä½¿ç”¨taskkillå¼ºåˆ¶ç»ˆæ­¢ï¼‰
    """
    å¼ºåˆ¶å…³é—­æ‰€æœ‰CADè¿›ç¨‹
    ä½¿ç”¨ç³»ç»Ÿtaskkillå‘½ä»¤ç¡®ä¿å³ä½¿æœ‰å¼¹çª—ä¹Ÿèƒ½å…³é—­
    è¿”å›: Trueè¡¨ç¤ºæˆåŠŸï¼ŒFalseè¡¨ç¤ºå¤±è´¥
    """
    import subprocess

    max_retries = 3
    for attempt in range(max_retries):
        try:
            # æ£€æŸ¥å½“å‰CADè¿›ç¨‹æ•°
            process_count = jingchengshu_wenjian()
            if process_count == 0:
                print("[ä¿¡æ¯] æ²¡æœ‰CADè¿›ç¨‹éœ€è¦å…³é—­")
                return True

            print(f"[æ¸…ç†] æ£€æµ‹åˆ° {process_count} ä¸ªCADè¿›ç¨‹ï¼Œæ­£åœ¨å¼ºåˆ¶å…³é—­...")

            # ä½¿ç”¨taskkillå¼ºåˆ¶ç»ˆæ­¢æ‰€æœ‰acad.exe
            result = subprocess.run(
                ["taskkill", "/F", "/IM", "acad.exe"],
                capture_output=True,
                text=True,
                timeout=10
            )

            # æ£€æŸ¥è¿”å›ç 
            if result.returncode == 0:
                print("[æˆåŠŸ] CADè¿›ç¨‹å·²å…³é—­")
                time.sleep(2)

                # éªŒè¯æ˜¯å¦çœŸæ­£å…³é—­
                process_count_after = jingchengshu_wenjian()
                if process_count_after == 0:
                    print("[éªŒè¯] æ‰€æœ‰CADè¿›ç¨‹å·²æˆåŠŸå…³é—­")
                    return True
                else:
                    print(f"[è­¦å‘Š] ä»æœ‰ {process_count_after} ä¸ªCADè¿›ç¨‹æœªå…³é—­")

            elif result.returncode == 128:
                # æ²¡æœ‰æ‰¾åˆ°è¿›ç¨‹ï¼ˆå·²ç»å…³é—­ï¼‰
                print("[ä¿¡æ¯] æ²¡æœ‰CADè¿›ç¨‹éœ€è¦å…³é—­")
                return True
            else:
                print(f"[è­¦å‘Š] taskkill è¿”å›ç : {result.returncode}")
                print(f"[è¾“å‡º] {result.stdout}")

        except subprocess.TimeoutExpired:
            print(f"[é”™è¯¯] ç¬¬ {attempt + 1} æ¬¡å…³é—­è¶…æ—¶")
        except Exception as e:
            print(f"[é”™è¯¯] ç¬¬ {attempt + 1} æ¬¡å…³é—­å¤±è´¥: {e}")

        if attempt < max_retries - 1:
            print(f"[é‡è¯•] ç­‰å¾… 2 ç§’åé‡è¯•...")
            time.sleep(2)

    print("[å¤±è´¥] å¤šæ¬¡å°è¯•å…³é—­CADè¿›ç¨‹å¤±è´¥ï¼Œè¯·æ‰‹åŠ¨æ£€æŸ¥")
    return False
    
def close_oldest_cad_process(process_name="acad.exe"):#å…³é—­ä¸Šä¸€ä¸ªè¿›ç¨‹
    cad_processes = [p for p in psutil.process_iter(['pid', 'name', 'create_time']) if p.info['name'] == process_name]
        
    # æ£€æŸ¥æ˜¯å¦æœ‰å¤šä¸ªCADè¿›ç¨‹
    if len(cad_processes) > 1:
        # æŒ‰å¯åŠ¨æ—¶é—´æ’åºï¼Œæœ€æ—©çš„è¿›ç¨‹åœ¨å‰
        oldest_process = sorted(cad_processes, key=lambda p: p.info['create_time'])[0]
            
        try:
            # å…³é—­æœ€æ—©çš„è¿›ç¨‹
            psutil.Process(oldest_process.pid).terminate()
            print(f"å·²å…³é—­æœ€æ—©çš„CADè¿›ç¨‹ï¼ŒPID: {oldest_process.pid}")
        except Exception as e:
            print(f"å…³é—­è¿›ç¨‹æ—¶å‡ºé”™: {e}")
            pass
    else:
        print("æ²¡æœ‰å¤šä¸ªCADè¿›ç¨‹è¿è¡Œã€‚")
        pass


#&&%  cadè¿æ¥ã€å…³é—­

def li():#è¿™é‡Œçš„li()ä¸éœ€è¦å†™æˆå¤šæ¬¡å°è¯•
    global acad, doc, mp, sp
    acad, doc = get_acad_doc()
    mp = doc.ModelSpace
    sp = doc.PaperSpace
    print("å½“å‰æ¡Œé¢æ–‡ä»¶ï¼š", doc.Name)
    print("win32å·²ç»è¿æ¥æ­£å¸¸â€”CADåŸºæœ¬æ“ä½œ")
    return True





_RPC_BUSY = (-2147417846, -2147418111)  # åº”ç”¨ç¨‹åºå¿™/Call rejected
_RPC_DOWN = (-2147023174,)              # RPC æœåŠ¡å™¨ä¸å¯ç”¨

def com_retry(fn, retries=30, delay=0.05):
    for _ in range(retries):
        try:
            return fn()
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if code in _RPC_BUSY + _RPC_DOWN:
                time.sleep(delay); continue
            raise
    return fn()

def _coinit_once():
    try: pythoncom.CoInitialize()
    except pythoncom.error: pass

def ensure_typelib_from_running():
    """ä»æ­£åœ¨è¿è¡Œçš„ AutoCAD ç›´æ¥ç”Ÿæˆ makepyï¼ˆä¸ä¾èµ–æ³¨å†Œè¡¨ï¼‰"""
    _coinit_once()
    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
    tlb, _ = app._oleobj_.GetTypeInfo().GetContainingTypeLib()
    makepy.GenerateFromTypeLibSpec(tlb)

def get_acad_doc(max_wait=8.0):
    _coinit_once()
    t0 = time.time()
    app = None
    while True:
        try:
            app = win32.gencache.EnsureDispatch("AutoCAD.Application")
            doc = app.ActiveDocument
            _ = doc.Name
            return app, doc
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if (code in _RPC_BUSY + _RPC_DOWN) and (time.time()-t0 < max_wait):
                time.sleep(0.05); continue
            # è‡ªæ„ˆï¼šè‹¥æ— æ–‡æ¡£åˆ™æ–°å»º
            try:
                if app is None:
                    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
                doc = app.Documents.Add()
                _ = doc.Name
                return app, doc
            except Exception:
                pass
            raise

class _ComLiveProxy:
    """æŠŠ getter åŒ…æˆâ€œæ´»ä½“â€å¯¹è±¡ï¼šæ¯æ¬¡è®¿é—®éƒ½æ‹¿æœ€æ–°/å¯ç”¨çš„ COM å®ä¾‹"""
    def __init__(self, getter): self._getter = getter
    def __getattr__(self, name):
        obj = self._getter()
        return getattr(obj, name)
    # å¯é€‰ï¼šå…è®¸ç›´æ¥ä½œä¸ºå¯è¿­ä»£æˆ– bool ä½¿ç”¨æ—¶æ›´è‡ªç„¶
    def __dir__(self): return dir(self._getter())

# ===== ä½ æ–°çš„ li()ï¼ˆè£…è½½ä»£ç†ï¼‰ï¼Œå…¼å®¹æ—§å†™æ³• =====


def li_new():#20250924æ–°ç‰ˆæœ¬
    """
    è¿æ¥/ä¿®å¤ CAD ä¼šè¯ï¼Œå¹¶æŠŠ acad/doc/mp/sp è®¾ç½®ä¸ºâ€œæ´»ä½“ä»£ç†â€ã€‚
    åŸæœ‰å¤§é‡å‡½æ•°é‡Œç»§ç»­å†™ acad./doc./mp./sp. å³å¯ï¼Œä½†å®ƒä»¬ä¼šè‡ªåŠ¨å–åˆ°å½“å‰æœ‰æ•ˆå¯¹è±¡ã€‚
    """
    global acad, doc, mp, sp
    # å…ˆç¡®ä¿ makepy å­˜åœ¨ï¼ˆCastTo/å¼ºç±»å‹æ›´ç¨³ï¼‰
    try:
        ensure_typelib_from_running()
    except Exception:
        pass

    # å®‰è£…ä»£ç†ï¼ˆgetter å†…éƒ¨åŒ…å«é‡è¯•ä¸è‡ªæ„ˆï¼‰
    acad = _ComLiveProxy(lambda: get_acad_doc()[0])
    doc  = _ComLiveProxy(lambda: get_acad_doc()[1])
    mp   = _ComLiveProxy(lambda: get_acad_doc()[1].ModelSpace)
    sp   = _ComLiveProxy(lambda: get_acad_doc()[1].PaperSpace)

    # è½»é‡æ ¡éªŒ & æç¤º
    name = com_retry(lambda: doc.Name)
    print("å½“å‰æ¡Œé¢æ–‡ä»¶ï¼š", name)
    print("win32å·²ç»è¿æ¥æ­£å¸¸â€”CADåŸºæœ¬æ“ä½œ")
    return True
    



"""
@require_docï¼Œå®ƒæœ€å¼ºå¤§çš„åœ°æ–¹åœ¨äºè‡ªåŠ¨è·Ÿéšcadæ–‡ä»¶çš„åˆ‡æ¢ï¼Œè·Ÿè¸ªæ¿€æ´»æ–‡ä»¶ï¼Ÿæ›´å¿«çš„_.doc=get_cad_doc()æ¨¡å¼åˆ™æ˜¯ç¡®å®šæ²¡æœ‰æ–‡ä»¶åˆ‡æ¢æ—¶æ›´å¿«




"""

#è£…é¥°å™¨åŠ¨æ€æ¨¡å¼

def require_doc(fn):
    def wrapper(*args, doc=None, **kw):
        if doc is None:
            _, doc = get_acad_doc()
        return fn(*args, doc=doc, **kw)
    return wrapper

def rpc_safe(fn):
    def wrapper(*a, **k):
        return com_retry(lambda: fn(*a, **k))
    return wrapper

#åªè¯»ç±»å‡½æ•°æ¨¡æ¿

@require_doc
@rpc_safe
def read_active_doc_info(doc=None):
    """ç¤ºä¾‹ï¼šè¯»å–å½“å‰å›¾åã€ç©ºé—´å®ä½“æ•°é‡"""
    name = com_retry(lambda: doc.Name)
    msp  = doc.ModelSpace
    count = com_retry(lambda: msp.Count)
    return {"name": name, "model_count": count}

#ä¿®æ”¹ç±»å‡½æ•°æ¨¡æ¿ï¼ˆå¸¦æ’¤é”€ã€å¤±è´¥ä¹Ÿèƒ½æ”¶å°¾
@require_doc
def move_selected_by(dx, dy, dz=0.0, doc=None):
    """ç¤ºä¾‹ï¼šäº¤äº’æ‹¾å–è‹¥å¹²å¯¹è±¡å¹¶å¹³ç§»"""
    ents = pmxz(prompt="\né€‰æ‹©è¦å¹³ç§»çš„å¯¹è±¡ï¼Œå›è½¦ç»“æŸï¼š")
    if not ents:
        print("æœªé€‰æ‹©å¯¹è±¡"); return 0

    com_retry(lambda: doc.StartUndoMark())
    moved = 0
    try:
        for e in ents:
            com_retry(lambda e=e: e.Move((0,0,0), (dx,dy,dz)))
            moved += 1
    finally:
        com_retry(lambda: doc.EndUndoMark())
    print(f"[OK] å·²å¹³ç§» {moved} ä¸ªå¯¹è±¡")
    return moved

#å‘½ä»¤è¡Œäº¤äº’æ¨¡æ¿ï¼ˆSendCommand + ç­‰å¾…ï¼‰
"""
ä¸»è¦æ˜¯ä¸ºäº†ç¨³å®šæ€§å’Œå¯æ§æ€§ï¼ŒåŒºåˆ«åœ¨äºï¼š

ç›´æ¥ç”¨ doc.SendCommand()

æœ€å¿«ï¼Œç›´æ¥æŠŠå­—ç¬¦ä¸²é€è¿› CAD å‘½ä»¤è¡Œã€‚

ä½†è¦è‡ªå·±åŠ  \n (chr(13))ï¼Œå¦åˆ™å‘½ä»¤ä¸ä¼šç”Ÿæ•ˆã€‚

æ²¡æœ‰ç­‰å¾…/ç¡®è®¤é€»è¾‘ï¼ŒCAD è¿˜æ²¡æ‰§è¡Œå®Œï¼Œä½ çš„ Python å¯èƒ½å°±ç»§ç»­å¾€ä¸‹è·‘äº†ã€‚

å°è£…è¿‡çš„æ¨¡æ¿ï¼ˆSendCommand + ç­‰å¾…ï¼‰

ä¼šåœ¨ SendCommand åè‡ªåŠ¨ time.sleep(x) æˆ–å¾ªç¯æ£€æµ‹ CAD æ˜¯å¦ç©ºé—²ï¼ˆå¦‚ While acad.GetAcadState().IsQuiescent == Falseï¼‰ã€‚

å¯ä»¥åŠ æ—¥å¿—ï¼šæ‰§è¡Œäº†ä»€ä¹ˆå‘½ä»¤ã€è€—æ—¶å¤šä¹…ã€‚

é‡åˆ° CAD å¿™ç¢Œ / æ‹’ç»è°ƒç”¨æ—¶ï¼Œå¯ä»¥è‡ªåŠ¨é‡è¯•ã€‚

ä¸€èˆ¬è¿˜ä¼šåšâ€œæ¢è¡Œæ¸…ç†â€ï¼Œé¿å…ä¸Šä¸€æ¡å‘½ä»¤æ®‹ç•™å½±å“ä¸‹ä¸€æ¡ã€‚

è¿”å›ç»“æ„ï¼šå°½é‡è¿”å›æ˜ç¡®çš„å­—å…¸/å…ƒç»„/æ•°é‡ï¼Œä¾¿äºä¸Šå±‚ç»„åˆä½¿ç”¨ã€‚

å¼‚å¸¸ï¼šæ™®é€š COM æ³¢åŠ¨äº¤ç»™ com_retryï¼›çœŸæ­£çš„é€»è¾‘é”™è¯¯ï¼ˆç±»å‹ä¸ç¬¦ã€æ•°æ®ä¸å…¨ï¼‰ç›´æ¥æŠ› RuntimeError æˆ–è¿”å› None å¹¶æ‰“å°å‹å¥½ä¿¡æ¯ã€‚





"""

def send_cmd(cmd: str, wait_s: float = 0.3):
    """å®‰å…¨å‘é€å‘½ä»¤å¹¶å°ç­‰å¾…ï¼Œé¿å…å‘½ä»¤å †ç§¯"""
    _, doc = get_acad_doc()
    doc.SendCommand(cmd if cmd.endswith("\n") else (cmd + "\n"))
    time.sleep(wait_s)

# CADååŒæœºåˆ¶å¢å¼ºåŠŸèƒ½ - å¯¼å…¥ååŒæœºåˆ¶æ¨¡å—
try:
    from CAD_coordination import (
        wait_quiescent,
        wait_document_opened,
        send_cmd_with_sync,
        start_cad_with_dialog_killer,
        ensure_single_process
    )
    print("[æˆåŠŸ] CADååŒæœºåˆ¶æ¨¡å—å·²åŠ è½½")
except ImportError as e:
    print(f"[è­¦å‘Š] CADååŒæœºåˆ¶æ¨¡å—åŠ è½½å¤±è´¥: {e}")

    # æä¾›åŸºç¡€å®ç°ä½œä¸ºåå¤‡
    def wait_quiescent(min_quiet: float = 0.4, timeout: float = 30.0) -> bool:
        """åŸºç¡€ç­‰å¾…CADç©ºé—²(åå¤‡å®ç°)"""
        time.sleep(min_quiet)
        return True

    def wait_document_opened(path: str, timeout: float = 120.0) -> bool:
        """åŸºç¡€ç­‰å¾…æ–‡æ¡£æ‰“å¼€(åå¤‡å®ç°)"""
        time.sleep(2.0)
        return True

    def send_cmd_with_sync(cmd: str, wait_after: float = 0.3, timeout: float = 30.0) -> bool:
        """åŸºç¡€åŒæ­¥å‘½ä»¤å‘é€(åå¤‡å®ç°)"""
        try:
            send_cmd(cmd, wait_after)
            return True
        except:
            return False

    def start_cad_with_dialog_killer() -> bool:
        """åŸºç¡€å¯åŠ¨CAD(åå¤‡å®ç°)"""
        return start_applicationV9() is not None

    def ensure_single_process() -> bool:
        """åŸºç¡€ç¡®ä¿å•è¿›ç¨‹(åå¤‡å®ç°)"""
        return True

@require_doc
def zoom_window(x1, y1, x2, y2, pad_ratio=0.1, doc=None):
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    pad = pad_ratio * ((x_hi-x_lo + y_hi-y_lo)/2.0)
    send_cmd(f"_.ZOOM\n_W\n{x_lo-pad},{y_lo-pad}\n{x_hi+pad},{y_hi+pad}\n", wait_s=0.6)



#é€‰æ‹© + Cast + è¯»å–å±æ€§æ¨¡æ¿

@require_doc
def pick_line_and_report(doc=None):
    """äº¤äº’é€‰ç›´çº¿å¹¶è¾“å‡ºèµ·ç»ˆç‚¹ã€é•¿åº¦"""
    ents = ss_select("onscreen", filter_types=[0], filter_data=["LINE"], autocast=True,
                     prompt="\nè¯·é€‰æ‹©ä¸€æ ¹ç›´çº¿ï¼š")
    if not ents:
        print("æœªé€‰åˆ°ç›´çº¿"); return None
    ln = ents[0]  # IAcadLineï¼ˆautocast å·²è½¬æ¢ï¼‰
    sp = com_retry(lambda: ln.StartPoint)
    ep = com_retry(lambda: ln.EndPoint)
    L  = com_retry(lambda: ln.Length)
    info = {"start": tuple(sp), "end": tuple(ep), "length": L}
    print(info)
    return info


#æ‰¹å¤„ç†æ¨¡æ¿ï¼ˆä¸äº¤äº’ã€æŒ‰è¿‡æ»¤æ‰¹é‡å¤„ç†ï¼‰

"""
ğŸ”¹ ä¸ºä»€ä¹ˆæˆ‘å•ç‹¬æå‡ºâ€œæ‰¹å¤„ç†æ¨¡æ¿â€ï¼Ÿ

å› ä¸ºå…‰æœ‰ for å¾ªç¯ è¿˜ä¸å¤Ÿï¼Œå®é™…åœ¨ CAD ä¸­ä¼šé‡åˆ°ä¸€äº›å‘ï¼š

æ’¤é”€ä¿æŠ¤
å¦‚æœä½ å¾ªç¯ 1000 ä¸ªå¯¹è±¡ï¼Œç”¨æˆ·åæ‚”æ—¶ï¼ŒCtrl+Z å¯èƒ½è¦ç‚¹ 1000 æ¬¡ã€‚
â†’ ç”¨ doc.StartUndoMark() / doc.EndUndoMark() åŒ…èµ·æ¥ï¼Œåªéœ€æ’¤é”€ä¸€æ¬¡ã€‚

RPC é”™è¯¯ / CAD å¿™
åœ¨å¾ªç¯é‡Œæœ‰æ—¶ä¼šæŠ¥ RPC æœåŠ¡å™¨ä¸å¯ç”¨ï¼Œæ‰€ä»¥æ¯ä¸ªå¯¹è±¡çš„å±æ€§è®¿é—®æœ€å¥½åŒ…ä¸€å±‚ com_retryã€‚

æ€§èƒ½
å¦‚æœæ¯æ¬¡å¾ªç¯éƒ½é‡æ–° get_acad_doc()ï¼Œä¼šå¾ˆæ…¢ã€‚
â†’ æ‰¹å¤„ç†æ¨¡æ¿é‡Œåªè·å–ä¸€æ¬¡ docï¼Œç„¶ååœ¨å¾ªç¯ä¸­ç›´æ¥ç”¨ã€‚

å®‰å…¨æ¸…ç†
SelectionSet è¦è®°å¾— Delete()ï¼Œå¦åˆ™ä¸‹ä¸€æ¬¡å¯èƒ½æŠ¥é”™ã€‚

ğŸ”¹ ä¸€ä¸ªæœ€ç®€å•çš„æ‰¹å¤„ç†ä¾‹å­
@require_doc
def batch_move_all_lines_to_layer0(doc=None):
    
    # 1) é€‰æ‹©
    ents = ss_select("all", filter_types=[0], filter_data=["LINE"], autocast=True)
    if not ents:
        print("[é”™è¯¯] æ²¡æœ‰æ‰¾åˆ°ç›´çº¿"); return 0

    # 2) æ’¤é”€å—
    com_retry(lambda: doc.StartUndoMark())
    moved = 0
    try:
        # 3) å¾ªç¯å¤„ç†
        for e in ents:
            com_retry(lambda: setattr(e, "Layer", "0"))
            moved += 1
    finally:
        # 4) ç»“æŸæ’¤é”€å—
        com_retry(lambda: doc.EndUndoMark())

    print(f"[OK] æ‰¹å¤„ç†å®Œæˆï¼Œå…±ç§»åŠ¨ {moved} æ¡ç›´çº¿åˆ°å›¾å±‚ 0")
    return moved


"""



@require_doc
def flatten_all_circles_to_z0(doc=None):
    """æŠŠæ‰€æœ‰åœ†çš„ Z ç½® 0ï¼ˆç¤ºä¾‹ï¼‰"""
    circles = ss_select("all", filter_types=[0], filter_data=["CIRCLE"], autocast=True)
    if not circles:
        print("æœªæ‰¾åˆ°åœ†"); return 0
    com_retry(lambda: doc.StartUndoMark())
    n = 0
    try:
        for c in circles:
            # ç§»åˆ° Z=0ï¼ˆCircle æ²¡æœ‰ Elevationï¼Œå°± Transformï¼‰
            sp = com_retry(lambda: c.Center)
            com_retry(lambda: c.Move(sp, (sp[0], sp[1], 0.0)))
            n += 1
    finally:
        com_retry(lambda: doc.EndUndoMark())
    print(f"[OK] å¤„ç†åœ† {n} ä¸ª")
    return n


#&&&% RGBè‰²å½©

def aci_to_rgb(ci: int):
    # åªç»™å‡ºå¸¸ç”¨ ACI çš„è¿‘ä¼¼è‰²ï¼›å…¶ä½™è¿”å› None
    table = {
        1:(255,0,0), 2:(255,255,0), 3:(0,255,0), 4:(0,255,255),
        5:(0,0,255), 6:(255,0,255), 7:(255,255,255)
    }
    return table.get(int(ci)) if ci is not None else None
def get_entity_rgb(ent):
    """è¿”å› (rgb_tuple_or_None, source_str)"""
    # 1) å…ˆçœ‹ TrueColor çš„æ–¹å¼
    tc = ent.TrueColor                 # IAcadAcCmColor
    method = getattr(tc, "ColorMethod", None)
    # acColorMethodByRGB å¤§å¤šä¸º 3ï¼ˆä¸åŒç‰ˆæœ¬æšä¸¾å€¼å¯èƒ½ä¸åŒï¼Œä½†é€»è¾‘ç›¸åŒï¼‰
    if method == 3 or (hasattr(tc, "Red") and (tc.Red or tc.Green or tc.Blue)):
        return (tc.Red, tc.Green, tc.Blue), "ByRGB(TrueColor)"

    # 2) çœ‹ ColorIndex / Colorï¼ˆACI / ByLayer / ByBlockï¼‰
    ci = getattr(ent, "ColorIndex", None)
    if ci is None:
        ci = getattr(ent, "Color", None)

    # ByLayer
    if ci == 256:
        layer_name = ent.Layer
        layer = ent.Document.Layers.Item(layer_name)
        ltc = layer.TrueColor
        return (ltc.Red, ltc.Green, ltc.Blue), f"ByLayer({layer_name})"

    # ByBlock
    if ci == 0:
        # çœŸæ­£çš„æ˜¾ç¤ºé¢œè‰²å–å†³äºå¼•ç”¨å®ƒçš„å—å‚ç…§/ä¸Šçº§å®¹å™¨
        return None, "ByBlock(éœ€æ ¹æ®ä¸Šçº§å—å‚ç…§è§£æ)"

    # 3) çº¯ ACIï¼ˆ1..255ï¼‰
    if isinstance(ci, int) and 1 <= ci <= 255:
        rgb = aci_to_rgb(ci)
        return (rgb if rgb else None), f"ACI({ci})"

    # å…œåº•
    return None, "Unknown"


"""
lj=pmxz()
pl=lj[0]
rgb, src = get_entity_rgb(pl)
print("RGB:", rgb, "| æ¥æº:", src)
RGB: (255, 0, 0) | æ¥æº: ByRGB(TrueColor)
"""

    
def  guanbi_cad_doc():#å…³é—­æ‰€æœ‰cadæ–‡ä»¶

        
    shuliang_cad = acad.Documents.Count

    print("å½“å‰æ¡Œé¢æ‰“å¼€çš„cadæ–‡ä»¶æ•°é‡:",shuliang_cad)

    for i in range(0,shuliang_cad):

        CourentDoc = acad.ActiveDocument

        CourentDoc.Close()

        li()
    
def guanbi_cad_one():#å…³é—­æ‰€æœ‰cadæ–‡ä»¶ä½†ä¿ç•™ä¸€ä¸ªå¹¶è¿æ¥
    shuliang_cad = acad.Documents.Count
    print("å½“å‰æ¡Œé¢æ‰“å¼€çš„cadæ–‡ä»¶æ•°é‡:", shuliang_cad)

    if shuliang_cad > 1:
        for i in range(shuliang_cad - 1):  # ä¿ç•™æœ€åä¸€ä¸ªæ–‡ä»¶
            courent_doc = acad.ActiveDocument
            courent_doc.Close()

        # é‡æ–°è¿æ¥åˆ°æœ€åä¸€ä¸ªæ–‡ä»¶
        if not li():
            print("é‡æ–°è¿æ¥åˆ°æœ€åä¸€ä¸ªæ–‡ä»¶å¤±è´¥ï¼Œè¯·æ£€æŸ¥ç³»ç»Ÿã€‚")
    else:
        print("åªæœ‰ä¸€ä¸ªCADæ–‡ä»¶æ‰“å¼€ï¼Œæ— éœ€å…³é—­å…¶ä»–æ–‡ä»¶ã€‚")

def st():#å¯åŠ¨æˆ–è¿ä¸ŠCADå¼€å§‹å·¥ä½œ
    """
    å½’é›¶é‡æ–°å½’äºæ ‡å‡†æ€çš„å¤„ç†

    """
    cad_process_count = jingchengshu_wenjian()

    if cad_process_count > 1:
        close_all_cad_processes()
        start_applicationV9()
        li()#ç”Ÿæˆå…¨å±€å˜é‡acad,mp,doc
    elif cad_process_count == 0:
        start_applicationV9()
        li()#ç”Ÿæˆå…¨å±€å˜é‡acad,mp,doc
    elif cad_process_count == 1:
        if not li():
            close_all_cad_processes()
            start_applicationV9()



def huifu_xitong():#å¯åŠ¨æˆ–è¿ä¸ŠCADå¼€å§‹å·¥ä½œ
    """
    å½’é›¶é‡æ–°å½’äºæ ‡å‡†æ€çš„å¤„ç†

    """
    close_all_cad_processes()

    start_applicationV9()




#&&&%  * é‡å¤å¤šæ¬¡è°ƒç”¨å‡½æ•°

def chongfu_caozuo(
    Fx,
    dwg_instance=None,
    args: tuple=(),
    kwargs: dict=None,
    max_retries: int=3,
    failure_value=None
):
    """
    å¯¹æŒ‡å®šå‡½æ•°/æ–¹æ³•è¿›è¡Œé‡å¤è°ƒç”¨ï¼Œç›´åˆ°æˆåŠŸæˆ–è€—å°½é‡è¯•æ¬¡æ•°ã€‚

    :param Fx:             å¾…è°ƒç”¨çš„å‡½æ•°å¯¹è±¡ï¼ˆæˆ–æ–¹æ³•ï¼‰ï¼Œä¸è¦åœ¨è¿™é‡Œå†™ Fx()
    :param dwg_instance:   å¦‚æœè¦è°ƒç”¨ç±»çš„æ–¹æ³•ï¼Œå°±æŠŠå®ä¾‹ä¼ è¿›æ¥ï¼›å¦åˆ™ä¸º None
    :param args:           ä½ç½®å‚æ•°ï¼ŒæŒ‰é¡ºåºä¼ ç»™ Fx
    :param kwargs:         å…³é”®å­—å‚æ•° dictï¼Œä¼ ç»™ Fx
    :param max_retries:    æœ€å¤šå°è¯•æ¬¡æ•°ï¼ˆ>=1ï¼‰
    :param failure_value:  å…¨éƒ¨å°è¯•å¤±è´¥æ—¶çš„è¿”å›å€¼

    :return: (result, attempts, error)
        - result:   Fx çš„è¿”å›å€¼ï¼Œæˆ– failure_value
        - attempts: æˆåŠŸæ—¶ä¸ºé‚£æ¬¡å°è¯•ç¼–å·ï¼ˆ1-basedï¼‰ï¼Œå¤±è´¥æ—¶ç­‰äº max_retries
        - error:    æœ€åä¸€æ¬¡æ•è·åˆ°çš„ Exception å®ä¾‹ï¼›æˆåŠŸæ—¶ä¸º None
    """
    if kwargs is None:
        kwargs = {}

    last_exception = None

    for attempt in range(1, max_retries + 1):
        try:
            # å†³å®šå¦‚ä½•è°ƒç”¨ Fx
            if dwg_instance is not None and not getattr(Fx, "__self__", None):
                # Fx å¾ˆå¯èƒ½æ˜¯ç±»ä¸­å®šä¹‰çš„å‡½æ•°ï¼Œä½†è¿˜æœªç»‘å®šåˆ°å®ä¾‹
                result = Fx(dwg_instance, *args, **kwargs)
            else:
                # Fx å·²ç»æ˜¯æ™®é€šå‡½æ•°æˆ–å·²ç»‘å®šçš„æ–¹æ³•
                result = Fx(*args, **kwargs)

            # æˆåŠŸï¼šè¿”å› (ç»“æœ, å°è¯•æ¬¡æ•°, None)
            return result, attempt, None

        except Exception as e:
            last_exception = e
            if attempt < max_retries:
                print(f"[è­¦å‘Š] æ“ä½œå‡ºé”™ï¼Œæ­£åœ¨ç¬¬ {attempt} æ¬¡é‡è¯•â€¦")
                # ï¼ˆå¯é€‰ï¼‰åœ¨é‡è¯•å‰åšäº›ç¯å¢ƒåˆ·æ–°ã€ç§»åŠ¨çª—å£ç­‰
                try:    li()  # ç¤ºä¾‹ï¼šé‡æ–°æ’åˆ— CAD/IDLE çª—å£
                except: pass
                time.sleep(2)
            else:
                print("[é”™è¯¯] å¤šæ¬¡å°è¯•å‡å¤±è´¥ï¼Œè¯·æ£€æŸ¥ç¯å¢ƒæˆ–å‚æ•°æ˜¯å¦æ­£ç¡®ã€‚")
                # è¿”å› (failure_value, ç”¨å°½çš„å°è¯•æ¬¡æ•°, æœ€åä¸€ä¸ªå¼‚å¸¸)
                return failure_value, attempt, last_exception


#  è‡ªåŠ¨è®¡æ—¶è£…é¥°å™¨

"""
ç”¨æ³•
@simple_timer
def heavy_func():
    ...

heavy_func()

"""
def simple_timer(func):
    def wrapper(*args, **kwargs):
        t0 = time.time()
        result = func(*args, **kwargs)
        t1 = time.time()
        print(f"â±ï¸ {func.__name__} è€—æ—¶ï¼š{t1 - t0:.4f} ç§’")
        return result
    return wrapper            









#&&&&%% ç¬¬äºŒéƒ¨åˆ† åˆ—è¡¨å¤„ç†


"""
è¯¥æ¨¡å—ç ”ç©¶åˆ—è¡¨çš„åŸºæœ¬é—®é¢˜ï¼Œç‰¹åˆ«æ˜¯comå¯¹è±¡åˆ—è¡¨ 


æœ‰ä¸€ä¸ªé’ˆå¯¹å…¨éƒ¨ç¨‹åºç³»ç»Ÿçš„åŸºæœ¬æ–‡ä»¶å¤„ç†æ–‡ä»¶Basic_oper.pyï¼Œä¾‹å¦‚å­—ç¬¦ä¸²ï¼Œåˆ—è¡¨ï¼Œé›†åˆçš„æ“ä½œç­‰ç­‰ï¼Œä¾¿äºæˆ‘ä»¬å±•å¼€åˆ«çš„ä¸“é¡¹å·¥ä½œæ—¶å¼•ç”¨è€Œä¸å¿…é‡å¤ç”Ÿæˆã€‚

åœ¨æ¯ä¸ªä¸“é¡¹å·¥ä½œæ–‡ä»¶ä¸­ï¼Œä»ç„¶æœ‰ä¸€å¥—è‡ªå·±çš„åŸºæœ¬å¤„ç†æ–‡ä»¶ï¼Œæ”¾åœ¨è„šæœ¬çš„æœ€å‰é¢ï¼Œè¿™ä¸¤è€…å¹¶ä¸é‡å¤ï¼Œè€Œä¸”å¾ˆåˆç†ã€‚æˆ‘ä»¬å°†åœ¨ä¸“é¡¹å·¥ä½œä¸­é›†ä¸­è¿™äº›åŸºæœ¬å¤„ç†ï¼Œ

å°†æ¥å†ç§»æ¤åˆ°Basic_oper.pyã€‚è™½ç„¶å‡½æ•°é‡å¤å‡ºç°åœ¨ä¸åŒçš„è„šæœ¬ï¼Œä½†å¹¶ä¸ä¼šæ··ä¹±ï¼Œä¹Ÿä¸ä¼šé™ä½æ•ˆç‡ã€‚è¿™åæ˜ äº†å†…åœ¨çš„æœ¬è´¨è§„å¾‹ï¼šBasic_oper.py

ä¸æ˜¯ä¸“é—¨å¼€å‘ç ”ç©¶çš„æˆæœï¼Œè€Œæ˜¯æ¥è‡ªå„ä¸ªä¸“é¡¹å·¥ä½œçš„æ€»ç»“ã€‚


"""


# æŒ‰comå®ä½“å¯¹è±¡ä¸­æå–çš„åæ ‡æ’åº


def sort_tuples(lst,cha_Y =2000):#å¯¹åˆ—è¡¨æŒ‰æ’å…¥ç‚¹xyåæ ‡æ’åº
    
    """
    è¿™æ˜¯å¾ˆæœ‰ç”¨çš„ä¸€ä¸ªåŒå€¼æ’åºå‡½æ•°ï¼Œå¯¹äºCOMå¯¹è±¡ï¼Œå¯ä»¥å…ˆå°†å…¶è½¬æ¢ä¸ºå…ƒç»„ï¼Œå³å¯ä½¿ç”¨è¿™ä¸ªå‡½æ•°

    å®ƒçš„ä»·å€¼åœ¨äºï¼Œå¾ˆå®¹æ˜“æ‹“å±•åˆ°nå€¼æ’åº
    """
    
    # å…ˆæŒ‰ç…§m[1]é™åºæ’åº
    lst.sort(key=lambda x: -x[2][1])

    i = 0
    while i < len(lst) - 1:
        j = i + 1
        # æŸ¥æ‰¾æ‰€æœ‰m[1]å·®è·åœ¨chaYä»¥å†…çš„å…ƒç´ 
        while j < len(lst) and abs(lst[i][2][1] - lst[j][2][1]) < cha_Y:
            j += 1
        
        # å¦‚æœæ‰¾åˆ°äº†m[1]å€¼ç›¸è¿‘çš„å…ƒç´ ï¼Œæ ¹æ®m[0]å€¼è¿›è¡Œæ’åº
        if j - i > 1:
            lst[i:j] = sorted(lst[i:j], key=lambda x: x[2][0])
        
        i = j

    return lst

def multi_dim_tolerance_sort(lst, key_index=2, tolerances=[10000, 1000, 0]):#é«˜ç»´æ’åº
    """
    å¯¹ lst åˆ—è¡¨ä¸­çš„å…ƒç»„æŒ‰å¤šç»´åæ ‡å­—æ®µæ’åºï¼Œè€ƒè™‘æ¯ä¸ªç»´åº¦çš„å®¹å·®è¿›è¡Œé€å±‚æ’åºã€‚

    å‚æ•°ï¼š
        lst: [(id, name, (x, y, z)), ...]
        key_index: åæ ‡åœ¨å…ƒç»„ä¸­çš„ç´¢å¼•ï¼ˆé»˜è®¤æ˜¯ç¬¬3é¡¹ï¼Œå³å…ƒç»„[2]ï¼‰
        tolerances: æ¯ä¸ªç»´åº¦å…è®¸çš„å®¹å·®ï¼Œä¾‹å¦‚ [Zå·®, Yå·®, Xå·®]

    è¿”å›ï¼š
        æ’å¥½åºçš„æ–°åˆ—è¡¨
    """
    # æœ€å¤–å±‚æ’åºï¼šæŒ‰æœ€é«˜ç»´é™åºæ’ï¼ˆZä»ä¸Šå¾€ä¸‹ï¼‰
    dim = len(tolerances)
    lst.sort(key=lambda x: -x[key_index][0])  # Z å€’åº

    def recursive_sort(sublist, level):
        if level >= dim - 1:
            # æœ€åä¸€çº§ç›´æ¥æŒ‰è¯¥ç»´å‡åº
            return sorted(sublist, key=lambda x: x[key_index][level + 1])
        
        i = 0
        while i < len(sublist) - 1:
            j = i + 1
            while j < len(sublist) and abs(sublist[i][key_index][level] - sublist[j][key_index][level]) < tolerances[level]:
                j += 1
            if j - i > 1:
                sublist[i:j] = recursive_sort(sublist[i:j], level + 1)
            i = j
        return sublist

    return recursive_sort(lst, 0)


def get_ll_pt(ent):#æå–å‡½æ•°ï¼Œå¯¹è±¡å·¦ä¸‹è§’ç‚¹
    minpt, _ = ent.GetBoundingBox()
    return minpt[0], minpt[1],0

def get_center(ent):#æå–å‡½æ•°ï¼Œä¸­å¿ƒç‚¹
    minpt, maxpt = ent.GetBoundingBox()
    return ((minpt[0]+maxpt[0])/2, (minpt[1]+maxpt[1])/2)

def sort_entities_by_position( entity_list, extract_func, cha_Y=2000):#å¯¹comå¯¹è±¡æŒ‰æå–åæ ‡åˆ†åˆ«æ²¿y,xæ–¹å‘æ’åº
        """
        å¯¹å®ä½“åˆ—è¡¨æ ¹æ®å…¶åæ ‡ï¼ˆé€šè¿‡ extract_func è·å–ï¼‰è¿›è¡Œæ’åºï¼š
        - å…ˆæŒ‰ Y å€¼é™åºï¼ˆä»ä¸Šåˆ°ä¸‹ï¼‰
        - Y å€¼æ¥è¿‘ï¼ˆå·®å€¼ < cha_Yï¼‰è€…å†æŒ‰ X å€¼å‡åºï¼ˆä»å·¦åˆ°å³ï¼‰

        å‚æ•°ï¼š
            entity_list: COM å®ä½“å¯¹è±¡åˆ—è¡¨
            extract_func: æå–åæ ‡å‡½æ•°ï¼Œè¿”å› (x, y) å…ƒç»„çš„å‡½æ•°å³å¯
            cha_Y: åŒä¸€è¡Œåˆ¤å®šçš„ Y æ–¹å‘å®¹å·®

        è¿”å›ï¼šæŒ‰åæ ‡é¡ºåºæ’åˆ—çš„æ–°å®ä½“å¯¹è±¡åˆ—è¡¨

        è°ƒç”¨ç¤ºä¾‹

        sorted_objs = sort_entities_by_position(LB, extract_func=get_ll_pt)
        
        """
        triples = [(ent, *extract_func(ent)) for ent in entity_list]

        # æŒ‰ Y å€¼é™åºæ’åˆ—
        triples.sort(key=lambda t: -t[2])

        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])  # æŒ‰ X å‡åº
            i = j

        return [t[0] for t in triples]

def get_line_start(ent):
    """
    æå–ä¸€æ¡ç›´çº¿çš„èµ·ç‚¹ (x, y)
    ent: AcDbLine æˆ–ç±»ä¼¼å¯¹è±¡ï¼Œå…·æœ‰ .StartPoint å±æ€§
    """
    pt = ent.StartPoint   # å‡è®¾æ˜¯ä¸€ä¸ª (x, y, z) æˆ– [x, y, z]
    return pt[0], pt[1]

#&&% * å¯¹åˆ—è¡¨å®ä½“è¿›è¡Œä»ä¸Šåˆ°ä¸‹ã€ä»å·¦åˆ°å³çš„æ’åº

def sort_coms_by_llcorner(com_list, cha_Y=2000):
    """
    æŒ‰ BoundingBox å·¦ä¸‹è§’åæ ‡æ’åºï¼š
      Â· å…ˆæŒ‰ y é™åºï¼ˆè¶Šå¤§è¶Šé å‰ï¼Œå³è‡ªä¸Šè€Œä¸‹ï¼‰
      Â· åŒä¸€è¡Œ(Î”y<cha_Y)å†…æŒ‰ x å‡åºï¼ˆè‡ªå·¦å‘å³ï¼‰
    """
    wrapped = []
    for ent in com_list:
        try:
            p1, _ = ent.GetBoundingBox()      # p1 å·²æ˜¯å·¦ä¸‹
            x_ll, y_ll = p1[0], p1[1]
        except Exception:
            x_ll = y_ll = float('-inf')       # å–ä¸åˆ°çš„ä¸€å¾‹æ”¾æœ€å
        wrapped.append((ent, x_ll, y_ll))     # (å®ä½“, x, y)

    # å…ˆæŒ‰ y é™åº
    wrapped.sort(key=lambda t: -t[2])

    i = 0
    while i < len(wrapped) - 1:
        j = i + 1
        while j < len(wrapped) and abs(wrapped[i][2] - wrapped[j][2]) < cha_Y:
            j += 1
        # è¡Œå†…å†æŒ‰ x å‡åº
        if j - i > 1:
            wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: t[1])
        i = j

    return [ent for ent, _, _ in wrapped]




def sort_coms_by_rbcorner(com_list, *, cha_X=100):
    """
    ç«–å‘å›¾æ¡†ï¼ˆæˆ–å·²æ•´ä½“æ—‹è½¬ -90Â° çš„å›¾çº¸ï¼‰ä½¿ç”¨ â€”â€”  
    Â· å…ˆæŒ‰åˆ—ï¼šxâ‚ é™åºï¼ˆå³ â†’ å·¦ï¼‰  
    Â· åˆ—å†…ï¼š  yâ‚‚ é™åºï¼ˆä¸Š â†’ ä¸‹ï¼‰

    å‚æ•°
    ----
    com_list : list[COM object]
        æ–‡å­—å®ä½“åˆ—è¡¨
    cha_X    : float
        åˆ¤å®šâ€œåŒä¸€åˆ—â€çš„å®¹å·®ï¼Œå•ä½ä¸å›¾çº¸åæ ‡ä¸€è‡´
    """
    wrapped = []
    for ent in com_list:
        try:
            (x1, _, _), (_, y2, _) = ent.GetBoundingBox()  # x1 = å·¦, y2 = ä¸Š
        except Exception:
            x1 = y2 = float("-inf")                        # å¤±è´¥çš„æ’æœ€å
        wrapped.append((ent, x1, y2))

    # â€” â‘  æŒ‰ xâ‚ é™åºï¼šæœ€å³åˆ—åœ¨å‰ â€”
    wrapped.sort(key=lambda t: -t[1])

    # â€” â‘¡ åŒåˆ—å†…å†æŒ‰ yâ‚‚ é™åºï¼šä¸Š â†’ ä¸‹ â€”
    i = 0
    while i < len(wrapped) - 1:
        j = i + 1
        while j < len(wrapped) and abs(wrapped[i][1] - wrapped[j][1]) < cha_X:
            j += 1
        wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: -t[2])
        i = j

    return [t[0] for t in wrapped]


def sort_coms_by_llcorner_custom(objs, tol_x=100):
    """
    æŒ‰å·¦ä¸‹è§’ x, y åæ ‡å¯¹ COM å¯¹è±¡åˆ—è¡¨ objs æ’åºï¼š
      1. å…ˆæŒ‰ x å‡åºåˆ†ç»„ï¼šç›¸é‚» x å·® â‰¤ tol_x è§†ä¸ºåŒä¸€ç»„
      2. æ¯ç»„å†…æŒ‰ y é™åºæ’åˆ—ï¼ˆy å¤§çš„æ’å‰ï¼‰
      3. æœ€ç»ˆæŠŠå„ç»„æŒ‰å®ƒä»¬çš„ x å‡åºä¾æ¬¡æ‹¼æ¥

    å‚æ•°
    ----
    objs : List[COMObject]
        è¦æ’åºçš„ COM å¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡å¿…é¡»æ”¯æŒ GetBoundingBox()
    tol_x : float
        x æ–¹å‘ä¸Šçš„å®¹å·®ï¼Œå°äºç­‰äºæ­¤å€¼çš„ x å·®è§†ä¸ºåŒç»„

    è¿”å›
    ----
    List[COMObject]
        æ’åºåçš„å¯¹è±¡åˆ—è¡¨
    """
    # æå– (obj, llx, lly)
    items = []
    for e in objs:
        ll, ur = e.GetBoundingBox()
        llx, lly, _ = ll
        items.append((e, llx, lly))

    # æŒ‰ x å‡åºåˆæ’
    items.sort(key=lambda t: t[1])

    # åˆ†ç»„ï¼šç›¸é‚» x å·® â‰¤ tol_x çš„å½’ä¸€ç»„
    clusters = []
    rep_x, current = items[0][1], [items[0]]
    for item in items[1:]:
        _, x, _ = item
        if abs(x - rep_x) <= tol_x:
            current.append(item)
        else:
            clusters.append((rep_x, current))
            rep_x, current = x, [item]
    clusters.append((rep_x, current))

    # æŒ‰ç»„çš„ x å‡åºæ’å¥½åï¼Œç»„å†…æŒ‰ y é™åºï¼Œå†å¹³é“º
    result = []
    for _, group in sorted(clusters, key=lambda c: c[0]):
        group_sorted = sorted(group, key=lambda t: t[2], reverse=True)
        result.extend([t[0] for t in group_sorted])

    return result


def sort_coms_by_center(objs, tol_x=100):
    """
    æŒ‰å¤–åŒ…ç›’ä¸­å¿ƒåæ ‡å¯¹ COM å¯¹è±¡åˆ—è¡¨ objs æ’åºï¼š
      1. å…ˆè®¡ç®—æ¯ä¸ªå¯¹è±¡çš„ bbox ä¸­å¿ƒ (cx, cy)
      2. æŒ‰ cx å‡åºåˆæ’åº
      3. å°†ç›¸é‚» cx å·® â‰¤ tol_x çš„å¯¹è±¡å½’åŒä¸€ç»„
      4. å„ç»„æŒ‰ cy é™åºæ’åˆ—ï¼ˆcy å¤§çš„æ’å‰ï¼‰
      5. æœ€åæŒ‰ç»„çš„ cx å‡åºä¾æ¬¡æ‹¼æ¥å„ç»„å†…å¯¹è±¡

    å‚æ•°
    ----
    objs : List[COMObject]
        è¦æ’åºçš„ COM å¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡å¿…é¡»æ”¯æŒ GetBoundingBox()
    tol_x : float
        x æ–¹å‘ä¸Šçš„å®¹å·®ï¼Œå°äºç­‰äºæ­¤å€¼çš„ cx å·®è§†ä¸ºåŒç»„

    è¿”å›
    ----
    List[COMObject]
        æ’åºåçš„å¯¹è±¡åˆ—è¡¨
    """
    # 1. æå– (obj, cx, cy)
    items = []
    for e in objs:
        ll, ur = e.GetBoundingBox()
        cx = (ll[0] + ur[0]) / 2.0
        cy = (ll[1] + ur[1]) / 2.0
        items.append((e, cx, cy))

    if not items:
        return []

    # 2. æŒ‰ cx å‡åºåˆæ’
    items.sort(key=lambda t: t[1])

    # 3. åˆ†ç»„ï¼šç›¸é‚» cx å·® â‰¤ tol_x çš„å½’ä¸€ç»„
    clusters = []
    rep_x, current = items[0][1], [items[0]]
    for item in items[1:]:
        _, x, _ = item
        if abs(x - rep_x) <= tol_x:
            current.append(item)
        else:
            clusters.append((rep_x, current))
            rep_x, current = x, [item]
    clusters.append((rep_x, current))

    # 4. å„ç»„å†…éƒ¨æŒ‰ cy é™åºæ’åˆ—ï¼Œå†å¹³é“º
    result = []
    for rep_x, group in sorted(clusters, key=lambda c: c[0]):
        group_sorted = sorted(group, key=lambda t: t[2], reverse=True)
        result.extend([t[0] for t in group_sorted])

    return result


##å¯¹åˆ—è¡¨å®ä½“è¿›è¡Œæ­£åºæˆ–é€†åºç¼–å·

def number_entities_by_order(entity_list, prefix="", start=1, k=0):
    """
    å¯¹æ’åºå¥½çš„ COM å®ä½“å¯¹è±¡åˆ—è¡¨è¿›è¡Œç¼–å·ã€‚

    å‚æ•°ï¼š
        entity_list: å®ä½“å¯¹è±¡åˆ—è¡¨
        prefix: ç¼–å·å‰ç¼€ï¼Œé»˜è®¤ä¸ºç©ºå­—ç¬¦ä¸²
        start: ç¼–å·èµ·å§‹æ•°å€¼ï¼Œé»˜è®¤ä¸º 1
        k: æ’åºæ–¹å‘æ§åˆ¶å˜é‡ï¼š
            - k = 0 æ­£åºç¼–å·
            - k = 1 é€†åºç¼–å·

    è¿”å›ï¼š
        ç¼–å·å­—ç¬¦ä¸²åˆ—è¡¨ï¼ˆå¦‚ ["1", "2", "3"] æˆ– ["å›¾1", "å›¾2", "å›¾3"]ï¼‰
    """
    n = len(entity_list)
    index_list = range(n) if k == 0 else reversed(range(n))
    
    result = []
    for i, idx in enumerate(index_list):
        label = f"{prefix}{start + i}"
        result.append(label)
    
    return result


# é‡å¤æ“ä½œåˆ—è¡¨å¯¹è±¡

def apply_to_each1(obj_list,  action_func):#é‡å¤æ“ä½œåˆ—è¡¨å¯¹è±¡
    """
    å¯¹ obj_list ä¸­çš„æ¯ä¸ªå¯¹è±¡ï¼Œ
    ä¼ å…¥ action_func ä¸­å¤„ç†ã€‚

    å‚æ•°ï¼š
        obj_list: å¯¹è±¡åˆ—è¡¨
        
        action_func: ç”¨äºå¤„ç†æå–ç»“æœçš„å‡½æ•°ï¼ˆå¦‚ srhdï¼‰
    """
    for obj in obj_list:
        
        action_func(obj)
        
def apply_to_each2(obj_list, extract_func, action_func):#åŒå±‚åµŒå¥—é‡å¤æ“ä½œåˆ—è¡¨å¯¹è±¡
    """
    å¯¹ obj_list ä¸­çš„æ¯ä¸ªå¯¹è±¡ï¼Œå…ˆé€šè¿‡ extract_func æå–å€¼ï¼Œ
    å†å°†è¯¥å€¼ä¼ å…¥ action_func ä¸­å¤„ç†ã€‚

    å‚æ•°ï¼š
        obj_list: å¯¹è±¡åˆ—è¡¨
        extract_func: ç”¨äºæå– (x, y) æˆ–å…¶ä»–å€¼çš„å‡½æ•°
        action_func: ç”¨äºå¤„ç†æå–ç»“æœçš„å‡½æ•°ï¼ˆå¦‚ srhdï¼‰
    """
    for obj in obj_list:
        value = extract_func(obj)
        action_func(value)

#&&&&%% ç¬¬ä¸‰éƒ¨åˆ† æ–‡ä»¶å¤¹æ–‡ä»¶å¤„ç†


"""
è¯¥æ¨¡å—æ˜¯æœ‰å…³ä¸€èˆ¬æ€§çš„æ–‡ä»¶å¤¹ã€æ–‡ä»¶ä¸Šçš„æ“ä½œ 

"""
#&&% ç¡®ä¿æ–‡ä»¶è¢«åˆ é™¤
def ensure_file_absent(save_path: str, ty: float = 1.0) -> None:
    """
    ç¡®ä¿æŒ‡å®šè·¯å¾„çš„æ–‡ä»¶ä¸å­˜åœ¨ã€‚å¦‚æœå­˜åœ¨ï¼Œåˆ™åˆ é™¤å¹¶ç­‰å¾… ty ç§’ã€‚

    :param save_path: å¸¦è·¯å¾„çš„æ–‡ä»¶åï¼Œä¾‹å¦‚ "D:/output/result.txt"
    :param ty: åˆ é™¤åç­‰å¾…çš„ç§’æ•°ï¼Œé»˜è®¤ä¸º 1 ç§’
    """
    try:
        if os.path.isfile(save_path):
            os.remove(save_path)
            # ç­‰å¾… ty ç§’ï¼Œä»¥ç¡®ä¿æ–‡ä»¶ç³»ç»Ÿå®Œæˆåˆ é™¤æ“ä½œ
            time.sleep(ty)
            print(f"[OK] å·²åˆ é™¤æ–‡ä»¶ï¼š{save_path}ï¼Œå¹¶ç­‰å¾…äº† {ty} ç§’")
        else:
            print(f"â„¹ æ–‡ä»¶ä¸å­˜åœ¨ï¼Œæ— éœ€åˆ é™¤ï¼š{save_path}")
    except Exception as e:
        print(f"[é”™è¯¯] åˆ é™¤æ–‡ä»¶æ—¶å‡ºé”™ï¼š{e}")


def traverse_with_os_walk(root_dir: str):
    """
    éå† root_dir åŠå…¶æ‰€æœ‰å­ç›®å½•ï¼Œæ‰“å°æ¯ä¸ªç›®å½•å’Œæ–‡ä»¶çš„å®Œæ•´è·¯å¾„ã€‚
    """
    for dirpath, dirnames, filenames in os.walk(root_dir):
        # dirpath: å½“å‰éå†åˆ°çš„ç›®å½•è·¯å¾„
        print(f"Directory: {dirpath}")
        for dirname in dirnames:
            print(f"  Sub-dir: {os.path.join(dirpath, dirname)}")
        for filename in filenames:
            print(f"  File   : {os.path.join(dirpath, filename)}")



def find_files_with_extensions(directory, extensions):
    #æ‰¾åˆ°ä»¥[".dwg"]ç»“å°¾çš„æ–‡ä»¶directoryä¸ºæ–‡ä»¶å¤¹åŠå…¶è·¯å¾„ï¼Œextensionsä¸ºåç¼€æˆ–ä¸­é—´ä½ç½®å­—ç¬¦åˆ—è¡¨
    matching_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if any(file.lower().endswith(ext) for ext in extensions):
                matching_files.append(os.path.join(root, file))
    return matching_files

def get_filename_without_extension(FileandPath):
    # ä»å®Œæ•´è·¯å¾„ä¸­è·å–æ–‡ä»¶åï¼ˆåŒ…å«æ‰©å±•åï¼‰
    filename_with_extension = os.path.basename(FileandPath)
    
    # åˆ†ç¦»æ–‡ä»¶åå’Œæ‰©å±•å
    filename_without_extension, _ = os.path.splitext(filename_with_extension)

    return filename_without_extension


def delete_files_with_patterns(folder_path, patterns):
    """
    åˆ é™¤æ–‡ä»¶å¤¹ä¸­ç¬¦åˆæŒ‡å®šæ¨¡å¼çš„æ–‡ä»¶ã€‚

    Args:
        folder_path (str): æ–‡ä»¶å¤¹çš„è·¯å¾„ã€‚
        patterns (list): åŒ…å«è¦åŒ¹é…çš„æ¨¡å¼çš„åˆ—è¡¨ï¼Œä¾‹å¦‚["_t7", "_t3", "_t8"]ã€‚

    Returns:
        None
    """
    # éå†æ–‡ä»¶å¤¹ä¸­çš„æ‰€æœ‰æ–‡ä»¶
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        # æ£€æŸ¥æ–‡ä»¶åæ˜¯å¦åŒ…å«ä»»ä½•ä¸€ä¸ªæŒ‡å®šçš„æ¨¡å¼
        if any(pattern in filename for pattern in patterns):
            # å¦‚æœç¬¦åˆæ¡ä»¶ï¼Œå°±åˆ é™¤è¯¥æ–‡ä»¶
            os.remove(file_path)
            print(f"Deleted file: {filename}")
            pass

    print("File deletion completed.")
    pass


#ç¡®ä¿æ–‡ä»¶å¤¹ä¸­åå­—å«æœ‰ç‰¹æ®Šå­—ç¬¦çš„æ–‡ä»¶è¢«æ¸…ç©º

def clear_files_with_prefix(folder: str, filename_prefix: str = "åŒºåŸŸå¯¼å‡º", delay: float = 0.5):
    """
    æ¸…é™¤æŒ‡å®šæ–‡ä»¶å¤¹ä¸­æ‰€æœ‰æ–‡ä»¶ååŒ…å«ç»™å®šå‰ç¼€çš„æ–‡ä»¶ã€‚è‹¥æ–‡ä»¶æ­£åœ¨è¢«å ç”¨æˆ–åˆ é™¤å¤±è´¥ï¼Œä¼šå°è¯•é‡è¯•ä¸€æ¬¡ã€‚

    :param folder:           è¦æ¸…ç†çš„æ–‡ä»¶å¤¹è·¯å¾„
    :param filename_prefix:  éœ€è¦åŒ¹é…çš„æ–‡ä»¶åå‰ç¼€ï¼ˆåªè¦æ–‡ä»¶åä¸­åŒ…å«è¯¥å­—ç¬¦ä¸²ï¼Œå°±ä¼šè¢«åˆ é™¤ï¼‰
    :param delay:            åˆ é™¤å¤±è´¥æ—¶çš„ç­‰å¾…æ—¶é—´ï¼ˆç§’ï¼‰ï¼Œé»˜è®¤ 0.5 ç§’
    """
    if not os.path.isdir(folder):
        print(f"[é”™è¯¯] ç›®æ ‡è·¯å¾„ä¸æ˜¯æœ‰æ•ˆæ–‡ä»¶å¤¹ï¼š{folder}")
        return

    # åˆ—å‡ºæ–‡ä»¶å¤¹ä¸­æ‰€æœ‰æ¡ç›®
    entries = os.listdir(folder)
    # è¿‡æ»¤å‡ºæ–‡ä»¶åä¸­åŒ…å«æŒ‡å®šå‰ç¼€çš„æ–‡ä»¶
    to_delete = [fname for fname in entries if filename_prefix in fname and os.path.isfile(os.path.join(folder, fname))]

    if not to_delete:
        print(f"â„¹ï¸ æ–‡ä»¶å¤¹ä¸­æœªå‘ç°æ–‡ä»¶ååŒ…å« â€œ{filename_prefix}â€ çš„æ–‡ä»¶ã€‚")
        return

    for fname in to_delete:
        full_path = os.path.join(folder, fname)
        try:
            os.remove(full_path)
            print(f"[OK] å·²åˆ é™¤ï¼š{fname}")
        except Exception as e:
            print(f"[è­¦å‘Š] åˆ é™¤å¤±è´¥ï¼ˆç¬¬ä¸€æ¬¡å°è¯•ï¼‰ï¼š{fname}ï¼Œé”™è¯¯ï¼š{e}ï¼Œç¨åé‡è¯•â€¦â€¦")
            time.sleep(delay)
            # å†è¯•ä¸€æ¬¡
            try:
                os.remove(full_path)
                print(f"[OK] é‡è¯•æˆåŠŸåˆ é™¤ï¼š{fname}")
            except Exception as e2:
                print(f"[é”™è¯¯] å†æ¬¡åˆ é™¤ä»å¤±è´¥ï¼š{fname}ï¼Œé”™è¯¯ï¼š{e2}")



def find_files_with_string(directory, search_string):
    #æ‰¾åˆ°æ–‡ä»¶å¤¹ä¸­å«æœ‰æŒ‡å®šå­—ç¬¦ä¸²çš„æ–‡ä»¶
    matched_files = []
    for file in os.listdir(directory):
        if search_string in file:
            matched_files.append(file)
    return matched_files

#ä½¿è·¯å¾„åå’Œæ–‡ä»¶ååˆå¹¶ååˆä¹é¢„æœŸ
def join_paths(p1, p2):
    # ä½¿ç”¨ os.path.join åˆå¹¶è·¯å¾„
    result = os.path.join(p1, p2)
    # æ›¿æ¢åæ–œæ ä¸ºæ­£æ–œæ 
    return result.replace("\\", "/")



#&&&&%% ç¬¬å››éƒ¨åˆ† é€‰æ‹©æ–¹æ³• 


# ===================== å‰ç½®å·¥å…·ï¼ˆç»Ÿä¸€å¤„ç†ï¼‰ =====================

# â€”â€” é€šç”¨é‡è¯•ï¼šåâ€œå¿™/æ‹’ç»/RPC downâ€
_RPC_BUSY = (-2147417846, -2147418111)
_RPC_DOWN = (-2147023174,)

def com_retry(fn, retries=30, delay=0.05):
    for _ in range(retries):
        try:
            return fn()
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if code in _RPC_BUSY + _RPC_DOWN:
                time.sleep(delay); continue
            raise
    return fn()

# â€”â€” ç°å–ç°ç”¨ï¼šè·å–å½“å‰æ¿€æ´»æ–‡æ¡£ï¼ˆæ— æ–‡æ¡£åˆ™æ–°å»ºï¼‰
def get_acad_doc(max_wait=8.0):
    try: pythoncom.CoInitialize()
    except pythoncom.error: pass
    t0 = time.time()
    app = None
    while True:
        try:
            app = win32.gencache.EnsureDispatch("AutoCAD.Application")
            doc = app.ActiveDocument
            _ = doc.Name
            return app, doc
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if (code in _RPC_BUSY + _RPC_DOWN) and (time.time()-t0 < max_wait):
                time.sleep(0.05); continue
            # æ— æ‰“å¼€å›¾ â†’ æ–°å»º
            try:
                if app is None:
                    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
                doc = app.Documents.Add(); _ = doc.Name
                return app, doc
            except Exception:
                pass
            raise

# â€”â€” è¿‡æ»¤æ•°ç»„ï¼ˆé¿å…ä¸ä½ åº“å·²æœ‰ vtInt/vtVariant å†²çªï¼‰
def to_vt_int(seq):
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, list(seq))

def to_vt_variant(seq):
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, list(seq))

# â€”â€” å¸¸è§å¯¹è±¡è‡ªåŠ¨ Castï¼šæ‹¿åˆ°ä¸“æ¥å£ï¼ˆStartPoint/Coordinates/Contentsâ€¦ï¼‰
_CAST_MAP = {
    # åŸºç¡€å‡ ä½•
    "AcDbLine":"IAcadLine", "AcDbCircle":"IAcadCircle", "AcDbArc":"IAcadArc","AcDbPoint":"IAcadPoint",
    "AcDbEllipse":"IAcadEllipse", "AcDbSpline":"IAcadSpline",
    # å¤šæ®µçº¿
    "AcDbPolyline":"IAcadLWPolyline", "AcDb2dPolyline":"IAcadPolyline", "AcDb3dPolyline":"IAcad3DPolyline",
    # æ–‡å­—
    "AcDbText":"IAcadText", "AcDbMText":"IAcadMText",
    # å—/å±æ€§
    "AcDbBlockReference":"IAcadBlockReference",
    "AcDbAttribute":"IAcadAttributeReference", "AcDbAttributeDefinition":"IAcadAttribute",
    # å¼•çº¿/æ ‡æ³¨ï¼ˆå¸¸ç”¨ï¼‰
    "AcDbLeader":"IAcadLeader", "AcDbMLeader":"IAcadMLeader",
    "AcDbDimension":"IAcadDimension", "AcDbAlignedDimension":"IAcadDimAligned",
    "AcDbRotatedDimension":"IAcadDimRotated", "AcDbRadialDimension":"IAcadDimRadial",
    "AcDbDiametricDimension":"IAcadDimDiametric", "AcDbArcDimension":"IAcadDimArc",
    "AcDb3PointAngularDimension":"IAcadDim3PointAngular", "AcDb2LineAngularDimension":"IAcadDim2LineAngular",
    "AcDbOrdinateDimension":"IAcadDimOrdinate",
    # å…¶å®ƒ
    "AcDbHatch":"IAcadHatch", "AcDbTable":"IAcadTable",
}

def _maybe_cast(ent):
    try:
        name = com_retry(lambda: ent.ObjectName)
        iface = _CAST_MAP.get(name)
        if iface:
            try: return CastTo(ent, iface)
            except Exception: return ent
        return ent
    except Exception:
        return ent

# â€”â€” ç»Ÿä¸€é€‰æ‹©å™¨ï¼šall/window/crossing/onscreen
# mode: "all" | "window" | "crossing" | "onscreen"
def ss_select(mode="all", p1=None, p2=None, filter_types=None, filter_data=None, autocast=True, prompt=None):
    _, doc = get_acad_doc()
    ss_name = f"SS_{uuid.uuid4().hex[:8]}"
    # æ¸…ç†æ—§å + æ–°å»º
    try: doc.SelectionSets.Item(ss_name).Delete()
    except Exception: pass
    ss = doc.SelectionSets.Add(ss_name)

    try:
        if mode == "onscreen":
            if prompt:
                try: doc.Utility.Prompt(prompt)
                except Exception: pass
            ss.SelectOnScreen()
        elif mode in ("all","window","crossing"):
            selmode = 5 if mode=="all" else (3 if mode=="window" else 4)
            ss.Select(
                selmode,
                p1 if mode!="all" else None,
                p2 if mode!="all" else None,
                to_vt_int(filter_types) if filter_types else None,
                to_vt_variant(filter_data) if filter_data else None
            )
        else:
            raise ValueError("æœªçŸ¥é€‰æ‹©æ¨¡å¼")

        n = com_retry(lambda: ss.Count)
        items = [com_retry(lambda i=i: ss.Item(i)) for i in range(n)]
        if autocast:
            items = [_maybe_cast(e) for e in items]
        return items
    finally:
        try: ss.Delete()
        except Exception: pass


# ===================== é€‰æ‹©å‡½æ•°ï¼ˆé‡å†™ç‰ˆï¼‰ =====================

# 1) é€‰æ‹©â€œåŒ…å›´æ¡†è¦†ç›–æŒ‡å®šç‚¹â€çš„å¯¹è±¡
def select_entities_through_point(p, tol=0.1):
    """
    æŸ¥æ‰¾æ‰€æœ‰å…¶ BoundingBox è¦†ç›–ç‚¹ p çš„å›¾å…ƒã€‚
    ä¸ä¾èµ–å…¨å±€ mpï¼›è‡ªåŠ¨é‡è¯•ï¼›å¤±è´¥çš„å¯¹è±¡è·³è¿‡ã€‚
    """
    _, doc = get_acad_doc()
    msp = doc.ModelSpace
    px, py, _ = p
    selected = []
    # éå†æ¨¡å‹ç©ºé—´ï¼›å¤§å›¾æ…ç”¨ï¼ˆå¿…è¦æ—¶å¯æ”¹ä¸ºå°çª—å£äº¤å‰é€‰æ‹©ï¼‰
    for ent in msp:
        try:
            p1, p2 = com_retry(lambda: ent.GetBoundingBox())
            x_min, x_max = sorted([p1[0], p2[0]])
            y_min, y_max = sorted([p1[1], p2[1]])
            x_min -= tol; x_max += tol; y_min -= tol; y_max += tol
            if x_min <= px <= x_max and y_min <= py <= y_max:
                selected.append(_maybe_cast(ent))
        except Exception:
            continue
    print(f"[OK] å…±æ‰¾åˆ° {len(selected)} ä¸ªå¯¹è±¡ç»è¿‡ç‚¹ {p}")
    return selected


# 2) æŒ‰å›¾å±‚é€‰æ‹©ï¼ˆæ”¯æŒå•ä¸ªæˆ–åˆ—è¡¨ï¼‰
def select_tuceng(layer_names, max_retries=5, delay=0.5, autocast=True):
    if isinstance(layer_names, str):
        layers = [layer_names]
    else:
        layers = list(layer_names)
    last = None
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(
                mode="all",
                filter_types=[8],                # 8 = Layer
                filter_data=[layers if len(layers)>1 else layers[0]],
                autocast=autocast
            )
            print(f"[OK] ç¬¬ {k} æ¬¡å°è¯•ï¼šé€‰åˆ°å›¾å±‚ {layers} ä¸Š {len(ents)} ä¸ªå¯¹è±¡")
            return ents
        except Exception as e:
            last = e
            print(f"[è­¦å‘Š] ç¬¬ {k} æ¬¡å¤±è´¥ï¼š{e!r}")
            try:
                _, doc = get_acad_doc()
                doc.SendCommand("RE\nZ\nE\n")
            except Exception:
                pass
            time.sleep(delay)
    print(f"[é”™è¯¯] é‡è¯• {max_retries} æ¬¡åä»å¤±è´¥ï¼š{last!r}")
    return []

def stc(layer_names, **kwargs):
    return select_tuceng(layer_names, **kwargs)


# 3) é€‰æ‹©æ‰€æœ‰å—ï¼ˆINSERTï¼‰
def select_kuai(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(
                mode="all",
                filter_types=[0],               # 0 = å®ä½“ç±»å‹
                filter_data=["INSERT"],
                autocast=autocast
            )
            print(f"[OK] select_kuai æˆåŠŸï¼ˆç¬¬ {k} æ¬¡ï¼‰ï¼Œè€—æ—¶ {time.time()-t0:.3f}sï¼Œå…± {len(ents)} ä¸ªå—")
            return ents
        except Exception as e:
            last = e
            print(f"[è­¦å‘Š] select_kuai ç¬¬ {k} æ¬¡å¤±è´¥ï¼š{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"[é”™è¯¯] select_kuai åœ¨ {max_retries} æ¬¡å°è¯•åä»å¤±è´¥ï¼š{last!r}")
    return []


# 4) é€‰æ‹©æ‰€æœ‰ TEXT
def select_text(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["TEXT"], autocast=autocast)
    print("è€—æ—¶ï¼š", time.time()-t0)
    return ents

# 5) é€‰æ‹©æ‰€æœ‰ MTEXT
def select_mtext(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["MTEXT"], autocast=autocast)
    print("è€—æ—¶ï¼š", time.time()-t0)
    return ents

# 6) PUB_TEXT å›¾å±‚ä¸Šçš„â€œå¤©æ­£æ–‡å­—â€åˆ†ç±»ï¼ˆæŒ‰ ObjectNameï¼‰
def select_pub_text_entities():
    LAYER_NAME = "PUB_TEXT"
    ents = select_tuceng(LAYER_NAME, autocast=False)  # å¤©æ­£å¯¹è±¡é€šå¸¸æ˜¯ä»£ç†/ä¸“æœ‰ç±»ï¼Œå…ˆåˆ« Cast
    tdb_texts, tdb_mtexts = [], []
    for ent in ents:
        name = getattr(ent, "ObjectName", None) or getattr(ent, "EntityName", "")
        if name == "TDbText":
            tdb_texts.append(ent)
        elif name == "TDbMText":
            tdb_mtexts.append(ent)
    return tdb_texts, tdb_mtexts

def collect_all_texts():
    """
    åŒæ—¶æ”¶é›†å¤©æ­£ä¸åŸç”Ÿ CAD æ–‡æœ¬ï¼Œå¹¶ç»Ÿä¸€å›¾å±‚ä¸º 'PUB_TEXT'ã€‚
    è¿”å›ï¼štianzheng_texts / tianzheng_mtexts / cad_texts / cad_mtexts
    """
    LAYER_NAME = "PUB_TEXT"
    tz_texts, tz_mtexts = select_pub_text_entities()
    cad_texts  = select_text(autocast=True)
    cad_mtexts = select_mtext(autocast=True)

    for ent in (tz_texts + tz_mtexts + cad_texts + cad_mtexts):
        try:
            ent.Layer = LAYER_NAME
        except Exception:
            pass
    return tz_texts, tz_mtexts, cad_texts, cad_mtexts


# 7) é€‰æ‹© LINE / CIRCLE / ELLIPSE / SPLINE
def select_line(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["LINE"], autocast=autocast)
    print("è€—æ—¶ï¼š", time.time()-t0)
    return ents

def select_circle(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["CIRCLE"], autocast=autocast)
    print("è€—æ—¶ï¼š", time.time()-t0)
    return ents

def select_ellipse(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["ELLIPSE"], autocast=autocast)
    print("è€—æ—¶ï¼š", time.time()-t0)
    return ents

def select_spline(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["SPLINE"], autocast=autocast)
    print("è€—æ—¶ï¼š", time.time()-t0)
    return ents


# 8) ä¼ ç»Ÿå¤šæ®µçº¿ï¼ˆPOLYLINEï¼‰ä¸è½»é‡å¤šæ®µçº¿ï¼ˆLWPOLYLINEï¼‰
def select_polyline_chuantong(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(mode="all", filter_types=[0], filter_data=["POLYLINE"], autocast=autocast)
            print(f"[OK] select_polyline_chuantong æˆåŠŸï¼ˆç¬¬ {k} æ¬¡ï¼‰ï¼Œè€—æ—¶ {time.time()-t0:.3f}sï¼Œå…± {len(ents)} æ¡")
            return ents
        except Exception as e:
            last = e
            print(f"[è­¦å‘Š] select_polyline_chuantong ç¬¬ {k} æ¬¡å¤±è´¥ï¼š{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"[é”™è¯¯] select_polyline_chuantong åœ¨ {max_retries} æ¬¡åä»å¤±è´¥ï¼š{last!r}")
    return []

def select_polyline(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(mode="all", filter_types=[0], filter_data=["LWPOLYLINE"], autocast=autocast)
            print(f"[OK] select_polyline æˆåŠŸï¼ˆç¬¬ {k} æ¬¡ï¼‰ï¼Œè€—æ—¶ {time.time()-t0:.3f}sï¼Œå…± {len(ents)} æ¡")
            return ents
        except Exception as e:
            last = e
            print(f"[è­¦å‘Š] select_polyline ç¬¬ {k} æ¬¡å¤±è´¥ï¼š{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"[é”™è¯¯] select_polyline åœ¨ {max_retries} æ¬¡åä»å¤±è´¥ï¼š{last!r}")
    return []


#&&% ##  6  å±å¹•é€‰æ‹©





# --- æ–°ç‰ˆé€‰æ‹©ï¼šä¸ä¾èµ–å…¨å±€ docï¼›ä¸ for-inï¼›è‡ªåŠ¨æ¸…ç†ï¼›å¯é€‰è‡ªåŠ¨ CastTo ---
def pmxz(prompt="\nè¯·åœ¨å±å¹•æ‹¾å–å›¾å…ƒï¼Œä»¥Enteré”®ç»“æŸï¼š", autocast=True):
    """
    äº¤äº’æ‹¾å–å¯¹è±¡ï¼ˆæ˜¾å¼ on-screen é€‰æ‹©ï¼‰ï¼Œè¿”å›å®ä½“åˆ—è¡¨ã€‚
    å†…éƒ¨ç›´æ¥å¤ç”¨ç»Ÿä¸€å†…æ ¸ ss_selectï¼Œä¿æŒä¸å…¶å®ƒé€‰æ‹©å‡½æ•°ä¸€è‡´çš„è¡Œä¸ºã€‚
    """
    return ss_select(
        mode="onscreen",
        p1=None, p2=None,
        filter_types=None,     # å¦‚éœ€é™åˆ¶ç±»å‹ï¼Œå¯ä¼  [0] å¹¶é…åˆ filter_data=["LINE"] ç­‰
        filter_data=None,
        autocast=autocast,
        prompt=prompt
    )

"""
å¦‚æœä½ æƒ³â€œåªå…è®¸é€‰æŸç±»å¯¹è±¡â€ï¼ˆæ¯”å¦‚ç›´çº¿ï¼‰ï¼Œå°±ä¼ è¿‡æ»¤å‚æ•°å³å¯ï¼š

def pmxz_line(prompt="é€‰ä¸€æ ¹ç›´çº¿ï¼š"):
    return ss_select("onscreen", filter_types=[0], filter_data=["LINE"], autocast=True, prompt=prompt)


"""


# 8) éšæ€§ â†’ æ˜¾æ€§é€‰æ‹©ï¼ˆå®‰å…¨ç‰ˆï¼‰
def yin_to_xian_xuanze(LB, wait_s=0.6):
    """
    å°† COM é€‰ä¸­çš„å¯¹è±¡åˆ—è¡¨ï¼ˆLBï¼‰è½¬æ¢ä¸ºå‘½ä»¤çª—å£ä¸­çš„è“è‰²â€œé«˜äº®é€‰ä¸­â€çŠ¶æ€ã€‚
    å®ç°ï¼šStartUndoMark â†’ æ‰¹é‡ Delete â†’ U æ’¤é”€ â†’ SELECT Pã€‚
    """
    _, doc = get_acad_doc()
    deleted = 0
    com_retry(lambda: doc.StartUndoMark())
    try:
        for x in LB:
            try:
                com_retry(lambda x=x: x.Delete())
                deleted += 1
            except Exception:
                pass
        print(f"ğŸ§¹ å°è¯•åˆ é™¤ {len(LB)} ä¸ªå¯¹è±¡ï¼ŒæˆåŠŸåˆ é™¤ {deleted} ä¸ªã€‚")

        # æ’¤é”€åˆ é™¤ï¼ˆæ¢å¤å¹¶å»ºç«‹ _Pï¼‰
        doc.SendCommand("_U\n\n")
        time.sleep(wait_s)

        # ä½¿ç”¨ Pï¼ˆprevious selectionï¼‰å»ºç«‹æ˜¾æ€§é€‰æ‹©
        doc.SendCommand("_.SELECT\nP\n\n")
        time.sleep(0.5)
        print("[OK] å·²å°†éšæ€§é€‰æ‹©è½¬æ¢ä¸ºæ˜¾æ€§ï¼ˆè“è‰²é«˜äº®ï¼‰")
    finally:
        try: com_retry(lambda: doc.EndUndoMark())
        except Exception: pass


# 9) æ˜¾æ€§ â†’ éšæ€§é€‰æ‹©ï¼ˆä¿ç•™ä½ åŸæ³•ï¼Œä½†æ”¶æ•›æˆå‡½æ•°ï¼‰
def xian_to_yin_xuanze():
    _, doc = get_acad_doc()
    # åŸºäºä½ çš„æµç¨‹ï¼šCopyBase â†’ å…¨é€‰ â†’ Erase â†’ PasteClip â†’ å– SelectionSetAll
    doc.SendCommand("_COPYBASE\n0,0,0\n\n"); time.sleep(3)
    doc.SendCommand("_AI_SELALL\n\n");       time.sleep(1.5)
    doc.SendCommand("_ERASE\n\n");           time.sleep(0.8)
    doc.SendCommand("_PASTECLIP\n0,0,0\n\n");time.sleep(2.0)

    # å–éšæ€§é€‰æ‹©ï¼ˆå…¨éƒ¨ï¼‰
    try:
        doc.SelectionSets.Item("MySelectionSet").Delete()
    except Exception:
        pass
    ss = doc.SelectionSets.Add("MySelectionSet")
    ss.Select(5)  # å…¨é€‰
    lb = [ss.Item(i) for i in range(ss.Count)]
    try: ss.Delete()
    except Exception: pass
    print(f"[OK] å…±è·å–å¯¹è±¡ {len(lb)} ä¸ªï¼Œè½¬æ¢ä¸ºå¯æ“ä½œé€‰æ‹©åˆ—è¡¨")
    return lb


# 10) éšæ€§çª—å£é€‰æ‹©ï¼ˆä¼˜å…ˆ SelectionSet çª—å£ï¼›å¤±è´¥å…œåº•éå†åŒ…å›´ç›’ï¼‰

# è¾…åŠ©å‡½æ•°ï¼šè§„èŒƒåŒ–çŸ©å½¢åæ ‡
def normalize_rect(x1, y1, x2, y2):
    """è§„èŒƒåŒ–çŸ©å½¢åæ ‡ï¼Œè¿”å›å·¦ä¸‹è§’å’Œå³ä¸Šè§’"""
    x_lo, x_hi = (x1, x2) if x1 < x2 else (x2, x1)
    y_lo, y_hi = (y1, y2) if y1 < y2 else (y2, y1)
    return (x_lo, y_lo), (x_hi, y_hi)

# è¾…åŠ©å‡½æ•°ï¼šåˆ›å»º3Dç‚¹
def pt3(x, y, z=0):
    """åˆ›å»º3Dç‚¹çš„VARIANT"""
    import win32com.client
    import pythoncom
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [x, y, z])

def select_objects_in_window_area(x1, y1, x2, y2, max_retry=5):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    doc.SetVariable("TILEMODE", 1)

    # ä¸´æ—¶é€‰æ‹©é›†
    try: doc.SelectionSets.Item("MyWindowSelection").Delete()
    except Exception: pass
    ss = doc.SelectionSets.Add("MyWindowSelection")

    entities = []
    # â‘  é‡è¯•çª—å£é€‰æ‹©
    p1, p2 = pt3(x_lo,y_lo), pt3(x_hi,y_hi)
    for _ in range(max_retry):
        try:
            ss.Clear()
            ss.Select(1, p1, p2)  # 1 = Window
            if ss.Count > 0: break
        except Exception:
            pass
        time.sleep(0.25)

    if ss.Count > 0:
        for i in range(ss.Count):
            try: entities.append(_maybe_cast(ss.Item(i)))
            except Exception: continue
        try: ss.Delete()
        except Exception: pass
        print(f"[OK] çª—å£é€‰æ‹©æˆåŠŸï¼Œå…± {len(entities)} ä¸ªå¯¹è±¡ã€‚")
        return entities

    # â‘¡ å…œåº•ï¼šéå† ModelSpace çš„åŒ…å›´ç›’è§’ç‚¹
    try:
        msp = doc.ModelSpace
        def in_win(pt): return x_lo <= pt[0] <= x_hi and y_lo <= pt[1] <= y_hi
        for ent in msp:
            try:
                a,b = com_retry(lambda: ent.GetBoundingBox())
                if in_win(a) or in_win(b):
                    entities.append(_maybe_cast(ent))
            except Exception:
                continue
        print(f"[FALLBACK] éå†æ¨¡å‹ç©ºé—´å¾—åˆ° {len(entities)} ä¸ªå¯¹è±¡ã€‚")
    except Exception as e:
        print(f"[é”™è¯¯] éå†æ¨¡å‹ç©ºé—´å¤±è´¥: {e}")

    try: ss.Delete()
    except Exception: pass
    return entities


# 11) éšæ˜¾ç»“åˆçš„åŒºåŸŸé€‰æ‹©ï¼ˆé«˜äº®é€‰æ‹©å¹¶è¿”å› PickfirstSelectionSetï¼‰
def select_entities_in_window(x1, y1, x2, y2, ty: float = 1.0, select_mode: str = "_W"):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)

    # æ¸…ç©º Pickfirst
    try: doc.Pickenabled = False
    except Exception: pass
    try: doc.PickfirstSelectionSet.Clear()
    except Exception: pass

    # Zoomï¼ˆåŠ  20% ç¼“å†²ï¼‰
    buf = 0.20 * ((x_hi-x_lo) + (y_hi-y_lo)) / 2.0
    doc.SendCommand(f"_.ZOOM\n_W\n{x_lo-buf},{y_lo-buf}\n{x_hi+buf},{y_hi+buf}\n")
    time.sleep(ty)

    # Selectï¼ˆæ˜¾æ€§ï¼Œè“è‰²é«˜äº®ï¼‰
    doc.SendCommand(f"_.SELECT\n{select_mode}\n{x_lo},{y_lo}\n{x_hi},{y_hi}\n\n")
    time.sleep(ty/2)

    selset = doc.PickfirstSelectionSet
    com_list = [ent for ent in selset]
    try: selset.Clear()
    except Exception: pass
    return com_list


# 12) æ˜¾æ€§åŒºåŸŸé€‰æ‹©ï¼ˆè“è‰²é«˜äº®ï¼‰
def highlight_entities_in_window(x1, y1, x2, y2):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    h = 0.1 * ((abs(x_hi-x_lo)+abs(y_hi-y_lo))/2.0)

    # ZOOM
    doc.SendCommand(f"_.ZOOM\n_W\n{x_lo-h},{y_lo-h}\n{x_hi+h},{y_hi+h}\n"); time.sleep(1)

    # æ˜¾æ€§ SELECT
    doc.SendCommand(f"_.SELECT\n_W\n{x_lo},{y_lo}\n{x_hi},{y_hi}\n\n"); time.sleep(0.5)
    print(f"[OK] å·²é«˜äº®é€‰æ‹©åŒºåŸŸ ({x_lo},{y_lo}) ~ ({x_hi},{y_hi}) çš„å¯¹è±¡")


# ====== è¾…åŠ©ï¼šæŒ‰åç§»é‡æ‰©å¼ åŒ…å›´æ¡†ï¼ˆp1,p2 éƒ½æ˜¯ (x,y,?) åºåˆ—ï¼‰======
def expand_rectangle(p1, p2, offset):
    x1, y1 = float(p1[0]), float(p1[1])
    x2, y2 = float(p2[0]), float(p2[1])
    x_lo, x_hi = sorted((x1, x2))
    y_lo, y_hi = sorted((y1, y2))
    return (x_lo - offset, y_lo - offset), (x_hi + offset, y_hi + offset)
# 1) éš”ç¦»å¯¹è±¡ï¼ˆæ¨¡å‹ç©ºé—´åŒºåŸŸï¼‰
def isolate_modelspace_area(x1, y1, x2, y2):
    """
    åœ¨æ¨¡å‹ç©ºé—´ï¼Œå°†çª—å£åŒºåŸŸå†…å¯¹è±¡éš”ç¦»æ˜¾ç¤ºï¼š
      - éšæ€§é€‰åŒºï¼ˆCOM SelectionSet/åŒ…å›´ç›’å…œåº•ï¼‰â†’
      - è½¬æ˜¾æ€§ï¼ˆè“è‰²é«˜äº®ï¼Œä¾› SendCommand è¯†åˆ«ï¼‰â†’
      - å‘é€ _IsolateObjects
    """
    _, doc = get_acad_doc()
    # ç¡®ä¿æ¨¡å‹ç©ºé—´
    try: doc.SetVariable("TILEMODE", 1)
    except Exception: pass

    LB = select_objects_in_window_area(x1, y1, x2, y2)
    if not LB:
        print("[é”™è¯¯] æ²¡æœ‰é€‰æ‹©åˆ°å¯¹è±¡ï¼Œç»ˆæ­¢æ“ä½œ")
        return

    print(f"[OK] é€‰ä¸­å¯¹è±¡ {len(LB)} ä¸ªï¼Œå‡†å¤‡éš”ç¦»")
    yin_to_xian_xuanze(LB, wait_s=0.6)  # å·²å®ç°ï¼šéšæ€§â†’æ˜¾æ€§
    time.sleep(0.4)
    doc.SendCommand("_.IsolateObjects\n")
    print("ğŸ¯ å·²å‘é€éš”ç¦»å‘½ä»¤")
# 2) å›¾çº¸ç©ºé—´çš„åŒºåŸŸé€‰æ‹©ï¼ˆéšæ€§åˆ—è¡¨è¿”å›ï¼‰
def select_paperspace_objects_in_window(x1, y1, x2, y2):
    """
    åœ¨å›¾çº¸ç©ºé—´å†…ï¼Œé€‰æ‹©çª—å£ (x1,y1)-(x2,y2) è¦†ç›–çš„å¯¹è±¡ï¼›ä¼˜å…ˆç”¨ SelectionSet çª—å£ï¼Œ
    è‹¥ä¸ºç©ºåˆ™éå† PaperSpace çš„åŒ…å›´ç›’å…œåº•ã€‚è¿”å› COM å¯¹è±¡åˆ—è¡¨ï¼ˆå·²è‡ªåŠ¨ Castï¼‰ã€‚
    """
    _, doc = get_acad_doc()
    try: doc.SetVariable("TILEMODE", 0)  # åˆ‡åˆ°å›¾çº¸ç©ºé—´
    except Exception: pass

    (x_lo, y_lo), (x_hi, y_hi) = normalize_rect(x1, y1, x2, y2)
    # ä¼˜å…ˆï¼šçª—å£é€‰æ‹©
    try:
        items = ss_select(
            mode="window",
            p1=pt3(x_lo, y_lo),
            p2=pt3(x_hi, y_hi),
            filter_types=None,
            filter_data=None,
            autocast=True
        )
        if items:
            print(f"[OK] å›¾çº¸ç©ºé—´çª—å£é€‰æ‹© {len(items)} ä¸ªå¯¹è±¡ã€‚")
            return items
    except Exception:
        pass

    # å…œåº•ï¼šéå† PaperSpace åŒ…å›´ç›’
    selected = []
    ps = doc.PaperSpace
    def intersects(a1, a2, b1, b2):
        return (a1 <= b2) and (a2 >= b1)
    for i in range(ps.Count):
        try:
            ent = com_retry(lambda i=i: ps.Item(i))
            p1, p2 = com_retry(lambda: ent.GetBoundingBox())
            ex1, ey1 = p1[0], p1[1]
            ex2, ey2 = p2[0], p2[1]
            if intersects(ex1, ex2, x_lo, x_hi) and intersects(ey1, ey2, y_lo, y_hi):
                selected.append(_maybe_cast(ent))
        except Exception:
            continue
    print(f"[FALLBACK] å›¾çº¸ç©ºé—´åŒ…å›´ç›’éå†å¾—åˆ° {len(selected)} ä¸ªå¯¹è±¡ã€‚")
    return selected
# 3) å°†â€œå•ä¸ªå¯¹è±¡â€æ˜¾æ€§é«˜äº®ï¼ˆè“è‰²å¤¹ç‚¹ï¼‰
def highlight_entity_by_bbox(entity):
    """
    å¯¹æŒ‡å®š COM å¯¹è±¡ entityï¼Œé€šè¿‡æ‰©å¤§å…¶ bounding box æ¥è¿›è¡Œæ˜¾æ€§é«˜äº®é€‰ä¸­ã€‚
    ä¼šè‡ªåŠ¨ Zoom åˆ°å¯¹è±¡é™„è¿‘ï¼Œç„¶åç”¨çª—å£é€‰æ‹©æŠŠå®ƒé«˜äº®å‡ºæ¥ã€‚

    æ³¨æ„ï¼šå¤©æ­£å¯¹è±¡ä¸æ”¯æŒ GetBoundingBox()ï¼Œä¼šæŠ›å‡ºå¼‚å¸¸
    """
    try:
        # ç”¨å¯¹è±¡è‡ªèº«çš„ App/Docï¼Œé¿å…å¤šæ–‡æ¡£åˆ‡æ¢å¹²æ‰°
        doc = com_retry(lambda: entity.Application.ActiveDocument)
        # GetBoundingBox è¿”å› (MinPoint, MaxPoint) å…ƒç»„
        p1, p2 = com_retry(lambda: entity.GetBoundingBox())

        x1, y1 = float(p1[0]), float(p1[1])
        x2, y2 = float(p2[0]), float(p2[1])

        # Zoom åˆ°å¯¹è±¡ + 10% ç¼“å†²
        h = 0.1 * ((abs(x1 - x2) + abs(y1 - y2)) / 2.0)
        doc.SendCommand(f"_.ZOOM\n_W\n{x1-h},{y1-h}\n{x2+h},{y2+h}\n")
        time.sleep(1.0)

        # è®¡ç®—ç•¥æ‰©å¼ çš„é€‰æ‹©çª—å£ï¼ˆå¢™ä½“ç‰¹æ®Šåç§»ï¼‰
        x_len = abs(x2 - x1); y_len = abs(y2 - y1)
        max_len = max(x_len, y_len)
        offset = 130 if (getattr(entity, "ObjectName", "") == "TDbWall") else (max_len * 0.1)
        # expand_rectangle è¿”å› 3D åæ ‡ï¼Œå–å‰ä¸¤ä¸ªå…ƒç´ 
        pt1, pt2 = expand_rectangle(p1, p2, offset)
        X1, Y1 = pt1[0], pt1[1]
        X2, Y2 = pt2[0], pt2[1]

        # æ˜¾æ€§çª—å£é€‰æ‹©
        highlight_entities_in_window(X1, Y1, X2, Y2)
        print("[OK] å·²é«˜äº®ç›®æ ‡å¯¹è±¡")
    except Exception as e:
        print("[é”™è¯¯] æ— æ³•é«˜äº®è¯¥å¯¹è±¡:", e)
        raise  # é‡æ–°æŠ›å‡ºå¼‚å¸¸ï¼Œè®©å¤–å±‚ try-except æ•è·
# 4) åˆ©ç”¨ Visible è¿›è¡Œâ€œæ˜¾é€‰â†’éšé€‰â€è½¬æ¢çš„ä¾¿æ·å°è£…ï¼ˆä¿æŒä½ çš„æµç¨‹ï¼‰
def select_visible(x1, y1, x2, y2):
    """
    å…ˆæ˜¾æ€§é«˜äº®çª—å£å¯¹è±¡ï¼Œå†è°ƒç”¨ xian_to_yin_xuanze() è½¬æ¢ä¸ºå¯æ“ä½œçš„éšæ€§åˆ—è¡¨ã€‚
    è¿”å›éšæ€§åˆ—è¡¨ï¼ˆCOM å¯¹è±¡é›†åˆï¼‰ã€‚
    """
    highlight_entities_in_window(x1, y1, x2, y2)
    time.sleep(1.0)
    return xian_to_yin_xuanze()
# 5) æ˜¾ç¤ºæŒ‡å®šç©ºé—´ä¸­æ‰€æœ‰éšè—å¯¹è±¡ï¼ˆå¯è¿‡æ»¤ç±»å‹ï¼Œå¯é€‰æ‹©é«˜äº®ï¼‰
def unhide_all(space=None, filter_names=None, highlight=False):
    """
    æ˜¾ç¤º space ä¸­æ‰€æœ‰ Visible=False çš„å¯¹è±¡ã€‚
    å‚æ•°ï¼š
      - space: å¯ä¼  doc.ModelSpace / doc.PaperSpace / None / "model" / "paper"
      - filter_names: ä»…æ¢å¤è¿™äº› ObjectName çš„å¯¹è±¡ï¼ˆä¾‹å¦‚ ["AcDbPolyline", "AcDbBlockReference"]ï¼‰
      - highlight: True åˆ™å¯¹æ¢å¤çš„å¯¹è±¡è°ƒç”¨ .Highlight(True)
    è¿”å›ï¼šrevealed åˆ—è¡¨ï¼ˆå·²è‡ªåŠ¨ Castï¼‰
    """
    _, doc = get_acad_doc()

    # è§£æç©ºé—´å‚æ•°
    if space is None:
        target = doc.ModelSpace
    elif isinstance(space, str):
        s = space.strip().lower()
        if s in ("model", "m", "modelspace"):
            target = doc.ModelSpace
        elif s in ("paper", "p", "paperspace", "layout"):
            target = doc.PaperSpace
        else:
            target = doc.ModelSpace
    else:
        # ä¼ å…¥çš„æ˜¯ COM é›†åˆï¼ˆæœ‰ Count/Itemï¼‰
        target = space

    revealed = []
    cnt = com_retry(lambda: target.Count)
    for i in range(cnt):
        try:
            obj = com_retry(lambda i=i: target.Item(i))
            # æœ‰äº›å¯¹è±¡æ²¡æœ‰ Visibleï¼Œè·³è¿‡
            if not hasattr(obj, "Visible"):
                continue
            if obj.Visible:
                continue
            name = getattr(obj, "ObjectName", "")
            if (filter_names is None) or (name in filter_names):
                obj.Visible = True  # è®¾ä¸ºå¯è§
                if highlight and hasattr(obj, "Highlight"):
                    try: obj.Highlight(True)
                    except Exception: pass
                revealed.append(_maybe_cast(obj))
                print(f"[OK] æ˜¾ç¤ºå¯¹è±¡ï¼š{name} | Handle={getattr(obj, 'Handle', '?')}")
        except Exception as e:
            print(f"[è­¦å‘Š]ï¸ è·³è¿‡ç´¢å¼• {i}ï¼š{e}")

    print(f"\nğŸ“Š å…±æ˜¾ç¤º {len(revealed)} ä¸ªéšè—å¯¹è±¡ã€‚")
    return revealed


#&&% #  ç»„
"""

åˆ›å»ºç»„
group = doc.Groups.Add("mygroup")
LB=pmxz()
è¯·åœ¨å±å¹•æ‹¾å–å›¾å…ƒï¼Œä»¥Enteré”®ç»“æŸ

group.AppendItems(vtobj([LB[0], LB[1], LB[2]]))


ä»ç»„åè·å–ç»„åˆç»„ä¸­å¯¹è±¡

group = doc.Groups.Item("G001")

entities = [group.Item(i) for i in range(group.Count)]  # éå†ç»„å†…å¯¹è±¡

entities[0].Handle
'2C3'

entities[0].Move(vtpnt(0,0,0),vtpnt(0,10000,0))


è§£é™¤ç»„
group.Delete()


group1 = doc.Groups.Add("mygroup")
group1.AppendItems(vtobj([LB[0], LB[1], LB[2]]))
group2 = doc.Groups.Add("mygroupA")
group2.AppendItems(vtobj([LB[3], LB[4], LB[5],LB[6]]))
group3 = doc.Groups.Add("mygroupB")
group3.AppendItems(vtobj([group1,group2]))éæ³•æ“ä½œ
group3.AppendItems(vtobj([LB[0],LB[1],LB[2],LB[3], LB[4], LB[5],LB[6]]))


# æå– group1 å’Œ group2 çš„æ‰€æœ‰æˆå‘˜
group1_entities = [group1.Item(i) for i in range(group1.Count)]
group2_entities = [group2.Item(i) for i in range(group2.Count)]

# åˆå¹¶æˆæ–°çš„åˆ—è¡¨
all_entities = group1_entities + group2_entities

# åˆ›å»º group3
group3 = doc.Groups.Add("mygroupB")
group3.AppendItems(vtobj(all_entities))

 get_boundingbox_from_objects(objs)
"""

# å»ºç«‹å…¨éƒ¨åˆ—è¡¨comå¯¹è±¡çš„æœ€å°è¾¹ç•Œæ¡†

def get_boundingbox_from_objects(objs):#ä»åˆ—è¡¨comå¯¹è±¡å»ºç«‹æœ€å°è¾¹ç•Œæ¡†
    """
    ä»ä¸€ç»„å›¾å½¢å¯¹è±¡ï¼ˆå¦‚ LBï¼‰ä¸­è·å–æ•´ä½“åŒ…å›´ç›’
    è¿”å›å€¼ï¼š(min_x, min_y, min_z), (max_x, max_y, max_z)
    """
    min_point, max_point = None, None

    for obj in objs:
        try:
            min_pt, max_pt = obj.GetBoundingBox()
            if min_point is None:
                min_point, max_point = list(min_pt), list(max_pt)
            else:
                min_point = [min(min_point[i], min_pt[i]) for i in range(3)]
                max_point = [max(max_point[i], max_pt[i]) for i in range(3)]
        except Exception as e:
            print(f"è·³è¿‡æ— æ³•è·å–è¾¹ç•Œçš„å¯¹è±¡: {obj.ObjectName}")
            continue

    return tuple(min_point), tuple(max_point)
        

# å»ºç«‹ç»„çš„æœ€å°è¾¹ç•Œæ¡†

def chuangjian_zu(group_name):

    group = doc.Groups.Add(group_name)

    return group

def nametogroup(group_name):#ä»ç»„åè·å–å®ä½“comç»„å¯¹è±¡
    group_obj = doc.Groups.Item(group_name)

    return group_obj

##è·å–æ‰€æœ‰ç»„

def get_all_group_names():
    """
    è·å–å½“å‰ DWG æ–‡æ¡£ä¸­æ‰€æœ‰ç»„çš„åç§°åˆ—è¡¨ã€‚
    
    è¿”å›:
      List[str] â€” åŒ…å«æ‰€æœ‰ç»„åç§°çš„åˆ—è¡¨
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    groups = doc.Groups
    return [groups.Item(i).Name for i in range(groups.Count)]

def get_all_groups():
    """
    è·å–å½“å‰ DWG æ–‡æ¡£ä¸­æ‰€æœ‰ç»„çš„ COM å¯¹è±¡åˆ—è¡¨åŠå…¶åç§°ã€‚
    
    è¿”å›:
      List[Tuple[str, COMObject]] â€” æ¯é¡¹ä¸º (ç»„åç§°, ç»„å¯¹è±¡)
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    groups = doc.Groups
    result = []
    for i in range(groups.Count):
        grp = groups.Item(i)
        result.append((grp.Name, grp))
    return result



#å°†å¤šä¸ªcomå¯¹è±¡å¯¹è±¡åŠ å…¥åä¸ºgroup_nameçš„ç»„
def add_objects_to_group(group_name, obj_list):

    """
    å°† obj_list ä¸­çš„æ‰€æœ‰å›¾å½¢å¯¹è±¡åŠ å…¥åä¸º group_name çš„ç»„ä¸­
    å¦‚æœç»„å·²å­˜åœ¨ï¼Œä½¿ç”¨åŸç»„ï¼›å¦åˆ™æ–°å»º
    è¿”å›ï¼šGroup å¯¹è±¡
    """
    groups = doc.Groups
    try:
        group = groups.Item(group_name)
    except:
        group = groups.Add(group_name)

    group.AppendItems(vtlist(obj_list))
    return group


#å°†å•ç‹¬comå¯¹è±¡å¯¹è±¡åŠ å…¥åä¸ºgroup_nameçš„ç»„ä¸­

def add_object_to_group(group_name, obj):
    """
    å°†å•ä¸ªå›¾å½¢å¯¹è±¡ obj åŠ å…¥åä¸º group_name çš„ç»„ä¸­ã€‚
    å¦‚æœç»„å·²å­˜åœ¨ï¼Œåˆ™ä½¿ç”¨è¯¥ç»„ï¼›å¦åˆ™æ–°å»ºä¸€ä¸ªæ–°ç»„ã€‚
    
    å‚æ•°ï¼š
      group_name (str)ï¼šç»„åç§°
      objï¼šè¦åŠ å…¥ç»„çš„ COM å¯¹è±¡ï¼ˆå¦‚å¤šæ®µçº¿ã€çº¿æ®µã€å—å‚ç…§ç­‰ï¼‰
    
    è¿”å›ï¼š
      COM Group å¯¹è±¡
    """
    groups = doc.Groups
    try:
        # å°è¯•è·å–å·²å­˜åœ¨çš„ç»„
        group = groups.Item(group_name)
    except Exception:
        # ä¸å­˜åœ¨åˆ™æ–°å»º
        group = groups.Add(group_name)
    
    # vtlist å·¥å…·å°† Python list è½¬æˆ VBA å¯æ¥å—çš„ SAFEARRAY
    group.AppendItems(vtlist([obj]))
    return group

#å°†å•ç‹¬comå¯¹è±¡å¯¹è±¡ç§»å‡ºåä¸ºgroup_nameçš„ç»„
def remove_object_from_group(group_name, obj):
    """
    å°†å•ä¸ª COM å¯¹è±¡ obj ä»åä¸º group_name çš„ç»„ä¸­ç§»å‡ºã€‚
    å¦‚æœç»„ä¸å­˜åœ¨æˆ–å¯¹è±¡ä¸åœ¨ç»„ä¸­ï¼Œåˆ™ä¼šæ‰“å°é”™è¯¯ä¿¡æ¯ä½†ä¸æŠ›å¼‚å¸¸ã€‚
    
    å‚æ•°:
      group_name: ç»„åï¼ˆå­—ç¬¦ä¸²ï¼‰
      obj:         è¦ç§»å‡ºçš„ COM å¯¹è±¡
    
    è¿”å›:
      å¦‚æœç»„å­˜åœ¨ï¼Œè¿”å›è¯¥ Group å¯¹è±¡ï¼›å¦åˆ™è¿”å› Noneã€‚
    """
    try:
        group = doc.Groups.Item(group_name)
    except Exception:
        print(f"[é”™è¯¯] ç»„ '{group_name}' ä¸å­˜åœ¨")
        return None

    # æŠŠå•ä¸ªå¯¹è±¡åŒ…è£…æˆé•¿åº¦ä¸º1çš„ COM SafeArray
    variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, [obj])
    try:
        group.RemoveItems(variant)
        print(f"[OK] å·²ä»ç»„ '{group_name}' ä¸­ç§»é™¤å¯¹è±¡ {obj.Handle}")
    except Exception as e:
        print(f"[é”™è¯¯] ä»ç»„ '{group_name}' ç§»é™¤å¯¹è±¡å¤±è´¥ï¼š{e}")

    return group

#å°†å¤šä¸ªcomå¯¹è±¡å¯¹è±¡ç§»å‡ºåä¸ºgroup_nameçš„ç»„
def remove_objects_from_group(group_name, obj_list):
    """
    å°† obj_list ä¸­çš„æ‰€æœ‰å›¾å½¢å¯¹è±¡ä»åä¸º group_name çš„ç»„ä¸­ç§»å‡ºã€‚
    å¦‚æœç»„ä¸å­˜åœ¨ï¼Œä¼šæ‰“å°æç¤ºå¹¶è¿”å› Noneï¼›å¦åˆ™è¿”å›è¯¥ç»„å¯¹è±¡ã€‚
    
    :param group_name: ç»„åç§°
    :param obj_list: è¦ç§»é™¤çš„ COM å¯¹è±¡åˆ—è¡¨
    :return: Group å¯¹è±¡ æˆ– None
    """
    groups = doc.Groups
    try:
        group = groups.Item(group_name)
    except Exception:
        print(f"ç»„ '{group_name}' ä¸å­˜åœ¨ï¼Œæ— æ³•ç§»é™¤å¯¹è±¡ã€‚")
        return None

    # æŠŠ Python åˆ—è¡¨åŒ…è£…æˆ VARIANT SafeArrayï¼ŒVT_DISPATCH è¡¨ç¤ºå¯¹è±¡ç±»å‹
    arr = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj_list)
    try:
        group.RemoveItems(arr)
        print(f"å·²ä»ç»„ '{group_name}' ä¸­ç§»é™¤ {len(obj_list)} ä¸ªå¯¹è±¡ã€‚")
    except Exception as e:
        print(f"ç§»é™¤å¯¹è±¡æ—¶å‘ç”Ÿé”™è¯¯ï¼š{e}")
    return group



#&&% ä»åä¸ºgroup_nameçš„ç»„è·å–å†…éƒ¨åŒ…å«çš„å®ä½“å¯¹è±¡


def get_com_from_groupname(group_name):
    """
    æ ¹æ®ç»„åè·å–å¯¹åº”å®ä½“åˆ—è¡¨ã€‚
    - è‹¥ç»„ä¸å­˜åœ¨ã€æˆ–ç»„ä¸­æ— å®ä½“ï¼Œå‡è¿”å›ç©ºåˆ—è¡¨ï¼Œä¸æŠ›å¼‚å¸¸ã€‚
    """

    try:
        group = nametogroup(group_name)
    except Exception:
        # nametogroup æœ¬èº«å¤±è´¥ï¼ˆç»„ä¸å­˜åœ¨ç­‰ï¼‰
        return []

    if not group:
        # group ä¸º None æˆ–ç©ºï¼Œä¹Ÿç›´æ¥è¿”å›ç©ºåˆ—è¡¨
        return []

    entities = [group.Item(i) for i in range(group.Count)] #ä»comç»„ä¸­è·å–å…¨éƒ¨å¯¹è±¡
    
    return entities

#ä»åä¸ºgroup_nameçš„ç»„è¿”å›æŒ‰ç±»å‹åˆ†ç±»çš„å­—å…¸
def get_com_from_groupname_by_type(group_name):
    """
    æ ¹æ®ç»„åè·å–å¯¹åº”å®ä½“ï¼Œå¹¶æŒ‰ç±»å‹ååˆ†ç±»è¿”å›ã€‚

    :param group_name: ç»„åç§°
    :return: dictï¼Œé”®ä¸ºå®ä½“ç±»å‹åï¼ˆObjectNameï¼‰ï¼Œå€¼ä¸ºè¯¥ç±»å‹çš„å®ä½“åˆ—è¡¨
    """
    # nametogroup æ˜¯ä½ å·²æœ‰çš„â€œç»„åâ†’Group å¯¹è±¡â€å‡½æ•°
    group = nametogroup(group_name)
    if group is None:
        print(f"ç»„ '{group_name}' ä¸å­˜åœ¨")
        return {}

    by_type = {}
    # Group.Count æ˜¯å®ä½“æ•°é‡ï¼ŒItem(i) å–å‡ºç¬¬ i ä¸ªå®ä½“
    for i in range(group.Count):
        ent = group.Item(i)
        # AutoCAD COM å¯¹è±¡ä¸€èˆ¬æœ‰ ObjectName å±æ€§
        typ = getattr(ent, "ObjectName", None) or getattr(ent, "EntityName", "Unknown")
        by_type.setdefault(typ, []).append(ent)

    # æ‰“å°ä¸€ä¸‹å„ç±»å‹æ•°é‡ï¼Œæ–¹ä¾¿è°ƒè¯•
    for typ, lst in by_type.items():
        print(f"  ç±»å‹ {typ} ï¼š{len(lst)} ä¸ªå®ä½“")

    return by_type

#ä»åä¸ºgroup_nameçš„ç»„è¿”å›æŒ‰ç±»å‹åˆ†ç±»çš„å­—å…¸ï¼Œä¸”ç±»å‹æŒ‰å„è‡ªä½ç½®æå–å‡½æ•°æ’å¥½åº
def get_group_entities_sorted(group_name, type_extractors, cha_Y=0.5):
    """
    ä»ç»„ä¸­æŒ‰ç±»å‹è·å–å®ä½“ï¼Œå¹¶å¯¹æŒ‡å®šç±»å‹æŒ‰åæ ‡æ’åºã€‚

    :param group_name: strï¼Œç»„å
    :param type_extractors: dictï¼Œ{ type_name: extract_func }ï¼Œ
           extract_func(ent) è¿”å› (x, y) åæ ‡ï¼Œç”¨äºæ’åºã€‚
    :param cha_Y: floatï¼ŒåŒä¸€è¡Œ Y æ–¹å‘çš„å®¹å·®
    :return: dictï¼Œ{ type_name: [ent, ...] }ï¼Œå·²æ’åºæˆ–åŸåº

    å¯¹å¤©æ­£å•è¡Œå¤šè¡Œæ–‡å­—å¯ä»¥ä½¿ç”¨GetBoundingBoxè·å–çš„å·¦ä¸‹è§’ç‚¹
    
    """
    def get_cadtext_pos(ent):
        # CADå•è¡Œã€å¤šè¡Œæ–‡å­—ï¼šPosition å±æ€§è¿”å› (x, y, z)
        return float(ent.InsertionPoint[0]), float(ent.InsertionPoint[1])


    # å…ˆæŒ‰ç±»å‹è·å–å…¨éƒ¨å®ä½“
    by_type = get_com_from_groupname_by_type(group_name)
    sorted_by_type = {}

    for typ, ents in by_type.items():
        if typ in type_extractors:
            extract_func = type_extractors[typ]
            # æ’åº
            sorted_list = sort_entities_by_position(ents, extract_func, cha_Y=cha_Y)
            sorted_by_type[typ] = sorted_list
            print(f"Type '{typ}' sorted with {len(sorted_list)} entities")
        else:
            # ä¿æŒåŸåº
            sorted_by_type[typ] = list(ents)
            print(f"Type '{typ}' left unsorted ({len(ents)} entities)")

    return sorted_by_type


#ä»åä¸ºgroup_nameçš„ç»„è¿”å›æŒ‰ç±»å‹åˆ†ç±»çš„å­—å…¸ï¼Œå„ç±»å‹ç»Ÿä¸€æŒ‰boundingboxä¸­å¿ƒæ’å¥½åº


def get_group_entities_sorted_by_type_and_bbox(group_name, cha_Y=0.5):
    """
    å°†ç»„ group_name ä¸­çš„å®ä½“æŒ‰ç±»å‹åˆ†ç±»ï¼Œå¹¶å¯¹æ¯ç§ç±»å‹å†…éƒ¨æŒ‰åŒ…å›´ç›’ä¸­å¿ƒæ’åºï¼š
      1) å…ˆæŒ‰ center_y é™åºï¼ˆä»ä¸Šåˆ°ä¸‹ï¼‰
      2) åŒä¸€â€œè¡Œâ€å†…ï¼ˆ|Î”Y|<cha_Yï¼‰å†æŒ‰ center_x å‡åºï¼ˆä»å·¦åˆ°å³ï¼‰

    å‚æ•°ï¼š
      group_name: è¦æ“ä½œçš„ç»„å
      cha_Y: åŒä¸€â€œè¡Œâ€Y æ–¹å‘å®¹å·®

    è¿”å›ï¼š
      ä¸€ä¸ª dictï¼Œkey=ç±»å‹å(ObjectName)ï¼Œvalue=æ’åºåçš„å®ä½“åˆ—è¡¨
    """
    # 1. å–ç»„
    group = nametogroup(group_name)
    ents = [group.Item(i) for i in range(group.Count)]

    # 2. æŒ‰ç±»å‹åˆ†ç»„
    by_type = {}
    for ent in ents:
        typ = getattr(ent, "ObjectName", None) or ent.EntityName
        by_type.setdefault(typ, []).append(ent)

    # 3. è¾…åŠ©ï¼šè®¡ç®—åŒ…å›´ç›’ä¸­å¿ƒç‚¹

    # 4. å¯¹æ¯ä¸ªç±»å‹å†…éƒ¨æ’åº
    for typ, lst in by_type.items():
        triples = [(e, *bbox_center_2(e)) for e in lst]
        # Y é™åº
        triples.sort(key=lambda t: -t[2])
        # åŒè¡Œå†…æŒ‰ X å‡åº
        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])
            i = j
        # è¦†ç›–åŸåˆ—è¡¨
        by_type[typ] = [t[0] for t in triples]

    return by_type


# è·å–ä¸¤ä¸ªç»„ä¸­å…±æœ‰çš„å®ä½“ï¼ŒæŒ‰ç±»å‹åˆ†ç±»å¹¶æŒ‰åŒ…å›´ç›’ä¸­å¿ƒæ’åº
def common_group_entities_sorted(group_name1, group_name2, cha_Y=0.5):
    """
    è·å–ä¸¤ä¸ªç»„ä¸­å…±æœ‰çš„å®ä½“ï¼ŒæŒ‰ç±»å‹åˆ†ç±»å¹¶æŒ‰åŒ…å›´ç›’ä¸­å¿ƒæ’åºã€‚

    å‚æ•°ï¼š
      group_name1, group_name2: è¦æ¯”è¾ƒçš„ä¸¤ä¸ªç»„å
      cha_Y: åŒä¸€è¡Œåˆ¤å®šçš„ Y æ–¹å‘å®¹å·®

    è¿”å›ï¼š
      dict => key: ObjectName ç±»å‹å, value: æ’åºåçš„å®ä½“åˆ—è¡¨
    """
    # 1. å–ç»„
    g1 = nametogroup(group_name1)
    g2 = nametogroup(group_name2)

    ents1 = [g1.Item(i) for i in range(g1.Count)]
    ents2 = [g2.Item(i) for i in range(g2.Count)]

    # 2. å»ºç«‹ handle->entity æ˜ å°„
    map1 = {e.Handle: e for e in ents1}
    map2 = {e.Handle: e for e in ents2}

    # 3. æ‰¾å…±æœ‰çš„ handles
    common_handles = set(map1.keys()) & set(map2.keys())

    # 4. æ”¶é›†å…±æœ‰å®ä½“ï¼ˆè¿™é‡Œå–è‡ª map1ï¼‰
    common_ents = [map1[h] for h in common_handles]

    # 5. æŒ‰ç±»å‹åˆ†ç»„
    by_type = {}
    for ent in common_ents:
        typ = getattr(ent, "ObjectName", None) or ent.EntityName
        by_type.setdefault(typ, []).append(ent)

    # 6. åŒ…å›´ç›’ä¸­å¿ƒç‚¹è®¡ç®—
    def bbox_center(e):
        min_pt, max_pt = e.GetBoundingBox()
        x1, y1, _ = tuple(min_pt)
        x2, y2, _ = tuple(max_pt)
        return ((x1 + x2) / 2, (y1 + y2) / 2)

    # 7. æ’åºï¼šY é™åºï¼ŒåŒä¸€è¡Œå†…(|Î”Y|<cha_Y)æŒ‰ X å‡åº
    for typ, lst in by_type.items():
        triples = [(e, *bbox_center(e)) for e in lst]
        triples.sort(key=lambda t: -t[2])
        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])
            i = j
        by_type[typ] = [t[0] for t in triples]

    return by_type

def get_boundingbox_from_group(group):#ä»comå¯¹è±¡groupå»ºç«‹æœ€å°è¾¹ç•Œæ¡†

    """
    å¹¶éç»„çš„å®é™…BoundingBOXæ•°æ®

    """

    entities = [group.Item(i) for i in range(group.Count)] #ä»ç»„ä¸­è·å–å…¨éƒ¨å¯¹è±¡

    p1,p2 = get_boundingbox_from_objects(entities)

    return p1,p2


def select_group_entities(group_obj): #é€‰æ‹©ç»„ä¸­å¯¹è±¡ï¼Œè½¬æ¢ä¸ºé«˜äº®æ˜¾æ€§é€‰æ‹©
    """
    è®©ç»„ä¸­çš„æ‰€æœ‰å¯¹è±¡è¿›å…¥è“è‰²é«˜äº®çŠ¶æ€ï¼ˆé€šè¿‡åé€‰æ³•ï¼‰
    """
    try:
        handles = [entity.Handle for entity in group_obj]
        objs = []

        for obj in mp:
            if obj.Handle in handles:
                objs.append(obj)

        yin_to_xian_xuanze(objs)  # ä½ å·²æœ‰çš„é«˜äº®å‡½æ•°
        print("[OK] å·²æ˜¾æ€§é€‰ä¸­ç»„å†…å›¾å…ƒ")
        return True
    except Exception as e:
        print("[é”™è¯¯] æ˜¾æ€§é€‰æ‹©å¤±è´¥:", e)
        return False

#&&% #  Handleå’ŒLabel


"""
Handleèº«ä»½ä¿¡æ¯å’Œå¤©æ­£æ ‡ç­¾Label

ä¸åˆ—è¡¨å’Œç»„ä¸åŒï¼Œèº«ä»½ä¿¡æ¯çš„æ ‡è¯†ä½¿æˆ‘ä»¬å¯ä»¥ä»å­—å…¸å­˜å‚¨çš„åç§°ä¿¡æ¯ç²¾å‡†æ§åˆ¶æ¯ä¸ªå¯¹è±¡ï¼Œè€Œä¸æ˜¯æ€»ä¾èµ–å¯¹å…¨éƒ¨å¯¹è±¡çš„éå†


å¯ä»¥åœ¨å¯¹è±¡åˆ›å»ºæ—¶å°±æ‰“ä¸Šæ ‡ç­¾ï¼Œä½¿å¯¹è±¡å½’ä¸ºæŸä¸ªç±»

å¯ä»¥åœ¨ä¸€æ‰¹å¯¹è±¡å®Œæˆåè°ƒæ•´æ ‡ç­¾

å¯ä»¥ä¸´æ—¶æ ‡è®°ä¸€ç»„æ ‡ç­¾

å¯ä»¥é€šè¿‡Labelç»™å¤©æ­£å¯¹è±¡åŠ æ ‡ç­¾
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)#è‡ªåŠ¨è·å–å›¾çº¸ç©ºé—´ä¹‹å‰æœ€åä¸€ä¸ªç»˜åˆ¶çš„å¯¹è±¡


ä»Handleå›æº¯comå¯¹è±¡

doc.HandleToObject('2BD')
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
q1.Label="A1"
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
q1.Label="B1"
LB=pmxz()
è¯·åœ¨å±å¹•æ‹¾å–å›¾å…ƒï¼Œä»¥Enteré”®ç»“æŸ
LB[0].Label
'B1'
LB[1].Label
'A1'


"""

# å°†åˆ—è¡¨å¯¹è±¡æŒ‰åˆ†ç±»å°†å…¶handleèº«ä»½æ ‡è¯†å­˜å…¥å­—å…¸

@alias("h")
def HandleToObject(ZF):#ä»Handleèº«ä»½ä¿¡æ¯å€¼å›æº¯comå¯¹è±¡

    """
    å¯¹è¿æ¥åœ¨å¢™ä¸Šçš„é—¨çª—æµ‹è¯•æ— æ•ˆ

    """

    obj = doc.HandleToObject(ZF)

    return obj

def print_coms_handle(LB):

    LC=[]

    for x in LB:

        LC.append(x.Handle)

    print(f"comå¯¹è±¡åˆ—è¡¨å¯¹åº”çš„Handleå¥æŸ„åˆ—è¡¨ï¼š{LC} ")




@alias("H")
def handles_to_coms(LB_handles):

    """
    å¯¹è¿æ¥åœ¨å¢™ä¸Šçš„é—¨çª—æµ‹è¯•æ— æ•ˆ

    """
    LC=[]

    for xx in LB_handles:

        obj = doc.HandleToObject(xx)
        LC.append(obj)

    return LC


def get_all_handles():#è·å–æ‰€æœ‰Handle
    """
    è·å–å½“å‰å›¾çº¸ä¸­æ‰€æœ‰å¯¹è±¡ï¼ˆé€šå¸¸åœ¨ ModelSpaceï¼‰çš„ Handle å€¼åˆ—è¡¨ã€‚

    è¿”å›ï¼š
        handle_list: æ‰€æœ‰å›¾å…ƒçš„ Handle å­—ç¬¦ä¸²åˆ—è¡¨
    """
    handle_list = []

    for obj in mp:  # ä½¿ç”¨ä½ å·²ç»å®šä¹‰å¥½çš„å…¨å±€ ModelSpace mp
        try:
            handle_list.append(obj.Handle)
        except:
            continue  # è·³è¿‡æ—  Handle æˆ–å¼‚å¸¸å¯¹è±¡

    print(f"[OK] å·²è·å– {len(handle_list)} ä¸ªå¯¹è±¡çš„ Handle")
    return handle_list

def find_entity_by_handle(handle_str):#ä»Handleè·å–å®ä½“ï¼ˆé€‚åˆåŒ…æ‹¬å¤©æ­£çš„æ–‡ä»¶ï¼‰
    """
    éå†å½“å‰å›¾çº¸æ‰€æœ‰å¯¹è±¡ï¼Œæ‰‹åŠ¨æ¯”å¯¹ Handle å€¼ï¼Œæ‰¾åˆ°æŒ‡å®šçš„å®ä½“å¯¹è±¡ã€‚

    å‚æ•°ï¼š
        handle_str: ç›®æ ‡ Handleï¼ˆå­—ç¬¦ä¸²ï¼‰

    è¿”å›ï¼š
        å¯¹è±¡ï¼ˆè‹¥æ‰¾åˆ°ï¼‰ï¼Œå¦åˆ™ None
    """
    for obj in mp:  # å¯æ‰©å±•ï¼šéå† sp ä¹Ÿè¡Œ
        try:
            if obj.Handle == handle_str:
                return obj
        except:
            continue

    return None



def group_objects_by_type_and_handle(LB):#å°†åˆ—è¡¨å¯¹è±¡çš„Handleèº«ä»½ä¿¡æ¯åˆ†ç±»å­˜å…¥å­—å…¸è¿”å›
    """
    å°†comå¯¹è±¡åˆ—è¡¨ LB ä¸­çš„å¯¹è±¡æŒ‰ ObjectName åˆ†ç±»ï¼Œå¹¶å­˜å‚¨å…¶ Handleã€‚
    æ¯ç±»æŒ‰ LB ä¸­å‡ºç°é¡ºåºç¼–å·ã€‚

    å‚æ•°ï¼š
        LB - AutoCAD å®ä½“å¯¹è±¡åˆ—è¡¨ï¼ˆå¦‚ select_objects_in_window_area() è¿”å›ï¼‰

    è¿”å›ï¼š
        ZD - dict æ ¼å¼ {ObjectName: [Handle1, Handle2, ...]}
    """
    ZD = {}  # åˆå§‹åŒ–å­—å…¸

    for obj in LB:
        try:
            obj_type = obj.ObjectName
            handle = obj.Handle

            if obj_type not in ZD:
                ZD[obj_type] = []

            ZD[obj_type].append(handle)

        except Exception as e:
            print(f"[è­¦å‘Š]ï¸ è·³è¿‡å¯¹è±¡ï¼ŒåŸå› : {e}")
            continue

    # è¾“å‡ºæç¤ºä¿¡æ¯
    for obj_type, handles in ZD.items():
        print(f"[OK] {obj_type}: å…± {len(handles)} ä¸ªå¯¹è±¡")

    return ZD

# é€šè¿‡åç§°å­˜å‚¨å¯¹è±¡ä¿¡æ¯åå›æº¯å¯¹è±¡

def record_handle_with_type(LB, typename, prefix="OBJ"):#å°†ä¸€æ‰¹å¯¹è±¡çš„ Handle å­˜å‚¨åˆ°ç»“æ„åŒ–çš„å­—å…¸ä¸­ï¼Œå¹¶è®°å½•ç±»å‹åå’Œç¼–å·
    """
    æ›¿ä»£ XData æ–¹æ³•ï¼šè®°å½•å¯¹è±¡ Handleã€ç±»å‹åã€ç¼–å·ï¼Œè¿”å›ç»“æ„åŒ–å­—å…¸ã€‚
    """
    ZD = {typename: {}}
    for i, obj in enumerate(LB, start=1):
        try:
            h = obj.Handle
            tag = f"{prefix}_{i:03d}"
            ZD[typename][h] = tag
        except:
            continue
    print(f"[OK] å·²è®°å½• {len(ZD[typename])} ä¸ªâ€œ{typename}â€å¯¹è±¡ï¼ˆHandle+ç¼–å·ï¼‰")
    return ZD

def convert_named_dict(ZD, typename):# æ„å»ºç¼–å· â†’ COM å¯¹è±¡ çš„æ˜ å°„å­—å…¸
    """
    å°† ZD["é—¨"] çš„ç»“æ„ç”± Handle: ç¼–å· è½¬æ¢ä¸º ç¼–å·: COMå¯¹è±¡
    è¿”å›ï¼šæ–°çš„å­—å…¸ {ç¼–å·: COMå®ä½“}
    """
    doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument
    named_dict = {}

    handle_map = ZD.get(typename, {})
    for handle, name in handle_map.items():
        try:
            obj = doc.HandleToObject(handle)
            named_dict[name] = obj
        except:
            print(f"[è­¦å‘Š]ï¸ æ— æ³•æ‰¾åˆ°å¯¹è±¡ï¼ˆhandle={handle}ï¼‰")
            continue

    return named_dict

def get_named_object(tag, ZD, typename="é—¨"):#ä»æ ‡ç­¾è·å–å¯¹è±¡
    named = convert_named_dict(ZD, typename)
    return named.get(tag)



def draw_tags_on_objects_fixed(named_dict, height=250, offset=(1000, 1000, 0)):#ç›´æ¥å°†ç¼–å·æ–‡å­—å†™åœ¨æ¯ä¸ªå›¾å…ƒä¸Šï¼Œé€šå¸¸å±…ä¸­æˆ–åç§»ä¸€ç‚¹ç‚¹
    """
    åœ¨æ¯ä¸ªå¯¹è±¡çš„ä¸­å¿ƒç‚¹é™„è¿‘ç»˜åˆ¶æ ‡æ³¨æ–‡å­—ã€‚
    
    å‚æ•°:
        named_dict - å¦‚ {"Men_001": <COMObject>, ...}
        height     - æ–‡å­—é«˜åº¦
        offset     - åç§»é‡ï¼ˆç”¨äºé˜²æ­¢æ–‡å­—ç›–ä½å¯¹è±¡ï¼‰
    """
    import win32com.client
    import pythoncom

    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    ms = doc.ModelSpace

    for name, obj in named_dict.items():
        try:
            # ä½¿ç”¨ GetBoundingBox è·å–ä¸­å¿ƒç‚¹
            min_pt, max_pt = obj.GetBoundingBox()
            center_pt = (
                (min_pt[0] + max_pt[0]) / 2,
                (min_pt[1] + max_pt[1]) / 2,
                (min_pt[2] + max_pt[2]) / 2
            )

            # åŠ ä¸Šåç§»é‡
            label_pt = (
                center_pt[0] + offset[0],
                center_pt[1] + offset[1],
                center_pt[2] + offset[2]
            )

            # ç¡®ä¿æ’å…¥ç‚¹æ˜¯ä¸‰ç»´ç‚¹
            pt = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, label_pt)
            
            # æ·»åŠ æ–‡å­—
            ms.AddText(name, pt, height)
            print(f"[OK] å·²æ ‡æ³¨å¯¹è±¡: {name}")

        except Exception as e:
            print(f"[è­¦å‘Š]ï¸ æ ‡æ³¨å¤±è´¥: {name}, é”™è¯¯: {e}")

# ç»™å¤©æ­£å¯¹è±¡æ‰“ä¸Šæ ‡ç­¾å­˜å…¥å­—å…¸ï¼Œç”¨äºä»¥åç§°åå‘å›æº¯æ“ä½œ

def label_tarch_doors(LB1, typename="é—¨", prefix="men"):#ç»™é€‰ä¸­çš„å¤©æ­£é—¨æ‰“ä¸Šæ ‡ç­¾å¹¶å­˜å…¥å­—å…¸è¿”å›ï¼Œéå¤©æ­£å›¾å…ƒæ²¡æœ‰Labelå±æ€§
    """
    ä»å¯¹è±¡åˆ—è¡¨ LB1 ä¸­ç­›é€‰å‡ºå¤©æ­£é—¨ (ObjectName == 'TDbOpening')ï¼Œ
    å¹¶ä¸ºå…¶æŒ‰é¡ºåºæ‰“ä¸Š .Label æ ‡ç­¾ï¼ˆå¦‚ men001, men002 ...ï¼‰ã€‚

    è¿”å›ï¼š
        ZD = {typename: {ç¼–å·: å¯¹è±¡}}
    """
    ZD = {typename: {}}
    LB2 = []

    for obj in LB1:
        try:
            if hasattr(obj, "ObjectName") and obj.ObjectName == "TDbOpening":
                LB2.append(obj)
        except:
            continue

    for i, obj in enumerate(LB2, start=1):
        try:
            tag = f"{prefix}{i:03d}"
            obj.Label = tag
            ZD[typename][tag] = obj
            print(f"[OK] å·²æ ‡è®°: {tag}")
        except Exception as e:
            print(f"[è­¦å‘Š]ï¸ è®¾ç½®æ ‡ç­¾å¤±è´¥ï¼š{e}")

    print(f"\nğŸ“¦ å…±æ‰¾åˆ°å¹¶æ ‡æ³¨ {len(LB2)} ä¸ªå¤©æ­£é—¨")
    return ZD


# è·å–æ¨¡å‹ç©ºé—´ä¸Šç»˜åˆ¶çš„æœ€åä¸€ä¸ªå›¾å…ƒ

def last_obj():
    """è·å–æ¨¡å‹ç©ºé—´æœ€åä¸€ä¸ªå¯¹è±¡"""
    _, doc = get_acad_doc()
    obj = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
    return obj

# è·å–æ¨¡å‹ç©ºé—´ä¸Šçš„Handle


"""
target_handles = ['5F', '60', '61']
map = get_handle_object_map(doc.ModelSpace)
objs = [map[h] for h in target_handles if h in map]
è¿™æ¯”æ¯æ¬¡éƒ½éå† ModelSpace å¿«å¾—å¤šï¼Œå°¤å…¶æ˜¯å¤§å›¾çº¸ä¸­ä¸Šåƒä¸ªå›¾å…ƒæ—¶ã€‚

"""

def get_handle_object_map(ms):
    """è¿”å› {handle: object} æ˜ å°„"""
    return {ent.Handle: ent for ent in ms}



#&&% XData

"""

åœ¨ RegAppTable ä¸­æ³¨å†Œ,åœ¨ç¬¬ä¸€æ¬¡ç»™å›¾å…ƒé™„åŠ  XData æ—¶ï¼ŒAutoCAD å†…éƒ¨ä¼šæ£€æŸ¥ RegAppTableï¼ˆæ³¨å†Œåº”ç”¨ç¨‹åºè¡¨ï¼‰ä¸­æ˜¯å¦å·²ç»å­˜åœ¨ â€œTestAppâ€ è¿™ä¸ªåç§°ã€‚

å¦‚æœä¸å­˜åœ¨ï¼ŒAutoCAD ä¼šè‡ªåŠ¨å¾€ RegAppTable é‡Œæ’å…¥ä¸€æ¡è®°å½•ï¼ŒæŠŠ â€œTestAppâ€ æ³¨å†Œè¿›å»ã€‚

å¦‚æœä½ å¸Œæœ›æ‰‹åŠ¨æ§åˆ¶ï¼Œä¹Ÿå¯ä»¥å…ˆè°ƒç”¨ doc.Application.RegistryModes.Add("TestApp")ï¼ˆæˆ–ä½¿ç”¨ AutoLISPï¼š(regapp "TestApp")ï¼‰

app_name    = "TestApp"            # è‡ªå®šä¹‰çš„åº”ç”¨ç¨‹åºå
data_types  = [1000, 1040, 1070]
data_values = ["ç¤ºä¾‹æ–‡å­—", 3.14159, 12345]
set_xdata(lineObj, app_name, data_types, data_values)
types_out, data_out = get_xdata(lineObj, app_name)
types_out
[1001, 1000, 1040, 1070]
data_out
['TestApp', 'ç¤ºä¾‹æ–‡å­—', 3.14159, 12345]


"""
def set_xdata(
    com_obj,
    app_name: str,
    data_types: list[int],
    data_values: list,
):
    """
    å‘ä»»æ„ AutoCAD COM å¯¹è±¡ï¼ˆå¦‚ Lineã€Circleã€BlockReference ç­‰ï¼‰é™„åŠ  XDataã€‚

    å‚æ•°ï¼š
      com_obj        -- ä»»æ„æ”¯æŒ SetXData() æ–¹æ³•çš„ COM å¯¹è±¡
      app_name       -- æ³¨å†Œè¿‡çš„åº”ç”¨ç¨‹åºåï¼ˆå­—ç¬¦ä¸²ï¼‰ï¼›ç¬¬ä¸€ä¸ª DataType ä¸€å®šæ˜¯ 1001ï¼Œå¯¹åº”çš„ç¬¬ä¸€ä¸ª Data å­˜æ”¾æ­¤ app_name
      data_types     -- åç»­çš„ DataType åˆ—è¡¨ï¼ˆä¸å«ç¬¬ä¸€é¡¹ 1001ï¼‰ï¼›ä¾‹å¦‚ [1000, 1040, 1070] ç­‰
      data_values    -- ä¸ data_types å¯¹åº”çš„æ•°æ®å€¼åˆ—è¡¨ï¼›é•¿åº¦ä¸ data_types ä¸€ä¸€å¯¹åº”ã€‚ä¾‹å¦‚ ["æ–‡å­—ä¸²", 3.14, 42]

    è¯´æ˜ï¼š
      AutoCAD è§„å®š XData çš„ç¬¬ä¸€å¯¹å…ƒç´ å¿…é¡»æ˜¯ (1001, åº”ç”¨ç¨‹åºå)ã€‚åé¢æ‰æ˜¯æŒ‰é¡ºåºå‡ºç°çš„å…¶ä»– (DataType, Data)ã€‚
      å› æ­¤å®é™…å‘é€ç»™ SetXData çš„ DataType æ•°ç»„ç¬¬ä¸€ä¸ªå…ƒç´ è¦æ”¾ 1001ï¼ŒData æ•°ç»„ç¬¬ä¸€ä¸ªå…ƒç´ è¦æ”¾ app_nameã€‚
    """
    def vtint(val):
        """
        å°† Python åˆ—è¡¨è½¬æ¢ä¸º VARIANT ç±»å‹çš„æ•´æ•°æ•°ç»„ï¼Œ
        ä»¥ä¾¿ä¼ ç»™ COM å¯¹è±¡ä½œä¸º XData çš„ DataTypeã€‚
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, val)

    def vtvariant(var):
        """
        å°† Python åˆ—è¡¨è½¬æ¢ä¸º VARIANT ç±»å‹çš„ VARIANT æ•°ç»„ï¼Œ
        ä»¥ä¾¿ä¼ ç»™ COM å¯¹è±¡ä½œä¸º XData çš„ Dataã€‚
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, var)


    # 1. æ‹¼æ¥å®Œæ•´çš„ DataType åˆ—è¡¨ï¼šç¬¬ä¸€ä¸ªä¸º 1001
    full_types = [1001] + data_types

    # 2. æ‹¼æ¥å®Œæ•´çš„ Data åˆ—è¡¨ï¼šç¬¬ä¸€ä¸ªä¸º app_name
    full_data = [app_name] + data_values

    # 3. å°† Python åˆ—è¡¨è½¬æ¢ä¸º VARIANT æ•°ç»„
    vt_types = vtint(full_types)
    vt_data  = vtvariant(full_data)

    # 4. è°ƒç”¨ COM çš„ SetXData æ–¹æ³•
    com_obj.SetXData(vt_types, vt_data)

def get_xdata(
    com_obj,
    app_name: str,
):
    """
    ä»ä»»æ„ AutoCAD COM å¯¹è±¡ï¼ˆå¦‚ Lineã€Circleã€BlockReference ç­‰ï¼‰è¯»å– XDataã€‚

    å‚æ•°ï¼š
      com_obj   -- ä»»æ„æ”¯æŒ GetXData() æ–¹æ³•çš„ COM å¯¹è±¡
      app_name  -- ç”³è¯·è¯»å–çš„åº”ç”¨ç¨‹åºåï¼ˆå¿…é¡»ä¸ set_xdata æ—¶ä½¿ç”¨çš„ç›¸åŒï¼‰

    è¿”å›ï¼š
      (type_codes, data_values) äºŒå…ƒç»„ï¼Œå…¶ä¸­
        type_codes: Python åˆ—è¡¨ï¼Œå¯¹åº”æ¯ä¸ª XData é¡¹ç›®çš„ DataTypeï¼ˆåŒ…æ‹¬ç¬¬ä¸€é¡¹ 1001ï¼‰
        data_values: Python åˆ—è¡¨ï¼Œå¯¹åº”æ¯ä¸ª XData é¡¹ç›®çš„ Dataï¼ˆåŒ…æ‹¬ç¬¬ä¸€é¡¹ app_nameï¼‰
    
    å¦‚æœè¯¥å¯¹è±¡æ²¡æœ‰é™„åŠ æ­¤ app_name ä¸‹çš„ XDataï¼Œåˆ™ GetXData ä¼šæŠ›å‡ºé”™è¯¯ï¼›å»ºè®®è°ƒç”¨å‰å…ˆç”¨ Error Handling åŒ…è£¹æˆ–
    é€šè¿‡ com_obj.GetXData(app_name) è¿›è¡Œæ•è·å¹¶è¿”å› Noneã€‚
    """
    def vtint(val):
        """
        å°† Python åˆ—è¡¨è½¬æ¢ä¸º VARIANT ç±»å‹çš„æ•´æ•°æ•°ç»„ï¼Œ
        ä»¥ä¾¿ä¼ ç»™ COM å¯¹è±¡ä½œä¸º XData çš„ DataTypeã€‚
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, val)

    def vtvariant(var):
        """
        å°† Python åˆ—è¡¨è½¬æ¢ä¸º VARIANT ç±»å‹çš„ VARIANT æ•°ç»„ï¼Œ
        ä»¥ä¾¿ä¼ ç»™ COM å¯¹è±¡ä½œä¸º XData çš„ Dataã€‚
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, var)


    try:
        type_codes, data_values = com_obj.GetXData(app_name)
        # æ³¨æ„ï¼šè¿”å›çš„ type_codes å’Œ data_values éƒ½æ˜¯ tupleï¼Œè½¬æ¢ä¸º list æ›´æ˜“å¤„ç†
        return list(type_codes), list(data_values)
    except pythoncom.com_error:
        # å¯¹è±¡ä¸Šä¸å­˜åœ¨è¯¥ app_name çš„ XDataï¼Œæˆ–è¯»å–å¤±è´¥
        return None, None


#&&% Xdataæ ‡è®°

def set_xdata_tab(entitycom):

    app_name    = "PrintApp"
    data_types  = [1000]
    data_values = ["å¢è¡¥ç›®å½•æ¨¡æ¿"]
    set_xdata(entitycom, app_name, data_types, data_values)

    return

def is_printApp_xdata_com(entitycom):

    try:

        get_xdata( entitycom, "PrintApp")

        return True

    except:

        return  False


#&&&&%%  ç¬¬äº”éƒ¨åˆ†  çº¿é¢åˆ†æ 


#_____________________________________________________________________________________________________________________________________________

#  æ¨¡å—ä½¿ç”¨è¯´æ˜

"""
è¯¥æ¨¡å—ç ”ç©¶dwgå›¾çº¸ä¸­çš„çº¿æ®µã€åœ†æ›²çº¿ã€å¹³é¢ç­‰åŸºæœ¬å‡ ä½•é—®é¢˜ 

"""
# çº¿æ®µåˆ†æ

def compute_line_angle(line):#æŒ‰ç»˜åˆ¶é¡ºåºåº¦é‡çº¿æ®µè§’åº¦
    """
    è®¡ç®—ç›´çº¿çš„æ–¹å‘è§’ï¼ˆå•ä½ï¼šåº¦ï¼‰ï¼ŒåŸºäº StartPoint / EndPointã€‚
    éç›´çº¿å¯¹è±¡å°†æŠ›å‡ºå¼‚å¸¸ã€‚
    0-360ï¼Œä»èµ·ç‚¹å¤„ç”»æ¨ªçº¿ï¼Œæ—‹è½¬åˆ°ç»ˆç‚¹çš„è§’åº¦
    """

    try:
        x1, y1, _ = line.StartPoint
        x2, y2, _ = line.EndPoint
        dx = x2 - x1
        dy = y2 - y1
        angle_rad = math.atan2(dy, dx)
        angle_deg = math.degrees(angle_rad)
        if angle_deg < 0:
            angle_deg += 360
        return angle_deg
    except AttributeError:
        print("[é”™è¯¯] è¯¥å¯¹è±¡ä¸å…·å¤‡ StartPoint / EndPoint")
        return None


def draw_point(pt):
    """
    åœ¨æ¨¡å‹ç©ºé—´ç»˜åˆ¶ä¸€ä¸ª AutoCAD ç‚¹å®ä½“ã€‚

    å‚æ•°ï¼š
        pt: (x, y, z) ä¸‰ç»´åæ ‡å…ƒç»„

    è¿”å›ï¼š
        æ–°åˆ›å»ºçš„ Point å¯¹è±¡ï¼›å¤±è´¥æ—¶è¿”å› None
    """
    try:
        # AutoCAD çš„â€œç‚¹â€ç”± AddPoint åˆ›å»ºï¼Œéœ€ä¼  VARIANT
        obj = mp.AddPoint(vtpnt(*pt))
        return obj
    except Exception as e:
        print(f"[é”™è¯¯] æ— æ³•ç»˜åˆ¶ç‚¹: {e}")
        return None

def draw_line(p1, p2):#ä»ä¸¤ç‚¹åæ ‡è¿”å›ç›´çº¿æ®µ
    """
    åœ¨æ¨¡å‹ç©ºé—´ä¸­ç»˜åˆ¶ä» p1 åˆ° p2 çš„ç›´çº¿æ®µã€‚

    å‚æ•°ï¼š
        p1, p2: ä¸‰ç»´åæ ‡å…ƒç»„ (x, y, z)

    è¿”å›ï¼š
        æ–°åˆ›å»ºçš„ç›´çº¿å¯¹è±¡ï¼ˆCOM å¯¹è±¡ï¼‰
    """
    try:
        line_obj = mp.AddLine(vtpnt(*p1), vtpnt(*p2))
        return line_obj
    except Exception as e:
        print(f"[é”™è¯¯] æ— æ³•ç»˜åˆ¶ç›´çº¿: {e}")
        return None


def draw_circle(center, radius):
    """
    ä»¥ center ä¸ºåœ†å¿ƒã€radius ä¸ºåŠå¾„ç»˜åˆ¶åœ†ã€‚

    å‚æ•°ï¼š
        center: (x, y, z)
        radius: æµ®ç‚¹åŠå¾„

    è¿”å›ï¼š
        æ–°åˆ›å»ºçš„ Circle å¯¹è±¡ï¼›å¤±è´¥æ—¶è¿”å› None
    """
    try:
        obj = mp.AddCircle(vtpnt(*center), radius)
        return obj
    except Exception as e:
        print(f"[é”™è¯¯] æ— æ³•ç»˜åˆ¶åœ†: {e}")
        return None


def draw_regular_polygon(center, radius, sides):
    """
    ç»˜åˆ¶æ­£å¤šè¾¹å½¢ï¼ˆLWPolylineï¼Œå·²é—­åˆï¼‰
    :param center: åœ†å¿ƒ (x,y,z)
    :param radius: å¤–æ¥åœ†åŠå¾„
    :param sides : è¾¹æ•° â‰¥3
    """
    if sides < 3:
        print("[é”™è¯¯] è¾¹æ•°å¿…é¡» â‰¥ 3"); return None
    cx, cy, cz = (center + (0.0,))[:3]

    pts_flat = []
    for k in range(sides):
        ang = 2 * math.pi * k / sides
        pts_flat.extend([cx + radius*math.cos(ang),
                         cy + radius*math.sin(ang)])

    try:
        v_pts = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            pts_flat
        )
        poly = mp.AddLightWeightPolyline(v_pts)
        poly.Closed = True
        return poly
    except Exception as e:
        print(f"[é”™è¯¯] æ— æ³•ç»˜åˆ¶æ­£å¤šè¾¹å½¢: {e}")
        return None



def prioritize_horizontal(lines, tol=0.5):
    """
    å°†åˆ—è¡¨ä¸­æ‰€æœ‰â€œæ°´å¹³â€ç›´çº¿æ®µï¼ˆèµ·ç‚¹å’Œç»ˆç‚¹çš„ y å·®å°äº tolï¼‰æ”¾åœ¨æœ€å‰é¢ï¼Œ
    å…¶å®ƒç›´çº¿ä¿ç•™åŸæœ‰ç›¸å¯¹é¡ºåºã€‚

    :param lines: ç›´çº¿å¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡å…·æœ‰ .StartPoint å’Œ .EndPoint å±æ€§ï¼Œ
                  è¿™ä¸¤ä¸ªå±æ€§åº”è¿”å› (x, y, z) æˆ–ç±»ä¼¼å¯ä¸‹æ ‡çš„ä¸‰å…ƒç»„ã€‚
    :param tol:   åˆ¤å®šä¸ºæ°´å¹³çš„ y æ–¹å‘å®¹å·®ï¼ˆé»˜è®¤ 0.5ï¼‰
    :return:      æ–°åˆ—è¡¨ï¼Œæ°´å¹³ç›´çº¿æ®µåœ¨å‰ï¼Œéæ°´å¹³åœ¨å
    """
    horizontals = []
    non_horizontals = []
    for ln in lines:
        y1 = ln.StartPoint[1]
        y2 = ln.EndPoint[1]
        if abs(y1 - y2) < tol:
            horizontals.append(ln)
        else:
            non_horizontals.append(ln)
    return horizontals,non_horizontals

    
def get_spline_length_by_conversion(spline_entity):#è¿”å›æ ·æ¡æ›²çº¿çš„é•¿åº¦ï¼ˆæŒ‰é»˜è®¤10åˆ†æ–­æ‹Ÿåˆï¼‰
    
    """
    å°†æ ·æ¡æ›²çº¿å¯¹è±¡å¤åˆ¶ã€é«˜äº®å¹¶é€šè¿‡ _SPLINEDIT è½¬ä¸ºå¤šæ®µçº¿ï¼Œ
    ç„¶åè¯»å–é•¿åº¦ï¼Œå¹¶åˆ é™¤è¯¥å¤šæ®µçº¿ã€‚

    è¿”å›ï¼š
        æ ·æ¡æ›²çº¿è½¬æ¢åçš„é•¿åº¦å€¼ï¼ˆfloatï¼‰
    """
    try:
        # Step 1ï¼šå¤åˆ¶ spline å¯¹è±¡
        new_spline = spline_entity.Copy()

        # Step 2ï¼šæ˜¾æ€§é«˜äº®
        highlight_entity_by_bbox(new_spline)

        # Step 3ï¼šæ¨¡æ‹Ÿå‘½ä»¤ SPLINEDIT â†’ P â†’ Enter â†’ Enter
        doc.SendCommand("_SPLINEDIT\nP\n\n\n")
        time.sleep(1.2)  # ç­‰å¾… CAD å®Œæˆå¤„ç†ï¼ˆå¯æ ¹æ®æœºå™¨é€Ÿåº¦è°ƒæ•´ï¼‰

        # Step 4ï¼šè·å–æ–°ç”Ÿæˆçš„å¯¹è±¡ï¼ˆæœ€åä¸€ä¸ªï¼‰
        last_index = doc.ModelSpace.Count - 1
        poly = doc.ModelSpace.Item(last_index)

        # Step 5ï¼šæ£€æŸ¥ Length å±æ€§
        if hasattr(poly, "Length"):
            length = poly.Length
            poly.Delete()  # åˆ é™¤ä¸´æ—¶ polyline
            return length
        else:
            print("[é”™è¯¯] è½¬æ¢åå¯¹è±¡æ²¡æœ‰ Length å±æ€§")
            return None

    except Exception as e:
        print(f"[é”™è¯¯] è·å–æ ·æ¡æ›²çº¿é•¿åº¦å¤±è´¥ï¼š{e}")
        return None

def estimate_ellipse_length(ellipse):#è¿”å›æ¤­åœ†é•¿åº¦
    """
    ä¼°ç®—æ¤­åœ†å¯¹è±¡çš„é•¿åº¦ï¼ˆå‘¨é•¿ï¼‰ï¼Œä½¿ç”¨ Ramanujan å…¬å¼ã€‚
    """
    try:
        a = ellipse.MajorRadius
        b = ellipse.MinorRadius

        pi = math.pi
        h = 3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b))
        length = pi * h
        return length
    except Exception as e:
        print(f"[é”™è¯¯] æ— æ³•ä¼°ç®—æ¤­åœ†é•¿åº¦: {e}")
        return None



def get_entity_geometry_info(obj):#è¿”å›å›¾å½¢å…³é”®å‡ ä½•ä¿¡æ¯
    """
    æ ¹æ®å›¾å…ƒç±»å‹è¿”å›å…¶å…³é”®å‡ ä½•ä¿¡æ¯ï¼š
    - ç‚¹ï¼šåæ ‡
    - ç›´çº¿ï¼šèµ·ç‚¹ã€ç»ˆç‚¹ã€é•¿åº¦
    - åœ†ï¼šåœ†å¿ƒã€åŠå¾„ã€é•¿åº¦ã€é¢ç§¯
    - æ¤­åœ†ï¼šä¸­å¿ƒã€ä¸»è½´ã€æ¬¡è½´ã€é•¿åº¦ã€é¢ç§¯
    - å¤šæ®µçº¿ï¼šèµ·ç‚¹ã€ç»ˆç‚¹ã€é•¿åº¦ã€é¢ç§¯ï¼ˆè‹¥é—­åˆï¼‰
    - æ ·æ¡æ›²çº¿ï¼šèµ·ç‚¹ã€ç»ˆç‚¹ã€é•¿åº¦ï¼ˆéœ€è½¬æ¢ï¼‰ï¼Œé¢ç§¯ï¼ˆè‹¥é—­åˆï¼‰
    """
    try:
        name = obj.ObjectName.lower()

        # ç‚¹
        if "point" in name:
            return {"type": "Point", "position": obj.Coordinates}

        # ç›´çº¿
        elif "line" in name and "xline" not in name:
            p1 = obj.StartPoint
            p2 = obj.EndPoint
            length = math.dist(p1, p2)
            return {"type": "Line", "start": p1, "end": p2, "length": length}

        # åœ†
        elif "circle" in name:
            center = obj.Center
            radius = obj.Radius
            length = 2 * math.pi * radius
            area = math.pi * radius ** 2
            return {"type": "Circle", "center": center, "radius": radius, "length": length, "area": area}

        # æ¤­åœ†
        elif "ellipse" in name:
            center = obj.Center
            a = obj.MajorRadius
            b = obj.MinorRadius
            area = math.pi * a * b
            h = 3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b))
            length = math.pi * h  # Ramanujan å…¬å¼
            return {
                "type": "Ellipse",
                "center": center,
                "major_radius": a,
                "minor_radius": b,
                "length": length,
                "area": area
            }

        # å¤šæ®µçº¿
        elif "polyline" in name:
            coords = obj.Coordinates
            start = (coords[0], coords[1], 0)
            end = (coords[-2], coords[-1], 0)
            length = getattr(obj, "Length", 0)
            area = obj.Area if obj.Closed else 0
            return {
                "type": "Polyline",
                "start": start,
                "end": end,
                "length": length,
                "area": area
            }

        # æ ·æ¡æ›²çº¿ï¼ˆéœ€è½¬æ¢æµ‹é‡ï¼‰
        elif "spline" in name:
            p1 = obj.GetFitPoint(0)
            p2 = obj.GetFitPoint(obj.NumberOfFitPoints - 1)
            length = get_spline_length_by_conversion(obj)
            area = obj.Area if obj.Closed else 0
            return {
                "type": "Spline",
                "start": p1,
                "end": p2,
                "length": length,
                "area": area
            }

        else:
            return {"type": "Unknown", "ObjectName": obj.ObjectName}

    except Exception as e:
        return {"type": "Error", "message": str(e)}



##åœ¨ä¸¤ç‚¹ç¡®å®šçš„æ–¹å‘ä¸Šï¼Œè¿”å›ä¸å¯¹è±¡ç‚¹æŒ‡å®šè·ç¦»çš„ç‚¹

def points_on_line_at_distance_3d(
    p1: Tuple[float, float, float],
    p2: Tuple[float, float, float],
    px: Tuple[float, float, float],
    distance: float
) -> List[Tuple[float, float, float]]:
    """
    å·²çŸ¥ px åœ¨ç”± p1->p2 ç¡®å®šçš„ç›´çº¿ä¸Šï¼Œè¿”å›åœ¨è¯¥ç›´çº¿ä¸Šä¸ px è·ç¦»ä¸º distance çš„ä¸¤ä¸ªç‚¹ã€‚

    :param p1: èµ·ç‚¹ (x1, y1, z1)
    :param p2: ç»ˆç‚¹ (x2, y2, z2)
    :param px: å‚è€ƒç‚¹ (x, y, z)ï¼Œä½äºç›´çº¿ä¸Š
    :param distance: ä¸ px çš„ç›®æ ‡è·ç¦»
    :return: [(ax, ay, az), (bx, by, bz)]ï¼Œåˆ†åˆ«ä¸ºæ­£å‘å’Œåå‘ç§»åŠ¨ distance åçš„ç‚¹
    """
    x1, y1, z1 = p1
    x2, y2, z2 = p2
    xx, yy, zz = px

    # 1) è®¡ç®—æ–¹å‘å‘é‡å¹¶å½’ä¸€åŒ–
    dx, dy, dz = x2 - x1, y2 - y1, z2 - z1
    length = math.sqrt(dx*dx + dy*dy + dz*dz)
    if length == 0:
        raise ValueError("p1 å’Œ p2 é‡åˆï¼Œæ— æ³•ç¡®å®šæ–¹å‘å‘é‡")
    ux, uy, uz = dx / length, dy / length, dz / length

    # 2) æ²¿æ­£å‘å’Œåå‘å„ç§»åŠ¨ distance
    ax = xx + ux * distance
    ay = yy + uy * distance
    az = zz + uz * distance

    bx = xx - ux * distance
    by = yy - uy * distance
    bz = zz - uz * distance

    return [(ax, ay, az), (bx, by, bz)]


# æ‰¾å‡ºä¸€ç»„ç›´çº¿æ®µå†…çš„ä¼ªç›¸äº¤åŒºåŸŸ

def find_fake_intersection_regions(lines, tol=10, real_tol=0.01):
    """
    æŸ¥æ‰¾ä¼ªç›¸äº¤åŒºåŸŸï¼šå¯¹äºä»»æ„çº¿æ®µ A çš„ç«¯ç‚¹ Pï¼Œè‹¥ï¼š
    - å­˜åœ¨å…¶ä»–çº¿æ®µ B æ»¡è¶³ P åˆ° B è·ç¦» < tolï¼Œä¸”
    - å¯¹æ‰€æœ‰ Bï¼ŒP åˆ° B çš„è·ç¦» >= real_tol
    åˆ™åˆ¤å®šä¸ºä¼ªç›¸äº¤ç‚¹ã€‚
    åœ¨æ¨¡å‹ç©ºé—´ä¸­ç»˜åˆ¶åœ†ï¼ˆåŠå¾„ 1000ï¼‰è¡¨ç¤ºè¿™äº›ç‚¹ã€‚
    """
    ensure_layer("æµ‹è¯•è¾…åŠ©")
    ms = doc.ModelSpace
    added = []

    def point_to_line_distance(p, a1, a2):
        x0, y0 = p[:2]
        x1, y1 = a1[:2]
        x2, y2 = a2[:2]
        dx, dy = x2 - x1, y2 - y1
        if dx == dy == 0:
            return math.hypot(x0 - x1, y0 - y1)
        t = max(0, min(1, ((x0 - x1) * dx + (y0 - y1) * dy) / (dx*dx + dy*dy)))
        proj_x = x1 + t * dx
        proj_y = y1 + t * dy
        return math.hypot(x0 - proj_x, y0 - proj_y)

    for A in lines:
        try:
            p1 = A.StartPoint
            p2 = A.EndPoint
        except Exception:
            continue

        for pt in [p1, p2]:
            pt_key = tuple(round(c, 3) for c in pt)
            if pt_key in added:
                continue

            min_dist = 1e10
            has_near = False
            has_real_near = False

            for B in lines:
                if B == A:
                    continue
                try:
                    b1, b2 = B.StartPoint, B.EndPoint
                    dist = point_to_line_distance(pt, b1, b2)
                    if dist < tol:
                        has_near = True
                    if dist < real_tol:
                        has_real_near = True
                        break
                except:
                    continue

            if has_near and not has_real_near:
                ms.AddCircle(vtpnt(*pt), 1000).Layer = "æµ‹è¯•è¾…åŠ©"
                added.append(pt_key)
                print(f"[OK] ä¼ªç›¸äº¤åŒºåŸŸç‚¹: {pt}")

    print("[OK] ä¼ªç›¸äº¤åŒºåŸŸç»˜åˆ¶å®Œæˆ")



# æŠŠåŒºåŸŸå†…çš„ç›´çº¿æ®µäº¤ç‚¹æ‰“æ–­

def lines_daduan(start_point,end_point):#å…¨éƒ¨è„šæœ¬ç»Ÿä¸€é‡‡ç”¨ä¸‰ç»´åæ ‡ç‚¹æ¨¡å¼

    """
    è¿™ä¸ªå‘½ä»¤å¯¹äºé¿å…å¤©æ­£å¢™ä½“æ²¡æœ‰å‡ºç°ä¸ç›¸äº¤çš„è¦†ç›–æ˜¯éå¸¸é‡è¦çš„ï¼Œç›´æ¥åº”ç”¨å¤©æ­£çš„tlinebk

    è¿˜è¦å…ˆå¤„ç†å‡ç›¸äº¤ç‚¹åŒºåŸŸå¾…ä¼˜åŒ–20250409

    """

    # ä½¿ç”¨ f-string è¯­æ³•å°†ä¸‰ç»´åæ ‡å˜é‡æ’å…¥å‘½ä»¤å­—ç¬¦ä¸²ä¸­
    start_point_str = f"{start_point[0]},{start_point[1]},{start_point[2]}"

    end_point_str = f"{end_point[0]},{end_point[1]},{end_point[2]}"

    command = f"tlinebk{chr(13)}{start_point_str}{chr(13)}{end_point_str}{chr(13)}{chr(13)}{chr(13)}"

    acad.ActiveDocument.SendCommand(command)






#æ‰¾å‡ºä¸€ç»„ç›´çº¿æ®µä¸­çš„æ‰€æœ‰ç›´çº¿æ®µä¸­æ‰€æœ‰é‡å¤çš„çº¿æ®µå¹¶åˆ é™¤

def delete_duplicate_lines(lines, tol=0.01):
    """
    åˆ é™¤é‡å¤çš„ç›´çº¿æ®µï¼Œä»…ä¿ç•™æ¯ç»„ä¸­ä¸€æ¡ã€‚

    å‚æ•°ï¼š
        lines: æ¨¡å‹ç©ºé—´ä¸­æ‰€æœ‰çº¿æ®µå¯¹è±¡åˆ—è¡¨ï¼ˆObjectName ä¸º 'AcDbLine'ï¼‰
        tol: è·ç¦»å®¹å·®ï¼Œå°äºæ­¤å€¼è®¤ä¸ºä¸¤ç‚¹é‡åˆ
    """
    def is_same_point(p1, p2):
        return all(abs(a - b) < tol for a, b in zip(p1, p2))

    def is_duplicate(line1, line2):
        try:
            a1, a2 = line1.StartPoint, line1.EndPoint
            b1, b2 = line2.StartPoint, line2.EndPoint
            return (
                (is_same_point(a1, b1) and is_same_point(a2, b2)) or
                (is_same_point(a1, b2) and is_same_point(a2, b1))
            )
        except:
            return False

    keep = []
    to_delete = []

    for i, line in enumerate(lines):
        is_dup = False
        for j in range(i):
            if is_duplicate(line, lines[j]):
                is_dup = True
                break
        if is_dup:
            to_delete.append(line)
        else:
            keep.append(line)

    count = 0
    for dup in to_delete:
        try:
            dup.Delete()
            count += 1
        except:
            continue

    print(f"[OK] åˆ é™¤äº† {count} æ¡é‡å¤ç›´çº¿æ®µï¼Œä¿ç•™ {len(keep)} æ¡ã€‚")
    return keep



#åˆ é™¤å®Œå…¨æˆ–å±€éƒ¨é‡å¤çº¿æ®µ

def delete_redundant_lines(lines, tol=0.01):
    """
    åˆ é™¤é‡å¤çº¿æ®µå’Œå±€éƒ¨é‡å¤çº¿æ®µï¼Œåªä¿ç•™æ¯ç»„ä¸­çš„ä¸€æ¡ã€‚
    """
    def is_same_point(p1, p2):
        return abs(p1[0] - p2[0]) < tol and abs(p1[1] - p2[1]) < tol

    def point_on_segment(p, a, b):
        ax, ay = a[:2]
        bx, by = b[:2]
        px, py = p[:2]
        cross = abs((bx - ax) * (py - ay) - (by - ay) * (px - ax))
        if cross > tol:
            return False
        dot = (px - ax) * (bx - ax) + (py - ay) * (by - ay)
        if dot < 0:
            return False
        sq_len = (bx - ax)**2 + (by - ay)**2
        if dot > sq_len:
            return False
        return True

    def is_completely_duplicate(l1, l2):
        try:
            p1, p2 = l1.StartPoint, l1.EndPoint
            q1, q2 = l2.StartPoint, l2.EndPoint
            return (
                (is_same_point(p1, q1) and is_same_point(p2, q2)) or
                (is_same_point(p1, q2) and is_same_point(p2, q1))
            )
        except:
            return False

    def is_locally_duplicate(short_line, long_line):
        try:
            p1, p2 = short_line.StartPoint, short_line.EndPoint
            q1, q2 = long_line.StartPoint, long_line.EndPoint
            return point_on_segment(p1, q1, q2) and point_on_segment(p2, q1, q2)
        except:
            return False

    to_delete_handles = set()
    total = len(lines)

    for i in range(total):
        l1 = lines[i]
        h1 = l1.Handle
        if h1 in to_delete_handles:
            continue
        for j in range(i + 1, total):
            l2 = lines[j]
            h2 = l2.Handle
            if h2 in to_delete_handles:
                continue

            if is_completely_duplicate(l1, l2):
                to_delete_handles.add(h2)
            elif is_locally_duplicate(l2, l1):
                to_delete_handles.add(h2)
            elif is_locally_duplicate(l1, l2):
                to_delete_handles.add(h1)
                break

    deleted = 0
    for ent in lines:
        if ent.Handle in to_delete_handles:
            try:
                ent.Delete()
                deleted += 1
            except:
                continue

    print(f"[OK] åˆ é™¤é‡å¤/å±€éƒ¨é‡å¤çº¿æ®µ {deleted} æ¡ï¼Œä¿ç•™ {total - deleted} æ¡ã€‚")

#æ‰¾å‡ºä¸€ç»„ç›´çº¿æ®µä¸­çš„å­¤ç«‹çº¿æ®µäº§ç”Ÿçš„äº¤ç‚¹


def find_isolated_intersections(LB, tol=0.5):
    """
    æ‰¾å‡ºçº¿æ®µåˆ—è¡¨ LB ä¸­çš„å­¤ç«‹çº¿æ®µï¼Œå¹¶è®¡ç®—å®ƒä»¬ä¸å…¶å®ƒçº¿æ®µçš„æ‰€æœ‰äº¤ç‚¹ã€‚

    ç”¨äºäººå·¥æ ‡è®°çš„é—¨çª—ä½ç½®

    å‚æ•°ï¼š
      LB:   çº¿æ®µåˆ—è¡¨ï¼Œæ¯ä¸ªå…ƒç´ æ˜¯ [(x1,y1,z1), (x2,y2,z2)]
      tol:  ç«¯ç‚¹é‡åˆåˆ¤æ–­å®¹å·®

    è¿”å›ï¼š
      intersections: äº¤ç‚¹åˆ—è¡¨ï¼Œæ¯ä¸ªå…ƒç´ æ˜¯ (x, y, z)
    """
    def segment_intersection(seg1, seg2, tol):
        """
        è®¡ç®—çº¿æ®µ seg1=(A,B) ä¸ seg2=(C,D) çš„äº¤ç‚¹ï¼ˆäºŒç»´ï¼‰ï¼Œ
        è‹¥ç›¸äº¤ä¸”å”¯ä¸€ï¼Œè¿”å› (x, y, z)ï¼Œå¦åˆ™è¿”å› Noneã€‚
        z å– seg1 ç¬¬ä¸€ç«¯ç‚¹çš„ zã€‚
        """
        (x1, y1, z1), (x2, y2, _) = seg1
        (x3, y3, _),  (x4, y4, _) = seg2

        # æ–¹å‘å‘é‡
        r = (x2-x1, y2-y1)
        s = (x4-x3, y4-y3)
        # å‰ç§¯ r Ã— s
        rxs = r[0]*s[1] - r[1]*s[0]
        if abs(rxs) < tol:
            return None  # å¹³è¡Œæˆ–å…±çº¿ï¼Œä¸å¤„ç†
        # è§£ t, u
        qp = (x3-x1, y3-y1)
        t = (qp[0]*s[1] - qp[1]*s[0]) / rxs
        u = (qp[0]*r[1] - qp[1]*r[0]) / rxs
        # åªè€ƒè™‘ä¸¥æ ¼äº¤äºæ®µå†…
        if -tol <= t <= 1+tol and -tol <= u <= 1+tol:
            xi = x1 + t*r[0]
            yi = y1 + t*r[1]
            zi = z1
            return (xi, yi, zi)
        return None

    # 1. æ‰¾å‡ºå­¤ç«‹çº¿æ®µ
    isolated = []
    for i, seg in enumerate(LB):
        p1, p2 = seg
        shared = False
        for j, other in enumerate(LB):
            if i == j:
                continue
            q1, q2 = other
            if same_point(p1, q1, tol) or same_point(p1, q2, tol) \
            or same_point(p2, q1, tol) or same_point(p2, q2, tol):
                shared = True
                break
        if not shared:
            isolated.append(seg)

    # 2. å¯¹æ¯æ ¹å­¤ç«‹çº¿æ®µï¼Œä¸å…¶ä½™çº¿æ®µæ±‚äº¤ç‚¹
    intersections = []
    for seg in isolated:
        for other in LB:
            if other is seg:
                continue
            ip = segment_intersection(seg, other, tol)
            if ip is not None:
                intersections.append(ip)

    #åˆ é™¤å­¤ç«‹çº¿æ®µ20250420
    for seg in isolated:

        seg.Delete()

    return intersections    


##doc.sendcommand("TSpOutline"+chr(13)+"41849.69465957, 12250.50102376, 0"+chr(13)+chr(13)+chr(13))

##doc.sendcommand("TRoflna"+chr(13)+"0"+chr(13))


def get_inner_point_of_polygon(polygon: Polygon):
    """
    è·å–ç»™å®š polygon çš„ä¸€ä¸ªä¿è¯åœ¨å…¶å†…éƒ¨çš„ç‚¹ã€‚

    å‚æ•°ï¼š
        polygon (shapely.geometry.Polygon): ç›®æ ‡å¤šè¾¹å½¢

    è¿”å›ï¼š
        (x, y): å†…éƒ¨ç‚¹åæ ‡å…ƒç»„
    """
    if not isinstance(polygon, Polygon):
        raise ValueError("[é”™è¯¯] è¾“å…¥å¿…é¡»æ˜¯ shapely.geometry.Polygon")

    inner_point = polygon.representative_point()
    return inner_point.x, inner_point.y


#&&%____________________________________________________       è·å–ä¸€ç»„ç›´çº¿æ®µæ‰€æœ‰çš„å°é—­å¤šè¾¹å½¢å’Œå¤–è½®å»“çº¿       ________________
#â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦
# çº¿é¢åˆ†æ - è·å–ä¸€ç»„ç›´çº¿æ®µæ‰€æœ‰çš„å°é—­å¤šè¾¹å½¢å’Œå¤–è½®å»“çº¿

#__________________________________________________________________________________________________________________________
#â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦â€¦
#  ä½¿ç”¨è¯´æ˜

"""
è·å–ä¸€ç»„ç›´çº¿æ®µæ‰€æœ‰çš„å°é—­å¤šè¾¹å½¢å’Œå¤–è½®å»“çº¿

ä¸»è¦å‡½æ•°æœ‰ï¼š
(1)å¯»æ‰¾ç›´çº¿æ®µç»„ä¸­æœ€å³ä¸‹è§’å°é—­å¤šè¾¹æ€§
(2)ä»comè¾¹æˆ–é¡¶ç‚¹åæ ‡åˆ—è¡¨ç”¨PLå¤çº¿ç»˜åˆ¶å¤šè¾¹å½¢
(3)è·å–ä¸€ç»„ç›´çº¿æ®µçš„å¤–è½®å»“çº¿
(4)è·å–æœ€å³ä¸‹è§’çš„å°é—­å¤šè¾¹å½¢ä¸å½±å“å…¶ä»–å°é—­å¤šè¾¹å½¢çš„è¿ç»­è¾¹çš„é¡¶ç‚¹åˆ—è¡¨
(5)è·å–å…¨éƒ¨å°é—­å¤šè¾¹å½¢ï¼Œä½†ä¸å®Œå…¨
(6)è·å–å…¨éƒ¨å°é—­å¤šè¾¹å½¢

"""
def get_room_outline_from_point(x, y, z=0):## è·å–è¾“å…¥ç‚¹æ‰€åœ¨æˆ¿é—´çš„è½®å»“
    """
    è‡ªåŠ¨å‘é€ TSpOutline å‘½ä»¤ï¼Œä»æŒ‡å®šç‚¹è·å–æˆ¿é—´è½®å»“ã€‚

    å‚æ•°:
        x, y, z: ç‚¹çš„åæ ‡ï¼Œz é»˜è®¤ä¸º 0ã€‚

    ä¹Ÿè®¸ä¼šæœ‰ç”¨å¤„ï¼Œèƒ½ä»å¤šè¾¹å½¢å†…ç‚¹è¿…é€Ÿå¾—åˆ°å¤šè¾¹å½¢
        
    """
    try:
        point_str = f"{x},{y},{z}"
        cmd = (
            "TSpOutline" + chr(13) +  # å¯åŠ¨å‘½ä»¤
            point_str + chr(13) +     # è¾“å…¥ç‚¹
            chr(13) +                 # ç¡®è®¤é»˜è®¤è®¾ç½®
            chr(13)                   # ç¡®è®¤ç”Ÿæˆ
        )
        doc.SendCommand(cmd)
        print(f"[OK] å·²è¯·æ±‚è·å–ç‚¹ ({x},{y},{z}) çš„æˆ¿é—´è½®å»“ã€‚")

    except Exception as e:
        print(f"[é”™è¯¯] è·å–æˆ¿é—´è½®å»“å¤±è´¥ï¼š{e}")


def connect_lines_to_polyline_if_closed(lines, tol=0.5):##åˆ¤æ–­ä¸€ç»„é—­åˆç›´çº¿æ®µæ˜¯å¦æ„æˆå°é—­å¤šæ®µçº¿ï¼Œæ˜¯å°±è¿”å›PLå¤çº¿
    """
    åˆ¤æ–­çº¿æ®µæ˜¯å¦é¦–å°¾è¿æ¥æˆé—­åˆå¤šè¾¹å½¢ï¼Œå¦‚æœæ˜¯ï¼Œåˆ™ç»˜åˆ¶PLå¤šæ®µçº¿ã€‚
    
    å‚æ•°:
        lines: AutoCADä¸­é€‰ä¸­çš„ AcDbLine å¯¹è±¡åˆ—è¡¨ã€‚
        tol: å®¹è®¸çš„ç«¯ç‚¹é—­åˆè¯¯å·®ã€‚
    
    è¿”å›:
        å¤šæ®µçº¿å¯¹è±¡ï¼Œæˆ– Noneã€‚
    """

    try:
        # æå–äºŒç»´ç«¯ç‚¹é›†åˆï¼ˆå¿½ç•¥zï¼‰
        segments = []
        for ln in lines:
            try:
                p1 = (ln.StartPoint[0], ln.StartPoint[1])
                p2 = (ln.EndPoint[0], ln.EndPoint[1])
                segments.append((p1, p2))
            except Exception:
                continue

        if not segments:
            print("[è­¦å‘Š]ï¸ æ— æœ‰æ•ˆçº¿æ®µ")
            return None

        # æ„å»ºè¿æ¥é“¾æ¡
        used = set()
        sequence = [segments[0][0]]
        current = segments[0][1]
        used.add(0)

        while True:
            found = False
            for idx, (a, b) in enumerate(segments):
                if idx in used:
                    continue
                if Point(current).distance(Point(a)) < tol:
                    sequence.append(current)
                    current = b
                    used.add(idx)
                    found = True
                    break
                elif Point(current).distance(Point(b)) < tol:
                    sequence.append(current)
                    current = a
                    used.add(idx)
                    found = True
                    break
            if not found:
                break

        # æ£€æŸ¥æ˜¯å¦é—­åˆ
        if Point(current).distance(Point(sequence[0])) > tol:
            print("[é”™è¯¯] çº¿æ®µæœªæ„æˆé—­åˆåŒºåŸŸ")
            return None
        sequence.append(sequence[0])  # é—­åˆç¯

        # æ„é€ äºŒç»´ç‚¹æ•°ç»„
        pts = []
        for pt in sequence:
            pts.extend([pt[0], pt[1]])

        # ç»˜åˆ¶PL
        poly = doc.ModelSpace.AddLightWeightPolyline(vtFloat(pts))
        poly.Closed = True
        print("[OK] æˆåŠŸç»˜åˆ¶å°é—­PLçº¿")
        return poly

    except Exception as e:
        print(f"[é”™è¯¯] Polyline åˆ›å»ºå¤±è´¥: {e}")
        return None


def is_closed_polygon_from_lines(lines, tol=0.5):##åˆ¤æ–­ä¸€ç»„é—­åˆç›´çº¿æ®µæ˜¯å¦æ„æˆå°é—­å¤šæ®µçº¿ï¼Œä¸è¿”å›PLå¤çº¿
    """
    åˆ¤æ–­ä¸€ç»„ AutoCAD ç›´çº¿æ®µæ˜¯å¦é¦–å°¾è¿æ¥å½¢æˆé—­åˆå¤šè¾¹å½¢ã€‚
    
    å‚æ•°:
        lines: AcDbLine ç±»å‹çš„ COM å¯¹è±¡åˆ—è¡¨
        tol: é—­åˆåˆ¤æ–­å®¹å·®ï¼Œå•ä½ä¸CADä¸€è‡´ï¼ˆå¦‚mmï¼‰

    è¿”å›:
        True è¡¨ç¤ºé¦–å°¾é—­åˆå½¢æˆå¤šè¾¹å½¢ï¼ŒFalse å¦åˆ™
    """
    try:
        # æå–äºŒç»´ç«¯ç‚¹ (x, y)
        segments = []
        for ln in lines:
            try:
                p1 = (ln.StartPoint[0], ln.StartPoint[1])
                p2 = (ln.EndPoint[0], ln.EndPoint[1])
                segments.append((p1, p2))
            except Exception:
                continue

        if not segments:
            return False

        # æ„é€ é¦–å°¾è¿æ¥é“¾
        used = set()
        sequence = [segments[0][0]]
        current = segments[0][1]
        used.add(0)

        while True:
            found = False
            for idx, (a, b) in enumerate(segments):
                if idx in used:
                    continue
                if Point(current).distance(Point(a)) < tol:
                    sequence.append(current)
                    current = b
                    used.add(idx)
                    found = True
                    break
                elif Point(current).distance(Point(b)) < tol:
                    sequence.append(current)
                    current = a
                    used.add(idx)
                    found = True
                    break
            if not found:
                break

        # åˆ¤æ–­æ˜¯å¦å›åˆ°èµ·ç‚¹ï¼ˆé—­åˆï¼‰
        if Point(current).distance(Point(sequence[0])) < tol and len(used) == len(segments):
            return True
        else:
            return False

    except Exception as e:
        print(f"[é”™è¯¯] åˆ¤æ–­å¤±è´¥: {e}")
        return False

def same_point(p1, p2, tol=0.5):
    """åˆ¤æ–­ä¸¤ä¸ªç‚¹æ˜¯å¦åœ¨å®¹å·®èŒƒå›´å†…ç›¸åŒï¼ˆåªæ¯”è¾ƒ Xã€Y åæ ‡ï¼‰"""
    return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol


def same_line(ln1, ln2, tol=0.5):
    """
    åˆ¤æ–­ä¸¤æ¡çº¿æ®µ ln1 å’Œ ln2 æ˜¯å¦â€œç›¸åŒâ€
    çº¿æ®µè¢«è®¤ä¸ºç›¸åŒçš„æ¡ä»¶æ˜¯ï¼š
      ln1 çš„ StartPoint ä¸ ln2 çš„ StartPoint è¿‘ä¼¼ç›¸åŒä¸” ln1 çš„ EndPoint ä¸ ln2 çš„ EndPoint è¿‘ä¼¼ç›¸åŒï¼Œ
      æˆ–è€… ln1 çš„ StartPoint ä¸ ln2 çš„ EndPoint è¿‘ä¼¼ç›¸åŒä¸” ln1 çš„ EndPoint ä¸ ln2 çš„ StartPoint è¿‘ä¼¼ç›¸åŒã€‚
    """
    p1 = tuple(ln1.StartPoint)
    p2 = tuple(ln1.EndPoint)
    q1 = tuple(ln2.StartPoint)
    q2 = tuple(ln2.EndPoint)

    return (same_point(p1, q1, tol) and same_point(p2, q2, tol)) or \
           (same_point(p1, q2, tol) and same_point(p2, q1, tol))


def calculate_absolute_angle(line, P, tol=0.5):
    """
    è®¡ç®—çº¿æ®µï¼ˆlineï¼‰ä»ç‚¹ P å‡ºå‘çš„ç»å¯¹è§’åº¦ï¼ˆ0-360Â°ï¼‰
    å¦‚æœ line çš„èµ·ç‚¹ç­‰äº Pï¼Œåˆ™è¿”å›ä» P åˆ°ç»ˆç‚¹çš„è§’åº¦ï¼Œå¦åˆ™è¿”å›ä» P åˆ°èµ·ç‚¹çš„è§’åº¦ã€‚
    """
    sp = tuple(line.StartPoint)
    ep = tuple(line.EndPoint)
    if same_point(sp, P, tol):
        dx = ep[0] - P[0]
        dy = ep[1] - P[1]
    else:
        dx = sp[0] - P[0]
        dy = sp[1] - P[1]
    return math.degrees(math.atan2(dy, dx)) % 360

def calculate_relative_angle(line, P, current_line, tol=0.5):
    """
    è®¡ç®—å½“å‰å‚è€ƒçº¿ï¼ˆcurrent_lineï¼‰ä¸å€™é€‰çº¿æ®µï¼ˆlineï¼‰ä¹‹é—´çš„ç›¸å¯¹è§’åº¦ï¼Œ
    è§’åº¦æ˜¯ä»¥å…±ç‚¹ P ä¸ºä¸­å¿ƒï¼Œå½“å‰çº¿ä» P åˆ°å…¶é P çš„ç«¯ç‚¹çš„ç»å¯¹è§’åº¦ä¸ºåŸºå‡†ï¼Œ
    è®¡ç®— candidate line ä¸è¯¥åŸºå‡†è§’åº¦ä¹‹é—´é¡ºæ—¶é’ˆæˆ–é€†æ—¶é’ˆçš„è§’åº¦å·®ï¼ˆé€†æ—¶é’ˆæ–¹å‘ï¼‰ã€‚
    ç»“æœä¸º 0 åˆ° 360 ä¹‹é—´çš„æ•°å€¼ã€‚
    """
    sp_current = tuple(current_line.StartPoint)
    ep_current = tuple(current_line.EndPoint)
    sp = tuple(line.StartPoint)
    ep = tuple(line.EndPoint)

    # é€‰æ‹©å½“å‰çº¿æ®µä¸­ä¸ç­‰äº P çš„ç«¯ç‚¹ä½œä¸ºå‚è€ƒ
    if same_point(sp_current, P, tol):
        current_point = ep_current
    else:
        current_point = sp_current

    def angle(p1, p2):
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        return math.degrees(math.atan2(dy, dx)) % 360

    angle_current = angle(P, current_point)
    # å¯¹å€™é€‰çº¿æ®µï¼Œé€‰æ‹©ä¸ç­‰äº P çš„ç«¯ç‚¹
    if same_point(sp, P, tol):
        target_point = ep
    else:
        target_point = sp
    angle_target = angle(P, target_point)
    angle_diff = (angle_target - angle_current) % 360
    return angle_diff

#####################
# å‡½æ•°ï¼šæŸ¥æ‰¾ç»™å®šç‚¹Pç»è¿‡çš„çº¿æ®µï¼ŒæŒ‰ç…§ç»å¯¹è§’åº¦æ’åº

def find_lines_angle(lines, P, tol=0.5):
    """
    æŸ¥æ‰¾ä¸æŒ‡å®šç‚¹ P å…±ç«¯ç‚¹çš„æ‰€æœ‰çº¿æ®µï¼Œå¹¶æŒ‰ä» P å‡ºå‘ç¦»å¼€çš„ç»å¯¹å‡ ä½•è§’åº¦æ’åºã€‚

    å‚æ•°:
        lines: ç›´çº¿æ®µå¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡è¦æ±‚å…·å¤‡ StartPoint, EndPoint, Handle å±æ€§ã€‚
        P: ä¸‰å…ƒç»„ (x, y, z)ï¼Œç›®æ ‡å…±ç‚¹ã€‚
        tol: åˆ¤æ–­å…±ç‚¹çš„å®¹å·®ï¼ˆä»…æ¯”è¾ƒ x å’Œ y åæ ‡ï¼‰ã€‚

    è¿”å›:
        æŒ‰ç»å¯¹è§’åº¦ä»å°åˆ°å¤§æ’åºçš„å…±ç‚¹çº¿æ®µåˆ—è¡¨ã€‚
    """
    shared_lines = []
    P = tuple(P)
    print(f"è°ƒè¯•ï¼šç›®æ ‡å…±ç‚¹ P = {P}")
    
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            dx_sp = abs(sp[0] - P[0])
            dy_sp = abs(sp[1] - P[1])
            dx_ep = abs(ep[0] - P[0])
            dy_ep = abs(ep[1] - P[1])
            print(f"çº¿æ®µ {ln.Handle}: sp={sp} å·®å€¼=({dx_sp:.4f},{dy_sp:.4f}), ep={ep} å·®å€¼=({dx_ep:.4f},{dy_ep:.4f})")
            if (dx_sp <= tol and dy_sp <= tol) or (dx_ep <= tol and dy_ep <= tol):
                shared_lines.append(ln)
        except Exception as e:
            print(f"[è­¦å‘Š]ï¸ è·³è¿‡æ— æ•ˆçº¿æ®µ {getattr(ln, 'Handle', 'æœªçŸ¥')} : {e}")
            continue

    shared_lines.sort(key=lambda ln: calculate_absolute_angle(ln, P, tol))
    print("è°ƒè¯•ï¼šæŒ‰ç»å¯¹è§’åº¦æ’åºåçš„å…±ç‚¹çº¿æ®µï¼š")
    for ln in shared_lines:
        print(f"  çº¿æ®µ {ln.Handle}ï¼šè§’åº¦ = {calculate_absolute_angle(ln, P, tol):.2f}Â°")
    return shared_lines

#####################
# å‡½æ•°ï¼šæŸ¥æ‰¾ä¸På…±ç‚¹çš„çº¿æ®µï¼ŒæŒ‰ç…§ä¸å½“å‰çº¿æ®µçš„ç›¸å¯¹è§’åº¦æ’åº

def find_lines_sharing_point(lines, P, current_line, tol=0.5):
    """
    æŸ¥æ‰¾ä¸æŒ‡å®šç‚¹ P å…±ç«¯ç‚¹çš„æ‰€æœ‰çº¿æ®µï¼Œå¹¶æŒ‰ä» current_line é€†æ—¶é’ˆæ—‹è½¬åˆ°å…¶ä»–çº¿æ®µçš„ç›¸å¯¹è§’åº¦æ’åºã€‚
    å…¶ä¸­ current_line çš„è§’åº¦å®šä¹‰ä¸º 0Â°ã€‚
    
    å‚æ•°:
        lines: ç›´çº¿æ®µå¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡è¦æ±‚å…·å¤‡ StartPoint, EndPoint, Handle å±æ€§ã€‚
        P: å…±ç‚¹ï¼Œä¸‰å…ƒç»„ (x, y, z)ã€‚
        current_line: å½“å‰å‚è€ƒçº¿æ®µï¼Œå¯¹åº”è§’åº¦å®šä¹‰ä¸º 0Â°ã€‚
        tol: åˆ¤æ–­å…±ç‚¹çš„å®¹å·®ï¼ˆåªæ¯”è¾ƒ x å’Œ y åæ ‡ï¼‰ã€‚
        
    è¿”å›:
        æŒ‰ç›¸å¯¹è§’åº¦ä»å°åˆ°å¤§æ’åºçš„ç»è¿‡ P çš„çº¿æ®µåˆ—è¡¨ï¼ˆcurrent_line äº¦åŒ…æ‹¬å…¶ä¸­ï¼‰ã€‚
    """
    shared_lines = []
    P = tuple(P)
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            if (abs(sp[0]-P[0])<=tol and abs(sp[1]-P[1])<=tol) or (abs(ep[0]-P[0])<=tol and abs(ep[1]-P[1])<=tol):
                shared_lines.append(ln)
        except Exception as e:
            print(f"[è­¦å‘Š]ï¸ è·³è¿‡æ— æ•ˆçº¿æ®µ {getattr(ln, 'Handle', 'æœªçŸ¥')} : {e}")
            continue
    # æ ¹æ®ä»å½“å‰çº¿æ®µæ—‹è½¬ï¼ˆé€†æ—¶é’ˆï¼‰çš„ç›¸å¯¹è§’åº¦æ’åº
    shared_lines.sort(key=lambda ln: calculate_relative_angle(ln, P, current_line, tol))
    return shared_lines

#####################
# å‡½æ•°ï¼šæ ¹æ®å½“å‰çº¿æ®µå’Œå…±ç‚¹ Pï¼Œé€‰æ‹©ä¸‹ä¸€æ¡åç»§çº¿æ®µï¼ˆé€‰æ‹©ç›¸å¯¹è§’åº¦æœ€å¤§çš„é‚£æ¡ï¼‰ï¼Œè¿”å› (åç»§çº¿æ®µ, æ–°å…±ç‚¹)

def find_successor_line_max(current_line, lines, P, tol=0.5):
    """
    åœ¨ç»™å®šå…±ç‚¹ P å¤„ï¼Œæ’é™¤å½“å‰çº¿æ®µï¼ˆcurrent_lineï¼‰åï¼Œ
    é€‰æ‹©åœ¨è¯¥ç‚¹å¤„ä¸å½“å‰çº¿æ®µç›¸å¯¹æ—‹è½¬è§’åº¦æœ€å¤§çš„é‚£æ¡çº¿æ®µä½œä¸ºåç»§çº¿æ®µï¼Œ
    è¿”å› (åç»§çº¿æ®µ, æ–°å…±ç‚¹)ã€‚

    å‚æ•°ï¼š
      current_line: å½“å‰çº¿æ®µå¯¹è±¡ã€‚
      lines: æ‰€æœ‰ç›´çº¿æ®µå¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡å¿…é¡»å…·å¤‡ StartPoint, EndPoint, Handle å±æ€§ã€‚
      P: å…±ç‚¹ï¼ˆä¸‰å…ƒç»„ï¼‰ã€‚
      tol: å…±ç‚¹åˆ¤æ–­çš„å®¹å·®ï¼ˆé»˜è®¤ 0.5ï¼‰ã€‚

    è¿”å›ï¼š
      (åç»§çº¿æ®µ, æ–°å…±ç‚¹)ï¼Œè‹¥æ‰¾ä¸åˆ°åˆé€‚çš„åç»§çº¿æ®µï¼Œåˆ™è¿”å› (None, P)ã€‚
    """
    # è°ƒç”¨ find_lines_sharing_point è·å–æ‰€æœ‰ç»è¿‡ P çš„çº¿æ®µï¼Œå¹¶ä½¿ç”¨ current_line çš„ç›¸å¯¹è§’åº¦æ’åº
    candidates = find_lines_sharing_point(lines, P, current_line, tol)
    # æ’é™¤å½“å‰çº¿æ®µ
    candidates = [ln for ln in candidates if ln.Handle != current_line.Handle]

    if not candidates:
        print(f"[é”™è¯¯] åœ¨ç‚¹ {P} å¤„æ‰¾ä¸åˆ°é™¤å½“å‰çº¿å¤–çš„å€™é€‰åç»§çº¿æ®µ")
        return None, P

    best_line = None
    max_angle = -1
    new_point = P
    for ln in candidates:
        relative_angle = calculate_relative_angle(ln, P, current_line, tol)
        if relative_angle > max_angle:
            max_angle = relative_angle
            if same_point(tuple(ln.StartPoint), P, tol):
                new_point = tuple(ln.EndPoint)
            else:
                new_point = tuple(ln.StartPoint)
            best_line = ln

    if best_line is None:
        print(f"[é”™è¯¯] æ²¡æœ‰æ‰¾åˆ°æœ‰æ•ˆçš„åç»§çº¿æ®µ")
        return None, P

    print(f"é€‰ä¸­åç»§çº¿æ®µ {best_line.Handle}ï¼Œæ–°å…±ç‚¹ {new_point}")
    return best_line, new_point

#&&%#####################
# è¾…åŠ©å‡½æ•°ï¼šä»æ‰€æœ‰çº¿æ®µä¸­æ‰¾å‡ºæœ€å³ä¸‹è§’çš„ç‚¹
def find_rightbottom_point(lines, tol=0.5):
    """
    ä»æ‰€æœ‰çº¿æ®µç«¯ç‚¹ä¸­ï¼Œæ‰¾å‡º y å€¼æœ€å°çš„ç‚¹ï¼›è‹¥æœ‰å¤šä¸ªï¼Œåˆ™é€‰ x æœ€å¤§çš„ç‚¹ä½œä¸ºæœ€å³ä¸‹è§’ç‚¹ã€‚
    """
    all_points = []
    for line in lines:
        if hasattr(line, "StartPoint") and hasattr(line, "EndPoint"):
            all_points.append(tuple(line.StartPoint))
            all_points.append(tuple(line.EndPoint))
    if not all_points:
        return None
    min_y = min(p[1] for p in all_points)
    candidates = [p for p in all_points if abs(p[1]-min_y) <= tol]
    rb = max(candidates, key=lambda p: p[0])
    print(f"[OK] æœ€å³ä¸‹è§’ç‚¹ä¸ºï¼š{rb}")
    return rb



#  ä¸»å‡½æ•°
#  (1)
# å¯»æ‰¾ç›´çº¿æ®µç»„ä¸­æœ€å³ä¸‹è§’å°é—­å¤šè¾¹æ€§

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°


"""
è¾…åŠ©å‡½æ•°

same_point(p1, p2, tol) åˆ¤æ–­ä¸¤ç‚¹æ˜¯å¦åœ¨å®¹å·®èŒƒå›´å†…ç›¸åŒ

calculate_absolute_angle(line, P, tol) è®¡ç®—ä»ç‚¹ P å‡ºå‘åˆ°æŸæ¡çº¿æ®µï¼ˆå–ä¸ P ä¸ç›¸åŒçš„ç«¯ç‚¹ï¼‰çš„ç»å¯¹è§’åº¦

calculate_relative_angle(line, P, current_line, tol) è®¡ç®—å½“å‰çº¿ï¼ˆä»¥ P ä¸ºèµ·ç‚¹ï¼Œé€‰å–ä¸ P ä¸ç›¸åŒçš„ç«¯ç‚¹ï¼‰

å’Œå€™é€‰çº¿ï¼ˆä»¥ P ä¸ºèµ·ç‚¹ï¼‰çš„è§’åº¦å·®ï¼ˆé€†æ—¶é’ˆæ–¹å‘ï¼Œç»“æœåœ¨ 0ï½360 ä¹‹é—´ï¼‰

å‡½æ•° find_lines_angle
ç”¨äºåˆå§‹é˜¶æ®µï¼šç»™å®šä¸€ä¸ªå…±ç‚¹ Pï¼Œæ‰¾åˆ°æ‰€æœ‰ç»è¿‡è¯¥ç‚¹çš„çº¿æ®µï¼Œå¹¶æŒ‰å®ƒä»¬çš„ç»å¯¹è§’åº¦æ’åºï¼ˆä»å°åˆ°å¤§ï¼‰ã€‚

å‡½æ•° find_lines_sharing_point
ç»™å®šå…±ç‚¹ P ä¸ä¸€ä¸ªâ€œå½“å‰çº¿æ®µâ€ï¼ˆä½œä¸ºå‚è€ƒï¼‰ï¼Œè¿”å›æ‰€æœ‰ç»è¿‡ P çš„çº¿æ®µï¼Œå¹¶æŒ‰ç…§ä»å½“å‰çº¿æ®µå‡ºå‘æ—‹è½¬ï¼ˆé€†æ—¶é’ˆï¼‰çš„ç›¸å¯¹è§’åº¦æ’åºã€‚

å‡½æ•° find_successor_line_max
æ ¹æ®å½“å‰çº¿æ®µå’Œå…±ç‚¹ Pï¼Œä»é€šè¿‡ P çš„å€™é€‰çº¿ä¸­é€‰å–â€œè½¬è§’â€æœ€å¤§çš„ä½œä¸ºåç»§çº¿æ®µï¼Œå¹¶è¿”å›åç»§çº¿æ®µåŠè¯¥çº¿æ®µå¦ä¸€ç«¯çš„æ–°å…±ç‚¹ã€‚

å‡½æ•° find_rightbottom_closed_polygon
åˆ©ç”¨ä¸Šè¿°å‡½æ•°æ„é€ å°é—­å¤šè¾¹å½¢ã€‚åˆå§‹æ—¶å…ˆæ ¹æ®æ‰€æœ‰çº¿æ®µç«¯ç‚¹ç¡®å®šâ€œæœ€å³ä¸‹è§’â€ç‚¹ï¼ˆå‡½æ•° find_rightbottom_pointï¼‰ï¼Œç„¶ååœ¨è¯¥ç‚¹å¤„è°ƒç”¨

find_lines_angle å¾—åˆ°æ‰€æœ‰ç»è¿‡è¯¥ç‚¹çš„çº¿æ®µï¼Œé€‰å–ç»å¯¹è§’åº¦æœ€å°çš„é‚£ä¸€æ ¹ä½œä¸ºç¬¬ä¸€æ¡è¾¹ï¼ˆåˆå§‹ current_lineï¼‰ï¼Œå†ä¾æ¬¡è°ƒç”¨

find_successor_line_max æ¨è¿›å¤šè¾¹å½¢ç›´åˆ°é—­åˆæˆ–è¾¾åˆ°æœ€å¤§æ­¥æ•°ã€‚


"""

def find_rightbottom_closed_polygon(lines, tol=0.5, max_steps=50):
    """
    åˆ©ç”¨æ‰€æœ‰çº¿æ®µæ„é€ å°é—­å¤šè¾¹å½¢ï¼š
      1. å…ˆæŸ¥æ‰¾æ‰€æœ‰çº¿æ®µç«¯ç‚¹ä¸­çš„æœ€å³ä¸‹è§’ç‚¹ï¼ˆRBï¼‰ã€‚
      2. åœ¨ RB å¤„è·å–æ‰€æœ‰ç»è¿‡è¯¥ç‚¹çš„çº¿æ®µï¼ˆæŒ‰ç…§ç»å¯¹è§’åº¦æ’åºï¼‰ï¼Œé€‰æ‹©ç»å¯¹è§’åº¦æœ€å°çš„é‚£æ ¹ä½œä¸ºåˆå§‹è¾¹ã€‚
      3. ä»¥è¯¥åˆå§‹è¾¹ä¸ºç¬¬ä¸€æ¡è¾¹ï¼Œä¹‹åä¾æ¬¡åˆ©ç”¨ find_successor_line_max é€‰æ‹©åç»§è¾¹ï¼Œ
         ç›´åˆ°æ–°å…±ç‚¹å›åˆ°åˆå§‹ç‚¹ï¼ˆé—­åˆï¼‰æˆ–è¶…è¿‡æœ€å¤§æ­¥æ•°ã€‚
         
    è¿”å›ï¼š
      æ„æˆå°é—­å¤šè¾¹å½¢çš„ç‚¹åˆ—è¡¨ï¼ˆä¾æ¬¡ä¸ºæ¯æ¡è¾¹çš„ç»ˆç‚¹ï¼‰ï¼Œè‹¥æ— æ³•æ„æˆåˆ™è¿”å› Noneã€‚
    """
    # å®šä½æœ€å³ä¸‹è§’ç‚¹ RB
    rb = find_rightbottom_point(lines, tol)
    if rb is None:
        print("[é”™è¯¯] æ— å³ä¸‹è§’ç‚¹")
        return None

    # åˆå§‹å…±ç‚¹
    current_point = rb
    # ä» RB å¤„æŒ‰ç»å¯¹è§’åº¦æ’åºå–å€™é€‰çº¿æ®µ
    candidates = find_lines_angle(lines, rb, tol)
    if not candidates:
        print(f"[é”™è¯¯] åœ¨å³ä¸‹è§’ç‚¹ {rb} å¤„æ²¡æœ‰æ‰¾åˆ°ç»è¿‡çš„çº¿æ®µ")
        return None
    # é€‰æ‹©ç»å¯¹è§’åº¦æœ€å°çš„çº¿æ®µä½œä¸ºåˆå§‹è¾¹
    current_line = candidates[0]
    print(f"è°ƒè¯•ï¼šé€‰ä¸­åˆå§‹çº¿æ®µ {current_line.Handle}ï¼Œç»å¯¹è§’åº¦ = {calculate_absolute_angle(current_line, rb, tol):.2f}Â°")

    # å¾—åˆ°åˆå§‹è¾¹çš„å¦ä¸€ç«¯
    sp = tuple(current_line.StartPoint)
    ep = tuple(current_line.EndPoint)
    if same_point(sp, rb, tol):
        next_point = ep
    else:
        next_point = sp

    polygon_points = [rb, next_point]
    visited_handles = {current_line.Handle}
    current_point = next_point
    steps = 1

    while steps < max_steps:
        # è°ƒç”¨ find_successor_line_max å¾—åˆ°ä¸‹ä¸€æ¡çº¿æ®µåŠå…¶å¦ä¸€ç«¯çš„ç‚¹
        successor, new_point = find_successor_line_max(current_line, lines, current_point, tol)
        if successor is None:
            print("[é”™è¯¯] æ— åç»§çº¿æ®µï¼Œæ„é€ å¤±è´¥")
            return None
        # æ£€æŸ¥æ˜¯å¦é—­åˆï¼ˆæ–°å…±ç‚¹ä¸èµ·å§‹ç‚¹æ¥è¿‘ï¼‰
        if same_point(new_point, rb, tol):
            polygon_points.append(rb)
            print(f"[OK] æˆåŠŸæ„å»ºå°é—­å¤šè¾¹å½¢ï¼Œæ­¥æ•° = {steps}")
            return polygon_points

        if successor.Handle in visited_handles:
            print("ğŸ” æ£€æµ‹åˆ°é‡å¤çº¿æ®µï¼Œæ„é€ å¤±è´¥")
            return None

        polygon_points.append(new_point)
        visited_handles.add(successor.Handle)
        current_line = successor
        current_point = new_point
        steps += 1

    print("[è­¦å‘Š]ï¸ è¾¾åˆ°æœ€å¤§æ­¥æ•°ï¼Œæœªèƒ½æ„é€ å‡ºé—­åˆå¤šè¾¹å½¢")
    return None


# ä»comè¾¹æˆ–é¡¶ç‚¹åæ ‡åˆ—è¡¨ç”¨PLå¤çº¿ç»˜åˆ¶å¤šè¾¹å½¢


def draw_polygon_as_polyline(polygon, layer_name="æµ‹è¯•è¾…åŠ©", tol=0.5):
    """
    å°†æ„é€ çš„å¤šè¾¹å½¢ï¼ˆpolygonï¼‰è½¬æ¢ä¸ºé¡¶ç‚¹åºåˆ—ï¼Œå¹¶åœ¨å½“å‰ AutoCAD æ–‡æ¡£ doc çš„ ModelSpace ä¸­æ·»åŠ 
    ä¸€ä¸ªå¤šæ®µçº¿ï¼ˆPLINEï¼‰ã€‚å¦‚æœé¡¶ç‚¹åºåˆ—çš„é¦–å°¾ä¸¤ç‚¹é‡åˆï¼Œåˆ™ç»˜åˆ¶é—­åˆå¤šæ®µçº¿ï¼Œå¦åˆ™ç»˜åˆ¶å¼€æ”¾å¤šæ®µçº¿ã€‚
    
    å‚æ•°:
      polygon:
        1. å¦‚æœæ˜¯ç”±çº¿æ®µç»„æˆçš„åˆ—è¡¨ï¼Œæ¯ä¸ªå…ƒç´ è¦æ±‚å…·å¤‡ StartPoint å’Œ EndPoint å±æ€§ï¼Œ
           åˆ™æŒ‰ç…§è¿™äº›çº¿æ®µçš„ç«¯ç‚¹é¡ºåºæ„é€ é¡¶ç‚¹åºåˆ—ã€‚
        2. å¦‚æœæ˜¯é¡¶ç‚¹åˆ—è¡¨ï¼Œä¾‹å¦‚ [(x, y, z), (x, y, z), ...]ï¼Œåˆ™ç›´æ¥ä½¿ç”¨ã€‚
      layer_name: ç»˜åˆ¶å¤šæ®µçº¿æ‰€åœ¨çš„å›¾å±‚åç§°ï¼ˆé»˜è®¤ä¸º "æµ‹è¯•è¾…åŠ©"ï¼‰ã€‚
      tol: åˆ¤æ–­ç‚¹æ˜¯å¦é‡åˆçš„å®¹å·®å€¼ï¼ˆä»…æ¯”è¾ƒ x å’Œ y åæ ‡ï¼‰ã€‚
      
    è¿”å›:
      æˆåŠŸæ—¶è¿”å›åˆ›å»ºçš„å¤šæ®µçº¿å¯¹è±¡ï¼ˆPLINEï¼‰ï¼Œå¦åˆ™è¿”å› Noneã€‚
    """

    # å†…éƒ¨å‡½æ•°ï¼šåˆ¤æ–­ä¸¤ä¸ªç‚¹æ˜¯å¦è¿‘ä¼¼ç›¸ç­‰ï¼ˆä»…æ¯”è¾ƒ x å’Œ y åæ ‡ï¼‰
##    def same_point(p1, p2, tol=tol):
##        return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol

    if not polygon:
        print("[é”™è¯¯] æœªæä¾›æœ‰æ•ˆçš„ polygon æ•°æ®")
        return None

    vertices = []
    is_closed = False  # æ˜¯å¦ç»˜åˆ¶é—­åˆå¤šæ®µçº¿

    # åˆ¤æ–­ polygon æ˜¯é¡¶ç‚¹åˆ—è¡¨è¿˜æ˜¯çº¿æ®µåˆ—è¡¨
    if isinstance(polygon[0], (tuple, list)):
        # åˆ¤æ–­ä¼ å…¥æ˜¯å¦ä¸ºé¡¶ç‚¹åˆ—è¡¨ï¼šæ£€æŸ¥ç¬¬ä¸€ä¸ªå…ƒç´ ä¸º tuple/list ä¸”é•¿åº¦>=3
        if len(polygon[0]) < 3:
            print("[é”™è¯¯] é¡¶ç‚¹æ•°æ®æ ¼å¼é”™è¯¯")
            return None
        # ç›´æ¥ä½¿ç”¨é¡¶ç‚¹åˆ—è¡¨ï¼ˆè½¬æ¢ä¸º tuple å½¢å¼ï¼‰
        vertices = [tuple(pt) for pt in polygon]
        # å¦‚æœé¦–å°¾å·²ç»é‡åˆï¼Œåˆ™è§†ä¸ºé—­åˆï¼›å¦åˆ™ä¿æŒå¼€æ”¾ï¼ˆä¸è‡ªåŠ¨è¡¥å……é¦–ç‚¹ï¼‰
        if same_point(vertices[0], vertices[-1], tol):
            is_closed = True
        else:
            is_closed = False
    else:
        # å‡å®š polygon æ˜¯çº¿æ®µåˆ—è¡¨ï¼Œæ¯ä¸ªå…ƒç´ å…·æœ‰ StartPoint å’Œ EndPoint å±æ€§
        first_line = polygon[0]
        start_pt = tuple(first_line.StartPoint)
        last_line = polygon[-1]
        # åˆ¤æ–­å“ªä¸€ç«¯ä¸èµ·å§‹ç‚¹ç›¸è¿ï¼Œä½œä¸ºåˆå§‹ç‚¹
        if same_point(start_pt, tuple(last_line.StartPoint), tol) or same_point(start_pt, tuple(last_line.EndPoint), tol):
            # start_pt å¯ä»¥ä½œä¸ºèµ·ç‚¹
            pass
        else:
            # å¦åˆ™é€‰ç”¨ç¬¬ä¸€æ¡çº¿æ®µçš„ EndPointä½œä¸ºèµ·å§‹ç‚¹
            start_pt = tuple(first_line.EndPoint)
        vertices.append(start_pt)
        current_pt = start_pt
        # éå†æ¯æ¡çº¿æ®µæ„é€ é¡¶ç‚¹åºåˆ—
        for line in polygon:
            sp = tuple(line.StartPoint)
            ep = tuple(line.EndPoint)
            if same_point(current_pt, sp, tol):
                next_pt = ep
            elif same_point(current_pt, ep, tol):
                next_pt = sp
            else:
                print(f"[é”™è¯¯] çº¿æ®µ {line.Handle} ä¸å½“å‰ç‚¹ {current_pt} æœªè¿æ¥ï¼Œæ„é€ å¤šè¾¹å½¢å¤±è´¥")
                return None
            vertices.append(next_pt)
            current_pt = next_pt
        # æ£€æŸ¥æ˜¯å¦é—­åˆï¼ˆé¦–å°¾é‡åˆï¼‰
        if same_point(vertices[0], vertices[-1], tol):
            is_closed = True
        else:
            is_closed = False

    # è¾“å‡ºè°ƒè¯•ä¿¡æ¯ï¼šæ‰“å°é¡¶ç‚¹åºåˆ—
    print("è°ƒè¯•ï¼šå¤šæ®µçº¿é¡¶ç‚¹åºåˆ—ï¼š")
    for i, pt in enumerate(vertices):
        print(f"  é¡¶ç‚¹ {i}: {pt}")

    # å°†é¡¶ç‚¹åºåˆ—è½¬æ¢ä¸ºä¸€ç»´åæ ‡æ•°ç»„ï¼š[x1, y1, z1, x2, y2, z2, â€¦]
    coords = []
    for pt in vertices:
        coords.extend([pt[0], pt[1], pt[2]])
    coords_tuple = tuple(coords)
    coords_variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, coords_tuple)

    # ç¡®ä¿å›¾å±‚å­˜åœ¨ï¼Œå¦åˆ™åˆ›å»ºæ–°å›¾å±‚
    try:
        _ = doc.Layers.Item(layer_name)
    except Exception:
        doc.Layers.Add(layer_name)

    # åœ¨ ModelSpace ä¸­æ·»åŠ å¤šæ®µçº¿å¹¶è®¾ç½®ç›¸åº”å±æ€§
    try:
        ms = doc.ModelSpace
        polyline = ms.AddPolyline(coords_variant)
        polyline.Closed = is_closed
        polyline.Layer = layer_name
        # å¯é€‰è®¾ç½®é¢œè‰²å’Œå®½åº¦ï¼ŒæŒ‰éœ€è¦è°ƒæ•´
        polyline.color = 1
        polyline.ConstantWidth = 20
        polyline.Update()
        doc.Regen(0)
        print(f"[OK] æˆåŠŸåœ¨å›¾å±‚ '{layer_name}' ç»˜åˆ¶å¤šæ®µçº¿, Closed={is_closed}")
        return polyline
    except Exception as e:
        print("[é”™è¯¯] ç»˜åˆ¶å¤šæ®µçº¿å¤±è´¥ï¼š", e)
        return None



#&&%###å¤–è½®å»“çº¿

# -----------------------------------------------------
# è¾…åŠ©å‡½æ•°ï¼šåˆ¤æ–­ä¸¤ä¸ªç‚¹æ˜¯å¦è¿‘ä¼¼ç›¸ç­‰ï¼ˆä»…æ¯”è¾ƒ x å’Œ y åæ ‡ï¼‰
def is_nearly_equal(p1, p2, tol):
    return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol



# -----------------------------------------------------
# å¯»æ‰¾åç»§çº¿æ®µï¼šåœ¨å…±ç‚¹ P å¤„ï¼Œä»å½“å‰è¾¹ä¹‹å¤–çš„å€™é€‰è¾¹ä¸­é€‰æ‹©ç›¸å¯¹è§’åº¦æœ€å°çš„è¾¹
def find_successor_line_min(current_line, lines, P, tol=0.5):
    # å…ˆè·å–å…±ç‚¹ P çš„æ‰€æœ‰çº¿æ®µï¼Œä½¿ç”¨ find_lines_sharing_point é€»è¾‘ï¼Œä½†è¿™é‡Œç›´æ¥é‡‡ç”¨ find_lines_angle æ’åºåç­›é€‰
    candidates = []
    P = tuple(P)
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            if (abs(sp[0]-P[0]) <= tol and abs(sp[1]-P[1]) <= tol) or (abs(ep[0]-P[0]) <= tol and abs(ep[1]-P[1]) <= tol):
                # æ’é™¤å½“å‰çº¿æ®µè‡ªèº«
                if ln.Handle == current_line.Handle:
                    continue
                candidates.append(ln)
        except Exception as e:
            continue

    if not candidates:
        print(f"[é”™è¯¯] åœ¨ç‚¹ {P} å¤„æ‰¾ä¸åˆ°å€™é€‰åç»§çº¿æ®µ")
        return None, P

    best_line = None
    min_angle = 360
    new_point = P
    for candidate in candidates:
        angle_diff = calculate_relative_angle(candidate, P, current_line, tol)
        # é€‰æ‹©ç›¸å¯¹è§’åº¦æœ€å°çš„å€™é€‰
        if angle_diff < min_angle:
            min_angle = angle_diff
            best_line = candidate
            # æ›´æ–°æ–°å…±ç‚¹ï¼šé€‰æ‹© candidate ä¸­ä¸ç­‰äº P çš„ç«¯ç‚¹
            sp_cand = tuple(candidate.StartPoint)
            ep_cand = tuple(candidate.EndPoint)
            if is_nearly_equal(sp_cand, P, tol):
                new_point = ep_cand
            else:
                new_point = sp_cand

    if best_line is None:
        print(f"[é”™è¯¯] æ²¡æœ‰æ‰¾åˆ°åˆé€‚çš„åç»§çº¿æ®µ")
        return None, P

    print(f"é€‰ä¸­åç»§çº¿æ®µ {best_line.Handle}ï¼Œæ–°å…±ç‚¹ {new_point}ï¼Œç›¸å¯¹è½¬è§’ = {min_angle:.2f}Â°")
    return best_line, new_point



#  ä¸»å‡½æ•°
#  (3)
# è·å–ä¸€ç»„ç›´çº¿æ®µçš„å¤–è½®å»“çº¿

def get_outer_contour(lines, tol=0.5, max_steps=50):
    """
    è·å–ä¸€ç»„ç›´çº¿æ®µçš„å¤–è½®å»“çº¿
    è§„åˆ™ï¼š
      1. è®¡ç®—æ‰€æœ‰çº¿æ®µç«¯ç‚¹çš„æœ€å³ä¸‹è§’ç‚¹ Pï¼ˆç»å¯¹æ–¹æ³•ï¼‰ã€‚
      2. åœ¨ P å¤„ï¼ŒæŒ‰ç»å¯¹è§’åº¦æ’åºï¼Œé€‰æ‹©ç»å¯¹è§’åº¦æœ€å°çš„è¾¹ä½œä¸ºç¬¬ä¸€æ¡è¾¹ï¼›
      3. ä»¥å½“å‰è¾¹çš„å¦ä¸€ç«¯ç‚¹ä½œä¸ºæ–°çš„å…±ç‚¹ï¼Œä»å½“å‰è¾¹å‡ºå‘ï¼Œé€‰æ‹©ç›¸å¯¹äºå½“å‰è¾¹é€†æ—¶é’ˆè½¬è§’æœ€å°çš„å€™é€‰è¾¹ï¼Œ
         ç›´åˆ°æ–°å…±ç‚¹å›åˆ°åˆå§‹ P ç‚¹ï¼ˆé—­åˆï¼‰æˆ–è¾¾åˆ°æœ€å¤§æ­¥æ•°ä¸ºæ­¢ã€‚
    è¿”å›ï¼š
      å¤–è½®å»“çº¿æ„æˆçš„çº¿æ®µåˆ—è¡¨ï¼Œå¦‚æœæ— æ³•æ„æˆå°é—­è½®å»“åˆ™è¿”å›ç©ºåˆ—è¡¨ã€‚
    """
    # å†…éƒ¨å‡½æ•°ï¼šè·å–æ‰€æœ‰çº¿æ®µçš„æœ€å³ä¸‹è§’ç‚¹
    rb = find_rightbottom_point(lines, tol)
    if rb is None:
        print("[é”™è¯¯] æ— æœ€å³ä¸‹è§’ç‚¹")
        return []

    # è®¾ç½®åˆå§‹å…±ç‚¹ä¸º rb
    P = rb
    print(f"è°ƒè¯•ï¼šæœ€å³ä¸‹è§’ç‚¹ä¸º {P}")

    # åœ¨ P å¤„ï¼ŒæŒ‰ç…§ç»å¯¹è§’åº¦æ’åºè·å–æ‰€æœ‰å…±ç‚¹çš„çº¿æ®µ
    candidate_lines = find_lines_angle(lines, P, tol)
    if not candidate_lines:
        print(f"[é”™è¯¯] åœ¨ç‚¹ {P} å¤„æœªæ‰¾åˆ°å…±ç‚¹çº¿æ®µ")
        return []

    # æ ¹æ®é¢˜ç›®è¦æ±‚ï¼Œç¬¬ä¸€æ¡è¾¹é€‰æ‹©ç»å¯¹è§’åº¦æœ€å°çš„çº¿æ®µï¼ˆå³æ’åºåç¬¬ 1 æ ¹çº¿æ®µï¼‰
    initial_line = candidate_lines[0]
    # ç¡®å®šåˆå§‹çº¿æ®µçš„å¦ä¸€ç«¯ç‚¹
    sp_init = tuple(initial_line.StartPoint)
    ep_init = tuple(initial_line.EndPoint)
    if is_nearly_equal(sp_init, P, tol):
        next_point = ep_init
    else:
        next_point = sp_init
    print(f"è°ƒè¯•ï¼šé€‰ä¸­åˆå§‹çº¿æ®µ {initial_line.Handle}ï¼Œèµ·ç‚¹ {P} -> ç»ˆç‚¹ {next_point}")
    
    contour_lines = [initial_line]
    visited_handles = {initial_line.Handle}
    current_line = initial_line
    current_point = next_point
    steps = 0

    # è¿­ä»£æ„é€ å¤–è½®å»“
    while steps < max_steps:
        print(f"è°ƒè¯•ï¼šç›®æ ‡å…±ç‚¹ P = {current_point}")
        # åœ¨å½“å‰å…±ç‚¹å¤„æŸ¥æ‰¾å€™é€‰çš„åç»§çº¿æ®µï¼Œä½¿ç”¨æœ€å°ç›¸å¯¹è§’åº¦ç­–ç•¥
        successor, new_point = find_successor_line_min(current_line, lines, current_point, tol)
        if successor is None:
            print("[é”™è¯¯] æ— åç»§çº¿æ®µï¼Œæ„é€ å¤±è´¥")
            return []
        if successor.Handle in visited_handles:
            print(f"ğŸ” æ£€æµ‹åˆ°é‡å¤çº¿æ®µ {successor.Handle}ï¼Œæ„é€ å¤±è´¥")
            return []
        contour_lines.append(successor)
        visited_handles.add(successor.Handle)
        print(f"æ­¥ {steps+1}: é€‰ä¸­çº¿æ®µ {successor.Handle}ï¼Œæ–°å…±ç‚¹ {new_point}")
        # æ£€æŸ¥æ˜¯å¦é—­åˆï¼šå¦‚æœæ–°å…±ç‚¹ä¸æœ€å³ä¸‹è§’ç‚¹è¿‘ä¼¼ç›¸ç­‰ï¼Œåˆ™è®¤ä¸ºé—­åˆ
        if is_nearly_equal(new_point, rb, tol):
            print("[OK] æˆåŠŸæ„é€ å°é—­è½®å»“çº¿")
            return contour_lines
        # æ›´æ–°å½“å‰çº¿æ®µåŠå…±ç‚¹
        current_line = successor
        current_point = new_point
        steps += 1

    print("[è­¦å‘Š]ï¸ è¾¾åˆ°æœ€å¤§æ­¥æ•°ï¼Œæœªèƒ½æ„é€ å°é—­è½®å»“çº¿")
    return []


#&&%##è·å–æ‰€æœ‰å°é—­å¤šè¾¹å½¢

#åˆ é™¤å¤šè¾¹å½¢çš„ä¸€äº›è¾¹

"""
æœ‰linesä¸­å¤šäºä¸¤æ¡ç›´çº¿æ®µç»è¿‡ï¼Œå°±ç§°è¯¥ç‚¹ä¸ºå¤šåˆ†æç‚¹

ç¼–åˆ¶å‡½æ•°ï¼ŒPLæ˜¯linesä¸­çš„å°é—­å¤šè¾¹å½¢ï¼Œç”¨é¡¶ç‚¹åˆ—è¡¨è¡¨ç¤ºï¼Œp1æ˜¯linesçš„æœ€å³ä¸‹è§’ç‚¹ï¼Œ

ä¸”p1æ˜¯PLçš„ä¸€ä¸ªé¡¶ç‚¹ï¼Œç”¨åæ ‡ç‚¹è¡¨ç¤ºã€‚å°†PLçš„é¡¶ç‚¹æ’åºï¼Œä»p1æŒ‰é€†æ—¶é’ˆæ–¹å‘æ¨è¿›ï¼Œ

å°†é‡åˆ°çš„éå¤šåˆ†æé¡¶ç‚¹æ”¾å…¥åˆ—è¡¨LNï¼Œç›´åˆ°é‡åˆ°ä¸€ä¸ªå¤šåˆ†æç‚¹D1ç»“æŸï¼›å†ä»p1æŒ‰é¡ºæ—¶

é’ˆæ–¹å‘æ¨è¿›ï¼Œå°†é‡åˆ°çš„é¡¶ç‚¹æ”¾å…¥åˆ—è¡¨LNï¼Œç›´åˆ°é‡åˆ°ä¸€ä¸ªå¤šåˆ†æç‚¹D2ç»“æŸã€‚é¡ºæ—¶é’ˆæ–¹

å‘æ¨è¿›æ—¶ï¼ŒæŠŠp1è€ƒè™‘åœ¨å†…ï¼Œå³p1å¦‚æœä¸ºå¤šåˆ†æç‚¹åˆ™ç«‹å³ç»“æŸï¼Œç»“æŸæ—¶é‡åˆ°çš„å¤šåˆ†æ

ç‚¹D2å°±æ˜¯p1ã€‚é€†æ—¶é’ˆæ–¹å‘æ¨è¿›æ—¶ï¼Œä¸æŠŠp1è€ƒè™‘åœ¨å†…ï¼Œå³p1ä¸ºå¤šåˆ†æç‚¹ä»ç„¶æ¨è¿›ï¼Œé‡

åˆ°ä¸€ä¸ªå¤šåˆ†æç‚¹D1ç»“æŸã€‚å‡½æ•°è¿”å›å¤šåˆ†æç‚¹D1ï¼ŒD2ï¼Œä»¥åŠå®ƒä»¬ä¹‹é—´çš„éå¤šåˆ†æé¡¶ç‚¹

åˆ—è¡¨LNã€‚


"""
def deduplicate_vertices(vertices, tol=0.5):
    """
    å»æ‰é¡¶ç‚¹åˆ—è¡¨ä¸­ç›¸é‚»é‡å¤çš„é¡¶ç‚¹ï¼š
    å¦‚æœä¸¤ä¸ªç›¸é‚»é¡¶ç‚¹ï¼ˆé¡ºåºå‡ºç°ï¼‰ä¹‹é—´çš„äºŒç»´è·ç¦»å°äº tolï¼Œåˆ™è®¤ä¸ºå®ƒä»¬é‡å¤ï¼Œåªä¿ç•™å‰ä¸€ä¸ªï¼Œ
    ä½†å¦‚æœè¿™ä¸ªé‡å¤ç‚¹æ˜¯åˆ—è¡¨ä¸­çš„æœ€åä¸€ä¸ªä¸”ä¸ç¬¬ä¸€ä¸ªé¡¶ç‚¹ç›¸åŒï¼ˆè¡¨ç¤ºé—­åˆå¤šè¾¹å½¢ï¼‰ï¼Œåˆ™ä¿ç•™æ­¤é‡å¤ç‚¹ã€‚
    
    å‚æ•°:
      vertices: é¡¶ç‚¹åˆ—è¡¨ï¼Œæ¯ä¸ªé¡¶ç‚¹æ ¼å¼ä¸º (x, y, z) çš„å…ƒç»„ã€‚
      tol: è·ç¦»é˜ˆå€¼ã€‚
      
    è¿”å›:
      å¤„ç†åçš„é¡¶ç‚¹åˆ—è¡¨ï¼Œåªæœ‰ä¸­é—´è¿ç»­é‡å¤çš„ç‚¹è¢«å»é™¤ï¼Œä¿ç•™é—­åˆå¤šè¾¹å½¢çš„é¦–å°¾é‡å¤ã€‚
    """
    if not vertices:
        return []
    
    deduped = [vertices[0]]
    n = len(vertices)
    
    for i in range(1, n):
        pt = vertices[i]
        prev = deduped[-1]
        dx = pt[0] - prev[0]
        dy = pt[1] - prev[1]
        dist = math.sqrt(dx*dx + dy*dy)
        
        # å¦‚æœè·ç¦»å¤§äºç­‰äº tolï¼Œåˆ™è®¤ä¸ºæ˜¯ä¸åŒçš„é¡¶ç‚¹ï¼Œä¿ç•™å®ƒ
        if dist >= tol:
            deduped.append(pt)
        else:
            # å¦‚æœå½“å‰é¡¶ç‚¹æ˜¯æœ€åä¸€ä¸ªï¼Œå¹¶ä¸”ä¸ç¬¬ä¸€ä¸ªé¡¶ç‚¹ç›¸åŒï¼Œåˆ™ä¿ç•™å®ƒï¼ˆè¡¨ç¤ºé—­åˆï¼‰
            if i == n - 1 and same_point(pt, vertices[0], tol):
                deduped.append(pt)
            # å¦åˆ™è·³è¿‡è¯¥ç‚¹
    return deduped


#  ä¸»å‡½æ•°
#  (4)
# è·å–æœ€å³ä¸‹è§’çš„å°é—­å¤šè¾¹å½¢ä¸å½±å“å…¶ä»–å°é—­å¤šè¾¹å½¢çš„è¿ç»­è¾¹çš„é¡¶ç‚¹åˆ—è¡¨ 

def analyze_polygon_branches(PL, lines, p1, tol=0.5):
    """
    åˆ†æå°é—­å¤šè¾¹å½¢ PL çš„åˆ†ææƒ…å†µã€‚PL ä¸ºå°é—­å¤šè¾¹å½¢çš„é¡¶ç‚¹åˆ—è¡¨ï¼ˆæŒ‰é€†æ—¶é’ˆé¡ºåºæ’åˆ—ï¼‰ï¼Œ
    p1 ä¸ºæœ€å³ä¸‹è§’ç‚¹ï¼Œå¿…åœ¨ PL ä¸­ã€‚
    
    è§„åˆ™ï¼š
      1. ä» p1 å‡ºå‘ï¼š
         - æ²¿é€†æ—¶é’ˆæ–¹å‘ï¼ˆä¸å°† p1 è§†ä¸ºå€™é€‰ï¼‰æ¨è¿›ï¼Œç´¯è®¡é‡åˆ°çš„éå¤šåˆ†æé¡¶ç‚¹ï¼Œç›´è‡³é‡åˆ°ç¬¬ä¸€ä¸ªå¤šåˆ†æç‚¹ D1ï¼›
         - æ²¿é¡ºæ—¶é’ˆæ–¹å‘ï¼ˆå°† p1 ä¹Ÿè€ƒè™‘åœ¨å†…ï¼‰æ¨è¿›ï¼Œç´¯è®¡é‡åˆ°çš„éå¤šåˆ†æé¡¶ç‚¹ï¼Œç›´è‡³é‡åˆ°ç¬¬ä¸€ä¸ªå¤šåˆ†æç‚¹ D2ã€‚
      2. å°†è¿™ä¸¤ä¸ªæ–¹å‘ç´¯è®¡å¾—åˆ°çš„éå¤šåˆ†æé¡¶ç‚¹ LNï¼ˆå…¶ä¸­é¡ºæ—¶é’ˆæ–¹å‘å¾—åˆ°çš„é¡¶ç‚¹éœ€è¦åè½¬åä¸é€†æ—¶é’ˆæ–¹å‘çš„é¡¶ç‚¹ç›¸æ¥ï¼‰æ‰“å°å‡ºæ¥ã€‚
      3. ç”±äº PL æ˜¯é—­åˆå¤šè¾¹å½¢ï¼Œä» D1 åˆ° D2æœ‰ä¸¤æ¡è¿ç»­é¡¶ç‚¹åºåˆ—ï¼Œä»ä¸­é€‰æ‹©åŒ…å« LN çš„é‚£æ¡ä½œä¸ºæœ€ç»ˆè¿”å›ç»“æœã€‚

    è¿”å›ï¼š
      è¿”å›ä» D1 åˆ° D2ï¼ˆåŒ…å« D1 å’Œ D2ï¼‰çš„è¿ç»­é¡¶ç‚¹åºåˆ—ï¼Œæ­¤åºåˆ—åŒ…å«äº† LN çš„é¡¶ç‚¹ã€‚
      å¦‚æœä»»ä¸€æ–¹å‘æœªèƒ½æ‰¾åˆ°å¤šåˆ†æç‚¹ï¼Œåˆ™æ‰“å°æç¤ºï¼Œå¹¶è¿”å› Noneã€‚
    """


    # å†…éƒ¨è¾…åŠ©ï¼šåˆ¤æ–­ä¸¤ä¸ªç‚¹æ˜¯å¦è¿‘ä¼¼ç›¸ç­‰ï¼ˆä»…æ¯”è¾ƒ x,y åæ ‡ï¼‰
##    def same_point(a, b, tol=tol):
##        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol
##


    # å†…éƒ¨è¾…åŠ©ï¼šåˆ¤æ–­æŸé¡¶ç‚¹æ˜¯å¦ä¸ºå¤šåˆ†æç‚¹
    def is_multi_branch(vertex):
        cnt = 0
        for ln in lines:
            try:
                sp = tuple(ln.StartPoint)
                ep = tuple(ln.EndPoint)
            except Exception:
                continue
            if same_point(vertex, sp, tol) or same_point(vertex, ep, tol):
                cnt += 1
        return cnt > 2

    # æ‰¾å‡º p1 åœ¨ PL ä¸­çš„ç´¢å¼•
    try:
        idx = PL.index(p1)
    except Exception as e:
        print("é”™è¯¯ï¼šp1 ä¸åœ¨ PL ä¸­")
        return None

    n = len(PL)

    # æ²¿é€†æ—¶é’ˆæ–¹å‘æ¨è¿›ï¼ˆä¸åŒ…æ‹¬ p1ï¼‰ï¼Œç´¯è®¡éå¤šåˆ†æé¡¶ç‚¹ LN_ccwï¼Œç›´åˆ°é‡åˆ°ç¬¬ä¸€ä¸ªå¤šåˆ†æç‚¹ D1
    LN_ccw = []
    D1 = None
    i = (idx + 1) % n
    while i != idx:
        v = PL[i]
        if is_multi_branch(v):
            D1 = v
            break
        else:
            LN_ccw.append(v)
        i = (i + 1) % n
    if D1 is None:
        print("è°ƒè¯•ï¼šé€†æ—¶é’ˆæ–¹å‘æœªé‡åˆ°å¤šåˆ†æç‚¹")
        return None

    # æ²¿é¡ºæ—¶é’ˆæ–¹å‘æ¨è¿›ï¼ˆåŒ…å« p1ï¼‰ï¼Œç´¯è®¡éå¤šåˆ†æé¡¶ç‚¹ LN_cwï¼Œç›´åˆ°é‡åˆ°ç¬¬ä¸€ä¸ªå¤šåˆ†æç‚¹ D2
    LN_cw = []
    if is_multi_branch(PL[idx]):
        D2 = PL[idx]
    else:
        D2 = None
        i = (idx - 1) % n
        while i != idx:
            v = PL[i]
            if is_multi_branch(v):
                D2 = v
                break
            else:
                LN_cw.append(v)
            i = (i - 1) % n
        if D2 is None:
            print("è°ƒè¯•ï¼šé¡ºæ—¶é’ˆæ–¹å‘æœªé‡åˆ°å¤šåˆ†æç‚¹")
            return None

    # åˆå¹¶é¡ºæ—¶é’ˆæ–¹å‘çš„é¡¶ç‚¹ï¼ˆéœ€è¦åè½¬ï¼‰å’Œé€†æ—¶é’ˆæ–¹å‘çš„é¡¶ç‚¹
    LN = list(reversed(LN_cw)) + LN_ccw

    # è°ƒè¯•æ‰“å°
    print("è°ƒè¯•ä¿¡æ¯ï¼š")
    print("ç›®æ ‡å…±ç‚¹ p1 =", p1)
    print("é€†æ—¶é’ˆæ–¹å‘æ”¶é›†çš„éå¤šåˆ†æé¡¶ç‚¹ (LN_ccw):")
    for pt in LN_ccw:
        print("  ", pt)
    print("é¡ºæ—¶é’ˆæ–¹å‘æ”¶é›†çš„éå¤šåˆ†æé¡¶ç‚¹ (LN_cw, åŸé¡ºåº):")
    for pt in LN_cw:
        print("  ", pt)
    print("åˆå¹¶åçš„éå¤šåˆ†æé¡¶ç‚¹ LN (ä» D2 åˆ° D1):")
    for pt in LN:
        print("  ", pt)
    print("é€†æ—¶é’ˆæ–¹å‘é‡åˆ°çš„å¤šåˆ†æç‚¹ D1:", D1)
    print("é¡ºæ—¶é’ˆæ–¹å‘é‡åˆ°çš„å¤šåˆ†æç‚¹ D2:", D2)

    # è®¡ç®—åœ¨ PL ä¸­ D1 å’Œ D2 çš„ç´¢å¼•
    try:
        idx_D1 = PL.index(D1)
        idx_D2 = PL.index(D2)
    except Exception as e:
        print("é”™è¯¯ï¼šæ— æ³•åœ¨ PL ä¸­æŸ¥æ‰¾ D1 æˆ– D2:", e)
        return None

    # ç”±äº PL æ˜¯é—­åˆçš„ï¼Œæˆ‘ä»¬æœ‰ä¸¤æ¡è¿æ¥ D1 å’Œ D2ï¼š
    if idx_D1 <= idx_D2:
        branch_A = PL[idx_D1: idx_D2 + 1]
        branch_B = PL[idx_D2:] + PL[:idx_D1 + 1]
    else:
        branch_A = PL[idx_D1:] + PL[:idx_D2 + 1]
        branch_B = PL[idx_D2: idx_D1 + 1]

    # é€‰æ‹©åŒ…å« LN ä¸­é¡¶ç‚¹çš„é‚£ä¸ªåˆ†æ
    selected_branch = None
    if LN:
        found_in_A = any(same_point(v, LN[0], tol) for v in branch_A)
        found_in_B = any(same_point(v, LN[0], tol) for v in branch_B)
        if found_in_A and not found_in_B:
            selected_branch = branch_A
        elif found_in_B and not found_in_A:
            selected_branch = branch_B
        elif found_in_A and found_in_B:
            # å¦‚æœä¸¤æ¡åˆ†æéƒ½åŒ…å«ï¼Œåˆ™é€‰æ‹©è¾ƒçŸ­çš„é‚£æ¡ï¼Œé€šå¸¸è¿™æ‰æ˜¯â€œå±€éƒ¨â€çš„åˆ†æ
            selected_branch = branch_A if len(branch_A) <= len(branch_B) else branch_B
        else:
            print("è°ƒè¯•ï¼šLN ä¸­çš„é¡¶ç‚¹ä¸åœ¨ä»»ä¸€åˆ†æä¸Šï¼Œé»˜è®¤é€‰æ‹© branch_A")
            selected_branch = branch_A
    else:
        print("è°ƒè¯•ï¼šLN ä¸ºç©ºï¼Œæ— æ³•ç¡®å®šåŒ…å« LN çš„åˆ†æï¼Œé»˜è®¤é€‰æ‹© branch_A")
        selected_branch = branch_A

    print("æœ€ç»ˆè¿”å›çš„ä» D1 åˆ° D2ï¼ˆåŒ…å« LN çš„åˆ†æï¼‰é¡¶ç‚¹åºåˆ—ï¼š")
    for pt in selected_branch:
        print("  ", pt)


    #å»é‡
    selected_branch=deduplicate_vertices(selected_branch, tol=tol)

    return selected_branch

##æ ¹æ®è¾“å…¥çš„é¡¶ç‚¹åˆ—è¡¨åˆ¤æ–­ï¼Œå°†linesçš„å…¶é¡¶ç‚¹åœ¨è¯¥é¡¶ç‚¹åˆ—è¡¨çš„çº¿æ®µç§»å‡ºåˆ—è¡¨


def remove_lines_in_LBv(lines, LB_v, tol=0.1):
    """
    ä» COM çº¿æ®µåˆ—è¡¨ lines ä¸­ç§»é™¤é‚£äº›å…¶ä¸¤ä¸ªé¡¶ç‚¹éƒ½åœ¨ LB_v ä¸­çš„çº¿æ®µã€‚
    
    å‚æ•°ï¼š
      lines: COM çº¿æ®µå¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡è¦æ±‚å…·æœ‰ StartPoint ä¸ EndPoint å±æ€§ï¼Œ
             å…¶å€¼ä¸ºå½¢å¦‚ (x, y, z) çš„å¯è¿­ä»£å¯¹è±¡ã€‚
      LB_v: é¡¶ç‚¹åˆ—è¡¨ï¼Œæ¯ä¸ªé¡¶ç‚¹ä¸º (x, y, z) çš„å…ƒç»„ã€‚
      tol: åˆ¤æ–­ä¸¤é¡¶ç‚¹è¿‘ä¼¼ç›¸ç­‰çš„å®¹å·®å€¼ï¼ˆé»˜è®¤ 0.1ï¼‰ã€‚
    
    è¿”å›ï¼š
      è¿”å›ç§»é™¤æ»¡è¶³æ¡ä»¶ï¼ˆå³ä¸¤ç«¯ç‚¹éƒ½åœ¨ LB_v ä¸­ï¼‰çš„çº¿æ®µåå‰©ä½™çš„çº¿æ®µåˆ—è¡¨ã€‚
    """

    def same_point(pt1, pt2, tol=tol):
        # åªæ¯”è¾ƒ x, y, z åæ ‡çš„å·®å€¼ï¼Œåˆ¤æ–­ä¸¤ç‚¹æ˜¯å¦è¿‘ä¼¼ç›¸ç­‰
        return abs(pt1[0] - pt2[0]) <= tol and abs(pt1[1] - pt2[1]) <= tol and abs(pt1[2] - pt2[2]) <= tol

    remaining_lines = []
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
        except Exception as e:
            print(f"[è­¦å‘Š]ï¸ è·³è¿‡æ— æ•ˆçº¿æ®µï¼ŒåŸå› : {e}")
            continue

        # åˆ¤æ–­è¯¥çº¿æ®µçš„èµ·ç‚¹å’Œç»ˆç‚¹æ˜¯å¦å‡å­˜åœ¨äº LB_v ä¸­
        sp_in = any(same_point(sp, lb, tol) for lb in LB_v)
        ep_in = any(same_point(ep, lb, tol) for lb in LB_v)
        if not (sp_in and ep_in):
            remaining_lines.append(ln)
        else:
            print(f"åˆ é™¤çº¿æ®µ {ln.Handle}ï¼Œå…¶ä¸¤ä¸ªç«¯ç‚¹å‡åœ¨ LB_v ä¸­ï¼šsp={sp}, ep={ep}")
    return remaining_lines

#  ä¸»å‡½æ•°
#  (5)
#&&% è·å–å…¨éƒ¨å°é—­å¤šè¾¹å½¢ï¼Œä½†ä¸å®Œå…¨

def process_polygons(lines, tol=0.5, max_steps=50, layer_name="æµ‹è¯•è¾…åŠ©"):
    """
    é€’å½’æå–å¹¶ç»˜åˆ¶ç›´çº¿æ®µé›† lines ä¸­æ‰€æœ‰å°é—­å¤šè¾¹å½¢ã€‚
    
    æµç¨‹ï¼š
      1. ä» lines ä¸­æå–æœ€å³ä¸‹è§’å°é—­å¤šè¾¹å½¢ï¼ˆè°ƒç”¨ find_rightbottom_closed_polygonï¼‰ã€‚
      2. å¯¹è¯¥å¤šè¾¹å½¢ï¼š
            - è·å–æœ€å³ä¸‹è§’ç‚¹ p1ï¼ˆè°ƒç”¨ find_rightbottom_pointï¼‰ï¼›
            - å°†å¤šè¾¹å½¢ï¼ˆpolygonï¼‰åŠ å…¥åˆ—è¡¨ LBï¼›
            - è°ƒç”¨ analyze_polygon_branches åˆ†æå¤šè¾¹å½¢åˆ†æï¼Œå¾—åˆ°ç”¨æ¥åˆ¤æ–­ç§»é™¤çº¿æ®µçš„é¡¶ç‚¹åˆ—è¡¨ Lvï¼›
            - è°ƒç”¨ remove_lines_in_LBvï¼Œå°†æ‰€æœ‰ä¸¤ç«¯ç‚¹å‡åœ¨ Lv ä¸­çš„ç›´çº¿æ®µä» lines ä¸­ç§»é™¤ï¼›
            - è°ƒç”¨ draw_polygon_as_polyline ç»˜åˆ¶è¯¥å¤šè¾¹å½¢ï¼Œå°†ç”Ÿæˆçš„å¤šæ®µçº¿ COM å¯¹è±¡åŠ å…¥ LBcomï¼›
            - ç„¶åè°ƒç”¨ draw_polygon_as_polyline ç»˜åˆ¶é¡¶ç‚¹åˆ—è¡¨ Lvï¼ˆè“è‰²ç»˜åˆ¶ï¼‰ï¼Œå°†ç”Ÿæˆçš„å¤šæ®µçº¿ COM å¯¹è±¡åŠ å…¥ LB_ycï¼›
      3. é‡å¤ä¸Šè¿°è¿‡ç¨‹ç›´è‡³æ— æ³•æå–å‡ºå°é—­å¤šè¾¹å½¢ï¼›
    
    è¿”å›ï¼š
      (LB, LBcom, LB_yc)ï¼Œå…¶ä¸­ï¼š
         LB: å°é—­å¤šè¾¹å½¢é¡¶ç‚¹åˆ—è¡¨é›†åˆï¼ˆæ¯ä¸ªä¸ºé¡¶ç‚¹åºåˆ—ï¼‰ã€‚
         LBcom: ç»˜åˆ¶å‡ºçš„å¤šæ®µçº¿ COM å¯¹è±¡åˆ—è¡¨ã€‚
         LB_yc: ç»˜åˆ¶è“è‰²è¾…åŠ©å¤šæ®µçº¿ï¼ˆç”¨äºæ£€æŸ¥ç§»é™¤çš„çº¿æ®µé¡¶ç‚¹ï¼‰çš„ COM å¯¹è±¡åˆ—è¡¨ã€‚
    """
    LB = []
    LBcom = []
    LB_yc = []  # å­˜å‚¨è“è‰²è¾…åŠ©å¤šæ®µçº¿ COM å¯¹è±¡ï¼Œç”¨äºæ£€æŸ¥ç§»é™¤çš„é¡¶ç‚¹
    Ly = []  # ç”¨æ¥è®°å½•æ¯æ¬¡ç§»é™¤çš„ç›´çº¿æ®µ
    iteration = 0

    while True:
        iteration += 1
        print(f"\nã€è¿­ä»£ {iteration}ã€‘å‰©ä½™ç›´çº¿æ®µæ•°é‡ = {len(lines)}")
        p1 = find_rightbottom_point(lines, tol)
        if p1 is None:
            print("æœªæ‰¾åˆ°æœ€å³ä¸‹è§’ç‚¹ï¼Œç»“æŸè¿­ä»£")
            break
        print(f"å½“å‰æœ€å³ä¸‹è§’ç‚¹ p1 = {p1}")
        
        polygon = find_rightbottom_closed_polygon(lines, tol=tol, max_steps=max_steps)
        if polygon is None:
            print("æ— æ³•æå–å°é—­å¤šè¾¹å½¢ï¼Œç»“æŸè¿­ä»£")
            break
        LB.append(polygon)
        print("æå–çš„å°é—­å¤šè¾¹å½¢é¡¶ç‚¹ï¼š")
        for pt in polygon:
            print("  ", pt)
        
        Lv = analyze_polygon_branches(polygon, lines, p1, tol=tol)
        if Lv is None:
            print("åˆ†æå¤šåˆ†æå¤±è´¥ï¼Œç»“æŸæœ¬æ¬¡è¿­ä»£")
        else:
            print("ç”¨äºç§»é™¤çº¿æ®µçš„é¡¶ç‚¹åˆ—è¡¨ Lvï¼š")
            for pt in Lv:
                print("  ", pt)
            # ç§»é™¤ lines ä¸­ä¸¤ç«¯ç‚¹å‡åœ¨ Lv ä¸­çš„ç›´çº¿æ®µï¼ˆè°ƒç”¨ remove_lines_in_LBvï¼‰
            new_lines = remove_lines_in_LBv(lines, Lv, tol=0.1)
            removed_count = len(lines) - len(new_lines)
            print(f"æœ¬æ¬¡ç§»é™¤ç›´çº¿æ®µæ•°: {removed_count}")
            # è®°å½•è¢«ç§»é™¤çš„ç›´çº¿æ®µ
            for ln in lines:
                if ln not in new_lines:
                    Ly.append(ln)
            lines = new_lines
        
        # ç»˜åˆ¶æå–å‡ºçš„å°é—­å¤šè¾¹å½¢
        polyline = draw_polygon_as_polyline(polygon, layer_name=layer_name, tol=tol)
        if polyline:
            LBcom.append(polyline)
        
        # è°ƒç”¨ draw_polygon_as_polyline ç»˜åˆ¶ Lvï¼ˆè¾…åŠ©å¤šæ®µçº¿ï¼Œæ³¨æ„Lvæ˜¯é¡¶ç‚¹åˆ—è¡¨ï¼‰
        if Lv is not None and len(Lv) > 0:
            poly_blue = draw_polygon_as_polyline(Lv, layer_name=layer_name, tol=tol)
            if poly_blue:
                LB_yc.append(poly_blue)
        
        if len(lines) < 3:
            print("å‰©ä½™ç›´çº¿æ®µä¸è¶³ä»¥æ„æˆå°é—­å¤šè¾¹å½¢ï¼Œé€€å‡º")
            break

    return LB, LBcom, LB_yc



def extract_polygon_from_lines(lines, tol=0.5):
    """
    å°†è¡¨ç¤ºå°é—­å¤šè¾¹å½¢è¾¹ç¼˜çš„çº¿æ®µï¼ˆCOMå¯¹è±¡åˆ—è¡¨ï¼‰è½¬æ¢ä¸ºé¡¶ç‚¹åˆ—è¡¨ï¼ˆæŒ‰é¡ºåºæ’åˆ—ï¼‰ï¼Œ
    æ¶ˆé™¤ä¸­é—´ç›¸é‚»çš„é‡å¤é¡¶ç‚¹ã€‚è‹¥èƒ½æˆåŠŸæ„æˆé—­åˆå¤šè¾¹å½¢ï¼Œåˆ™è¿”å›é¡¶ç‚¹åˆ—è¡¨ï¼ˆé—­ç¯ä¸é‡å¤ï¼‰ï¼Œå¦åˆ™è¿”å› Noneã€‚
    
    å‚æ•°ï¼š
      lines: ç›´çº¿æ®µå¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡è¦æ±‚å…·æœ‰ StartPoint å’Œ EndPoint å±æ€§
      tol: åˆ¤æ–­é¡¶ç‚¹ç›¸ç­‰çš„å®¹å·®ï¼ˆä»…æ¯”è¾ƒ x å’Œ y åæ ‡ï¼‰ã€‚
    
    è¿”å›ï¼š
      é¡¶ç‚¹åˆ—è¡¨ï¼Œå¦‚ [p1, p2, ..., pn]ï¼Œå…¶ä¸­ p1 è¡¨ç¤ºå¤šè¾¹å½¢çš„èµ·ç‚¹ï¼ˆä¸é‡å¤åˆ—å‡ºé—­åˆé¡¶ç‚¹ï¼‰ã€‚
    """
    if not lines:
        return None
    
    # å®šä¹‰åŒç‚¹åˆ¤æ–­å‡½æ•°
##    def same_point(a, b, tol=tol):
##        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol
    
    # æ‹·è´çº¿æ®µåˆ—è¡¨ï¼Œæ–¹ä¾¿ä¿®æ”¹ï¼ˆæ³¨æ„ï¼šæ­¤å¤„ä¸ä¼šå¤åˆ¶COMå¯¹è±¡ï¼Œä»…å¤åˆ¶å¼•ç”¨ï¼‰
    remaining_lines = list(lines)
    
    try:
        start_pt = tuple(remaining_lines[0].StartPoint)
    except Exception as e:
        print("æå–ç¬¬ä¸€æ¡çº¿æ®µèµ·ç‚¹å¤±è´¥ï¼š", e)
        return None
    vertices = [start_pt]
    current_pt = start_pt

    # ç”±äºå¤šè¾¹å½¢æ˜¯é—­åˆçš„ï¼Œæœ€å¤šå°è¯• 2*len(remaining_lines) æ¬¡
    for _ in range(len(remaining_lines) * 2):
        found_edge = False
        for ln in remaining_lines:
            try:
                sp = tuple(ln.StartPoint)
                ep = tuple(ln.EndPoint)
            except Exception:
                continue

            nxt = None
            if same_point(current_pt, sp) and (not same_point(current_pt, ep)):
                nxt = ep
            elif same_point(current_pt, ep) and (not same_point(current_pt, sp)):
                nxt = sp

            if nxt is None:
                continue

            # ä¸€æ—¦æ‰¾åˆ°ä¸å½“å‰ç‚¹ç›¸è¿çš„è¾¹ï¼Œç§»é™¤æ­¤è¾¹
            remaining_lines.remove(ln)
            # å¦‚æœå¾—åˆ°çš„ nxt ä¸èµ·ç‚¹ç›¸åŒï¼Œåˆ™è¯´æ˜é—­åˆç»“æŸ
            if same_point(nxt, start_pt):
                vertices.append(start_pt)
                return deduplicate_vertices(vertices, tol)
            else:
                vertices.append(nxt)
                current_pt = nxt
                found_edge = True
                break
        if not found_edge:
            # æœªèƒ½æ‰¾åˆ°ä¸å½“å‰ç‚¹ç›¸è¿çš„è¾¹ï¼Œæ„æˆé—­åˆå¤±è´¥
            break
    # å¦‚æœå¾ªç¯ç»“æŸåé—­åˆæœªå®Œæˆï¼Œåˆ™è¿”å› None
    return None



#å°†å¤šæ®µçº¿åˆ—è¡¨ç‚¸å¼€ä¸ºçº¿æ®µï¼Œè¿”å›çº¿æ®µåˆ—è¡¨



def explode_polylines(LB):
    """
    å¯¹å¤šæ®µçº¿åˆ—è¡¨ LB ä¸­çš„æ¯ä¸€ä¸ªå¤šæ®µçº¿ï¼Œè°ƒç”¨ .Explode() æ–¹æ³•ï¼Œ
    å¹¶å°†ç‚¸å¼€å¾—åˆ°çš„æ‰€æœ‰çº¿æ®µåˆå¹¶ä¸ºä¸€ä¸ªæ–°çš„çº¿æ®µåˆ—è¡¨è¿”å›ã€‚

    å‚æ•°:
      LB: å¤šæ®µçº¿ COM å¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå¯¹è±¡åº”æ”¯æŒ .Explode() æ–¹æ³•ï¼Œ
          å¦‚ pl = polyline_explode_object.Explode() è¿”å›è¯¥å¤šæ®µçº¿ç‚¸å¼€åçš„çº¿æ®µé›†åˆã€‚
          
    è¿”å›:
      ä¸€ä¸ªåŒ…å«æ‰€æœ‰ç‚¸å¼€åç›´çº¿æ®µ COM å¯¹è±¡çš„åˆ—è¡¨ã€‚
    """
    exploded_lines = []
    for pl in LB:
        try:
            # è°ƒç”¨ Explode() æ–¹æ³•ï¼Œè¿”å›ä¸€ä¸ªé›†åˆï¼ˆä¾‹å¦‚ COM Collectionï¼‰
            exploded = pl.Explode()
            # éå†è¿”å›çš„é›†åˆï¼Œå°†æ¯æ ¹çº¿æ®µæ·»åŠ åˆ°åˆ—è¡¨ä¸­
            # æ³¨æ„ï¼šéå† COM Collection çš„æ–¹æ³•å¯èƒ½å› ç¯å¢ƒä¸åŒè€Œä¸åŒï¼Œæ­¤å¤„å‡è®¾å¯ä»¥ç›´æ¥è¿­ä»£
            for ln in exploded:
                exploded_lines.append(ln)
        except Exception as e:
            handle = getattr(pl, 'Handle', 'æœªçŸ¥')
            print(f"[è­¦å‘Š]ï¸ å¤„ç†å¤šæ®µçº¿ {handle} æ—¶å‡ºé”™: {e}")
    return exploded_lines


#lines1 ä¸­é‚£äº›ä¸åœ¨ lines2 ä¸­çš„çº¿æ®µ

def subtract_line_sets(lines1, lines2, tol=0.5):
    """
    æ¯”è¾ƒä¸¤ä¸ªçº¿æ®µé›†åˆ lines1 å’Œ lines2ï¼Œè¿”å› lines1 ä¸­é‚£äº›ä¸åœ¨ lines2 ä¸­çš„çº¿æ®µã€‚

    å‚æ•°ï¼š
      lines1: ç¬¬ä¸€ç»„çº¿æ®µï¼ˆå€™é€‰é›†åˆï¼‰ï¼Œæ¯ä¸ªå¯¹è±¡é¡»å…·æœ‰ StartPoint å’Œ EndPoint å±æ€§ã€‚
      lines2: ç¬¬äºŒç»„çº¿æ®µï¼Œå‚ç…§é›†åˆã€‚
      tol: åˆ¤æ–­ç‚¹æ˜¯å¦ç›¸åŒçš„å®¹å·®ï¼Œé»˜è®¤å€¼ä¸º 0.5ã€‚
      
    è¿”å›ï¼š
      lines1 ä¸­æ‰€æœ‰ä¸ä¸ lines2 ä¸­ä»»æ„çº¿æ®µç›¸åŒçš„çº¿æ®µæ„æˆçš„åˆ—è¡¨ã€‚
    """
    result = []
    for ln in lines1:
        found = False
        for ln2 in lines2:
            if same_line(ln, ln2, tol):
                found = True
                break
        if not found:
            result.append(ln)
    return result

#  ä¸»å‡½æ•°
#  (6)
#&&% è·å–å…¨éƒ¨å°é—­å¤šè¾¹å½¢


def process_final(lines, tol=0.5, max_steps=50, layer_name="æµ‹è¯•è¾…åŠ©"):

    print("len(lines):",len(lines))

    L1, L2, L3 = process_polygons(lines, tol=tol, max_steps=max_steps, layer_name=layer_name)

    print(f"process_polygons å®Œæˆï¼š\n  L1 æ•°é‡ = {len(L1)}\n  L2 æ•°é‡ = {len(L2)}\n  L3 æ•°é‡ = {len(L3)}")
    
    print(">>> å¯¹ L3 ä¸­çš„å¤šæ®µçº¿æ‰§è¡Œ Explode() æ“ä½œ...")

    exploded_lines=explode_polylines(L3)



    shengyu = subtract_line_sets(lines, exploded_lines, tol=tol)

    for x in shengyu:

        print(x.StartPoint,x.EndPoint)

    L_shengyu=extract_polygon_from_lines(shengyu, tol=tol)
   

    ld = draw_polygon_as_polyline(L_shengyu, layer_name=layer_name, tol=tol)

    L2.append(ld)

    L1.append(L_shengyu)
    for x in exploded_lines:
        x.Delete()

    return L1, L2, L3



#&&&% PLå¤çº¿å¤„ç†


"""
ç ”ç©¶PLå¤çº¿å¤„ç†çš„é—®é¢˜

(1)å¤šæ®µçº¿çš„åŸºæœ¬æ“ä½œ get_unique_vertices_from_pl_com

(2)PLæ‰“å°çº¿ generate_name_and_ratio_from_polyline(polyline,A3dy=0)

(3)å°†æ­£äº¤å…­è¾¹å½¢å¤šæ®µçº¿åˆ†æˆä¸¤ä¸ªçŸ©å½¢åŒºåŸŸ split_hexagon_combined(polygon, tol=0.1, simplify_tol=0.5)

(4)è·å–å¤šæ®µçº¿çš„ä¸Šä¸‹å·¦å³è¾¹ç•Œçš„ç›´çº¿æ®µï¼Œè¿”å›çº¿æ®µç«¯ç‚¹åˆ—è¡¨ get_bbox_edge_segments(pl, tol=0.5)

(5)è·å–å¤šæ®µçº¿çš„å†…éƒ¨çš„æ–‡å­— get_texts_in_polyline(com_pl, tol=0.5)

(6)å¤šæ®µçº¿ä¸Šçš„å‡åˆ†æ’å…¥ distribute_points_on_entity(entity, n, block, scale_factor, ys)

(7)è¿”å› pl1 ä¸­ä¸ pl2 â€œå…±çº¿ä¸”æœ‰é‡å â€çš„åŒºæ®µåˆ—è¡¨ common_segments_between_polylines(pl1, pl2, tol=0.5)

(8)æ‰¾åˆ°å…¨éƒ¨â€œä¸¤æ ¹å¤šæ®µçº¿è€¦åˆæˆä¸€ä¸ªçŸ©å½¢â€çš„å¤šæ®µçº¿ two_plines_making_rectangle(pl1, pl2, tol=0.5)

"""
#  ä¸»å‡½æ•°
#  (1)
# å¤šæ®µçº¿çš„åŸºæœ¬æ“ä½œ

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°
"""

æ ‡å‡†é¡¶ç‚¹åæ ‡åˆ—è¡¨,æ˜¯åƒ[(0, 0, 0), (0, 100, 0), (0, 900, 0)]è¿™æ ·çš„åˆ—è¡¨

å®ƒè¡¨ç¤º3ä¸ªè¿ç»­é¡¶ç‚¹2æ ¹è¿ç»­çº¿æ®µï¼Œå…¬å…±ç‚¹åæ ‡ä¸é‡å¤

"""
## 0 è½»é‡å¤šæ®µçº¿å’Œä¸€èˆ¬ä¼ ç»Ÿå¤šæ®µçº¿å„æœ‰ç”¨å¤„ï¼Œåè€…æ‰èƒ½åœ¨ä¸‰ç»´ä¸­ä½¿ç”¨ï¼Œæ›´å¹¿æ³›




@alias("ç”»è½»é‡å¤šæ®µçº¿")
def draw_lwpolyline(
    coords3d: list[tuple[float, float, float]],
    layer_name: str = "0",
    width: float = 0.0,
    color: int = 256,
    closed: bool = False
):
    """
    æ ¹æ®ä¸€ç»„ (x, y, z) åæ ‡ç‚¹ç»˜åˆ¶è½»é‡çº§å¤šæ®µçº¿ï¼ˆLWPOLYLINEï¼‰ã€‚

    :param coords3d: å½¢å¦‚ [(x1, y1, z1), (x2, y2, z2), â€¦] çš„ç‚¹åˆ—è¡¨ï¼Œ
                     ä»…ä½¿ç”¨ x,y åæ ‡ï¼Œå¿½ç•¥ zã€‚
    :param layer_name: ç›®æ ‡å›¾å±‚åç§°ï¼Œä¸å­˜åœ¨åˆ™è‡ªåŠ¨åˆ›å»ºã€‚
    :param width:      å¤šæ®µçº¿æ’å®½ (ConstantWidth)ã€‚
    :param color:      é¢œè‰²ç´¢å¼• (AutoCAD Color Index)ï¼Œ256=BYLAYERã€‚
    :param closed:     æ˜¯å¦é—­åˆå¤šæ®µçº¿ï¼ˆé¦–å°¾ç›¸è¿ï¼‰ã€‚

    :return:           æ–°å»ºçš„è½»é‡çº§å¤šæ®µçº¿å¯¹è±¡ (COM AddLightWeightPolyline)ã€‚

    pts = [
        (0.0, 0.0, 0.0),
        (100.0, 0.0, 0.0),
        (100.0, 50.0, 0.0),
        (0.0, 50.0, 0.0),
    ]
    poly = draw_lwpolyline(
        coords3d=pts,
        layer_name="dy_quyu",
        width=0.0,
        color=1,      # çº¢è‰²
        closed=True
    )
    poly.Coordinates
    (0.0, 0.0, 100.0, 0.0, 100.0, 50.0, 0.0, 50.0)

    len((0.0, 0.0, 100.0, 0.0, 100.0, 50.0, 0.0, 50.0))
    8

    """
    # 1ï¸âƒ£ è¿æ¥ AutoCAD

    # 2ï¸âƒ£ ç¡®ä¿å›¾å±‚å­˜åœ¨
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    # Optional: å¼€å¯å›¾å±‚
    lyr.LayerOn = True

    # 3ï¸âƒ£ å‡†å¤‡åæ ‡æ•°ç»„ï¼šæ‰å¹³åŒ– x,y
    raw = []
    for x, y, _ in coords3d:
        raw.extend((x, y))
    # è½¬ COM VARIANT æ•°ç»„
    arr = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        raw
    )

    # 4ï¸âƒ£ ç»˜åˆ¶è½»é‡çº§å¤šæ®µçº¿

    try:

        pline = mp.AddLightWeightPolyline(arr)
        pline.Layer         = layer_name
        pline.ConstantWidth = width
        pline.color         = color
        pline.Closed        = bool(closed)

        print(f"[OK] å·²åœ¨å›¾å±‚ã€{layer_name}ã€ç»˜åˆ¶å¤šæ®µçº¿ï¼ŒClosed = {closed}")
        return pline
    except Exception as e:
        print("[é”™è¯¯] ç»˜åˆ¶å¤šæ®µçº¿å¤±è´¥:", e) 

   # 5ï¸âƒ£ è¿”å›æ–°å¯¹è±¡

    return pline



# 1 ä»comå¤çº¿è·å–æ ‡å‡†é¡¶ç‚¹åæ ‡åˆ—è¡¨

def get_unique_vertices_from_pl_com(pl_com):
    """
    æå–å¤šæ®µçº¿çš„é¡¶ç‚¹åˆ—è¡¨ï¼Œä¸é‡å¤è¿ç»­çº¿æ®µçš„å…¬å…±é¡¶ç‚¹ï¼Œè¿”å›é¡¶ç‚¹åˆ—è¡¨ã€‚
    
    å‚æ•°:
        pl_com: AutoCAD ä¸­çš„ Polyline COM å¯¹è±¡ï¼ˆAcDbPolylineï¼‰
        
    è¿”å›:
        é¡¶ç‚¹åˆ—è¡¨ï¼Œæ¯ä¸¤ä¸ªç‚¹æ„æˆä¸€æ¡çº¿æ®µï¼Œæ ¼å¼ï¼š[ (x1, y1, z1), (x2, y2, z2), ... ]
        
    """
    # è·å–äºŒç»´åæ ‡æ•°æ®
    coords = pl_com.Coordinates
    vertices = []

    # ä»¥æ¯ä¸¤ä¸ªåæ ‡ä¸ºä¸€ç»„ï¼Œæ„æˆçº¿æ®µ
    for i in range(0, len(coords) - 1, 2):
        x1, y1 = coords[i], coords[i + 1]
        z1 = 0  # å‡è®¾zåæ ‡ä¸º0ï¼Œå¦‚æœéœ€è¦ï¼Œå¯ä»¥é€šè¿‡æŸç§æ–¹å¼è·å–çœŸå®zåæ ‡
        if not vertices:
            vertices.append((x1, y1, z1))
        else:
            # å¦‚æœå½“å‰ç‚¹ä¸ä¸Šä¸€ä¸ªç‚¹ä¸é‡å¤ï¼Œæ·»åŠ åˆ°é¡¶ç‚¹åˆ—è¡¨
            if (x1, y1, z1) != vertices[-1]:
                vertices.append((x1, y1, z1))
    
    # æ·»åŠ æœ€åä¸€ä¸ªç‚¹ï¼Œé¿å…é—æ¼
    last_x, last_y = coords[-2], coords[-1]
    z_last = 0  # åŒæ ·ï¼Œå‡è®¾æœ€åçš„zåæ ‡ä¸º0
    if (last_x, last_y, z_last) != vertices[-1]:
        vertices.append((last_x, last_y, z_last))
    
    return vertices

# 2 å°†comçº¿æ®µè½¬æˆé¡¶ç‚¹åæ ‡åˆ—è¡¨ï¼Œä¸€æ ¹çº¿æ®µä¸€ä¸ªåˆ—è¡¨

def convert_lines_to_points(segments):
    """
    å°†comçº¿æ®µåˆ—è¡¨è½¬æ¢ä¸ºé¡¶ç‚¹åˆ—è¡¨ï¼Œæ¯æ¡çº¿æ®µçš„ä¸¤ä¸ªç«¯ç‚¹ä½œä¸ºä¸€ä¸ªç‹¬ç«‹çš„åˆ—è¡¨ã€‚

    å‚æ•°:
      segments: çº¿æ®µå¯¹è±¡åˆ—è¡¨ï¼Œæ¯ä¸ªå…ƒç´ å…·æœ‰ StartPoint å’Œ EndPoint å±æ€§ã€‚

    è¿”å›:
      åŒ…å«å¤šä¸ªçº¿æ®µé¡¶ç‚¹çš„åˆ—è¡¨ï¼Œæ¯ä¸ªçº¿æ®µæ˜¯ä¸€ä¸ªåŒ…å«ä¸¤ä¸ªç«¯ç‚¹åæ ‡çš„åˆ—è¡¨ã€‚
    """
    points_list = []

    for segment in segments:
        # æå–çº¿æ®µçš„ä¸¤ä¸ªç«¯ç‚¹
        start_point = tuple(segment.StartPoint)
        end_point = tuple(segment.EndPoint)

        # å°†çº¿æ®µçš„ä¸¤ä¸ªç«¯ç‚¹å­˜å…¥ä¸€ä¸ªåˆ—è¡¨ï¼Œæ·»åŠ åˆ°ç»“æœåˆ—è¡¨
        points_list.append([start_point, end_point])

    return points_list

# 3 åˆå¹¶é¡¶ç‚¹åˆ—è¡¨è¡¨ç¤ºçš„è¿ç»­çº¿æ®µï¼Œå…è®¸å¤šæ ¹æ–­å¼€çš„è¿ç»­çº¿æ®µ

def merge_segments_new(LB, tol=0.5):
    """
    ä½¿ç”¨convert_lines_to_points å°†çº¿æ®µå®ä½“è½¬æˆé¡¶ç‚¹åˆ—è¡¨è¡¨è¾¾å¼åå°±å¯ä»¥ä½¿ç”¨æ­¤å‘½ä»¤

    ä¸æ–­åˆå¹¶è¿æ¥ï¼Œèƒ½è¿æ¥çš„éƒ½ä¼šè¿æ¥   

    """
    def same(p, q):
        return abs(p[0]-q[0]) <= tol and abs(p[1]-q[1]) <= tol

    # 1) ä¸ºç«¯ç‚¹åšå“ˆå¸Œ â€” ç”¨ round() æŠŠ tol çº³å…¥ keyï¼Œé¿å…æµ®ç‚¹å­—å…¸é”®éš¾æ¯”è¾ƒ
    def key(pt):
        return (round(pt[0]/tol), round(pt[1]/tol))   # åª hash XY

    buckets = defaultdict(list)   # key(pt)  ->  [(seg_index, dir), ...]
    for idx, seg in enumerate(LB):
        a, b = seg[0], seg[-1]
        buckets[key(a)].append((idx, +1))   #  +1 è¡¨ç¤º seg[0] æ–¹å‘
        buckets[key(b)].append((idx, -1))   #  -1 è¡¨ç¤º seg[-1] æ–¹å‘

    used = [False]*len(LB)
    sequences = []

    while True:
        # æ‰¾åˆ°å°šæœªä½¿ç”¨çš„ç¬¬ä¸€æ¡çº¿æ®µ
        try:
            seed_idx = next(i for i,u in enumerate(used) if not u)
        except StopIteration:
            break                           # å…¨éƒ¨ç”¨å®Œ
        used[seed_idx] = True
        seq = deque(LB[seed_idx])           # åŒç«¯é˜Ÿåˆ—ä¾¿äºé¦–å°¾å¢é•¿

        # å‡½æ•°: æŠŠå¯æ¥çš„çº¿æ®µæ‹¼åˆ° deque çš„ä¸€å¤´
        def grow(at_tail: bool):
            while True:
                end_pt = seq[-1] if at_tail else seq[0]
                bucket = buckets[key(end_pt)]
                # ç§»é™¤å·²ç”¨å®Œçš„éª¨ç‰Œ
                bucket[:] = [pair for pair in bucket if not used[pair[0]]]
                if not bucket:              # å†ä¹Ÿæ¥ä¸ä¸Š
                    break
                idx, direction = bucket.pop()
                used[idx] = True
                seg = LB[idx]
                # æ ¹æ® direction å†³å®šæ­£å‘è¿˜æ˜¯åå‘åŠ å…¥
                if direction == +1:         # bucket ç‚¹æ˜¯ seg[0]
                    add = seg[1:]           # å»æ‰å…¬å…±ç‚¹å†æ‹¼
                else:                       # bucket ç‚¹æ˜¯ seg[-1]
                    add = seg[-2::-1]       # åå‘ã€å»æ‰å…¬å…±ç‚¹
                if at_tail:
                    seq.extend(add)
                else:
                    seq.extendleft(add[::-1])   # extendleft è¦åè½¬

        # å…ˆå¾€å°¾å·´æ‹¼ï¼Œå†å¾€å¤´æ‹¼ï¼ˆé¡ºåºæ— æ‰€è°“ï¼Œéƒ½ä¼šåšåˆ°æé™ï¼‰
        grow(True)
        grow(False)
        sequences.append(list(seq))

    return sequences

# 4 ç»˜åˆ¶è¿ç»­PLå¤šæ®µçº¿ï¼Œæ–­å¼€çš„PLå¤šæ®µçº¿è¦åˆ†å¼€ç»˜åˆ¶

def draw_polyline(vertices,
                  layer_name="æµ‹è¯•è¾…åŠ©",
                  tol=0.5,
                  width=20,
                  color=1):
    """
    å¤çº¿å’Œå¤šè¾¹å½¢éƒ½åº”è¯¥æŒ‰æ ‡å‡†é¡¶ç‚¹åæ ‡åˆ—è¡¨è¡¨è¾¾
    æ ¹æ®é¡¶ç‚¹åºåˆ— vertices åœ¨å½“å‰ AutoCAD æ–‡æ¡£ (å…¨å±€å˜é‡ doc) çš„ ModelSpace
    ç»˜åˆ¶å¤šæ®µçº¿ (PLine)ã€‚

    å‚æ•°
    ----
    vertices : list[tuple]
        ä»…æ”¯æŒé¡¶ç‚¹åˆ—è¡¨å½¢å¼ï¼š[(x, y, z), (x, y, z), ...]ã€‚
    layer_name : str
        ç›®æ ‡å›¾å±‚åï¼›ä¸å­˜åœ¨åˆ™è‡ªåŠ¨åˆ›å»ºã€‚
    tol : float
        åˆ¤æ–­é¦–å°¾æ˜¯å¦é—­åˆçš„å®¹å·®(åªæ¯”è¾ƒ Xâ€†/â€†Y)ã€‚
    width : float
        ConstantWidth è®¾ç½®ã€‚
    color : int
        AutoCAD é¢œè‰²å·ã€‚

    è¿”å›
    ----
    acad_polyline : PLine COM å¯¹è±¡ | None
    """

    # ----------------- å†…éƒ¨å·¥å…· -----------------
##    def same_point(p1, p2, _tol=tol):
##        """åªæ¯”è¾ƒ xã€y åæ ‡çš„è¿‘ä¼¼ç›¸ç­‰"""
##        return abs(p1[0] - p2[0]) <= _tol and abs(p1[1] - p2[1]) <= _tol
    # --------------------------------------------

    # ---------- è¾“å…¥æ ¡éªŒ ----------
    if not vertices or not isinstance(vertices, (list, tuple)):
        print("[é”™è¯¯] è¯·è¾“å…¥æœ‰æ•ˆçš„é¡¶ç‚¹åˆ—è¡¨")
        return None
    if not isinstance(vertices[0], (list, tuple)) or len(vertices[0]) < 3:
        print("[é”™è¯¯] é¡¶ç‚¹æ ¼å¼é”™è¯¯ï¼Œåº”ä¸º (x, y, z)")
        return None

    # --------- å¤„ç†é—­åˆæ€§ ---------
    is_closed = same_point(vertices[0], vertices[-1])

    # --------- æ‰“å°è°ƒè¯•ä¿¡æ¯ --------
    print("è°ƒè¯•ï¼šç»˜åˆ¶å¤šæ®µçº¿çš„é¡¶ç‚¹åºåˆ—")
    for idx, pt in enumerate(vertices):
        print(f"  {idx}: {pt}")

    # --------- ç”Ÿæˆ SAFEARRAY ------
    flat = []
    for x, y, z in vertices:
        flat.extend([x, y, z])
    coords_var = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, tuple(flat))

    # ---------- ç¡®ä¿å›¾å±‚å­˜åœ¨ -------
    try:
        _ = doc.Layers.Item(layer_name)
    except Exception:
        doc.Layers.Add(layer_name)

    # ---------- ç»˜åˆ¶å¤šæ®µçº¿ ----------
    try:
        pl = doc.ModelSpace.AddPolyline(coords_var)
        pl.Closed = is_closed
        pl.Layer = layer_name
        pl.color = color
        pl.ConstantWidth = width
        pl.Update()
        doc.Regen(0)
        print(f"[OK] å·²åœ¨å›¾å±‚ã€{layer_name}ã€ç»˜åˆ¶å¤šæ®µçº¿ï¼ŒClosed = {is_closed}")
        return pl
    except Exception as e:
        print("[é”™è¯¯] ç»˜åˆ¶å¤šæ®µçº¿å¤±è´¥:", e)
        return None



# 5 è·å–å¤šæ®µçº¿åçš„çº¿æ®µåˆ—è¡¨ï¼ŒåŸæ¥çš„å¤šæ®µçº¿ä»ç„¶å­˜åœ¨


"""
pl=  Pl_obj.Explode()

"""



# 6 å°†å¤šæ¡ç›´çº¿æ®µï¼ˆå…è®¸ä¸è¿ç»­ï¼‰è¿æ¥æˆPLå¤çº¿

def lines_to_polylines(Lc,
                       tol=0.5,
                       layer_name="æŸæŸå›¾å±‚",
                       width=20,
                       color=1):
    """
    å°†è‹¥å¹²ç›´çº¿æ®µ (COM Line å¯¹è±¡) è¿ç»­åˆå¹¶æˆå¤šæ®µçº¿ï¼š
      1. ç›´çº¿æ®µ â†’ é¡¶ç‚¹å¯¹åˆ—è¡¨  (convert_lines_to_points)
      2. è¿ç»­æ®µåˆå¹¶         (merge_segments_new)
      3. ç”Ÿæˆå¤šæ®µçº¿         (draw_polyline)
      4. åˆ é™¤åŸç›´çº¿æ®µ

    å‚æ•°
    ----
    Lc : list[Line COM]
        è¦åˆå¹¶çš„ AcadLine å¯¹è±¡åˆ—è¡¨ã€‚
    tol : float
        é¡¶ç‚¹åˆ¤åŒå®¹å·®ã€‚
    layer_name / width / color
        ä¼ é€’ç»™ draw_polyline çš„æ§åˆ¶å‚æ•°ã€‚

    è¿”å›
    ----
    PLs : list[PLine COM]
        æ–°ç”Ÿæˆçš„å¤šæ®µçº¿å¯¹è±¡åˆ—è¡¨ã€‚
    """

    # ---------- 0. è¾¹ç•Œæ£€æŸ¥ ----------
    if not Lc:
        print("[é”™è¯¯] è¾“å…¥çº¿æ®µåˆ—è¡¨ä¸ºç©º")
        return []

    # ---------- 1. çº¿æ®µ â†’ é¡¶ç‚¹å¯¹ --------
    # convert_lines_to_points åº”è¿”å›å½¢å¦‚ [[p1,p2],[p3,p4] ...]
    LB = convert_lines_to_points(Lc)

    # ---------- 2. åˆå¹¶è¿ç»­é¡¶ç‚¹ --------
    # merge_segments_new ä¼šæŠŠ LB åˆå¹¶æˆè‹¥å¹²è¿ç»­é¡¶ç‚¹åºåˆ—
    LK = merge_segments_new(LB, tol=tol)

    # ---------- 3. ç”Ÿæˆå¤šæ®µçº¿ ----------
    PLs = []
    for verts in LK:
        pl = draw_polyline(verts,
                           layer_name=layer_name,
                           tol=tol,
                           width=width,
                           color=color)
        if pl:
            PLs.append(pl)

    # ---------- 4. åˆ é™¤åŸç›´çº¿ ----------
    for ln in Lc:
        try:
            ln.Delete()
        except Exception:
            pass          # è‹¥å·²è¢«åˆ é™¤æˆ–æ— æ•ˆåˆ™å¿½ç•¥

    print(f"[OK] å·²ç”Ÿæˆ {len(PLs)} æ¡å¤šæ®µçº¿ï¼Œå¹¶åˆ é™¤ {len(Lc)} æ¡åŸç›´çº¿")
    return PLs





# 7 æ‰¾åˆ°å¤šæ®µçº¿çš„æœ€å·¦ä¸‹è§’çš„ç‚¹

def find_min_point(obj):
    """
    è·å–ä»»æ„å¯¹è±¡çš„å·¦ä¸‹è§’åæ ‡ï¼ˆé€šè¿‡å…¶å¤–åŒ…ç›’ï¼‰ã€‚

    :param obj: æ”¯æŒ GetBoundingBox() æ–¹æ³•çš„ COM å¯¹è±¡ï¼ˆå¦‚å¤šæ®µçº¿ã€å—å‚ç…§ç­‰ï¼‰ã€‚
    :return:    (min_x, min_y) å…ƒç»„ï¼Œè¡¨ç¤ºå¯¹è±¡å¤–åŒ…ç›’çš„å·¦ä¸‹è§’åæ ‡
    """
    try:
        ll_point, _ = obj.GetBoundingBox()
        min_x, min_y, _ = ll_point
        return min_x, min_y
    except Exception as e:
        print(f"[é”™è¯¯] è·å–å¤–åŒ…ç›’å¤±è´¥: {e}")
        return None, None

# 8 æ‰¾åˆ°å¤šæ®µçº¿çš„æœ€å³ä¸Šè§’çš„ç‚¹

def find_max_point(obj):
    """
    è·å–ä»»æ„å¯¹è±¡çš„å³ä¸Šè§’åæ ‡ï¼ˆé€šè¿‡å…¶å¤–åŒ…ç›’ï¼‰ã€‚

    :param obj: æ”¯æŒ GetBoundingBox() æ–¹æ³•çš„ COM å¯¹è±¡ï¼ˆå¦‚å¤šæ®µçº¿ã€å—å‚ç…§ç­‰ï¼‰ã€‚
    :return:    (max_x, max_y) å…ƒç»„ï¼Œè¡¨ç¤ºå¯¹è±¡å¤–åŒ…ç›’çš„å³ä¸Šè§’åæ ‡
    """
    try:
        _, ur_point = obj.GetBoundingBox()
        max_x, max_y, _ = ur_point
        return max_x, max_y
    except Exception as e:
        print(f"[é”™è¯¯] è·å–å¤–åŒ…ç›’å¤±è´¥: {e}")
        return None, None


def distance(point1, point2):
    """è®¡ç®—ä¸¤ç‚¹ä¹‹é—´çš„è·ç¦»"""
    return ((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)**0.5

# 9 åˆ é™¤å¤šæ®µçº¿åˆ—è¡¨ä¸­é‡å¤çš„å¤šæ®µçº¿

def remove_duplicate_polylines(LB_1):
    """å¤„ç†å¤šæ®µçº¿åˆ—è¡¨ï¼Œåˆ é™¤é‡å¤çš„å¤šæ®µçº¿ï¼Œå¹¶å°†å¤„ç†åçš„å¤šæ®µçº¿æ·»åŠ åˆ°æ–°åˆ—è¡¨ä¸­ã€‚"""
    LB_2 = []

    for XX in LB_1:

        biaoji=0
        try:
            XX.color = 6  # å°è¯•è®¾ç½®é¢œè‰²ï¼Œå¦‚æœå‡ºé”™åˆ™è·³è¿‡
        except :
            biaoji=1

        if biaoji == 0:
           
            # æ£€æŸ¥æ˜¯å¦æœ‰é‡å¤çš„å¤šæ®µçº¿
            for YY in LB_1:

                try:
                    
                    if XX != YY:
                        
                        min_distance = distance(find_min_point(XX), find_min_point(YY))
                        
                        max_distance = distance(find_max_point(XX), find_max_point(YY))

                        if min_distance < 10 and max_distance < 10:
                            
                            YY.Delete()  # å°è¯•åˆ é™¤é‡å¤çš„å¤šæ®µçº¿
                           
                except:

                    pass
           
            LB_2.append(XX)

    return LB_2

# 10 å®šä¹‰çŸ©å½¢

def define_rectangle_by_diagonal(p1, p2):
    """
    ä½¿ç”¨ä¸¤ä¸ªå¯¹è§’é¡¶ç‚¹å®šä¹‰çŸ©å½¢ï¼ŒçŸ©å½¢çš„è¾¹ä¸åæ ‡è½´å¹³è¡Œæˆ–å‚ç›´ã€‚
    p1, p2: åˆ†åˆ«ä¸ºçŸ©å½¢çš„ä¸€å¯¹å¯¹è§’é¡¶ç‚¹ï¼Œæ ¼å¼ä¸º (x, y)ã€‚
    è¿”å›çŸ©å½¢çš„å››ä¸ªé¡¶ç‚¹ã€é•¿å’Œå®½ã€‚
    """
    x1, y1 = p1
    x2, y2 = p2

    # è®¡ç®—ä¸¤ä¸ªè¾¹é•¿
    length_x = abs(x2 - x1)
    length_y = abs(y2 - y1)

    # ç¡®å®šé•¿å’Œå®½
    length = max(length_x, length_y)
    width = min(length_x, length_y)

    # ç¡®å®šçŸ©å½¢çš„å››ä¸ªé¡¶ç‚¹
    rectangle_points = [(x1, y1), (x1, y2), (x2, y2), (x2, y1)]

    return rectangle_points, length, width

def define_rectangle_by_diagonal_x(p1, p2):
    """
    ä½¿ç”¨ä¸¤ä¸ªå¯¹è§’é¡¶ç‚¹å®šä¹‰çŸ©å½¢ï¼ŒçŸ©å½¢çš„è¾¹ä¸åæ ‡è½´å¹³è¡Œæˆ–å‚ç›´ã€‚
    p1, p2: åˆ†åˆ«ä¸ºçŸ©å½¢çš„ä¸€å¯¹å¯¹è§’é¡¶ç‚¹ï¼Œæ ¼å¼ä¸º (x, y)ã€‚
    è¿”å›çŸ©å½¢çš„å››ä¸ªé¡¶ç‚¹ã€é•¿å’Œå®½ã€‚
    """
    x1, y1 = p1
    x2, y2 = p2

    # è®¡ç®—ä¸¤ä¸ªè¾¹é•¿
    length_x = abs(x2 - x1)
    length_y = abs(y2 - y1)

    # ç¡®å®šé•¿å’Œå®½
    length = max(length_x, length_y)
    width = min(length_x, length_y)

    # ç¡®å®šçŸ©å½¢çš„å››ä¸ªé¡¶ç‚¹
    rectangle_points = [x1, y1, x2, y1, x2, y2, x1, y2]

    return rectangle_points




# 11 çŸ©å½¢æ¡†çš„æ‰©å¼ 

def expand_rectangle(p1, p2, offset=130):
    """
    ç»™å®šçŸ©å½¢æ¡†çš„ä¸¤ä¸ªå¯¹è§’ç‚¹ï¼ˆp1 å’Œ p2ï¼‰ï¼Œ
    è¿”å›åœ¨å››ä¸ªæ–¹å‘æ‰©å±• offset åçš„æ–°å¯¹è§’ç‚¹ P1 å’Œ P2ã€‚
    """

    x1, y1, z1 = p1
    x2, y2, z2 = p2

    # ç¡®ä¿ p1 æ˜¯å·¦ä¸‹è§’ï¼Œp2 æ˜¯å³ä¸Šè§’ï¼ˆå³ä½¿è¾“å…¥åäº†ï¼‰
    x_min, x_max = sorted([x1, x2])
    y_min, y_max = sorted([y1, y2])
    z = z1  # z åæ ‡ä¿æŒä¸€è‡´

    # å››å‘æ‰©å±• offsetï¼Œæ„é€ æ–°çš„çŸ©å½¢æ¡†
    P1 = (x_min - offset, y_min - offset, z)
    P2 = (x_max + offset, y_max + offset, z)

    return P1, P2


# 12 çŸ©å½¢æ ‡å‡†åŒ–

def parse_rectangle_points(*args):
    """
    æ¥æ”¶å¤šç§åæ ‡æ ¼å¼è¾“å…¥ï¼Œç»Ÿä¸€è§£æä¸ºçŸ©å½¢å››è§’ç‚¹ï¼š
    è¿”å›ï¼š
        (å·¦ä¸‹, å³ä¸Š, å·¦ä¸Š, å³ä¸‹)ï¼Œæ¯ä¸ªä¸ºä¸‰å…ƒç»„ (x, y, z)
    
    åˆæ³•è¾“å…¥å½¢å¼ï¼š
        - (x1, y1, z1), (x2, y2, z2)
        - [(x1, y1, z1), (x2, y2, z2)]
        - (x1, y1, x2, y2)
        - (x1, y1, 0, x2, y2, 0)
    """
    try:
        # è§£åŒ…åˆ—è¡¨è¾“å…¥
        if len(args) == 1 and isinstance(args[0], list):
            args = args[0]

        # æ ‡å‡†åŒ–ä¸ºä¸¤ä¸ªä¸‰ç»´ç‚¹
        if len(args) == 2 and all(isinstance(pt, (tuple, list)) and len(pt) == 3 for pt in args):
            p1, p2 = args
        elif len(args) == 4:
            x1, y1, x2, y2 = args
            p1, p2 = (x1, y1, 0), (x2, y2, 0)
        elif len(args) == 6:
            x1, y1, z1, x2, y2, z2 = args
            p1, p2 = (x1, y1, z1), (x2, y2, z2)
        else:
            raise ValueError("è¾“å…¥æ ¼å¼ä¸åˆæ³•")

        # è§£æ min/max ç‚¹åæ ‡
        x_min = min(p1[0], p2[0])
        x_max = max(p1[0], p2[0])
        y_min = min(p1[1], p2[1])
        y_max = max(p1[1], p2[1])
        z = p1[2] if len(p1) > 2 else 0

        # å››è§’åæ ‡
        left_bottom = (x_min, y_min, z)
        right_top = (x_max, y_max, z)
        left_top = (x_min, y_max, z)
        right_bottom = (x_max, y_min, z)

        return left_bottom, right_top, left_top, right_bottom

    except Exception as e:
        print(f"[é”™è¯¯] è§£æçŸ©å½¢ç‚¹å¤±è´¥: {e}")
        return None





#  ä¸»å‡½æ•°
#  (2)
#  PLæ‰“å°çº¿

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°


#&&%  1 åˆ¤æ–­å¤šæ®µçº¿æ˜¯å¦åˆä¹æ‰“å°è¦æ±‚çš„å¤šæ®µçº¿è¿”å›å…¶å›¾å·å’Œæ¯”ä¾‹æˆ–0

def generate_name_and_ratio_from_polyline(comobj,A3dy=0,Fandy=("ISO_A3_(420.00_x_297.00_MM)","0:0","A3"),tol=10):



    """
    å‡½æ•°å·²ç»ä¿®æ”¹ä¸ºé’ˆå¯¹ä»»æ„å¯¹è±¡ï¼ŒæŒ‰å…¶å¤–åŒ…ç›’æ¥åˆ†æ

    å½“ä¸€èˆ¬é€‰æ‹©ä¸ºç©ºæ—¶è‡ªåŠ¨ä½œæ¯”ä¾‹é€‰æ‹©ã€‚ä¸å†é‡æ–°ç¼–åˆ¶å‡½æ•°å’Œä¿®æ”¹è¯¥å‡½æ•°ï¼Œç›´æ¥å–å€¼A3dy=1å³å¯


    1)åœ¨ä½¿ç”¨è¯¥å‡½æ•°ä¹‹å‰ï¼Œåº”ä½¿ç”¨å¤šæ®µçº¿å¿«é€Ÿé€‰æ‹©å¹¶é€æ­¥å»æ‰å¤šä½™çš„PLçº¿

    2)è¿™ä¸€æ®µè¡¨è¾¾è¿™æ ·çš„æ€æƒ³ï¼šå¯¹äºé‚£äº›ä¸æ˜¯å¾ˆæ¥è¿‘ä½†è¯¯å·®ä¹Ÿä¸æ˜¯å¾ˆå¤§çš„æƒ…å†µï¼Œæˆ‘ä»¬ä¼šä½¿ç”¨å·²æœ‰å›¾æ¡†ä¸­æœ€æ¥è¿‘çš„ä¸€ä¸ªå»ä½œä¸ºç»“è®ºï¼Œä½†è¿™ä¸ªå¯èƒ½ä¹Ÿæ˜¯é”™è¯¯çš„
    elif diff1 < 200 and diff2 < 200:
                    total_diff = diff1 + diff2  # è®¡ç®—æ€»å·®å€¼
                    if total_diff < closest_diff:
                        closest_diff = total_diff
                        index_pl = i


    index_plå’Œclosest_difféƒ½åœ¨ä¸æ–­åœ°æ›´æ–°ç›´åˆ°å¾—åˆ°æœ€æ¥è¿‘çš„.å½“48åŸºå…ƒæœ‰é è¿‘å¯¹è±¡æ—¶ï¼Œä¼šè®°å½•ä¸‹é è¿‘ç—•è¿¹ï¼Œä½†ä¸ä¼šå½±å“å–å€¼ã€‚ä»…å½“å…¶æœªåŒ¹é…åå·®åˆ°æ›´å‡†ç¡®çš„åŸºå…ƒæ—¶èœæœ‰å¿…è¦æ‰“å°è¿™ä¸ªå‚è€ƒæ¶ˆæ¯ã€‚

    å¦å¤–ï¼ŒåŒ…å«ç€5ç‚¹åˆæ³•æ‰“å°çº¿çš„å¯èƒ½ã€‚å› æ­¤é€»è¾‘ä»å¾…ä¼˜åŒ–ï¼Œä½†ä»æ ¹æœ¬ä¸Šæ²¡é—®é¢˜äº†ã€‚20250507

    3)define_rectangle_by_diagonalæ‰€å®šä¹‰çš„lengthå°±æ˜¯é•¿è¾¹ï¼Œä¸æ˜¯xæ–¹å‘çš„è·ç¦»

    4)Fandy=("ISO_A3_(420.00_x_297.00_MM)","0:0","A3"),("ISO_A2_(594.00_x_420.00_MM)","0:0","A2"),("ISO_A1_(841.00_x_594.00_MM)","0:0","A1"),("ISO_A0_(1189.00_x_841.00_MM)","0:0","A0")
        
    """
    #å‡½æ•°è¿”å›æ‰“å°å¤šæ®µçº¿çš„å›¾å¹…ï¼Œæ¯”ä¾‹ï¼Œå›¾å·ä¿¡æ¯ã€‚å¯¹äºä¸åˆæ³•çš„PLçº¿è¿”å›0å€¼ã€‚

    # 1ï¼‰ å®šä¹‰åŸºæœ¬åˆ—è¡¨
    
    LB_dayingkuang = [
            (118900,84100,100),(178350,126150,150),(59450,42050,50),(29725,21025,25),
            (133763,84100,100),(200644.5,126150,150),(66881.5,42050,50),(33440.75,21025,25),
            (148625,84100,100),(222937.5,126150,150),(74312.5,42050,50),(37156.25,21025,25),
            (84100,59400,100),(126150,89100,150),(42050,29700,50),(21025,14850,25),
            (105125,59400,100),(157687.5,89100,150),(52562.5,29700,50),(26281.25,14850,25),            
            (126150,59400,100),(189225,89100,150),(63075,29700,50),(31537.5,14850,25),            
            (147175,59400,100),(220762.5,89100,150),(73587.5,29700,50),(36793.75,14850,25),
            (59400,42000,100),(89100,63000,150),(29700,21000,50),(14850,10500,25),            
            (74250,42000,100),(111375,63000,150),(37125,21000,50),(18562.5,10500,25),            
            (89100,42000,100),(133650,63000,150),(44550,21000,50),(22275,10500,25),            
            (103950,42000,100),(155925,63000,150),(51975,21000,50),(25987.5,10500,25),            
            (42000,29700,100),(63000,44550,150),(21000,14850,50),(10500,7425,25)            
        ]
    
    drawing_map_ml = [("A0","1:100"),("A0","1:150"),("A0","1:50"),("A0","1:25"),
                      ("A0+1/8","1:100"),("A0+1/8","1:150"),("A0+1/8","1:50"),("A0+1/8","1:25"),
                      ("A0+1/4","1:100"),("A0+1/4","1:150"),("A0+1/4","1:50"),("A0+1/4","1:25"),
                      ("A1","1:100"),("A1","1:150"),("A1","1:50"),("A1","1:25"),
                      ("A1+1/4","1:100"),("A1+1/4","1:150"),("A1+1/4","1:50"),("A1+1/4","1:25"),
                      ("A1+1/2","1:100"),("A1+1/4","1:150"),("A1+1/2","1:50"),("A1+1/2","1:25"),
                      ("A1+3/4","1:100"),("A1+3/4","1:150"),("A1+3/4","1:50"),("A1+3/4","1:25"),
                      ("A2","1:100"),("A2","1:150"),("A2","1:50"),("A2","1:25"),
                      ("A2+1/4","1:100"),("A2+1/4","1:150"),("A2+1/4","1:50"),("A2+1/4","1:25"),
                      ("A2+1/2","1:100"),("A2+1/2","1:150"),("A2+1/2","1:50"),("A2+1/2","1:25"),
                      ("A2+3/4","1:100"),("A2+3/4","1:150"),("A2+3/4","1:50"),("A2+3/4","1:25"),
                      ("A3","1:100"),("A3","1:150"),("A3","1:50"),("A3","1:25")
        ]                      
                      
    drawing_map = ["ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)",
                   "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)", "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)","UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)","UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)",
                   "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)","UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)","UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)","UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)",
                   "ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)",
                   "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)","UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)","UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)","UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)",
                   "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)","UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)","UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)","UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)",
                   "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)","UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)","UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)","UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)",
                   "ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)",
                   "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)","UserDefinedMetric (742.50 x 420.00æ¯«ç±³)","UserDefinedMetric (742.50 x 420.00æ¯«ç±³)","UserDefinedMetric (742.50 x 420.00æ¯«ç±³)",                   
                   "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)","UserDefinedMetric (891.00 x 420.00æ¯«ç±³)","UserDefinedMetric (891.00 x 420.00æ¯«ç±³)","UserDefinedMetric (891.00 x 420.00æ¯«ç±³)",
                   "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)","UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)","UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)","UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)",
                   "ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)"]
                      
    # 2ï¼‰ ç¡®å®šå¤šæ®µçº¿çš„é•¿ä¸å®½ï¼ˆé•¿åº¦å€¼å¤§çš„ä¸ºlengthï¼‰
    PL_min = find_min_point(comobj)
    
    PL_max = find_max_point(comobj)
    
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)


    # 3ï¼‰ ç¡®å®šåºå·

    index_pl = ""
    closest_diff = float('inf')  # åˆå§‹åŒ–æœ€æ¥è¿‘çš„å·®å€¼ä¸ºæ— ç©·å¤§
    if A3dy == 0:

        print("length,width",length,width)
        for i in range(len(LB_dayingkuang)):
            obj = LB_dayingkuang[i]
            obj_length, obj_width, _ = obj
            diff1 = abs(length - obj_length)
            diff2 = abs(width - obj_width)

            if diff1 < tol and diff2 < tol:
                index_pl = i
                break  # å¦‚æœæ‰¾åˆ°äº†éå¸¸æ¥è¿‘çš„ï¼Œå°±ä¸å†ç»§ç»­æŸ¥æ‰¾


            

            elif diff1 < 200 and diff2 < 200:

                print("åå·®å€¼åœ¨10--200ä¹‹é—´,è¯·æ³¨æ„è¯¥å°ºå¯¸")
                total_diff = diff1 + diff2  # è®¡ç®—æ€»å·®å€¼
                if total_diff < closest_diff:
                    closest_diff = total_diff
                    index_pl = i

        # é€€å‡ºå¾ªç¯åç»Ÿä¸€åˆ¤æ–­
        if index_pl != "":

            print("ä½ç½®ï¼š",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            print("æ•°æ®ï¼š",drawing_map[index_pl], drawing_map_ml[index_pl][1], drawing_map_ml[index_pl][0])
            
            return drawing_map[index_pl], drawing_map_ml[index_pl][1], drawing_map_ml[index_pl][0]
        else:


            print("ä½ç½®ï¼š",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            #å¯¹åŸ0å€¼å¯¹è±¡åŠ å¹¿ä¹‰åˆ¤æ–­

            fanhui = get_print_template_info(comobj, tol=tol)

            return fanhui 

    # 4ï¼‰ è¿”å›å€¼

    elif A3dy == 1:

        print("length,width",length,width)
        
        if abs(width/length - 0.707) <= 0.01:

            print("ä½ç½®ï¼š",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            return Fandy[0],Fandy[1],Fandy[2]

        else:            

            print("ä½ç½®ï¼š",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            return 0

    else:

        print("å‚æ•°è¾“å…¥é”™è¯¯")

        return 0
##æ–°çš„å››å…ƒä¿¡æ¯åˆ†ææ£€æµ‹å‡½æ•°

def generate_name_and_ratio_from_com(
    comobj,
    A3dy=0,
    Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3",0),
    tol=10
):
    """
    å‡½æ•°å·²ç»ä¿®æ”¹ä¸ºé’ˆå¯¹ä»»æ„å¯¹è±¡ï¼ŒæŒ‰å…¶å¤–åŒ…ç›’æ¥åˆ†æã€‚
    è¿”å›å€¼ç”±åŸæ¥çš„ä¸‰å…ƒç»„ (å›¾å¹…, æ¯”ä¾‹, å›¾å·) æ‰©å±•ä¸ºå››å…ƒç»„ (å›¾å¹…, æ¯”ä¾‹, å›¾å·, ç«–å‘æ ‡å¿—)ï¼š
      - ç«–å‘æ ‡å¿— = 1 è¡¨ç¤ºå¤–åŒ…ç›’ç«–å‘ (height > width)
      - å¦åˆ™è¿”å› 0ï¼ˆæ¨ªå‘æˆ–æ­£æ–¹å½¢ï¼‰

    å‚æ•°ï¼š
      comobj -- ä»»æ„æ”¯æŒ GetBoundingBox() çš„å¤šæ®µçº¿ COM å¯¹è±¡
      A3dy   -- å½“ A3dy=1 æ—¶ï¼Œç›´æ¥æŒ‰ç…§ Fandy è¿”å›ã€‚å¦åˆ™ä½¿ç”¨ LB_dayingkuang è¿›è¡ŒåŒ¹é…ã€‚
      Fandy  -- é»˜è®¤ ("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3",0)ï¼Œè§„æ¨¡ A3 æƒ…å†µä¸‹è¿”å›çš„ä¿¡æ¯
      tol    -- ä¸æ ‡å‡†å°ºå¯¸æ¯”è¾ƒçš„å®¹å·®ï¼Œé»˜è®¤ä¸º 10

    è¿”å›ï¼š
      (å›¾å¹…, æ¯”ä¾‹, å›¾å·, ç«–å‘æ ‡å¿—) æˆ–è€… 0
    """
    # 1ï¼‰å®šä¹‰åŸºæœ¬å°ºå¯¸åº“
    LB_dayingkuang = [
        (118900, 84100, 100),   (178350, 126150, 150),  (59450, 42050, 50),   (29725, 21025, 25),
        (133763, 84100, 100),   (200644.5, 126150, 150), (66881.5, 42050, 50), (33440.75, 21025, 25),
        (148625, 84100, 100),   (222937.5, 126150, 150),(74312.5, 42050, 50), (37156.25, 21025, 25),
        (84100, 59400, 100),    (126150, 89100, 150),   (42050, 29700, 50),   (21025, 14850, 25),
        (105125, 59400, 100),   (157687.5, 89100, 150), (52562.5, 29700, 50), (26281.25, 14850, 25),
        (126150, 59400, 100),   (189225, 89100, 150),   (63075, 29700, 50),   (31537.5, 14850, 25),
        (147175, 59400, 100),   (220762.5, 89100, 150), (73587.5, 29700, 50), (36793.75, 14850, 25),
        (59400, 42000, 100),    (89100, 63000, 150),    (29700, 21000, 50),   (14850, 10500, 25),
        (74250, 42000, 100),    (111375, 63000, 150),   (37125, 21000, 50),   (18562.5, 10500, 25),
        (89100, 42000, 100),    (133650, 63000, 150),   (44550, 21000, 50),   (22275, 10500, 25),
        (103950, 42000, 100),   (155925, 63000, 150),   (51975, 21000, 50),   (25987.5, 10500, 25),
        (42000, 29700, 100),    (63000, 44550, 150),    (21000, 14850, 50),   (10500, 7425, 25)
    ]

    drawing_map_ml = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"),  ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"),  ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"),  ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"),  ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"),  ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"),  ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"),  ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"),  ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"),  ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"),  ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"),  ("A3", "1:25")
    ]

    drawing_map = [
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)", "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)",
        "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)", "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)",
        "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)", "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)",
        "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)", "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)", "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)", "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)", "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)", "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)", "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)", "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)", "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)", "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)", "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)", "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)", "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)", "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)"
    ]

    # 2ï¼‰è®¡ç®—å¤–åŒ…ç›’çš„é•¿ã€å®½ï¼Œä»¥åŠæ¨ªç«–ä¿¡æ¯
    PL_min = find_min_point(comobj)
    PL_max = find_max_point(comobj)
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)

    # è®¡ç®— dx, dy ä»¥åˆ¤æ–­ç«–å‘æˆ–æ¨ªå‘
    dx = abs(PL_max[0] - PL_min[0])
    dy = abs(PL_max[1] - PL_min[1])
    orientation_flag = 1 if dy > dx else 0

    # 3ï¼‰ç¡®å®šåŒ¹é…åºå· index_pl
    index_pl = ""
    closest_diff = float('inf')

    if A3dy == 0:
        print("length, width =", length, width)
        for i, (obj_length, obj_width, _) in enumerate(LB_dayingkuang):
            diff1 = abs(length - obj_length)
            diff2 = abs(width - obj_width)

            # å®Œå…¨åŒ¹é…
            if diff1 < tol and diff2 < tol:
                index_pl = i
                break

            # è¿‘ä¼¼åŒ¹é…çš„å€™é€‰
            elif diff1 < 200 and diff2 < 200:
                total_diff = diff1 + diff2
                if total_diff < closest_diff:
                    closest_diff = total_diff
                    index_pl = i

        # åˆ¤æ–­å¹¶è¿”å›
        if index_pl != "":
            print("ä½ç½®ï¼š", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return (
                drawing_map[index_pl],
                drawing_map_ml[index_pl][1],
                drawing_map_ml[index_pl][0],
                orientation_flag
            )
        else:
            print("ä½ç½®ï¼š", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            fanhui = get_print_template_info(comobj, tol=tol)
            # å¦‚æœ fanhui æ˜¯ä¸‰å…ƒç»„ï¼Œåˆ™æ‹¼æ¥ orientation_flagï¼›å¦åˆ™è¿”å› 0
            if isinstance(fanhui, tuple) and len(fanhui) == 3:
                return (*fanhui, orientation_flag)
            return 0

    # 4ï¼‰å½“ A3dy == 1 æ—¶ï¼Œç›´æ¥æ ¹æ®é•¿å®½æ¯”è¿”å› Fandy å¹¶é™„åŠ  orientation_flag
    elif A3dy == 1:
        print("length, width =", length, width)
        if abs(width / length - 0.707) <= 0.01:
            print("ä½ç½®ï¼š", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return (*Fandy, orientation_flag)
        else:
            print("ä½ç½®ï¼š", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return 0

    else:
        print("å‚æ•°è¾“å…¥é”™è¯¯")
        return 0


##20250711ä¿®æ”¹
def generate_name_and_ratio_from_com(
    comobj,
    A3dy=0,
    Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0),
    tol=10
):
    """
    è¿”å› (å›¾å¹…, æ¯”ä¾‹, å›¾å·, ç«–å‘æ ‡å¿—)ã€‚æœªå‘½ä¸­æ—¶è¿”å› 0ã€‚

    â‘  ç²¾ç¡®å‘½ä¸­ â†’ åŸæ ·è¿”å›  
    â‘¡ å‘½ä¸­ â€œÃ—1.2â€  â†’ è¿”å›åŸæ ‡å‡†æ•°æ®ï¼Œå¹¶æŠŠå¯¹è±¡æ”¹æˆè“è‰²  
    â‘¢ è¿‘ä¼¼å‘½ä¸­     â†’ è¿”å›åŸæ ‡å‡†æ•°æ®ï¼Œå¹¶æŠŠå¯¹è±¡æ”¹æˆçº¢è‰² + å®½åº¦ (200 / 2)

    å¯¹æ‹Ÿåˆæ ‡å‡†æ¡†çº¿ï¼ŒåŠ äº†çº¢è‰²è­¦å‘Šå’ŒåŠ ç²—æç¤ºï¼Œå¯¹æ”¾å¤§1.2å€çš„æ‰“å°æ¡†çº¿åŠ äº†è“è‰²æç¤º

    """

    # â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€” 1. åŸºç¡€æ•°æ®åŒº â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”
    LB_dayingkuang = [
        (118900, 84100, 100),   (178350, 126150, 150),  (59450, 42050, 50),   (29725, 21025, 25),
        (133763, 84100, 100),   (200644.5, 126150, 150), (66881.5, 42050, 50), (33440.75, 21025, 25),
        (148625, 84100, 100),   (222937.5, 126150, 150),(74312.5, 42050, 50), (37156.25, 21025, 25),
        (84100, 59400, 100),    (126150, 89100, 150),   (42050, 29700, 50),   (21025, 14850, 25),
        (105125, 59400, 100),   (157687.5, 89100, 150), (52562.5, 29700, 50), (26281.25, 14850, 25),
        (126150, 59400, 100),   (189225, 89100, 150),   (63075, 29700, 50),   (31537.5, 14850, 25),
        (147175, 59400, 100),   (220762.5, 89100, 150), (73587.5, 29700, 50), (36793.75, 14850, 25),
        (59400, 42000, 100),    (89100, 63000, 150),    (29700, 21000, 50),   (14850, 10500, 25),
        (74250, 42000, 100),    (111375, 63000, 150),   (37125, 21000, 50),   (18562.5, 10500, 25),
        (89100, 42000, 100),    (133650, 63000, 150),   (44550, 21000, 50),   (22275, 10500, 25),
        (103950, 42000, 100),   (155925, 63000, 150),   (51975, 21000, 50),   (25987.5, 10500, 25),
        (42000, 29700, 100),    (63000, 44550, 150),    (21000, 14850, 50),   (10500, 7425, 25)
    ]

    drawing_map_ml = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"),  ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"),  ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"),  ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"),  ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"),  ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"),  ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"),  ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"),  ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"),  ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"),  ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"),  ("A3", "1:25")
    ]

    drawing_map = [
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)", "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)",
        "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)", "UserDefinedMetric (1337.63 x 841.00æ¯«ç±³)",
        "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)", "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)",
        "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)", "UserDefinedMetric (1486.25 x 841.00æ¯«ç±³)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)", "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)", "UserDefinedMetric (1051.25 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)", "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)", "UserDefinedMetric (1261.50 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)", "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)",
        "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)", "UserDefinedMetric (1471.75 x 594.00æ¯«ç±³)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)", "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)", "UserDefinedMetric (742.50 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)", "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)", "UserDefinedMetric (891.00 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)", "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)",
        "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)", "UserDefinedMetric (1039.50 x 420.00æ¯«ç±³)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)"
    ]

    # â€”â€”â€”â€”â€”â€”â€”â€”â€”â€” 2. å¤–åŒ…ç›’ä¸æœå‘ â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”
    PL_min = find_min_point(comobj)
    PL_max = find_max_point(comobj)
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)
    dx, dy = abs(PL_max[0] - PL_min[0]), abs(PL_max[1] - PL_min[1])
    orientation_flag = 1 if dy > dx else 0

    # â€”â€”â€”â€”â€”â€”â€”â€”â€”â€” 3. åŒ¹é…ç±»å‹åˆ¤å®š â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”
    best_i, best_type = None, None      # 'exact' | 'scale12' | 'approx'
    best_score = float('inf')           # ä»…ç”¨äº approx

    if A3dy == 0:
        for i, (std_len, std_wid, _) in enumerate(LB_dayingkuang):
            # ç²¾ç¡®
            if abs(length - std_len) < tol and abs(width - std_wid) < tol:
                best_i, best_type = i, 'exact'
                break

            # Ã—1.2
            if abs(length - 1.2*std_len) < tol and abs(width - 1.2*std_wid) < tol:
                best_i, best_type = i, 'scale12'
                break

            # è¿‘ä¼¼
            diff = max(abs(length - std_len), abs(width - std_wid))
            if diff < 200 and diff < best_score:
                best_i, best_type = i, 'approx'
                best_score = diff

        # â€”â€”â€”â€”â€”â€”â€”â€”â€”â€” 4. å‘½ä¸­åçš„å¤„ç† â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”
        if best_i is not None:
            zhi = (
                drawing_map[best_i],
                drawing_map_ml[best_i][1],
                drawing_map_ml[best_i][0],
                orientation_flag
            )

            if best_type == 'scale12':          # è“
                try:
                    comobj.color = 5
                except Exception:
                    pass

            elif best_type == 'approx':         # çº¢ + åŠ ç²—
                try:
                    comobj.color = 1
                except Exception:
                    pass
                try:
                    total_len = float(getattr(comobj, "Length", 0))
                except Exception:
                    total_len = 0
                new_wid = 2 if total_len and total_len < 10000 else 200
                try:
                    comobj.ConstantWidth = new_wid
                except Exception:
                    try:
                        comobj.SetWidth(new_wid, new_wid)
                    except Exception:
                        pass

            return zhi

        # â€”â€” æœªå‘½ä¸­ï¼šè°ƒç”¨å¤–éƒ¨æ¨¡æ¿åˆ¤å®š â€”â€” 
        fanhui = get_print_template_info(comobj, tol=tol)
        return (*fanhui, orientation_flag) if isinstance(fanhui, tuple) else 0

    # â€”â€”â€”â€”â€”â€”â€”â€”â€”â€” 5. A3dy == 1 ç›´æ¥ç”¨é•¿å®½æ¯”åˆ¤å®š â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”
    elif A3dy == 1:
        if abs(width / length - 0.707) <= 0.01:
            return (*Fandy, orientation_flag)
        return 0

    # â€”â€”â€”â€”â€”â€”â€”â€”â€”â€” 6. å‚æ•°é”™è¯¯ â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”
    else:
        print("å‚æ•°è¾“å…¥é”™è¯¯")
        return 0





#  2 comå¤çº¿ä»ä¸Šåˆ°ä¸‹ä»å·¦åˆ°å³æ’åº


def polyline_sort(polyline_list):
    """å¯¹comå¤šæ®µçº¿æŒ‰ç…§ç‰¹å®šè§„åˆ™è¿›è¡Œæ’åº"""

    # å­˜å‚¨å¤šæ®µçº¿åŠå…¶æœ€å·¦ä¸‹è§’ç‚¹
    polylines_with_min_points = [(pl, find_min_point(pl)) for pl in polyline_list]

    # å…ˆæŒ‰ç…§yå€¼é™åºæ’åº
    polylines_with_min_points.sort(key=lambda item: -item[1][1])

    # å†å¯¹yå€¼å·®è·åœ¨1000ä»¥å†…çš„å¤šæ®µçº¿æŒ‰ç…§xå€¼å‡åºæ’åº
    i = 0
    while i < len(polylines_with_min_points) - 1:
        j = i + 1
        # æŸ¥æ‰¾æ‰€æœ‰yå€¼å·®è·åœ¨1000ä»¥å†…çš„å¤šæ®µçº¿
        while j < len(polylines_with_min_points) and abs(polylines_with_min_points[i][1][1] - polylines_with_min_points[j][1][1]) < 1000:
            j += 1
        
        # å¦‚æœæ‰¾åˆ°äº†yå€¼ç›¸è¿‘çš„å¤šæ®µçº¿ï¼Œæ ¹æ®xå€¼è¿›è¡Œæ’åº
        if j - i > 1:
            polylines_with_min_points[i:j] = sorted(polylines_with_min_points[i:j], key=lambda item: item[1][0])
        
        i = j

    # åªè¿”å›å¤šæ®µçº¿å¯¹è±¡
    return [item[0] for item in polylines_with_min_points]




#&&&%  *** 3 å°†PLcomçº¿åˆ—è¡¨çš„åæ ‡ä¿¡æ¯å­˜å‚¨


def plcom_to_coor(plines):
    """
    æ¥å—å¤šæ ¹è½»é‡çº§å¤šæ®µçº¿æˆ–å¸¸è§„å¤šæ®µçº¿çš„ COM å¯¹è±¡åˆ—è¡¨ï¼Œè¿”å›å®ƒä»¬çš„åæ ‡åˆ—è¡¨åŠé—­åˆçŠ¶æ€ã€‚

    :param plines: å¯è¿­ä»£çš„ä¸€ç»„ LWPOLYLINE æˆ– POLYLINE COM å¯¹è±¡
    :return: åˆ—è¡¨ï¼Œæ¯ä¸ªå…ƒç´ ä¸º (pts, closed_flag)ï¼š
             - pts: é¡¶ç‚¹åˆ—è¡¨ [(x0, y0), (x1, y1), â€¦]
             - closed_flag: 1 è¡¨ç¤ºé—­åˆï¼Œ0 è¡¨ç¤ºæœªé—­åˆ

    å…¼å®¹ä¸¤ç§æƒ…å†µï¼š
      1. LWPOLYLINE.Coordinates â†’ å¶æ•°é•¿åº¦ï¼Œä¾‹å¦‚ [x0,y0, x1,y1, x2,y2, â€¦]
      2. POLYLINE.Coordinates  â†’ 3 çš„å€æ•°é•¿åº¦ï¼Œä¾‹å¦‚ [x0,y0,z0, x1,y1,z1, â€¦]

    å¦‚æœæ—¢ä¸æ˜¯å¶æ•°ä¹Ÿä¸æ˜¯ 3 çš„å€æ•°ï¼Œå°†è·³è¿‡è¯¥æ¡å¤šæ®µçº¿å¹¶æ‰“å° WARN æç¤ºã€‚
    """
    plines = ensure_list(plines)
    all_info = []

    for pl in plines:
        raw = list(pl.Coordinates)  # å¯èƒ½æ˜¯å¶æ•°é•¿åº¦ (LWPOLYLINE) æˆ– 3 çš„å€æ•°é•¿åº¦ (POLYLINE)

        pts = []
        # â€”â€” æƒ…å†µ Aï¼šå¦‚æœé•¿åº¦èƒ½è¢« 3 æ•´é™¤ï¼Œè®¤ä¸ºæ˜¯å¸¸è§„ POLYLINE â†’ æ¯ 3 ä¸ªæ•°ä¸€ç»„ (x,y,z)
        if len(raw) % 3 == 0 and len(raw) > 0:
            for i in range(0, len(raw), 3):
                x = raw[i]
                y = raw[i + 1]
                pts.append((x, y))

        # â€”â€” æƒ…å†µ Bï¼šå¦åˆ™å¦‚æœé•¿åº¦èƒ½è¢« 2 æ•´é™¤ï¼Œè®¤ä¸ºæ˜¯è½»é‡çº§ LWPOLYLINE â†’ æ¯ 2 ä¸ªæ•°ä¸€ç»„ (x,y)
        elif len(raw) % 2 == 0 and len(raw) > 0:
            for i in range(0, len(raw), 2):
                x = raw[i]
                y = raw[i + 1]
                pts.append((x, y))

        else:
            # æ—¢ä¸æ˜¯ 2 çš„å€æ•°ä¹Ÿä¸æ˜¯ 3 çš„å€æ•°ï¼šåæ ‡æ•°æ®å¼‚å¸¸ï¼Œè·³è¿‡è¿™ä¸€æ¡ï¼Œæ‰“å° WARN
            handle = getattr(pl, "Handle", "<unknown>")
            print(f"[WARN] plcom_to_coorï¼šè·³è¿‡ Handle={handle} çš„å¤šæ®µçº¿ï¼Œ"
                  f"Coordinates é•¿åº¦={len(raw)} æ—¢é 2 çš„å€æ•°ä¹Ÿé 3 çš„å€æ•°ã€‚")
            continue

        # è¯»å– Closed å±æ€§ï¼ŒTrue è¡¨ç¤ºé—­åˆ
        closed_flag = 1 if getattr(pl, "Closed", False) else 0

        all_info.append((pts, closed_flag))

    return all_info


# 4 ä»åæ ‡ä¿¡æ¯åˆ—è¡¨è¿”å›PLcomçº¿åˆ—è¡¨

    
def plcoor_to_com(coord_info, layer_name="æµ‹è¯•è¾…åŠ©", width=0, color=256):
    """
    åœ¨å½“å‰ DWG ä¸­æ ¹æ®åæ ‡å’Œå°é—­æ ‡å¿—ç»˜åˆ¶å¤šæ¡è½»é‡çº§å¤šæ®µçº¿ã€‚

    :param coord_info: åˆ—è¡¨ï¼Œæ¯ä¸ªå…ƒç´ ä¸º (pts, closed_flag)ï¼Œ
                       pts ä¸º [(x0,y0),â€¦] é¡¶ç‚¹åˆ—è¡¨ï¼Œ
                       closed_flag ä¸º 1ï¼ˆé—­åˆï¼‰æˆ– 0ï¼ˆä¸é—­åˆï¼‰ã€‚
    :param layer_name: ç›®æ ‡å›¾å±‚åç§°ï¼ˆä¸å­˜åœ¨åˆ™åˆ›å»ºï¼‰ï¼Œé»˜è®¤ "æµ‹è¯•è¾…åŠ©"
    :param width:      å¤šæ®µçº¿å®½åº¦ï¼Œé»˜è®¤ 0
    :param color:      é¢œè‰²ç´¢å¼•ï¼Œé»˜è®¤ 256ï¼ˆBYLAYERï¼‰
    :return:           ç»˜åˆ¶çš„å¤šæ®µçº¿å¯¹è±¡åˆ—è¡¨
    """
    # 1) è¿æ¥ AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    ms   = doc.ModelSpace

    # 2) ç¡®ä¿å›¾å±‚å­˜åœ¨
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    lyr.LayerOn = True

    created = []
    for pts, closed_flag in coord_info:
        # å°† pts å±•å¹³ä¸º [x0,y0,x1,y1,â€¦]
        raw = []
        for x, y in pts:
            raw.extend((x, y))
        # è½¬ä¸º COM æ•°ç»„
        arr = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            raw
        )
        # æ·»åŠ è½»é‡çº§å¤šæ®µçº¿
        lw = ms.AddLightWeightPolyline(arr)
        lw.Layer         = layer_name
        lw.ConstantWidth = width
        lw.color         = color
        lw.Closed        = bool(closed_flag)
        created.append(lw)

    # å¯é€‰ï¼šç¼©æ”¾åˆ°å¯è§èŒƒå›´
    acad.ZoomExtents()
    print(f"[OK] å·²ç»˜åˆ¶ {len(created)} æ¡è½»é‡çº§å¤šæ®µçº¿åˆ°å›¾å±‚ â€œ{layer_name}â€")
    return created






# 5 ç¡®å®šå¤šæ®µçº¿æ‰“å°æ¡†æ˜¯å¦ç«–å‘

def panduan_shuxiangkuang(polyline):

    PL_min = find_min_point(polyline)
    
    PL_max = find_max_point(polyline)

    cha_x  = abs(PL_max[0] - PL_min[0])   

    cha_y  = abs(PL_max[1] - PL_min[1])

    if  cha_y > cha_x:

        return True

    else :

        return False


# 6 ç»Ÿä¸€ä¸ºA2çš„å›¾å¹…æ‰“å°

def tongyi_tufu(LB,TFname):


    """
    å°†æ‰“å°çº¿åˆ—è¡¨çš„æ¯æ ¹çº¿å¯¹åº”çš„å›¾çº¸å°ºå¯¸ç»Ÿä¸€ä¸ºä¸€ä¸ªTFname
        
    """

    LB_xin = []

    for ob in LB:

        LB_xin.append(TFname)#"ISO_A2_(594.00_x_420.00_MM)","ISO_A3_(420.00_x_297.00_MM)"

    return LB_xin



#  ä¸»å‡½æ•°
#  (3)
#  å°†æ­£äº¤å…­è¾¹å½¢å¤šæ®µçº¿åˆ†æˆä¸¤ä¸ªçŸ©å½¢åŒºåŸŸ

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

"""
è¯¥å‡½æ•°å°†æ­£äº¤å…­è¾¹å½¢å¤šæ®µçº¿åˆ†æˆä¸¤ä¸ªçŸ©å½¢åŒºåŸŸï¼Œåªé’ˆå¯¹æ ‡å‡†å…­ç‚¹å…­è¾¹çš„PLå¤šè¾¹å½¢
åœ¨å¤„ç†å…­ç‚¹å…­è¾¹çš„æ­£äº¤PLå½¢å¤šè¾¹å½¢ä¹‹å‰ï¼Œä½¿ç”¨simplify_polygon(polygon, tol=0.5)å°†ä¼ªè¾¹ç‚¹æ¸…é™¤
            
"""

# æ¶ˆé™¤å¤šè¾¹å½¢çš„ä¼ªè¾¹ç‚¹

def simplify_polygon(poly, tol=0.5):
    """
    ç®€åŒ–å¤šè¾¹å½¢é¡¶ç‚¹åˆ—è¡¨ï¼šå¦‚æœæŸé¡¶ç‚¹ P ä¸å…¶å‰åä¸¤ç‚¹å…±çº¿ï¼ˆåœ¨å®¹å·® tol èŒƒå›´å†…ï¼‰ï¼Œåˆ™å°†å…¶ç§»é™¤ã€‚
    é¦–å°¾ç›¸è¿å¤„ç†ï¼šç¬¬ä¸€ä¸ªç‚¹çš„â€œå‰ç‚¹â€æ˜¯æœ€åä¸€ä¸ªç‚¹ï¼Œæœ€åä¸€ä¸ªç‚¹çš„â€œæ¬¡ç‚¹â€æ˜¯ç¬¬ä¸€ä¸ªç‚¹ã€‚
    
    å‚æ•°:
        poly: [(x, y, z), â€¦]  åŸå§‹é¡¶ç‚¹åˆ—è¡¨ï¼ˆå¯èƒ½é¦–å°¾é‡å¤æˆ–æœ‰å¤šä½™é¡¶ç‚¹ï¼‰
        tol:  å…±çº¿åˆ¤æ–­çš„å®¹å·®ï¼ˆå¯¹åº”å‰ç§¯ç»å¯¹å€¼ï¼‰
    
    è¿”å›:
        ç®€åŒ–åçš„é¡¶ç‚¹åˆ—è¡¨ï¼ˆåŒæ ·ä¿ç•™é¦–å°¾æ˜¯å¦é—­åˆçš„å½¢å¼ï¼Œä¸ä¼šè‡ªåŠ¨å»é‡é¦–å°¾ï¼‰
    """
    # å…ˆåšä¸€æ¬¡â€œå»æ‰é¦–å°¾é‡å¤â€ä»¥å…æ— é™å¾ªç¯
    if len(poly) > 1 and poly[0] == poly[-1]:
        poly = poly[:-1]

    def is_colinear(p_prev, p, p_next):
        # åªæ¯”è¾ƒ x,yï¼Œè®¡ç®— (p - p_prev) Ã— (p_next - p) çš„â€œå‰ç§¯â€
        x1, y1 = p[0] - p_prev[0], p[1] - p_prev[1]
        x2, y2 = p_next[0] - p[0],   p_next[1] - p[1]
        cross = x1 * y2 - y1 * x2
        return abs(cross) <= tol

    # é‡å¤æ‰«ä¸€éèƒ½åˆ å°±åˆ ï¼Œç›´åˆ°ä¸€è½®ä¸‹æ¥æ²¡æœ‰åˆ é™¤
    changed = True
    while changed and len(poly) >= 3:
        changed = False
        n = len(poly)
        for i in range(n):
            prev_idx = (i - 1) % n
            next_idx = (i + 1) % n
            if is_colinear(poly[prev_idx], poly[i], poly[next_idx]):
                # åˆ é™¤ç¬¬ i ä¸ªç‚¹ï¼Œé‡ç½®ä¸€è½®æ‰«æ
                del poly[i]
                changed = True
                break
    return poly



# 1. æ ‡å‡†åŒ–å¤šè¾¹å½¢é¡¶ç‚¹åˆ—è¡¨ï¼Œå»æ‰ç›¸é‚»å’Œé¦–å°¾é‡å¤ç‚¹
def normalize_polygon(polygon):
    """
    æ ‡å‡†åŒ–å¤šè¾¹å½¢é¡¶ç‚¹åˆ—è¡¨ï¼š  
      - å»æ‰ä»»æ„ç›¸é‚»é‡å¤çš„ç‚¹  
      - å¦‚æœé¦–å°¾ç›¸åŒï¼Œåˆ™å»æ‰æœ«å°¾é‚£ä¸ª  
      
    å‚æ•°:
        polygon: åŸå§‹é¡¶ç‚¹åˆ—è¡¨ï¼Œæ¯ä¸ªç‚¹ä¸º (x, y, z)
    è¿”å›:
        å»é‡åçš„é¡¶ç‚¹åˆ—è¡¨
    """
    if not polygon:
        return []
    normalized = [polygon[0]]
    for pt in polygon[1:]:
        if pt != normalized[-1]:
            normalized.append(pt)
    # é¦–å°¾ç›¸åŒåˆ™åˆ å°¾
    if len(normalized) > 1 and normalized[0] == normalized[-1]:
        normalized.pop()
    return normalized


# 2. æ‰¾åˆ°æŸé¡¶ç‚¹çš„å‰é©±/åç»§ï¼ˆæŒ‰å¾ªç¯å¤šè¾¹å½¢ï¼‰
def get_adjacent_points(polygon, p):
    """
    åœ¨å¤šè¾¹å½¢ polygon ä¸­è¿”å›é¡¶ç‚¹ p çš„å‰åç›¸é‚»ç‚¹ï¼ˆæ”¯æŒå¾ªç¯ï¼‰ã€‚
    ä¼šå…ˆè°ƒç”¨ normalize_polygon æ¸…ç†é‡å¤ç‚¹ã€‚
    """
    poly = normalize_polygon(polygon)
    if not poly:
        raise ValueError("å¤šè¾¹å½¢ä¸ºç©º")
    try:
        idx = poly.index(p)
    except ValueError:
        raise ValueError(f"ç‚¹ {p} ä¸åœ¨å¤šè¾¹å½¢é¡¶ç‚¹åˆ—è¡¨ä¸­")
    prev_pt = poly[idx - 1] if idx > 0 else poly[-1]
    next_pt = poly[idx + 1] if idx < len(poly) - 1 else poly[0]
    return prev_pt, next_pt


# 3. ç‚¹æ˜¯å¦åœ¨å¤šè¾¹å½¢å†…éƒ¨ï¼ˆå°„çº¿æ³•ï¼Œä»…åœ¨ XY å¹³é¢åˆ¤æ–­ï¼‰
def point_in_polygon(pt, polygon):
    """
    åˆ¤æ–­ä¸‰ç»´ç‚¹ pt=(x,y,z) åœ¨å¤šè¾¹å½¢ polygon çš„ XY æŠ•å½±å†…å¦ã€‚
    polygon ä¸­ç‚¹æ ¼å¼ä¸º (x,y,z)ï¼Œé¦–å°¾å¯é‡å¤æˆ–ä¸é‡å¤å‡å¯ã€‚
    """
    x, y, _ = pt
    poly2d = [(q[0], q[1]) for q in normalize_polygon(polygon)]
    inside = False
    n = len(poly2d)
    for i in range(n):
        x1, y1 = poly2d[i]
        x2, y2 = poly2d[(i + 1) % n]
        if (y1 > y) != (y2 > y):
            xint = x1 + (y - y1) * (x2 - x1) / (y2 - y1)
            if xint > x:
                inside = not inside
    return inside


# 4. æ— ç©·ç›´çº¿ vs çº¿æ®µåœ¨ XY å¹³é¢ç›¸äº¤
def line_segment_intersection_2d(p, d, a, b, tol=1e-8):
    """
    è®¡ç®—å°„çº¿ L(t)=p + tÂ·d ä¸çº¿æ®µ AB åœ¨ XY å¹³é¢ä¸Šçš„äº¤ç‚¹ã€‚
    p,d,a,b çš†ä¸º (x,y,z)ï¼Œä½†åªå– x,y åˆ†é‡å‚ä¸è®¡ç®—ã€‚
    è¿”å› (xi, yi, t) æˆ– Noneã€‚
    """
    px, py, _ = p
    dx, dy, _ = d
    ax, ay, _ = a
    bx, by, _ = b
    ux, uy = bx - ax, by - ay

    det = dx * (-uy) - dy * (-ux)
    if abs(det) < tol:
        return None

    rhsx, rhsy = ax - px, ay - py
    t_param = (rhsx * (-uy) - rhsy * (-ux)) / det
    u_param = (dx * rhsy - dy * rhsx) / det

    if -tol <= u_param <= 1 + tol:
        xi = px + t_param * dx
        yi = py + t_param * dy
        return xi, yi, t_param
    return None


# 5. è®¡ç®— p å’Œå…¶ç›¸é‚»ç‚¹ä¸­ç‚¹ cï¼Œå¦‚æœ c å†…éƒ¨åˆ™è¿”å› cï¼Œå¦åˆ™æ²¿ p->c çš„å°„çº¿æ‰¾åˆ°ç¬¬ä¸€ä¸ªäº¤ç‚¹
def get_auxiliary_point(p, p_prev, p_next, polygon, tol=1e-8):
    """
    å¯¹äºå¤šè¾¹å½¢é¡¶ç‚¹ p åŠå…¶å‰åç›¸é‚»ç‚¹ p_prev, p_nextï¼Œ
    è¿”å›ä¸€ä¸ªä½äºå¤šè¾¹å½¢å†…éƒ¨çš„è¾…åŠ©ç‚¹ qï¼š
      1. å…ˆå– c = (p_prev + p_next)/2ï¼›è‹¥ c åœ¨å†…éƒ¨ï¼Œåˆ™è¿”å› c
      2. å¦åˆ™æ²¿å°„çº¿ p->c ä¸å¤šè¾¹å½¢å…¶å®ƒè¾¹æ±‚æœ€é è¿‘ p çš„äº¤ç‚¹
    è¿”å›ç‚¹æ ¼å¼ä¸º (x,y,z)ã€‚
    """
    # 1) ä¸­ç‚¹ c
    cx = (p_prev[0] + p_next[0]) / 2
    cy = (p_prev[1] + p_next[1]) / 2
    cz = (p_prev[2] + p_next[2]) / 2
    c = (cx, cy, cz)

    if point_in_polygon(c, polygon):
        return c

    # 2) æ„é€ æ–¹å‘ d = c - p å¹¶å½’ä¸€åŒ–
    dx, dy = cx - p[0], cy - p[1]
    mag = math.hypot(dx, dy)
    if mag < tol:
        raise RuntimeError("è¾…åŠ©ç‚¹æ–¹å‘å‘é‡è¿‡å°")
    d = (dx / mag, dy / mag, 0.0)

    # åœ¨æ¯æ¡ä¸å« p çš„è¾¹ä¸Šæ±‚äº¤
    poly = normalize_polygon(polygon)
    intersects = []
    for i in range(len(poly)):
        a = poly[i]
        b = poly[(i + 1) % len(poly)]
        if a == p or b == p:
            continue
        res = line_segment_intersection_2d(p, d, a, b, tol)
        if res:
            xi, yi, t_param = res
            if abs(t_param) > tol:
                # æ’å€¼å‡ºå¯¹åº”çš„ z
                zi = p[2] + t_param * d[2]
                intersects.append((xi, yi, zi, t_param))

    if not intersects:
        raise RuntimeError("æœªæ‰¾åˆ°æœ‰æ•ˆäº¤ç‚¹")
    # å–æœ€å° |t| çš„é‚£ä¸€ä¸ª
    intersects.sort(key=lambda it: abs(it[3]))
    xi, yi, zi, _ = intersects[0]
    return (xi, yi, zi)


# 6. è®¡ç®— p ç‚¹çš„â€œå‡¹å‡¸åº¦é‡è§’â€
def concavity_measure(p, p_prev, p_next, q):
    """
    ç»™å®š p, p_prev, p_next, qï¼ˆå‡ä¸º (x,y,z)ï¼‰ï¼Œ
    è®¡ç®—åº¦é‡è§’ï¼š
      angle = 360 - larger_angle + smaller_angle  
    å…¶ä¸­ smaller/larger æ˜¯ pqâ†’p_prev ä¸ pqâ†’p_next çš„é€†æ—¶é’ˆå¤¹è§’ã€‚
    å‡¸ç‚¹çº¦ ~90Â°ï¼Œå‡¹ç‚¹çº¦ ~270Â°ã€‚
    """
    def angle_of(vx, vy):
        a = math.degrees(math.atan2(vy, vx))
        return a if a >= 0 else a + 360

    # æ„é€  2D å‘é‡
    vq = (q[0] - p[0], q[1] - p[1])
    v1 = (p_prev[0] - p[0], p_prev[1] - p[1])
    v2 = (p_next[0] - p[0], p_next[1] - p[1])

    a_q = angle_of(*vq)
    a1  = angle_of(*v1)
    a2  = angle_of(*v2)

    d1 = (a1 - a_q) % 360
    d2 = (a2 - a_q) % 360

    small, large = (d1, d2) if d1 < d2 else (d2, d1)
    return 360 - large + small


# 7. ç›´æ¥ç»™å‡º p åœ¨å¤šè¾¹å½¢ä¸Šçš„åº¦é‡è§’
def concavity_angle(p, polygon):
    """
    ç›´æ¥è®¡ç®—å¤šè¾¹å½¢ polygon ä¸Šé¡¶ç‚¹ p çš„å‡¹å‡¸åº¦é‡è§’ã€‚
    """
    p_prev, p_next = get_adjacent_points(polygon, p)
    q = get_auxiliary_point(p, p_prev, p_next, polygon)
    return concavity_measure(p, p_prev, p_next, q)





# 8.åˆç†åˆ†å‰²PLæ­£äº¤å…­è¾¹å½¢

def split_orthogonal_hexagon(polygon, tol=0.1):#æ°´å¹³åˆ†å‰²
    """
    å°†æ­£äº¤å…­è¾¹å½¢ polygon æŒ‰å‡¹é¡¶ç‚¹æ‰€åœ¨æ°´å¹³çº¿åˆ‡æˆä¸¤ä¸ªçŸ©å½¢ã€‚
    polygon: 6 ä¸ª (x,y,z) é¡¶ç‚¹çš„åˆ—è¡¨ï¼Œå…è®¸é¦–å°¾é‡åˆæˆ–ç›¸é‚»é‡å¤ï¼Œä¼šè‡ªåŠ¨è§„èŒƒåŒ–ã€‚
    """
    # 1. è§„èŒƒåŒ–ï¼Œå»æ‰ç›¸é‚»é‡å¤å’Œé¦–å°¾åŒç‚¹
    poly = normalize_polygon(polygon)
    if len(poly) != 6:
        raise ValueError("å¿…é¡»ä¼ å…¥6ç‚¹æ­£äº¤å¤šè¾¹å½¢")
    # 2. æ‰¾å‡ºå”¯ä¸€çš„å‡¹ç‚¹ p
    concaves = [pt for pt in poly
                if abs(concavity_angle(pt, poly) - 270) < tol]
    if len(concaves) != 1:
        raise RuntimeError(f"æ²¡èƒ½å”¯ä¸€å®šä½å‡¹ç‚¹ï¼Œæ‰¾åˆ° {len(concaves)} ä¸ª")
    p = concaves[0]
    y0 = p[1]

    # 3. åªå¯¹çœŸæ­£â€œè·¨è¶Šâ€ y=y0 çš„è¾¹æ±‚äº¤
    intersections = []
    n = len(poly)
    for i in range(n):
        a = poly[i]
        b = poly[(i+1) % n]
        y1, y2 = a[1], b[1]
        # ä»…å½“ä¸¥æ ¼è·¨è¶Šæ‰ç®—ï¼šä¸€ç«¯åœ¨ä¸Š (y>y0)ï¼Œä¸€ç«¯åœ¨ä¸‹ (y<y0)
        if (y1 - y0) * (y2 - y0) < -tol**2:
            t = (y0 - y1) / (y2 - y1)
            xi = a[0] + t * (b[0] - a[0])
            zi = a[2] + t * (b[2] - a[2])
            intersections.append((xi, y0, zi, i))

    # 4. åº”è¯¥åªå‰©ä¸€ä¸ªçœŸçš„ crossing äº¤ç‚¹
    if len(intersections) != 1:
        raise RuntimeError(f"æ²¡èƒ½å”¯ä¸€å®šä½ qï¼Œæ‰¾åˆ° {len(intersections)} ä¸ªå€™é€‰ç‚¹")
    xi, yi, zi, edge_idx = intersections[0]
    q = (xi, yi, zi)

    # 5. æŠŠ q æ’å›é‚£æ¡è¾¹ä¹‹å
    newpoly = []
    for i in range(n):
        newpoly.append(poly[i])
        if i == edge_idx:
            newpoly.append(q)
    # now len(newpoly)==7

    # 6. æ‰¾ p, q çš„ç´¢å¼•
    i_p = newpoly.index(p)
    i_q = newpoly.index(q)

    # 7. åˆ†å‰²æˆä¸¤æ®µå¤šè¾¹å½¢
    if i_q < i_p:
        rect1 = newpoly[i_q:i_p+1]
        rect2 = newpoly[i_p:] + newpoly[:i_q+1]
    else:
        rect1 = newpoly[i_p:i_q+1]
        rect2 = newpoly[i_q:] + newpoly[:i_p+1]

    # 8. è®¡ç®—2Dé¢ç§¯
    def area2d(pts):
        s = 0
        m = len(pts)
        for j in range(m):
            x1,y1,_ = pts[j]
            x2,y2,_ = pts[(j+1)%m]
            s += x1*y2 - x2*y1
        return abs(s)/2

    A1, A2 = area2d(rect1), area2d(rect2)
    # é¢ç§¯å°çš„æ”¾å‰é¢
    return (rect1, rect2) if A1 <= A2 else (rect2, rect1)


def split_orthogonal_hexagon_vertical(polygon, tol=0.1):#ç«–å‘åˆ†å‰²
    """
    å°†æ­£äº¤å…­è¾¹å½¢ polygon æŒ‰å‡¹é¡¶ç‚¹æ‰€åœ¨ç«–çº¿åˆ‡æˆä¸¤ä¸ªçŸ©å½¢ã€‚
    polygon: 6 ä¸ª (x,y,z) é¡¶ç‚¹çš„åˆ—è¡¨ï¼Œå…è®¸é¦–å°¾é‡åˆæˆ–ç›¸é‚»é‡å¤ï¼Œä¼šè‡ªåŠ¨è§„èŒƒåŒ–ã€‚
    tol: ç”¨äºè¯†åˆ«å‡¹ç‚¹å’Œè·¨è¶Šåˆ¤æ–­çš„å®¹å·®ã€‚
    è¿”å›: (rect1, rect2)ï¼Œé¢ç§¯å°çš„æ”¾å‰é¢ã€‚
    """
    # è§„èŒƒåŒ–ï¼šå»æ‰ç›¸é‚»é‡å¤å’Œé¦–å°¾åŒç‚¹
    poly = normalize_polygon(polygon)
    if len(poly) != 6:
        raise ValueError("å¿…é¡»ä¼ å…¥6ç‚¹æ­£äº¤å¤šè¾¹å½¢")

    # æ‰¾å”¯ä¸€å‡¹ç‚¹
    concaves = [pt for pt in poly
                if abs(concavity_angle(pt, poly) - 270) < tol]
    if len(concaves) != 1:
        raise RuntimeError(f"æ²¡èƒ½å”¯ä¸€å®šä½å‡¹ç‚¹ï¼Œæ‰¾åˆ° {len(concaves)} ä¸ª")
    p = concaves[0]
    x0 = p[0]

    # æ±‚ç«–çº¿ x=x0 ä¸çœŸæ­£è·¨è¶Šè¾¹çš„äº¤ç‚¹
    intersections = []
    n = len(poly)
    for i in range(n):
        a = poly[i]
        b = poly[(i+1)%n]
        x1, x2 = a[0], b[0]
        # ä»…å¯¹ä¸¥æ ¼è·¨è¶Šçš„è¾¹æ±‚äº¤
        if (x1 - x0)*(x2 - x0) < -tol**2:
            # æ’å€¼æ¯”ä¾‹
            t = (x0 - x1)/(x2 - x1)
            yi = a[1] + t*(b[1] - a[1])
            zi = a[2] + t*(b[2] - a[2])
            intersections.append((x0, yi, zi, i))

    if len(intersections) != 1:
        raise RuntimeError(f"æ²¡èƒ½å”¯ä¸€å®šä½ qï¼Œæ‰¾åˆ° {len(intersections)} ä¸ªå€™é€‰ç‚¹")
    xi, yi, zi, edge_idx = intersections[0]
    q = (xi, yi, zi)

    # æŠŠ q æ’å›é‚£æ¡è¾¹
    newpoly = []
    for i in range(n):
        newpoly.append(poly[i])
        if i == edge_idx:
            newpoly.append(q)
    # newpoly é•¿åº¦åº”ä¸º7

    # å®šä½ pã€q ç´¢å¼•
    i_p = newpoly.index(p)
    i_q = newpoly.index(q)

    # åˆ†å‰²ä¸¤æ®µ
    if i_q < i_p:
        rect1 = newpoly[i_q:i_p+1]
        rect2 = newpoly[i_p:] + newpoly[:i_q+1]
    else:
        rect1 = newpoly[i_p:i_q+1]
        rect2 = newpoly[i_q:] + newpoly[:i_p+1]

    # é¢ç§¯è®¡ç®—ï¼ˆ2Dï¼‰
    def area2d(pts):
        s = 0
        m = len(pts)
        for j in range(m):
            x1,y1,_ = pts[j]
            x2,y2,_ = pts[(j+1)%m]
            s += x1*y2 - x2*y1
        return abs(s)*0.5

    A1, A2 = area2d(rect1), area2d(rect2)
    return (rect1, rect2) if A1 <= A2 else (rect2, rect1)


# åˆç†åˆ†å‰²PLæ­£äº¤å…­è¾¹å½¢

def area_of(verts):
    """å¤šè¾¹å½¢é¢ç§¯è®¡ç®—ï¼ˆé¡¶ç‚¹é¦–å°¾é—­åˆæˆ–ä¸é—­åˆå‡å¯ï¼‰"""
    s = 0
    n = len(verts)
    for i in range(n):
        x1, y1, *_ = verts[i]
        x2, y2, *_ = verts[(i+1) % n]
        s += x1 * y2 - x2 * y1
    return abs(s) * 0.5



def split_hexagon_combined(polygon, tol=0.1, simplify_tol=0.5):# åˆç†åˆ†å‰²PLæ­£äº¤å…­è¾¹å½¢
    """
    åˆç†åˆ†å‰²ä¸€ä¸ªæ­£äº¤ï¼ˆè¿‘ä¼¼ï¼‰å…­è¾¹å½¢ PLï¼š
      1) å¦‚æœä¼ å…¥çš„æ˜¯ COM PL å¯¹è±¡ï¼Œåˆ™å…ˆæå–å”¯ä¸€é¡¶ç‚¹åˆ—è¡¨ï¼›
      2) å¯¹é¡¶ç‚¹åˆ—è¡¨åšç®€åŒ–ï¼ˆå»é™¤ä¼ªé¡¶ç‚¹/ä¼ªè¾¹ï¼‰ï¼›
      3) å…ˆåšæ¨ªå‘åˆ†å‰²ï¼Œå†åšç«–å‘åˆ†å‰²ï¼Œæ¯”è¾ƒä¸¤ç§åˆ†å‰²æœ€å°çŸ©å½¢é¢ç§¯ï¼Œ
         å°†æœ€å°çš„é‚£å¯¹çŸ©å½¢æ’åœ¨å‰é¢ï¼Œè¿”å›å››ä¸ªçŸ©å½¢é¡¶ç‚¹åˆ—è¡¨ï¼š
         [min1, partner1, min2, partner2]
    å‚æ•°:
      polygon: è¦åˆ†å‰²çš„å¤šæ®µçº¿ï¼Œæ—¢å¯ä»¥æ˜¯ [(x,y,z),...] é¡¶ç‚¹åˆ—è¡¨ï¼Œä¹Ÿå¯ä»¥æ˜¯ COM PL å¯¹è±¡
      tol: åˆ†å‰²æ—¶åˆ¤æ–­å‡¹è§’çš„å®¹å·®
      simplify_tol: ç®€åŒ–å¤šè¾¹å½¢æ—¶å»é™¤â€œä¼ªé¡¶ç‚¹/ä¼ªè¾¹â€çš„å®¹å·®
    è¿”å›:
      å››ä¸ªçŸ©å½¢çš„é¡¶ç‚¹åˆ—è¡¨ï¼šæœ€å°çŸ©å½¢ã€å…¶åŒç»„çŸ©å½¢ã€æ¬¡å°çŸ©å½¢ã€å…¶åŒç»„çŸ©å½¢
    """
    # â€”â€” 1. å¦‚æœæ˜¯ COM PL å¯¹è±¡ï¼Œå…ˆæå–é¡¶ç‚¹åˆ—è¡¨ â€”â€” 
    # ï¼ˆå‡å®šå·²æœ‰ get_unique_vertices_from_pl_com(com_pl) -> [(x,y,z),...]ï¼‰
    if not isinstance(polygon, list):
        polygon = get_unique_vertices_from_pl_com(polygon)

    # â€”â€” 2. ç®€åŒ–é¡¶ç‚¹åˆ—è¡¨ â€”â€” 
    # ï¼ˆå‡å®šå·²æœ‰ simplify_polygon(verts, tol) -> æ¸…ç†åçš„é¡¶ç‚¹åˆ—è¡¨ï¼‰
    polygon = simplify_polygon(polygon, simplify_tol)

    # â€”â€” 3. æ¨ªå‘ä¸ç«–å‘åˆ†å‰² â€”â€” 
    rects_h = split_orthogonal_hexagon(polygon, tol)
    rects_v = split_orthogonal_hexagon_vertical(polygon, tol)

    A_h1, A_h2 = rects_h
    A_v1, A_v2 = rects_v

    # è®¡ç®—æ¯å¯¹çš„â€œæ›´å°â€çŸ©å½¢å’Œâ€œé…å¯¹â€çŸ©å½¢
    if area_of(A_h1) <= area_of(A_h2):
        min_h, partner_h = A_h1, A_h2
    else:
        min_h, partner_h = A_h2, A_h1

    if area_of(A_v1) <= area_of(A_v2):
        min_v, partner_v = A_v1, A_v2
    else:
        min_v, partner_v = A_v2, A_v1

    # â€”â€” 4. æ ¹æ®æœ€å°é¢ç§¯å€¼å†³å®šè¾“å‡ºé¡ºåº â€”â€” 
    if area_of(min_h) <= area_of(min_v):
        return [min_h, partner_h, min_v, partner_v]
    else:
        return [min_v, partner_v, min_h, partner_h]



#  ä¸»å‡½æ•°
#  (4)
#  è·å–å¤šæ®µçº¿çš„ä¸Šä¸‹å·¦å³è¾¹ç•Œçš„ç›´çº¿æ®µï¼Œè¿”å›çº¿æ®µç«¯ç‚¹åˆ—è¡¨

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

def get_bbox_edge_segments(pl, tol=0.5):
    """
    è·å–å¯¹è±¡ pl çš„åŒ…å›´ç›’å››æ¡è¾¹ï¼Œåˆ†åˆ«ä½œä¸ºç‹¬ç«‹çš„åˆ—è¡¨è¿”å›ï¼š
      top    = [(xmin, ymax, z), (xmax, ymax, z)]
      bottom = [(xmin, ymin, z), (xmax, ymin, z)]
      left   = [(xmin, ymin, z), (xmin, ymax, z)]
      right  = [(xmax, ymin, z), (xmax, ymax, z)]
    å¹¶æ‰“å°è°ƒè¯•ä¿¡æ¯ã€‚
    """
    # ----- 1. è°ƒç”¨ GetBoundingBox -----
    try:
        min_pt, max_pt = pl.GetBoundingBox()
    except Exception:
        mins = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))
        maxs = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))
        pl.GetBoundingBox(mins, maxs)
        min_pt = tuple(mins.value)
        max_pt = tuple(maxs.value)

    xmin, ymin, zmin = min_pt
    xmax, ymax, _    = max_pt

    print(f"â–¶ BoundingBox Min ç‚¹: {min_pt}")
    print(f"â–¶ BoundingBox Max ç‚¹: {max_pt}")

    # ----- 2. æ„é€ å››ä¸ªé¡¶ç‚¹ï¼ˆé¡ºæ—¶é’ˆï¼‰ -----
    p1 = (xmin, ymin, zmin)
    p2 = (xmax, ymin, zmin)
    p3 = (xmax, ymax, zmin)
    p4 = (xmin, ymax, zmin)

    print("â–¶ çŸ©å½¢å››ä¸ªé¡¶ç‚¹ (é¡ºæ—¶é’ˆ):")
    for i, pt in enumerate((p1, p2, p3, p4), 1):
        print(f"   {i}: {pt}")

    # ----- 3. å››æ¡è¾¹ï¼Œå„è‡ªç”¨åˆ—è¡¨è¡¨è¾¾ -----
    top    = [p4, p3]    # y = ymax
    bottom = [p1, p2]    # y = ymin
    left   = [p1, p4]    # x = xmin
    right  = [p2, p3]    # x = xmax

    return top, bottom, left, right



#  ä¸»å‡½æ•°
#  (5)
#&&%  è·å–å¤šæ®µçº¿çš„å†…éƒ¨çš„æ–‡å­—

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°



def get_texts_in_polyline(com_pl, tol=0.5):
    """
    åœ¨å¤šæ®µçº¿ com_pl å†…éƒ¨ç­›é€‰æ–‡å­—ï¼Œå¹¶è¿”å›æ–‡å­—å¯¹è±¡åˆ—è¡¨å’Œå¯¹åº”çš„æ–‡å­—å†…å®¹åˆ—è¡¨ã€‚

    å‚æ•°:
      com_pl:  COM å¤šæ®µçº¿å¯¹è±¡
      tol:     ç‚¹-in-å¤šè¾¹å½¢æ—¶çš„å®¹å·®ï¼ˆç›®å‰æœªç”¨åˆ°ï¼Œå¯ç•™ä½œå°†æ¥æ‰©å±•ï¼‰

    è¿”å›:
      inside:   è½åœ¨ com_pl å†…éƒ¨çš„æ–‡å­— COM å¯¹è±¡åˆ—è¡¨
      contents: å¯¹åº” inside ä¸­æ¯ä¸ªå¯¹è±¡çš„æ–‡å­—å†…å®¹åˆ—è¡¨
    """
    # 1) å¤šæ®µçº¿è½¬é¡¶ç‚¹åˆ—è¡¨ï¼ˆæ ‡å‡†åŒ–åçš„ä¸‰ç»´ç‚¹ï¼‰
    poly = get_unique_vertices_from_pl_com(com_pl)

    # 2) æ”¶é›†æ‰€æœ‰æ–‡å­—å®ä½“ï¼ˆå¤©æ­£ï¼‹åŸç”Ÿ CADï¼‰
    tzh_text, tzh_mtext, cad_text, cad_mtext = collect_all_texts()
    all_texts = tzh_text + tzh_mtext + cad_text + cad_mtext

    inside = []
    contents = []

    for txt in all_texts:
        # ç”¨å·¦ä¸‹è§’ç‚¹åˆ¤æ–­æ˜¯å¦åœ¨å¤šæ®µçº¿å†…
        min_pt, _ = txt.GetBoundingBox()
        if point_in_polygon(min_pt, poly):
            inside.append(txt)

            # æ ¹æ®å¯¹è±¡ç±»å‹å–å‡ºå†…å®¹
            name = getattr(txt, "ObjectName", "") or getattr(txt, "EntityName", "")
            if name in ("AcDbText", "AcDbMText"):
                # AutoCAD åŸç”Ÿå•/å¤šè¡Œ
                contents.append(txt.TextString)
            elif name == "TDbText":
                # å¤©æ­£å•è¡Œ
                contents.append(txt.Text)
            elif name == "TDbMText":
                # å¤©æ­£å¤šè¡Œï¼Œéœ€è¦ç‚¸å¼€å–å†…å®¹
                contents.append( TDbMText_content(txt))
            else:
                # å…¶å®ƒï¼ˆä¸‡ä¸€æœ‰ï¼‰ï¼Œç•™ç©ºæˆ–ç”¨ repr
                contents.append("")

    print(f"æ€»å…±æ‰¾åˆ° {len(inside)} æ¡è½åœ¨å¤šæ®µçº¿å†…éƒ¨çš„æ–‡å­—ã€‚")
    return inside, contents




# è·å–å•ç‹¬ä¸€è¡Œçš„å¤©æ­£å¤šè¡Œæ–‡å­—å†…å®¹

def TDbMText_content(comobj):

    """
    å¯¹å¤æ‚å¤šè¡Œæ–‡æœ¬å†…å®¹ï¼Œéœ€è¦ä¸“é—¨åˆ†æ

    æœ¬å‡½æ•°ä»…é’ˆå¯¹å•ä¸ªæ–‡å­—ï¼Œè·å–å…¶æ–‡å­—å†…å®¹

    ä¸æ”¹å˜å…¶æœ¬èº«å±æ€§å’Œå…¶å®ƒå¯¹è±¡çš„å…³ç³»

    """

    highlight_entity_by_bbox( comobj)
    cmd = "x\n"
    doc.SendCommand(cmd)

    ob=doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
    neirong=ob.TextString

    cmd = "u\n"

    doc.SendCommand(cmd)

    return neirong







#  ä¸»å‡½æ•°
#  (6)
#  å¤šæ®µçº¿ä¸Šçš„å‡åˆ†æ’å…¥


"""
è¯¥å‡½æ•°ç”¨äºåœ¨dwgæ–‡ä»¶æ²¿ç€PLçº¿å¿«é€Ÿå‡è¡¡æ”¾ç½®æ ‘æœ¨ç­‰å›¾å—

            
"""
def distribute_points_on_entity(entity, n, block, scale_factor, ys):

    def distance(p1, p2):
        return math.sqrt((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2)

    model_space = doc.ModelSpace
    block_name = block.Name  # è·å–å—çš„åç§°
    
    # å¦‚æœå®ä½“æ˜¯ç›´çº¿
    if entity.ObjectName == "AcDbLine":
        start_point = entity.StartPoint
        end_point = entity.EndPoint
        for i in range(n):
            x = start_point[0] + i * (end_point[0] - start_point[0]) / (n - 1)
            y = start_point[1] + i * (end_point[1] - start_point[1]) / (n - 1)
            inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
            inserted_block.color = ys  # è®¾ç½®é¢œè‰²ä¸ºçº¢è‰²

    # å¦‚æœå®ä½“æ˜¯åœ†å¼§
    elif entity.ObjectName == "AcDbArc":
        start_angle = entity.StartAngle
        end_angle = entity.EndAngle
        center = entity.Center
        radius = entity.Radius
        for i in range(n):
            angle = start_angle + i * (end_angle - start_angle) / (n - 1)
            x = center[0] + radius * math.cos(angle)
            y = center[1] + radius * math.sin(angle)
            inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
            inserted_block.color = ys  # è®¾ç½®é¢œè‰²ä¸ºçº¢è‰²

    # å¦‚æœå®ä½“æ˜¯å¤šæ®µçº¿
    elif entity.ObjectName == "AcDbPolyline":
        coords = entity.Coordinates
        points = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        total_length = sum(distance(points[i], points[i+1]) for i in range(len(points)-1))
        segment_length = total_length / n
        current_length = 0

        for i in range(n):
            accumulated_length = 0
            for j in range(len(points) - 1):
                segment = distance(points[j], points[j+1])
                if accumulated_length + segment > current_length:
                    ratio = (current_length - accumulated_length) / segment
                    x = points[j][0] + ratio * (points[j+1][0] - points[j][0])
                    y = points[j][1] + ratio * (points[j+1][1] - points[j][1])
                    inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
                    inserted_block.color = ys  # è®¾ç½®é¢œè‰²ä¸ºçº¢è‰²
                    break
                accumulated_length += segment
            current_length += segment_length
            




#  ä¸»å‡½æ•°
#  (7)
# è¿”å› pl1 ä¸­ä¸ pl2 â€œå…±çº¿ä¸”æœ‰é‡å â€çš„åŒºæ®µåˆ—è¡¨

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

# 1 åˆ¤æ–­ä¸€æ¡ç›´çº¿æ˜¯å¦å®Œå…¨åœ¨å¦ä¸€æ¡ç›´çº¿ä¸Š

def is_segment_contained(seg_a, seg_b, tol=1e-4):
    """
    åˆ¤æ–­ seg_a æ˜¯å¦å®Œå…¨ä½äº seg_b ä¸Šï¼ˆåŒ…å«ç«¯ç‚¹ï¼‰ã€‚
    
    å‚æ•°
    ----
    seg_a, seg_b :  ( (x1,y1,z1), (x2,y2,z2) )   æˆ–   AcDbLine
        ä¸¤æ¡å¾…æ¯”è¾ƒçš„çº¿æ®µã€‚å…ˆåˆ¤æ–­ seg_a æ˜¯å¦è¢« seg_b è¦†ç›–ï¼›
        è‹¥æƒ³åŒå‘åˆ¤æ–­ï¼Œè°ƒç”¨ä¸¤æ¬¡å¹¶å¯¹è°ƒé¡ºåºå³å¯ã€‚
    tol : float
        è·ç¦»ä¸æŠ•å½±è¯¯å·®å®¹å·®ï¼Œé»˜è®¤ 1eâ€‘4 (CAD å•ä½)ã€‚

    è¿”å›
    ----
    bool
        True  â€”â€” seg_a æ•´æ®µè½åœ¨ seg_b ä¸Š  
        False â€”â€” å¦åˆ™
    """
    # -------- æŠŠè¾“å…¥ç»Ÿä¸€è½¬æˆç«¯ç‚¹å…ƒç»„ ----------
    def get_endpoints(entity):
        if hasattr(entity, "StartPoint"):            # COM çº¿æ®µ
            return (tuple(entity.StartPoint), tuple(entity.EndPoint))
        else:                                        # çº¯åæ ‡äºŒå…ƒç»„
            return (tuple(entity[0]), tuple(entity[1]))

    a1, a2 = get_endpoints(seg_a)
    b1, b2 = get_endpoints(seg_b)

    # -------- åŸºæœ¬å‡ ä½•å·¥å…· ----------
    def dist(p, q):
        return math.hypot(p[0]-q[0], p[1]-q[1])

    def dot(u, v):
        return u[0]*v[0] + u[1]*v[1]

    def colinear(p, q, r, tol):
        """ä¸‰ç‚¹æ˜¯å¦è¿‘ä¼¼å…±çº¿ï¼ˆé¢ç§¯â‰ˆ0ï¼‰"""
        return abs( (q[0]-p[0])*(r[1]-p[1]) - (q[1]-p[1])*(r[0]-p[0]) ) <= tol

    # -------- å…ˆæ£€æµ‹å…±çº¿æ€§ ----------
    if not (colinear(b1, b2, a1, tol) and colinear(b1, b2, a2, tol)):
        return False          # seg_a ç«¯ç‚¹ä¸ä¸ seg_b å…±çº¿ â†’ ä¸å¯èƒ½åŒ…å«

    # -------- å†æ£€æµ‹æŠ•å½±æ˜¯å¦åœ¨ b1â€‘b2 åŒºé—´å†… ----------
    # ä»¤ b1 ä¸ºåŸç‚¹ï¼Œb_dir ä¸ºæ–¹å‘å‘é‡
    b_dir = (b2[0]-b1[0], b2[1]-b1[1])
    b_len2 = dot(b_dir, b_dir)
    if b_len2 == 0:           # seg_b é•¿åº¦ä¸º 0 â†’ æ— æ³•åŒ…å«ä»–æ®µ
        return False

    def proj_param(p):
        # æ ‡å‡†åŒ–æŠ•å½±å‚æ•° tï¼Œè‹¥ 0<=t<=1 åˆ™æŠ•å½±è½åœ¨ seg_b ä¸Š
        return dot( (p[0]-b1[0], p[1]-b1[1]), b_dir ) / b_len2

    t_a1 = proj_param(a1)
    t_a2 = proj_param(a2)

    # å…è®¸ä»‹äº 0Â±tol åˆ° 1Â±tol ä¹‹é—´
    inside_1 = -tol <= t_a1 <= 1+tol
    inside_2 = -tol <= t_a2 <= 1+tol

    return inside_1 and inside_2

#&&% 2 è¿”å› PLçº¿pl1 ä¸­ä¸ pl2 â€œå…±çº¿ä¸”æœ‰é‡å â€çš„åŒºæ®µåˆ—è¡¨

def common_segments_between_polylines(pl1, pl2, tol=0.5):
    """
    è¿”å› pl1 ä¸­ä¸ pl2 â€œå…±çº¿ä¸”æœ‰é‡å â€çš„åŒºæ®µåˆ—è¡¨ï¼Œæ¯ä¸ªåŒºæ®µç”¨
      [(x1,y1,0.0),(x2,y2,0.0)] è¡¨ç¤ºã€‚

    å‚æ•°
    ----
    pl1, pl2 : AutoCAD AcDbPolyline COM å¯¹è±¡ (æˆ–ä¼ªé€ å¯¹è±¡ï¼Œåªéœ€æœ‰ .Coordinates / .Closed)
    tol      : å…±çº¿ & è·ç¦»å®¹å·® (åŒ CAD å•ä½)

    è¿”å›
    ----
    overlaps : list[ list[(x,y,z),(x,y,z)] ]
    """

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ å†…éƒ¨å°å·¥å…· â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    def coords_to_xy_pairs(coords):
        """æŠŠ (x1,y1,x2,y2,â€¦) è½¬æˆ [(x,y),â€¦]ï¼›è‡ªåŠ¨ä¸¢å¼ƒå°¾éƒ¨æ®‹å€¼"""
        pairs = []
        for i in range(0, len(coords) - 1, 2):
            pairs.append((coords[i], coords[i + 1]))
        return pairs

    def build_segments(verts, closed=False):
        """ç”±é¡¶ç‚¹é¡ºåºç”Ÿæˆçº¿æ®µ [P,Q] åˆ—è¡¨"""
        segs = []
        for i in range(len(verts) - 1):
            segs.append([verts[i], verts[i + 1]])
        if closed and len(verts) > 2:
            segs.append([verts[-1], verts[0]])
        return segs

    def dist(p, q):
        return math.hypot(p[0] - q[0], p[1] - q[1])

    def colinear(p, q, r):
        """ä¸‰ç‚¹å…±çº¿åˆ¤å®š |å‰ç§¯| < tol Â· max(è¾¹é•¿)"""
        return abs((q[0] - p[0]) * (r[1] - p[1]) -
                   (q[1] - p[1]) * (r[0] - p[0])) <= tol * max(dist(p, q), dist(p, r), dist(q, r), 1)

    def project(p, axis):
        """æ ¹æ® axis==0 ç”¨ xï¼Œå¦åˆ™ y"""
        return p[axis]

    def segment_overlap(seg_a, seg_b):
        """
        è‹¥å…±çº¿ä¸”åŒºé—´æœ‰é‡å ï¼Œè¿”å›å®é™…é‡å åŒºé—´ç«¯ç‚¹ (pa, pb)ï¼ˆäºŒç»´ç‚¹ï¼‰ã€‚
        å¦åˆ™è¿”å› None
        """
        (p1, p2), (q1, q2) = seg_a, seg_b
        # å…±çº¿æ£€æµ‹
        if not (colinear(p1, p2, q1) and colinear(p1, p2, q2)):
            return None

        # é€‰æŠ•å½±è½´
        axis = 0 if abs(p2[0] - p1[0]) >= abs(p2[1] - p1[1]) else 1
        a1, a2 = project(p1, axis), project(p2, axis)
        b1, b2 = project(q1, axis), project(q2, axis)
        # ä½¿ a1 <= a2, b1 <= b2
        if a1 > a2:
            p1, p2 = p2, p1
            a1, a2 = a2, a1
        if b1 > b2:
            q1, q2 = q2, q1
            b1, b2 = b2, b1

        # è®¡ç®— 1â€‘D é‡å åŒºé—´
        left, right = max(a1, b1), min(a2, b2)
        if right - left <= tol:          # â€œé•¿åº¦â€â‰ˆ0 è§†ä¸ºæ— é‡å 
            return None

        # æŠŠæŠ•å½±ç‚¹è¿˜åŸåˆ° 2D ç«¯ç‚¹ â€”â€” åœ¨çº¿æ®µ p1â€‘p2 ä¸ŠæŒ‰æ¯”ä¾‹å–å€¼
        def interp(t):
            """t ä¸º 0~1"""
            return (p1[0] + t * (p2[0] - p1[0]),
                    p1[1] + t * (p2[1] - p1[1]))

        len_p = a2 - a1 if a2 != a1 else 1e-9
        pa = interp((left - a1) / len_p)
        pb = interp((right - a1) / len_p)
        return pa, pb

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ ä¸»æµç¨‹ â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    v1 = coords_to_xy_pairs(pl1.Coordinates)
    v2 = coords_to_xy_pairs(pl2.Coordinates)

    # è‹¥ä¸ºé—­åˆ polylineï¼Œè¡¥ä¸€ä¸ªé¦–å°¾é¡¶ç‚¹
    if getattr(pl1, "Closed", False) and v1 and v1[0] != v1[-1]:
        v1.append(v1[0])
    if getattr(pl2, "Closed", False) and v2 and v2[0] != v2[-1]:
        v2.append(v2[0])

    segs1 = build_segments(v1, closed=False)
    segs2 = build_segments(v2, closed=False)

    overlaps = []

    for s1 in segs1:
        for s2 in segs2:
            ov = segment_overlap(s1, s2)
            if ov:
                pa, pb = ov
                overlaps.append([(pa[0], pa[1], 0.0), (pb[0], pb[1], 0.0)])
                break        # ä¸€æ¡ s1 æ‰¾åˆ°é‡å å°±å¤Ÿäº†ï¼Œå¯è·³å‡º

    # --- æ‰“å°æ‘˜è¦ ---
    print("â˜… ä¸ pl2 é‡å  (æˆ–è¢«åŒ…å«) çš„ pl1 çº¿æ®µæ•°ï¼š", len(overlaps))
    for idx, seg in enumerate(overlaps, 1):
        print(f"  {idx}. {seg[0]}  â†’  {seg[1]}")

    return overlaps




#  ä¸»å‡½æ•°
#  (8)
# æ‰¾åˆ°å…¨éƒ¨â€œä¸¤æ ¹å¤šæ®µçº¿è€¦åˆæˆä¸€ä¸ªçŸ©å½¢â€çš„å¤šæ®µçº¿

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

"""
å‡½æ•°æ˜¯ç”¨æ¥åˆ†æä¸»æˆ¿é—´å¸¦å«ç”Ÿé—´è¿™ç§æƒ…å†µçš„ï¼Œå› æ­¤å¯¹è¾“å…¥å˜é‡æ˜¯æœ‰è¾ƒä¸¥æ ¼å‡å®šçš„ï¼Œå¹¶éé’ˆå¯¹ä»»æ„æƒ…å†µ

"""
# 1 åˆ¤æ–­çŸ©å½¢æ˜¯å¦åŒ…å«å¦ä¸€ä¸ªçŸ©å½¢

def is_rect_inside_rect(rect_outer, rect_inner, tol=1e-6):
    """
    åˆ¤å®š axisâ€‘aligned çš„çŸ©å½¢ rect_inner æ˜¯å¦è¢«ï¼ˆå«è¾¹ç•Œï¼‰å®Œå…¨åŒ…åœ¨ rect_outer å†…ã€‚

    å‚æ•°
    ----
    rect_outer : ((xmin, ymin), (xmax, ymax))
    rect_inner : ((xmin, ymin), (xmax, ymax))
        ä¸¤ä¸ªå…ƒç»„åˆ†åˆ«ç»™å‡ºçŸ©å½¢å·¦ä¸‹ã€å³ä¸Šåæ ‡ï¼ˆå‡å®š Z å…¨ 0ï¼‰ã€‚
    tol        : float
        å®¹å·®ï¼ˆå…è®¸è½»å¾®æ•°å€¼è¯¯å·®ï¼›AutoCAD double è½¬ python float æ—¶æ¨è 1eâ€‘6ï½1eâ€‘4ï¼‰ã€‚

    è¿”å›
    ----
    bool   â€”â€”  rect_inner âŠ† rect_outer ï¼Ÿ
    """
    (ox0, oy0), (ox1, oy1) = rect_outer
    (ix0, iy0), (ix1, iy1) = rect_inner

    return (
        ix0 >= ox0 - tol and
        iy0 >= oy0 - tol and
        ix1 <= ox1 + tol and
        iy1 <= oy1 + tol
    )




# 2 åˆ¤æ–­ä¸¤æ¡æ­£äº¤å¤šæ®µçº¿æ‹¼åœ¨ä¸€èµ·åæ˜¯å¦æ­£å¥½æ˜¯ä¸€ä¸ªçŸ©å½¢


def two_plines_making_rectangle(pl1, pl2, tol=0.5):#
    """
    åˆ¤æ–­ä¸¤æ¡æ­£äº¤å¤šæ®µçº¿æ‹¼åœ¨ä¸€èµ·åæ˜¯å¦æ­£å¥½æ˜¯ä¸€ä¸ªçŸ©å½¢ã€‚
    å‡è®¾ï¼šä¸¤ PLine æ²¡æœ‰é¢ç§¯é‡å ï¼Œåªå¯èƒ½å…±ç”¨å®Œæ•´è¾¹æˆ–è¾¹çš„ä¸€éƒ¨åˆ†ã€‚
    """

    import math

    def same_point(a, b, tol):
        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol

    def pline_vertices(pl):
        # ä» pl.Coordinates æå– (x,y) é¡¶ç‚¹åˆ—è¡¨ï¼Œè‡ªåŠ¨é—­åˆ
        c = pl.Coordinates
        verts = []
        for i in range(0, len(c), 2):
            verts.append((c[i], c[i+1]))
        if not same_point(verts[0], verts[-1], tol):
            verts.append(verts[0])
        return verts

    def poly_area(verts):
        # è®¡ç®—é¦–å°¾é—­åˆçš„å¤šè¾¹å½¢é¢ç§¯
        s = 0
        for i in range(len(verts)-1):
            x1,y1 = verts[i]
            x2,y2 = verts[i+1]
            s += x1*y2 - x2*y1
        return abs(s)*0.5

    def collect_segments(verts):
        # ä»é¡¶ç‚¹åˆ—è¡¨ç”Ÿæˆçº¿æ®µåˆ—è¡¨ [((x1,y1),(x2,y2)), ...]
        segs = []
        for i in range(len(verts)-1):
            segs.append((verts[i], verts[i+1]))
        return segs

    def covers_edge(edge, segs):
        """
        åˆ¤æ–­ç»™å®šçš„ bbox è¾¹ edge=((x0,y0),(x1,y1)) èƒ½å¦è¢« segs ä¸­è‹¥å¹²å…±çº¿çº¿æ®µ
        è¿ç»­è¦†ç›–ï¼ˆæ— ç¼éš™ï¼‰ã€‚
        """
        (x0,y0),(x1,y1) = edge
        # ç®—å‡ºæ–¹å‘å‘é‡å’Œé•¿åº¦
        dx, dy = x1-x0, y1-y0
        L = math.hypot(dx, dy)
        if L < tol:
            return False
        ux, uy = dx/L, dy/L  # å•ä½æ–¹å‘å‘é‡

        # æŠ•å½±æ¯æ¡ seg åˆ° [0,L] å‚æ•°åŒºé—´
        intervals = []
        for (ax,ay),(bx,by) in segs:
            # åˆ¤æ–­ç«¯ç‚¹æ˜¯å¦åœ¨åŒä¸€ç›´çº¿ä¸Š
            # cross == 0
            cross = (ax-x0)*dy - (ay-y0)*dx
            if abs(cross) > tol*L:  
                continue
            # è®¡ç®—ä¸¤ç«¯æŠ•å½±å‚æ•°
            t1 = (  (ax-x0)*ux + (ay-y0)*uy )
            t2 = (  (bx-x0)*ux + (by-y0)*uy )
            a, b = min(t1,t2), max(t1,t2)
            # åªä¿ç•™ä¸ [0,L] æœ‰äº¤é›†çš„éƒ¨åˆ†
            if b < -tol or a > L+tol:
                continue
            intervals.append((max(0.0, a), min(L, b)))

        if not intervals:
            return False

        # åˆå¹¶æ‰€æœ‰åŒºé—´
        intervals.sort(key=lambda iv: iv[0])
        cur_start, cur_end = intervals[0]
        for a,b in intervals[1:]:
            if a > cur_end + tol:
                # å‡ºç°é—´éš™
                return False
            cur_end = max(cur_end, b)
        # æœ€åæ£€æµ‹æ˜¯å¦è¦†ç›–äº†æ•´ä¸ª [0, L]
        return (cur_start <= tol) and (cur_end >= L - tol)

    # 1) æå–é¡¶ç‚¹åŠé¢ç§¯
    v1 = pline_vertices(pl1)
    v2 = pline_vertices(pl2)
    A1 = poly_area(v1)
    A2 = poly_area(v2)

    # 2) è®¡ç®—å…¬å…±å¤–åŒ…çŸ©å½¢
    xs = [p[0] for p in v1[:-1]] + [p[0] for p in v2[:-1]]
    ys = [p[1] for p in v1[:-1]] + [p[1] for p in v2[:-1]]
    xmin, xmax = min(xs), max(xs)
    ymin, ymax = min(ys), max(ys)
    A_bbox = (xmax-xmin)*(ymax-ymin)

    # 3) é¢ç§¯æ ¡éªŒ
    if abs((A1 + A2) - A_bbox) > tol:
        return False

    # 4) æ£€æŸ¥å››æ¡è¾¹è¢«è¦†ç›–
    bbox_edges = [
        ((xmin,ymin),(xmax,ymin)),
        ((xmax,ymin),(xmax,ymax)),
        ((xmax,ymax),(xmin,ymax)),
        ((xmin,ymax),(xmin,ymin))
    ]
    segs = collect_segments(v1) + collect_segments(v2)

    hits = 0
    for edge in bbox_edges:
        if covers_edge(edge, segs):
            hits += 1
    return hits == 4



#  ä¸»å‡½æ•°
#  (9)
#  åˆ¤æ–­å¤šæ®µçº¿PL2æ˜¯å¦åœ¨å¤šæ®µçº¿PL1å¤šè¾¹å½¢ä¸­

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

def are_all_vertices_inside(pl1, pl2):
    """
    åˆ¤æ–­å¤šæ®µçº¿ pl2 çš„æ‰€æœ‰é¡¶ç‚¹æ˜¯å¦éƒ½åœ¨å¤šæ®µçº¿ pl1 æ„æˆçš„å¤šè¾¹å½¢å†…éƒ¨ã€‚

    å‚æ•°ï¼š
      pl1, pl2: COM å¤šæ®µçº¿å¯¹è±¡ï¼ˆLWPOLYLINE æˆ– POLYLINEï¼‰ï¼Œè§†ä¸ºå°é—­æ­£å¤šè¾¹å½¢ã€‚
    è¿”å›ï¼š
      (all_inside, outside_pts)
      - all_inside: å¦‚æœ pl2 çš„æ¯ä¸ªé¡¶ç‚¹éƒ½åœ¨ pl1 å†…éƒ¨ï¼Œè¿”å› Trueï¼Œå¦åˆ™ Falseã€‚
      - outside_pts: åˆ—è¡¨ï¼ŒåŒ…å«æ‰€æœ‰è½åœ¨ pl1 å¤–éƒ¨çš„ pl2 é¡¶ç‚¹ (x, y, z)ã€‚
    """
    # å…ˆæŠŠ COM å¤šæ®µçº¿è½¬æˆé¡¶ç‚¹åˆ—è¡¨ [(x,y,z), ...]
    verts1 = get_unique_vertices_from_pl_com(pl1)
    verts2 = get_unique_vertices_from_pl_com(pl2)

    outside_pts = []
    for pt in verts2:
        if not point_in_polygon(pt, verts1):
            outside_pts.append(pt)

    all_inside = len(outside_pts) == 0
    if all_inside:
        print("[OK] pl2 çš„æ‰€æœ‰é¡¶ç‚¹éƒ½åœ¨ pl1 çš„å†…éƒ¨ã€‚")
    else:
        print(f"[é”™è¯¯] pl2 æœ‰ {len(outside_pts)} ä¸ªé¡¶ç‚¹ä¸åœ¨ pl1 å†…éƒ¨ï¼š")
        for p in outside_pts:
            print("   ", p)

    return all_inside, outside_pts




#&&&&%% ç¬¬å…­éƒ¨åˆ† ç»¼åˆæ§åˆ¶ç®¡ç†

#  æ¨¡å—ä½¿ç”¨è¯´æ˜

"""



ä¼°è®¡é©¬å·¥é‚£æ ·çš„æ–‡ä»¶æ‰“å¼€æ—¶ä¼šé‡åˆ°å¤§é‡çš„çª—å£ï¼Œéœ€è¦æŒ‰ç…§æ’å…¥å›¾ç­¾å‡½æ•°é‚£æ ·é‡‡ç”¨åŒçº¿ç¨‹è¿›è¡Œ7æ¬¡å–æ¶ˆå¤„ç† 20250720





æ²¡æœ‰é‚£ä¹ˆå¤æ‚ï¼Œæ ¹æœ¬ä¸Šè®²ï¼Œä½¿ç”¨ä»£ç å°†ä¸€ä¸ªä½œä¸šè¿‡ç¨‹å®Œæˆï¼Œå†å°†å…¶æµç¨‹æ ‡å‡†åŒ–æ¢³ç†ï¼Œè¿™ä¸ªè¿‡ç¨‹çš„è‡ªåŠ¨åŒ–å°±å®Œæˆäº†ã€‚ä¸éœ€è¦å¤ªå¤šè€—è´¹å¿ƒç¥ç ”ç©¶è¿™ä¸ªè¿‡ç¨‹æœ¬èº«ã€‚





é€šè¿‡è¾“å…¥è¾“å‡ºå­—å…¸ï¼Œä¸ä»…æ”¶é›†äº†ä¿¡æ¯ï¼Œç»Ÿä¸€äº†ä¿¡æ¯å¤„ç†å¹³å°ï¼Œå°†æ¥åœ¨æ„å»ºå¤§å‡½æ•°èŠ‚ç‚¹æ—¶ï¼Œä¹Ÿå¯ä»¥é€šè¿‡æŸ¥çœ‹å­—å…¸çš„ä¿¡æ¯è¾“å‡ºï¼Œåˆ¤æ–­æ¯ä¸€ä¸ªç¯èŠ‚çš„é‡è¦å‚æ•°

æ§åˆ¶ç³»ç»Ÿè¿è½¬

æŠ“ä½ä¸»è¦æ ¸å¿ƒéƒ¨åˆ†ï¼Œå¿½ç•¥æ¬¡è¦çš„è¦ç´ ï¼Œå…å¾—é™·å…¥æ— æ„ä¹‰çš„åŠ³ä½œ

å…ˆæŠŠæ•´ä½“å¤§æ¡†æ¶éª¨æ¶å»ºç«‹å¥½

å…³é”®æ¦‚å¿µï¼Œæ€æƒ³ï¼ŒèŠ‚ç‚¹ï¼Œæ¶ˆæ¯è¦æ‰“å°å‡ºæ¥åˆ†æ

åŠ å¼ºæ•°æ®ç»“æ„å’Œç®—æ³•çš„è®­ç»ƒï¼Œè¿™æ˜¯è¦ç‚¹

æˆ‘ä»¬ä¸ä»…ä»…æ˜¯è§£å†³äº†æŸä¸ªé—®é¢˜ï¼Œæ›´æ˜¯ç ”ç©¶æ¸…æ¥šäº†å…¶åŒ…å«çš„å†…åœ¨è§„å¾‹ã€‚è¿™äº›éª¨æ¶ï¼Œç»“æ„ï¼Œç›¸å…³çš„æ€æƒ³æ–¹æ³•æ¦‚å¿µå°±æ˜¯å†…åœ¨è§„å¾‹çš„è¡¨è¾¾ã€‚

è¦æ·±åˆ»ç¡®å®šæ¯ä¸ªåŠŸèƒ½å‡½æ•°çš„è®¾è®¡æ€æƒ³å’Œéª¨æ¶ç»“æ„ï¼Œä¸èƒ½å«ç³Šä¾èµ–äººå·¥æ™ºèƒ½ï¼Œå¦åˆ™å°±ä¼šé£˜å¿½ä¸ç¡®å®šï¼Œæ²¡åŠ›åº¦ã€‚æœ‰äº†æœ¬è´¨çš„æ˜ç¡®çš„è®¤çŸ¥ï¼Œéœ€è¦çš„æ—¶å€™ï¼Œæ— æƒ§é‡æ¥


ç³»ç»Ÿå°½å¯èƒ½å®Œå…¨è‡ªåŠ¨åŒ–   å°½å¯èƒ½å°‘å¹²é¢„ç³»ç»Ÿè‡ªåŠ¨åŒ– 

å…³äºæ•´ä¸ªç³»ç»Ÿå¤„ç†çš„ä¼˜åŒ–åˆ†æ

æ–‡ä»¶å¤„ç†ä¹‹å‰ï¼Œåº”è¯¥å¯¹æ‰€æœ‰å¯¹è±¡é€‰æ‹©ååˆ†ç±»æ”¾å…¥å„è‡ªçš„é¢„è®¾å›¾å±‚ï¼Œè¿™æ ·å°†æ¥å¯ä»¥å¿«é€Ÿæ‰¾åˆ°å®ƒä»¬
åŒæ ·æ–‡ä»¶æ“ä½œä¸­ï¼Œå°†ç›¸åº”å¯¹è±¡æ”¾å…¥å®ƒä»¬çš„å›¾å±‚ï¼Œä¹Ÿä¾¿äºæ‰¾åˆ°å®ƒä»¬
å›¾å±‚çš„é€‰æ‹©å¾ˆå¿«ï¼Œä¹Ÿå¯ä»¥ä½¿ç”¨Handleæ ‡ç­¾ç­‰
å¦‚æœå¯¹è±¡æœ‰å¾ˆå¼ºçš„é€»è¾‘å…³è”ï¼Œå¯ä»¥å»ºç«‹ç»„ç”¨æ¥å…³è”å®ƒä»¬

20250423

å¯¹è‡ªåŠ¨åŒ–æ“ä½œæœ‰äº†æ–°çš„æå‡ã€‚æˆ‘ä»¬æœ‰pywin32çš„APIæ§åˆ¶ï¼Œè¿˜æœ‰SendCommanå‘é€çª—å£å‘½ä»¤æ§åˆ¶ï¼Œè¿˜æœ‰pyautoguiç­‰æ‹Ÿäººæ“ä½œæ§åˆ¶ï¼Œ
è¿˜æœ‰autolispæ§åˆ¶ã€‚åŒæ—¶è¿˜è¦æ³¨æ„æ•´ä¸ªç”µè„‘ç³»ç»Ÿ çš„æ§åˆ¶è®¾ç½®ã€‚å®ƒä»¬æ˜¯ç»¼åˆè§£å†³é—®é¢˜çš„ã€‚

æŠŠCADçª—å£è®¾ç½®ä¸º1/4ï¼ŒæŠŠIDLEè®¾ç½®åˆ°å³ä¾§ï¼Œè¿™æ˜¯ä½¿ç”¨æ‹Ÿäººæ“ä½œçš„åŸºç¡€ã€‚

select_objects_in_window_area å®é™…é€‰æ‹©ä¸å¯é ä¸”æ…¢ï¼Œå¯è¡Œæ–¹æ³•æ˜¯ä½¿ç”¨é«˜äº®æ˜¾ç¤ºï¼Œå†é€šè¿‡pickæ–¹æ³•è½¬éšæ€§é€‰æ‹© pickæ–¹æ³•è½¬éšæ€§é€‰æ‹©
ä¹Ÿä¼šæ”¹å˜æ–‡ä»¶ä¹‹é—´çš„å¤åˆ¶ç²˜è´´

åŸºäºåœ¨ä¸€æ¬¡æ“ä½œè¿ç®—ä¸­è·å¾—çš„ä¿¡æ¯åº”è¯¥è¿›è¡Œå……åˆ†çš„æ•´åˆåˆ†æï¼Œå‡å°‘é‡å¤é€‰æ‹©è®¡ç®—å’Œåœ¨CADå¤æ‚ç¯å¢ƒæ“ä½œçš„é£é™©


åº”è¯¥æ€»ç»“ç»éªŒæ•™è®­ï¼Œé’ˆå¯¹ä¸€ä¸ªä»»åŠ¡çš„æ·±å…¥åˆ†æï¼Œåˆ†è§£æˆå‡½æ•°ï¼Œæ•°æ®ç»“æ„ï¼Œå„ä¸ªç»“æ„çš„å’¬åˆï¼Œç²¾å‡†åˆ°ä½ã€‚æ¯”å¦‚ç¼–ç›®å½•ï¼Œå®ƒçš„è¦ç‚¹ï¼Œæ¯ä¸ªä»»åŠ¡çš„å®ç°ï¼Œç»„åˆå®ç°ã€‚è½¬åœºï¼Œå‚ç…§èµ·ç‚¹ï¼Œé›·è¾¹ï¼Œåˆ é™¤ï¼Œé€‰æ‹©

æ’å›¾çº¸ç›®å½•æ¨¡æ¿æ—¶è·³å‡ºshxå­—ä½“å¯¹è¯æ¡† 

activate_window_by_title("ç¼ºå°‘ SHX æ–‡ä»¶")
[OK] å·²æ¿€æ´»çª—å£ï¼šâ€œç¼ºå°‘ SHX æ–‡ä»¶â€ ä½ç½®(1060,510) å¤§å°424Ã—318
(1060, 510, 424, 318)

å·²è·å–åæ ‡ï¼š(1232, 733)  ç‚¹å‡»å¿½ç•¥
(1232, 733)




activate_window_by_title("AutoCAD é”™è¯¯ä¸­æ–­", click_titlebar= True)
[OK] å·²æ¿€æ´»çª—å£ï¼šâ€œAutoCAD é”™è¯¯ä¸­æ–­â€ ä½ç½®(485,227) å¤§å°364Ã—211
(485, 227, 364, 211)

ceshubiao_weizhi()
è¯·åœ¨ 5 ç§’é’Ÿå†…ï¼Œå°†é¼ æ ‡ç²¾ç¡®åœ°æ”¾åœ¨ AutoCAD å‘½ä»¤æ çš„è¾“å…¥ä½ç½®â€¦ ç¡®å®š
å·²è·å–åæ ‡ï¼š(775, 404)
(775, 404)


activate_window_by_title("AutoCAD è­¦å‘Š", click_titlebar= True)
[OK] å·²æ¿€æ´»çª—å£ï¼šâ€œAutoCAD è­¦å‘Šâ€ ä½ç½®(392,247) å¤§å°497Ã—227
(392, 247, 497, 227)

ceshubiao_weizhi()
è¯·åœ¨ 5 ç§’é’Ÿå†…ï¼Œå°†é¼ æ ‡ç²¾ç¡®åœ°æ”¾åœ¨ AutoCAD å‘½ä»¤æ çš„è¾“å…¥ä½ç½®â€¦  å¦
å·²è·å–åæ ‡ï¼š(811, 434)
(811, 434)


[OK] å·²æ¿€æ´»çª—å£ï¼šâ€œ AutoCAD é”™è¯¯æŠ¥å‘Šâ€ ä½ç½®(1055,425) å¤§å°450Ã—549
(1055, 425, 450, 549)

ceshubiao_weizhi()
è¯·åœ¨ 5 ç§’é’Ÿå†…ï¼Œå°†é¼ æ ‡ç²¾ç¡®åœ°æ”¾åœ¨ AutoCAD å‘½ä»¤æ çš„è¾“å…¥ä½ç½®â€¦  xï¼ˆå…³é—­ï¼‰
å·²è·å–åæ ‡ï¼š(1483, 440)
(1483, 440)

activate_window_by_title("é”™è¯¯æŠ¥å‘Š - å·²å–æ¶ˆ", click_titlebar= True)
[OK] å·²æ¿€æ´»çª—å£ï¼šâ€œé”™è¯¯æŠ¥å‘Š - å·²å–æ¶ˆâ€ ä½ç½®(1127,610) å¤§å°307Ã—179
(1127, 610, 307, 179)

ceshubiao_weizhi()
è¯·åœ¨ 5 ç§’é’Ÿå†…ï¼Œå°†é¼ æ ‡ç²¾ç¡®åœ°æ”¾åœ¨ AutoCAD å‘½ä»¤æ çš„è¾“å…¥ä½ç½®â€¦ ç¡®å®š
å·²è·å–åæ ‡ï¼š(1364, 753)
(1364, 753)


å‡ºé—®é¢˜ä»¥åçš„CADé‡æ–°æ‰“å¼€

activate_window_by_title("å›¾å½¢ä¿®å¤", click_titlebar = True)
[OK] å·²æ¿€æ´»çª—å£ï¼šâ€œå›¾å½¢ä¿®å¤â€ ä½ç½®(1089,562) å¤§å°366Ã—213
(1089, 562, 366, 213)

ceshubiao_weizhi()
è¯·åœ¨ 5 ç§’é’Ÿå†…ï¼Œå°†é¼ æ ‡ç²¾ç¡®åœ°æ”¾åœ¨ AutoCAD å‘½ä»¤æ çš„è¾“å…¥ä½ç½®â€¦å…³é—­(æ±‰å­—æŒ‰é’®)
å·²è·å–åæ ‡ï¼š(1397, 724)
(1397, 724)



"""
#&&&% å¦‚æœè·å–ä¸åˆ°cadæ–‡å­—çš„å¤„ç†åŠæ³•


from win32com.client import CastTo

# â€”â€” å¯é€‰ï¼šç®€å•æ¸…æ´— MText æ§åˆ¶ç ï¼ˆä¿ç•™æ•°å­—ç”¨ï¼‰
#_MTEXT_CTRL_RE = re.compile(r"(\\[A-Za-z])|({|})|(;)|(\\\S[^;]*;)")
def _clean_mtext(s: str) -> str:
    # æŠŠ \P æ¢æˆç©ºæ ¼ï¼Œå»æ‰å¸¸è§æ§åˆ¶ç ï¼›å¦‚éœ€ä¿ç•™æ¢è¡Œï¼Œå¯æ”¹æˆ '\n'
    s = str(s).replace("\\P", " ")
    s = _MTEXT_CTRL_RE.sub("", s)
    return s.strip()

def get_text_and_ip(ent):
    """
    é€šåƒ AcDbText / AcDbMText / Attribute(Reference)ï¼š
    è¿”å› (text, (x,y,z)) æˆ– (None, None)
    """
    try:
        name = ent.ObjectName
    except Exception:
        name = ""

    # 1) å…ˆå°è¯•â€œå°±åœ°è®¿é—®â€ï¼Œå…¼å®¹ä»¥å‰ç‰ˆæœ¬ï¼ˆæœ‰äº›ç¯å¢ƒç›´æ¥èƒ½ç”¨ï¼‰
    for attr in ("TextString", "Contents"):  # TEXT / MTEXT
        try:
            txt = getattr(ent, attr)
            ip  = ent.InsertionPoint  # è¿™ä¸€æ­¥æœ‰æ—¶ä¹Ÿéœ€è¦ CastToï¼Œä¸‹é¢ä¼šå…œåº•
            if txt:
                return str(txt), (float(ip[0]), float(ip[1]), float(ip[2]) if len(ip) > 2 else 0.0)
        except Exception:
            pass

    # 2) æŒ‰ç±»å‹ CastTo åˆ°å…·ä½“æ¥å£å†å–
    try:
        if name == "AcDbText":
            obj = CastTo(ent, "IAcadText")
            return str(obj.TextString), tuple(map(float, obj.InsertionPoint))
        elif name == "AcDbMText":
            obj = CastTo(ent, "IAcadMText")
            return _clean_mtext(obj.Contents), tuple(map(float, obj.InsertionPoint))
        elif name in ("AcDbAttribute", "AcDbAttributeReference"):
            # å±æ€§æ–‡å­—ï¼ˆå—é‡Œçš„æ ‡ç­¾ï¼‰
            # ä¸¤ç§æ¥å£éƒ½è¯•ä¸€ä¸‹
            try:
                obj = CastTo(ent, "IAcadAttributeReference")
            except Exception:
                obj = CastTo(ent, "IAcadAttribute")
            return str(obj.TextString), tuple(map(float, obj.InsertionPoint))
    except Exception:
        pass

    return None, None


# â€”â€” ä½ åŸæ¥â€œæŒ‰æ’å…¥ç‚¹ y ä»ä¸Šåˆ°ä¸‹é…å¯¹ J1..Jnâ€çš„å°è£…ï¼ˆé€‚é… get_text_and_ipï¼‰
#_num_re = re.compile(r"-?\d+(?:\.\d+)?")

def _first_number(s: str) -> float:
    m = _num_re.search(str(s))
    if not m:
        raise ValueError(f"æ— æ³•è§£ææ•°å­—: {s!r}")
    return float(m.group(0))

def build_J_points_from_selected_texts(LB, n_points=61, prefix_x="30", prefix_y="37"):
    """
    è¾“å…¥ï¼šLB=select_text() è¿”å›çš„é€‰æ‹©é›†ï¼ˆå« TEXT/MTEXT/å±æ€§æ–‡å­—ç­‰ï¼‰
    è§„åˆ™ï¼š
      - æ–‡æœ¬ä»¥ '30' å¼€å¤´çš„æ˜¯ Xï¼›ä»¥ '37' å¼€å¤´çš„æ˜¯ Y
      - æŒ‰â€œæ’å…¥ç‚¹ yâ€ä»å¤§åˆ°å°ï¼ˆå³ä»ä¸Šåˆ°ä¸‹ï¼‰æ’åºï¼›åŒä¸€è¡ŒæŒ‰ x å‡åºç¨³å®š
      - ä¾åºé…ç»™ J1..Jn
    è¿”å›ï¼š
      pts_dict: {'J1': (X, Y), ...}
      pts_list: [('J1', X, Y), ...]  # å·²æŒ‰ç¼–å·æ’åº
    """
    items = []
    for ent in LB:
        txt, ip = get_text_and_ip(ent)
        if not txt or not ip:
            continue
        x0, y0 = float(ip[0]), float(ip[1])
        t = str(txt).strip().replace(" ", "")

        if t.startswith(prefix_x):
            items.append(("x", _first_number(t), y0, x0))
        elif t.startswith(prefix_y):
            items.append(("y", _first_number(t), y0, x0))

    # ä¸Š->ä¸‹ï¼šy é™åºï¼›åŒè¡Œ x å‡åº
    items.sort(key=lambda it: (-it[2], it[3]))

    xs = [v for typ, v, _, _ in items if typ == "x"]
    ys = [v for typ, v, _, _ in items if typ == "y"]

    if len(xs) != n_points or len(ys) != n_points:
        raise RuntimeError(f"X={len(xs)} / Y={len(ys)} ä¸æœŸæœ› {n_points} ä¸ç¬¦ï¼Œè¯·æ£€æŸ¥é€‰æ‹©é›†æˆ–å‰ç¼€ã€‚")

    pts_dict = {f"J{i+1}": (xs[i], ys[i]) for i in range(n_points)}
    pts_list = [(f"J{i+1}", xs[i], ys[i]) for i in range(n_points)]
    return pts_dict, pts_list

"""

æŠŠä½ é¡¹ç›®é‡Œæ‰€æœ‰ç›´æ¥ç”¨ï¼š

obj.TextString

obj.Contents

obj.InsertionPoint

çš„åœ°æ–¹ï¼Œç»Ÿä¸€æ›¿æ¢ä¸ºï¼š

txt, ip = get_text_and_ip(obj)
# txt ä¸ºå­—ç¬¦ä¸²ï¼ˆå·²åšäº†åŸºæœ¬æ¸…æ´—ï¼‰ï¼Œip ä¸º (x, y, z)


"""
import math
import traceback
from win32com.client import CastTo, gencache, Dispatch

def _flatten_coords(obj):
    """å°†å„ç§è¿”å›æ ¼å¼ï¼ˆtuple/list of tuples OR flat listï¼‰ç»Ÿä¸€ä¸ºå¹³é¢åæ ‡ [x,y,x,y,...]"""
    if obj is None:
        return None
    # å¦‚æœæ˜¯æ‰å¹³æ•°å€¼åˆ—è¡¨
    try:
        if isinstance(obj, (list, tuple)) and obj and all(isinstance(x, (int,float)) for x in obj):
            return list(obj)
    except Exception:
        pass
    # å¦‚æœæ˜¯ list/tuple of points [(x,y,z),...]
    try:
        coords = []
        for p in obj:
            if p is None:
                continue
            # p å¯ä»¥æ˜¯ (x,y,z)
            coords.extend([float(p[0]), float(p[1])])
        if coords:
            return coords
    except Exception:
        pass
    return None

def poly_length_from_coords_flat(coords):
    if not coords or len(coords) < 4:
        return 0.0
    total = 0.0
    for i in range(0, len(coords)-2, 2):
        dx = coords[i+2] - coords[i]
        dy = coords[i+3] - coords[i+1]
        total += math.hypot(dx, dy)
    return total

def try_read_coords_by_props(ent):
    """å°è¯•å„ç§å¸¸è§å±æ€§/æ–¹æ³•è¯»å–å¤šæ®µçº¿é¡¶ç‚¹æˆ–åæ ‡åˆ—è¡¨"""
    cand_props = ("Coordinates","coordinates","GetCoordinates","Points","PointsXY","Vertexes","Vertices")
    for p in cand_props:
        try:
            val = getattr(ent, p)
            coords = val() if callable(val) else val
            flat = _flatten_coords(coords)
            if flat:
                return flat
        except Exception:
            continue
    return None

def try_cast_and_read(ent):
    """å°è¯• CastTo åˆ°å¸¸è§ Polyline æ¥å£å¹¶è¯»å– .Coordinates"""
    try:
        # IAca dLWPolyline æˆ– IAcadPolylineï¼ˆåœ¨ makepy ä¸­å¸¸è§ï¼‰
        for iface in ("IAcadLWPolyline","IAcadPolyline"):
            try:
                pev = CastTo(ent, iface)
                # å¸¸è§å±æ€§
                for p in ("Coordinates","coordinates","GetCoordinates"):
                    try:
                        val = getattr(pev, p)
                        coords = val() if callable(val) else val
                        flat = _flatten_coords(coords)
                        if flat:
                            return flat
                    except Exception:
                        pass
            except Exception:
                continue
    except Exception:
        pass
    return None

def get_length_via_explode(ent, doc):
    """é€šè¿‡å¤åˆ¶+Explodeçš„æ–¹å¼æ”¶é›†çˆ†ç‚¸ååˆ†æ®µçš„é•¿åº¦ï¼ˆå¯¹çº¿æ®µ/å¼§æ®µå¤„ç†ï¼‰ï¼Œå¹¶æ¸…ç†ä¸´æ—¶å¯¹è±¡
       è¿”å› (length, had_temp_entities_flag)
    """
    created = []
    total = 0.0
    try:
        # 1) å¤åˆ¶å®ä½“ï¼ˆCopy ä¼šè¿”å›æ–°çš„ COM å¯¹è±¡ï¼‰
        try:
            ent_copy = ent.Copy()
            created.append(ent_copy)
        except Exception:
            # æœ‰äº› wrapper Copy æ–¹æ³•éœ€è¦æ–‡æ¡£æ“ä½œï¼Œå°è¯• alternativeï¼šAdd copy via clone? è‹¥å¤±è´¥ç›´æ¥ explode åŸå¯¹è±¡ï¼ˆå±é™©ï¼‰
            ent_copy = ent  # é€€è€Œæ±‚å…¶æ¬¡ï¼ˆä½†æˆ‘ä»¬ä¸æƒ³ç›´æ¥ explode åŸä»¶ï¼‰
        # 2) Explodeï¼ˆå¯èƒ½è¿”å›ä¸€ä¸ªé›†åˆæˆ–äº§ç”Ÿæ–°çš„å®ä½“åœ¨ ModelSpaceï¼‰
        try:
            res = ent_copy.Explode()
        except Exception as e:
            # å¦‚æœ Explode å¤±è´¥ï¼Œå°è¯•ç›´æ¥è¯»å– Length å±æ€§ï¼ˆæœ€åçš„å…œåº•ï¼‰
            try:
                return float(ent.Length), False
            except Exception:
                raise

        # res å¯èƒ½æ˜¯ä¸€ä¸ª sequenceï¼ˆç”Ÿæˆçš„å®ä½“é›†åˆï¼‰æˆ– Noneï¼ˆæœ‰äº›æ¥å£ç›´æ¥åœ¨ ModelSpace æ’å…¥ï¼‰
        segs = []
        try:
            # å¦‚æœ res æœ¬èº«æ˜¯ä¸€ä¸ª Python iterable
            for item in (res or []):
                segs.append(item)
        except Exception:
            # å½“ res ä¸º None æ—¶ï¼Œå°è¯•æ‰«ææœ€è¿‘åˆ›å»ºçš„å®ä½“ï¼ˆä¸èƒ½å¯é ï¼‰
            segs = []

        # å¦‚æœæ²¡æœ‰é€šè¿‡è¿”å›çš„ res è·å–åˆ°åˆ†æ®µï¼Œå°è¯•åœ¨ ModelSpace ä¸­æŸ¥æ‰¾ç”± Copy äº§ç”Ÿçš„åç»­å¯¹è±¡
        # ä¸ºç®€å•ç¨³å¦¥ï¼Œæˆ‘ä»¬ä¹ŸæŠŠ ent_copy åŠ å…¥ createdï¼ˆè‹¥ ent_copy != entï¼‰
        for s in segs:
            # å¯¹æ¯ä¸ªåˆ†æ®µï¼Œä¼˜å…ˆè¯»å– Length å±æ€§ï¼ˆåœ†å¼§ä¹Ÿå¯èƒ½æœ‰ Lengthï¼‰
            try:
                L = getattr(s, "Length")
                if L is not None:
                    total += float(L)
                    created.append(s)
                    continue
            except Exception:
                pass
            # å¯¹ç›´çº¿æ®µï¼Œä½¿ç”¨ StartPoint/EndPoint
            try:
                sp = getattr(s, "StartPoint", None)
                ep = getattr(s, "EndPoint", None)
                if sp and ep:
                    dx = float(ep[0]) - float(sp[0])
                    dy = float(ep[1]) - float(sp[1])
                    total += math.hypot(dx, dy)
                    created.append(s)
                    continue
            except Exception:
                pass
            # å¯¹å¼§æ®µï¼Œå°è¯•ç”¨ Radius å’Œ StartAngle/EndAngle è®¡ç®—
            try:
                r = getattr(s, "Radius", None)
                sa = getattr(s, "StartAngle", None)
                ea = getattr(s, "EndAngle", None)
                if r is not None and sa is not None and ea is not None:
                    # è§’åº¦å¯èƒ½ä»¥å¼§åº¦ç»™å‡º
                    ang = abs(float(ea) - float(sa))
                    total += abs(float(r)) * ang
                    created.append(s)
                    continue
            except Exception:
                pass
            # è‹¥æ— æ³•è¯†åˆ«ï¼Œå¿½ç•¥è¯¥æ®µ
        # 3) æ¸…ç†ï¼šæ“¦é™¤ä¸´æ—¶åˆ›å»ºçš„å®ä½“ï¼ˆæ³¨æ„ï¼šå¦‚æœ ent_copy is original ent, ä¸è¦æ“¦é™¤åŸå®ä½“ï¼‰
        for obj in created:
            try:
                # ä¸æ“¦é™¤åŸå¯¹è±¡
                if obj is ent:
                    continue
                obj.Erase()
            except Exception:
                pass
        # å¦‚æœ ent_copy was the original ent (we didn't want that), we didn't erase it above
        return total, (len(created) > 0)
    except Exception as e:
        # å°è¯•å…œåº•è¯» Length
        try:
            return float(ent.Length), False
        except Exception:
            # æœ€ç»ˆå¤±è´¥
            # æ‰“å° traceback ä»¥ä¾¿è°ƒè¯•
            traceback.print_exc()
            return None, False

def get_polyline_coords_and_length(ent, doc=None):
    """
    ä¸»è¦å¯¹å¤–å‡½æ•°ï¼šå…ˆå°è¯•ç›´æ¥å–é¡¶ç‚¹åæ ‡ï¼ˆè¿”å› flat coords listï¼‰ï¼Œå¹¶è®¡ç®—é•¿åº¦ï¼›
    å¦‚æœæ— æ³•å¾—åˆ°é¡¶ç‚¹ï¼Œåˆ™ç”¨ explode å›é€€è®¡ç®—é•¿åº¦ã€‚
    è¿”å› (coords_flat_or_None, length_or_None, used_explode_flag)
    """
    # 1) ç›´æ¥æŒ‰å¸¸è§å±æ€§åå°è¯•
    coords = try_read_coords_by_props(ent)
    if coords:
        length = poly_length_from_coords_flat(coords)
        return coords, length, False

    # 2) å°è¯• CastTo å¸¸è§æ¥å£
    coords = try_cast_and_read(ent)
    if coords:
        length = poly_length_from_coords_flat(coords)
        return coords, length, False

    # 3) Explode å›é€€ï¼ˆéœ€è¦æä¾› doc ç”¨äºæ¸…ç†ï¼‰
    if doc is None:
        # å°è¯•è·å– docï¼ˆå®ä½“æœ‰ Document å±æ€§ï¼‰
        try:
            doc = ent.Document
        except Exception:
            doc = None
    length, used_temp = get_length_via_explode(ent, doc)
    if length is not None:
        return None, length, used_temp

    # 4) æœ€åå…œåº•ï¼šå°è¯•è¯»å– Length å±æ€§ç›´æ¥è¿”å›
    try:
        L = getattr(ent, "Length")
        return None, float(L), False
    except Exception:
        return None, None, False



import win32com.client as win32

def ensure_acad():
    # æ—©ç»‘å®šï¼šç”Ÿæˆå¹¶åŠ è½½ç±»å‹åº“ï¼ˆæ¯” Dispatch ç¨³å®šï¼Œå±æ€§/æ–¹æ³•æ›´é½å…¨ï¼‰
    acad = win32.gencache.EnsureDispatch("AutoCAD.Application")
    return acad

def cast_polyline(ent):
    """æŠŠé€‰åˆ°çš„å¤šæ®µçº¿å®ä½“è½¬æˆæ›´å…·ä½“çš„æ¥å£ï¼Œæ–¹ä¾¿å–ä¸“æœ‰å±æ€§ã€‚"""
    name = ent.ObjectName  # å¦‚ 'AcDbPolyline' / 'AcDb2dPolyline' / 'AcDb3dPolyline'
    if name == "AcDbPolyline":     # è½»é‡ 2D å¤šæ®µçº¿
        try:
            return win32.CastTo(ent, "IAcadLWPolyline")
        except Exception:
            return ent  # é€€åŒ–ä¸ºåŠ¨æ€è°ƒåº¦
    elif name == "AcDb2dPolyline":
        try:
            return win32.CastTo(ent, "IAcadPolyline")  # æ—§ 2D å¤šæ®µçº¿
        except Exception:
            return ent
    elif name == "AcDb3dPolyline":
        try:
            return win32.CastTo(ent, "IAcad3DPolyline")
        except Exception:
            return ent
    else:
        return ent

def get_ent_truecolor_rgb(ent):
    """
    è¿”å›å®ä½“å®é™…æ˜¾ç¤ºé¢œè‰²(å°½åŠ›)ï¼š(r,g,b)
    å¦‚æœæ˜¯ ByLayer / ByBlockï¼Œä¼šå›é€€å»æŸ¥å›¾å±‚æˆ–ç»™å‡º Noneã€‚
    """
    try:
        tc = ent.TrueColor  # AcadAcCmColor
        # ColorMethodï¼š0=ByBlock,1=ByLayer,2=ByColor,3=ByACI,4=ByRGB,5=ByPen
        cm = getattr(tc, "ColorMethod", None)
        if cm in (4,):  # ByRGB
            return (tc.Red, tc.Green, tc.Blue)
        if cm in (2, 3):  # ç›´æ¥é¢œè‰²/ACI
            # ç›´æ¥ç»™å‡º ACI å¯¹åº”çš„ RGB å¯èƒ½éœ€è¦æ˜ å°„ï¼›ç®€å•è¿”å› None
            return None
        if cm == 1:  # ByLayer
            try:
                layer = ent.Document.Layers.Item(ent.Layer)
                ltc = layer.TrueColor
                if getattr(ltc, "ColorMethod", None) == 4:
                    return (ltc.Red, ltc.Green, ltc.Blue)
            except Exception:
                pass
        # å…¶å®ƒæƒ…å†µå›é€€
        return None
    except Exception:
        return None

def polyline_vertices(pl):
    """
    è¿”å›å¤šæ®µçº¿é¡¶ç‚¹åˆ—è¡¨ï¼š
    - LWPolyline: ç”¨ Coordinatesï¼ˆæ‰å¹³æ•°ç»„ x,y[,x,y...]ï¼‰
    - Polyline/3DPolyline: ç”¨ Coordinate(i)
    """
    name = pl.ObjectName
    pts = []
    try:
        if name == "AcDbPolyline":  # LW
            coords = list(pl.Coordinates)  # [x1,y1, x2,y2, ...]
            # LWPolyline å›ºå®šåœ¨å¹³é¢ï¼ŒZ ä¸€èˆ¬ç­‰äº Elevation
            z = getattr(pl, "Elevation", 0.0)
            for i in range(0, len(coords), 2):
                pts.append((coords[i], coords[i+1], z))
        elif name in ("AcDb2dPolyline", "AcDb3dPolyline"):
            n = pl.NumberOfVertices
            for i in range(n):
                p = pl.Coordinate(i)  # (x, y, z)
                pts.append((p[0], p[1], p[2]))
        else:
            # å…œåº•å°è¯• Coordinates
            coords = list(pl.Coordinates)
            z = getattr(pl, "Elevation", 0.0)
            for i in range(0, len(coords), 2):
                pts.append((coords[i], coords[i+1], z))
    except Exception:
        pass
    return pts

def inspect_polyline(ent):
    """æ‰“å°å¤šæ®µçº¿çš„å…³é”®ä¿¡æ¯ï¼šç±»å‹ã€å›¾å±‚ã€é¢œè‰²ã€æ˜¯å¦é—­åˆã€é•¿åº¦ã€èµ·ç‚¹ã€é¡¶ç‚¹åˆ—è¡¨ç­‰ã€‚"""
    pl = cast_polyline(ent)
    info = {}
    info["ObjectName"] = ent.ObjectName
    info["Layer"] = ent.Layer
    # Color: 256 è¡¨ç¤º ByLayerï¼›æ›´å¯é çš„æ˜¯ TrueColor
    try:
        info["ColorIndex_or_Int"] = ent.color
    except Exception:
        info["ColorIndex_or_Int"] = None
    info["TrueColorRGB"] = get_ent_truecolor_rgb(ent)

    # æ˜¯å¦é—­åˆ
    closed = None
    for key in ("Closed", "Closed2d", "Closed3d"):
        if hasattr(pl, key):
            try:
                closed = bool(getattr(pl, key))
                break
            except Exception:
                pass
    info["Closed"] = closed

    # é•¿åº¦/é¢ç§¯
    for key in ("Length", "Area"):
        try:
            info[key] = float(getattr(pl, key))
        except Exception:
            info[key] = None

    # é¡¶ç‚¹ä¸èµ·ç‚¹
    verts = polyline_vertices(pl)
    info["VertexCount"] = len(verts)
    info["StartPoint"] = verts[0] if verts else None
    info["Vertices"] = verts

    return info
















#&&&% é‡è¦åŸºç¡€çŸ¥è¯†ï¼ˆpythonã€ç®—æ³•ã€æ•°æ®ç»“æ„ï¼‰

#&&% ***ğŸ§   æŠ›å‡ºå¼‚å¸¸çš„å¤„ç†é€»è¾‘

"""
1 try+except+finaly

æˆåŠŸåæ‰§è¡Œfinalyåè¯­å¥ï¼Œå¤±è´¥åä¹Ÿä¼šæ‰§è¡Œfinalyåè¯­å¥ï¼Œå¿…å®šé¢„æœŸæ•´ä¸ªä»£ç æ®µè¦åšçš„å°±æ˜¯finalyæä¾›çš„åŠŸèƒ½


2 try+except+else

æˆåŠŸåæ‰§è¡Œelseåè¯­å¥ï¼Œå¤±è´¥åä¸ä¼šä¼šæ‰§è¡Œelseåè¯­å¥ï¼Œå¿…å®šé¢„æœŸæˆåŠŸæƒ…å†µä¸‹ è¦åšçš„å°±æ˜¯elseæä¾›çš„åŠŸèƒ½

3
try:
    risky_operation()
except SomeError as e:
    handle_error(e)
    # è¿™æ®µåªåœ¨ try å¤±è´¥æ—¶è¿è¡Œ
    try:
        do_special_on_failure()
    except Exception as e2:
        log("failure-only block failed:", e2)
else:
    # æˆåŠŸæ—¶èµ°è¿™é‡Œ
    do_if_success()
finally:
    cleanup()

do_special_on_failure() å°±æ˜¯ä¸“é—¨ä¸ºä»£ç æ®µå¤±è´¥æ—¶é¢„æœŸå¿…å®šè¦æ‰§è¡Œçš„è¯­å¥ã€‚å®ƒä¸åŒäºç›´æ¥æ”¾åœ¨exceptä¸­ï¼Œå› ä¸ºexceptå¯èƒ½çš„å¤±è´¥æ“ä½œä¸èƒ½ä¿è¯å®ƒè¢«é¢„æœŸå¿…ç„¶æ‰§è¡Œã€‚

æˆ‘ä»¬éœ€è¦åœ¨æŸä¸ªæ“ä½œå¤±è´¥æ—¶ï¼Œå¿…é¡»æ‰§è¡ŒæŸä¸ªæ“ä½œï¼Œå°±ä½¿ç”¨è¿™ä¸ªé€»è¾‘



"""




#&&&% è„šæœ¬æ³¨é‡Šæ ‡å‡†æ¨¡å—
# â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- -- -- -- -- â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- --

#ğŸ“Œ èŠ‚ç‚¹ 1ï¼šå½¢æˆæ­£å¼â€œæ‰“å°æ¡†çº¿â€å¹¶é‡æ’å›¾å½¢
#â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

"""ğŸ” è¾“å…¥:"""

"""ğŸ”§ å…³é”®å‡½æ•°ï¼š"""
"""   > F1: `find_standard_printframes()` â€“ è¿”å›æ ‡å‡†å›¾æ¡†åˆ—è¡¨"""

"""ğŸ§  å¤„ç†é€»è¾‘:"""

"""ğŸ“¤ è¾“å‡º:"""


 

"""
ğŸ„  ğŸ„‘  ğŸ„’  ğŸ„“  ğŸ„”  ğŸ„•  ğŸ„–  ğŸ„—  ğŸ„˜  ğŸ„™  ğŸ„š  ğŸ„›  ğŸ„œ  ğŸ„  ğŸ„  ğŸ„Ÿ  ğŸ„   ğŸ„¡  ğŸ„¢  ğŸ„£  ğŸ„¤  ğŸ……  ğŸ…†  ğŸ…‡   ğŸ…ˆ   ğŸ…‰







â“¿ â‘  â‘¡ â‘¢ â‘£ â‘¤ â‘¥ â‘¦ â‘§ â‘¨ â‘© â‘ª â‘« ï¸âƒ£  

# â–â–â–â€”â€” é‡è¦å‡½æ•°ï¼šfoo() â€”â€”â–â–â–

# âœºâœºâœºâ€”â€” å…³é”®é€»è¾‘ï¼šprocess_data()â€”â€”âœºâœºâœº

# âš¡âš¡âš¡â€”â€” é«˜ä¼˜å…ˆï¼šinit_config() â€”â€”âš¡âš¡âš¡

# â˜…â˜…â˜…â€”â€” æ ¸å¿ƒå…¥å£ï¼šmain() â€”â€”â˜…â˜…â˜…
å¸¸ç”¨ç¬¦å·æ¨èï¼š

â– ï¼ˆå››è§’è±å½¢ï¼‰

âœº ï¼ˆæ”¾å°„çŠ¶èŠ±ç“£ï¼‰

âš¡ â˜… ï¼ˆé—ªç”µï¼‰

â˜… ï¼ˆå®å¿ƒæ˜Ÿï¼‰

âœ¦ ï¼ˆä¸­å®è±å½¢ï¼‰

â— ï¼ˆç²—æ„Ÿå¹å·ï¼‰

 
+--------------------------+
| æ–¹æ³•1ï¼šä»æ‰“å°å›¾ç­¾å—æˆ–æ‰“å°çº¿è·å–DX |
+--------------------------+
      â†“ æˆåŠŸ          â†“ å¤±è´¥
+------------------+   +--------------------------+
| ä»å›¾ç­¾å—é‡å»ºæ¡†çº¿ |   | æ–¹æ³•2ï¼šä»"å…¬å¸å›¾ç­¾"å›¾å—æå–DX |
+------------------+   +--------------------------+
           â†“                     â†“ æˆåŠŸ    â†“ å¤±è´¥


+--------------------------------+
| åˆå§‹çŠ¶æ€ï¼šæœªå¼€å§‹æ‰“å°           |
+--------------------------------+
               â†“


        +---------+
        | F3å‡½æ•°  | â† è¯†åˆ«æ‹Ÿåˆæ¡†çº¿å¯¹åº”å›¾å½¢
        +---------+
             â†“
 
ğŸ“Œ èŠ‚ç‚¹ 1ï¼šç¡®å®šâ€œå‡†æ‰“å°æ¡†çº¿â€
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
ğŸ” åˆå§‹åŒ–é˜¶æ®µï¼š
    > æ¸…ç©º "å‡†æ‰“å°æ¡†çº¿" å›¾å±‚
    > è‹¥å·²å­˜åœ¨æ—§å¯¹è±¡ï¼Œåˆ™å…¨éƒ¨åˆ é™¤

ğŸ” æå–æ‰“å°æ¡†å€™é€‰é›†åˆ DX çš„ä¸‰ç§æ–¹æ³•ï¼š

â‘  æ–¹æ³•ä¸€ï¼ˆæ¨èæ–¹å¼ï¼‰
    > æŸ¥æ‰¾ â€œå›¾ç­¾å—â€ + â€œæ‰“å°æ¡†çº¿â€ å—
    > è‹¥å­˜åœ¨å›¾ç­¾å—ï¼š
        > æå–åŒ…å›´ç›’ â†’ é‡ç»˜ä¸ºçŸ©å½¢ â†’ å­˜å…¥â€œå‡†æ‰“å°æ¡†çº¿â€
        > æ¸…é™¤åŸâ€œæ‰“å°æ¡†çº¿â€å¯¹è±¡
        > [OK] å®Œæˆæœ¬èŠ‚ç‚¹ç›®æ ‡


ğŸ“¦ è¾“å‡ºï¼šDX å¯¹è±¡é›†åˆ + â€œå‡†æ‰“å°æ¡†çº¿â€å›¾å±‚ç»˜åˆ¶

------------------------------------------------------

ğŸ“Œ èŠ‚ç‚¹ 2ï¼šå½¢æˆæ­£å¼â€œæ‰“å°æ¡†çº¿â€å¹¶é‡æ’å›¾å½¢
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
ğŸ” è¾“å…¥ï¼šDX é›†åˆï¼ˆæ¥è‡ªä¸ŠèŠ‚ç‚¹ï¼‰

ğŸ§  å›¾å½¢åˆ†æä¸å¤„ç†ï¼š
    > è¯†åˆ«æ ‡å‡†æ‰“å°æ¡†ï¼ˆF1ï¼‰
    > è¯†åˆ«æ‹Ÿåˆæ‰“å°æ¡†ï¼ˆF2ï¼‰

ğŸ§± å›¾æ¡†æ›¿æ¢ä¸æ’ç‰ˆï¼š
    > æ‹Ÿåˆæ¡† â†’ æ›¿æ¢ä¸ºæœ€è¿‘ A1 æ ‡å‡†æ¯”ä¾‹å›¾æ¡†
    > é‡æ–°æ’åˆ—æ‰“å°å›¾æ¡†ï¼Œé¿å…é‡å 

ğŸ“¦ è¾“å‡ºï¼š
    > æ­£å¼â€œæ‰“å°æ¡†çº¿â€å†™å…¥ç›®æ ‡å›¾å±‚
    > å‡†æ‰“å°æ¡†å›¾å±‚æ¸…ç©º
    > Handle â†’ ä¿¡æ¯å­˜å‚¨ç”¨äºåç»­åŒ¹é…

ğŸ”§ å‡½æ•°ï¼š
    > F1: æ ‡å‡†æ¡†è¯†åˆ«
    > F2: æ‹Ÿåˆæ¡†è¯†åˆ«
    > F3: å›¾å½¢è¯†åˆ«å½’å±ï¼ˆå±€éƒ¨ï¼‰
    > F4: å›¾å½¢é‡æ’æ•´ä½“é€»è¾‘

â”Œâ”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”> F1: æ ‡å‡†æ¡†è¯†åˆ« 
â”‚ èŠ‚ç‚¹ 1ï¼š   â”‚
â”‚ å‡†æ‰“å°æ¡†çº¿ â”‚
â””â”€â”€â”€â”€â”¬â”€â”€â”€â”€â”€â”€â”€â”˜
     â”‚ æ¸…ç©ºå›¾å±‚
     â–¼
â”Œâ”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”
â”‚ æ–¹æ³•ä¸€ï¼šå›¾ç­¾å—æˆ–æ¡†çº¿ â”‚â—„â”€â”€â”€â”€â”€â”
â””â”€â”€â”€â”€â”¬â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜      â”‚
     â”‚ æ‰¾ä¸åˆ°             â”‚
     â–¼                    â”‚

â”Œâ”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”
â”‚ æ’åº + ç§»åŠ¨å›¾å½¢ F3/F4 â”‚
â””â”€â”€â”€â”€â”¬â”€â”€â”€â”€â”€â”€â”€â”€â”˜
     â–¼
â”Œâ”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”
â”‚ å†™å…¥â€œæ‰“å°æ¡†çº¿â€å›¾å±‚ â”‚
â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜

ğŸ“Œ èŠ‚ç‚¹ Xï¼š<ç®€è¦æ ‡é¢˜>
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
ğŸ” è¾“å…¥ï¼š
    > è¾“å…¥æ•°æ®æˆ–å›¾å…ƒè¯´æ˜ï¼ˆå¦‚ DX é›†åˆï¼‰
    > æ¥æºè¯´æ˜ï¼ˆå¦‚æ¥è‡ªä¸Šæ¸¸èŠ‚ç‚¹ï¼‰
    
ğŸ§  å¤„ç†é€»è¾‘ï¼š
    > æ­¥éª¤è¯´æ˜ 1ï¼ˆå¦‚ï¼šè¯†åˆ«æ ‡å‡†æ‰“å°æ¡†ï¼‰
    > æ­¥éª¤è¯´æ˜ 2ï¼ˆå¦‚ï¼šåˆ¤æ–­æ‹Ÿåˆæ¡†ï¼‰

ğŸ”§ å…³é”®å‡½æ•°ï¼š
    > F1: å‡½æ•°å â€“ åŠŸèƒ½è¯´æ˜
    > F2: å‡½æ•°å â€“ åŠŸèƒ½è¯´æ˜

ğŸ“¤ è¾“å‡ºï¼š
    > è¾“å‡ºæ•°æ®ç›®æ ‡ï¼ˆå¦‚å†™å…¥æ­£å¼å›¾å±‚ï¼‰
    > å‰¯ä½œç”¨ï¼ˆå¦‚æ¸…é™¤ä¸­é—´å±‚ã€æ›´æ–° handle è®°å½•ï¼‰

[è­¦å‘Š]ï¸ å¼‚å¸¸å¤„ç†ï¼š
    > æœªåŒ¹é…å›¾æ¡† â†’ è¾“å‡ºè­¦å‘Š
    > å¤šå›¾é‡å  â†’ æ’åºå¹¶ä¿®æ­£é‡æ’

ğŸ—‚ ç¤ºä¾‹è°ƒç”¨ï¼š
    process_printframes(DX)


ğŸ“Œ	èŠ‚ç‚¹èµ·å§‹	ğŸ“Œ èŠ‚ç‚¹ 2ï¼šå½¢æˆæ­£å¼â€œæ‰“å°æ¡†çº¿â€å¹¶é‡æ’å›¾å½¢
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€	åˆ†å‰²çº¿	â”€â”€â”€â”€â”€â”€â”€â”€ èŠ‚ç‚¹åˆ†éš” â”€â”€â”€â”€â”€â”€â”€â”€
ğŸ§­	é€»è¾‘å¯¼å‘	ğŸ§­ ä¸‹ä¸€æ­¥è¯†åˆ«æ‹Ÿåˆå›¾æ¡†
ğŸ”€	è·¯å¾„åˆ†æ”¯	ğŸ”€ åˆ†æµå¤„ç†ï¼šæ ‡å‡†æ¡† vs æ‹Ÿåˆæ¡†
â¤µ / â¤´	å­æµç¨‹è·³è½¬	â¤µ è¿›å…¥å­æµç¨‹ F3
ğŸ”‚	å¾ªç¯å¤„ç†	ğŸ”‚ é’ˆå¯¹æ‰€æœ‰ F2 æ¡†æ‰§è¡Œæ›¿æ¢é€»è¾‘

ğŸ”§ åŠŸèƒ½æˆ–æ¨¡å—è¯´æ˜
ç¬¦å·	ç”¨é€”	ç¤ºä¾‹
ğŸ”§	å‡½æ•°å®šä¹‰	ğŸ”§ F1: æ ‡å‡†æ¡†è¯†åˆ«å‡½æ•°
ğŸ§ 	å¤„ç†é€»è¾‘	ğŸ§  å›¾å½¢åˆ†æä¸å¤„ç†
ğŸ§ª	æµ‹è¯•/è°ƒè¯•å—	ğŸ§ª ä¸´æ—¶æ‰“å° bbox ä¿¡æ¯
ğŸ“¦	è¾“å‡ºç»“æœ	ğŸ“¦ è¾“å‡ºå›¾å±‚å¯¹è±¡
ğŸ“¥ / ğŸ“¤	è¾“å…¥/è¾“å‡º	ğŸ“¤ è¾“å‡ºå†™å…¥ LAYER_PRINT æ­£å¼å›¾æ¡†
ğŸ§±	æ•°æ®ç»“æ„/å®ä½“å¯¹è±¡	ğŸ§± å®ä½“å¯¹è±¡æ›¿æ¢æ“ä½œ

ğŸ” åˆ†æä¸è¯†åˆ«ç›¸å…³
ç¬¦å·	ç”¨é€”	ç¤ºä¾‹
ğŸ”	æœç´¢/è¯†åˆ«	ğŸ” æŸ¥æ‰¾ç¬¦åˆ A1 æ¯”ä¾‹çš„å›¾æ¡†
ğŸ§¬	ç‰¹å¾æå–	ğŸ§¬ æå–å›¾å½¢å¤–åŒ…ç›’åæ ‡ä¸å®½é«˜æ¯”
ğŸ”	ç²¾ç»†è¯†åˆ«	ğŸ” åˆ¤æ–­æ˜¯å¦ä¸ºå®Œæ•´å°é—­çŸ©å½¢

[OK] çŠ¶æ€åé¦ˆ
ç¬¦å·	ç”¨é€”	ç¤ºä¾‹
[OK]	æˆåŠŸ	[OK] å›¾æ¡†æ›¿æ¢å®Œæˆ
[è­¦å‘Š]ï¸	è­¦å‘Š	[è­¦å‘Š]ï¸ æ‰¾ä¸åˆ°åˆé€‚å›¾æ¡†åŒ¹é…é¡¹
[é”™è¯¯]	é”™è¯¯	[é”™è¯¯] æ— æ³•è¯†åˆ«å¤šæ®µçº¿è¾¹ç•Œ
â³	ç­‰å¾…	â³ ç­‰å¾… CAD å¯¹è±¡å“åº”

ğŸ” æ§åˆ¶ç»“æ„ï¼ˆå¯åµŒå…¥æ³¨é‡Šä¸­ï¼‰
ç¬¦å·	ç”¨é€”	ç¤ºä¾‹
â¬…ï¸ â¡ï¸ â¬†ï¸ â¬‡ï¸	æ–¹ä½ã€ç§»åŠ¨æ–¹å‘	â¬†ï¸ ä¸Šç§»å›¾æ¡† 1000 å•ä½é¿å…é‡å 
ğŸ”„	é‡å¤æ‰§è¡Œ	ğŸ”„ å¯¹æ‰€æœ‰å›¾æ¡†æ‰§è¡Œä¸€æ¬¡æ—‹è½¬æ£€æµ‹
â†ªï¸	è¿”å›	â†ªï¸ è¿”å›åŸå›¾å±‚å¤„ç†çŠ¶æ€

ğŸ“Š æ•°æ®æ“ä½œä¸è®¡ç®—
ç¬¦å·	ç”¨é€”	ç¤ºä¾‹
ğŸ“	å‡ ä½•è®¡ç®—	ğŸ“ è®¡ç®—å›¾æ¡†å®½é«˜æ¯”
ğŸ“	æµ‹é‡	ğŸ“ æµ‹é‡è·ç¦»ä¸åç§»
ğŸ§®	æ•°å€¼å¤„ç†	ğŸ§® æ€»è®¡å›¾æ¡†æ•°é‡ = len(F1) + len(F2)



"""


#&&%  *** å¸¸ç”¨åŠŸèƒ½ä»£ç å—

## 1 å¦‚æœæŸä¸ªæ“ä½œæµç¨‹å‡ºé”™ï¼Œæœ€å¤šé—´éš”1ç§’å°è¯•3æ¬¡å³æ— è§†ä¹‹ï¼Œè¿›å…¥ä¸‹ä¸€æµç¨‹


##try:
##    feng_lines = stc("bianmulu_lp")
##    print(f"ğŸ—‘ å°† {len(split_pairs)} æ¡åˆ†è£‚çº¿æ ‡è®°ä¸ºçº¢è‰² (å°è¯• {attempt})")
##    for A, B in split_pairs:
##        _, yA, _ = bbox_center_3(A)
##        _, yB, _ = bbox_center_3(B)
##        y_mid = (yA + yB) / 2.0
##
##        for ent in feng_lines:
##            try:
##                obj = dyn.Dispatch(ent._oleobj_)
##            except:
##                continue
##            if obj.ObjectName not in ("AcDbLine", "AcDbPolyline", "AcDb2dPolyline"):
##                continue
##            _, cy, _ = bbox_center_3(obj)
##            if abs(cy - y_mid) <= y_tol:
##                print(f"  ğŸ”´ æ ‡è®° Handle={obj.Handle} centerY={cy:.1f}")
##                try:
##                    obj.Color = 1
##                    obj.Layer = "æµ‹è¯•è¾…åŠ©"
##                except:
##                    try:
##                        obj.TrueColor = 0xFF0000
##                    except:
##                        print(f"    [è­¦å‘Š]ï¸ ä¸Šè‰²å¤±è´¥ Handle={obj.Handle}")
##    # å¦‚æœæ‰§è¡Œåˆ°è¿™é‡Œä¸å‡ºå¼‚å¸¸ï¼Œå°±é€€å‡ºé‡è¯•å¾ªç¯
##    break
##except Exception as e:
##    print(f"[è­¦å‘Š] åˆ†è£‚çº¿æ ‡è®°ç¬¬ {attempt} æ¬¡å¤±è´¥: {e}")
##    if attempt < 3:
##        time.sleep(1)
##    else:
##        print("â„¹ åˆ†è£‚çº¿æ ‡è®°è·³è¿‡ï¼Œè¿›å…¥ä¸‹ä¸€æµç¨‹")


def kill_wps(verbose: bool = False):
    """
    ç»“æŸæ‰€æœ‰ WPS/é‡‘å±±åŠå…¬ç›¸å…³è¿›ç¨‹ï¼Œç‰¹åˆ«æ˜¯ wpspdf.exeã€‚
    verbose=True æ—¶æ‰“å°è¢«ç»ˆæ­¢çš„æ˜ åƒåã€‚
    """
    # éœ€è¦å…³æ‰çš„æ˜ åƒæ¸…å•ï¼ˆå…¨éƒ¨å°å†™ï¼‰
    targets = {
        "wps.exe",          # WPS ä¸»è¿›ç¨‹
        "wpspdf.exe",       # â˜… PDF é¢„è§ˆå™¨
        "wpp.exe",          # æ¼”ç¤º
        "et.exe",           # è¡¨æ ¼
        "ksolaunch.exe",    # å¯åŠ¨å™¨
        "wpscloudsvr.exe",  # äº‘åŒæ­¥
        "wpsupdate.exe",    # æ›´æ–°å™¨
        "wpscenter.exe",    # æ¶ˆæ¯ä¸­å¿ƒ
    }

    killed = set()
    for name in targets:
        # /T è¿å­è¿›ç¨‹ä¸€èµ·ï¼›/F å¼ºåˆ¶ï¼›/IM æŒ‰æ˜ åƒå
        rc = subprocess.call(
            f'taskkill /T /F /IM {name}',
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        if rc == 0:      # è¿”å› 0 è¡¨ç¤ºç¡®æœ‰æ­¤è¿›ç¨‹å¹¶å·²ç»“æŸ
            killed.add(name)

    if verbose:
        if killed:
            print("[OK] å·²ç»“æŸè¿›ç¨‹:", ", ".join(sorted(killed)))
        else:
            print("â„¹ï¸ æœªæ£€æµ‹åˆ° WPS ç›¸å…³è¿›ç¨‹")


#&&% äººæœºç»“åˆçš„å‡½æ•°


"""

é€šè¿‡inputè¾“å…¥ç­‰å¾…ï¼Œåœ¨å¤æ‚ç¯èŠ‚å¼•å…¥äººå·¥æ“ä½œï¼Œè€Œå…¶ä½™éƒ¨åˆ†ä»äº¤ç»™è®¡ç®—æœºè‡ªåŠ¨å®Œæˆ
å› ä¸ºæœ€ç»ˆçš„ç›®çš„æ˜¯èŠ‚çœæ—¶é—´
ä»æ™®é€šå›¾ç­¾å—åˆ¶ä½œæ ‡å‡†å±æ€§å—å°±æ˜¯ä¸€ä¸ªå…¸èŒƒ
"""

#&&% å‡½æ•°è¿è¡ŒèŠ‚ç‚¹æ§åˆ¶


"""
é€šè¿‡æ‰“å°å¯¹è±¡çš„Handleæˆ‘ä»¬èƒ½å¾ˆæ–¹ä¾¿æ£€æµ‹å‡½æ•°è¿è¡Œè¿‡ç¨‹ä¸­çš„ç»“æœï¼Œè€Œä¸éœ€è¦ä»å‡½æ•°è¿”å›å€¼æˆ–éº»çƒ¦çš„è°ƒè¯•ä¸­æ£€æŸ¥é”™è¯¯
print(f"ä»å—{blk.Name}ä¸­å¯¹è±¡æŒ‰æ¯”ä¾‹å’Œé¢ç§¯ç‰¹å¾è¿‡æ»¤å¾—åˆ°çš„comå¯¹è±¡Handleå¥æŸ„")

print_coms_handle(filtered)

"""


#&&% å‡½æ•°çŠ¶æ€æ§åˆ¶

# å‡½æ•°è¿è¡ŒçŠ¶æ€çš„æ ‡å‡†åŒ–é‡æ„(é‡è£…è½¯ä»¶ä½¿é‡å¯çŠ¶æ€æ ‡å‡†åŒ–)

"""
å¦‚æœé‡å¯å¤©æ­£CADç³»ç»Ÿ5æ¬¡éƒ½æœªèƒ½è¿è¡Œï¼Œåˆ™ç³»ç»Ÿå°†è‡ªå·±é‡æ–°å®‰è£…å¤©æ­£å’ŒCAD

å¤©æ­£ç”»å›¾é”™è¯¯è¢«å¼ºåˆ¶ä¸­æ–­CADè¿›ç¨‹æ¶ˆæ¯çª—å£
activate_window_by_title("å›¾å½¢ä¿®å¤", click_titlebar= True)
[OK] å·²æ¿€æ´»çª—å£ï¼šâ€œå›¾å½¢ä¿®å¤â€ ä½ç½®(1089,562) å¤§å°366Ã—213
(1089, 562, 366, 213)
åŒä¸€æµ‹è¯•çš„å…¶çª—å£â€œå…³é—­â€æŒ‰é’®çš„ç‚¹å‡»ä½ç½® (1394, 724)ï¼Œç‚¹å‡»åå³æ­£å¸¸è¿è¡Œ



"""

# å‡½æ•°æ—¶é—´æ§åˆ¶

"""
å¦‚æœä¸€ä¸ªç³»ç»Ÿå‡½æ•°è¿è½¬å—é˜»ï¼Œå°†ä¼šè¢«æ—¶é—´ç›‘æ§ä¸­æ–­è¿›ç¨‹ï¼Œå¤©æ­£CADç³»ç»Ÿä¼šè¢«ä¸­æ–­é‡å¯


"""
def cad_zhongduan_ceshi():

    i=0

    while True:

        LP=[(0, i, 0.0), (90509.17232155753, 38080.72757883748, 0.0), (72659.40193916814, 48550.64173869454, 0.0), (61087.391734572855, 35617.55453017057, 0.0), (62182.914098626905, 23340.76471079199, 0.0), (43443.713317261485, 21669.276832727846, 0.0), (32018.9079127094, 29207.203355399004, 0.0), (22430.211826160492, 19081.580148860856, 0.0)]

        draw_lwpolyline(
    LP,
    layer_name= "0",
    width = 0.0,
    color= 256,
    closed = False
)

    i=i+10000




# å‡½æ•°è¿è¡Œæ¶ˆæ¯æ‰“å°å’Œç»“æ„æ ‡å‡†æ§åˆ¶

"""
åœ¨ç³»ç»Ÿçš„æ›´é«˜çº§å±‚é¢ï¼Œå‡½æ•°ä¸éœ€è¦æ˜¾ç¤ºåº•å±‚å‡½æ•°çš„è¿è½¬ä¿¡æ¯ï¼Œä»æé€Ÿçš„è§’åº¦è®²åº”è¯¥å»æ‰printè¯­å¥ï¼Œä½†è¿™ä¸ªå¯ä»¥åœ¨ç¨³å®šè¿è½¬ä¸€æ®µæ—¶é—´ä¹‹åå¤„ç†

"""


# å‡½æ•°é€»è¾‘ç»“æ„

"""
åº”è¯¥è®¾ç½®å‡½æ•°çš„æ˜ç¡®é€»è¾‘æ„æˆï¼Œåœ¨å„ä¸ªç¯èŠ‚æ‰“å°æ¶ˆæ¯ç”¨äºåˆ†æè¿è½¬æ˜¯å¦é”™è¯¯ã€‚é€šè¿‡å®šä¹‰å†…éƒ¨å‡½æ•°æˆ–å°å‡½æ•°ï¼Œå»ºç«‹æ­£ç¡®çš„æ¦‚å¿µå’Œæ¨¡å—ï¼Œä½¿å¾—äººå·¥æ™ºèƒ½ç³»ç»Ÿæ›´å®¹æ˜“è¢«é©¾é©­

"""


# å‡½æ•°æ¡ˆä¾‹ç”Ÿæˆ

"""
ç”Ÿæˆè¶³å¤Ÿå¤šçš„æµ‹è¯•æ–‡ä»¶ï¼ŒèŠ‚çœç¨‹åºæµ‹è¯•æ—¶é—´ï¼Œå¹¶é€šè¿‡å®Œæ•´æ¡ˆä¾‹ä½¿å‡½æ•°é€»è¾‘æ›´åˆç†

"""

# æ‰“å°æ¡ˆä¾‹ç”Ÿæˆå™¨

"""
generate_test_cases(
    r"D:/Myprogramsystem/BaiduSyncdisk/å®‹å²³/è‡ªåŠ¨åŒ–(åŠ¨æ€)/ç›®å½•å›¾ç­¾/ç”Ÿæˆæµ‹è¯•æ–‡ä»¶/ç›®å½•å›¾ç­¾.dwg",
    num = 3,
    jianju_x = 10000,
    juli_tukuang = 6000,
    juli_y = 400000,
)

"""

# æ–‡å­—å€™é€‰
TEXT_OPTIONS = [
    "ä¸€å±‚å¹³é¢å›¾", "äºŒå±‚å¹³é¢å›¾", "1-18è½´ç«‹é¢å›¾", "2-2å‰–é¢å›¾",
    "1-1å‰–é¢å›¾", "æ¥¼æ¢¯å¤§æ ·å›¾", "é—¨çª—è¯¦å›¾å›¾"
]

# è¾…åŠ©ï¼šåˆ›å»ºæ–‡å­—å®ä½“
def create_text(txt: str, position: Tuple[float, float, float], height: float, layer: str):
    vpt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, list(position))
    ent = mp.AddText(txt, vpt, height)
    try:
        ent.Layer = layer
    except Exception:
        pass
    return ent

# éšæœºç»˜åˆ¶åŸºæœ¬å›¾å½¢ï¼Œç½®äºlayer_tu
def draw_random_shape(xmin: float, xmax: float, ymin: float, ymax: float, layer_tu: str):
    """
    åœ¨ç»™å®šèŒƒå›´å†…éšæœºç»˜åˆ¶ä¸€ç‚¹ã€çº¿æ®µã€åœ†æˆ–æ­£å¤šè¾¹å½¢ï¼Œ
    å¹¶æ”¾ç½®åˆ°æŒ‡å®šå›¾å±‚ã€‚
    """
    typ = random.choice(["point", "line", "circle", "polygon"])
    shape = None
    if typ == "point":
        pt = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        shape = draw_point(pt)
    elif typ == "line":
        p1 = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        p2 = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        shape = draw_line(p1, p2)
    elif typ == "circle":
        center = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        r = random.uniform(1000, min(xmax - xmin, ymax - ymin) / 4)
        shape = draw_circle(center, r)
    else:
        center = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        r = random.uniform(1000, min(xmax - xmin, ymax - ymin) / 4)
        sides = random.randint(3, 8)
        shape = draw_regular_polygon(center, r, sides)
    # è®¾ç½®å›¾å±‚
    try:
        if shape is not None:
            shape.Layer = layer_tu
    except Exception:
        pass
def generate_test_cases(
    output_path: str,
    num: int = 3,
    jianju_x: float = 10000,
    juli_tukuang: float = 6000,
    juli_y: float = 400000,
):
    """
    ç”Ÿæˆæµ‹è¯• DWGï¼šå¤šè¡Œå›¾æ¡†+å›¾å½¢+è¾¹ç¼˜æ–‡å­—ï¼Œæ–‡å­—åŒºåŸŸç”± get_frame_edge è®¡ç®—ã€‚
    """
    sizes = {"A3": (42000, 29700), "A2": (59400, 42000), "A1": (84000, 59400)}
    seq_fixed = [
        ("A3", False), ("A2", False), ("A1", False),
        ("A3", True),  ("A2", False), ("A1", True),
        ("A2", False), ("A2", False),
    ]

    layer_kuang = "æµ‹è¯•ç»˜åˆ¶_å›¾æ¡†"
    layer_tu    = "æµ‹è¯•ç»˜åˆ¶_å›¾å½¢"
    layer_wz    = "æµ‹è¯•ç»˜åˆ¶_æ–‡å­—"
    #ä¸è¦†ç›–ä»å½“å‰æ—¶é—´ç»™å¸¦è·¯å¾„æ–‡ä»¶åå‘½å
    output_path = rename_time(output_path)
    
    create_new_dwg_file_no(output_path)
    time.sleep(1)

    base_x = random.uniform(0, 1000000)
    base_y = random.uniform(0, 1000000)

    for row in range(num):
        y0 = base_y + row * juli_y
        x = base_x
        seq = seq_fixed if row == 0 else [
            (random.choice(list(sizes.keys())), random.choice([False, True]))
            for _ in seq_fixed
        ]

        for name, vertical in seq:
            w, h = sizes[name]
            if vertical:
                pts = [(x, y0, 0), (x, y0 + w, 0), (x + h, y0 + w, 0),
                       (x + h, y0, 0), (x, y0, 0)]
            else:
                pts = [(x, y0, 0), (x + w, y0, 0), (x + w, y0 + h, 0),
                       (x, y0 + h, 0), (x, y0, 0)]
            # ç»˜åˆ¶å¹¶è·å–å¤šæ®µçº¿å®ä½“
            poly_ent = draw_polyline(pts, layer_name=layer_kuang, tol=0.5, width=20, color=1)
            # è®¡ç®—è¾¹åŸŸå››è§’ç‚¹
            edge_pts = get_frame_edge(poly_ent, juli_tukuang)
            if not edge_pts:
                continue
            ex0, ey0, _ = edge_pts[0]
            _, ey1, _ = edge_pts[1]
            ex1, _, _ = edge_pts[2]

            # å›¾æ¡†å†…éƒ¨éšæœºå›¾å½¢ï¼Œä¿ç•™å†…è¾¹ç¼˜
            # è®¡ç®—å†…éƒ¨åŒºåŸŸ
            if vertical:
                inner_xmin, inner_xmax = x, x + h
                inner_ymin, inner_ymax = y0 + juli_tukuang, y0 + w
                x_shift = h
            else:
                inner_xmin, inner_xmax = x, x + w
                inner_ymin, inner_ymax = y0, y0 + h - juli_tukuang
                x_shift = w
            # éšæœºå›¾å½¢
            for _ in range(random.randint(5, 10)):
                ensure_layer_current(layer_name="æµ‹è¯•ç»˜åˆ¶_å›¾å½¢")                
                draw_random_shape(
                    inner_xmin, inner_xmax,
                    inner_ymin, inner_ymax,
                    layer_tu
                )

            # è¾¹åŸŸæ–‡å­—ï¼š1-6 æ–‡æœ¬ï¼Œåˆ† 1-3 æ’ï¼Œé™å®šåœ¨ edge åŒºåŸŸ
            n_texts = random.randint(1, 6)
            rows_t = random.randint(1, min(3, n_texts))
            counts = [n_texts // rows_t] * rows_t
            for i in range(n_texts % rows_t):
                counts[i] += 1
            text_height = 300
            line_spacing = text_height + 100

            for r, cnt in enumerate(counts):
                choices = random.sample(TEXT_OPTIONS, cnt)
                line_str = 'ã€'.join(choices)
                if not vertical:
                    tx = random.uniform(ex0, ex1 - text_height)
                    ty = ey1 - (r + 1) * line_spacing
                    ty = min(max(ty, ey0), ey1 - text_height)
                else:
                    ty = random.uniform(ey0 + text_height, ey1)
                    tx = ex0 + r * line_spacing
                    tx = min(max(tx, ex0), ex1 - text_height)
                print(f"[DEBUG] å†™å…¥æ–‡å­— '{line_str}' åœ¨ ({tx:.2f},{ty:.2f}) å›¾å±‚ {layer_wz}")
                ensure_layer_current("æµ‹è¯•ç»˜åˆ¶_æ–‡å­—")                
                create_text(line_str, (tx, ty, 0), text_height, layer_wz)

            x += x_shift + jianju_x

    time.sleep(1)
    doc.SendCommand("Z\nE\n")
    savefile()
    guanbifile()

def get_frame_edge(poly_ent, juli_tukuang: float = 6000.0):
    """
    ç»™å®šä¸€ä¸ªé—­åˆå¤šæ®µçº¿å®ä½“ï¼Œåˆ¤æ–­å…¶æ¨ªå‘/ç«–å‘ï¼Œå¹¶è¿”å›è¾¹åŸŸå››è§’åæ ‡ã€‚

    :param poly_ent:    é—­åˆçš„ polyline å®ä½“ï¼ˆAcDbPolylineï¼‰
    :param juli_tukuang:è¾¹åŸŸå®½åº¦
    :return: [(x1,y1,0), (x1,y2,0), (x2,y2,0), (x2,y1,0)]
    """
    try:
        (pmin, pmax) = poly_ent.GetBoundingBox()
    except Exception as e:
        print(f"[ERROR] è·å–åŒ…å›´ç›’å¤±è´¥: {e}")
        return None
    xmin, ymin, _ = pmin
    xmax, ymax, _ = pmax
    w = xmax - xmin
    h = ymax - ymin
    if w >= h:
        orientation = 'horizontal'
        # æ¨ªå‘ï¼šè¾¹åŸŸåœ¨å³ä¾§
        x1 = xmax - juli_tukuang
        x2 = xmax
        y1 = ymin
        y2 = ymax
    else:
        orientation = 'vertical'
        # ç«–å‘ï¼šè¾¹åŸŸåœ¨åº•éƒ¨
        x1 = xmin
        x2 = xmax
        y1 = ymin
        y2 = ymin + juli_tukuang
    print(f"[DEBUG] æ¡†ä¸º {orientation}ï¼Œè¾¹åŸŸåæ ‡: ({x1:.2f},{y1:.2f}), ({x1:.2f},{y2:.2f}), ({x2:.2f},{y2:.2f}), ({x2:.2f},{y1:.2f})")
    return [(x1, y1, 0), (x1, y2, 0), (x2, y2, 0), (x2, y1, 0)]

#&&% æ‰“å°è®¾ç½®å’ŒpyautoguiåŸºæœ¬æ“ä½œ

"""
run_py("æµ‹é¼ æ ‡ä½ç½®.py")#æµ‹ä½ç½®

pyautogui.moveTo(1415, 796)#å®šä½

pyautogui.write("EXPLODE", interval=0.1)#å†™æ–‡å­—

pyautogui.press("enter")#å›è½¦

pyautogui.doubleClick()#åŒå‡»

pyautogui.hotkey('ctrl', 'a')#å…¨é€‰

pyautogui.press('delete')#åˆ é™¤

# ç„¶åä»è¯¥ä½ç½®å‘ä¸Šæ»šåŠ¨ 200
pyautogui.scroll(200)

# å‡è®¾å·²ç» moveTo(x,y)
pyautogui.mouseDown()
pyautogui.dragRel(200, 100, duration=0.3)  # å‘å³200ï¼Œä¸‹100çš„æ¡†é€‰
pyautogui.mouseUp()

å¦‚æœä½ çš„â€œé¢„å…ˆæœ‰çš„æ•°æ®â€ä¸æ˜¯å…¨åœ¨ä¸€ä¸ªæ–‡æœ¬æ¡†é‡Œï¼Œè€Œæ˜¯ä¸€ä¸ªåˆ—è¡¨/è¡¨æ ¼ä¸­çš„å¤šè¡Œï¼Œä½ å¯ä»¥
åœ¨å•å‡»ä¸€æ¬¡ä¹‹åï¼Œç”¨ pyautogui.dragTo() æˆ– pyautogui.dragRel() åšæ¡†é€‰
ç„¶åæŒ‰ delete

# æ–¹æ³•ä¸€ï¼šç›´æ¥å³é”®
pyautogui.click(button='right')
# æˆ–è€…ä½¿ç”¨å¿«æ·å‡½æ•°
# pyautogui.rightClick()

# å¦‚æœéœ€è¦åœ¨å¼¹å‡ºçš„ä¸Šä¸‹æ–‡èœå•é‡Œé€‰æ‹©æŸé¡¹ï¼Œæ¯”å¦‚å‘ä¸‹ç§»åŠ¨ä¸¤æ¬¡å†å›è½¦ï¼š
pyautogui.press('down', presses=2, interval=0.1)
pyautogui.press('enter')


# å‘é€ Ctrl+1
pyautogui.hotkey('ctrl', '1')
# å‘é€ Ctrl+9
pyautogui.hotkey('ctrl', '9')
# å‘é€ Shift+PrintScreen
# æ³¨æ„ PyAutoGUI é‡Œ PrintScreen é”®é€šå¸¸å« 'printscreen' æˆ– 'prtsc'
pyautogui.hotkey('shift', 'printscreen')

# Ctrl+1 æ‰‹åŠ¨ç‰ˆ
pyautogui.keyDown('ctrl')
pyautogui.press('1')
pyautogui.keyUp('ctrl')
pyautogui.press("esc")

currentLayout = acad.ActiveDocument.ActiveLayout
currentLayout.CanonicalMediaName = "ISO_A1_(841.00_x_594.00_MM)"
currentLayout.CanonicalMediaName#æŸ¥çœ‹å½“å‰å›¾çº¸è®¾ç½®


æ ‡å‡†A0,A1,A2,A3çš„å¯æ‰“å°åŒºåŸŸæ˜¯å¯å•ç‹¬ä¿®æ”¹çš„ï¼Œå…¶0-0-0-0æ•ˆæœå·²ç»æ‰“å°å‡ºæ¥ 


"""

dy_yonghu=[

                                           #A0 1189.00_x_841.00

    ("1337.63","841.00","17","17","5","5"),#A0+1/8


    ("1486.25","841.00","17","17","5","5"),#A0+1/4

                                           #A1 841.00_x_594.00

    ("1051.25","594.00","17","17","5","5"),#A1+1/4


    ("1261.50","594.00","17","17","5","5"),#A1+1/2


    ("1471.75","594.00","17","17","5","5"),#A1+3/4

                                           #A2 594.00_x_420.00


    ("742.50","420.00","17","17","5","5"), #A2+1/4

                                         

    ("891.00","420.00","17","17","5","5"), #A2+1/2


    ("1039.50","420.00","17","17","5","5") #A2+3/4


                                           #A3 420.00_x_297.00
]



def æ‰¹é‡è®¾ç½®ç”¨æˆ·æ‰“å°å°ºå¯¸(dy_yonghu):
    """
    å…ˆè¿æ¥ï¼Œå†äººå·¥æ‰“å¼€ctr+Pé€‰æ‹©DWG TO PDFæ‰“å°åå†è¿è¡Œæ­¤å‡½æ•°

    """

    #æŒ‰ç”¨æˆ·ç»™å®šæ•°æ®æ‰¹é‡ç”Ÿæˆç”¨æˆ·æ‰“å°å°ºå¯¸
    for i in range(0,len(dy_yonghu)):
        dyshu = dy_yonghu[i]
 
        pyautogui.moveTo(1331, 569)#ç‚¹å‡»æ‰“å°çª—å£çš„ ç‰¹æ€§ æŒ‰é’®
        pyautogui.click(1331, 569)        
        time.sleep(2)


        pyautogui.moveTo(1189, 627)#ç‚¹å‡»æ‰“å°çª—å£çš„ è‡ªå®šä¹‰å›¾çº¸å°ºå¯¸ æŒ‰é’®
        pyautogui.click(1189, 627)        
        time.sleep(2)

        pyautogui.moveTo(1415, 796)#ç‚¹å‡»æ‰“å°çª—å£çš„ ç¼–è¾‘ æŒ‰é’®
        pyautogui.click(1415, 796)        
        time.sleep(2)

        pyautogui.moveTo(1161, 690)#ç‚¹å‡»å®½åº¦çª—å£ä½ç½®
        pyautogui.click(1161, 690)        
        time.sleep(2)

        pyautogui.doubleClick()#é€‰é¢„è®¾æ•°æ®
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[0], interval=0.1)#å†™å®½åº¦
        time.sleep(2)

        pyautogui.moveTo(1171, 745)#ç‚¹å‡»é«˜åº¦çª—å£ä½ç½®
        pyautogui.click(1171, 745)        
        time.sleep(2)

        pyautogui.doubleClick()#é€‰é¢„è®¾æ•°æ®
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[1], interval=0.1)#å†™é«˜åº¦
        time.sleep(2)
        
        pyautogui.moveTo(1452, 915)#ç‚¹å‡»ä¸‹ä¸€é¡µ
        pyautogui.click(1452, 915)        
        time.sleep(2)

        pyautogui.moveTo(1144, 684)#ç‚¹å‡»ä¸Šè¾¹è·çª—å£ä½ç½®
        pyautogui.click(1144, 684)        
        time.sleep(2)

        pyautogui.doubleClick()#é€‰é¢„è®¾æ•°æ®
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[2], interval=0.1)#å†™ä¸Šè¾¹è·
        time.sleep(2)

        pyautogui.moveTo(1144, 729)#ç‚¹å‡»ä¸‹è¾¹è·çª—å£ä½ç½®
        pyautogui.click(1144, 729)        
        time.sleep(2)


        pyautogui.doubleClick()#é€‰é¢„è®¾æ•°æ®
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[3], interval=0.1)#å†™ä¸‹è¾¹è·
        time.sleep(2)

        pyautogui.moveTo(1143, 770)#ç‚¹å‡»å·¦è¾¹è·çª—å£ä½ç½®
        pyautogui.click(1143, 770)        
        time.sleep(2)

        pyautogui.doubleClick()#é€‰é¢„è®¾æ•°æ®
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[4], interval=0.1)#å†™å·¦è¾¹è·
        time.sleep(2)

        pyautogui.moveTo(1147, 809)#ç‚¹å‡»å³è¾¹è·çª—å£ä½ç½®
        pyautogui.click(1147, 809)        
        time.sleep(2)

        pyautogui.doubleClick()#é€‰é¢„è®¾æ•°æ®
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[5], interval=0.1)#å†™å³è¾¹è·
        time.sleep(2)

        pyautogui.moveTo(1443, 919)#ç‚¹å‡»ä¸‹ä¸€é¡µ
        pyautogui.click(1443, 919)        
        time.sleep(2)

        pyautogui.moveTo(1443, 922)#ç‚¹å‡»ä¸‹ä¸€é¡µ
        pyautogui.click(1443, 922)        
        time.sleep(2)

        pyautogui.moveTo(1460, 921)#ç‚¹å‡»å®Œæˆ
        pyautogui.click(1460, 921)        
        time.sleep(2)


        pyautogui.moveTo(1239, 920)#ç‚¹å‡»ç¡®å®š
        pyautogui.click(1239, 920)        
        time.sleep(2)

##ç¡®ä¿åˆ é™¤å¯¹è±¡

def safe_delete(ob, retries: int = 5, delay: float = 1.0) -> bool:
    """
    å°è¯•åˆ é™¤ CAD å¯¹è±¡ obï¼Œæœ€å¤šé‡è¯• retries æ¬¡ï¼Œæ¯æ¬¡é—´éš” delay ç§’ã€‚
    åªæ•è· COM é”™è¯¯ï¼ŒæˆåŠŸè¿”å› Trueï¼Œå¦åˆ™è¿”å› Falseã€‚
    """


    for attempt in range(1, retries + 1):
        try:
            ob.Delete()
            return True
        except pywintypes.com_error:
            time.sleep(delay)
    return False





#&&&%  * é«˜äº®é€‰æ‹©è½¬éšæ€§ç§»åŠ¨åŒºåŸŸå†…å…¨éƒ¨å¯¹è±¡


"""
é«˜äº®é€‰æ‹©çª—å£æ“ä½œæ›´å¯é ä»è€Œæ›´å¿«ï¼Œä½†è¦è€ƒè™‘çª—å£ä¸èƒ½æŒ¡

"""

def move_entities_in_region(coms, target=(0,0,0), ty=1, max_iter=3):
    """
    å°† `coms` å¯¹è±¡çš„åŒ…å›´ç›’å†…æ‰€æœ‰å®ä½“ï¼Œæ²¿å‘é‡ (target - å·¦ä¸‹è§’) ç§»åŠ¨ï¼Œ
    æ¯è½®ç­‰å¾… `ty` ç§’ï¼Œæœ€å¤šå¾ªç¯ `max_iter` æ¬¡ã€‚

    å‚æ•°ï¼š
      coms      -- æ”¯æŒ GetBoundingBox() çš„ COM å¯¹è±¡ï¼ˆå¦‚å¤šæ®µçº¿ã€å—å‚ç…§ç­‰ï¼‰
      target    -- ç›®æ ‡åŸºç‚¹åæ ‡ï¼Œé»˜è®¤ä¸º (0,0,0)
      ty        -- æ¯è½®ç§»åŠ¨åç­‰å¾…ç§’æ•°
      max_iter  -- æœ€å¤šå°è¯•çš„è½®æ•°
    """
    # 1. è¯»å–åŒ…å›´ç›’ï¼Œè®¡ç®—å·¦ä¸‹è§’ (x1,y1) å’Œå³ä¸Šè§’ (x2,y2)
    p1, p2 = coms.GetBoundingBox()
    x1, y1 = min(p1[0], p2[0]), min(p1[1], p2[1])
    x2, y2 = max(p1[0], p2[0]), max(p1[1], p2[1])

    # 2. è®¡ç®—ä½ç§»å‘é‡ Î” = target - (x1,y1)
    dx = target[0] - x1
    dy = target[1] - y1
    dz = target[2] - 0.0

    for i in range(1, max_iter + 1):
        # æ¸…ç©ºæ˜¾æ€§é€‰æ‹©é›†
        try:
            doc.PickfirstSelectionSet.Clear()
        except Exception:
            pass

        # 3. åœ¨ CAD é‡Œç”¨çª—å£é€‰æ‹©é«˜äº®åŒºåŸŸå†…å¯¹è±¡
        highlight_entities_in_window(x1, y1, x2, y2)

        # 4. æ‹¿åˆ°æ˜¾æ€§é€‰æ‹©é›†é‡Œçš„å®ä½“
        pick = doc.PickfirstSelectionSet
        count = pick.Count if hasattr(pick, 'Count') else len(list(pick))
        if count == 0:
            print(f"[OK] ç¬¬ {i} è½®ï¼šåŒºåŸŸå·²æ¸…ç©ºï¼Œåœæ­¢ã€‚")
            break

        print(f"â™»ï¸ ç¬¬ {i} è½®ï¼šæ£€æµ‹åˆ° {count} ä¸ªå¯¹è±¡ï¼Œæ­£åœ¨ç§»åŠ¨â€¦")
        # 5. å¯¹æ¯ä¸ªå®ä½“æŒ‰è®¡ç®—å¥½çš„å‘é‡æ‰§è¡Œ Move
        for ent in pick:
            try:
                # ä» (x1,y1,0) ç§»åŠ¨åˆ° (x1+dx, y1+dy, dz)
                ent.Move(vtpnt(x1, y1, 0.0),
                         vtpnt(x1 + dx, y1 + dy, dz))
            except Exception as e:
                print(f"  [è­¦å‘Š] å¯¹è±¡ {ent.Handle} ç§»åŠ¨å¤±è´¥ï¼š{e}")

        # 6. ç­‰å¾… CAD å®Œæˆå‘½ä»¤å†è¿›å…¥ä¸‹ä¸€è½®
        time.sleep(ty)
    else:
        print("[è­¦å‘Š] è¾¾åˆ°æœ€å¤§è¿­ä»£æ¬¡æ•°ï¼Œå¯èƒ½ä»æœ‰æ®‹ç•™å¯¹è±¡ã€‚")





###çŸ©å½¢æ ‡å‡†åŒ–åŠæ•´ä½“ä¼˜åŒ–
"""
ä½¿ç”¨parse_rectangle_pointså‡½æ•°æ¥å—ç”¨æˆ·å¯¹çŸ©å½¢å¯¹è§’ç‚¹ä¸¤ç§æ–¹å¼çš„è‡ªç”±ç†è§£,ä½¿ç”¨å®ƒ

æ”¹é€ å·²æœ‰å‡½æ•°ï¼Œä»è€Œä¿®å¤è¿™ç§å› ä¸ºå¯¹è§’ç‚¹ä¸åŒç†è§£çš„æ··ä¹±

jd()åº”è¯¥åœ¨è¿æ¥æ—¶å°±è®¾ç½®ï¼Œæ§åˆ¶ç»˜å›¾ç²¾åº¦

æœ‰äº›å‡½æ•°å¤„ç†è¿˜å¯ä»¥ä¼˜åŒ–ï¼Œæ˜¾æ€§é€‰æ‹©ä¸è¦åœ¨çª—å£ä¸Šé®æŒ¡ï¼Œä¸è¦åŠ¨çª—å£

ç¨³å®šæ€§å’Œå¼ºå£®æ€§éšä½¿ç”¨åé¦ˆå˜åŒ–

å¤åˆ¶ä¸€ä¸ªå‚ç…§ç‰ˆæœ¬ä¾¿äºæŸ¥é˜…å‡½æ•°

ä»¥æ•°æ®åº“ä¸ºåŸºç¡€ä¸æ–­æ‰©å±•æ•°æ®ï¼Œæ‰©å¼ æ•´ä¸ªç»˜å›¾ç³»ç»Ÿçš„æƒ³æ³•æ˜¯å¯¹çš„


åº”è¯¥æœ‰ä¸€ä¸ªæ–‡ä»¶åˆå§‹åŒ–å‡½æ•°ï¼Œå°†å¤„ç†æ–‡ä»¶è½¬å…¥æ ‡å‡†è®¾ç½®çŠ¶æ€ï¼Œä»ç²¾åº¦ï¼Œå­—ä½“ï¼Œå¢™çº¿æ˜¾ç¤ºåŠ ç²—ï¼Œæ‰“å°ç©ºé—´å®‰å…¨ï¼Œæ‰“å°é…ç½®ï¼Œè§†å›¾è¾…åŠ©ç¯å¢ƒç­‰ç­‰ã€‚


å…è®¸è¾“å…¥ä¸€æ¡å¤šæ®µçº¿æˆ–å¤šæ ¹å¤šæ®µçº¿ï¼Œæ”¹å–„å‡½æ•°æ¥å£ï¼Œè¿™ç§é”™è¯¯æ˜¯åˆä¹äººç±»æ€ç»´çš„


åœ¨ä¸¤ä¸ªå­—å…¸ä¹‹é—´å®šä¹‰å‡½æ•°ï¼Œé€šè¿‡å­—å…¸æŸ¥æ‰¾ç¡®å®šä¿¡æ¯








"""


#&&% å­—å…¸çš„ç”Ÿæˆã€ä½¿ç”¨ã€æ‰©å±•

"""
##{
##  # é¡¶å±‚é”®ï¼šæ¯ä¸ªå›¾ç­¾æ¡†çš„ BlockReference.Handleï¼ˆå­—ç¬¦ä¸²ï¼‰
##  "2BC3": {
##      # â‘  è¿™æ¡è®°å½•å¯¹åº”çš„æ‰“å°æ¡†ä¿¡æ¯ï¼Œç»Ÿä¸€æ”¾åˆ°ä¸€ä¸ªå­å­—å…¸é‡Œ
##      "frame_info": {
##          # 4 ä¸ªè§’ç‚¹ï¼ˆ3Dï¼‰    
##          "corners": [
##              (x1, y1, z1),
##              (x1, y2, z1),
##              (x2, y2, z1),
##              (x2, y1, z1),
##          ],
##          # ISO ç¼–å·æˆ–åç§°
##          "format": "ISO_A2_(594.00_x_420.00_MM)",
##          # è¡¨ç¤ºæ¯”ä¾‹
##          "scale":  "1:100",
##          # çº¸å¼ è§„æ ¼ï¼ˆA2/A3/A1â€¦ï¼‰
##          "paper_size": "A2",
##      },
##      # â‘¡ æ’å…¥æˆåŠŸåçš„æ ‡é¢˜å—å¼•ç”¨  
##      "title_block_handle": "2C10",         # Handle å­—ç¬¦ä¸²
##      "title_block_entity": <COMObject ...> # ï¼ˆå¯é€‰ï¼Œæ–¹ä¾¿ç«‹åˆ»æ“ä½œï¼‰
##  },
##
##  "2BC4": {
##    # â€¦â€¦
##  }
##}
##
##åœ¨ä¸Šé¢ç¤ºä¾‹é‡Œï¼š
##
##é¡¶å±‚ key ç”¨ Handleï¼ˆå­—ç¬¦ä¸²å½¢å¼ï¼‰æ¥ç´¢å¼•æ¯ä¸€æ¡è®°å½•ï¼›
##
##frame_infoï¼šä¸€ä¸ªâ€œæ‰“å°æ¡†â€ç›¸å…³çš„å­å­—å…¸ï¼Œé‡Œé¢å†ç»†åˆ†
##
##cornersï¼šå››ä¸ª 3D åæ ‡ç‚¹
##
##formatï¼šå›¾å¹…åç§°ï¼ˆISO_â€¦ï¼‰
##
##scaleï¼šæ‰“å°æ¯”ä¾‹ï¼ˆæ¯”å¦‚ "1:100"ï¼‰
##
##paper_sizeï¼šè§„æ ¼ï¼ˆæ¯”å¦‚ "A2"ã€"A3"ï¼‰
##
##title_block_handleï¼šæ–°æ’å…¥çš„å›¾ç­¾å—å¼•ç”¨çš„ Handleï¼Œç”¨äºå¿«é€ŸæŸ¥å›æˆ–äºŒæ¬¡å¤„ç†ï¼›
##
##title_block_entityï¼ˆå¯é€‰ï¼‰ï¼šä¿ç•™ä¸€ä»½ COM å¯¹è±¡ï¼Œä»¥å…ä½ åé¢è¿˜æƒ³é©¬ä¸Šè°ƒç”¨å®ƒçš„å±æ€§ï¼æ–¹æ³•ã€‚
##
##ä¸ºä»€ä¹ˆè¿™æ ·å‘½åæ›´æ˜“ç»´æŠ¤ï¼Ÿ
##è¯­ä¹‰æ˜ç¡®
##
##formatã€scaleã€paper_size ä¸€çœ‹å°±çŸ¥é“åˆ†åˆ«ä»£è¡¨å›¾å¹…ã€æ¯”ä¾‹ã€è§„æ ¼ã€‚
##
##å±‚æ¬¡åˆ†æ˜
##
##æ‰€æœ‰è·Ÿâ€œæ‰“å°æ¡†â€ç›¸å…³çš„å†…å®¹éƒ½æ”¾ frame_infoï¼Œä»¥åè¦æ‰©å±•ï¼ˆæ¯”å¦‚åŠ  â€œåç§»é‡ offsetâ€ï¼‰ä¹Ÿå¯ä»¥æ”¾è¿›å»ã€‚
##
##ä¸ä¹‹å¹³è¡Œçš„æ˜¯æ ‡é¢˜å—ä¿¡æ¯ (title_block_â€¦)ã€‚
##
##Handle åšä¸»é”®
##
##å¤–å±‚ç›´æ¥ç”¨ BlockReference.Handle å­—ç¬¦ä¸²ï¼Œæ—¢ä¸æ˜“å†²çªï¼Œä¹Ÿæ–¹ä¾¿åºåˆ—åŒ–ã€å­˜ç›˜ã€‚
##
##ä¿ç•™å®ä½“æˆ–å¥æŸ„
##
##å­˜ä¸‹ title_block_entity å¯ç›´æ¥è°ƒç”¨ï¼›å­˜ä¸‹ title_block_handle å¯åœ¨é‡è½½æ–‡æ¡£åé‡æ–°æŸ¥æ‰¾å®ä¾‹ã€‚
##
##ä½ å¯ä»¥æ®æ­¤æŠŠå·²æœ‰çš„ zd_tuqian åšä¸€æ¬¡é‡å‘½åï¼é‡æ„ï¼Œä¾‹å¦‚ï¼š
##
##python
##å¤åˆ¶
##ç¼–è¾‘
##def normalize_tuqian_dict(raw: dict) -> dict:
##    
##    å°†åŸå§‹çš„ zd_tuqian è½¬æ¢ä¸ºæ›´æ¸…æ™°çš„ç»“æ„ï¼š
##    {
##      handle: {
##        "frame_info": {
##          "corners":    â€¦,
##          "format":     â€¦,
##          "scale":      â€¦,
##          "paper_size": â€¦,
##        },
##        "title_block_handle": â€¦,
##        "title_block_entity": â€¦
##      }, â€¦
##    }
##    
##    new = {}
##    for handle, info in raw.items():
##        fi = info["frame_info"]
##        new[handle] = {
##            "frame_info": {
##                "corners":    fi["corners"],
##                "format":     fi["drawing_frame"],
##                "scale":      fi["ratio"],
##                "paper_size": fi["spec"],
##            },
##            "title_block_handle": getattr(info["title_block"], "Handle", None),
##            "title_block_entity": info["title_block"],
##        }
##    return new
##è¿™æ ·ï¼Œåç»­æ‰€æœ‰åŸºäº zd_tuqian è¯»å–æ¯”ä¾‹ã€è§„æ ¼ç­‰ï¼Œéƒ½èƒ½å†™æˆï¼š
##
##entry = zd_tuqian[handle]
##scale     = entry["frame_info"]["scale"]
##paper_sz  = entry["frame_info"]["paper_size"]
##blk_handle = entry["title_block_handle"]
##å¸Œæœ›è¿™ä¸ªå‘½åè§„èŒƒèƒ½å¸®åŠ©ä½ æŠŠå­—å…¸ç»“æ„ç†å¾—æ›´æ¸…æ™°ï¼Œä¹Ÿæ›´å®¹æ˜“ç»´æŠ¤å’Œæ‰©å±•ã€‚
##
##
##entities = [info["entity"] for info in res.values()]
##specs     = [info["spec"]   for info in res.values()]
##ratios    = [info["ratio"]  for info in res.values()]
##
##info =bindy.get("2D07")
##info['title_block']
##å­—å…¸çš„äº§ç”Ÿï¼Œä¿å­˜ï¼Œè°ƒç”¨éå¸¸æœ‰ä»·å€¼ï¼Œå®ƒå°†ä¸ºæˆ‘ä»¬å­˜å–å¤§é‡ä¿¡æ¯ï¼Œä¹Ÿæ”¹å˜äº†æ•´ä¸ªç¨‹åºç»“æ„
##
### æ‹¿åˆ°æŸä¸ª title_block_handle çš„ info
##info = bind_dict[handle]
##
### ç§»åŠ¨å›¾ç­¾å—
##tb = info['title_block']
##tb.Move(vtpnt(0,0,0), vtpnt(0,10000,0))
##
### ç§»åŠ¨å¯¹åº”çš„æ‰“å°æ¡† polyline
##pl = info['frame_info']['entity']
##pl.Move(vtpnt(0,0,0), vtpnt(0,10000,0))


"""

def åœ†ç‚¹(tz=1):


    """
    æ§åˆ¶ç‚¹çš„æ˜¾ç¤º

    """

    zhi=0

    if tz ==1:

        zhi=35

    acad.ActiveDocument.SetVariable("PDMODE", zhi)#å°±æ˜¯ç‚¹æœ€å¥½çš„åœ†åŠ åå­—å½¢æ˜¾ç¤º



def å›¾çº¸èƒŒæ™¯(zhi = 16777215 ):

    acad.ActiveDocument.Application.preferences.Display.GraphicsWinModelBackgrndColor = zhi#0å³å˜æˆé»‘è‰²





#&&&%  *** æŒ‰åŒºåŸŸè°ƒæ•´è§†å›¾

def  shitu_region(x1,y1,x2,y2):

    """
    æŒ‰æŒ‰å¯¹è±¡å¤–åŒ…ç›’è°ƒæ•´è§†å›¾
    ä½¿ç”¨æ—¶å¯shitu_region(*p),p=(x1,y1,x2,y2)

    """

    h=0.3*(abs(x1-x2)+abs(y1-y2))/2

    # 1ï¸âƒ£ ç¼©æ”¾è§†å›¾åˆ°åˆé€‚çª—å£
    zoom_cmd = (
        f"_.ZOOM\n"      # è°ƒç”¨ Zoom
        f"_W\n"          # çª—å£é€‰é¡¹
        f"{x1-h},{y1-h}\n"   # ç¬¬ä¸€ä¸ªè§’ç‚¹
        f"{x2+h},{y2+h}\n"   # ç¬¬äºŒä¸ªè§’ç‚¹
    )
    doc.SendCommand(zoom_cmd)

#æŒ‰å¯¹è±¡å¤–åŒ…ç›’è°ƒæ•´è§†å›¾

def  shitu_entity(obj):

    """
    æŒ‰æŒ‰å¯¹è±¡å¤–åŒ…ç›’è°ƒæ•´è§†å›¾

    """
    p1,p2=obj.GetBoundingBox()

    x1,y1=p1[0],p1[1]

    x2,y2=p2[0],p2[1]

    h=0.3*(abs(x1-x2)+abs(y1-y2))/2

    # 1ï¸âƒ£ ç¼©æ”¾è§†å›¾åˆ°åˆé€‚çª—å£
    zoom_cmd = (
        f"_.ZOOM\n"      # è°ƒç”¨ Zoom
        f"_W\n"          # çª—å£é€‰é¡¹
        f"{x1-h},{y1-h}\n"   # ç¬¬ä¸€ä¸ªè§’ç‚¹
        f"{x2+h},{y2+h}\n"   # ç¬¬äºŒä¸ªè§’ç‚¹
    )
    doc.SendCommand(zoom_cmd)


def ensure_list(obj, element_type=None):
    """
    å¦‚æœ obj å·²ç»æ˜¯åˆ—è¡¨ï¼Œåˆ™åŸæ ·è¿”å›ï¼›
    å¦åˆ™ï¼Œå¦‚æœæ²¡æœ‰æŒ‡å®š element_typeï¼Œæˆ– obj æ˜¯ element_type çš„å®ä¾‹ï¼Œ
      åˆ™å°†å…¶åŒ…è£…æˆå•å…ƒç´ åˆ—è¡¨è¿”å›ï¼›
    å¦åˆ™æŠ›å‡º TypeErrorã€‚

    :param obj: å•ä¸ªå…ƒç´ æˆ–å…ƒç´ åˆ—è¡¨
    :param element_type: åˆ—è¡¨å…ƒç´ åº”æœ‰çš„ç±»å‹ï¼ˆå¯é€‰ï¼‰
    :return: å…ƒç´ åˆ—è¡¨
    """
    # 1) å¦‚æœå·²ç»æ˜¯åˆ—è¡¨ï¼Œç›´æ¥è¿”å›
    if isinstance(obj, list):
        return obj

    # 2) è¦ä¹ˆä¸é™åˆ¶ç±»å‹ï¼Œè¦ä¹ˆ obj æ˜¯æŒ‡å®šç±»å‹
    if element_type is None or isinstance(obj, element_type):
        return [obj]

    # 3) å…¶å®ƒæƒ…å†µæŠ¥é”™
    raise TypeError(
        f"æœŸæœ›ç±»å‹ {element_type.__name__} æˆ– List[{element_type.__name__}]ï¼Œ"
        f"ä½†æ”¶åˆ° {type(obj).__name__}"
    )


#  ä¸»å‡½æ•°
#  (1)
# æŸæŸåŠŸèƒ½ä½œç”¨çš„å‡½æ•°

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

"""
æ–‡ä»¶å¤„ç†ä¹‹å‰æ‰§è¡Œ
collect_all_texts()
å°†æ‰€æœ‰æ–‡å­—æ”¾å…¥å¤©æ­£"PUB_TEXT"å›¾å±‚ï¼Œåç»­æ— éœ€å†å»é‡å¤

æ‰§è¡Œå¤§æ–‡ä»¶å˜æ–‡ä»¶å¤¹+æ¯å¼ å›¾ä¸€ä¸ªæ–‡ä»¶ï¼Œå‡å°‘ä¸å¿…è¦çš„é€‰æ‹©è¿ç®—


"""


#&&% çª—å£é”®ç›˜æ§åˆ¶
"""
# 1. æˆªå–å…¨å±ï¼Œè¿”å›ä¸€ä¸ª PIL.Image å¯¹è±¡
img = pyautogui.screenshot()

# 2. å¦‚æœæƒ³ç›´æ¥ä¿å­˜åˆ°æ–‡ä»¶ï¼š
pyautogui.screenshot('full_screen.png')

# 3. åªæˆªå–å±å¹•çš„ä¸€éƒ¨åˆ†ï¼ˆx, y, width, heightï¼‰
#    ä¾‹å¦‚ï¼šä»å·¦ä¸Šè§’ (100,100) å¼€å§‹ï¼Œæˆªå– 300Ã—200 çš„åŒºåŸŸ
region_img = pyautogui.screenshot(region=(100, 100, 300, 200))
region_img.save('partial.png')


"""

#GIFå½•å±

def record_screen_gif(
    output_path: str,
    duration: float = 5.0,
    fps: int = 10,
    region: tuple[int,int,int,int] | None = None
):
    """
    å½•åˆ¶å±å¹•å¹¶ä¿å­˜ä¸º GIFã€‚

    :param output_path: è¾“å‡º GIF æ–‡ä»¶è·¯å¾„ï¼Œæ¯”å¦‚ "demo.gif"
    :param duration:    å½•åˆ¶æ—¶é•¿ï¼ˆç§’ï¼‰
    :param fps:         å¸§ç‡ï¼ˆæ¯ç§’æˆªå¤šå°‘å¸§ï¼‰
    :param region:      å¯é€‰ï¼Œ(left, top, width, height) åªå½•åˆ¶è¿™ä¸€åŒºåŸŸ
    """
    import imageio 

    frames = []
    interval = 1.0 / fps
    end_time = time.time() + duration

    print(f"â–¶ å¼€å§‹å½•åˆ¶ï¼šæ—¶é•¿={duration}sï¼Œå¸§ç‡={fps}fpsï¼ŒåŒºåŸŸ={region or 'å…¨å±'}")
    while time.time() < end_time:
        img = pyautogui.screenshot(region=region)  # PIL.Image å¯¹è±¡
        frames.append(img)
        time.sleep(interval)

    print(f"ğŸ›‘ å½•åˆ¶ç»“æŸï¼Œå…±æ•è· {len(frames)} å¸§ï¼Œæ­£åœ¨åˆæˆ GIFâ€¦â€¦")
    # imageio èƒ½ç›´æ¥æ¥å— PIL.Image
    imageio.mimsave(output_path, frames, fps=fps)
    print(f"[OK] å·²ä¿å­˜ä¸º {output_path}")


#åŠ¨ç”»å½•å±










def minimize_all_windows():
    """
    æ¨¡æ‹ŸæŒ‰ä¸‹ Win+Mï¼Œå°†æ‰€æœ‰çª—å£æœ€å°åŒ–ã€‚

    """

    import ctypes

    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B   # å·¦ Win é”®
    VK_M    = 0x4D   # M é”®
    KEYEVENTF_KEYUP = 0x2

    # æŒ‰ä¸‹ Win
    user32.keybd_event(VK_LWIN, 0, 0, 0)
    # æŒ‰ä¸‹ M
    user32.keybd_event(VK_M,    0, 0, 0)
    # æ¾å¼€ M
    user32.keybd_event(VK_M,    0, KEYEVENTF_KEYUP, 0)
    # æ¾å¼€ Win
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)

    # ç»™ç³»ç»Ÿä¸€ç‚¹ååº”æ—¶é—´
    time.sleep(0.1)



#&&%#æ§åˆ¶CADå±å¹•çª—å£åœ¨å·¦ä¸Šè§’


def set_autocad_window_to_top_left(resize_half: bool = True):
    """
    å°† AutoCAD çª—å£è¿˜åŸå¹¶ç§»åŠ¨åˆ°å±å¹•å·¦ä¸Šè§’ï¼Œå¯é€‰å°†å…¶è°ƒæ•´ä¸ºåŠå±å¤§å°ã€‚
    """
    # 1ï¸âƒ£ æ‰¾åˆ°å¯è§æˆ–æœ€å°åŒ–çš„ AutoCAD çª—å£
    windows = [w for w in gw.getWindowsWithTitle('AutoCAD')]
    if not windows:
        print("[é”™è¯¯] æœªæ‰¾åˆ° AutoCAD çª—å£ï¼")
        return
    win = windows[0]

    # 2ï¸âƒ£ å¦‚æœçª—å£æœ€å°åŒ–ï¼Œå…ˆè¿˜åŸ
    if win.isMinimized:
        win.restore()
        time.sleep(0.3)

    # 3ï¸âƒ£ æ¿€æ´»çª—å£ï¼Œç¡®ä¿æˆ‘ä»¬æ“ä½œçš„æ˜¯çœŸæ­£çš„å‰å°çª—å£
    try:
        win.activate()
    except Exception:
        # æœ‰æ—¶ activate ä¹Ÿä¼šå¤±è´¥ï¼ŒçŸ­æš‚ç­‰å¾…å†è¯•
        time.sleep(0.2)
        win.activate()
    time.sleep(0.2)

    # 4ï¸âƒ£ ç§»åŠ¨åˆ°å·¦ä¸Šè§’ (0,0)
    win.moveTo(0, 0)
    time.sleep(0.2)

    # 5ï¸âƒ£ å¯é€‰ï¼šå°†çª—å£è°ƒæ•´ä¸ºå±å¹•ä¸€åŠå¤§å°
    if resize_half:
        screen_w, screen_h = pyautogui.size()
        win.resizeTo(screen_w // 2, screen_h // 2)
        time.sleep(0.2)
        print(f"[OK] AutoCAD çª—å£å·²æ¢å¤å¹¶ç§»åŠ¨åˆ°å·¦ä¸Šè§’ï¼Œå°ºå¯¸è®¾ä¸º {screen_w//2} x {screen_h//2}")
    else:
        print("[OK] AutoCAD çª—å£å·²æ¢å¤å¹¶ç§»åŠ¨åˆ°å·¦ä¸Šè§’")



def l():

    set_autocad_window_to_top_left()



#&&% æ›´åˆç†æ§åˆ¶çª—å£å‡½æ•°
# â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- -- -- -- -- â€” â€” â€” â€” -- -- -- -- --  â€” â€” â€” â€” -- --

def minimize_all_windows_d():
    """
    æ¨¡æ‹Ÿ Win + Dï¼Œå°†æ‰€æœ‰çª—å£æœ€å°åŒ–ï¼ˆåˆ‡æ¢ï¼‰ã€‚
    """
    # VK_LWIN = 0x5B, VK_D = 0x44, KEYEVENTF_KEYUP = 0x2
    ctypes.windll.user32.keybd_event(0x5B, 0, 0, 0)       # æŒ‰ä¸‹ Win
    ctypes.windll.user32.keybd_event(0x44, 0, 0, 0)       # æŒ‰ä¸‹ D
    ctypes.windll.user32.keybd_event(0x44, 0, 0x2, 0)     # æ¾å¼€ D
    ctypes.windll.user32.keybd_event(0x5B, 0, 0x2, 0)     # æ¾å¼€ Win
    time.sleep(0.3)


def minimize_all_windows_m():
    """
    æ¨¡æ‹ŸæŒ‰ä¸‹ Win+Mï¼Œå°†æ‰€æœ‰çª—å£æœ€å°åŒ–ã€‚
    """
    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B   # å·¦ Win é”®
    VK_M    = 0x4D   # M é”®
    KEYEVENTF_KEYUP = 0x2

    # æŒ‰ä¸‹ Win
    user32.keybd_event(VK_LWIN, 0, 0, 0)
    # æŒ‰ä¸‹ M
    user32.keybd_event(VK_M,    0, 0, 0)
    # æ¾å¼€ M
    user32.keybd_event(VK_M,    0, KEYEVENTF_KEYUP, 0)
    # æ¾å¼€ Win
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)

    # ç»™ç³»ç»Ÿä¸€ç‚¹ååº”æ—¶é—´
    time.sleep(0.1)


def restore_and_position(
    name: str = "AutoCAD",
    width_ratio: float = 0.5,
    height_ratio: float = 0.5,
    x: int = 0,
    y: int = 0
) -> bool:
    """
    å°†ç¬¬ä¸€ä¸ªæ ‡é¢˜ä¸­åŒ…å« name çš„çª—å£æ¢å¤ã€æ¿€æ´»ï¼Œå¹¶è°ƒæ•´åˆ°æŒ‡å®šä½ç½®å’Œå¤§å°ã€‚

    :param name:          çª—å£æ ‡é¢˜å…³é”®å­—ï¼Œé»˜è®¤ "AutoCAD"
    :param width_ratio:   çª—å£å®½åº¦å å±å¹•å®½åº¦çš„æ¯”ä¾‹ (0 < ratio â‰¤ 1)
    :param height_ratio:  çª—å£é«˜åº¦å å±å¹•é«˜åº¦çš„æ¯”ä¾‹ (0 < ratio â‰¤ 1)
    :param x:             çª—å£å·¦ä¸Šè§’ X åæ ‡ï¼Œé»˜è®¤ 0
    :param y:             çª—å£å·¦ä¸Šè§’ Y åæ ‡ï¼Œé»˜è®¤ 0
    :return:              æ‰¾åˆ°å¹¶æ“ä½œæˆåŠŸè¿”å› Trueï¼Œå¦åˆ™ False
    """
    # â‘  æŸ¥æ‰¾çª—å£
    candidates = [w for w in gw.getWindowsWithTitle(name) if w.title]
    if not candidates:
        print(f"[é”™è¯¯] æœªæ‰¾åˆ°æ ‡é¢˜åŒ…å« â€œ{name}â€ çš„çª—å£ã€‚")
        return False
    win = candidates[0]

    # â‘¡ å¦‚æœæœ€å°åŒ–ï¼Œåˆ™è¿˜åŸ
    if win.isMinimized:
        try:
            win.restore()
            time.sleep(0.2)
        except Exception as e:
            print(f"[è­¦å‘Š] æ— æ³•è¿˜åŸçª—å£ï¼š{e}")

    # â‘¢ æ¿€æ´»çª—å£ï¼ˆç½®äºæœ€å‰ï¼‰
    try:
        win.activate()
    except Exception:
        time.sleep(0.1)
        try:
            win.activate()
        except Exception as e:
            print(f"[è­¦å‘Š] æ¿€æ´»çª—å£å¤±è´¥ï¼š{e}")

    time.sleep(0.2)

    # â‘£ ç§»åŠ¨åˆ°æŒ‡å®šä½ç½®
    try:
        win.moveTo(x, y)
    except Exception as e:
        print(f"[è­¦å‘Š] ç§»åŠ¨çª—å£å¤±è´¥ï¼š{e}")

    time.sleep(0.2)

    # â‘¤ è°ƒæ•´çª—å£å¤§å°
    sw, sh = pyautogui.size()
    # é™åˆ¶æ¯”ä¾‹èŒƒå›´
    wr = max(0.01, min(1.0, width_ratio))
    hr = max(0.01, min(1.0, height_ratio))
    new_w = int(sw * wr)
    new_h = int(sh * hr)
    try:
        win.resizeTo(new_w, new_h)
        time.sleep(0.2)
    except Exception as e:
        print(f"[è­¦å‘Š] è°ƒæ•´çª—å£å¤§å°å¤±è´¥ï¼š{e}")

    print(f"[OK] å·²å°†çª—å£â€œ{win.title}â€ç§»åŠ¨åˆ° ({x},{y})ï¼Œå¹¶è°ƒæ•´ä¸º {new_w}Ã—{new_h}ï¼ˆå å±å¹• {wr*100:.0f}% Ã— {hr*100:.0f}%ï¼‰")

    return True


"""
restore_and_position_cad(
    "å¾®ä¿¡",
    width_ratio = 0.5,
    height_ratio = 0.5,
    x= 0,
    y= 0
)

"""

def list_open_window_titles() -> list[str]:
    """
    è·å–å½“å‰æ‰€æœ‰å¯è§çª—å£çš„æ ‡é¢˜åˆ—è¡¨åŒ…æ‹¬å­çª—å£ã€‚

    :return: ä¸€ä¸ªå­—ç¬¦ä¸²åˆ—è¡¨ï¼Œæ¯ä¸€é¡¹éƒ½æ˜¯ä¸€ä¸ªéç©ºçª—å£æ ‡é¢˜ã€‚
    """
    titles = []
    for w in gw.getAllWindows():
        title = w.title.strip()
        if title:
            titles.append(title)
    return titles
#&&% * æµ‹é¼ æ ‡ä½ç½®
def ceshubiao_weizhi():
    """
    æç¤ºç”¨æˆ· 5 ç§’å†…å°†é¼ æ ‡ç§»åŠ¨åˆ° AutoCAD å‘½ä»¤æ è¾“å…¥ä½ç½®ï¼Œ
    ç„¶åé‡‡é›†å½“å‰é¼ æ ‡åæ ‡å¹¶è¿”å› (x, y)ã€‚
    """
    print("è¯·åœ¨ 5 ç§’é’Ÿå†…ï¼Œå°†é¼ æ ‡ç²¾ç¡®åœ°æ”¾åœ¨ AutoCAD å‘½ä»¤æ çš„è¾“å…¥ä½ç½®â€¦")
    time.sleep(5)
    x, y = pyautogui.position()
    print(f"å·²è·å–åæ ‡ï¼š({x}, {y})")
    return x, y

def run_idle_background(script_path: str):
    """
    ç”¨åå°æ¨¡å¼å¯åŠ¨ IDLE å»è¿è¡ŒæŸä¸ªè„šæœ¬ï¼Œè¿”å› Popen å®ä¾‹ã€‚
    """
    # Windows ä¸Šéšè—çª—å£çš„ flag
    CREATE_NO_WINDOW = 0x08000000
    proc = subprocess.Popen([
            sys.executable, "-m", "idlelib", "-r", script_path
        ],
        creationflags=CREATE_NO_WINDOW,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
    )
    return proc



"""
idle_proc = run_idle_background(r"D:/Myprogramsystem/cad/CADåŸºæœ¬æ“ä½œ.py")
# ç»“æŸåˆšæ‰å¯åŠ¨çš„ IDLE
idle_proc.terminate()
idle_proc.wait(timeout=5)
"""

#æŒ‰ä¸‹é¼ æ ‡å·¦é”®å†æ‹–åŠ¨

def click_and_drag(x: int, y: int, juli: int):
    """
    åœ¨å±å¹•åæ ‡ (x, y) æŒ‰ä¸‹å·¦é”®ï¼Œç„¶åå‘çºµå‘æ‹–åŠ¨è·ç¦» juliã€‚
    è‹¥ juli ä¸ºæ­£å€¼ï¼Œåˆ™å‘ä¸Šæ‹–åŠ¨ï¼›è‹¥ä¸ºè´Ÿå€¼ï¼Œåˆ™å‘ä¸‹æ‹–åŠ¨ã€‚
    """
    # 1. ç§»åŠ¨åˆ° (x, y)
    pyautogui.moveTo(x, y)

    time.sleep(2) 

    # 2. æŒ‰ä½å·¦é”®
    pyautogui.mouseDown(button='left')

    # 3. æ‹–åŠ¨ï¼šæ‹–åŠ¨åˆ° (x, y + juli)
    #    å¦‚æœä»…éœ€ç›¸å¯¹æ‹–åŠ¨ï¼Œä¹Ÿå¯ç”¨ dragRel(0, juli)ã€‚
    dest_x = x
    dest_y = y - juli  # æ³¨æ„ï¼šå±å¹•åæ ‡é‡Œï¼Œå‘ä¸Šæ˜¯ y å‡å°ï¼Œå‘ä¸‹æ˜¯ y å¢å¤§
    pyautogui.moveTo(dest_x, dest_y, duration=0.2)

    # 4. æ¾å¼€å·¦é”®
    pyautogui.mouseUp(button='left')

#å¯»æ‰¾æŒ‡å®šå›¾ç‰‡å½¢çŠ¶
def click_and_find_image_shape(x: int, y: int, tupian_path: str, timeout: float = 10.0):
    """
    åœ¨ (x, y) å•å‡»ä¸€æ¬¡ï¼Œç„¶åä¸æ–­åœ¨æ•´ä¸ªå±å¹•ä¸ŠæŸ¥æ‰¾ä¸ tupian_path å¯¹åº”çš„å›¾ç‰‡å½¢çŠ¶ï¼Œ
    ä¸€æ—¦æ‰¾åˆ°ï¼Œå°±æŠŠé¼ æ ‡ç§»åˆ°å®ƒçš„ä¸­å¿ƒå¹¶è¿”å›ä¸­å¿ƒåæ ‡ (Python int)ï¼›è¶…æ—¶ä»æœªæ‰¾åˆ°åˆ™è¿”å› Noneã€‚

    :param x, y:        å•å‡»çš„åˆå§‹åæ ‡ï¼ˆå¯è§¦å‘ç•Œé¢æ›´æ–°ï¼‰
    :param tupian_path: è¦è¯†åˆ«çš„ç›®æ ‡å›¾ç‰‡è·¯å¾„ï¼ˆå¦‚å¾®ä¿¡ç¬‘è„¸ï¼‰
    :param timeout:     æœ€é•¿ç­‰å¾…æ—¶é—´ï¼ˆç§’ï¼‰
    :return:            æ‰¾åˆ°æ—¶è¿”å› (x, y)ï¼›å¦åˆ™è¿”å› None
    """
    if not os.path.isfile(tupian_path):
        raise FileNotFoundError(f"æ‰¾ä¸åˆ°å›¾ç‰‡æ–‡ä»¶ï¼š{tupian_path}")

    # 1. å³é”®ç‚¹å‡» (x, y)
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.3)  # ç­‰å¾…ç•Œé¢åˆ·æ–°

    start = time.time()
    while True:
        # 2. æœç´¢å›¾ç‰‡å¹¶è·å–ä¸­å¿ƒåæ ‡
        loc = pyautogui.locateCenterOnScreen(tupian_path, confidence=0.9)
        if loc:
            cx, cy = loc
            # è½¬ä¸º Python int å†ç§»åŠ¨
            cx, cy = int(cx), int(cy)
            pyautogui.moveTo(cx, cy)
            return cx, cy

        # 3. æ£€æŸ¥è¶…æ—¶
        if time.time() - start > timeout:
            print("[é”™è¯¯] è¶…æ—¶ï¼Œæœªèƒ½è¯†åˆ«åˆ°æŒ‡å®šå›¾ç‰‡å½¢çŠ¶ã€‚")
            return None

        time.sleep(0.2)

#å³é”®èœå•
def right_click_and_move(x: int, y: int, x_xiangdui: int, y_xiangdui: int):
    """
    åœ¨å±å¹•åæ ‡ (x, y) å¤„æ‰§è¡Œå³é”®ç‚¹å‡»ï¼Œç„¶åå°†é¼ æ ‡ç›¸å¯¹äº (x, y) 
    ç§»åŠ¨åˆ° (x + x_xiangdui, y + y_xiangdui) å¹¶åœç•™ã€‚

    :param x:            åˆå§‹æ°´å¹³åæ ‡
    :param y:            åˆå§‹å‚ç›´åæ ‡
    :param x_xiangdui:   ç›¸å¯¹æ°´å¹³åç§»ï¼ˆæ­£å€¼å‘å³ï¼Œè´Ÿå€¼å‘å·¦ï¼‰
    :param y_xiangdui:   ç›¸å¯¹å‚ç›´åç§»ï¼ˆæ­£å€¼å‘ä¸‹ï¼Œè´Ÿå€¼å‘ä¸Šï¼‰
    """
    # ç§»åŠ¨åˆ°ç›®æ ‡ä½ç½®å¹¶å³é”®ç‚¹å‡»
    pyautogui.moveTo(x, y)
    pyautogui.click(button='right')
    time.sleep(2)
    # è®¡ç®—ç›¸å¯¹ç›®æ ‡ä½ç½®
    dest_x = x + x_xiangdui
    dest_y = y + y_xiangdui
    # å¹³æ»‘ç§»åŠ¨åˆ°æ–°ä½ç½®
    pyautogui.moveTo(dest_x, dest_y, duration=0.2)

def kill_all_idle():
    """
    ç»ˆæ­¢æ‰€æœ‰åä¸º 'idle' æˆ– 'idle.exe' çš„è¿›ç¨‹ï¼ˆä¸å†éœ€è¦ä»»åŠ¡ç®¡ç†å™¨ï¼‰ã€‚
    """
    for p in psutil.process_iter(("pid", "name")):
        name = (p.info["name"] or "").lower()
        if name.startswith("idle"):
            try:
                p.terminate()
            except Exception:
                pass






##æ§åˆ¶IDLEå±å¹•çª—å£åœ¨å³ä¸Šè§’
def set_idle_window_to_top_right():

    # è·å–IDLEçª—å£å¥æŸ„
    windows = [w for w in gw.getWindowsWithTitle('IDLE') if w.visible]
    
    if not windows:
        print("[é”™è¯¯] æœªæ‰¾åˆ°IDLEçª—å£ï¼")
        return
    
    # è·å–IDLEä¸»çª—å£
    win = windows[0]
    
    # è·å–å±å¹•åˆ†è¾¨ç‡
    screen_width, screen_height = pyautogui.size()

    # è®¡ç®—çª—å£å¤§å°ï¼šæ¨ªç«–å„å 1/2
    half_width = screen_width // 2
    half_height = screen_height // 2

    # å°†çª—å£ç§»åŠ¨åˆ°å±å¹•çš„å³ä¸Šè§’ (å±å¹•å®½åº¦çš„ä¸€åŠï¼Œé¡¶éƒ¨ä¸º0)
    win.moveTo(half_width, 0)

    # è°ƒæ•´çª—å£å¤§å°ï¼Œè®¾ç½®ä¸ºå±å¹•çš„1/2å®½åº¦å’Œ1/2é«˜åº¦
    win.resizeTo(half_width, half_height)
    
    print(f"[OK] IDLEçª—å£å·²è°ƒæ•´åˆ°å±å¹•å³ä¸Šè§’ï¼Œå°ºå¯¸ä¸º {half_width} x {half_height}")

def r():

    set_idle_window_to_top_right()

##æ§åˆ¶OBSçª—å£åœ¨å³ä¸‹è§’
def place_obs_bottom_right():
    """
    å°† OBS Studio ä¸»çª—å£ç§»åŠ¨åˆ°å±å¹•å³ä¸‹è§’ï¼Œå¹¶ç¼©æ”¾ä¸ºå±å¹•å®½é«˜çš„ä¸€åŠã€‚
    - è‹¥æ‰¾ä¸åˆ° OBS çª—å£ï¼Œä¼šæ‰“å°é”™è¯¯ä¿¡æ¯ã€‚
    - è‹¥æœ‰å¤šä¸ª OBS çª—å£ï¼Œä»…æ“ä½œç¬¬ä¸€ä¸ªå¯è§çª—å£ã€‚
    """
    # 1ï¸âƒ£ è·å– OBS ä¸»çª—å£
    obs_windows = [w for w in gw.getWindowsWithTitle('OBS') if w.visible]
    if not obs_windows:
        print('[é”™è¯¯] æœªæ‰¾åˆ°å¯è§çš„ OBS Studio çª—å£')
        return

    obs = obs_windows[0]            # å–ç¬¬ä¸€ä¸ª
    print(f'ğŸ” æ‰¾åˆ°çª—å£: {obs.title}')

    # 2ï¸âƒ£ è®¡ç®—ç›®æ ‡å°ºå¯¸ä¸ä½ç½® â€”â€” å³ä¸‹è§’ 1/4 åŒºåŸŸ
    screen_w, screen_h = pyautogui.size()
    half_w, half_h = screen_w // 2, screen_h // 2
    target_left = screen_w - half_w      # å³åŠå±èµ·ç‚¹
    target_top  = screen_h - half_h      # ä¸‹åŠå±èµ·ç‚¹

    # 3ï¸âƒ£ ç§»åŠ¨å¹¶ç¼©æ”¾
    obs.moveTo(target_left, target_top)
    time.sleep(0.1)                      # ç»™ç³»ç»Ÿä¸€ç‚¹ç¼“å†²
    obs.resizeTo(half_w, half_h)

    print(f'[OK] å·²å°† OBS çª—å£è°ƒæ•´åˆ°å³ä¸‹è§’ {half_w}Ã—{half_h}')

def r2():

    place_obs_bottom_right()




##æœ€å°ã€æœ€å¤§åŒ–çª—å£
def minimize_window(window_keyword: str = 'OBS') -> bool:
    """
    é€šç”¨ï¼šæœ€å°åŒ–ç¬¬ä¸€ä¸ªæ ‡é¢˜åŒ…å« window_keyword çš„å¯è§çª—å£ã€‚

    :param window_keyword: è¦åŒ¹é…çš„çª—å£æ ‡é¢˜å…³é”®å­—ï¼ˆå­ä¸²åŒ¹é…ï¼‰ï¼Œé»˜è®¤ 'OBS'
    :return: å¦‚æœæˆåŠŸæœ€å°åŒ–è¿”å› Trueï¼Œå¦åˆ™è¿”å› False
    """
    # 1) æ‰¾åˆ°æ‰€æœ‰åŒ¹é…çš„å¯è§çª—å£
    windows = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not windows:
        print(f'[é”™è¯¯] æœªæ‰¾åˆ°æ ‡é¢˜åŒ…å« "{window_keyword}" çš„å¯è§çª—å£')
        return False

    # 2) å–ç¬¬ä¸€ä¸ªå¹¶æœ€å°åŒ–
    win = windows[0]
    print(f'ğŸ” æ‰¾åˆ°çª—å£: "{win.title}"ï¼Œæ‰§è¡Œæœ€å°åŒ–')
    win.minimize()
    return True

def maximize_autocad_window(window_keyword: str = 'AutoCAD') -> bool:
    """
    å¼ºåˆ¶æœ€å¤§åŒ–ç¬¬ä¸€ä¸ªæ ‡é¢˜åŒ…å« window_keyword çš„å¯è§çª—å£ã€‚
    ä¼˜å…ˆå°è¯•ä½¿ç”¨ win32guiï¼Œå¦‚ä¸å¯ç”¨åˆ™é€€å› ctypes è°ƒç”¨ user32ã€‚
    """
    # 1) æ‰¾åˆ°ç›®æ ‡çª—å£
    wins = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not wins:
        print(f"[é”™è¯¯] æœªæ‰¾åˆ°æ ‡é¢˜åŒ…å« â€œ{window_keyword}â€ çš„å¯è§çª—å£")
        return False

    win = wins[0]
    hWnd = win._hWnd

    # 2) å…ˆæ¢å¤ï¼ˆé¿å…æœ€å°åŒ–çŠ¶æ€ï¼‰ï¼Œå†æœ€å¤§åŒ–
    #    å°è¯•ä½¿ç”¨ win32gui
    try:
        import win32gui, win32con
        win32gui.ShowWindow(hWnd, win32con.SW_RESTORE)
        time.sleep(0.1)
        win32gui.ShowWindow(hWnd, win32con.SW_MAXIMIZE)
    except ImportError:
        # å¦‚æœæ²¡æœ‰ pywin32ï¼Œå°±é€€å› ctypes
        SW_RESTORE  = 9
        SW_MAXIMIZE = 3
        ctypes.windll.user32.ShowWindow(hWnd, SW_RESTORE)
        time.sleep(0.1)
        ctypes.windll.user32.ShowWindow(hWnd, SW_MAXIMIZE)

    time.sleep(0.2)  # ç¡®ä¿çª—å£å®Œæˆæœ€å¤§åŒ–
    print(f"[OK] å·²å°†çª—å£ â€œ{win.title}â€ æœ€å¤§åŒ–")
    return True

##æ§åˆ¶å±å¹•ä½ç½®å¼€å§‹å½•åˆ¶æˆ–å…³é—­OBS
def start_obs_recording_by_click(x: int = 1768, y: int = 872,
                                 button: str = 'left',
                                 clicks: int = 1,
                                 move_duration: float = 0.2):
    """
    é€šè¿‡é¼ æ ‡ç‚¹å‡»å±å¹•ä¸Š (x,y) åæ ‡æ¥æ§åˆ¶ OBS å¼€å§‹/åœæ­¢å½•åˆ¶ã€‚
    
    :param x: ç›®æ ‡ç‚¹å‡»ç‚¹ X åæ ‡
    :param y: ç›®æ ‡ç‚¹å‡»ç‚¹ Y åæ ‡
    :param button: é¼ æ ‡æŒ‰é’®ï¼Œ'left'ã€'right' æˆ– 'middle'
    :param clicks: ç‚¹å‡»æ¬¡æ•°ï¼Œé»˜è®¤ä¸ºå•å‡»
    :param move_duration: ä»å½“å‰ä½ç½®ç§»åŠ¨åˆ°ç›®æ ‡çš„è€—æ—¶ï¼ˆç§’ï¼‰
    """
    # å®‰å…¨è®¾ç½®ï¼ˆå¯é€‰ï¼‰
    pyautogui.FAILSAFE = True
    pyautogui.PAUSE    = 0.1

    # 1) å¹³æ»‘ç§»åŠ¨åˆ°ç›®æ ‡
    pyautogui.moveTo(x, y, duration=move_duration)
    # 2) ç‚¹å‡»
    pyautogui.click(x=x, y=y, clicks=clicks, button=button)
    print(f"[OK] å·²ç‚¹å‡» ({x}, {y})ï¼Œè¯·æ£€æŸ¥ OBS æ˜¯å¦å·²å¼€å§‹/åœæ­¢å½•åˆ¶ã€‚")


#&&&% *** å½•å±
def fs(x1,y1):
    """
    å¾®ä¿¡è°ƒåˆ°0.5çª—å£

    """
    
    pyautogui.moveTo(x1+595,y1+190)
    pyautogui.click(x1+595,y1+190)
    pyautogui.press("enter")

def xuanqun(x1,y1,neirong):
    
    copy_to_clipboard(neirong)

    pyautogui.moveTo(x1+128,y1+33) 
    pyautogui.click(x1+128,y1+33)   
    time.sleep(2)
    activate_window_by_title("å¾®ä¿¡")

    time.sleep(2)

    pyautogui.moveTo(x1+128,y1+33)

    activate_window_by_title("å¾®ä¿¡")

    time.sleep(2)

    pyautogui.click(x1+128,y1+33)
  

    pyautogui.hotkey('ctrl', 'v') 


    pyautogui.moveTo(x1+158,y1+49) 
    pyautogui.click(x1+158,y1+49)   
 

    pyautogui.press("enter")


def copy_to_clipboard(text: str):
    """
    å°†ä¼ å…¥çš„ text æ–‡æœ¬å†™å…¥ç³»ç»Ÿå‰ªè´´æ¿ï¼Œä¾›åç»­å³é”®â†’ç²˜è´´ä½¿ç”¨ã€‚
    
    """
    
    # åˆ›å»ºä¸€ä¸ªéšè—çš„ tk æ ¹çª—å£
    r = tkinter.Tk()
    r.withdraw()  # éšè—ä¸»çª—å£

    # æ¸…ç©ºå‰ªè´´æ¿å¹¶å†™å…¥æ–°çš„æ–‡æœ¬
    r.clipboard_clear()
    r.clipboard_append(text)
    # å¿…é¡» update() ä¸€ä¸‹ï¼Œç¡®ä¿æ•°æ®çœŸæ­£å­˜åˆ°å‰ªè´´æ¿
    r.update()

    # é”€æ¯éšè—çª—å£
    r.destroy()

def xieweixin(x1,y1,neirong):
        
        copy_to_clipboard(neirong)

        time.sleep(2)
   
        pyautogui.moveTo(x1+532,y1+151)

        time.sleep(2)
        activate_window_by_title("å¾®ä¿¡")

        
        pyautogui.moveTo(x1+532,y1+151)

        time.sleep(2)

        pyautogui.click(x1+532,y1+151)
      

        pyautogui.hotkey('ctrl', 'v')        

        pyautogui.click(x1+532,y1+151)
    
def ä¸»æ“ä½œå‡½æ•°():
    restore_and_position(
        name = "å¾®ä¿¡",
        width_ratio = 0.5,
        height_ratio = 0.5,
        x = 0,
        y = 0
    )  
 
    time.sleep(1)


    x1,y1,_,_ = activate_window_by_title("å¾®ä¿¡")

    xuanqun(x1,y1,"åæ–°å·¥ä½œç¾¤")

    time.sleep(3)

    x2,y2=click_and_find_image_shape(358, 594, r"D:/Myprogramsystem/XT/weixin_xiaolian.png", timeout = 10.0)

    time.sleep(5)

    pyautogui.click(x2,y2) 

    time.sleep(5)

    x3,y3=click_and_find_image_shape(358, 594, r"D:/Myprogramsystem/XT/weixin_daixao_biaoqingbao.png", timeout = 10.0)

    time.sleep(2)
    pyautogui.click(x3,y3) 
    time.sleep(2)


 
    xieweixin(x1,y1,"Hello!æˆ‘æ˜¯å…¬å¸ç®¡ç†å‘˜å°åŒ–èº«ï¼Œä»ä»Šå¤©èµ·å°†å‚ä¸å…¬å¸ç®¡ç†ï¼Œä¸å„ä½ä¸€èµ·å¥‹è¿›ï¼")

    fs(x1,y1)

    
    time.sleep(1)


def  main_func(folder_path=r"D:/Myprogramsystem/BaiduSyncdisk/å®‹å²³/å·¥ä¸šå›­æ•´ç†/ä¸‰æœŸ/æµ‹è¯•"):

    æ‰“å°è¾“å‡ºPDF() 




def luping_1(main_func):
    """
    1) å¼€å§‹ OBS å½•åˆ¶
    2) æœ€å°åŒ–æ‰€æœ‰çª—å£ï¼›æ¢å¤å¹¶å®šä½ OBS
    3) è°ƒç”¨ä¼ å…¥çš„ main_func() æ‰§è¡Œä¸»æ“ä½œ
    4) æ¢å¤ OBSï¼Œåœæ­¢å½•åˆ¶

    main_func: é›¶å‚æ•°å‡½æ•°ï¼Œè´Ÿè´£æ‰§è¡Œæ’å…¥æˆ–å…¶ä»–ä¸»æ“ä½œ
    """
    # æœ€å°åŒ–æ‰€æœ‰çª—å£
    minimize_all_windows_d()

    # æ¢å¤å¹¶å®šä½ OBS çª—å£
    restore_and_position(
        name="OBS",
        width_ratio=0.5,
        height_ratio=0.5,
        x=0,
        y=0
    )
    time.sleep(1)

    # ç‚¹å‡» OBS â€œå¼€å§‹å½•åˆ¶â€æŒ‰é’®ï¼ˆç›¸å¯¹åæ ‡ï¼‰
    x0, y0, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x0 + 824, y0 + 328)
    time.sleep(1)
    pyautogui.click(x0 + 824, y0 + 328)

    # æœ€å°åŒ–æ‰€æœ‰çª—å£ï¼Œåˆ‡æ¢åˆ° CAD
    minimize_all_windows_d()

    # 3ï¸âƒ£ æ‰§è¡Œä¸»æ“ä½œ

    time.sleep(1)
    restore_and_position(
        name="AutoCAD",
        width_ratio=1,
        height_ratio=1,
        x=0,
        y=0
    )
 
    
    main_func()

    print("ä¸»æ“ä½œå‡½æ•°æ‰§è¡Œå®Œæ¯•")

    # 4ï¸âƒ£ æ¢å¤å¹¶æ¿€æ´» OBSï¼Œåœæ­¢å½•åˆ¶
    time.sleep(0.5)
    minimize_all_windows_d()
    restore_and_position(
        name="OBS",
        width_ratio=0.5,
        height_ratio=0.5,
        x=0,
        y=0
    )
    time.sleep(0.2)

    x4, y4, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x4 + 824, y4 + 328)
    time.sleep(1)
    pyautogui.click(x4 + 824, y4 + 328)

    print("å½•åˆ¶å·²åœæ­¢")


def luping(main_func, *args, **kwargs):
    """
    1) å¼€å§‹ OBS å½•åˆ¶  
    2) æœ€å°åŒ–æ‰€æœ‰çª—å£ï¼›æ¢å¤å¹¶å®šä½ OBS  
    3) åˆ‡æ¢åˆ° CAD çª—å£ï¼Œè°ƒç”¨ main_func(*args, **kwargsï¼‰  
    4) åˆ‡æ¢å› OBSï¼Œåœæ­¢å½•åˆ¶  

    main_func : ä»»æ„å¯è°ƒç”¨å¯¹è±¡  
    *args, **kwargs : ä¼šåŸæ ·ä¼ ç»™ main_func  
    """
    # â€”â€” æœ€å°åŒ–æ‰€æœ‰çª—å£ â€”â€”  
    minimize_all_windows_d()

    # â€”â€” å®šä½å¹¶æ¿€æ´» OBS â€”â€”  
    restore_and_position(name="OBS", width_ratio=0.5, height_ratio=0.5, x=0, y=0)
    time.sleep(1)

    # ç‚¹å‡» OBS â€œå¼€å§‹å½•åˆ¶â€  
    x0, y0, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x0 + 824, y0 + 328)
    time.sleep(0.2)
    pyautogui.click(x0 + 824, y0 + 328)

    # â€”â€” åˆ‡æ¢åˆ° CAD â€”â€”  
    minimize_all_windows_d()
    time.sleep(0.5)
    restore_and_position(name="AutoCAD", width_ratio=1.0, height_ratio=1.0, x=0, y=0)
    time.sleep(1)

    # â€”â€” 3ï¸âƒ£ æ‰§è¡Œä¸»æ“ä½œ â€”â€”  
    main_func(*args, **kwargs)
    print("ä¸»æ“ä½œå‡½æ•°æ‰§è¡Œå®Œæ¯•")

    # â€”â€” åœæ­¢å½•åˆ¶ â€”â€”  
    time.sleep(0.5)
    minimize_all_windows_d()
    restore_and_position(name="OBS", width_ratio=0.5, height_ratio=0.5, x=0, y=0)
    time.sleep(0.2)

    x4, y4, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x4 + 824, y4 + 328)
    time.sleep(0.2)
    pyautogui.click(x4 + 824, y4 + 328)

    print("å½•åˆ¶å·²åœæ­¢")













def mccs():
    
    dw=DWG()

    MC_com =[]

    shitu_entity(kuang)

    for x in LD:

        Ent=dw.insert_door(x,width=600,btn_x=351, btn_y=129, cmd_x=471, cmd_y=455)    

        MC_com.append(Ent)
        
    time.sleep(10)

    for yuan in LByuan:
        
        for x in MC_com:

            transfer_props_by_matchprop(yuan, x, delay=0.5)

            shitu_entity(x)

            time.sleep(2)

        name = (yuan.Layer)[3:]

        print("************************************************************")

        print(f"**************       æ›¿æ¢é—¨æˆ{name} æ„ä»¶       *************")

        print("************************************************************")


        shitu_entity(kuang)        

        time.sleep(10)


def é­”æ–¹():

    import é­”æ–¹åˆ†æ

    é­”æ–¹åˆ†æ.é­”æ–¹æ§åˆ¶å°(r"D:/Myprogramsystem/BaiduSyncdisk/å®‹å²³/è‡ªåŠ¨åŒ–(åŠ¨æ€)/é­”æ–¹/åˆ†æ")



#&&%#è¿è¡ŒæŒ‡å®šç¨‹åºåçš„ç¨‹åº

def run_py(pyname):
    try:
        # è¿è¡ŒæŒ‡å®šçš„ Python ç¨‹åºï¼Œéšè—å‘½ä»¤è¡Œçª—å£
        result = subprocess.run(
            [sys.executable, pyname],
            check=True,
            text=True,
            capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        print(f"[OK] ç¨‹åº {pyname} æ‰§è¡ŒæˆåŠŸã€‚è¾“å‡º:\n{result.stdout}")
    except subprocess.CalledProcessError as e:
        print(f"[é”™è¯¯] è¿è¡Œ {pyname} æ—¶å‘ç”Ÿé”™è¯¯: {e}")
        print(f"é”™è¯¯ä¿¡æ¯: {e.stderr}")
    except FileNotFoundError:
        print(f"[é”™è¯¯] æœªæ‰¾åˆ°ç¨‹åº {pyname}ã€‚è¯·æ£€æŸ¥æ–‡ä»¶åå’Œè·¯å¾„ã€‚")




##æŠŠé¼ æ ‡ç§»åŠ¨åˆ°å‘½ä»¤è¡Œçª—å£
def focus_cmdline(cmd_x, cmd_y, delay=0.2):
    """
    æŠŠé¼ æ ‡ç§»åˆ°å‘½ä»¤è¡Œå¹¶å•å‡»ï¼Œç¡®ä¿ç„¦ç‚¹å›åˆ° AutoCAD å‘½ä»¤æ ã€‚

    
    """
    pyautogui.moveTo(cmd_x, cmd_y, duration=delay)
    pyautogui.click()



#&&% æ¿€æ´»çª—å£å’Œå­çª—å£

def activate_window_by_title(title: str, click_titlebar: bool = True) -> bool:
    """
    æ¿€æ´»ä¸€ä¸ªæŒ‡å®šæ ‡é¢˜çš„çª—å£ã€‚
    
    1. ç”¨ pygetwindow æ‰¾åˆ°çª—å£å¯¹è±¡ï¼›
    2. å¦‚æœçª—å£è¢«æœ€å°åŒ–ï¼Œå…ˆè¿˜åŸï¼›
    3. å°†çª—å£ç½®ä¸ºå‰å°ï¼ˆä¼˜å…ˆä½¿ç”¨ win.activate()ï¼Œå¤±è´¥æ—¶ç”¨ ctypes å¼ºåˆ¶ï¼‰ï¼›
    4. ï¼ˆå¯é€‰ï¼‰åœ¨æ ‡é¢˜æ ä¸­ç‚¹å‡»ä¸€æ¬¡ï¼Œç¡®ä¿ç„¦ç‚¹ã€‚
    
    :param title: è¦æ¿€æ´»çš„çª—å£æ ‡é¢˜ï¼ˆéƒ¨åˆ†åŒ¹é…ï¼‰ã€‚
    :param click_titlebar: æ˜¯å¦æ¨¡æ‹Ÿä¸€æ¬¡ç‚¹å‡»æ ‡é¢˜æ ï¼Œä»¥ç¡®ä¿çª—å£è·å¾—ç„¦ç‚¹ã€‚
    :return: True=æ¿€æ´»æˆåŠŸï¼ŒFalse=æœªæ‰¾åˆ°çª—å£ã€‚
    """
    USER32 = ctypes.windll.user32
    SW_RESTORE = 9


    # â‘  æŸ¥æ‰¾æ ‡é¢˜åŒ…å«å…³é”®å­—çš„çª—å£
    wins = [w for w in gw.getWindowsWithTitle(title) if title in w.title]
    if not wins:
        print(f"[é”™è¯¯] æœªæ‰¾åˆ°æ ‡é¢˜åŒ…å« â€œ{title}â€ çš„çª—å£")
        return False
    win = wins[0]

    # â‘¡ å¦‚æœæœ€å°åŒ–ï¼Œå…ˆè¿˜åŸ
    if win.isMinimized:
        win.restore()
        time.sleep(0.2)

    
    # â‘¢ ç½®ä¸ºå‰å°ï¼šå…ˆå°è¯• pygetwindow.activate()
    try:
        win.activate()
        time.sleep(0.2)
    except Exception:
        # ctypes å¼ºåˆ¶è¿˜åŸå¹¶ç½®å‰
        hwnd = win._hWnd
        USER32.ShowWindow(hwnd, SW_RESTORE)          # è¿˜åŸ
        USER32.SetForegroundWindow(hwnd)             # ç½®å‰
        time.sleep(0.2)

    # â‘£ å¯é€‰ï¼šåœ¨æ ‡é¢˜æ ä¸­ç‚¹ä¸€ä¸‹
    if click_titlebar:
        x = win.left + win.width // 2
        y = win.top + max(1, min(30, win.height // 10))
        pyautogui.click(x, y)
        time.sleep(0.1)

    print(f"[æˆåŠŸ] å·²æ¿€æ´»çª—å£: {win.title} ä½ç½®({win.left},{win.top}) å¤§å°{win.width}x{win.height}")

    return  win.left,win.top,win.width,win.height


def click_in_window(title: str, offset_x: float, offset_y: float, click_titlebar: bool = False) -> bool:
    """
    åœ¨æŒ‡å®šçª—å£çš„æŸä¸ªç›¸å¯¹åƒç´ ä½ç½®ç‚¹å‡»ä¸€æ¬¡ï¼ˆç›¸å¯¹äºçª—å£å·¦ä¸Šè§’çš„åç§»é‡ï¼‰ã€‚

    :param title:         çª—å£æ ‡é¢˜å…³é”®å­—ï¼ˆéƒ¨åˆ†åŒ¹é…ï¼‰
    :param offset_x:      ä»çª—å£å·¦ä¸Šè§’ç®—èµ·çš„æ°´å¹³åƒç´ åç§»é‡ï¼ˆ0 è¡¨ç¤ºå·¦è¾¹ç¼˜ï¼Œwidth è¡¨ç¤ºå³è¾¹ç¼˜ï¼‰
    :param offset_y:      ä»çª—å£å·¦ä¸Šè§’ç®—èµ·çš„å‚ç›´åƒç´ åç§»é‡ï¼ˆ0 è¡¨ç¤ºé¡¶è¾¹ç¼˜ï¼Œheight è¡¨ç¤ºåº•è¾¹ç¼˜ï¼‰
    :param click_titlebar: æ˜¯å¦å…ˆåœ¨æ ‡é¢˜æ ç‚¹å‡»ä¸€æ¬¡ä»¥ç¡®ä¿çª—å£è·å–ç„¦ç‚¹
    :return:              True=ç‚¹å‡»æˆåŠŸï¼ŒFalse=æœªæ‰¾åˆ°çª—å£
    """
    # 1) æŸ¥æ‰¾çª—å£
    wins = [w for w in gw.getWindowsWithTitle(title) if title in w.title]
    if not wins:
        print(f"[é”™è¯¯] æœªæ‰¾åˆ°æ ‡é¢˜åŒ…å« â€œ{title}â€ çš„çª—å£")
        return False
    win = wins[0]

    # 2) å¦‚æœæœ€å°åŒ–ï¼Œå°±å…ˆè¿˜åŸ
    if win.isMinimized:
        win.restore()
        time.sleep(0.2)

    # 3) å°è¯•æ¿€æ´»
    try:
        win.activate()
        time.sleep(0.2)
    except Exception:
        # ctypes å¼ºåˆ¶è¿˜åŸå¹¶ç½®å‰
        hwnd = win._hWnd
        SW_RESTORE = 9
        ctypes.windll.user32.ShowWindow(hwnd, SW_RESTORE)
        ctypes.windll.user32.SetForegroundWindow(hwnd)
        time.sleep(0.2)

    # 4) å¯é€‰ï¼šç‚¹å‡»æ ‡é¢˜æ è®©ç„¦ç‚¹çœŸæ­£åˆ°çª—å£
    if click_titlebar:
        tx = win.left + win.width // 2
        ty = win.top + 10  # æ ‡é¢˜æ ä¸€èˆ¬åœ¨çª—å£é¡¶éƒ¨ 10px å·¦å³
        pyautogui.click(tx, ty)
        time.sleep(0.1)

    # 5) è®¡ç®—ç›®æ ‡ç‚¹çš„ç»å¯¹å±å¹•åæ ‡
    abs_x = int(win.left + offset_x)
    abs_y = int(win.top  + offset_y)

    # 6) ç‚¹å‡»
    pyautogui.moveTo(abs_x, abs_y, duration=0.2)
    pyautogui.click(abs_x, abs_y)
    time.sleep(0.1)
    print(f"ğŸ”˜ å·²åœ¨çª—å£ â€œ{win.title}â€ å†…éƒ¨ç‚¹å‡» ({offset_x}, {offset_y}) â†’ ç»å¯¹ ({abs_x}, {abs_y})")
    return True

"""
click_in_window("å›¾å½¢å¯¼å‡º", offset_x=600-10, offset_y=600-10, click_titlebar=True)

"""

#&&% é“¾æ¥å’Œå…³é—­è‰¾å¯äº‘

def activate_and_click_aikeyun():


    try:
    
        left, top, width, height = activate_window_by_title("è‰¾å¯äº‘", click_titlebar=True)
        
        # Calculate relative offsets from provided data
        rel_x1 = 887 - 682  # 205
        rel_y1 = 542 - 93   # 449
        
        # Move to first click position and left click
        pyautogui.moveTo(left + rel_x1, top + rel_y1)
        pyautogui.click(button='left')
        
        # Wait for 2 seconds
        time.sleep(2)
        
        # Calculate close position: using width - 22 for x, as it aligns with top-right close button
        rel_x2 = width - 22  # Equivalent to 400 - 22 = 378 in data
        rel_y2 = 18          # 111 - 93
        
        # Move to close position and left click
        pyautogui.moveTo(left + rel_x2, top + rel_y2)
        pyautogui.click(button='left')
    
    
    
    except:
    
        pass
    
    









def drag_in_window_simple(
    title: str,
    start: tuple[float,float],
    offset: tuple[float,float],
    duration: float = 0.3,
    button: str = 'left',
    absolute_start: bool = False
):
    """
    æ‹–æ‹½å‡½æ•°ï¼Œæ”¯æŒç›¸å¯¹æˆ–ç»å¯¹èµ·ç‚¹ï¼š

    :param title: çª—å£æ ‡é¢˜å…³é”®å­—
    :param start: èµ·ç‚¹åæ ‡ï¼Œ(x, y)ï¼›
                  å¦‚æœ absolute_start==Falseï¼Œåˆ™å½“æˆâ€œçª—å£å†…éƒ¨â€åæ ‡
                  å¦‚æœ absolute_start==Trueï¼Œåˆ™å½“æˆâ€œå±å¹•â€åæ ‡
    :param offset: æ‹–æ‹½å‘é‡ (dx, dy)
    :param duration: æ‹–åŠ¨æ—¶é•¿
    :param button: 'left' æˆ– 'right'
    :param absolute_start: True åˆ™ start è§†ä¸ºå±å¹•ç»å¯¹åæ ‡
    """
    left, top, w, h = activate_window_by_title(title)

    if absolute_start:
        x1, y1 = start
    else:
        x1 = left + start[0]
        y1 = top  + start[1]

    x2 = x1 + offset[0]
    y2 = y1 + offset[1]

    pyautogui.moveTo(x1, y1)
    pyautogui.mouseDown(button=button)
    pyautogui.moveTo(x2, y2, duration=duration)
    pyautogui.mouseUp(button=button)
    time.sleep(0.1)
    print(f"å·²åœ¨â€œ{title}â€çª—å£æ‹–æ‹½ï¼šèµ·ç‚¹{(x1,y1)} â†’ ç»ˆç‚¹{(x2,y2)}")


"""
drag_in_window_simple(
    "å›¾å½¢å¯¼å‡º",
    start=(10,10),
    offset=(100,50),
    absolute_start=False
)

"""

#&&% å±å¹•æˆªå›¾å¿«ç…§

"""
img = pyautogui.screenshot()

img.save('D:/Myprogramsystem/BaiduSyncdisk/å®‹å²³/å·¥ä¸šå›­æ•´ç†/ä¸‰æœŸ/CADæ‰“å°/01å»ºç­‘/screenshot.png')

å±€éƒ¨

# region=(left, top, width, height)
img = pyautogui.screenshot(region=(0, 0, 800, 600))

left=0 è¡¨ç¤ºä»å±å¹•æœ€å·¦è¾¹å¼€å§‹
top=0 è¡¨ç¤ºä»å±å¹•æœ€ä¸Šè¾¹å¼€å§‹
æˆªå–ä»å·¦ä¸Šè§’å¾€å³ 800 åƒç´ ã€å¾€ä¸‹ 600 åƒç´ èŒƒå›´å†…çš„åŒºåŸŸã€‚

åº”è¯¥æ˜¯å¯¹è§’ç‚¹ æµ‹ç®—å¤šæ¬¡å¾—CADå±å¹•
img = pyautogui.screenshot(region=(175,101,2350,1250))
img.save('D:/Myprogramsystem/BaiduSyncdisk/å®‹å²³/å·¥ä¸šå›­æ•´ç†/ä¸‰æœŸ/CADæ‰“å°/01å»ºç­‘/screenshot.png')


"""


##è‡ªåŠ¨ç‚¸å¼€åŒºåŸŸå†…å¯¹è±¡
    
def run_auto_explode_area(x1, y1, x2, y2, cmd_x, cmd_y, delay=2):

    """
    è¿™æ˜¯ä¸€ä¸ªæœªä½¿ç”¨pywin32APIæ§åˆ¶å¤©æ­£CADçš„å…¸å‹å‡½æ•°
    å®ƒç‚¸å¼€åŒºåŸŸx1, y1, x2, y2çš„æ‰€æœ‰å¯¹è±¡ï¼Œä¸é€‚åˆåå¤æ“ä½œ

    åˆç†çš„å¤„ç†æ–¹å¼è¿˜æ˜¯é…åˆè§†çª—è°ƒæ•´ï¼ˆå‘é€z\naæˆ–e\nï¼‰å…ˆslectæ˜¾æ€§é€‰æ‹©å¯¹è±¡ï¼Œå†å‘é€å‘½ä»¤

    """
    
    script = Path(__file__).with_name('auto_EXPLODE.py')
    cmd = [
        sys.executable, str(script),
        f'--x1={x1}', f'--y1={y1}',
        f'--x2={x2}', f'--y2={y2}',
        f'--cmd_x={cmd_x}', f'--cmd_y={cmd_y}',
        f'--delay={delay}'
    ]
    # éšè—å­è¿›ç¨‹çª—å£
    flags = subprocess.CREATE_NO_WINDOW
    try:
        subprocess.run(cmd, check=True, creationflags=flags)
        print('â–¶ å­ç¨‹åºå®Œæˆ')
    except subprocess.CalledProcessError as e:
        print(f'[é”™è¯¯] å­ç¨‹åºé€€å‡ºç  {e.returncode}')





##ç¡®ä¿CADçª—å£è¾“å…¥æ³•ä¸ºè‹±æ–‡
"""
æœ¬å‡½æ•°ç ”ç©¶äº†ctr+ç©ºæ ¼çš„è½¬æ¢ï¼Œå¦‚éœ€è¦å¯è¿›ä¸€æ­¥åˆ†æctr+shiftè½¬æ¢çš„é—®é¢˜        

å®ƒçœŸæ­£çš„æ„ä¹‰æ˜¯æä¾›äº†ocræˆªå›¾åˆ†ææ–¹æ³•ï¼Œè¿™è®©è‡ªåŠ¨åŒ–å¤„ç†çª—å£é—®é¢˜æ›´ä¸ºå¼ºå¤§

"""
def ensure_english_input_method(cmd_x: int, cmd_y: int, delay: float = 0.2):
    """
    èšç„¦ AutoCAD å‘½ä»¤è¡Œ -> è¾“å…¥ a -> OCR åˆ¤æ–­æ˜¯å¦å‡ºç° A(ARC)/AA(AREA)ã€‚
    è‹¥æœªå‡ºç°åˆ™æ‰§è¡Œ Win+Space å¼ºåˆ¶åˆ‡æ¢è¾“å…¥æ³•ä¸€æ¬¡ã€‚

    å‚æ•°
    ----
    cmd_x, cmd_y : å‘½ä»¤è¡Œè¾“å…¥æ¡†çš„å±å¹•åæ ‡
    delay        : é¼ æ ‡ç§»åŠ¨è€—æ—¶ï¼ˆç§’ï¼‰
    """

    # ---------- è¾…åŠ© ---------- #
    def click_cmdline():
        pyautogui.moveTo(cmd_x, cmd_y, duration=delay)
        pyautogui.click()

    def set_to_english():
        pyautogui.hotkey("win", "space")       # ç³»ç»Ÿè¾“å…¥æ³•åˆ‡æ¢
        time.sleep(2.0)                        # ç»™ç³»ç»Ÿå……è¶³æ—¶é—´

    def is_english() -> bool:
        """è¾“å…¥ a æˆªå›¾ â†’ OCR â†’ å®½æ¾åŒ¹é…"""
        pyautogui.write("a")                   # æµ‹è¯•å­—ç¬¦
        time.sleep(1.2)                        # ç­‰åˆ—è¡¨å¼¹å‡º
        img = ImageGrab.grab(bbox=(310, 381, 510, 581))   # æŒ‰éœ€è°ƒæ•´
        raw = pytesseract.image_to_string(img)
        print(f"OCR æ•è·: {raw!r}")

        # å»ç©ºç™½â†’å¤§å†™ï¼Œç»Ÿä¸€åŒ¹é…
        cleaned = re.sub(r"\s+", "", raw).upper()
        return ("A(ARC)" in cleaned) or ("AA(AREA)" in cleaned)

    # ---------- ä¸»æµç¨‹ ---------- #
    click_cmdline()                # èšç„¦å‘½ä»¤æ 

    if is_english():
        print("[OK] å·²æ˜¯è‹±æ–‡è¾“å…¥æ³•ï¼Œç›´æ¥ç»“æŸã€‚")
        return

    print("[é”™è¯¯] æ£€æµ‹åˆ°éè‹±æ–‡ï¼Œæ‰§è¡Œåˆ‡æ¢ â€¦")
    set_to_english()
    click_cmdline()                # åˆ‡æ¢åé‡æ–°èšç„¦

    # åˆ‡æ¢ä¸€æ¬¡åä¸å†å¾ªç¯ï¼›å¦‚ä»å¤±è´¥å¯æ‰‹åŠ¨æ£€æŸ¥
    if is_english():
        print("[OK] åˆ‡æ¢æˆåŠŸã€‚")
    else:
        print("[è­¦å‘Š]ï¸ åˆ‡æ¢åä»æœªæ£€æµ‹åˆ°è‹±æ–‡ï¼Œå¯èƒ½ OCR å¤±æ•ˆæˆ–åæ ‡éœ€è°ƒæ•´ã€‚")



##åˆ—å‡ºæ‰€æœ‰å½“å‰çª—å£

def list_all_windows():
    # è·å–æ‰€æœ‰å¯è§çš„çª—å£
    windows = gw.getWindowsWithTitle('')
    
    if not windows:
        print("[é”™è¯¯] æ²¡æœ‰æ‰¾åˆ°ä»»ä½•å¯è§çª—å£ã€‚")
    else:
        print("å½“å‰æ¡Œé¢ä¸Šçš„æ‰€æœ‰çª—å£ï¼š")
        for win in windows:
            print(f"çª—å£æ ‡é¢˜: {win.title}, çª—å£å¤§å°: {win.width}x{win.height}, ä½ç½®: ({win.left}, {win.top})")


#å…³é—­åƒåœ¾å¹²æ‰°çª—å£
def close_360_popup_window():
    # è·å–æ‰€æœ‰çª—å£æ ‡é¢˜
    windows = gw.getWindowsWithTitle('360')  # '360' ä¸ºå¼¹å‡ºçª—å£çš„éƒ¨åˆ†æ ‡é¢˜ï¼Œå…·ä½“æ ¹æ®å®é™…æƒ…å†µè°ƒæ•´

    # éå†æ‰€æœ‰çª—å£ï¼Œæ‰¾åˆ°åŒ¹é…çš„çª—å£
    for window in windows:
        print(f"æ‰¾åˆ°çª—å£: {window.title}")

        if '360' in window.title:  # ç¡®è®¤çª—å£å±äº360è½¯ä»¶ï¼ˆä½ å¯ä»¥æ ¹æ®å®é™…çª—å£æ ‡é¢˜è°ƒæ•´ï¼‰
            # æ¿€æ´»çª—å£
            window.activate()
            time.sleep(1)

            # é€šè¿‡æ¨¡æ‹ŸæŒ‰é”®å…³é—­çª—å£ï¼ˆå¦‚æœæ˜¯å¼¹å‡ºçš„æç¤ºæ¡†ï¼Œå¸¸è§çš„å…³é—­æŒ‰é’®æ˜¯ 'ESC' æˆ– 'Alt+F4'ï¼‰
            pyautogui.hotkey('alt', 'f4')  # æ¨¡æ‹Ÿ Alt + F4 æ¥å…³é—­çª—å£
            print(f"[OK] å…³é—­äº†çª—å£: {window.title}")
            return

    print("[é”™è¯¯] æ²¡æœ‰æ‰¾åˆ°åŒ¹é…çš„360çª—å£")







#  ä¸»å‡½æ•°
#  (1)
# å¤©æ­£å¢™ä¸­è½´çº¿æ˜¾ç¤ºä¸éšè—

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

"""
å¢™æ˜¯å¦åŠ ç²—è¾¹çº¿ï¼Œå¢™ä¸­çº¿æ˜¾ç¤ºï¼Œå¢™ä¸­çº¿éšè—æ˜¯æ–‡ä»¶å±æ€§ï¼Œå› æ­¤æˆ‘ä»¬ç”¨D:/Myprogramsystem/cad/xitongjicuwenjian/å¢™åŸºçº¿æ‰“å¼€.dwgï¼Œ

D:/Myprogramsystem/cad/xitongjicuwenjian/å¢™åŸºçº¿å…³é—­.dwgï¼ŒD:/Myprogramsystem/cad/xitongjicuwenjian/å¢™çº¿åŠ ç²—.dwgä¸‰ä¸ªç©ºæ–‡ä»¶æ¥

æ§åˆ¶æ–‡ä»¶å’Œç¨‹åºçš„è¿è¡Œã€‚å› ä¸ºè·å–å¤©æ­£å¢™ä¿¡æ¯çš„å‡½æ•°ä¾èµ–å¢™ä¸­çº¿çš„æ˜¾ç¤ºï¼Œæ‰€ä»¥å¯¹ä¸ç†Ÿæ‚‰çš„æ–‡ä»¶è¿›è¡Œå¤„ç†æ—¶ï¼Œå°±å¯ä»¥é€šè¿‡è¿™å‡ ä¸ªåŸºæœ¬æ–‡ä»¶å®ç°ç¡®å®š

çš„æ§åˆ¶ï¼Œå³æˆ‘ä»¬é¢„æœŸè¦æ–‡ä»¶çš„å¢™ä¸­çº¿æ˜¾ç¤ºå‡ºæ¥ï¼Œæˆ–è€…å¢™è¾¹çº¿è¦åŠ ç²—ï¼Œä¸åŠ ç²—ï¼Œä¾èµ–å®é™…çš„éœ€è¦ã€‚




"""
## æµ‹è¯•ç¤ºä¾‹

##__________







def toggle_wall_centerline(conf=0.8, wait=2):
    """
    è‹¥ CAD åº•æ ä¸­ â€œå¢™ä¸­çº¿æ˜¾ç¤ºâ€ å›¾æ ‡ä¸ºç°è‰²åˆ™ç‚¹å‡»å¯ç”¨ï¼Œ
    å¦åˆ™ç›´æ¥æŠ¥å‘Šå·²å¼€å¯ã€‚
    """
    import pyautogui as pg

    pg.FAILSAFE = True
    pg.PAUSE    = 0.1           # å…¨å±€è½»å¾®åœé¡¿

    # æ¨¡æ¿æ–‡ä»¶
    BASE = Path(__file__).parent
    BAR_IMG  = BASE / 'bar_block.png'     # åº•æ å—
    OFF_IMG  = BASE / 'icon_off.png'      # ç°  â‰¡
    ON_IMG   = BASE / 'icon_on.png'       # è“  â‰¡


    # 1ï¸âƒ£ æ‰¾åˆ°åº•æ å¤§å—åæ ‡
    region = pg.locateOnScreen(str(BAR_IMG), confidence=conf)
    if not region:
        print('[é”™è¯¯] æœªæ‰¾åˆ°åº•éƒ¨çŠ¶æ€æ å—ï¼Œè¯·ç¡®è®¤æ¨¡æ¿ bar_block.png')
        return
    bar_region = (region.left, region.top, region.width, region.height)

    # 2ï¸âƒ£ å…ˆæŸ¥â€œè“è‰²å·²å¼€â€å›¾æ ‡
    if pg.locateOnScreen(str(ON_IMG), region=bar_region, confidence=conf):
        print('[OK] å¤©æ­£å¢™ä¸­çº¿å·²ç»æ˜¾ç¤ºï¼ˆè“è‰²ï¼‰ã€‚')
        return

    # 3ï¸âƒ£ æŸ¥ç°è‰²å›¾æ ‡
    off = pg.locateOnScreen(str(OFF_IMG), region=bar_region, confidence=conf)
    if not off:
        print('[é”™è¯¯] æœªæ‰¾åˆ°ç°è‰²å›¾æ ‡ï¼Œå¯èƒ½ç•Œé¢çš®è‚¤ä¸åŒæˆ–æ¨¡æ¿éœ€é‡æˆªã€‚')
        return

    # 4ï¸âƒ£ ç‚¹å‡»ç°è‰²å›¾æ ‡
    pg.click(pg.center(off))
    time.sleep(wait)

    # 5ï¸âƒ£ å†æ¬¡éªŒè¯
    if pg.locateOnScreen(str(ON_IMG), region=bar_region, confidence=conf):
        print('[OK] å·²ç‚¹å‡»ï¼Œå¤©æ­£å¢™ä¸­çº¿ç°åœ¨å·²æ˜¾ç¤ºï¼ˆè“è‰²ï¼‰ã€‚')
    else:
        print('[è­¦å‘Š]ï¸ ç‚¹å‡»åä»æœªæ£€æµ‹åˆ°è“è‰²å›¾æ ‡ï¼Œè¯·æ‰‹åŠ¨æ£€æŸ¥ã€‚')





##ä½¿ç”¨è¾…åŠ©å¤§çŸ©å½¢æ§åˆ¶è§†å›¾èŒƒå›´

def draw_shitu_rectangle_lw(length=500000.0, width=500000.0, center=(0.0, 0.0), layer_name="shitu"):
    """
    åœ¨ AutoCAD ä¸­ç»˜åˆ¶ä¸€ä¸ªä»¥ center ä¸ºä¸­å¿ƒã€é•¿ lengthã€å®½ width çš„é—­åˆè½»é‡çº§å¤šæ®µçº¿çŸ©å½¢ï¼Œ
    å¹¶æ”¾åˆ° layer_name å›¾å±‚ï¼ˆä¸å­˜åœ¨åˆ™æ–°å»ºï¼‰ã€‚

    å‚æ•°
    ----
    length, width : float
        çŸ©å½¢çš„é•¿å’Œå®½
    center : tuple(float, float)
        çŸ©å½¢ä¸­å¿ƒç‚¹ (x, y)
    layer_name : str
        æŒ‡å®šå›¾å±‚åç§°
    """

    # 2ï¸âƒ£ ç¡®ä¿ç›®æ ‡å›¾å±‚å­˜åœ¨
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
        lyr.color = 3  # è®¾ç½®å›¾å±‚é¢œè‰²ï¼ˆå¯é€‰ï¼‰

    # 3ï¸âƒ£ è®¡ç®—çŸ©å½¢é¡¶ç‚¹ï¼ˆè½»é‡çº§å¤šæ®µçº¿åªè¦äºŒç»´åæ ‡ï¼‰
    cx, cy = center
    hl = length / 2.0
    hw = width  / 2.0

    pts = [
        cx - hl, cy - hw,
        cx + hl, cy - hw,
        cx + hl, cy + hw,
        cx - hl, cy + hw,
        cx - hl, cy - hw,  # å›åˆ°èµ·ç‚¹é—­åˆ
    ]

    # 4ï¸âƒ£ æŠŠ Python åˆ—è¡¨è½¬æ¢ä¸º COM-safe çš„åŒç²¾åº¦æ•°ç»„
    variant_pts = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        pts
    )

    # 5ï¸âƒ£ æ·»åŠ è½»é‡çº§å¤šæ®µçº¿å¹¶è®¾ç½®å±æ€§
    lwpoly = mp.AddLightWeightPolyline(variant_pts)
    lwpoly.Closed = True
    lwpoly.Layer  = layer_name

    # 6ï¸âƒ£ å¯é€‰ï¼šç¼©æ”¾åˆ°å›¾å½¢ extents
    acad.ZoomExtents()

    print(f"[OK] å·²åœ¨å›¾å±‚ â€œ{layer_name}â€ ä¸Šç»˜åˆ¶ {length}Ã—{width} çŸ©å½¢ï¼ˆè½»é‡çº§å¤šæ®µçº¿ï¼‰")














#&&%è®¾ç½®é•¿åº¦è§’åº¦å•ä½ç²¾åº¦

def set_dwg_units_precision():
    """
    è®¾ç½®å½“å‰ DWG æ–‡ä»¶çš„å•ä½åŠç²¾åº¦ï¼š
    - é•¿åº¦å•ä½ï¼šå•ä½ç±»å‹ä¸å˜ï¼Œä»…è®¾ç½®ç²¾åº¦ä¸º 0.00000000
    - è§’åº¦å•ä½ï¼šå•ä½ç±»å‹ä¸å˜ï¼Œç²¾åº¦ä¸º 0.00000000
    """

    try:
        vars = doc.GetVariable

        # è®¾ç½®é•¿åº¦ç²¾åº¦ï¼ˆLUPREC = 8 è¡¨ç¤º 8 ä½å°æ•°ï¼‰
        doc.SetVariable("LUPREC", 8)

        # è®¾ç½®è§’åº¦ç²¾åº¦ï¼ˆAUPREC = 8 è¡¨ç¤º 8 ä½å°æ•°ï¼‰
        doc.SetVariable("AUPREC", 8)

        print("[OK] å·²å°†é•¿åº¦å’Œè§’åº¦å•ä½ç²¾åº¦è®¾ç½®ä¸º 8 ä½å°æ•° (0.00000000)")
    except Exception as e:
        print(f"[é”™è¯¯] è®¾ç½®å¤±è´¥: {e}")

def jd():
    set_dwg_units_precision()


##æ ‡æ³¨æ ·å¼

def list_dim_styles():
    """
    åˆ—å‡ºå½“å‰ DWG æ–‡ä»¶ä¸­æ‰€æœ‰æ ‡æ³¨æ ·å¼åç§°ã€‚
    """
    try:
        styles = doc.DimStyles
        names = [styles.Item(i).Name for i in range(styles.Count)]
        print("ğŸ“ å½“å‰æ ‡æ³¨æ ·å¼åˆ—è¡¨ï¼š")
        for name in names:
            print(" -", name)
        return names
    except Exception as e:
        print(f"[é”™è¯¯] è·å–æ ‡æ³¨æ ·å¼å¤±è´¥ï¼š{e}")
        return []



def set_current_dimstyle_via_command(style_name="_TCH_ARCH"):
    """
    ä½¿ç”¨å‘½ä»¤è¡Œæ–¹å¼è®¾ç½®å½“å‰æ ‡æ³¨æ ·å¼ï¼Œå…¼å®¹å¤©æ­£ã€‚

    å‚æ•°:
        style_name (str): è¦è®¾ä¸ºå½“å‰æ ‡æ³¨æ ·å¼çš„åç§°ï¼ˆå¦‚ "_TCH_ARCH"ï¼‰
    """
    try:
        doc.SendCommand(f"-DIMSTYLE\nR\n{style_name}\n")
        print(f"[OK] å·²å°è¯•é€šè¿‡å‘½ä»¤è¡Œè®¾ç½®æ ‡æ³¨æ ·å¼ä¸ºï¼š{style_name}")
    except Exception as e:
        print(f"[é”™è¯¯] å‘½ä»¤è¡Œè®¾ç½®æ ‡æ³¨æ ·å¼å¤±è´¥ï¼š{e}")



#&&% æ–‡å­—å·¦å¯¹é½åˆ°å‚ç›´çº¿

def align_texts_to_x_by_llcorner(texts, x_target):
    """
    å°† texts åˆ—è¡¨ä¸­æ¯ä¸ªæ–‡å­—å¯¹è±¡çš„å¤–åŒ…ç›’å·¦ä¸‹è§’ X å¯¹é½åˆ° x_targetï¼ŒYã€Z ä¸å˜ã€‚

    texts: List[COMObject]    å•è¡Œæˆ–å¤šè¡Œæ–‡å­—å¯¹è±¡
    x_target: float           ç›®æ ‡ X åæ ‡
    """
    base_pt = vtpnt(0, 0, 0)
    for txt in texts:
        try:
            # å–å·¦ä¸‹è§’ç‚¹
            ll_pt, _ = txt.GetBoundingBox()
            x0, y0, z0 = ll_pt
            # è®¡ç®—å¹³ç§»å‘é‡
            dx = x_target - x0
            move_vec = vtpnt(dx, 0, 0)
            # å¹³ç§»
            txt.Move(base_pt, move_vec)
        except Exception as e:
            h = getattr(txt, "Handle", "?")
            print(f"[è­¦å‘Š] å¯¹è±¡ {h} å¯¹é½å¤±è´¥ï¼š{e}")




#&&% è·å–æ–‡å­—å†…å®¹

def extract_text_content(ent):
    """
    æ ¹æ®å®ä½“ç±»å‹è·å–æ–‡æœ¬å†…å®¹ï¼Œå¹¶ä»…ç§»é™¤æ–‡å­—é¦–å°¾çš„ç©ºæ ¼ï¼š
      - AcDbText æˆ– AcDbMText: ent.TextString
      - TDbText: ent.Text
      - TDbMText: TDbMText_content(ent)
    åªè°ƒç”¨ strip() åˆ é™¤é¦–å°¾ç©ºç™½ï¼Œä¸å½±å“ä¸­é—´ç©ºæ ¼ã€‚
    """
    obj = ent.ObjectName
    if obj in ("AcDbText", "AcDbMText"):
        raw = ent.TextString
    elif obj == "TDbText":
        raw = ent.Text
    elif obj == "TDbMText":
        raw = TDbMText_content(ent)
    else:
        return ""
    # ä»…åˆ é™¤é¦–å°¾ç©ºç™½
    return raw.strip()





#&&% è®¾ç½®å½“å‰ä½¿ç”¨çš„å­—ä½“æ ·å¼

        
def set_current_text_style(style_name="Standard"):
    """
    è®¾ç½®å½“å‰æ–‡å­—æ ·å¼ï¼ˆé€šè¿‡ COM æ¥å£æ–¹å¼ï¼‰ã€‚

    å‚æ•°:
        style_name (str): è¦è®¾ç½®ä¸ºå½“å‰çš„æ–‡å­—æ ·å¼åç§°
    """
    try:
        text_styles = doc.TextStyles
        style = text_styles.Item(style_name)  # è·å–æŒ‡å®šæ ·å¼
        doc.ActiveTextStyle = style           # è®¾ç½®ä¸ºå½“å‰æ ·å¼
        print(f"[OK] å½“å‰æ–‡å­—æ ·å¼å·²è®¾ç½®ä¸ºï¼š{style_name}")
    except Exception as e:
        print(f"[é”™è¯¯] è®¾ç½®æ–‡å­—æ ·å¼å¤±è´¥ï¼š{e}")

#&&% è·å–æ‰€æœ‰å½“å‰å­—ä½“æ ·å¼

def huoqu_ziti_style():

    styles = {ts.Name for ts in doc.TextStyles}
    
    return styles




#&&% è®¾ç½®å­—ä½“æ ·å¼

"""
1ï¼Œå¯å¤šæ¬¡å¯¹æŸç§æ ·å¼è®¾ç½®å†æ›´æ–°

2ï¼Œå¯¹æ±‰å­—å­—ä½“è®¾ç½®ï¼Œä¸å­˜åœ¨è¦è®¾å¤§å­—ä½“ã€‚å¯ç”¨å­—ä½“åœ¨C:/Windows/FontsæŸ¥æ‰¾ï¼Œç›´æ¥ç”¨æ±‰å­—åï¼Œä½¿ç”¨create_text_styleå‘½ä»¤

3ï¼Œ å¯ä»¥å•ç‹¬è®¾ç½®shxå­—ä½“ï¼Œä½¿ç”¨set_text_style_onlyshx(style_name="style01", font_file="gbenor.shx", big_font_file=None)

æ‰€é€ å­—ä½“ç¼ºå¤±å¤§å­—ä½“ï¼Œå¯¹æ±‰å­—ä¼šäº§ç”Ÿé—®å·

4ï¼Œå¦‚ä¸æ˜¯å•ç‹¬è®¾ç½®æ±‰å­—å­—ä½“ï¼Œå®Œæ•´çš„å­—ä½“è®¾ç½®åº”åŒ…æ‹¬è‹±æ–‡éƒ¨åˆ†çš„è®¾ç½®å’Œå¤§å­—ä½“è®¾ç½®ï¼Œç”¨äºä¸­æ–‡æ˜¾ç¤ºè€Œä¸”æ˜¯shxå­—ä½“

 

"""
def create_text_style(sty_name="style01", ziti="å®‹ä½“"):
    """
    åœ¨å½“å‰ DWG ä¸­åˆ›å»ºï¼ˆæˆ–æ›´æ–°ï¼‰ä¸€ä¸ªä¸­æ–‡æ–‡å­—æ ·å¼ã€‚
    

    :param sty_name: æ ·å¼åç§°ï¼Œé»˜è®¤ "style01"
    :param ziti:     å­—ä½“åç§°ï¼ŒAutoCAD ä¼šåœ¨ç³»ç»Ÿå­—ä½“ä¸­æŸ¥æ‰¾ï¼Œä¾‹å¦‚ "å®‹ä½“"
    # acad.ActiveDocument.ActiveTextStyle.SetFont(Typeface, Bold, Italic, charSet, PitchandFamily)
    # Typeface å­—ä½“åç§°ï¼›
    # Bold åŠ ç²—ï¼Œå¸ƒå°”å€¼ï¼ŒFalseä¸ºä¸åŠ ç²—å­—ä½“ï¼›
    # Italic å€¾æ–œï¼Œå¸ƒå°”å€¼ï¼ŒFalseä¸ºå€¾æ–œå­—ä½“ï¼›
    # CharSet å­—ä½“å­—ç¬¦é›†ï¼Œ1ä¸ºé»˜è®¤å­—ç¬¦é›†ï¼›
    # PitchAndFamily å­—èŠ‚åŠç¬”ç”»å½¢å¼ã€‚

    ts = doc.TextStyles.Item("HIT_TxtStyle")å¯¹shx

    ts.fontFile = "bigfont.shx"



    """

    # 1ï¸âƒ£ ç¡®ä¿æ ·å¼å­˜åœ¨
    styles = doc.TextStyles
    try:
        ts = styles.Item(sty_name)
        print(f"[è­¦å‘Š] æ ·å¼ '{sty_name}' å·²å­˜åœ¨ï¼Œæ­£åœ¨æ›´æ–°å…¶å±æ€§ã€‚")
    except Exception:
        ts = styles.Add(sty_name)
        print(f"[OK] å·²åˆ›å»ºæ–‡å­—æ ·å¼ '{sty_name}'ã€‚")

    # 2ï¸âƒ£ è®¾ç½®å­—ä½“
    try:
        acad.ActiveDocument.TextStyles.Item(sty_name).SetFont(ziti, False, False, 1, 0 or 0)
    except Exception:
        try:
            acad.ActiveDocument.TextStyles.Item(sty_name).SetFont(ziti, False, False, 1, 0 or 0)
        except Exception as e:
            print(f"[è­¦å‘Š] æ— æ³•è®¾ç½®å­—ä½“ä¸º '{ziti}'ï¼š{e}")

    # 3ï¸âƒ£ ç½®ä¸ºå½“å‰

    acad.ActiveDocument.ActiveTextStyle = acad.ActiveDocument.TextStyles.Item(sty_name)
    
    # 5ï¸âƒ£ é€šçŸ¥ç”¨æˆ·
    print(f"[OK] æ ·å¼ '{sty_name}' å±æ€§å·²æ›´æ–°")




def set_text_style_onlyshx(style_name="style01", font_file="gbenor.shx", big_font_file=None):
    """
    C:/Program Files/Autodesk/AutoCAD 2021/FontsæŸ¥æ‰¾å¯ç”¨shxå­—ä½“    
    è®¾ç½® AutoCAD çš„æ–‡å­—æ ·å¼ï¼š
    - font_fileï¼šè‹±æ–‡å­—ä½“ï¼ˆ.shx æˆ– .ttfï¼‰
    - big_font_fileï¼šå¯é€‰çš„å¤§å­—ä½“ï¼ˆå¦‚ gbcbig.txtï¼‰ï¼Œé»˜è®¤ä¸º None è¡¨ç¤ºä¸è®¾ç½®ï¼Œä»…è®¾ç½®è‹±æ–‡å­—ä½“
    set_text_style_onlyshx(style_name="TEST_STYLE", font_file="gbxxx.shx", big_font_file=None)
    ä¸æˆåŠŸæ¶ˆæ¯å¯ç”¨äºåˆ¤å®šshxå­—ä½“åœ¨ä¸åœ¨å·¦ä¾§è‹±æ–‡å­—ä½“


    """
    try:
        import win32com.client

        acad = win32com.client.Dispatch("AutoCAD.Application")
        doc = acad.ActiveDocument
        styles = doc.TextStyles

        # æŸ¥æ‰¾æˆ–åˆ›å»ºå­—ä½“æ ·å¼
        if style_name in [s.Name for s in styles]:
            text_style = styles.Item(style_name)
        else:
            text_style = styles.Add(style_name)

        text_style.FontFile = font_file

        # åªæœ‰ä¼ å…¥åˆæ³•çš„å¤§å­—ä½“è·¯å¾„æ‰è®¾ç½® BigFontFileï¼Œå¦åˆ™è·³è¿‡è®¾ç½®
        if big_font_file and isinstance(big_font_file, str):
            text_style.BigFontFile = big_font_file

        print(f"å­—ä½“æ ·å¼ '{style_name}' è®¾ç½®æˆåŠŸï¼šè‹±æ–‡å­—ä½“ = {font_file}ï¼Œå¤§å­—ä½“ = {big_font_file or 'æœªè®¾ç½®'}")
        return True

    except Exception as e:
        print(f"è®¾ç½®å­—ä½“æ ·å¼å¤±è´¥ï¼š{e}")
        return False



def set_text_style(style_name="style01", font_file="gbenor.shx", big_font_file="gbcbig.shx"):
    """
    è®¾ç½®CADä¸­æ–‡å­—æ ·å¼ï¼šè‹±æ–‡shxæ–‡ä»¶ + ä¸­æ–‡å¤§å­—ä½“æ–‡ä»¶
    """
    try:

        styles = doc.TextStyles

        # åˆ¤æ–­æ˜¯å¦å·²æœ‰è¯¥æ ·å¼
        if style_name in [s.Name for s in styles]:
            text_style = styles.Item(style_name)
        else:
            text_style = styles.Add(style_name)

        # è®¾ç½®å­—ä½“å’Œå¤§å­—ä½“
        text_style.FontFile = font_file
        text_style.BigFontFile = big_font_file

        print(f"å­—ä½“æ ·å¼ '{style_name}' è®¾ç½®æˆåŠŸï¼Œè‹±æ–‡å­—ä½“: {font_file}, ä¸­æ–‡å¤§å­—ä½“: {big_font_file}")
        return True

    except Exception as e:
        print(f"è®¾ç½®å­—ä½“æ ·å¼å¤±è´¥ï¼š{e}")
        return False

#&&% åˆ—å‡ºå¯ç”¨shxéå¤§å­—è¡¨ shxå¤§å­—è¡¨ 

"""

åˆ°CADå­—ä½“è®¾ç½®ä¸‹æ‹‰èœå•æ‰¾

"""

##&&% é—®å·å­—ä½“æ›¿æ¢
"""
ç½‘å‹çš„ftstæ–¹æ¡ˆå·²ç»å½»åº•è§£å†³æ­¤é—®é¢˜ï¼Œæˆ‘ä»¬éœ€è¦çš„æ˜¯å¤‡ä»½å¥½æ–‡ä»¶å’Œè¿›ä¸€æ­¥ç¼–åˆ¶ç‚¹å‡»çª—å£çš„å‡½æ•°
æˆ‘å»ºè®®ä½¿ç”¨Standardæ ·å¼çš„gbenor.shxå’Œå¤§å­—gbcbig.shxæ›¿æ¢æ›´å®‰å…¨
"""



#&&% ## å¤„ç†å­—ä½“æ ·å¼åŒåé—®é¢˜

def rename_conflicting_text_styles(file1_path: str,
                                   file2_path: str,
                                   suffix: str = "_1",
                                   retry_delay: float = 0.2,
                                   max_retries: int = 10):
    """
    åœ¨ä¸¤ä¸ª DWG ä¸­æ‰¾å‡ºåŒåï¼ˆç”¨æˆ·ï¼‰æ–‡å­—æ ·å¼ï¼Œ
    å¹¶åœ¨ç¬¬ä¸€ä¸ªæ–‡ä»¶ä¸­å°†å®ƒä»¬é‡å‘½åï¼ˆåŸå + suffixï¼‰ï¼š
      1) é€šè¿‡ -RENAME å‘½ä»¤é‡å‘½åæ ·å¼
      2) ç¡®è®¤æ ·å¼è¡¨é‡Œæ—§åå·²æ¶ˆå¤±ã€æ–°åå·²å‡ºç°
      3) REGEN å¼ºåˆ¶åˆ·æ–°
      4) éå† ModelSpaceï¼Œå°†å¼•ç”¨æ—§æ ·å¼çš„å®ä½“æŒ‡å‘æ–°æ ·å¼
      5) ä¿å­˜å¹¶å…³é—­
    ç³»ç»Ÿé»˜è®¤æ ·å¼ä¼šè¢«è‡ªåŠ¨è·³è¿‡ã€‚
    """
    SYSTEM_STYLES = {"Standard", "ASHADE", "Annotative", "BigFont"}

    acad = win32com.client.Dispatch("AutoCAD.Application")
    acad.Visible = True

    doc1 = acad.Documents.Open(os.path.abspath(file1_path))
    doc2 = acad.Documents.Open(os.path.abspath(file2_path))

    try:
        # 1) æ”¶é›†ä¸¤æ–‡ä»¶çš„æ ·å¼
        styles1 = {ts.Name for ts in doc1.TextStyles}
        styles2 = {ts.Name for ts in doc2.TextStyles}
        conflicts = (styles1 & styles2) - SYSTEM_STYLES
        if not conflicts:
            print("[OK] æœªå‘ç°éœ€è¦é‡å‘½åçš„ç”¨æˆ·æ ·å¼ã€‚")
            return

        print(f"[è­¦å‘Š] å‘ç°åŒåç”¨æˆ·æ ·å¼ï¼š{conflicts}ï¼Œå°†åœ¨ â€œ{os.path.basename(file1_path)}â€ ä¸­é‡å‘½åï¼š")
        ms = doc1.ModelSpace

        for old_name in conflicts:
            new_name = old_name + suffix
            # å¦‚æœæ–°åä¹Ÿå†²çªï¼Œå°±å¤šåŠ åç¼€ï¼Œç›´åˆ°ç‹¬ä¸€
            while new_name in styles1:
                new_name += suffix

            # 2) å‘é€ RENAME å‘½ä»¤
            cmd = f"-RENAME\nStyle\n{old_name}\n{new_name}\n\n"
            doc1.SendCommand(cmd)
            # ç­‰å¾…å‘½ä»¤è¢«å¤„ç†ã€æ ·å¼è¡¨æ›´æ–°
            for attempt in range(max_retries):
                time.sleep(retry_delay)
                # å¼ºåˆ¶åˆ·æ–°å›¾å½¢çŠ¶æ€
                doc1.SendCommand("REGEN\n")
                try:
                    # å¦‚æœæ—§åå·²ä¸å­˜åœ¨å¹¶ä¸”æ–°åå­˜åœ¨ï¼Œå°±è·³å‡ºé‡è¯•
                    doc1.TextStyles.Item(new_name)
                    try:
                        doc1.TextStyles.Item(old_name)
                        # æ—§åä»ç„¶å­˜åœ¨ï¼Œç»§ç»­ç­‰
                        continue
                    except Exception:
                        # æ—§åè¢«æ­£ç¡®ç§»é™¤
                        break
                except Exception:
                    # æ–°åè¿˜æ²¡å‡ºç°ï¼Œç»§ç»­ç­‰
                    continue
            else:
                print(f"  â— é‡å‘½å â€œ{old_name}â€ â†’ â€œ{new_name}â€ å¯èƒ½æœªç”Ÿæ•ˆï¼ˆè¶…æ—¶ï¼‰ã€‚")

            # 3) å†æŠŠæ‰€æœ‰å®ä½“ä¸­å¼•ç”¨æ—§æ ·å¼çš„æ”¹æˆæ–°æ ·å¼
            for ent in ms:
                try:
                    ename = getattr(ent, "EntityName", "").upper()
                    oname = getattr(ent, "ObjectName", "")
                    # åŸç”Ÿ TEXT/MTEXT
                    if ename in ("TEXT", "MTEXT"):
                        if ent.TextStyle == old_name:
                            ent.TextStyle = new_name
                    # å¤©æ­£æ–‡å­—
                    elif oname in ("TDbText", "TDbMText"):
                        if ent.TextStyle == old_name:
                            ent.TextStyle = new_name
                except Exception:
                    # æœ‰äº›å®ä½“å¯èƒ½ä¸å…è®¸æ”¹æ ·å¼ï¼Œå¿½ç•¥å®ƒä»¬
                    pass

            # æ›´æ–°æœ¬åœ°æ ·å¼é›†åˆ
            styles1.discard(old_name)
            styles1.add(new_name)
            print(f"  Â· æ ·å¼ â€œ{old_name}â€ â†’ â€œ{new_name}â€")

        # 4) æœ€åä¿å­˜å¹¶åé¦ˆ
        doc1.Save()
        print(f"[OK] å·²ä¿å­˜æ”¹åŠ¨åˆ° â€œ{os.path.basename(file1_path)}â€ã€‚")

    finally:
        # å…³é—­æ–‡æ¡£ï¼Œä¸ä¿å­˜å¯¹ç¬¬äºŒä¸ªæ–‡ä»¶çš„ä»»ä½•æ”¹åŠ¨
        doc1.Close(False)
        doc2.Close(False)




















##å°†ä¸€ä¸ªå¯¹è±¡å±æ€§ä¼ ç»™å¤šä¸ªå¯¹è±¡

def transfer_props_by_matchprop(entity, Ob, max_try=3, delay=0.4):

    CR = chr(13)

    def wait_idle(acad, dt=0.2, n=50):
        """è½®è¯¢ IsQuiescentï¼Œæœ€å¤š n æ¬¡ï¼Œæ¯æ¬¡ dt ç§’"""
        for _ in range(n):
            if acad.GetAcadState().IsQuiescent:
                return
            time.sleep(dt)

    def expand_rectangle(p1, p2, offset):
        return (p1[0]-offset, p1[1]-offset), (p2[0]+offset, p2[1]+offset)


    """
    æŠŠ entity çš„å±æ€§æ‰¹é‡å¤åˆ¶åˆ° Obã€‚è‹¥ Layer æœªå˜åŒ–åˆ™é‡è¯•ï¼Œæœ€å¤š 3 æ¬¡ã€‚
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    src_layer   = entity.Layer
    orig_layer  = Ob.Layer        # å¤åˆ¶å‰ç›®æ ‡çš„å›¾å±‚

    # ç›®æ ‡åŒ…å›´ç›’çª—å£
    p1, p2 = Ob.GetBoundingBox()
    x1, y1, x2, y2 = p1[0], p1[1], p2[0], p2[1]
    h = 0.1 * (abs(x1 - x2) + abs(y1 - y2)) / 2
    (wx1, wy1), (wx2, wy2) = expand_rectangle(p1, p2, h)

    match_cmd = (
        "_MATCHPROP" + CR +
        "P"          + CR +            # Previous ä½œä¸ºæº
        "_W"         + CR +
        f"{wx1},{wy1}" + CR +
        f"{wx2},{wy2}" + CR + CR
    )

    for attempt in range(1, max_try + 1):
        try:
            # â€”â€”â€” 1. è®¾å®šæºå¯¹è±¡ä¸º Previous â€”â€”â€”
            try:
                highlight_entity_by_bbox(entity)
            except Exception as e:
                # é€€è€Œæ±‚å…¶æ¬¡ï¼šä½¿ç”¨ LISP é€šè¿‡ Handle é€‰æ‹©å¯¹è±¡
                print(f"[FALLBACK] highlight_entity_by_bbox å¤±è´¥ï¼Œä½¿ç”¨ Handle é€‰æ‹©: {e}")
                handle = com_retry(lambda: entity.Handle)
                # ä½¿ç”¨ LISP çš„ ssget åˆ›å»ºé€‰æ‹©é›†å¹¶æ˜¾ç¤º
                doc.SendCommand(f"(sssetfirst nil (ssget \"_X\" '((5 . \"{handle}\"))))\n")
                time.sleep(0.5)
            time.sleep(delay)

            # â€”â€”â€” 2. å‘é€ MATCHPROP â€”â€”â€”
            doc.SendCommand(match_cmd)
            wait_idle(acad)

            # â€”â€”â€” 3. åˆ¤æ–­æ˜¯å¦æˆåŠŸ â€”â€”â€”
            if Ob.Layer == src_layer:
                print(f"[OK] ç¬¬ {attempt} æ¬¡åŒ¹é…æˆåŠŸï¼ŒLayer æ”¹ä¸º {src_layer}")
                return True

            print(f"[WARN] ç¬¬ {attempt} æ¬¡å Layer æœªå˜ï¼Œé‡è¯•â€¦")
            time.sleep(delay)

        except Exception as e:
            print(f"[ERR] ç¬¬ {attempt} æ¬¡åŒ¹é…å¼‚å¸¸ï¼š{e}")

    print(f"[FAIL] è¿ç»­ {max_try} æ¬¡ä»æœªæŠŠå±æ€§å¤åˆ¶ç»™ç›®æ ‡")
    return False


#è§†å›¾åˆç†åŒ–æ§åˆ¶
"""
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_swiso"+chr(13))#è¥¿å—è½´æµ‹
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_seiso"+chr(13))#ä¸œå—è½´æµ‹
acad.ActiveDocument.SendCommand(
    "_-view" + chr(13) +
    "_nwiso" + chr(13)
)#è¥¿åŒ—è½´æµ‹
acad.ActiveDocument.SendCommand(
    "_-view" + chr(13) +
    "_neiso" + chr(13)
)#ä¸œåŒ—è½´æµ‹
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_top"+chr(13))#ä¿¯è§†å›¾
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_bottom"+chr(13))#ä»°è§†å›¾
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_front"+chr(13))#å‰è§†å›¾
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Right"+chr(13))#å³è§†å›¾
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Back"+chr(13))#åè§†å›¾
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Left"+chr(13))#å·¦è§†å›¾
acad.ActiveDocument.SendCommand("_vscurrent"+chr(13)+"_R"+chr(13))#çœŸå®è§†è§‰æ ·å¼

##åˆç†æ˜¾ç¤ºå¯¹è±¡

acad.ActiveDocument.SendCommand(
    "_z" + chr(13) +
    "_e" + chr(13)
)

acad.ActiveDocument.SendCommand(
    "_zoom" + chr(13) +
    "s"     + chr(13) +
    "0.8xp" + chr(13)
)




"""
#&&% åŒçº¿ç¨‹ç”Ÿæˆå™¨

def run_dual_threads_1(f1,                     # çº¿ç¨‹1 å‡½æ•°

                     f2,                     # çº¿ç¨‹2 å‡½æ•°

                     f1_args=(), f1_kwargs=None,

                     f2_args=(), f2_kwargs=None,

                     timeout_sec=180):



    """
    é€šç”¨â€œåŒçº¿ç¨‹-GUIâ€è°ƒåº¦å™¨

    - f1 å¿…é¡»è´Ÿè´£è§¦å‘ä¸€ä¸ªé˜»å¡æ€§çª—å£ï¼ˆå¦‚ SendCommandã€ShowModalDlg ç­‰ï¼‰ã€‚
    - f2 è´Ÿè´£ä¾¦æµ‹å¹¶è‡ªåŠ¨åŒ–å¤„ç†è¯¥çª—å£ï¼Œå½“å¤„ç†å®Œæ¯•åé€šçŸ¥ f1 ç»§ç»­æˆ–é€€å‡ºã€‚
    - f1ã€f2 çš„å‡½æ•°ç­¾åéƒ½åº”ä»¥ (timeout_event, done_event, â€¦) å¼€å¤´ï¼š
    
      def f1(timeout_event, done_event, â€¦):
          pythoncom.CoInitialize()
          try:
              # â€¦â€¦ä¸»ä½“ä»£ç ï¼šä¾‹å¦‚è§¦å‘æ‰“å°å¯¹è¯æ¡†
              timeout_event.wait()    # ç­‰å¾…çº¿ç¨‹2é€šçŸ¥æˆ–è¶…æ—¶
          except Exception:
              # æ‰“å°æ—¥å¿—æˆ–å…¶ä»–å¤„ç†
          finally:
              pythoncom.CoUninitialize()
              done_event.set()        # é€šçŸ¥â€œåŒçº¿ç¨‹â€æ€»æ§ï¼šæˆ‘ç»“æŸäº†

      def f2(timeout_event, done_event, â€¦):
          pythoncom.CoInitialize()
          try:
              # â€¦â€¦ä¸»ä½“ä»£ç ï¼šä¾‹å¦‚ç­‰å¾…â€œåˆ›å»ºæ‰“å°æ–‡ä»¶â€å¯¹è¯æ¡†å‡ºç°å¹¶ç‚¹å‡»â€œå›è½¦â€é”®
          except Exception:
              # æ‰“å°æ—¥å¿—æˆ–å…¶ä»–å¤„ç†
          finally:
              pythoncom.CoUninitialize()
              timeout_event.set()    # é€šçŸ¥ f1 çº¿ç¨‹ï¼šçª—å£å·²å¤„ç†å®Œï¼Œå¯ä»¥ç»“æŸç­‰å¾…
              done_event.set()       # é€šçŸ¥â€œåŒçº¿ç¨‹â€æ€»æ§ï¼šæˆ‘ç»“æŸäº†

    - å‚æ•°è¯´æ˜ï¼š
        f1, f2ï¼šä¼ å…¥ä¸Šé¢é‚£ç§ç­¾åçš„å‡½æ•°
        f1_argsã€f1_kwargsï¼šç»™ f1 ä¼ é€’çš„ä½ç½®å‚æ•°å’Œå…³é”®å­—å‚æ•°
        f2_argsã€f2_kwargsï¼šç»™ f2 ä¼ é€’çš„ä½ç½®å‚æ•°å’Œå…³é”®å­—å‚æ•°
        timeout_secï¼šæœ€å¤šç­‰å¾…å¤šå°‘ç§’ï¼Œå¦‚æœè¶…è¿‡åˆ™æŠ¥å‘Šè¶…æ—¶å¹¶è¿”å› False

    - è¿”å›å€¼ï¼š
        True  ï¼šä¸¤ä¸ªå­çº¿ç¨‹åœ¨ time_limit å†…éƒ½å®Œæˆäº†
        False ï¼šè¶…æ—¶ï¼Œè‡³å°‘ä¸€ä¸ªçº¿ç¨‹æ²¡åŠæ—¶è°ƒç”¨ done_event.set()
    """
    if f1_kwargs is None:
        f1_kwargs = {}
    if f2_kwargs is None:
        f2_kwargs = {}

    # â‘  ä¸ºçº¿ç¨‹ 1ã€2 å‡†å¤‡åŒä¸€ä¸ªäº‹ä»¶å¯¹
    timeout_event = threading.Event()
    done_event    = threading.Event()

    # â‘¡ å¯åŠ¨â€œçº¿ç¨‹1â€ï¼šè´Ÿè´£å¼¹å‡ºã€é˜»å¡çª—å£
    t1 = threading.Thread(
        target=f1,
        args=(timeout_event, done_event, *f1_args),
        kwargs=f1_kwargs,
        daemon=True
    )
    # â‘¢ å¯åŠ¨â€œçº¿ç¨‹2â€ï¼šè´Ÿè´£ä¾¦æµ‹çª—å£å¹¶ç‚¹å‡»ã€å…³é—­
    t2 = threading.Thread(
        target=f2,
        args=(timeout_event, done_event, *f2_args),
        kwargs=f2_kwargs,
        daemon=True
    )

    start_time = time.time()
    t1.start()
    t2.start()

    # â‘£ ç­‰å¾…çº¿ç¨‹1 åœ¨ timeout_sec ç§’å†…ç»“æŸ
    t1.join(timeout=timeout_sec)
    
    t2.join(timeout=timeout_sec)
    
        
        

    # â‘¥ æ£€æŸ¥ done_event æ˜¯å¦å·²è¢« set()
    if not done_event.is_set():
        # è¶…æ—¶ï¼šç”±è°ƒåº¦å™¨è´Ÿè´£ç»™çº¿ç¨‹1 å‘é€é€€å‡ºä¿¡å·
        timeout_event.set()
        print(f"[è­¦å‘Š] åŒçº¿ç¨‹æ‰§è¡Œè¶…è¿‡ {timeout_sec}s â€”â€” å·²è§¦å‘ timeout_event")
        return False
    else:
        print("[OK] åŒçº¿ç¨‹ä»»åŠ¡åœ¨æ—¶é™å†…å®Œæˆ")
        return True



#&&% CADå–æ¶ˆé€‰æ‹©æ“ä½œ



def cancel_cad_selection(attempts: int = 3, delay: float = 0.5) -> bool:

    for i in range(1, attempts + 1):
        try:
            highlight_entities_in_window(0, 0, 0, 0)
            print(f"[OK] ç¬¬{i}æ¬¡å°è¯•ï¼šcancel_cad_selection æˆåŠŸ")
            return True
        except Exception as e:
            print(f"[è­¦å‘Š] ç¬¬{i}æ¬¡å°è¯•å¤±è´¥ï¼š{e}")
            if i < attempts:
                time.sleep(delay)
    print("[é”™è¯¯] å·²é‡è¯•å¤šæ¬¡ï¼Œä»æœªèƒ½æ‰§è¡Œ cancel_cad_selection")
    return False






#&&&&%% ç¬¬ä¸ƒéƒ¨åˆ†  æµ‹è¯•è¾…åŠ©




#   CADåŸºæœ¬æ“ä½œ-æµ‹è¯•è¾…åŠ©
#####&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&




#æœ€å°åŒ–çª—å£
def min_w():
    import ctypes
    
    VK_MENU = 0x12  # Alté”®
    VK_TAB = 0x09   # Tabé”®
    VK_LWIN = 0x5B  # å·¦Winé”®
    VK_M = 0x4D     # Mé”®
    KEYEVENTF_KEYUP = 0x2

    # æ¨¡æ‹ŸAlt + Tab
    ctypes.windll.user32.keybd_event(VK_MENU, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_TAB, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_TAB, 0, KEYEVENTF_KEYUP, 0)
    ctypes.windll.user32.keybd_event(VK_MENU, 0, KEYEVENTF_KEYUP, 0)

    # æ¨¡æ‹ŸWin + M
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_M, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_M, 0, KEYEVENTF_KEYUP, 0)
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)




def ql():#æ¸…é™¤æµ‹è¯•è¾…åŠ©å›¾å±‚ä¸Šçš„å¯¹è±¡

    ensure_layer("æµ‹è¯•è¾…åŠ©")

    




    
def srhd(*args):#æ ¹æ®è¾“å…¥åæ ‡åœ¨æ¨¡å‹ç©ºé—´ç”»ç‚¹
    """
    åœ¨æ¨¡å‹ç©ºé—´ç»˜åˆ¶ç‚¹å¹¶æ ‡æ³¨åºå·ï¼Œæ”¯æŒä»¥ä¸‹è°ƒç”¨å½¢å¼ï¼š
    - srhd((x1,y1,z1), (x2,y2,z2))   # å¤šä¸ªç‚¹å…ƒç»„
    - srhd([(x1,y1,z1), (x2,y2,z2)]) # åˆ—è¡¨å½¢å¼
    - srhd((x1,y1,z1))               # å•ç‚¹
    """
    doc = acad.ActiveDocument
    ms = doc.ModelSpace
    layer_name = "æµ‹è¯•è¾…åŠ©"

    # åˆ›å»ºå›¾å±‚ï¼ˆå¦‚æœä¸å­˜åœ¨ï¼‰
    try:
        doc.Layers.Item(layer_name)
    except:
        doc.Layers.Add(layer_name)
        print(f"[OK] å·²åˆ›å»ºå›¾å±‚ï¼š{layer_name}")

    # ç»Ÿä¸€æ ¼å¼å¤„ç†ï¼šå…è®¸å•ç‚¹ã€å¤šä¸ªç‚¹ã€åˆ—è¡¨ä¼ å…¥
    if len(args) == 1:
        if isinstance(args[0], (list, tuple)):
            if len(args[0]) == 3 and all(isinstance(i, (int, float)) for i in args[0]):
                points = [args[0]]
            else:
                points = args[0]
        else:
            print("[é”™è¯¯] è¾“å…¥æ ¼å¼ä¸æ­£ç¡®")
            return
    else:
        points = args

    # ç»˜åˆ¶ç‚¹ä¸ç¼–å·
    for idx, P in enumerate(points, 1):
        try:
            pt = vtpnt(*P)
            point = ms.AddPoint(pt)
            point.Layer = layer_name

            text = ms.AddText(str(idx), pt, 10)
            text.Layer = layer_name
        except Exception as e:
            print(f"[é”™è¯¯] æ·»åŠ ç‚¹å¤±è´¥: {e}")

    return "[OK] ç‚¹ä¸ç¼–å·å·²ç»˜åˆ¶"



def srhd_p(*args):#æ ¹æ®è¾“å…¥åæ ‡åœ¨å›¾çº¸ç©ºé—´ç”»ç‚¹
    """
    åœ¨å›¾çº¸ç©ºé—´ç»˜åˆ¶ç‚¹å’Œç¼–å·ï¼Œæ”¯æŒï¼š
    - å¤šä¸ªåæ ‡å…ƒç»„ï¼šsrhd_p((x1,y1,z1), (x2,y2,z2))
    - å•ä¸ªåˆ—è¡¨ï¼šsrhd_p([(x1,y1,z1), (x2,y2,z2)])
    - å•ä¸ªç‚¹ï¼šsrhd_p((x1,y1,z1))
    """
    ps = doc.PaperSpace
    layer_name = "æµ‹è¯•è¾…åŠ©"

    # åˆ›å»ºå›¾å±‚ï¼ˆå¦‚æœä¸å­˜åœ¨ï¼‰
    try:
        doc.Layers.Item(layer_name)
    except:
        doc.Layers.Add(layer_name)
        print(f"[OK] å·²åˆ›å»ºå›¾å±‚ï¼š{layer_name}")

    # åˆ¤æ–­å‚æ•°ç»“æ„
    if len(args) == 1:
        if isinstance(args[0], (list, tuple)):
            if len(args[0]) == 3 and all(isinstance(i, (int, float)) for i in args[0]):
                # å•ä¸ªç‚¹å…ƒç»„
                points = [args[0]]
            else:
                # åˆ—è¡¨å½¢å¼
                points = args[0]
        else:
            print("[é”™è¯¯] è¾“å…¥æ ¼å¼ä¸æ­£ç¡®")
            return
    else:
        # å¤šä¸ªç‚¹å…ƒç»„
        points = args

    for idx, P in enumerate(points, 1):
        try:
            pt = vtpnt(*P)
            point = ps.AddPoint(pt)
            point.Layer = layer_name

            text = ps.AddText(str(idx), pt, 10)
            text.Layer = layer_name
        except Exception as e:
            print(f"[é”™è¯¯] æ·»åŠ ç‚¹å¤±è´¥: {e}")

    return "[OK] å›¾çº¸ç©ºé—´ä¸­çš„ç‚¹ä¸ç¼–å·å·²ç»˜åˆ¶"


def comtomath(LBcom):#å°†comç‚¹åˆ—è¡¨è½¬ä¸ºæ•°å­¦ç‚¹åˆ—è¡¨

    LB_point=[]

    for i in range(0,len(LBcom)):

        point = LBcom[i].Coordinates

        LB_point.append(point)

    return LB_point        

        

#&&% éš”è¿œæŸ¥çœ‹

def fuzhi_chakan(LBcom,K=1):#Kä¸ºæ”¾å¤§å€æ•°

    LK=[]

    for xx in LBcom:

        try:

            copy_obj = xx.Copy()
    
            point1 = vtpnt(0,0,0)
    
            point2 = vtpnt(0,K*1000000,0)
    
            copy_obj.Move(point1,point2)


            LK.append(copy_obj)

        except Exception as e:

            print(f"[é”™è¯¯] ç§»åŠ¨å¯¹è±¡{xx.Handle}å¤±è´¥: {e}")               

    return LK



#æµ‹é‡å·²æœ‰æ–‡å­—é•¿åº¦

def celiang_wenzichangdu(TEXTCOM):

    text_copy = TEXTCOM.Copy()

    text_copy.Alignment = 2

    text_copy.TextAlignmentPoint =vtpnt(0,0,0)

    chang = abs(text_copy.InsertionPoint[0])

    text_copy.Delete()

    return chang

#æµ‹é‡æ–°å†™æ–‡å­—é•¿åº¦

def celiang_wenzichangdu_write(ZF,style="å›¾ç­¾",height=270,scalefactor=0.8):

    #æ ¹æ®å­—ç¬¦ä¸²æŒ‰æ ·å¼å­—é«˜å®½åº¦å› å­å†™å…¥cadåçš„æµ‹é‡é•¿åº¦

    text_obj = acad.ActiveDocument.ModelSpace.AddText(ZF, vtpnt(0,0,0), height)

    text_obj.StyleName = style

    text_obj.ScaleFactor =scalefactor #å®½åº¦å› å­

    chang = celiang_wenzichangdu(text_obj)

    text_obj.Delete()

    return chang




##æ¸…ç©ºæ–‡ä»¶å¤¹
def qingkong_wenjianjia(FolderPath):

     #æ¸…ç©ºæ–‡ä»¶å¤¹B
    folder_path_1 = FolderPath 

    # éå†æ–‡ä»¶å¤¹ä¸­çš„æ‰€æœ‰æ–‡ä»¶
    for filename in os.listdir(folder_path_1):
        
        file_pathx = os.path.join(folder_path_1, filename)
        
        # ç¡®ä¿å®ƒæ˜¯ä¸€ä¸ªæ–‡ä»¶è€Œä¸æ˜¯æ–‡ä»¶å¤¹
        if os.path.isfile(file_pathx):
            
            os.remove(file_pathx)  # åˆ é™¤æ–‡ä»¶

        print(f"{FolderPath}æ–‡ä»¶å¤¹å·²æ¸…ç©º")


#&&% è¿”å›å¯¹è±¡å¤–åŒ…ç›’çš„é•¿ï¼Œå®½ï¼Œæ¨ªç«–å‘ï¼Œè§’ç‚¹ä¿¡æ¯

def get_bbox_info(com_obj):
    """
    è·å–ä¼ å…¥ AutoCAD COM å¯¹è±¡ï¼ˆå¦‚ Lineã€Circleã€BlockReference ç­‰ï¼‰çš„å¤–åŒ…ç›’ä¿¡æ¯ï¼Œ
    å¹¶è®¡ç®—å…¶é•¿ï¼ˆlengthï¼‰ã€å®½ï¼ˆwidthï¼‰ä»¥åŠæ¨ªå‘æˆ–ç«–å‘ï¼ˆorientationï¼‰çŠ¶æ€ã€‚

    å‚æ•°ï¼š
      com_obj -- ä»»æ„æ”¯æŒ GetBoundingBox() æ–¹æ³•çš„ AutoCAD COM å¯¹è±¡

    è¿”å›ï¼š
      (length, width, orientation)
        length      -- å¤–åŒ…ç›’è¾ƒé•¿ä¸€è¾¹çš„é•¿åº¦ï¼ˆåœ¨ X/Y å¹³é¢ä¸Šï¼‰
        width       -- å¤–åŒ…ç›’è¾ƒçŸ­ä¸€è¾¹çš„é•¿åº¦ï¼ˆåœ¨ X/Y å¹³é¢ä¸Šï¼‰
        orientation -- å­—ç¬¦ä¸²ï¼š
                         "horizontal" è¡¨ç¤º X æ–¹å‘è·¨åº¦ â‰¥ Y æ–¹å‘è·¨åº¦ï¼Œ
                         "vertical"   è¡¨ç¤º Y æ–¹å‘è·¨åº¦ >  X æ–¹å‘è·¨åº¦
    å¦‚æœå¯¹è±¡ä¸æ”¯æŒ GetBoundingBoxï¼Œä¼šæŠ›å‡ºå¼‚å¸¸ï¼›ä¹Ÿå¯æ ¹æ®éœ€è¦è‡ªè¡Œæ•è·å¤„ç†ã€‚
    """
 


   # è°ƒç”¨ GetBoundingBox æ–¹æ³•ï¼Œè¿”å›ä¸¤ä¸ªç‚¹ï¼šminPtã€maxPt
    # minPtã€maxPt éƒ½æ˜¯ 3 å…ƒç´ çš„ tuple æˆ– listï¼Œå½¢å¦‚ (x, y, z)
    




    try:

        minPt, maxPt = com_obj.GetBoundingBox()

    except Exception as e:

        print(f"è·å–å¤–åŒ…ç›’å¤±è´¥: {e}")               

        return None


    # è®¡ç®— Xã€Y æ–¹å‘è·¨åº¦
    dx = maxPt[0] - minPt[0]
    dy = maxPt[1] - minPt[1]

    # å°†è¾ƒå¤§å€¼å®šä¹‰ä¸º lengthï¼Œè¾ƒå°å€¼å®šä¹‰ä¸º width
    length = max(dx, dy)
    width  = min(dx, dy)

    # åˆ¤æ–­æ¨ªå‘ï¼ˆX è·¨åº¦ â‰¥ Y è·¨åº¦ï¼‰è¿˜æ˜¯ç«–å‘ï¼ˆY è·¨åº¦ > X è·¨åº¦ï¼‰
    if dx >= dy:
        orientation = "horizontal"
    else:
        orientation = "vertical"

    return minPt, maxPt,length, width, orientation

#&&% åˆ¤æ–­å¯¹è±¡å¤–åŒ…ç›’çš„æ¨ªç«–

def bbox_orientation_flag(com_obj):
    """
    åˆ¤æ–­ä»»æ„ COM å¯¹è±¡çš„å¤–åŒ…ç›’æ˜¯ç«–å‘ã€æ¨ªå‘è¿˜æ˜¯æ­£æ–¹å½¢ï¼š
      - å¦‚æœ Y æ–¹å‘è·¨åº¦ > X æ–¹å‘è·¨åº¦ï¼Œè¿”å› 1ï¼ˆç«–å‘ï¼‰
      - å¦åˆ™ï¼ˆåŒ…æ‹¬ X æ–¹å‘è·¨åº¦ >= Y æ–¹å‘è·¨åº¦ï¼‰ï¼Œè¿”å› 0
        â€”â€” å³å½“å¤–åŒ…ç›’ä¸ºæ­£æ–¹å½¢ï¼ˆX è·¨åº¦ == Y è·¨åº¦ï¼‰æ—¶ï¼Œä¹Ÿè¿”å› 0

    å‚æ•°ï¼š
      com_obj -- æ”¯æŒ GetBoundingBox() çš„ AutoCAD COM å¯¹è±¡

    è¿”å›ï¼š
      int -- ç«–å‘è¿”å› 1ï¼Œæ¨ªå‘æˆ–æ­£æ–¹å½¢è¿”å› 0
    """
    # è·å–å¤–åŒ…ç›’çš„æœ€å°ç‚¹å’Œæœ€å¤§ç‚¹
    min_pt, max_pt = com_obj.GetBoundingBox()
    # è®¡ç®— Xã€Y æ–¹å‘è·¨åº¦
    dx = abs(max_pt[0] - min_pt[0])
    dy = abs(max_pt[1] - min_pt[1])
    # å¦‚æœæ˜¯ç«–å‘ï¼ˆdy > dxï¼‰ï¼Œè¿”å› 1ï¼›å¦åˆ™ï¼ˆæ¨ªå‘æˆ–æ­£æ–¹å½¢ï¼‰è¿”å› 0
    return 1 if dy > dx else 0

#&&% è·å–å¤šä¸ªå¯¹è±¡çš„å¤–åŒ…ç›’æ•°æ®
def group_bbox_corners(com_objs):
    """
    è®¡ç®—ä¸€ç»„ COM å¯¹è±¡çš„æ•´ä½“å¤–åŒ…ç›’ï¼Œå¹¶æŒ‰é¡ºåºè¿”å›å››ä¸ªè§’ç‚¹åæ ‡ï¼š
      1. å·¦ä¸‹è§’ (minX, minY)
      2. å³ä¸Šè§’ (maxX, maxY)
      3. å·¦ä¸Šè§’ (minX, maxY)
      4. å³ä¸‹è§’ (maxX, minY)

    å‚æ•°ï¼š
      com_objs -- å¯è¿­ä»£çš„ä¸€ç»„æ”¯æŒ GetBoundingBox() æ–¹æ³•çš„ COM å¯¹è±¡åˆ—è¡¨

    è¿”å›ï¼š
      å››å…ƒç»„ï¼š(
        (minX, minY, z),
        (maxX, maxY, z),
        (minX, maxY, z),
        (maxX, minY, z)
      )
      å…¶ä¸­ z å–è‡ªå„å¯¹è±¡å¤–åŒ…ç›’çš„ z å€¼èŒƒå›´ï¼Œç»Ÿä¸€ä½¿ç”¨æœ€å° zï¼ˆå¦‚éœ€ä¸åŒï¼Œå¯æŒ‰éœ€è°ƒæ•´ï¼‰
    """
    # åˆå§‹åŒ–ä¸ºæç«¯å€¼
    global_min_x = float('inf')
    global_min_y = float('inf')
    global_max_x = float('-inf')
    global_max_y = float('-inf')
    global_min_z = float('inf')  # å¦‚æœéœ€è¦ç»Ÿä¸€ z å€¼ï¼Œå¯ä½¿ç”¨æœ€å° z
    # å¦‚æœä¸å…³å¿ƒ zï¼Œåªè¿”å› 0 å³å¯ã€‚è¿™é‡Œä»¥æœ€å° z ä½œä¸ºç»Ÿä¸€ z
    for obj in com_objs:
        try:
            min_pt, max_pt = obj.GetBoundingBox()
        except Exception:
            # å¦‚æœå¯¹è±¡ä¸æ”¯æŒ GetBoundingBoxï¼Œè·³è¿‡
            continue

        x1, y1, z1 = min_pt
        x2, y2, z2 = max_pt

        # æ›´æ–° X/Y extremes
        if x1 < global_min_x:
            global_min_x = x1
        if y1 < global_min_y:
            global_min_y = y1
        if x2 > global_max_x:
            global_max_x = x2
        if y2 > global_max_y:
            global_max_y = y2

        # æ›´æ–° Z extremesï¼ˆå¦‚æœéœ€è¦ç»Ÿä¸€ä½¿ç”¨æœ€å° zï¼‰
        if z1 < global_min_z:
            global_min_z = z1

    # å¦‚æœæ‰€æœ‰å¯¹è±¡éƒ½è¢«è·³è¿‡ï¼ˆåˆ—è¡¨ä¸ºç©ºæˆ–éƒ½ä¸æ”¯æŒ GetBoundingBoxï¼‰ï¼Œç›´æ¥è¿”å› None
    if global_min_x == float('inf'):
        return None

    # é‡‡ç”¨ global_min_z ä½œä¸ºæ‰€æœ‰è§’ç‚¹çš„ z åˆ†é‡
    z = global_min_z

    # å·¦ä¸‹ã€å³ä¸Šã€å·¦ä¸Šã€å³ä¸‹
    bottom_left  = (global_min_x, global_min_y, z)
    top_right    = (global_max_x, global_max_y, z)
    top_left     = (global_min_x, global_max_y, z)
    bottom_right = (global_max_x, global_min_y, z)

    return bottom_left, top_right, top_left, bottom_right

#&&% zipçš„ç”¨æ³•
"""
seq_x0  = (x0, y0, z0)
P_start = (dx, dy, dz)
è¦å¾—åˆ°æ–°åæ ‡ (x0+dx, y0+dy, z0+dz)ï¼Œå°±å¯ä»¥å†™ï¼š
seq_x0 = tuple(a + b for a, b in zip(seq_x0, P_start))
zip(seq_x0, P_start) ä¼šäº§å‡º (x0, dx), (y0, dy), (z0, dz)
a + b for a, b in ... å°±å¯¹æ¯ä¸€å¯¹åšç›¸åŠ 
æœ€åç”¨ tuple(...) æŠŠç»“æœæ”¶æˆä¸‰å…ƒç»„
å¯¹ä¸¤ä¸ªåˆ—è¡¨æ±‚å’Œ
xs = [10, 20, 30]
ys = [1, 2, 3]
sums = [x + y for x, y in zip(xs, ys)]
# sums == [11, 22, 33]
å¹¶è¡Œéå†ä¸‰ç»„æ•°æ®
names = ["A", "B", "C"]
ages  = [30, 25, 40]
scores= [85, 92, 78]
for n, a, s in zip(names, ages, scores):
    print(f"{n} å¹´é¾„{a} åˆ†æ•°{s}")
è§£å‹
å¦‚æœæœ‰ä¸€ä¸ªåˆ—è¡¨ pairs = [(1,4),(2,5),(3,6)]ï¼Œè¦æ‹†æˆä¸¤ä¸ªåˆ—è¡¨ï¼š
a, b = zip(*pairs)
# a == (1,2,3), b == (4,5,6)

"""

#&&% ä»ä¸¤ç‚¹ç»˜åˆ¶çŸ©å½¢

def draw_rectangle_by_corners(p1: tuple[float, float, float],
                              p2: tuple[float, float, float],
                              layer_name: str = "æµ‹è¯•è¾…åŠ©",
                              width: float = 0.0,
                              color: int = 256) -> object:
    """
    åŸºäºä¸¤ç‚¹ç»˜åˆ¶ä¸€ä¸ªé—­åˆçŸ©å½¢ï¼š
      - p1: (x_min, y_min, z)
      - p2: (x_max, y_max, z)
    è°ƒç”¨ draw_lwpolyline ç»˜åˆ¶å››è¾¹é—­åˆå¤šæ®µçº¿ï¼Œè¿”å›æ–°åˆ›å»ºçš„å¤šæ®µçº¿å¯¹è±¡ã€‚
    """
    # å·¦ä¸‹ã€å³ä¸‹ã€å³ä¸Šã€å·¦ä¸Š
    x1, y1, z1 = p1
    x2, y2, z2 = p2
    coords = [
        (x1, y1, z1),
        (x2, y1, z1),
        (x2, y2, z1),
        (x1, y2, z1)
    ]
    # ç»˜åˆ¶é—­åˆçŸ©å½¢
    rect = draw_lwpolyline(
        coords3d=coords,
        layer_name=layer_name,
        width=width,
        color=color,
        closed=True
    )
    return rect

#&&% è·å–å¤–åŒ…ç›’ä¸­å¿ƒ
def bbox_center_2(e):
    # GetBoundingBox è¿”å›ä¸¤ä¸ª Point (minPt, maxPt)
    min_pt, max_pt = e.GetBoundingBox()
    x1, y1, _ = tuple(min_pt)
    x2, y2, _ = tuple(max_pt)
    return ((x1 + x2) / 2, (y1 + y2) / 2)

def bbox_center_3(ent):
    """è¿”å›å®ä½“å¤–åŒ…ç›’ä¸­å¿ƒ (cx, cy, cz)"""
    mn, mx = ent.GetBoundingBox()
    return ((mn[0] + mx[0]) / 2.0,
            (mn[1] + mx[1]) / 2.0,
            (mn[2] + mx[2]) / 2.0)


#&&&&%% ç¬¬å…«éƒ¨åˆ†   CADæ–‡ä»¶æ“ä½œ


#_________________________________________________________________________________________________________________________

#  æ¨¡å—ä½¿ç”¨è¯´æ˜

"""
è¯¥æ¨¡å—è§£å†³CADæ–‡ä»¶çš„è½¬æ¢ã€æ‰“å¼€ã€å…³é—­ã€æ–‡ä»¶ä¹‹é—´çš„å¤åˆ¶ã€åˆ†è§£ç»„åˆç­‰é—®é¢˜ 

"""

#  ä¸»å‡½æ•°
#  (1)
# æ–‡ä»¶è½¬æˆt7ã€t3æ ¼å¼

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°
"""
zhuancheng_t7()
zhuancheng_t3()
åœ¨li()è¿æ¥æ¿€æ´»æ–‡ä»¶åï¼Œç›´æ¥æ‰§è¡Œè¯¥å‘½ä»¤å³å¯è½¬æ¢
ä¸€èˆ¬ç†è§£æ‰€å¾—æ–‡ä»¶å’Œå½“å‰æ–‡ä»¶åŒæ–‡ä»¶å¤¹ï¼Œä½†è¿™æ¬¡æµ‹è¯•ç»“æœå´åœ¨æ–‡ä»¶å¤¹
C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
"""    
#&&% å¯¹è·å¾—çš„ä¸¤è¡Œå¤§åœ°åæ ‡æ•°æ®å»ºç«‹å­—å…¸

"""
å¤§é‡æ•°æ®çš„æ•´ç†è¯»å–ï¼ŒAIä¸èƒ½ç»å¯¹ä¿è¯æ­£ç¡®

"""

def build_J_points_from_selected_texts(LB, n_points=61, prefix_x="30", prefix_y="37"):
    num_re = re.compile(r"-?\d+(?:\.\d+)?")

    def first_number(s):
        m = num_re.search(str(s))
        if not m:
            raise ValueError(f"æ— æ³•è§£ææ•°å­—: {s!r}")
        return float(m.group(0))

    items = []
    for ent in LB:
        txt, ip = get_text_and_ip(ent)
        if not txt or not ip:
            continue
        txt = str(txt).strip().replace(" ", "")
        x0, y0, *_ = ip
        if txt.startswith(prefix_x):
            items.append(("x", first_number(txt), float(y0), float(x0)))
        elif txt.startswith(prefix_y):
            items.append(("y", first_number(txt), float(y0), float(x0)))

    # ä»ä¸Šåˆ°ä¸‹ï¼ˆyé™åºï¼‰ï¼ŒåŒè¡ŒæŒ‰xå‡åºç¨³å®š
    items.sort(key=lambda t: (-t[2], t[3]))
    xs = [v for typ, v, _, _ in items if typ == "x"]
    ys = [v for typ, v, _, _ in items if typ == "y"]

    if len(xs) != n_points or len(ys) != n_points:
        raise RuntimeError(f"X={len(xs)} / Y={len(ys)} ä¸æœŸæœ› {n_points} ä¸ç¬¦ã€‚")

    pts_dict = {f"J{i+1}": (xs[i], ys[i]) for i in range(n_points)}
    pts_list = [(f"J{i+1}", xs[i], ys[i]) for i in range(n_points)]
    return pts_dict, pts_list


#ä»ç•Œç‚¹å¤§åœ°åæ ‡è®¡ç®—ç»çº¬åº¦
def convert_pts_dict_to_latlon(pts_dict, central_lon=111):
    """
    è¾“å…¥: pts_dict = {'J1': (N, E), 'J2': (N, E), ...}
          N=åŒ—åæ ‡, E=ä¸œåæ ‡ (CGCS2000 é«˜æ–¯-å…‹å•æ ¼æŠ•å½±)
    è¾“å‡º: geo_dict = {'J1': (lon, lat), 'J2': (lon, lat), ...}  (ç»åº¦, çº¬åº¦, å•ä½:åº¦)
    """
    a = 6378137.0
    f = 1 / 298.257222101
    FE = 500000
    FN = 0
    e2 = 2*f - f*f
    e_prime_2 = e2 / (1.0 - e2)

    def one_point(N, E):
        m = N - FN
        mu = m / (a * (1 - e2/4 - 3*e2**2/64 - 5*e2**3/256))
        e1 = (1 - math.sqrt(1 - e2)) / (1 + math.sqrt(1 - e2))
        Bf = (mu
              + (3*e1/2 - 27*e1**3/32) * math.sin(2*mu)
              + (21*e1**2/16 - 55*e1**4/32) * math.sin(4*mu)
              + (151*e1**3/96) * math.sin(6*mu)
              + (1097*e1**4/512) * math.sin(8*mu))
        x_prime = E - FE
        cosBf = math.cos(Bf)
        tanBf = math.tan(Bf)
        eta2 = e_prime_2 * (cosBf**2)
        Nf = a / math.sqrt(1 - e2 * (math.sin(Bf)**2))
        Mf = a * (1 - e2) / (1 - e2 * (math.sin(Bf)**2))**1.5
        lat = (Bf
               - tanBf/(2* Mf* Nf) * (x_prime**2)
               + tanBf/(24* Mf* Nf**3) * (5 + 3*tanBf**2 + eta2 - 9*tanBf**2 * eta2) * (x_prime**4)
               - tanBf/(720* Mf* Nf**5) * (61 + 90*tanBf**2 + 45*tanBf**4) * (x_prime**6))
        lon = (math.radians(central_lon)
               + x_prime/(Nf* cosBf)
               - (1 + 2*tanBf**2 + eta2) * (x_prime**3)/(6* Nf**3* cosBf)
               + (5 + 28*tanBf**2 + 24*tanBf**4 + 6* eta2 + 8*tanBf**2 * eta2) * (x_prime**5)/(120* Nf**5* cosBf))
        return math.degrees(lon), math.degrees(lat)

    geo_dict = {}
    for k, (N, E) in pts_dict.items():
        geo_dict[k] = one_point(N, E)

    return geo_dict

# æ–‡ä»¶åŸºæœ¬æ“ä½œ



"""
ä½¿ç”¨å…±åŒçš„æ–‡ä»¶å…¨å±€å˜é‡acad,mp,doc,spæ˜¯æˆ‘ä»¬ç¼–åˆ¶è„šæœ¬æ§åˆ¶ä¸åŒæ–‡ä»¶çš„åŸºç¡€            


"""


    

def rename_time(output_path):#ä¸è¦†ç›–ä»å½“å‰æ—¶é—´ç»™å¸¦è·¯å¾„æ–‡ä»¶åå‘½å

    from datetime import datetime
    now = datetime.now()
    time_str = now.strftime("%m-%d-%H-%M")

    dir_name, file_name = os.path.split(output_path)
    name, ext = os.path.splitext(file_name)
    new_file_name = f"{name}_{time_str}{ext}"
    new_path = os.path.join(dir_name, new_file_name)
    return new_path

#ä¾›å‚è€ƒä¼˜åŒ–
def open_dwg(path: str, visible: bool = True):
    # åˆå§‹åŒ– COMï¼ˆå°¤å…¶åœ¨å¤šçº¿ç¨‹æˆ–éè„šæœ¬äº¤äº’ç¯å¢ƒä¸‹æ¨èè°ƒç”¨ï¼‰
    pythoncom.CoInitialize()

    # å¯åŠ¨å¹¶è¿æ¥åˆ° AutoCAD åº”ç”¨
    # Dispatch ä¼šåœ¨å·²æœ‰å®ä¾‹ä¸Šå¤ç”¨ï¼ŒDispatchEx ä¼šå¼ºåˆ¶æ–°å¼€ä¸€ä¸ªå®ä¾‹
    acad = Dispatch("AutoCAD.Application")  # æˆ–è€… DispatchEx("AutoCAD.Application")

    # å¯è§æ€§ï¼ˆTrue æ—¶ä¼šå¼¹å‡ºç•Œé¢ï¼‰
    acad.Visible = visible

    # æ‰“å¼€ DWG æ–‡æ¡£
    doc = acad.Documents.Open(path)

    print(f"å·²æ‰“å¼€ï¼š{doc.Name}")

    return acad, doc









##æ‰“å¼€æ–‡ä»¶
def Open_By_Omission_wenjian(file_path):

    """
    ä¸èƒ½æŒ‡æœ›é€šè¿‡è¿™å‡½æ•°æ§åˆ¶æ‰€æœ‰æ‰“å¼€æ–‡ä»¶æ—¶é‡åˆ°çš„çª—å£è·³å‡ºï¼ŒæŠ¥é”™ç­‰ï¼Œå®ƒä»¬åº”è¯¥åœ¨åˆ«çš„åœ°æ–¹è§£å†³ï¼Œä¾‹å¦‚å­—ä½“é—®é¢˜ï¼ŒæŠ¥é”™çª—å£ï¼Œä»£ç†é”™è¯¯ç­‰ç­‰

    """
    
    t1 = time.time()
    
    max_retries = 3

    for attempt in range(max_retries):
        
        try:
            
            # å°è¯•æ‰“å¼€æ–‡ä»¶
            new_doc = acad.Documents.Open(file_path)
            
            print(f"Opened file: {file_path}")
            
            t2 = time.time()
            
            print("æ–‡ä»¶æ‰“å¼€è€—æ—¶:", t2 - t1, "ç§’")

            li() #æ‰€æœ‰çš„å‡½æ•°éƒ½ä½¿ç”¨åŒæ ·çš„å…¨å±€å˜é‡acad,mp,docç­‰ï¼Œè¿™æ ·åœ¨ä¸åŒçš„æ–‡ä»¶ä¸Šå‡½æ•°ä»ç„¶é€šç”¨
            
            return new_doc
        
        except Exception as e:
            
            print(f"Attempt {attempt + 1} failed: {e}")
            pass
            
            time.sleep(2)  # ç­‰å¾…2ç§’åå†æ¬¡å°è¯•

    # å¦‚æœæ‰€æœ‰å°è¯•éƒ½å¤±è´¥
    t2 = time.time()
    
    print("æ–‡ä»¶æ‰“å¼€è€—æ—¶:", t2 - t1, "ç§’")
    
    return  None



##æ‰“å¼€æ–‡ä»¶æ—¶è·³è¿‡ç¼ºå­—ä½“çª—å£


def f1_openfile_getwindaow(timeout_event, done_event,
                     file_path):
    pythoncom.CoInitialize()
    try:

        

        print("çº¿ç¨‹1å¯åŠ¨")

        li()

        time.sleep(2)

        Open_By_Omission_wenjian(file_path)

        time.sleep(10)

        timeout_event.wait()          # ç­‰å¾…çº¿ç¨‹2å®Œå·¥ / æˆ–è°ƒåº¦å™¨è¶…æ—¶

    except Exception as e:
        print("f1_openfile_getwindaow:", e, traceback.format_exc())
    finally:
        pythoncom.CoUninitialize()
        done_event.set()
        


def f2_delwindaow(timeout_event, done_event):
    pythoncom.CoInitialize()
    try:
        print("çº¿ç¨‹2å¯åŠ¨")
        # é—ªåŠ¨çª—å£ï¼ˆç¤ºä¾‹ï¼šåˆ·æ–°ä¸€ä¸‹çª—å£åˆ—è¡¨ï¼‰
       
        time.sleep(1)

        BT = list_open_window_titles()
        print("å½“å‰çª—å£æ ‡é¢˜ï¼š", BT)

        if 'ç¼ºå°‘ SHX æ–‡ä»¶' in BT:
            # ç‚¹å‡»å¯¹è¯æ¡†ä¸­çš„â€œå¿½ç•¥â€æŒ‰é’®ï¼Œå‡è®¾å®ƒåœ¨å¯¹è¯æ¡†ä¸­çš„å¤§è‡´åæ ‡ (148, 220)
            click_in_window("ç¼ºå°‘ SHX æ–‡ä»¶", 148, 220, click_titlebar=True)
            print("ğŸ–± å·²ç‚¹å‡»â€œå¿½ç•¥â€æŒ‰é’®")
        else:
            print("â„¹ï¸ æœªæ£€æµ‹åˆ°â€œç¼ºå°‘ SHX æ–‡ä»¶â€å¯¹è¯æ¡†")

    except Exception as e:
        print("f2_delwindaow å¼‚å¸¸:", e, traceback.format_exc())
    finally:
        # å…ˆå‘ä¿¡å·ç»™çº¿ç¨‹1 é€€å‡ºé˜»å¡ï¼Œå†é€šçŸ¥æ€»æ§å·²ç»å®Œæˆ
        timeout_event.set()
        pythoncom.CoUninitialize()
        done_event.set()




def Open_file_nic(file_path ):

    """
    å®æµ‹ä¼¼ä¹äººå·¥æ‰“å¼€æ–‡ä»¶æ—¶ä¼šè·³å‡ºSHXå­—ä½“ç¼ºå¤±çª—å£ï¼Œè€ŒOpen_By_Omission_wenjianå‘½ä»¤ä¸ä¼šï¼Œä¹Ÿè®¸è¿™ä¸ªå‘½ä»¤æ˜¯å¤šä½™çš„ï¼Œä½†ä»ç„¶ä¿ç•™è¿™ä¸ªæœºåˆ¶ï¼Œå› ä¸ºæˆ‘ä»¬ä¸èƒ½æ–­å®š
    
    """

    ok = run_dual_threads(

        f1=f1_openfile_getwindaow,

        f2=f2_delwindaow,

        f1_args=(file_path,),

        f2_args=(),

        timeout_sec=600
    )
    if ok:
        print(f"ğŸ‰ æˆåŠŸæ‰“å¼€æ–‡ä»¶ â†’ {file_path}")
    else:
        print("ğŸš¨ æ‰“å¼€æ–‡ä»¶è¶…æ—¶ / å¤±è´¥")





##ä¿å­˜æ–‡ä»¶#å¦å­˜å°±æ˜¯doc.SaveAs()

def savefile():

    doc.Save()


##å…³é—­æ–‡ä»¶(åˆ«ä¹±åˆ )

def guanbifile():

    doc.Close()



##ç¡®ä¿å…³é—­å½“å‰æ–‡ä»¶


def close_current_drawing_safely():
    """
    å®‰å…¨å…³é—­å½“å‰ DWG æ–‡ä»¶ï¼Œç¡®ä¿ç¡®å®å…³é—­å¹¶é‡æ–°è¿æ¥ã€‚
    æœ€å¤šå°è¯• 3 æ¬¡ã€‚
    """
    
    try:
        Name1 = doc.Name
    except:
        print("[è­¦å‘Š]ï¸ å½“å‰ doc æ— æ³•è·å–åç§°ï¼Œå¯èƒ½æœªè¿æ¥ã€‚")

        li()

        Name1 = doc.Name
    
    for attempt in range(1, 4):  # æœ€å¤šå°è¯•3æ¬¡
        print(f"ğŸ”„ ç¬¬ {attempt} æ¬¡å°è¯•å…³é—­ '{Name1}'")
        close_dwg_by_name(Name1)

        li()  # é‡æ–°è¿æ¥ acadã€docã€mpã€sp ç­‰
        try:
            Name2 = doc.Name
        except:
            Name2 = None

        if Name2 != Name1:
            print(f"ğŸŸ¢ å·²ç¡®è®¤æ–‡ä»¶ '{Name1}' å…³é—­ï¼Œå½“å‰æ‰“å¼€æ–‡ä»¶ä¸º '{Name2}'")
            li()  # å†æ‰§è¡Œä¸€æ¬¡ï¼Œç¡®ä¿å˜é‡æ­£ç¡®
            return
        else:
            print("[è­¦å‘Š]ï¸ æ–‡ä»¶ä»æœªå…³é—­ï¼Œç»§ç»­å°è¯•...")

    print(f"[é”™è¯¯] å¤šæ¬¡å°è¯•ä»æœªæˆåŠŸå…³é—­ '{Name1}'ï¼Œè¯·æ‰‹åŠ¨æ£€æŸ¥ã€‚")



#&&% 2 ä¸¤ä¸ªæ–‡ä»¶A,Bçš„æ“ä½œ
"""
æ–‡ä»¶ä½œä¸ºå—æ’å…¥ä¸ç¨³å®šï¼Œç”šè‡³åœ¨äº¤äº’æ¨¡å¼å¯ä»¥åŒæ ·çš„ä»£ç è½¬ä¸ºå‡½æ•°æ¨¡å¼å´ä¸è¡Œï¼Œåå€’æ˜¯ä½¿ç”¨å…¨é€‰å¤åˆ¶ç²˜è´´æ¯”è¾ƒé è°±

é™¤äº†ä½¿ç”¨è¿™ä¸ªå‘½ä»¤åˆå¹¶æ–‡ä»¶ï¼Œè¿˜æœ‰è½¬æ¢ä¸ºçº¯æ•°æ®å‘½ä»¤

        
    """

#æ–°å»ºä¸€ä¸ªç©ºç™½æ–‡ä»¶(ä¸æ‰“å¼€)
def create_new_dwg_file(name_with_path):
    """
    åˆ›å»ºä¸€ä¸ªæ–°çš„ DWG æ–‡ä»¶ï¼Œå¹¶ä¿å­˜ä¸ºæŒ‡å®šå®Œæ•´è·¯å¾„ name_with_pathï¼ˆåº”ä»¥ .dwg ç»“å°¾ï¼‰ã€‚

    å‚æ•°ï¼š
        name_with_path - å®Œæ•´çš„æ–‡ä»¶ä¿å­˜è·¯å¾„ï¼ˆä¾‹å¦‚ D:\CADXT\Export\æ–°å›¾01.dwgï¼‰
    """
   
    # æ£€æŸ¥å¹¶åˆ›å»ºç›®å½•ï¼ˆå¦‚æœä¸å­˜åœ¨ï¼‰
    folder = os.path.dirname(name_with_path)
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"ğŸ“ å·²åˆ›å»ºæ–‡ä»¶å¤¹ï¼š{folder}")

    # æ–°å»º DWG æ–‡ä»¶
    acad.Documents.Add()
    time.sleep(1)

    # è¿æ¥å½“å‰æ–‡æ¡£
    new_doc = acad.ActiveDocument

    # ä¿å­˜åˆ°æŒ‡å®šè·¯å¾„
    new_doc.SaveAs(name_with_path)

    new_doc.Close()
    li()
    print(f"[OK] æ–°å»ºå¹¶ä¿å­˜ DWG æ–‡ä»¶ï¼š{name_with_path}")

#æ–°å»ºä¸€ä¸ªç©ºç™½æ–‡ä»¶
def create_new_dwg_file_no(name_with_path):
    """
    åˆ›å»ºä¸€ä¸ªæ–°çš„ DWG æ–‡ä»¶ï¼Œå¹¶ä¿å­˜ä¸ºæŒ‡å®šå®Œæ•´è·¯å¾„ name_with_pathï¼ˆåº”ä»¥ .dwg ç»“å°¾ï¼‰ã€‚

    å‚æ•°ï¼š
        name_with_path - å®Œæ•´çš„æ–‡ä»¶ä¿å­˜è·¯å¾„ï¼ˆä¾‹å¦‚ D:\CADXT\Export\æ–°å›¾01.dwgï¼‰
    """
   
    # æ£€æŸ¥å¹¶åˆ›å»ºç›®å½•ï¼ˆå¦‚æœä¸å­˜åœ¨ï¼‰
    folder = os.path.dirname(name_with_path)
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"ğŸ“ å·²åˆ›å»ºæ–‡ä»¶å¤¹ï¼š{folder}")

    # æ–°å»º DWG æ–‡ä»¶
    acad.Documents.Add()
    time.sleep(1)

    # è¿æ¥å½“å‰æ–‡æ¡£
    new_doc = acad.ActiveDocument

    # ä¿å­˜åˆ°æŒ‡å®šè·¯å¾„
    new_doc.SaveAs(name_with_path)
  
    li()
    print(f"[OK] æ–°å»ºå¹¶ä¿å­˜ DWG æ–‡ä»¶ï¼š{name_with_path}")



#  ä¸»å‡½æ•°
#  (1)
# å…³é—­æ–‡ä»¶

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

def close_all_except_active_safe():#å…³é—­é™¤å½“å‰æ¿€æ´»æ–‡æ¡£å¤–çš„æ‰€æœ‰ DWG æ–‡ä»¶
    """
    æ›´ç¨³å®šåœ°å…³é—­é™¤å½“å‰æ¿€æ´»æ–‡æ¡£å¤–çš„æ‰€æœ‰ DWG æ–‡ä»¶ï¼Œé¿å… COM å¯¹è±¡æ–­é“¾ã€‚
    """
    try:
        active_name = acad.ActiveDocument.Name
        all_names = [acad.Documents.Item(i).Name for i in range(acad.Documents.Count)]
        closed = 0

        for name in all_names:
            if name != active_name:
                try:
                    doc = acad.Documents.Item(name)
                    doc.Close(False)  # ä¸ä¿å­˜ç›´æ¥å…³é—­
                    print(f"[å·²å…³é—­] {name}")
                    closed += 1
                except Exception as e:
                    print(f"[è­¦å‘Š] æ— æ³•å…³é—­ {name}: {e}")

        print(f"[OK] æˆåŠŸå…³é—­ {closed} ä¸ªæ–‡æ¡£ï¼Œä»…ä¿ç•™ {active_name}")

    except Exception as e:
        print(f"[é”™è¯¯] å®‰å…¨å…³é—­å¤±è´¥ï¼š{e}")


## æµ‹è¯•ç¤ºä¾‹
##close_all_except_active_safe()
##ğŸ—‚ï¸ å·²å…³é—­ï¼šDrawing3.dwg
##ğŸ—‚ï¸ å·²å…³é—­ï¼šDrawing4.dwg
##ğŸ—‚ï¸ å·²å…³é—­ï¼šç©ºç™½.dwg
##ğŸ—‚ï¸ å·²å…³é—­ï¼šæµ‹è¯•1.dwg
##[OK] æˆåŠŸå…³é—­ 4 ä¸ªæ–‡æ¡£ï¼Œä»…ä¿ç•™ cs.dwg
##__________
    

# åŒæ–‡ä»¶æ“ä½œ
"""
åˆšå¼€å§‹å¯åŠ¨å¤©æ­£ï¼Œé»˜è®¤æ‰“å¼€äº†ä¸€ä¸ªdrawing1çš„æ–‡ä»¶ï¼Œä½¿ç”¨create_new_dwg_fileæ–°å»ºä¸€ä¸ªæ–‡ä»¶åï¼Œæ–°å»ºçš„æ–‡ä»¶ä¼šå…³é—­ï¼Œä»ç„¶æ˜¯drawing1åœ¨ï¼Œä»éœ€li()è¿æ¥ä¸€ä¸‹

æ­¤æ—¶ä½¿ç”¨Open_By_Omission_wenjianæ‰“å¼€æ–°æ–‡ä»¶ï¼Œå°±ä¼šåŒæ—¶è‡ªåŠ¨å…³é—­drawing1ã€‚åŒæ ·éœ€è¦li()è¿æ¥ä¸€ä¸‹
å› æ­¤æˆ‘ä»¬é¢ä¸´ä¸¤ç§å¯èƒ½ï¼Œä¸€ä¸ªæ˜¯å¤©æ­£çš„drawing1ä½œä¸ºåŸºæœ¬è¿æ¥çŠ¶æ€ï¼Œæ­¤æ—¶æ–‡ä»¶æ•°ä¸º1ï¼Œä½†è¦æ³¨æ„create_new_dwg_fileæ–°å»º ä¸æ”¹å˜è¿™ç§çŠ¶æ€ï¼Œè€Œ Open_By_Omission_wenjianæ‰“å¼€æ¡Œé¢å°±ä¼šå˜æˆåªæœ‰1ä¸ªæ–‡ä»¶

å¦ä¸€ç§æƒ…å†µæ—¶æˆ‘ä»¬ç”¨è‡ªå·±ç³»ç»Ÿçš„â€œç©ºç™½.dwgâ€ä½œä¸ºå½“å‰æ¡Œé¢çš„å”¯ä¸€åŸºç¡€æ–‡ä»¶ã€‚
äº‹å®ä¸Šï¼Œåœ¨è¿™ä¸ªæ¡Œé¢ä¸Šåªæœ‰ä¸€ä¸ªæ–‡ä»¶çš„æƒ…å†µä¸‹ï¼Œæˆ‘ä»¬éœ€è¦æœ‰ä¸¤ä¸ªæ–‡ä»¶çš„çŠ¶æ€ã€‚ç»§ç»­æ‰“å¼€ä¸€ä¸ªæ–°æ–‡ä»¶å³å¯ã€‚ä½†æ­¤æ—¶ï¼Œå°±éœ€è¦åœ¨æ‰“å¼€æ–°æ–‡ä»¶ä¹‹å‰ï¼Œè¿è¡Œdoc1=acad.ActiveDocumentï¼Œç”¨æ¥æ ‡è®°å½“å‰æ–‡ä»¶ã€‚å†æ‰“å¼€æ–°æ–‡ä»¶
li()æ–°æ–‡ä»¶ï¼Œè¿è¡Œdoc2=acad.ActiveDocumentï¼Œè®°å½•å½“å‰æ–‡ä»¶ã€‚æ­¤æ—¶ï¼Œè¿æ¥çš„æ˜¯doc2,ä½†æˆ‘ä»¬å¯ä»¥æ‰§è¡Œdoc1.Close()å…³é—­å‰é¢çš„æ–‡ä»¶ï¼Œå°±æ¢å¤åˆ°äº†å•ä¸€æ–‡ä»¶çŠ¶æ€ã€‚
å½“ç„¶ï¼Œé€šè¿‡dir(doc1)å¯ä»¥æŸ¥é˜…æ–‡ä»¶çš„å±æ€§å’Œæ–¹æ³•ï¼Œä¸ä»…å¯ä»¥æŸ¥çœ‹æ–‡ä»¶çš„åå­—ï¼Œè¿˜å¯ä»¥è·å–è¿™ä¸ªæ–‡ä»¶çš„å…¶å®ƒå±æ€§ã€‚ä¾‹å¦‚doc1.Groupsï¼ŒPlotConfigurations,æˆ‘ä»¬æ˜¯å¦å¯ä»¥åœ¨ä¸¤ä¸ªæ–‡ä»¶éƒ½åœ¨çš„æƒ…å†µä¸‹æ‹·è´ä¸€ä¸ªåˆ°å¦ä¸€ä¸ªæ¥ã€‚


#å½“å‰æ¿€æ´»æ¡Œé¢æ–‡ä»¶çš„æ¿€æ´»å¸ƒå±€
acad.ActiveDocument.SetVariable("TILEMODE",0)

layouts = doc.Layouts

existing_names = [layout.Name for layout in layouts]

doc.ActiveLayout = layouts.Item("å¸ƒå±€1")
    
layout = source_doc.ActiveLayout

#è·å–å½“å‰æ¿€æ´»æ–‡ä»¶çš„æ‰“å°é…ç½®

zg_pdf_config = source_doc.PlotConfigurations.Item("ZG_PDF")

# ä½¿ç”¨ CopyFrom æ–¹æ³•å°† "ZG_PDF" çš„è®¾ç½®å¤åˆ¶åˆ°æŒ‡å®šå¸ƒå±€

layout.CopyFrom(zg_pdf_config)

ä¹Ÿå°±æ˜¯è¯´ï¼Œä¸¤ä¸ªæ–‡ä»¶åŒæ—¶æ‰“å¼€çš„æ“ä½œæ˜¯å¿…ä¸å¯å°‘çš„ã€‚è™½ç„¶å…·ä½“è¿è¡Œæ—¶æˆ‘ä»¬å°½é‡ä¿æŒä¸€ä¸ªæ–‡ä»¶ã€‚ä½†ä¸¤ä¸ªæ–‡ä»¶éœ€è¦äº¤æ¢ä¿¡æ¯ã€‚

"""

def copy_doc_to_current(file_path):#å°†ç›®æ ‡æ–‡ä»¶åŸä½å¤åˆ¶ç²˜è´´åˆ°å½“å‰æ–‡ä»¶

    """
    è‹¥æ“ä½œä¸å½“ï¼Œç«Ÿè‡³ä¸èƒ½å¤åˆ¶ç²˜è´´
    å°†ç›®æ ‡æ–‡ä»¶ç²˜è´´è¿›æ¥ï¼Œå½“ç„¶æ˜¯æœ€è‡ªç„¶çš„éœ€æ±‚
    ä¸è¦ç”¨â€œç©ºç™½.dwgâ€ä½œæµ‹è¯•ï¼Œä¹Ÿä¸è¦ç”¨å¤©æ­£å¼€å¯æ—¶é»˜è®¤ç”Ÿæˆçš„æ–‡ä»¶ï¼Œå› ä¸ºå®ƒä»¬éƒ½ä¼šç ´åæ•´ä¸ªæµç¨‹
    å¾ˆé—æ†¾ï¼Œå—æ’å…¥å‘½ä»¤ä¸èƒ½ç”Ÿæ•ˆ

    """   

    li()
    doc.Save()
    file_path1=file_path
    file_path2=doc.FullName#è·å–å½“å‰æ–‡ä»¶å¸¦è·¯å¾„çš„æ–‡ä»¶å
    
    
    name2=doc.Name
    doc2=get_doc_by_name(name2)   
    
    #æ‰“å¼€è¾…åŠ©æ–‡ä»¶
    
    Open_By_Omission_wenjian("D:/Myprogramsystem/cad/xitongjicuwenjian/ç©ºç™½.dwg")

    li()

    name3=doc.Name

    doc3=get_doc_by_name(name3)     

    doc2.Close()

    insert_dwg_2(file_path2,file_path1, insert_point=(0, 0, 0))

    time.sleep(2)

    Open_By_Omission_wenjian(file_path2)

    li()

    doc.SendCommand("z\ne\n")

    doc3.Close()


def insert_dwg_2(new_file_path_A,new_file_path_B, insert_point=(0, 0, 0)):#æ–‡ä»¶Båˆæˆåˆ°æ–‡ä»¶Aä¸­

    """
    å¦‚æœåˆšæ‰“å¼€å¤©æ­£ ï¼Œä¼šé»˜è®¤äº§ç”Ÿä¸€ä¸ªdwg1æ–‡ä»¶ï¼Œä¹Ÿä¼šæ¿€æ´»ï¼Œç”¨æ¥è¿›è¡Œåç»­æ“ä½œã€‚å¯å½“æˆ‘ä»¬æ“ä½œåˆ°åé¢å…³é—­æ–°å»ºçš„æ–‡ä»¶Aæ—¶ï¼Œæ³¨æ„æ–°æ‰“å¼€Aæ–°å»ºAæ—¶è¿™ä¸ªdwg1ä¹Ÿå…³é—­äº†ã€‚?

    """

    

    li()

    Open_By_Omission_wenjian(new_file_path_B) #æ‰“å¼€æ–‡ä»¶B       

    li()

    acad.ActiveDocument.SendCommand('_ai_selall'+chr(13))

    time.sleep(4)

    acad.ActiveDocument.SendCommand('_copybase'+chr(13)+'0,0,0'+chr(13)+chr(13))#å…¨é€‰å¤åˆ¶

    time.sleep(4)

    guanbifile()
    
    li()

    Open_By_Omission_wenjian(new_file_path_A)#æ‰“å¼€æ–‡ä»¶A

    li()        
    
    acad.ActiveDocument.SendCommand('TPasteClip'+chr(13)+'0,0,0'+chr(13)+chr(13))#ç²˜è´´

    time.sleep(4)

    doc.save()

    time.sleep(4)
   
    doc.close()
    
    li()



def copy_group_S1_from_doc1_to_doc2(doc1, doc2, group_name="S1"):#å°†åä¸ºâ€œS1â€çš„ç»„å¤åˆ¶åˆ°å½“å‰æ¡Œé¢ä¸Šå¦ä¸€ä¸ªæ–‡æ¡£
    """
    å°† doc1 ä¸­åä¸º group_name çš„ç»„å¤åˆ¶åˆ° doc2 ä¸­ï¼Œå¹¶é‡æ–°ç»„è£…ç»„ã€‚
    ç²˜è´´ç‚¹ä¸º 0,0,0ã€‚ç²˜è´´åé€šè¿‡ handle å·®é›†è¯†åˆ«æ–°å¯¹è±¡ã€‚

    mpè¦ä¸æ–­æ›´æ–°

    def refresh_modelspace(doc):
        return doc.ModelSpace
    é€»è¾‘æ›´æ¸…æ™°

    ç›´æ¥å‘é€å‘½ä»¤å¤åˆ¶ç²˜è´´æˆ–è®¸æ›´å¥½
    """
    try:
        # 1. æ¿€æ´»æºæ–‡æ¡£
        set_active_doc(doc1)
        li()

        group = doc.Groups.Item(group_name)
        handles = [ent.Handle for ent in group]
        objs = [ent for ent in mp if ent.Handle in handles]

        yin_to_xian_xuanze(objs)

        doc.SendCommand("_copybase\n0,0,0\n")
        time.sleep(0.5)
        doc.SendCommand("_copyclip\n\n")
        time.sleep(1)

        # 2. æ¿€æ´»ç›®æ ‡æ–‡æ¡£
        set_active_doc(doc2)
        li()
        ms2 = doc.ModelSpace

        # 3. ç²˜è´´å‰è®°å½•å·²æœ‰å¯¹è±¡
        pre_map = get_handle_object_map(ms2)

        # 4. ç²˜è´´
        doc.SendCommand("_pasteclip"+chr(13)+"0,0,0"+chr(13)+chr(13))#
        time.sleep(1.5)

        # 5. ç²˜è´´åé‡æ–°è®°å½•å¯¹è±¡
        ms2 = doc.ModelSpace
        
        post_map = get_handle_object_map(ms2)

        # 6. å–å‡ºæ–°å¯¹è±¡ï¼ˆé€šè¿‡ handle å·®é›†ï¼‰
        new_handles = set(post_map) - set(pre_map)
        new_objs = [post_map[h] for h in new_handles]

        print(f"[OK] ç²˜è´´å®Œæˆï¼Œè¯†åˆ«å‡º {len(new_objs)} ä¸ªæ–°å›¾å…ƒ")

        # 7. æ·»åŠ è¿™äº›å¯¹è±¡åˆ°ç»„ä¸­ï¼ˆä½¿ç”¨ä½ æä¾›çš„æ–¹æ³•ï¼‰
        add_objects_to_group(group_name, new_objs)

        print(f"[OK] æˆåŠŸå°†ç²˜è´´å¯¹è±¡åŠ å…¥ç»„ '{group_name}'")

    except Exception as e:
        print(f"[é”™è¯¯] å¤åˆ¶ç»„å¤±è´¥: {e}")

        



#&&% åˆå¹¶ä¸¤ä¸ªä¸é‡å çš„æ–‡ä»¶    
def insert_dwg_pyautocad(new_file_path_A,new_file_path_B, insert_point=(0, 0, 0)):#æ–‡ä»¶Båˆæˆåˆ°æ–°å»ºæ–‡ä»¶Aä¸­

    """
    å¦‚æœåˆšæ‰“å¼€å¤©æ­£ ï¼Œä¼šé»˜è®¤äº§ç”Ÿä¸€ä¸ªdwg1æ–‡ä»¶ï¼Œä¹Ÿä¼šæ¿€æ´»ï¼Œç”¨æ¥è¿›è¡Œåç»­æ“ä½œã€‚å¯å½“æˆ‘ä»¬æ“ä½œåˆ°åé¢å…³é—­æ–°å»ºçš„æ–‡ä»¶Aæ—¶ï¼Œæ³¨æ„æ–°æ‰“å¼€Aæ–°å»ºAæ—¶è¿™ä¸ªdwg1ä¹Ÿå…³é—­äº†ã€‚?

    """

    create_new_dwg_file(new_file_path_A)#æ–°å»ºæ–‡ä»¶å‘½ä»¤å·²ç»é»˜è®¤äº†å…³é—­æ–°å»ºçš„è¿™ä¸ªæ–‡ä»¶çš„æ“ä½œ

    li()

    Open_By_Omission_wenjian(new_file_path_B) #æ‰“å¼€æ–‡ä»¶B       

    li()

    acad.ActiveDocument.SendCommand('_ai_selall'+chr(13))

    time.sleep(4)

    acad.ActiveDocument.SendCommand('_copybase'+chr(13)+'0,0,0'+chr(13)+chr(13))#å…¨é€‰å¤åˆ¶

    time.sleep(4)

    guanbifile()
    
    li()

    Open_By_Omission_wenjian(new_file_path_A)#æ‰“å¼€æ–‡ä»¶A

    li()        
    
    acad.ActiveDocument.SendCommand('TPasteClip'+chr(13)+'0,0,0'+chr(13)+chr(13))#ç²˜è´´

    time.sleep(4)

    doc.save()

    time.sleep(4)
   
    doc.close()
    
    li()

#æ‰¹é‡åˆæˆæ–‡ä»¶
def insert_multiple_dwgs_to_new_file(new_file_path_A, source_files_list):
    """
    åˆ›å»ºä¸€ä¸ªæ–° DWG æ–‡ä»¶ Aï¼Œå¹¶å°†å¤šä¸ªå·²æœ‰ DWG æ–‡ä»¶ï¼ˆB1, B2, ...ï¼‰çš„å†…å®¹
    åŸä½å¤åˆ¶ç²˜è´´åˆ° A ä¸­ï¼Œæœ€ç»ˆä¿å­˜å¹¶å…³é—­ Aï¼Œè¿”å›åˆå§‹ dwg çŠ¶æ€ã€‚

    å‚æ•°ï¼š
        new_file_path_A      - è¦åˆ›å»ºçš„æ–° DWG æ–‡ä»¶è·¯å¾„ï¼ˆå®Œæ•´è·¯å¾„ï¼‰
        source_files_list    - è¦å¤åˆ¶è¿›æ¥çš„å¤šä¸ª DWG æ–‡ä»¶è·¯å¾„åˆ—è¡¨
    """
    # Step 1: åˆ›å»ºæ–°æ–‡ä»¶ Aï¼ˆæ–°å»ºåè‡ªåŠ¨å…³é—­ï¼‰
    create_new_dwg_file(new_file_path_A)
    time.sleep(1)

    # Step 2: é€ä¸ªå¤„ç†æ¯ä¸ªæºæ–‡ä»¶ B
    for idx, fileB in enumerate(source_files_list):
        print(f"ğŸ“¥ å¤„ç†ç¬¬ {idx+1} ä¸ªæºæ–‡ä»¶ï¼š{fileB}")
        Open_By_Omission_wenjian(fileB)
        li()
        time.sleep(1)

        acad.ActiveDocument.SendCommand('_ai_selall' + chr(13))
        time.sleep(2)

        acad.ActiveDocument.SendCommand('_copybase' + chr(13) + '0,0,0' + chr(13) + chr(13))
        time.sleep(3)

        guanbifile()
        time.sleep(1)

        Open_By_Omission_wenjian(new_file_path_A)
        li()
        time.sleep(1)

        acad.ActiveDocument.SendCommand('TPasteClip' + chr(13) + '0,0,0' + chr(13) + chr(13))
        time.sleep(3)

        acad.ActiveDocument.Save()
        time.sleep(1)

        acad.ActiveDocument.Close(False)
        time.sleep(1)

    # Step 3: æ“ä½œç»“æŸåå›åˆ° dwg1 åˆå§‹çŠ¶æ€
    li()
    print(f"[OK] å·²æˆåŠŸå°† {len(source_files_list)} ä¸ª DWG åˆå¹¶è‡³æ–‡ä»¶ï¼š{new_file_path_A}")





    

def xianshi_yincangtuxing():#  æ˜¾ç¤ºæ–‡ä»¶ä¸­å¯èƒ½éšè—çš„å¯¹è±¡
        
    acad.ActiveDocument.SendCommand("HFKJ"+chr(13)+chr(13))#åœ¨V4çŠ¶æ€ä¸‹å¯èƒ½è¦æ”¹æˆ"HFKJ"+chr(13)+"Y"+chr(13)

    doc.Save()



# åˆ›å»ºä¸€ä¸ªäº‹ä»¶å¯¹è±¡

timeout_event = threading.Event()

event = threading.Event()

def run_cad_program(timeout_event, event):
    pythoncom.CoInitialize()
    
    try:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        print(acad.ActiveDocument.Name)
        
        acad.ActiveDocument.SendCommand("TSaveAs"+chr(13))
        
        print("CADå‘½ä»¤å·²å‘é€ï¼Œç­‰å¾…çª—å£æ“ä½œå®Œæˆ...")
        
        # ç­‰å¾…timeout_eventä¿¡å·ï¼Œå¦‚æœæ”¶åˆ°ä¿¡å·ï¼Œåˆ™é€€å‡º
        timeout_event.wait()

    except Exception as e:
        print(f"run_cad_program å‡ºç°é”™è¯¯: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        event.set()  # é€šçŸ¥ä¸»çº¿ç¨‹å®Œæˆ

        
def automate_window_with_pywinauto_t7(timeout_event, event):
    pythoncom.CoInitialize()

    try:
        jincheng_acad = get_acad_process_id('acad.exe')
        handle = findwindows.find_windows(process=jincheng_acad)[0]

        app = Application().connect(handle=handle)
        time.sleep(2)

        try:
            window = app.window(title="å›¾å½¢å¯¼å‡º", class_name="#32770")
        except Exception as e:
            print("å°è¯•é‡æ–°è·å–çª—å£:", e)
            time.sleep(2)
            window = app.window(title="å›¾å½¢å¯¼å‡º", class_name="#32770")

        if window.exists():
            print("çª—å£å­˜åœ¨")
            child_windows = window.children()

            for child_window in child_windows:
                try:
                    child_window.select("å¤©æ­£7æ–‡ä»¶ (*.dwg) ")
                    print("æ­£åœ¨å¯¼å‡ºt7")
                    pass
                except Exception as e:
                    pass
        else:
            print("çª—å£ä¸å­˜åœ¨")
            pass

        save_button = window.child_window(title="ä¿å­˜(&S)", class_name="Button")

        try:
            save_button.set_focus()
            
            save_button.click()


        except Exception as e:
            print("pywinautoå¤„ç†çª—å£å‡ºç°é—®é¢˜:", e)
            time.sleep(2)
            save_button.set_focus()
            save_button.click()

        print("å·²ä½¿ç”¨pywinautoè‡ªåŠ¨åŒ–çª—å£")
        pass

    except Exception as e:
        print(f"automate_window_with_pywinauto_t7 å‡ºç°é”™è¯¯: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # é€šçŸ¥ä¸»çº¿ç¨‹å®Œæˆ

#&&% # å¤©æ­£è½¬t7    
def zhuancheng_t7():
    
    """
    åœ¨li()è¿æ¥æ¿€æ´»æ–‡ä»¶åï¼Œç›´æ¥æ‰§è¡Œè¯¥å‘½ä»¤å³å¯è½¬æ¢
    ä¸€èˆ¬ç†è§£æ‰€å¾—æ–‡ä»¶å’Œå½“å‰æ–‡ä»¶åŒæ–‡ä»¶å¤¹ï¼Œä½†è¿™æ¬¡æµ‹è¯•ç»“æœå´åœ¨æ–‡ä»¶å¤¹
    C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
    """    
    t1 = time.time()

    chaoshibiaoji = 1

    # åˆ›å»ºä¸€ä¸ªEventï¼Œç”¨äºé€šçŸ¥å­çº¿ç¨‹ç»ˆæ­¢
    timeout_event = threading.Event()
    
    # åˆ›å»ºä¸€ä¸ªEventï¼Œç”¨äºé€šçŸ¥ä¸»çº¿ç¨‹å®Œæˆ
    event = threading.Event()

    # åˆ›å»ºä¸¤ä¸ªçº¿ç¨‹ï¼Œåˆ†åˆ«æ‰§è¡Œrun_cad_programå’Œautomate_window_with_pywinauto_t7
    thread1 = threading.Thread(target=run_cad_program, args=(timeout_event, event))
    thread2 = threading.Thread(target=automate_window_with_pywinauto_t7, args=(timeout_event, event))

    thread1.start()
    thread2.start()

    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("æ“ä½œè¶…æ—¶ï¼Œæ­£åœ¨ä¸­æ–­...")

        pass

        
        # è®¾ç½®timeout_eventï¼Œé€šçŸ¥run_cad_program()ç»ˆæ­¢
        timeout_event.set()

    print("æ–‡ä»¶è½¬T7æ ¼å¼æ“ä½œç»“æŸ")

    t2 = time.time()
    print("æ–‡ä»¶è½¬T7æ ¼å¼æ“ä½œæ€»å…±ç”¨æ—¶", t2 - t1, "ç§’")


# åˆ›å»ºä¸€ä¸ªäº‹ä»¶å¯¹è±¡

timeout_event = threading.Event()

event = threading.Event()


def automate_window_with_pywinauto_t3(timeout_event, event):
    pythoncom.CoInitialize()

    try:
        jincheng_acad = get_acad_process_id('acad.exe')
        handle = findwindows.find_windows(process=jincheng_acad)[0]

        app = Application().connect(handle=handle)
        time.sleep(2)

        try:
            window = app.window(title="å›¾å½¢å¯¼å‡º", class_name="#32770")
        except Exception as e:
            print("å°è¯•é‡æ–°è·å–çª—å£:", e)
            time.sleep(2)
            window = app.window(title="å›¾å½¢å¯¼å‡º", class_name="#32770")

        if window.exists():
            print("çª—å£å­˜åœ¨")
            pass
            child_windows = window.children()

            for child_window in child_windows:
                try:
                    child_window.select("å¤©æ­£3æ–‡ä»¶ (*.dwg) ")
                    print("æ­£åœ¨å¯¼å‡ºt3")
                    pass
                except Exception as e:
                    pass
        else:
            print("çª—å£ä¸å­˜åœ¨")
            pass

        save_button = window.child_window(title="ä¿å­˜(&S)", class_name="Button")

        try:
            save_button.set_focus()
            save_button.click()
        except Exception as e:
            print("pywinautoå¤„ç†çª—å£å‡ºç°é—®é¢˜:", e)
            pass
            time.sleep(2)
            save_button.set_focus()
            save_button.click()

        print("å·²ä½¿ç”¨pywinautoè‡ªåŠ¨åŒ–çª—å£")
        pass

    except Exception as e:
        print(f"automate_window_with_pywinauto_t3 å‡ºç°é”™è¯¯: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # é€šçŸ¥ä¸»çº¿ç¨‹å®Œæˆ
#&&% # å¤©æ­£è½¬t3
        
def zhuancheng_t3():

    """
    åœ¨li()è¿æ¥æ¿€æ´»æ–‡ä»¶åï¼Œç›´æ¥æ‰§è¡Œè¯¥å‘½ä»¤å³å¯è½¬æ¢
    ä¸€èˆ¬ç†è§£æ‰€å¾—æ–‡ä»¶å’Œå½“å‰æ–‡ä»¶åŒæ–‡ä»¶å¤¹ï¼Œä½†è¿™æ¬¡æµ‹è¯•ç»“æœå´åœ¨æ–‡ä»¶å¤¹
    C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
    """    

    t1 = time.time()

    chaoshibiaoji = 1

    # åˆ›å»ºä¸€ä¸ªEventï¼Œç”¨äºé€šçŸ¥å­çº¿ç¨‹ç»ˆæ­¢
    timeout_event = threading.Event()
    
    # åˆ›å»ºä¸€ä¸ªEventï¼Œç”¨äºé€šçŸ¥ä¸»çº¿ç¨‹å®Œæˆ
    event = threading.Event()

    # åˆ›å»ºä¸¤ä¸ªçº¿ç¨‹ï¼Œåˆ†åˆ«æ‰§è¡Œrun_cad_programå’Œautomate_window_with_pywinauto_t3
    thread1 = threading.Thread(target=run_cad_program, args=(timeout_event, event))
    thread2 = threading.Thread(target=automate_window_with_pywinauto_t3, args=(timeout_event, event))

    thread1.start()
    thread2.start()

    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("æ“ä½œè¶…æ—¶ï¼Œæ­£åœ¨ä¸­æ–­...")

        pass

        
        # è®¾ç½®timeout_eventï¼Œé€šçŸ¥run_cad_program()ç»ˆæ­¢
        timeout_event.set()

    print("æ–‡ä»¶è½¬T3æ ¼å¼æ“ä½œç»“æŸ")

    t2 = time.time()
    print("æ–‡ä»¶è½¬T3æ ¼å¼æ“ä½œæ€»å…±ç”¨æ—¶", t2 - t1, "ç§’")


#  ä¸»å‡½æ•°
#  (2)
# ç¡®å®šä¸€ä¸ªæ–‡ä»¶ä¸­æ‰€æœ‰çš„å¯¹è±¡ç±»å‹

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°


def get_all_object_types():
    """
    è·å–å½“å‰æ–‡ä»¶ä¸­æ¨¡å‹ç©ºé—´ä¸­æ‰€æœ‰å›¾å½¢å¯¹è±¡çš„ç±»å‹åç§°ï¼ˆå¦‚ AcDbText, AcDbLine ç­‰ï¼‰

    è¿”å›ï¼š
        ç±»å‹ååˆ—è¡¨ï¼ˆå»é‡ï¼Œå·²æ’åºï¼‰
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    modelspace = doc.ModelSpace

    types_set = set()

    for obj in modelspace:
        try:
            types_set.add(obj.ObjectName)
        except:
            continue

    types_list = sorted(types_set)
    print(f"[OK] å…±å‘ç° {len(types_list)} ç§å¯¹è±¡ç±»å‹ï¼š")
    for t in types_list:
        print(f" - {t}")

    return types_list



#  ä¸»å‡½æ•°
#  (3)
# å…³é—­æ–‡ä»¶åä¸ºNameçš„æ–‡ä»¶

#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

def close_dwg_by_name(Name):
    """
    å…³é—­å½“å‰æ¡Œé¢ä¸­åä¸º Name çš„ DWG æ–‡ä»¶ã€‚
    å¦‚æœæ–‡ä»¶å·²æ‰“å¼€ï¼Œåˆ™å…³é—­è¯¥æ–‡ä»¶ã€‚
    
    å‚æ•°ï¼š
        Name: è¦å…³é—­çš„ DWG æ–‡ä»¶çš„åç§°ï¼ˆå¦‚ "example.dwg"ï¼‰ä¸å«è·¯å¾„çš„å
    """
    try:
        acad = win32com.client.Dispatch("AutoCAD.Application")  # è·å– AutoCAD åº”ç”¨
        doc = acad.Documents.Item(Name)  # è·å–æŒ‡å®šåç§°çš„æ–‡æ¡£
        
        if doc:
            doc.Close(False)  # å…³é—­æ–‡ä»¶ï¼Œä¸æç¤ºä¿å­˜
            print(f"[OK] æ–‡ä»¶ '{Name}' å·²å…³é—­")
        else:
            print(f"[é”™è¯¯] æœªæ‰¾åˆ°åä¸º '{Name}' çš„æ–‡ä»¶")
    except Exception as e:
        print(f"[é”™è¯¯] å…³é—­æ–‡ä»¶ '{Name}' å¤±è´¥: {e}")




def set_active_doc(doc):# è®¾ç½®æ–‡ä»¶docä¸ºæ¿€æ´»å¯¹è±¡

    """
    docä¸ºcomå¯¹è±¡
    å°†æŒ‡å®šæ–‡æ¡£ doc è®¾ç½®ä¸ºå½“å‰æ¿€æ´»çª—å£ã€‚
    å°†å…¶ä¸­ä¸€ä¸ªæ‰“å¼€æ—¶ï¼Œè¦ä½¿ç”¨li()ï¼Œç„¶åæ‰§è¡Œdoc1=acad.ActiveDocument
    å†æ‰“å¼€å¦ä¸€ä¸ªï¼Œä½¿ç”¨li()è¿æ¥ï¼Œæ‰§è¡Œdoc2=acad.ActiveDocumentå°±å®é™…ä¸Šè·å–äº†ä¸¤ä¸ªæ–‡ä»¶comå®ä½“å¯¹è±¡
    ç„¶åå°±å¯ä»¥ä½¿ç”¨è¿™ä¸ªå‡½æ•°éšæ—¶æ¿€æ´»å…¶ä¸­ä¸€ä¸ªæ–‡ä»¶ï¼Œæ¿€æ´»åä»ç„¶è¦ä½¿ç”¨li()è¿æ¥æ‰èƒ½æ­£ç¡®è¿è¡Œ
    """
    try:
        doc.Activate()
        time.sleep(0.3)  # ç¨ä½œå»¶æ—¶ï¼Œç¡®ä¿æ¿€æ´»ç”Ÿæ•ˆ
        print("[OK] å½“å‰æ¿€æ´»æ–‡æ¡£ï¼š", doc.Name)
        return True
    except Exception as e:
        print("[é”™è¯¯] æ¿€æ´»æ–‡æ¡£å¤±è´¥ï¼š", e)
        return False


def get_doc_by_name(name): #ä»æ–‡ä»¶åè·å–æ–‡ä»¶å¯¹è±¡
    """
    é€šè¿‡æ–‡ä»¶åè·å– AutoCAD æ–‡æ¡£å¯¹è±¡ï¼Œä¾‹å¦‚ 'ç©ºç™½.dwg'
    å¦‚æœæœªæ‰¾åˆ°ï¼Œè¿”å› None
    """
    import win32com.client
    acad = win32com.client.GetActiveObject("AutoCAD.Application")
    for doc in acad.Documents:
        if doc.Name.lower() == name.lower():
            return doc
    return None

def get_open_document_names():#è¿”å›æ‰€æœ‰æ‰“å¼€çš„æ–‡ä»¶å
    import win32com.client
    acad = win32com.client.GetActiveObject("AutoCAD.Application")
    return [doc.Name for doc in acad.Documents]



#&&% å½“å‰CADæ–‡ä»¶æ•°

def dwgs_count():#å½“å‰æ¡Œé¢çš„dwgæ–‡ä»¶æ•°é‡

    shu = acad.Documents.Count

    return shu



#  ä¸»å‡½æ•°
#  (3)
# è·¨æ–‡ä»¶å¤åˆ¶ç²˜è´´

#&&%  æ–‡ä»¶æŒ‰æ‰“å°åŒºåŸŸåˆ†è§£æˆå¤šä¸ªæ–‡ä»¶


def export_region_to_new_file(x1, y1, x2, y2, filename,ty=1):##å°†åŒºåŸŸ(x1,y1,x2,y2)å†…çš„å¯¹è±¡åŸä½å‰ªåˆ‡ç²˜è´´åˆ°ä¸€ä¸ªæ–°æ–‡ä»¶ä¸­ã€‚
    """
    å°†å½“å‰æ–‡ä»¶ä¸­æŒ‡å®šçª—å£åŒºåŸŸå†…çš„å›¾å½¢å¯¹è±¡ï¼Œå‰ªåˆ‡ç²˜è´´åˆ°ä¸€ä¸ªæ–° DWG æ–‡ä»¶ä¸­ï¼Œå¹¶ä¿å­˜ä¸º filename.dwgã€‚
    æœ€åå…³é—­æ–°å»ºæ–‡ä»¶ã€é‡æ–°è¿æ¥å½“å‰æ–‡ä»¶ã€‚
    """


    # èšç„¦åŒºåŸŸ
    shitu_region(x1,y1,x2,y2)


    # æ­¥éª¤ 1ï¼šé€‰æ‹©åŒºåŸŸå¯¹è±¡å¹¶è½¬ä¸ºè“è‰²å¤¹ç‚¹é€‰ä¸­


    
    li()
    LB = select_objects_in_window_area(x1, y1, x2, y2)
    if not LB:
        print("[é”™è¯¯] åŒºåŸŸå†…æœªæ‰¾åˆ°ä»»ä½•å¯¹è±¡ï¼Œæ“ä½œä¸­æ­¢")
        return

    yin_to_xian_xuanze(LB,ty=ty)  # å…³é”®ï¼šè½¬æ¢ä¸ºå‘½ä»¤è¡ŒçŠ¶æ€é€‰ä¸­

    # æ­¥éª¤ 2ï¼šå¤åˆ¶é€‰ä¸­å¯¹è±¡
    
    
    original_doc = doc  # è®°å½•åŸå§‹æ–‡æ¡£å¥æŸ„

    doc.SendCommand("_copybase" + chr(13) + "0,0,0" + chr(13) + chr(13))
    time.sleep(3)

    # æ­¥éª¤ 3ï¼šæ–°å»º DWG æ–‡ä»¶å¹¶è¿æ¥
    acad.Documents.Add()
    time.sleep(1)
    li()
    new_doc = acad.ActiveDocument

    # æ­¥éª¤ 4ï¼šç²˜è´´å¯¹è±¡
    new_doc.SendCommand("TPasteClip" + chr(13) + "0,0,0" + chr(13) + chr(13))
    time.sleep(3)

    #è§†ç‚¹èšç„¦
    shitu_region(x1,y1,x2,y2)

    # æ­¥éª¤ 5ï¼šä¿å­˜æ–°æ–‡ä»¶
    original_path = original_doc.FullName
    folder = os.path.dirname(original_path)
    save_path = os.path.join(folder, filename + ".dwg")


    new_doc.SaveAs(save_path)
    print(f"[OK] å¯¹è±¡å·²å¯¼å‡ºåˆ°ï¼š{save_path}")

    # [OK] æ–°å¢éƒ¨åˆ† â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“â†“

    # æ­¥éª¤ 6ï¼šå…³é—­æ–°å»ºæ–‡ä»¶
    new_doc.Close(False)  # ä¸æç¤ºä¿å­˜

    # æ­¥éª¤ 7ï¼šé‡æ–°è¿æ¥åŸå§‹æ–‡ä»¶
    li()
    time.sleep(1)
    active_doc = acad.ActiveDocument

    



def batch_export_regions_to_files(region_list, filename_prefix="åŒºåŸŸå¯¼å‡º",ty=1):##æ–‡ä»¶åˆ†è§£æˆå¤šä¸ªå°æ–‡ä»¶å­˜å‚¨
    """
    æ‰¹é‡å¯¼å‡ºå¤šä¸ªåŒºåŸŸä¸º DWG æ–‡ä»¶ã€‚

    å‚æ•°ï¼š
        region_list: åˆ—è¡¨ï¼Œæ¯é¡¹ä¸ºä¸€ä¸ª (x1, y1, x2, y2) å…ƒç»„
        filename_prefix: ä¿å­˜æ–‡ä»¶åå‰ç¼€
        ty---é€šè¿‡åˆ é™¤å†æ¢å¤å¯¹è±¡ä»¥å°†éšæ€§é€‰æ‹©è½¬å˜ä¸ºæ˜¾æ€§é€‰æ‹©ï¼Œç•™ä¸€ä¸ªå»¶è¿Ÿæ—¶é—´æ§åˆ¶å‚æ•°
    """
    #ç¡®ä¿æ–‡ä»¶å¤¹ä¸­åå«"åŒºåŸŸå¯¼å‡º"çš„æ–‡ä»¶è¢«åˆ é™¤

    folder_1 = current_dwg_folder()
    clear_files_with_prefix(folder_1, filename_prefix="åŒºåŸŸå¯¼å‡º", delay = ty)
    time.sleep(ty)



    name1 = current_dwg_basename()

    for idx, region in enumerate(region_list):
        if len(region) != 4:
            print(f"[é”™è¯¯] åŒºåŸŸç¬¬ {idx+1} ä¸ªæ ¼å¼é”™è¯¯ï¼Œè·³è¿‡ã€‚")
            continue

        x1, y1, x2, y2 = region
        filename = f"{filename_prefix}_{idx+1:02d}"
        filename = name1 + filename

        chongfu_caozuo(
            export_region_to_new_file,
            dwg_instance=None,
            args=(x1, y1, x2, y2, filename),
            kwargs={'ty':ty},
            max_retries=3,
            failure_value=None
        )




#&&% #åŒæ–‡ä»¶æ“ä½œ

def copy_region_from_doc1_to_doc2_absolute(doc1, doc2, x1, y1, x2, y2):#ä»doc1çš„ä¸€ä¸ªåŒºåŸŸåŸä½ç²˜è´´åˆ°doc2
    """
    å°† doc1 ä¸­æŒ‡å®šåŒºåŸŸ (x1, y1)-(x2, y2) å†…çš„å¯¹è±¡å¤åˆ¶ï¼Œ
    ç²˜è´´åˆ° doc2 çš„ç›¸åŒä½ç½® (0,0,0)ã€‚ï¼ˆåŸä½å¤åˆ¶ï¼‰

    å½“æœªèƒ½å®Œå…¨æ§åˆ¶æ¿€æ´»æ–‡ä»¶æ—¶ï¼Œè¿™äº›å‡½æ•°è¿è¡Œå¹¶ä¸ç¨³å®š
    
    è¦æ±‚ï¼š
    - li() å¯æ›´æ–°å½“å‰å…¨å±€è¿æ¥
    - select_objects_in_window_area() è¿”å› COM å¯¹è±¡åˆ—è¡¨
    - yin_to_xian_xuanze() å°†å¯¹è±¡è½¬ä¸ºå‘½ä»¤è¡Œé€‰ä¸­çŠ¶æ€
    """
    try:
        # Step 1: æ¿€æ´»å¹¶è¿æ¥æºæ–‡æ¡£
        set_active_doc(doc1)
        li()

        # Step 2: é€‰æ‹©åŒºåŸŸå†…å¯¹è±¡
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("[é”™è¯¯] åŒºåŸŸå†…æœªæ‰¾åˆ°å¯¹è±¡ï¼Œå¤åˆ¶ä¸­æ­¢")
            return

        # Step 3: å°†å¯¹è±¡è“è‰²é€‰ä¸­çŠ¶æ€
        yin_to_xian_xuanze(LB)

        # Step 4: æ‰§è¡Œ _copybase å’Œ _copyclip
        doc1.SendCommand("_copybase" + chr(13) + "0,0,0" + chr(13) + chr(13))
        time.sleep(2)
        
        """        
        å¦‚æœå‡ºç°ç²˜è´´å›æºæ–‡æ¡£çš„æ··ä¹±ï¼Œå¯ä»¥è€ƒè™‘å°†æºæ–‡ä»¶åœ¨å·²ç»æ‰§è¡Œäº†set_active_doc(doc2)åå…³é—­doc1

        """        

        # Step 5: æ¿€æ´»å¹¶è¿æ¥ç›®æ ‡æ–‡æ¡£
        set_active_doc(doc2)
        li()

        # Step 6: ç²˜è´´åˆ°ç›¸åŒä½ç½®
        doc2.SendCommand("TPasteClip" + chr(13) + "0,0,0" + chr(13) + chr(13))#_pastclipæ•ˆæœä¸€æ ·
        time.sleep(1.5)

        print(f"[OK] åŒºåŸŸå¯¹è±¡å·²ä» {doc1.Name} ç²˜è´´åˆ° {doc2.Name}")

    except Exception as e:
        print(f"[é”™è¯¯] å¤åˆ¶åŒºåŸŸå¯¹è±¡å¤±è´¥: {e}")

def copy_region_from_doc1_to_doc2_relative(doc1, doc2, x1, y1, x2, y2):#ä»doc1åŒºåŸŸé›¶ç‚¹ç²˜è´´åˆ°doc2æ•´ä¸ªç©ºé—´é›¶ç‚¹
    """
    å°† doc1 ä¸­æŒ‡å®šåŒºåŸŸ (x1, y1)-(x2, y2) å†…çš„å¯¹è±¡å¤åˆ¶ï¼Œ
    ç²˜è´´åˆ° doc2 çš„ç›¸åŒä½ç½® (0,0,0)ã€‚ï¼ˆåŸä½å¤åˆ¶ï¼‰
    
    è¦æ±‚ï¼š
    - li() å¯æ›´æ–°å½“å‰å…¨å±€è¿æ¥
    - select_objects_in_window_area() è¿”å› COM å¯¹è±¡åˆ—è¡¨
    - yin_to_xian_xuanze() å°†å¯¹è±¡è½¬ä¸ºå‘½ä»¤è¡Œé€‰ä¸­çŠ¶æ€
    """
    try:
        # Step 1: æ¿€æ´»å¹¶è¿æ¥æºæ–‡æ¡£
        set_active_doc(doc1)
        li()

        # Step 2: é€‰æ‹©åŒºåŸŸå†…å¯¹è±¡
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("[é”™è¯¯] åŒºåŸŸå†…æœªæ‰¾åˆ°å¯¹è±¡ï¼Œå¤åˆ¶ä¸­æ­¢")
            return

        # Step 3: å°†å¯¹è±¡è“è‰²é€‰ä¸­çŠ¶æ€
        yin_to_xian_xuanze(LB)

        # Step 4: æ‰§è¡Œ _copybase å’Œ _copyclip
        doc1.SendCommand("_copybase\n0,0,0\n\n")
        time.sleep(0.5)
        doc1.SendCommand("_copyclip\n\n")
        time.sleep(2)

        
        """        
        å¦‚æœå‡ºç°ç²˜è´´å›æºæ–‡æ¡£çš„æ··ä¹±ï¼Œå¯ä»¥è€ƒè™‘å°†æºæ–‡ä»¶åœ¨å·²ç»æ‰§è¡Œäº†set_active_doc(doc2)åå…³é—­doc1

        """        


        # Step 5: æ¿€æ´»å¹¶è¿æ¥ç›®æ ‡æ–‡æ¡£
        set_active_doc(doc2)
        li()

        # Step 6: ç²˜è´´åˆ°ç›¸åŒä½ç½®
        doc2.SendCommand("_pasteclip\n0,0,0\n\n")
        time.sleep(1.5)

        print(f"[OK] åŒºåŸŸå¯¹è±¡å·²ä» {doc1.Name} ç²˜è´´åˆ° {doc2.Name}")

    except Exception as e:
        print(f"[é”™è¯¯] å¤åˆ¶åŒºåŸŸå¯¹è±¡å¤±è´¥: {e}")


def copy_region_from_doc1_to_doc2_at_point(doc1, doc2, x1, y1, x2, y2, x0, y0, z0=0):#å°†doc1æŒ‡å®šåŒºåŸŸçš„æœ€ä½ç‚¹ï¼Œç²˜è´´åˆ°doc2çš„ç›®æ ‡ç‚¹
    """
    å°† doc1 ä¸­æŒ‡å®šåŒºåŸŸ (x1, y1)-(x2, y2) å†…çš„å¯¹è±¡å¤åˆ¶ï¼Œ
    ç²˜è´´åˆ° doc2 çš„æŒ‡å®šä½ç½® (x0, y0, z0)ã€‚
    """
    try:
        # Step 1: æ¿€æ´»å¹¶è¿æ¥æºæ–‡æ¡£
        set_active_doc(doc1)
        li()

        # Step 2: é€‰æ‹©åŒºåŸŸå†…å¯¹è±¡
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("[é”™è¯¯] åŒºåŸŸå†…æœªæ‰¾åˆ°å¯¹è±¡ï¼Œå¤åˆ¶ä¸­æ­¢")
            return

        # Step 3: è“è‰²é«˜äº®
        yin_to_xian_xuanze(LB)

        # Step 4: æ‹·è´
        doc1.SendCommand("_copybase\n0,0,0\n\n")
        time.sleep(0.5)
        doc1.SendCommand("_copyclip\n\n")
        time.sleep(2)

        # Step 5: æ¿€æ´»ç›®æ ‡æ–‡æ¡£
        set_active_doc(doc2)

        # å°è¯•å¤šæ¬¡è¿æ¥ç›®æ ‡æ–‡æ¡£
        for _ in range(5):
            try:
                li()
                break
            except:
                time.sleep(0.5)
        else:
            print("[é”™è¯¯] æ¿€æ´»ç›®æ ‡æ–‡æ¡£å¤±è´¥ï¼Œæ— æ³•è¿æ¥")
            return

        # Step 6: ç²˜è´´åˆ°æŒ‡å®šç‚¹
        paste_cmd = f"_pasteclip\n{x0},{y0},{z0}\n\n"
        doc2.SendCommand(paste_cmd)
        time.sleep(1.5)

        print(f"[OK] åŒºåŸŸå¯¹è±¡å·²ä» {doc1.Name} ç²˜è´´åˆ° {doc2.Name} çš„æŒ‡å®šç‚¹ ({x0}, {y0}, {z0})")

    except Exception as e:
        print(f"[é”™è¯¯] å¤åˆ¶åŒºåŸŸå¯¹è±¡å¤±è´¥ï¼š{e}")






def insert_as_block(p,block_path = r"D:/Myprogramsystem/XT/MC_yuan.dwg"):#ä»¥å—çš„æ’å…¥å°†MC_yuan.dwgæ’å…¥å½“å‰æ¿€æ´»æ–‡ä»¶
    """
    å°† MC_yuan.dwg æ–‡ä»¶ä»¥å—å½¢å¼æ’å…¥å½“å‰æ–‡æ¡£ï¼Œå¹¶ç«‹å³åˆ†è§£ã€‚

    MC_yuan.dwg æ–‡ä»¶åˆ†åˆ«å°†9ç§é—¨çª—åŸºå…ƒæ”¾å…¥å„è‡ªçš„å›¾å±‚

    jz-danmen jz-shuangmen jz-tuilamen jz-juanlianmen jz-zimumen jz-pingchuang jz-tuchuang jz-baiyechuang jz-gaochuang

    å¢™éƒ½åœ¨jizhunwall
    p   (x,y,z)ä¸‰ç»´åæ ‡
    
    """
    
    if not os.path.exists(block_path):
        print("[é”™è¯¯] æ–‡ä»¶ä¸å­˜åœ¨ï¼š", block_path)
        return

    ensure_layer(layer_name="jizhunwall")

    x=p[0]

    y=p[1]

    z=p[2]

    # æ’å…¥å—ï¼ˆæ’å…¥ç‚¹0,0,0ï¼Œæ¯”ä¾‹1ï¼‰
    cmd = f"-insert\n{block_path}\n{x},{y},{z}\n1\n1\n0\n"
    doc.SendCommand(cmd)

    # ç¨³å®šç­‰å¾…å›¾å±‚å¯¹è±¡å‡ºç°ï¼ˆæœ€å¤šç­‰ 3 ç§’ï¼‰
    LB = []
    for _ in range(30):
        time.sleep(0.1)
        LB = select_tuceng("jizhunwall")
        if LB:
            break
    else:
        print("[é”™è¯¯] å›¾å±‚ä¸­æœªèƒ½åŠæ—¶æ£€æµ‹åˆ°å¯¹è±¡")
        return

    print("é€‰åˆ°çš„jizhutucengå¯¹è±¡æ•°é‡", len(LB))

    block = LB[0]

    for attempt in range(3):
        try:
            block.Explode()
            print(f"[OK] ç¬¬ {attempt+1} æ¬¡ç‚¸å—æˆåŠŸ")
            break
        except Exception as e:
            print(f"[è­¦å‘Š]ï¸ ç¬¬ {attempt + 1} æ¬¡ç‚¸å—å¤±è´¥ï¼š{e}")
            time.sleep(0.2)
    else:
        print("[é”™è¯¯] å¤šæ¬¡å°è¯•ç‚¸å—å¤±è´¥")
        return

    try:
        block.Delete()
        print("ğŸ—‘ï¸ åŸå—å¯¹è±¡å·²åˆ é™¤")
    except:
        print("[è­¦å‘Š]ï¸ åˆ é™¤åŸå—å¤±è´¥")



#&&&&%%  ç¬¬ä¹éƒ¨åˆ† CADå›¾å— 

##ç»™å±æ€§å—èµ‹äºˆæ–°å€¼ï¼Œå¯ä»¥å±€éƒ¨èµ‹å€¼


def huoqukuai_shuxing_zhi(XX):#XXä¸ºå±æ€§å—å®ä½“


    attributes = XX.GetAttributes()

    tags=[]
    values=[]

    for attr in attributes:
               
        tag = attr.TagString

            
        value = attr.TextString

        tags.append(tag)

        values.append(value)

    return tags,values


def set_attributes_values(block, tags_order, new_values):
    """
    ä¸ºå±æ€§å—çš„æ ‡ç­¾è®¾ç½®æ–°çš„å€¼ã€‚

    å‚æ•°:
    - block: è¦ä¿®æ”¹çš„å±æ€§å—ï¼ˆCOM BlockReference å¯¹è±¡ï¼‰ã€‚
    - tags_order: ä¸€ä¸ªåˆ—è¡¨ï¼ŒåŒ…å«ä½ å¸Œæœ›æŒ‰ç…§å“ªç§é¡ºåºä¸ºå±æ€§å—çš„æ ‡ç­¾è®¾ç½®æ–°çš„å€¼ã€‚
    - new_values: ä¸€ä¸ªåˆ—è¡¨ï¼ŒåŒ…å«æŒ‰æ ‡ç­¾é¡ºåºæ’åˆ—çš„æ–°å€¼ã€‚

    è¿”å›:
    None
    """
    # å…ˆå°è¯•è·å–å±æ€§åˆ—è¡¨
    try:
        attributes = block.GetAttributes()
    except Exception as e:
        print(f"[è­¦å‘Š] å®ä½“ {block.ObjectName}({getattr(block, 'Handle', '?')}) æ— æ³•è·å–å±æ€§ï¼Œè·³è¿‡: {e}")
        return

    index = 0
    for tag in tags_order:
        # æ‰¾åˆ°å¯¹åº”æ ‡ç­¾çš„å±æ€§
        found = False
        for attr in attributes:
            if attr.TagString == tag:
                found = True
                try:
                    attr.TextString = new_values[index]
                    print(f"æ ‡ç­¾: {tag}  æ–°å€¼: {new_values[index]}")
                except Exception as e:
                    print(f"[è­¦å‘Š] è®¾ç½®æ ‡ç­¾ '{tag}' æ—¶å‡ºé”™: {e}")
                index += 1
                break
        if not found:
            print(f"[è­¦å‘Š] æœªæ‰¾åˆ°å±æ€§æ ‡ç­¾ '{tag}'ï¼Œå·²è·³è¿‡")

    # æ›´æ–°å—
    try:
        block.Update()
    except Exception as e:
        print(f"[è­¦å‘Š] æ›´æ–°å—æ—¶å‡ºé”™: {e}")              

##>>> tags_order=["1.0","æ–½å·¥å›¾","2023.10","1:100","1.0","ä¸“ä¸šåç§°"]
##>>> new_values=["1.2ç‰ˆ","åˆæ­¥è®¾è®¡","2021.07","1:25","JS-09","å»ºæ–½"]



def resize_block_attribute(block_ref, tag: str, *, height: float = 200.0, width: float = 4500.0):
    """
    å°†å—å‚ç…§ block_ref ä¸­æŒ‡å®š Tag çš„å±æ€§æ–‡å­—æ”¹æˆç»™å®šå­—é«˜å¹¶è®¾ç½®è¾¹ç•Œå®½åº¦ã€‚

    é€‚ç”¨å¯¹è±¡
    --------
    block_ref : AcadBlockReference
        å¿…é¡»æ˜¯åŒ…å«å±æ€§ (HasAttributes=True) çš„å—å‚ç…§
    tag       : str
        ç›®æ ‡å±æ€§ TagStringï¼ˆä¸åŒºåˆ†å¤§å°å†™ï¼‰
    height    : float
        ç›®æ ‡å­—é«˜ï¼ˆDrawing Unitsï¼‰
    width     : float
        å¤šè¡Œå±æ€§ (MText attribute) çš„è¾¹ç•Œå®½åº¦ï¼›è‹¥å±æ€§ä¸æ˜¯å¤šè¡Œï¼Œ
        å°è¯•è®¾ç½® WidthFactor ä»¥è¿‘ä¼¼æ•ˆæœã€‚

    è¿”å›
    ----
    bool
        True  : è‡³å°‘æ‰¾åˆ°å¹¶ä¿®æ”¹äº†ä¸€ä¸ªå±æ€§
        False : æ²¡æœ‰æ‰¾åˆ°æŒ‡å®š tag æˆ–è°ƒæ•´å¤±è´¥
    """
    if (getattr(block_ref, "ObjectName", "") != "AcDbBlockReference"
            or not getattr(block_ref, "HasAttributes", False)):
        print("[è­¦å‘Š] ä¼ å…¥å¯¹è±¡ä¸æ˜¯å¸¦å±æ€§çš„å—å‚ç…§")
        return False

    modified = False
    target_tag = tag.strip().upper()

    try:
        for attr in block_ref.GetAttributes():
            if attr.TagString.strip().upper() != target_tag:
                continue

            # â€”â€”â€” å­—é«˜ â€”â€”â€”
            try:
                attr.Height = height
            except Exception as e:
                print(f"[è­¦å‘Š] è®¾ç½® Height å¤±è´¥: {e}")

            # â€”â€”â€” è¾¹ç•Œå®½åº¦ / å®½åº¦å› å­ â€”â€”â€”
            # å¤šè¡Œå±æ€§æ˜¯ AcDbAttributeReferenceï¼Œä½†å†…æ ¸ä¸­ä»å¸¦ MTextï¼Œ
            # COM æš´éœ² 'Width'ï¼›è‹¥æ²¡æœ‰å°±é€€è€Œæ±‚å…¶æ¬¡æ”¹ WidthFactor
            if hasattr(attr, "Width"):
                try:
                    attr.Width = width
                except Exception as e:
                    print(f"[è­¦å‘Š] è®¾ç½® Width å¤±è´¥: {e}")
            else:
                try:
                    # ä¼°ç®—ä¸€ä¸ªå®½åº¦å› å­ä½¿å•è¡Œæ–‡æœ¬å æ®è¿‘ä¼¼å®½åº¦
                    # ã€ç»éªŒã€‘WidthFactor * å­—ç¬¦æ•° * å­—é«˜ â‰ˆ å®½åº¦
                    char_count = max(len(attr.TextString.replace("\\P", "")), 1)
                    wf = width / (char_count * height)
                    attr.WidthFactor = wf
                except Exception as e:
                    print(f"[è­¦å‘Š] è®¾ç½® WidthFactor å¤±è´¥: {e}")

            modified = True
    except Exception as e:
        print(f"[è­¦å‘Š] GetAttributes() å¤±è´¥: {e}")

    return modified










## 3 è·å–å±æ€§å—é‡Œå¤šæ®µçº¿çŸ©å½¢çš„åæ ‡å€¼

def huoqu_kuai_pl(blocka):#è¾“å…¥å®ä½“å—ï¼Œå¾—åˆ°å®ä½“å—ä¸­å¤šæ®µçº¿çŸ©å½¢çš„åæ ‡ï¼Œå…¶åæ ‡ä»¥æ’å…¥ç‚¹çš„å®šä¹‰ç‚¹ä¸ºåŸç‚¹
    # è¿æ¥åˆ°AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    kuaiming=blocka.Name

    # è·å–å—å®šä¹‰
    block_def = doc.Blocks.Item(str(kuaiming))

    # è·å–å—å®šä¹‰ä¸­çš„æ‰€æœ‰å¯¹è±¡
    block_objects = list(block_def)

    # æŸ¥æ‰¾ä¸‰è§’å½¢å¹¶åˆ é™¤
    for obj in block_objects:

##        print(obj.ObjectName)
        
        if obj.ObjectName == "AcDbPolyline":

            print(obj.Coordinates)

##>>> huoqu_kuai_pl(kuai[0])
##AcDbPolyline
##(-6000.0, 0.0, 0.0, 0.0, 0.0, 57400.0, -6000.0, 57400.0)
## 



#å®šä¹‰åŸºç‚¹çš„å—

def create_block_with_basepoint():
    # è¿æ¥åˆ°AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # å®šä¹‰å—çš„åŸºç‚¹ä½ç½®
    base_point = vtpnt(10, 10, 0)

    # åˆ›å»ºä¸€ä¸ªæ–°çš„å—
    block = doc.Blocks.Add(base_point, "MyBlock")

    # åœ¨å—ä¸­æ·»åŠ ä¸€ä¸ªåœ†å½¢å®ä½“
    block.AddCircle(base_point, 5)

#å—çš„æ·»åŠ 


def create_block_with_triangle_and_text():
    # è¿æ¥åˆ°AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # åˆ›å»ºæ–°å—
    grip = vtpnt(0, 0)
    blockObj = doc.Blocks.Add(grip, "MyBlock")

    # åœ¨å—ä¸­æ·»åŠ ä¸‰è§’å½¢
    pt1 = vtpnt(0, 0, 0)
    pt2 = vtpnt(10, 0, 0)
    pt3 = vtpnt(5, 10, 0)
    blockObj.AddLine(pt1, pt2)
    blockObj.AddLine(pt2, pt3)
    blockObj.AddLine(pt3, pt1)

    # åœ¨å—ä¸­æ·»åŠ æ–‡å­—å¯¹è±¡
    text_point = vtpnt(2, 2, 0)
    blockObj.AddText("å¤ªç¾äº†", text_point, 2)

    print("å— 'MyBlock' åˆ›å»ºæˆåŠŸ")



def huoqu_kuai_pl(blocka):
    # è¿æ¥åˆ°AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    kuaiming=blocka.Name

    # è·å–å—å®šä¹‰
    block_def = doc.Blocks.Item(str(kuaiming))

    # è·å–å—å®šä¹‰ä¸­çš„æ‰€æœ‰å¯¹è±¡
    block_objects = list(block_def)

    # æŸ¥æ‰¾ä¸‰è§’å½¢å¹¶åˆ é™¤
    for obj in block_objects:

        print(obj.ObjectName)
        if obj.ObjectName == "AcDbPolyline":

            print(obj.Coordinates)

    

# å—çš„è¾¹ç•Œ

def get_bounding_box_of_block(block_name):
    # è¿æ¥åˆ°AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # è·å–å—å®šä¹‰
    block_def = doc.Blocks.Item(block_name)

    min_x, min_y, min_z = float('inf'), float('inf'), float('inf')
    max_x, max_y, max_z = float('-inf'), float('-inf'), float('-inf')

    # éå†å—å®šä¹‰ä¸­çš„æ‰€æœ‰å¯¹è±¡
    for obj in block_def:
        try:
            # å°è¯•è·å–å¯¹è±¡çš„è¾¹ç•Œæ¡†
            lower_left, upper_right = obj.GetBoundingBox()
            
            min_x = min(min_x, lower_left[0])
            min_y = min(min_y, lower_left[1])
            min_z = min(min_z, lower_left[2])
            
            max_x = max(max_x, upper_right[0])
            max_y = max(max_y, upper_right[1])
            max_z = max(max_z, upper_right[2])
        except:
            pass

    return ((min_x, min_y, min_z), (max_x, max_y, max_z))


def create_new_block_with_insert_and_line():
    # è¿æ¥åˆ°AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # æ£€æŸ¥å—åç§°"å—1"æ˜¯å¦å·²ç»å­˜åœ¨
    if "å—3" in [blk.Name for blk in doc.Blocks]:
        print("å—åç§°'å—3'å·²ç»å­˜åœ¨ã€‚è¯·é€‰æ‹©ä¸€ä¸ªæ–°çš„åç§°æˆ–åˆ é™¤ç°æœ‰çš„å—ã€‚")
        return

    # åˆ›å»ºæ–°å—çš„æ’å…¥ç‚¹
    grip = vtpnt(0, 0, 0)
    blockObj1 = doc.Blocks.Add(grip, "å—3")

    # åœ¨å—1ä¸­æ’å…¥MyBlockå—
    insertion_point_for_myblock = vtpnt(10, 10, 0)
    blockObj1.InsertBlock(insertion_point_for_myblock, "MyBlock", 1, 1, 1, 0)

    # åœ¨å—1ä¸­æ·»åŠ ä¸€æ ¹ç›´çº¿æ®µ
    start_point = vtpnt(0, 0, 0)
    end_point = vtpnt(50, 50, 0)
    blockObj1.AddLine(start_point, end_point)

    print("å—1å·²åˆ›å»ºå¹¶æ·»åŠ äº†MyBlockå’Œç›´çº¿æ®µ")

##å—çš„æœç´¢
    
def copy_and_move_blocks_from_layer(layer_name, block_prefix):
    
        
    # ä½¿ç”¨select_tucengå‡½æ•°é€‰æ‹©æŒ‡å®šå›¾å±‚ä¸Šçš„æ‰€æœ‰å¯¹è±¡
    all_objects = select_tuceng(layer_name)
    
    # è¿‡æ»¤å‡ºå—å¯¹è±¡ï¼Œä¸”å—åçš„å‰ä¸¤ä¸ªå­—æ¯ä¸æŒ‡å®šçš„å‰ç¼€åŒ¹é…
    blocks = [obj for obj in all_objects if obj.ObjectName == "AcDbBlockReference" and obj.Name[:2] == block_prefix]
    
    # å®šä¹‰ç§»åŠ¨çš„èµ·å§‹ç‚¹å’Œç»“æŸç‚¹
    vtpnt_from = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 0, 0])
    vtpnt_to = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 2000000, 0])
    
    # å¯¹æ¯ä¸ªå—è¿›è¡Œå¤åˆ¶å’Œç§»åŠ¨æ“ä½œ
    for block in blocks:
        # å¤åˆ¶å—
        copied_block = block.Copy()
        
        # ç§»åŠ¨å¤åˆ¶çš„å—
        copied_block.Move(vtpnt_from, vtpnt_to)

    print(f"Copied and moved {len(blocks)} blocks from layer {layer_name} with prefix {block_prefix}.")

#å—åçš„æ¸…é™¤

def delete_block_name(block_name):#å…ˆåˆ é™¤å—åä¸ºblock_nameçš„å®ä½“å†æ‰§è¡Œè¯¥å‘½ä»¤åˆ é™¤å—åï¼Œå…å¾—å°†æ¥å‘ç”Ÿæ›¿æ¢è­¦å‘Š

    t1=time.time()

    # è·å–å—å®šä¹‰çš„é›†åˆ
    blocks = acad.ActiveDocument.Database.Blocks

    print("len(blocks)",len(blocks))

    # éå†æ‰€æœ‰å—å®šä¹‰
    for block in blocks:
        # æ£€æŸ¥å—åç§°æ˜¯å¦åŒ¹é…
        if block.Name == block_name:
            try:
                # å°è¯•åˆ é™¤å—
                block.Delete()
                print(f"å— '{block_name}' å·²è¢«åˆ é™¤ã€‚")
            except Exception as e:
                print(f"åˆ é™¤å— '{block_name}' æ—¶å‘ç”Ÿé”™è¯¯: {str(e)}")
            break
    else:
        # å¦‚æœæœªæ‰¾åˆ°å—
        print(f"æœªæ‰¾åˆ°åä¸º '{block_name}' çš„å—ã€‚")

    t2 = time.time()

    print("åˆ é™¤å—åè€—æ—¶ï¼š",t2-t1)

#æ›´æ”¹å®ä½“å—å

def rename_block_entity(ent, new_name):
    """
    å°†ç»™å®šå—å‚ç…§å®ä½“ ent çš„å—åæ”¹ä¸º new_nameã€‚
    å¦‚æœ new_name åœ¨ Block è¡¨ä¸­å·²å­˜åœ¨ï¼Œåˆ™è¯¥å®ä½“å°†æŒ‡å‘å·²æœ‰å®šä¹‰ï¼›
    å¦åˆ™å°†é‡å‘½åå®ƒå½“å‰æ‰€å¼•ç”¨çš„å—å®šä¹‰ã€‚
    
    å‚æ•°ï¼š
      ent       -- ä¸€ä¸ª COM å—å‚ç…§å¯¹è±¡ï¼ˆå¦‚ BlockReferenceï¼‰
      new_name  -- ç›®æ ‡å—åï¼ˆå­—ç¬¦ä¸²ï¼‰
    """
    # è·å–å½“å‰æ–‡æ¡£å’Œå—è¡¨

    blocks = doc.Blocks
    old_name = ent.Name

    try:
        # å°è¯•æŸ¥æ‰¾ new_name æ˜¯å¦å·²å­˜åœ¨
        blocks.Item(new_name)
        # å­˜åœ¨ï¼šç›´æ¥è®©è¯¥å®ä½“å¼•ç”¨æ­¤å®šä¹‰
        ent.Name = new_name
    except Exception:
        # ä¸å­˜åœ¨ï¼šé‡å‘½åå®ƒå½“å‰æ‰€å¼•ç”¨çš„å®šä¹‰
        blk_def = blocks.Item(old_name)
        blk_def.Name = new_name

#&&% ç”±å—åé€‰æ‹©å®ä¾‹

def get_block_instances(block_name: str, max_retries: int = 5):
    """
    æ ¹æ®ç»™å®šçš„å—å®šä¹‰åï¼Œæ£€ç´¢å½“å‰å›¾å½¢ä¸­æ‰€æœ‰å¯¹åº”çš„å—å‚ç…§å®ä¾‹ï¼ˆBlockReferenceï¼‰ï¼Œ
    è¿”å›è¿™äº› COM å¯¹è±¡çš„åˆ—è¡¨ã€‚

    å‚æ•°ï¼š
      block_name     â€“ å—å®šä¹‰çš„åç§°ï¼ˆå­—ç¬¦ä¸²ï¼‰ï¼Œå¦‚ "MyBlock"
      max_retries    â€“ è°ƒç”¨ select_kuai æ—¶çš„æœ€å¤§é‡è¯•æ¬¡æ•°

    è¿”å›ï¼š
      instances     â€“ åŒ…å«æ‰€æœ‰åŒ¹é…å—å‚ç…§çš„åˆ—è¡¨ï¼ˆå¦‚æœæœªæ‰¾åˆ°æˆ–è€…å‡ºé”™ï¼Œè¿”å›ç©ºåˆ—è¡¨ï¼‰
    """
    # â‘  å…ˆè°ƒç”¨ select_kuai() æ‹¿åˆ°æ‰€æœ‰å—å®ä¾‹ COM å¯¹è±¡ï¼ˆæ‰å¹³åˆ—è¡¨ï¼‰
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        print(f"[é”™è¯¯] è°ƒç”¨ select_kuai å¤±è´¥ï¼š{e}")
        return []

    instances = []
    # â‘¡ select_kuai å·²ç»æ˜¯æ‰€æœ‰å—å®ä¾‹çš„åˆ—è¡¨ï¼Œç›´æ¥éå†
    for ent in all_blocks:
        try:
            # å¯¹äºå—å‚ç…§ï¼ŒEntityName é€šå¸¸æ˜¯ "AcDbBlockReference"
            # ä¸”æˆ‘ä»¬è¦ç­›é€‰ Name æ°å¥½ç­‰äº block_name çš„é‚£äº›
            if getattr(ent, "EntityName", "") == "AcDbBlockReference" and getattr(ent, "Name", "") == block_name:
                instances.append(ent)
        except Exception:
            # å¦‚æœæŸä¸ª COM å¯¹è±¡æ²¡æœ‰ EntityName/Name å±æ€§ï¼Œå°±è·³è¿‡
            continue

    print(f"é€‰æ‹©åˆ°åä¸º{block_name}çš„å®ä¾‹å—{len(instances)}ä¸ª")

    return instances



#&&% ä»å—å®ä½“å¯¹è±¡è·å–å…¶å†…éƒ¨comå¯¹è±¡

def get_entities_from_block_reference(block_ref):
    """
    è·å–å—å¼•ç”¨å¯¹è±¡ä¸­çš„æ‰€æœ‰å­å®ä½“ï¼ˆCOMå¯¹è±¡å½¢å¼ï¼‰ã€‚

    å‚æ•°ï¼š
        block_ref: å—å¼•ç”¨å¯¹è±¡ï¼ˆAcDbBlockReferenceï¼‰
        doc: å½“å‰ AutoCAD æ–‡æ¡£å¯¹è±¡ï¼ˆDocumentï¼‰

    è¿”å›ï¼š
        entities: å­å®ä½“åˆ—è¡¨
    """
    try:
        block_name = block_ref.EffectiveName
        block_def = doc.Blocks.Item(block_name)
        entities = [ent for ent in block_def]
        print(f"[OK] è·å–åˆ° {len(entities)} ä¸ªå­å¯¹è±¡")
        return entities
    except Exception as e:
        print(f"[é”™è¯¯] è·å–å¤±è´¥ï¼š{e}")
        return []

    




#ä»¥å—çš„æ–¹å¼æ’å…¥æ–‡ä»¶

def insert_block_into_autocad(block_file_path, insertion_point=(0, 0, 0), scale=(1, 1, 1), rotation=0):
    """
    ä»¥å—çš„æ–¹å¼æ’å…¥ DWG æ–‡ä»¶åˆ° AutoCAD ä¸­

    :param block_file_path: å—æ–‡ä»¶çš„è·¯å¾„ï¼Œé€šå¸¸ä¸º DWG æ–‡ä»¶è·¯å¾„
    :param insertion_point: æ’å…¥ä½ç½®çš„ä¸‰å…ƒç»„ (x, y, z)
    :param scale: å—çš„ç¼©æ”¾æ¯”ä¾‹ä¸‰å…ƒç»„ (sx, sy, sz)
    :param rotation: å—çš„æ—‹è½¬è§’åº¦ï¼ˆå¼§åº¦ï¼‰
    """
    try:

        # å®šä¹‰å—æ’å…¥ç‚¹
        insertion_point = (insertion_point[0], insertion_point[1], insertion_point[2])

        # æ’å…¥å—ï¼Œä½¿ç”¨ InsertBlock æ–¹æ³•
        block = ms.InsertBlock(insertion_point, block_file_path, scale[0], scale[1], scale[2], rotation)

        print(f"[OK] å—å·²æ’å…¥ï¼Œæ–‡ä»¶ï¼š{block_file_path}ï¼Œæ’å…¥ç‚¹ï¼š{insertion_point}ï¼Œç¼©æ”¾ï¼š{scale}ï¼Œæ—‹è½¬è§’åº¦ï¼š{rotation}")

    except Exception as e:
        print(f"[é”™è¯¯] æ’å…¥å—æ—¶å‡ºé”™ï¼š{e}")



#ä¸ç‚¸å¼€

def insert_standard_block(block_dwg,
                          insertion_point=(0, 0, 0),
                          scale=(1, 1, 1),
                          rotation=0,
                          wait=0.3):
    """
    ä¸ç‚¸å¼€ï¼Œ
    å…¨ç¨‹æ— äº¤äº’å¯¹è¯æ¡†ã€‚

    block_dwg: æ ‡å‡†å— DWG è·¯å¾„
    insertion_point: (x,y,z)
    scale: (sx,sy,sz)
    rotation: æ—‹è½¬è§’åº¦ï¼ˆåº¦ï¼‰
    """
    before = select_kuai()
    before_handles = {b.Handle for b in before}


    if not os.path.isfile(block_dwg):
        raise FileNotFoundError(block_dwg)

    # å‡†å¤‡å‚æ•°
    x, y, z    = insertion_point
    sx, sy, sz = scale
    path       = os.path.abspath(block_dwg).replace("\\", "/")

    # 1) æ’å…¥å—
    insert_cmd = (
        "-INSERT\n"
        f"\"{path}\"\n"    # æ–‡ä»¶è·¯å¾„è¦åŠ åŒå¼•å·
        f"{x},{y},{z}\n"
        f"{sx}\n"
        f"{sy}\n"
        f"{sz}\n"
        f"{rotation}\n"
    )
    doc.SendCommand(insert_cmd)
    time.sleep(wait)

    doc.SendCommand("RE\n")
    doc.SendCommand("Z\nE\n")
    time.sleep(wait)

    after = select_kuai()
    new_refs = [b for b in after if b.Handle not in before_handles]
    if not new_refs:
        print("[è­¦å‘Š] æœªæ£€æµ‹åˆ°ä»»ä½•æ–°æ’å…¥çš„å—å¼•ç”¨")
        return []

    results = []
    for blk in new_refs:
        # 5. å…ˆå°†å®ƒæ—‹è½¬å½’é›¶
        try:
            blk.Rotation = 0
        except Exception:
            pass

        # 7. å–å®ƒçš„åŒ…å›´ç›’å››è§’
        p1, p2 = blk.GetBoundingBox()
        minx, miny, minz = p1
        maxx, maxy, maxz = p2
        corners = [
            (minx, miny, minz),  # å·¦ä¸‹
            (minx, maxy, minz),  # å·¦ä¸Š
            (maxx, maxy, minz),  # å³ä¸Š
            (maxx, miny, minz),  # å³ä¸‹
        ]

        results.append((blk, corners))

    return results


#ç‚¸å¼€
def insert_and_explode_dwg(block_dwg,
                           insertion_point=(0, 0, 0),
                           scale=(1, 1, 1),
                           rotation=0,
                           wait=0.3):
    """
    å°†ä¸€ä¸ª WBLOCK å¯¼å‡ºçš„æ ‡å‡†å— DWG æ’å…¥åˆ°å½“å‰å›¾ï¼Œ
    å¹¶ç«‹å³ EXPLODE æˆæ™®é€šå›¾å…ƒï¼ˆä¸ä¿ç•™å—å¼•ç”¨ï¼‰ã€‚

    å‚æ•°:
        block_dwg: æ ‡å‡†å— DWG è·¯å¾„
        insertion_point: æ’å…¥ç‚¹ (x,y,z)
        scale: (sx,sy,sz)
        rotation: æ—‹è½¬è§’åº¦ï¼ˆåº¦ï¼‰
        wait: æ¯æ­¥å‘½ä»¤åç­‰å¾…ç§’æ•°
    """

    before = select_kuai()
    before_handles = {b.Handle for b in before}


    if not os.path.isfile(block_dwg):
        raise FileNotFoundError(block_dwg)

    # å‡†å¤‡å‚æ•°
    x, y, z    = insertion_point
    sx, sy, sz = scale
    path       = os.path.abspath(block_dwg).replace("\\", "/")

    # 1) æ’å…¥å—
    insert_cmd = (
        "-INSERT\n"
        f"\"{path}\"\n"    # æ–‡ä»¶è·¯å¾„è¦åŠ åŒå¼•å·
        f"{x},{y},{z}\n"
        f"{sx}\n"
        f"{sy}\n"
        f"{sz}\n"
        f"{rotation}\n"
    )
    doc.SendCommand(insert_cmd)
    time.sleep(wait)

    doc.SendCommand("RE\n")
    doc.SendCommand("Z\nE\n")
    time.sleep(wait)

    # 2) EXPLODE â€œLastâ€   ï¼ˆç‚¸å¼€æœ€æ–°æ’å…¥çš„å—å¼•ç”¨ï¼‰
    explode_cmd = (
        "EXPLODE\n"
        "L\n"    # Last
        "\n"     # å®Œæˆé€‰æ‹©
    )
    doc.SendCommand(explode_cmd)
    time.sleep(wait)

    print(f"[OK] å·²æ’å…¥å¹¶ç‚¸å¼€ï¼š{os.path.basename(path)} @ ({x},{y},{z})")


    after = select_kuai()
    new_refs = [b for b in after if b.Handle not in before_handles]
    if not new_refs:
        print("[è­¦å‘Š] æœªæ£€æµ‹åˆ°ä»»ä½•æ–°æ’å…¥çš„å—å¼•ç”¨")
        return []

    results = []
    for blk in new_refs:
        # 5. å…ˆå°†å®ƒæ—‹è½¬å½’é›¶
        try:
            blk.Rotation = 0
        except Exception:
            pass

        # 7. å–å®ƒçš„åŒ…å›´ç›’å››è§’
        p1, p2 = blk.GetBoundingBox()
        minx, miny, minz = p1
        maxx, maxy, maxz = p2
        corners = [
            (minx, miny, minz),  # å·¦ä¸‹
            (minx, maxy, minz),  # å·¦ä¸Š
            (maxx, maxy, minz),  # å³ä¸Š
            (maxx, miny, minz),  # å³ä¸‹
        ]

        results.append((blk, corners))

    return results,blk


#&&% åŒçº¿ç¨‹æ’å…¥å—


##ç¡®ä¿ç‚¸å¼€å—å¹¶è·å–å—å†…å¯¹è±¡
def safe_explode_and_delete(bk, ci=3, delay=1.0):
    """
    å¯¹å—å¯¹è±¡ bk æ‰§è¡Œå®‰å…¨çš„ Explode ä¸ Delete æ“ä½œï¼š
      1. æœ€å¤šå°è¯• ci æ¬¡è°ƒç”¨ bk.Explode()ï¼Œ
         æ¯æ¬¡è°ƒç”¨åç­‰å¾… delay ç§’ï¼Œå†æ£€æŸ¥è¿”å›çš„å®ä½“åˆ—è¡¨é•¿åº¦æ˜¯å¦ > 1ã€‚
      2. å¦‚æœ ci æ¬¡éƒ½æ²¡æœ‰æˆåŠŸï¼ˆlen(LP) <= 1ï¼‰ï¼ŒæŠ›å‡º RuntimeErrorã€‚
      3. å¦‚æœ Explode æˆåŠŸï¼Œæœ€å¤šå°è¯• ci æ¬¡è°ƒç”¨ bk.Delete()ï¼Œé‡é”™åˆ™ç»§ç»­é‡è¯•ã€‚
      4. è¿”å›ç¬¬ä¸€æ¬¡æˆåŠŸ Explode æ—¶å¾—åˆ°çš„å®ä½“åˆ—è¡¨ LPã€‚

    :param bk:    è¦ç‚¸å¼€çš„å—å¼•ç”¨ COM å¯¹è±¡
    :param ci:    æœ€å¤§å°è¯•æ¬¡æ•°ï¼Œé»˜è®¤ 3
    :param delay: æ¯æ¬¡ Explode åçš„ç­‰å¾…æ—¶é—´ï¼ˆç§’ï¼‰ï¼Œé»˜è®¤ 1.0
    :return:      æˆåŠŸ Explode åè¿”å›çš„æ–°å®ä½“åˆ—è¡¨ LP
    :raises:      RuntimeError å¦‚æœ Explode åœ¨ ci æ¬¡å†…éƒ½å¤±è´¥
    """
    LP = []
    # 1ï¸âƒ£ é‡è¯• Explode
    for attempt in range(1, ci + 1):
        try:
            LP = bk.Explode()
        except Exception as e:
            LP = []  # å¦‚æœè°ƒç”¨å¤±è´¥ï¼Œè§†ä¸ºæ²¡æœ‰ç‚¸å¼€
        time.sleep(delay)

        # å°è¯•è·å–é•¿åº¦
        try:
            count = len(LP)
        except Exception:
            # å¦‚æœ LP ä¸æ˜¯æ ‡å‡†åºåˆ—ï¼Œå°±æŠŠå®ƒè½¬æ¢ä¸€ä¸‹
            LP = list(LP)
            count = len(LP)

        if count > 1:
            # ç‚¸å¼€æˆåŠŸ
            break
        else:
            # ç¬¬ attempt æ¬¡ Explode æœªæˆåŠŸï¼Œç»§ç»­é‡è¯•
            continue
    else:
        raise RuntimeError(f"Explode åœ¨ {ci} æ¬¡å°è¯•åä»æœªæˆåŠŸ (len(LP)={len(LP)})")

    # 2ï¸âƒ£ Explode æˆåŠŸåï¼Œé‡è¯• Delete
    for attempt in range(1, ci + 1):
        try:
            bk.Delete()
            break
        except Exception:
            # Delete å‡ºé”™ï¼Œç»§ç»­ä¸‹æ¬¡é‡è¯•
            continue
    # å¦‚æœ ci æ¬¡éƒ½å¤±è´¥ï¼Œä¹Ÿä¸æŠ›é”™è¯¯ï¼Œåªæ˜¯æ”¾å¼ƒåˆ é™¤
    return LP





#&&% è·å–é¢ç§¯è¶³å¤Ÿå¤§çš„å…¨éƒ¨éåŒåå—å®ä¾‹

def get_large_block_instances(
    area_threshold: float,
    tol: float = 100.0,
    max_retries: int = 5
) -> list:
    """
    è·å–æ¨¡å‹ç©ºé—´ä¸­æ‰€æœ‰å—å®ä¾‹ï¼Œç­›é€‰å‡ºâ€œåŒ…å›´ç›’é¢ç§¯å¤§äº area_thresholdâ€ çš„å—ï¼Œ
    å¹¶æŒ‰é¢ç§¯å»é‡ï¼šå¦‚æœä¸¤ä¸ªå—çš„é¢ç§¯å·®å€¼å°äº tolï¼Œåˆ™åªä¿ç•™å…¶ä¸­ä¸€ä¸ªã€‚
    è¿”å›ä¸€ä¸ªæŒ‰å‡ºç°é¡ºåºå»é‡åçš„ COM å¯¹è±¡åˆ—è¡¨ã€‚

    :param area_threshold: é¢ç§¯é˜ˆå€¼ï¼ˆå•ä½ä¸ CAD åæ ‡ç›¸åŒï¼Œä¾‹å¦‚ä»¥å¹³æ–¹å›¾çº¸å•ä½è®¡ï¼‰ã€‚
    :param tol:            é¢ç§¯å»é‡çš„å®¹å·®å€¼ï¼ˆé»˜è®¤ 100ï¼‰ï¼Œè‹¥ä¸¤ä¸ªå—é¢ç§¯å·®å°äº tolï¼Œè§†ä¸ºåŒä¸€å—ã€‚
    :param max_retries:    è°ƒç”¨ select_kuai æ—¶çš„æœ€å¤§é‡è¯•æ¬¡æ•°ï¼Œé»˜è®¤ 5 æ¬¡ã€‚
    :return:               åŒ…å«æ‰€æœ‰â€œé¢ç§¯ > area_thresholdâ€ä¸”å»é‡åçš„å—å¼•ç”¨ COM å¯¹è±¡åˆ—è¡¨ï¼Œ
                          å¦‚æœ select_kuai è°ƒç”¨å¤±è´¥ï¼Œä¼šè¿”å›ç©ºåˆ—è¡¨ã€‚
    """
    # â‘  è°ƒç”¨ select_kuaiï¼Œæ¯éš” max_retries æ¬¡é‡è¿ä¸€æ¬¡
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        print(f"[é”™è¯¯] è°ƒç”¨ select_kuai å¤±è´¥ï¼š{e}")
        return []

    # ä¸´æ—¶åˆ—è¡¨ï¼Œç”¨äºæ”¶é›†å·²åŠ å…¥ç»“æœçš„å—é¢ç§¯ï¼ˆç”¨äºå»é‡ï¼‰
    seen_areas = []
    large_blocks = []

    for blk in all_blocks:
        try:
            # GetBoundingBox è¿”å› (ll_point, ur_point)
            ll_point, ur_point = blk.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point
            width = abs(x2 - x1)
            height = abs(y2 - y1)
            area = width * height
        except Exception:
            # å¦‚æœæŸäº›å—æ— æ³•è·å–åŒ…å›´ç›’ï¼Œåˆ™è·³è¿‡
            continue

        # é¢ç§¯åˆ¤æ–­
        if area > area_threshold:
            # æ£€æŸ¥å½“å‰ area æ˜¯å¦å·²è¿‘ä¼¼å‡ºç°åœ¨ seen_areas ä¸­
            duplicate = False
            for existing_area in seen_areas:
                if abs(area - existing_area) < tol:
                    duplicate = True
                    break

            if not duplicate:
                seen_areas.append(area)
                large_blocks.append(blk)

    return large_blocks

#ä»comå¯¹è±¡ä¸­ï¼Œæ ¹æ®å…¶å¤–åŒ…ç›’çš„çŸ©å½¢çš„é•¿è¾¹ä¸çŸ­è¾¹çš„æ¯”å€¼å’Œé¢ç§¯åœ¨160000000åˆ°1000000000ä¸¤ä¸ªæ¡ä»¶ç­›ç®—

#&&% ç¡®å®šåˆä¹æ ‡å‡†æ‰“å°è¦æ±‚çš„è‡ªå»ºå¤šæ®µçº¿åŒºåŸŸ

def get_large_block_instances_with_tolerance(max_retries: int = 5, area_threshold: float = 70 ):
    """
    è·å–å½“å‰ DWG ä¸­æ‰€æœ‰å¤§å°ºå¯¸å—å®ä¾‹
    A3 1:100çš„æ­£å¸¸å€¼> 1240000000.000000
   
    """
    # â‘  å°è¯•è·å–æ‰€æœ‰å—å®ä¾‹
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        print(f"[é”™è¯¯] è°ƒç”¨ select_kuai å¤±è´¥ï¼š{e}")
        return []

    # â‘¡ å…ˆç­›é€‰é¢ç§¯ >= area_threshold çš„å—
    LB = []
    for blk in all_blocks:
        try:
            ll_point, ur_point = blk.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point
            width = abs(x2 - x1)
            height = abs(y2 - y1)
            area = width * height
        except Exception:
            # è¯»å–åŒ…å›´ç›’å¤±è´¥ï¼Œè·³è¿‡
            continue

        if area >= area_threshold:
            # é™„å¸¦é¢ç§¯ä¿¡æ¯ï¼Œä¾¿äºåç»­æ¯”è¾ƒ
            LB.append((blk, area))

    return LC


#&&% ç¡®å®šåˆä¹å¹¿ä¹‰æ ‡å‡†æ‰“å°è¦æ±‚çš„è‡ªå»ºå¤šæ®µçº¿åŒºåŸŸ




#&&% å—å†…åæ ‡è½¬æ¢æˆä¸–ç•Œåæ ‡ï¼ˆé€‚åˆå¹³é¢ä¸Šçš„ä¸€èˆ¬å—ï¼‰

def transform_point_by_block(block_ref, local_pt):
    """
    å°†å—å†…éƒ¨åæ ‡ local_pt = (lx, ly, lz) è½¬æ¢ä¸ºä¸–ç•Œåæ ‡ï¼š
      1. æŒ‰ block_ref çš„æ¯”ä¾‹ç¼©æ”¾ (XScaleFactor, YScaleFactor, ZScaleFactor)
      2. æŒ‰ block_ref.Rotation æ—‹è½¬ï¼ˆç»• Z è½´ï¼Œå¼§åº¦åˆ¶ï¼‰
      3. å¹³ç§»åˆ° block_ref.InsertionPoint = (ix, iy, iz)

    è¿”å› (wx, wy, wz)ï¼š
      wx = ix + (lx * sx * cosÎ¸ - ly * sy * sinÎ¸)
      wy = iy + (lx * sx * sinÎ¸ + ly * sy * cosÎ¸)
      wz = iz + (lz * sz)
    """
    ix, iy, iz = block_ref.InsertionPoint
    sx = block_ref.XScaleFactor
    sy = block_ref.YScaleFactor
    sz = block_ref.ZScaleFactor
    theta = block_ref.Rotation  # å•ä½ï¼šå¼§åº¦

    lx, ly, lz = local_pt
    # ç¼©æ”¾åå†æ—‹è½¬
    x_scaled = lx * sx
    y_scaled = ly * sy

    cos_t = math.cos(theta)
    sin_t = math.sin(theta)
    xr = x_scaled * cos_t - y_scaled * sin_t
    yr = x_scaled * sin_t + y_scaled * cos_t
    zr = lz * sz

    wx = ix + xr
    wy = iy + yr
    wz = iz + zr
    return (wx, wy, wz)


def select_block_by_name(block_name: str, max_retries: int = 5):
    """
    ä» *æ¨¡å‹ç©ºé—´* å¿«é€Ÿé€‰å‡ºæŒ‡å®šå—åçš„æ‰€æœ‰å®ä¾‹ï¼Œè¿”å›å®ä½“åˆ—è¡¨ã€‚
    ä¸éå† ModelSpaceï¼Œé€Ÿåº¦ä¸ SelectionSet ç›¸åŒé‡çº§ã€‚
    """
    t0, last_exc = time.time(), None

    for attempt in range(1, max_retries + 1):
        try:
            # 1) åˆ æ—§é€‰æ‹©é›†
            with suppress(Exception):
                doc.SelectionSets.Item("SS_block_by_name").Delete()

            # 2) æ–°å»ºé€‰æ‹©é›†
            ss = doc.SelectionSets.Add("SS_block_by_name")

            # 3) è¿‡æ»¤å™¨ï¼š0=INSERT, 2=å—å, 410=å¸ƒå±€å"Model"
            filterType  = vtInt([0, 2, 410])
            filterData  = vtVariant(["INSERT", block_name, "Model"])

            # acSelectionSetAll = 5
            ss.Select(5, 0, 0, filterType, filterData)

            lb = list(ss)
            print(f"[OK] é€‰åˆ° {len(lb)} ä¸ª â€œ{block_name}â€ï¼ˆ{time.time() - t0:.3f}sï¼Œç¬¬{attempt}æ¬¡ï¼‰")
            return lb

        except Exception as e:
            last_exc = e
            print(f"[è­¦å‘Š] ç¬¬ {attempt} æ¬¡å¤±è´¥ï¼š{e!r}")
            time.sleep(0.3)

    print(f"[é”™è¯¯] {max_retries} æ¬¡ä»å¤±è´¥ï¼š{last_exc!r}")
    return []

def get_all_block_definitions(doc=None):
    """
    è¿”å›å½“å‰ DWG æ–‡æ¡£ä¸­æ‰€æœ‰å—å®šä¹‰ï¼ˆBlockTableRecordï¼‰åˆ—è¡¨ã€‚

    :param doc: AutoCAD Document å¯¹è±¡ï¼Œè‹¥ä¸º None åˆ™ä»å½“å‰æ¿€æ´»æ–‡æ¡£è·å–
    :return: list of BlockTableRecord COM å¯¹è±¡
    blk.Nameè·å–å—å

    """
    # å¦‚æœå¤–éƒ¨æ²¡ä¼ å…¥ docï¼Œå°±ä»å½“å‰æ¿€æ´»æ–‡æ¡£æ‹¿
    if doc is None:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        doc  = acad.ActiveDocument

    blocks = []
    count = doc.Blocks.Count  # å—å®šä¹‰æ€»æ•°
    # Blocks é›†åˆåœ¨ COM ä¸­æ˜¯ 0â€¦Countâˆ’1 ç¼–å·
    for i in range(count):
        try:
            blk = doc.Blocks.Item(i)
            blocks.append(blk)
        except Exception:
            # è·³è¿‡ä»»ä½•è®¿é—®ä¸æˆåŠŸçš„ç´¢å¼•
            continue

    return blocks

def purge_block(block_name: str, quiet: bool = False):
    """
    åˆ é™¤æŒ‡å®šå—çš„æ‰€æœ‰å®ä¾‹ï¼Œå¹¶å½»åº•æ¸…é™¤è¯¥å—å®šä¹‰ã€‚
    
    æ­¥éª¤ï¼š
      1. åœ¨æ¨¡å‹ç©ºé—´åˆ é™¤æ‰€æœ‰åŒå INSERT å®ä¾‹
      2. åœ¨æ¯ä¸ªå¸ƒå±€çš„ Block (PaperSpace) åˆ é™¤æ‰€æœ‰åŒå INSERT å®ä¾‹
      3. è°ƒç”¨ PurgeAll() æ¸…ç†æœªç”¨å®šä¹‰
      4. å†æ¬¡å°è¯•åˆ é™¤å—å®šä¹‰
    
    :param block_name: è¦æ¸…ç†çš„å—åç§°ï¼ˆåŒºåˆ†å¤§å°å†™ï¼‰
    :param quiet: True åˆ™ä¸æ‰“å°è¿‡ç¨‹ä¿¡æ¯
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    # --- 1) æ¨¡å‹ç©ºé—´ ---
    removed = 0
    for ent in list(doc.ModelSpace):
        if ent.ObjectName in ("AcDbBlockReference", "AcDbMInsertBlock") and getattr(ent, "Name", "") == block_name:
            with suppress(Exception):
                ent.Delete()
                removed += 1

    # --- 2) å„å¸ƒå±€å—ç©ºé—´ï¼ˆPaperSpaceï¼‰ ---
    for layout in doc.Layouts:
        with suppress(Exception):
            block_space = layout.Block
            for ent in list(block_space):
                if ent.ObjectName in ("AcDbBlockReference", "AcDbMInsertBlock") and getattr(ent, "Name", "") == block_name:
                    with suppress(Exception):
                        ent.Delete()
                        removed += 1

    time.sleep(0.2)
    if not quiet:
        print(f"â„¹ å…±åˆ é™¤ {removed} ä¸ª â€œ{block_name}â€ å®ä¾‹")

    # --- 3) PurgeAll æ¸…ç†æ‰€æœ‰æœªè¢«å¼•ç”¨çš„å®šä¹‰ ---
    with suppress(Exception):
        doc.PurgeAll()
        if not quiet:
            print("[OK] PurgeAll æ¸…ç†æœªç”¨å®šä¹‰")

    # --- 4) åˆ é™¤å—å®šä¹‰ ---
    try:
        blk = doc.Blocks.Item(block_name)
        blk.Delete()
        if not quiet:
            print(f"[OK] å·²åˆ é™¤å—å®šä¹‰ï¼š{block_name}")
    except Exception as e:
        if not quiet:
            print(f"[è­¦å‘Š] åˆ é™¤å—å®šä¹‰å¤±è´¥ï¼ˆä»æœ‰éšæ€§å¼•ç”¨ï¼Ÿï¼‰ï¼š{e}")

    if not quiet:
        print(f"â„¹ å®Œæˆå¯¹ '{block_name}' çš„æ¸…ç†ã€‚")

def purge_unused_blocks(quiet: bool = False):
    """
    ä¸€æ¬¡æ€§æ¸…é™¤æ‰€æœ‰æœªè¢«ä»»ä½• INSERT å®ä¾‹å¼•ç”¨çš„å—å®šä¹‰ã€‚
    é€Ÿåº¦å¿«ã€å¯é æ€§é«˜ï¼ˆä¸ç”¨é€ä¸ª SelectionSet æ£€æµ‹ï¼‰ã€‚
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    # 1) è®°å½•æ¸…ç†å‰çš„å—ååˆ—è¡¨
    before = []
    for i in range(doc.Blocks.Count):
        with suppress(Exception):
            name = doc.Blocks.Item(i).Name
            before.append(name)

    # 2) è°ƒç”¨ PurgeAll ä¸€æ¬¡æ€§æ¸…ç†
    t0 = time.time()
    with suppress(Exception):
        doc.PurgeAll()
    t1 = time.time()

    # 3) è®°å½•æ¸…ç†åçš„å—ååˆ—è¡¨
    after = []
    for i in range(doc.Blocks.Count):
        with suppress(Exception):
            name = doc.Blocks.Item(i).Name
            after.append(name)

    # 4) è®¡ç®—å·®é›†
    removed = [name for name in before if name not in after]

    if not quiet:
        print(f"[OK] PurgeAll æ¸…ç†å®Œæˆï¼Œè€—æ—¶ {t1 - t0:.3f}s")
        print(f"â„¹ å…±ç§»é™¤ {len(removed)} ä¸ªæœªä½¿ç”¨å—ï¼š")
        for nm in removed:
            print("   Â·", nm)

    return removed


def get_selected_blockreference_names():
    """
    ä½¿ç”¨ pmxz() é€‰æ‹©å®ä½“ï¼Œå¹¶è¿”å›æ‰€æœ‰å—å¼•ç”¨ï¼ˆAcDbBlockReferenceï¼‰çš„å—ååˆ—è¡¨ã€‚

    :return: list of strï¼Œæ‰€é€‰å—å¼•ç”¨çš„ Name å±æ€§åˆ—è¡¨
    """
    try:
        # è°ƒç”¨å·²æœ‰å‡½æ•°è·å–å½“å‰é€‰æ‹©é›†ï¼ˆè¿”å› COM å¯¹è±¡åˆ—è¡¨ï¼‰
        entities = pmxz()
    except Exception as e:
        print(f"[é”™è¯¯] è°ƒç”¨ pmxz() å¤±è´¥ï¼š{e}")
        return []

    block_names = []
    for ent in entities:
        try:
            # åªå¤„ç†å—å¼•ç”¨
            if getattr(ent, "ObjectName", "") == "AcDbBlockReference":
                # Name å±æ€§å°±æ˜¯å—å
                name = getattr(ent, "Name", None)
                if name:
                    block_names.append(name)
        except Exception as e:
            # æŸäº› COM å¯¹è±¡å¯èƒ½ä¸æ”¯æŒä¸Šè¿°å±æ€§ï¼Œå¿½ç•¥å®ƒ
            continue

    return block_names



def delete_layer(layername: str):
    """
    åˆ é™¤å½“å‰ DWG æ–‡ä»¶ä¸­åä¸º layername çš„å›¾å±‚ã€‚
    - å¦‚æœå›¾å±‚ä¸å­˜åœ¨ï¼Œç›´æ¥è¿”å›ã€‚
    - å¦‚æœå›¾å±‚æ˜¯å½“å‰å±‚ï¼Œåˆ™åˆ‡æ¢åˆ° 0 å±‚åå†åˆ é™¤ã€‚
    - åˆ é™¤å‰ä¼šå°è¯•è§£é”ã€å»æ‰å†»ç»“/æ‰“å°é”ã€‚
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    layers = doc.Layers

    try:
        layer = layers.Item(layername)
    except Exception:
        print(f"[è­¦å‘Š] å›¾å±‚ {layername} ä¸å­˜åœ¨ã€‚")
        return

    # å¦‚æœè¯¥å›¾å±‚æ˜¯å½“å‰å±‚ï¼Œåˆ‡æ¢åˆ° "0"
    if doc.ActiveLayer.Name == layername:
        doc.ActiveLayer = layers.Item("0")

    # è§£é”ã€è§£å†»ã€å»æ‰æ‰“å°é”å®šï¼Œé¿å…åˆ é™¤å¤±è´¥
    try:
        layer.Lock = False
        layer.Freeze = False
        layer.Plottable = True
    except Exception:
        pass

    try:
        layer.Delete()
        print(f"[OK] å›¾å±‚ {layername} å·²åˆ é™¤ã€‚")
    except Exception as e:
        print(f"[é”™è¯¯] åˆ é™¤å¤±è´¥ï¼š{e}")

# ä½¿ç”¨æ–¹æ³•
# delete_layer("MyLayer")



#&&&&%% ç¬¬åéƒ¨åˆ†  éå›¾å½¢å¯¹è±¡å¤„ç†


"""
è¯¥æ¨¡å—ç ”ç©¶éå›¾å½¢å¯¹è±¡å¤„ç†é—®é¢˜ 

"""


#  ä¸»å‡½æ•°
#  (1)
# å›¾å±‚æ“ä½œ


#&&% *** å°†å±å¹•æ‰€é€‰å¯¹è±¡èµ‹äºˆåˆ°æŒ‡å®šå›¾å±‚

@alias("s1")
def sc_objs_to_layer(layer_name,cl=256):

    def pmxz_new():
        """
        äººå·¥é€‰æ‹©å¯¹è±¡ï¼Œè¿”å›æ‰€é€‰å®ä½“å¯¹è±¡åˆ—è¡¨ã€‚
        è‡ªåŠ¨æ¸…ç†å·²æœ‰çš„ "SS1" é€‰æ‹©é›†ã€‚
        """
        try:
            # å¦‚æœ "SS1" å·²å­˜åœ¨ï¼Œå…ˆåˆ é™¤
            try:
                ss = doc.SelectionSets.Item("SS1")
                ss.Delete()
            except:
                pass  # å¦‚æœä¸å­˜åœ¨å°±å¿½ç•¥

            selection = doc.SelectionSets.Add("SS1")
            selection.SelectOnScreen()
            objs = [selection.Item(i) for i in range(selection.Count)]
            selection.Delete()
            return objs
        except Exception as e:
            print("[é”™è¯¯] é€‰æ‹©å¤±è´¥:", e)
            return []



    layers = doc.Layers

    try:
        layer = layers.Item(layer_name)
    except:
        layer = layers.Add(layer_name)
        print(f"ğŸŸ¢ å·²æ–°å»ºå›¾å±‚ï¼š{layer_name}")

    LB=pmxz_new()

    for x in LB:

        x.Layer = layer_name

        x.color = cl


    return LB



#  è¯¥å‡½æ•°ç³»åˆ—åŒ…æ‹¬å¦‚ä¸‹ä¸€äº›å‡½æ•°

def create_layers_from_list(layer_names):
    """
    åˆ›å»ºåˆ—è¡¨ä¸­æŒ‡å®šçš„å›¾å±‚ï¼Œå¦‚æœå›¾å±‚å·²å­˜åœ¨åˆ™è·³è¿‡ã€‚

    å‚æ•°ï¼š
        layer_names: åŒ…å«å›¾å±‚åç§°çš„å­—ç¬¦ä¸²åˆ—è¡¨
    """
    try:
        _, doc = get_acad_doc()
        layers = doc.Layers
        created = 0
        skipped = 0

        for name in layer_names:
            try:
                _ = layers.Item(name)  # æ£€æŸ¥æ˜¯å¦å·²å­˜åœ¨
                print(f"â© å›¾å±‚å·²å­˜åœ¨ï¼š{name}")
                skipped += 1
            except:
                layers.Add(name)
                print(f"[OK] æ–°å»ºå›¾å±‚ï¼š{name}")
                created += 1

        print(f"\nğŸ“Š æ€»è®¡ï¼šæ–°å»º {created} ä¸ªå›¾å±‚ï¼Œè·³è¿‡ {skipped} ä¸ªå·²æœ‰å›¾å±‚")

    except Exception as e:
        print("[é”™è¯¯] åˆ›å»ºå›¾å±‚æ—¶å‡ºé”™ï¼š", e)


def delete_layers_from_list(layer_names):
    """
    åˆ é™¤åˆ—è¡¨ä¸­æŒ‡å®šçš„å›¾å±‚

    å‚æ•°ï¼š
        layer_names: åŒ…å«å›¾å±‚åç§°çš„å­—ç¬¦ä¸²åˆ—è¡¨

    è¿”å›ï¼š
        dict: {'deleted': åˆ é™¤æˆåŠŸçš„å›¾å±‚åˆ—è¡¨, 'failed': åˆ é™¤å¤±è´¥çš„å›¾å±‚åˆ—è¡¨}
    """
    try:
        _, doc = get_acad_doc()
        layers = doc.Layers
        deleted = []
        failed = []

        for name in layer_names:
            try:
                layer = layers.Item(name)
                # æ£€æŸ¥æ˜¯å¦ä¸ºå½“å‰å›¾å±‚
                if doc.ActiveLayer.Name == name:
                    print(f"[è­¦å‘Š] å›¾å±‚ '{name}' æ˜¯å½“å‰å›¾å±‚ï¼Œæ— æ³•åˆ é™¤")
                    failed.append(name)
                    continue

                # å°è¯•åˆ é™¤å›¾å±‚
                layer.Delete()
                print(f"[æˆåŠŸ] å·²åˆ é™¤å›¾å±‚ï¼š{name}")
                deleted.append(name)
            except Exception as e:
                print(f"[å¤±è´¥] æ— æ³•åˆ é™¤å›¾å±‚ '{name}'ï¼š{e}")
                failed.append(name)

        print(f"\n[ç»Ÿè®¡] æˆåŠŸåˆ é™¤ {len(deleted)} ä¸ªå›¾å±‚ï¼Œå¤±è´¥ {len(failed)} ä¸ª")
        return {'deleted': deleted, 'failed': failed}

    except Exception as e:
        print(f"[é”™è¯¯] åˆ é™¤å›¾å±‚æ—¶å‡ºé”™ï¼š{e}")
        return {'deleted': [], 'failed': layer_names}


def dim_by_points(*args):
    """
    ä½¿ç”¨å¤©æ­£é€ç‚¹æ ‡æ³¨å‘½ä»¤å¯¹å€¾æ–œå¯¹è±¡è¿›è¡Œæ ‡æ³¨

    å‚æ•°ï¼š
        *args: ä¸‰ä¸ªç‚¹åæ ‡ (x1,y1,z1), (x2,y2,z2), (x3,y3,z3)
               æˆ– p1, p2, p3 å…¶ä¸­p1,p2æ˜¯è¢«æ ‡æ³¨å¯¹è±¡çš„èµ·å§‹ç‚¹å’Œç»ˆç‚¹ï¼Œp3æ˜¯æ ‡æ³¨ä½ç½®ç‚¹

    è¿”å›ï¼š
        bool: æˆåŠŸè¿”å›True
    """
    import pyautogui
    import time

    try:
        # è§£æå‚æ•°
        if len(args) == 3:
            p1, p2, p3 = args
        else:
            print("[é”™è¯¯] éœ€è¦3ä¸ªç‚¹åæ ‡")
            return False

        # æœ€å°åŒ–æ‰€æœ‰çª—å£
        pyautogui.hotkey('win', 'd')
        time.sleep(0.5)

        # æ¿€æ´»AutoCADçª—å£
        activate_window_by_title("AutoCAD", click_titlebar=True)
        time.sleep(0.5)

        # å‘é€å¤©æ­£é€ç‚¹æ ‡æ³¨å‘½ä»¤
        _, doc = get_acad_doc()
        cmd = f"zdbz\n{p1[0]},{p1[1]}\n{p2[0]},{p2[1]}\n{p3[0]},{p3[1]}\n\n"
        doc.SendCommand(cmd)

        print("[æˆåŠŸ] å·²æ‰§è¡Œé€ç‚¹æ ‡æ³¨")
        return True

    except Exception as e:
        print(f"[é”™è¯¯] æ ‡æ³¨å¤±è´¥ï¼š{e}")
        return False


def ensure_layer(layer_name="jizhunwall"):
    """
    ç¡®ä¿å›¾å±‚å­˜åœ¨å¹¶è®¾ä¸ºå½“å‰å›¾å±‚ï¼ŒåŒæ—¶åˆ é™¤è¯¥å›¾å±‚ä¸Šæ‰€æœ‰å¯¹è±¡ï¼ˆæœ€å¤šé‡è¯• 3 æ¬¡ï¼‰ã€‚
    """
    try:
        layers = doc.Layers
        # 1) è·å–æˆ–æ–°å»ºå›¾å±‚
        try:
            layer = layers.Item(layer_name)
        except Exception:
            layer = layers.Add(layer_name)
            print(f"ğŸŸ¢ å·²æ–°å»ºå›¾å±‚ï¼š{layer_name}")
        # 2) åˆ‡æ¢åˆ°å›¾å±‚
        doc.ActiveLayer = layer
        print(f"[OK] å½“å‰å›¾å±‚å·²è®¾ç½®ä¸ºï¼š{layer_name}")

        # 3) åˆ é™¤å›¾å±‚ä¸­å…¨éƒ¨å¯¹è±¡ï¼Œé‡è¯• up to 5
        for attempt in range(1, 6):
            ents = select_tuceng(layer_name)
            if not ents:
                # å·²ç»æ²¡æœ‰å¯¹è±¡ï¼Œæå‰é€€å‡º
                print(f"ğŸ§¹ å›¾å±‚ '{layer_name}' å·²æ¸…ç©ºï¼ˆå…±å°è¯• {attempt - 1} æ¬¡ï¼‰")
                break

            deleted = 0
            for ent in ents:
                try:
                    ent.Delete()
                    deleted += 1
                except:
                    continue
            print(f"  ç¬¬ {attempt} æ¬¡åˆ é™¤ï¼šå…±åˆ é™¤ {deleted} ä¸ªå¯¹è±¡")

            time.sleep(0.1)  # çŸ­æš‚ç­‰å¾…ï¼Œç¡®ä¿å¯¹è±¡è¢«ç§»é™¤


            #åˆ·æ–°

            doc.SendCommand("RE\n")
            doc.SendCommand("Z\nE\n")

        else:
            # äº”æ¬¡éƒ½è¿˜æœ‰æ®‹ç•™
            remaining = len(select_tuceng(layer_name))
            print(f"[è­¦å‘Š] é‡è¯• 3 æ¬¡åï¼Œå›¾å±‚ '{layer_name}' ä»æœ‰ {remaining} ä¸ªå¯¹è±¡æœªèƒ½åˆ é™¤")

    except Exception as e:
        print("[é”™è¯¯] åˆ›å»º/åˆ‡æ¢å›¾å±‚æˆ–æ¸…ç†å¤±è´¥ï¼š", e)


@alias("s2")

def ensure_layer_current(layer_name="jizhunwall", max_retries=3):
    """
    ç¡®ä¿å›¾å±‚å­˜åœ¨å¹¶è®¾ä¸ºå½“å‰å›¾å±‚ï¼Œå¤±è´¥æ—¶æœ€å¤šé‡è¯• max_retries æ¬¡ã€‚
    """
    layers = doc.Layers
    for attempt in range(1, max_retries + 1):
        try:
            # è·å–æˆ–æ–°å»ºå›¾å±‚
            try:
                layer = layers.Item(layer_name)
            except Exception:
                layer = layers.Add(layer_name)
                print(f"ğŸŸ¢ å·²æ–°å»ºå›¾å±‚ï¼š{layer_name}")
            # åˆ‡æ¢åˆ°å›¾å±‚
            doc.ActiveLayer = layer
            print(f"[OK] å½“å‰å›¾å±‚å·²è®¾ç½®ä¸ºï¼š{layer_name} (å°è¯• {attempt})")
            return True
        except Exception as e:
            print(f"[é”™è¯¯] å°è¯• {attempt} åˆ›å»º/è®¾ç½®å›¾å±‚å¤±è´¥ï¼š{e}")
    print(f"[é”™è¯¯] è¾¾åˆ°æœ€å¤§é‡è¯•æ¬¡æ•° ({max_retries})ï¼Œæ— æ³•åˆ›å»ºæˆ–åˆ‡æ¢åˆ°å›¾å±‚ï¼š{layer_name}")
    return False


# è®¾ç½®æŒ‡å®šå›¾å±‚çš„é¢œè‰²ã€çº¿å‹ã€å¼€å…³çŠ¶æ€å’Œå†»ç»“çŠ¶æ€

@alias("s3")

def set_layer_properties(layer_name, color_index=9, linetype="Continuous", on=True, frozen=False):
    """
    è®¾ç½®æŒ‡å®šå›¾å±‚çš„é¢œè‰²ã€çº¿å‹ã€å¼€å…³çŠ¶æ€å’Œå†»ç»“çŠ¶æ€ã€‚

    å‚æ•°ï¼š
        layer_name (str): å›¾å±‚åç§°
        color_index (int): å›¾å±‚é¢œè‰²ç´¢å¼•ï¼ˆé»˜è®¤ 9ï¼‰
        linetype (str): å›¾å±‚çº¿å‹åç§°ï¼ˆé»˜è®¤ 'Continuous'ï¼‰
        on (bool): å›¾å±‚æ˜¯å¦æ‰“å¼€ï¼ˆé»˜è®¤ Trueï¼‰
        frozen (bool): å›¾å±‚æ˜¯å¦å†»ç»“ï¼ˆé»˜è®¤ Falseï¼‰
    """
    try:
        layers = doc.Layers
        try:
            layer = layers.Item(layer_name)
        except:
            layer = layers.Add(layer_name)
            print(f"[OK] å·²æ–°å»ºå›¾å±‚ï¼š{layer_name}")

        # è®¾ç½®é¢œè‰²
        layer.color = color_index

        # è®¾ç½®çº¿å‹
        try:
            ltype = doc.Linetypes.Item(linetype)
        except:
            doc.Linetypes.Load(linetype, linetype)
            ltype = doc.Linetypes.Item(linetype)
        layer.Linetype = linetype

        # è®¾ç½®å¼€å…³çŠ¶æ€
        layer.LayerOn = on

        # è®¾ç½®å†»ç»“çŠ¶æ€
        layer.Freeze = frozen

        print(f"ğŸ”§ å›¾å±‚å±æ€§å·²æ›´æ–°ï¼š{layer_name} | é¢œè‰²={color_index} | çº¿å‹={linetype} | å¼€å…³={'å¼€' if on else 'å…³'} | å†»ç»“={'æ˜¯' if frozen else 'å¦'}")

    except Exception as e:
        print(f"[é”™è¯¯] è®¾ç½®å›¾å±‚å±æ€§å¤±è´¥ï¼š{e}")




#&&% å°†åˆ—è¡¨ä¸­çš„å¯¹è±¡å›¾å±‚è®¾ä¸ºç›®æ ‡å›¾å±‚

def set_layer_with_retry(LB, layername, ci=3):
    """
    å°†ç»™å®š COM å¯¹è±¡åˆ—è¡¨ LB ä¸­çš„æ¯ä¸ªå¯¹è±¡çš„ Layer å±æ€§è®¾ä¸º layernameã€‚
    æ¯æ¬¡å¤±è´¥åç­‰å¾… 1 ç§’ï¼Œæœ€å¤šå°è¯• ci æ¬¡ã€‚

    å‚æ•°ï¼š
      LB         -- å¯è¿­ä»£çš„ COM å¯¹è±¡åºåˆ—ï¼Œæ¯ä¸ªå¯¹è±¡æ”¯æŒè®¾ç½® .Layer
      layername  -- è¦è®¾ç½®çš„ç›®æ ‡å›¾å±‚åç§°ï¼ˆå­—ç¬¦ä¸²ï¼‰
      ci         -- æ¯ä¸ªå¯¹è±¡æœ€å¤§é‡è¯•æ¬¡æ•°ï¼ˆé»˜è®¤ 3ï¼‰

    è¿”å›ï¼š
      æˆåŠŸè®¾ç½®å›¾å±‚çš„å¯¹è±¡åˆ—è¡¨ï¼Œä»¥åŠå¤±è´¥æœªèƒ½è®¾ç½®çš„å¯¹è±¡åˆ—è¡¨ï¼š
        (success_list, failed_list)
    """
    success = []
    failed = []

    for obj in LB:
        for attempt in range(1, ci + 1):
            try:
                obj.Layer = layername
                success.append(obj)
                break
            except Exception as e:
                if attempt == ci:
                    # æœ€åä¸€æ¬¡ä»å¤±è´¥
                    failed.append(obj)
                    print(f"[è­¦å‘Š]ï¸ å¯¹è±¡ {getattr(obj, 'Handle', '<unknown>')} è®¾ç½®å›¾å±‚â€œ{layername}â€å¤±è´¥ï¼š{e}")
                else:
                    time.sleep(1)
    return success, failed








#_____________________________________________________________________________________________________________________________________________________________________________
#_____________________________________________________________________________________________________________________________________________________________________________



print("__________________  CADåŸºæœ¬æ“ä½œå¼€å§‹è¿è¡Œ _________________________")


































































































#&&&&%% ç¬¬åä¸€éƒ¨åˆ†  å¯¹è±¡å±æ€§è®¿é—®

"""
ç»Ÿä¸€çš„CADå¯¹è±¡å’Œå¤©æ­£å¯¹è±¡å±æ€§è®¿é—®
æ”¯æŒCADæ ‡å‡†å¯¹è±¡ï¼ˆé€šè¿‡CastToï¼‰å’Œå¤©æ­£å¯¹è±¡ï¼ˆé€šè¿‡IDispatch.Invokeï¼‰
"""

import pythoncom
from win32com.client import CastTo

# CADæ ‡å‡†å¯¹è±¡ç±»å‹æ˜ å°„ï¼ˆä¸ç¬¬å››éƒ¨åˆ†ä¿æŒä¸€è‡´ï¼‰
_CAST_MAP = {
    # åŸºç¡€å‡ ä½•
    "AcDbLine":"IAcadLine", "AcDbCircle":"IAcadCircle", "AcDbArc":"IAcadArc","AcDbPoint":"IAcadPoint",
    "AcDbEllipse":"IAcadEllipse", "AcDbSpline":"IAcadSpline",
    # å¤šæ®µçº¿
    "AcDbPolyline":"IAcadLWPolyline", "AcDb2dPolyline":"IAcadPolyline", "AcDb3dPolyline":"IAcad3DPolyline",
    # æ–‡å­—
    "AcDbText":"IAcadText", "AcDbMText":"IAcadMText",
    # å—/å±æ€§
    "AcDbBlockReference":"IAcadBlockReference",
    "AcDbAttribute":"IAcadAttributeReference", "AcDbAttributeDefinition":"IAcadAttribute",
    # å¼•çº¿/æ ‡æ³¨ï¼ˆå¸¸ç”¨ï¼‰
    "AcDbLeader":"IAcadLeader", "AcDbMLeader":"IAcadMLeader",
    "AcDbDimension":"IAcadDimension", "AcDbAlignedDimension":"IAcadDimAligned",
    "AcDbRotatedDimension":"IAcadDimRotated", "AcDbRadialDimension":"IAcadDimRadial",
    "AcDbDiametricDimension":"IAcadDimDiametric", "AcDbArcDimension":"IAcadDimArc",
    "AcDb3PointAngularDimension":"IAcadDim3PointAngular", "AcDb2LineAngularDimension":"IAcadDim2LineAngular",
    "AcDbOrdinateDimension":"IAcadDimOrdinate",
    # å…¶å®ƒ
    "AcDbHatch":"IAcadHatch", "AcDbTable":"IAcadTable",
}

# å¤©æ­£å¯¹è±¡å±æ€§DISPIDæ˜ å°„
_TARCH_PROPERTY_MAP = {
    'TDbOpening': {'Offset': 1, 'Width': 2, 'Type': 3, 'Direction': 7,
                   'Angle': 8, 'Height': 10, 'Name': 11},
    'TDbWall': {'Offset1': 1, 'Thickness': 2, 'Thickness2': 3, 'Length': 4,
                'WallType': 11, 'Material': 13, 'Hatch': 21, 'Surface': 22},
}

def _maybe_cast(ent):
    """å®‰å…¨åœ°è½¬æ¢CADå¯¹è±¡åˆ°ä¸“ç”¨æ¥å£ï¼ˆä½¿ç”¨com_retryï¼‰"""
    try:
        name = com_retry(lambda: ent.ObjectName)
        iface = _CAST_MAP.get(name)
        if iface:
            try:
                return CastTo(ent, iface)
            except Exception:
                return ent
        return ent
    except Exception:
        return ent

def cast_object(obj):
    """è½¬æ¢CADå¯¹è±¡åˆ°ä¸“ç”¨æ¥å£ï¼ˆå…¼å®¹æ—§ä»£ç ï¼Œå†…éƒ¨è°ƒç”¨_maybe_castï¼‰"""
    return _maybe_cast(obj)

def get_object_property(obj, property_name):
    """ç»Ÿä¸€è·å–å¯¹è±¡å±æ€§ï¼ˆè‡ªåŠ¨è¯†åˆ«CADå¯¹è±¡æˆ–å¤©æ­£å¯¹è±¡ï¼‰"""
    try:
        obj_name = com_retry(lambda: obj.ObjectName)
        # å¤©æ­£å¯¹è±¡ï¼šä½¿ç”¨DISPIDæ–¹å¼è®¿é—®
        if obj_name in _TARCH_PROPERTY_MAP:
            dispid = _TARCH_PROPERTY_MAP[obj_name].get(property_name)
            if dispid:
                return obj._oleobj_.Invoke(dispid, 0, pythoncom.DISPATCH_PROPERTYGET, True)
        # CADæ ‡å‡†å¯¹è±¡ï¼šå…ˆCastå†è®¿é—®å±æ€§
        obj = _maybe_cast(obj)
        return getattr(obj, property_name)
    except Exception as e:
        return None

def set_object_property(obj, property_name, value):
    """ç»Ÿä¸€è®¾ç½®å¯¹è±¡å±æ€§ï¼ˆè‡ªåŠ¨è¯†åˆ«CADå¯¹è±¡æˆ–å¤©æ­£å¯¹è±¡ï¼‰"""
    try:
        obj_name = com_retry(lambda: obj.ObjectName)
        # å¤©æ­£å¯¹è±¡ï¼šä½¿ç”¨DISPIDæ–¹å¼è®¾ç½®
        if obj_name in _TARCH_PROPERTY_MAP:
            dispid = _TARCH_PROPERTY_MAP[obj_name].get(property_name)
            if dispid:
                obj._oleobj_.Invoke(dispid, 0, pythoncom.DISPATCH_PROPERTYPUT, True, value)
                return True
        # CADæ ‡å‡†å¯¹è±¡ï¼šå…ˆCastå†è®¾ç½®å±æ€§
        obj = _maybe_cast(obj)
        setattr(obj, property_name, value)
        return True
    except Exception as e:
        return False


